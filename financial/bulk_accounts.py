from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db.models import Q
from openpyxl import Workbook, load_workbook

from catalog.models import ProductBulkJob
from entity.models import Entity
from financial.models import account
from financial.serializers_ledger import AccountProfileV2WriteSerializer

SHEET = "accounts"


def _normalize_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _is_placeholder(value: str) -> bool:
    return bool(re.fullmatch(r"val-\d+", value, flags=re.IGNORECASE))


def _to_bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    text = _normalize_text(value).lower()
    if _is_placeholder(text):
        return default
    return text in {"1", "true", "yes", "y"}


def _to_int(value: Any) -> int | None:
    text = _normalize_text(value)
    if not text or text in {"-", "--"} or _is_placeholder(text):
        return None
    return int(float(text))


def _to_decimal(value: Any, default: Decimal | None = None) -> Decimal | None:
    text = _normalize_text(value)
    if not text or text in {"-", "--"} or _is_placeholder(text):
        return default
    if text.endswith("%"):
        text = text[:-1].strip()
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def _to_datetime_string(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, time.min).isoformat()

    text = _normalize_text(value)
    if not text or _is_placeholder(text):
        return None

    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.isoformat()
        except ValueError:
            continue
    # Let DRF serializer decide if this is acceptable/unacceptable.
    return text


def _read_xlsx(content: bytes) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    sheet_lookup = {str(name).strip().lower(): name for name in wb.sheetnames}
    actual = sheet_lookup.get(SHEET)
    if not actual:
        return {SHEET: []}
    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {SHEET: []}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    payload: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        payload.append({headers[i]: row[i] for i in range(len(headers))})
    return {SHEET: payload}


def _read_csv_zip(content: bytes) -> dict[str, list[dict[str, Any]]]:
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        name_lookup = {str(name).strip().lower(): name for name in zf.namelist()}
        fname = name_lookup.get(f"{SHEET}.csv")
        if not fname:
            return {SHEET: []}
        data = zf.read(fname).decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(data))
        return {SHEET: [dict(r) for r in reader if any((v or "").strip() for v in r.values())]}


def _write_xlsx(data: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = Workbook()
    for idx, (sheet_name, rows) in enumerate(data.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name[:31] or f"sheet_{idx + 1}"
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(col) for col in headers])
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _write_csv_zip(data: dict[str, list[dict[str, Any]]]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        for sheet_name, rows in data.items():
            stream = io.StringIO()
            headers = list(rows[0].keys()) if rows else []
            writer = csv.DictWriter(stream, fieldnames=headers)
            if headers:
                writer.writeheader()
                writer.writerows(rows)
            zf.writestr(f"{sheet_name}.csv", stream.getvalue())
    return buff.getvalue()


def template_payload() -> dict[str, list[dict[str, Any]]]:
    return {
        SHEET: [
            {
                "id": "",
                "ledger_code": "",
                "accountname": "ABC Traders",
                "legalname": "ABC Traders Pvt Ltd",
                "emailid": "abc@example.com",
                "contactno": "9876543210",
                "isactive": True,
                "accounthead": "",
                "creditaccounthead": "",
                "accounttype": "",
                "openingbdr": "0.00",
                "openingbcr": "0.00",
                "partytype": "Customer",
                "gstno": "29ABCDE1234F1Z5",
                "pan": "ABCDE1234F",
                "creditlimit": "0.00",
                "creditdays": "0",
                "paymentterms": "Net30",
                "currency": "INR",
                "line1": "Address Line 1",
                "line2": "",
                "floor_no": "",
                "street": "",
                "country": "",
                "state": "",
                "district": "",
                "city": "",
                "pincode": "560001",
            }
        ]
    }


def export_payload(entity: Entity, search: str = "") -> dict[str, list[dict[str, Any]]]:
    qs = account.objects.select_related("compliance_profile", "commercial_profile").filter(entity=entity)
    if search:
        qs = qs.filter(
            Q(accountname__icontains=search)
            | Q(legalname__icontains=search)
            | Q(compliance_profile__gstno__icontains=search)
            | Q(compliance_profile__pan__icontains=search)
        )
    rows: list[dict[str, Any]] = []
    for row in qs.order_by("accountname"):
        compliance = getattr(row, "compliance_profile", None)
        commercial = getattr(row, "commercial_profile", None)
        primary_address = row.addresses.filter(isprimary=True, isactive=True).first()
        rows.append(
            {
                "id": row.id,
                "ledger_code": row.accountcode,
                "accountname": row.accountname,
                "legalname": row.legalname,
                "emailid": row.emailid,
                "contactno": row.contactno,
                "isactive": row.isactive,
                "accounthead": row.accounthead_id,
                "creditaccounthead": row.creditaccounthead_id,
                "accounttype": row.accounttype_id,
                "openingbdr": str(row.openingbdr or "0.00"),
                "openingbcr": str(row.openingbcr or "0.00"),
                "partytype": getattr(commercial, "partytype", None),
                "gstno": getattr(compliance, "gstno", None),
                "pan": getattr(compliance, "pan", None),
                "creditlimit": str(getattr(commercial, "creditlimit", Decimal("0.00")) or "0.00"),
                "creditdays": getattr(commercial, "creditdays", None),
                "paymentterms": getattr(commercial, "paymentterms", None),
                "currency": getattr(commercial, "currency", None),
                "line1": getattr(primary_address, "line1", None),
                "line2": getattr(primary_address, "line2", None),
                "floor_no": getattr(primary_address, "floor_no", None),
                "street": getattr(primary_address, "street", None),
                "country": getattr(primary_address, "country_id", None),
                "state": getattr(primary_address, "state_id", None),
                "district": getattr(primary_address, "district_id", None),
                "city": getattr(primary_address, "city_id", None),
                "pincode": getattr(primary_address, "pincode", None),
            }
        )
    return {SHEET: rows}


@dataclass
class ImportResult:
    summary: dict[str, Any]
    errors: list[dict[str, Any]]


def parse_payload(file_bytes: bytes, fmt: str) -> dict[str, list[dict[str, Any]]]:
    if fmt == "xlsx":
        return _read_xlsx(file_bytes)
    return _read_csv_zip(file_bytes)


def render_payload(payload: dict[str, list[dict[str, Any]]], fmt: str) -> bytes:
    if fmt == "xlsx":
        return _write_xlsx(payload)
    return _write_csv_zip(payload)


def _flatten_errors(errors: Any, prefix: str = "") -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    if isinstance(errors, dict):
        for key, value in errors.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            out.extend(_flatten_errors(value, next_prefix))
    elif isinstance(errors, list):
        for item in errors:
            if isinstance(item, (dict, list)):
                out.extend(_flatten_errors(item, prefix))
            else:
                out.append((prefix or "non_field_errors", str(item)))
    else:
        out.append((prefix or "non_field_errors", str(errors)))
    return out


def _resolve_existing_account(entity: Entity, row: dict[str, Any]) -> account | None:
    row_id = _to_int(row.get("id"))
    if row_id:
        return account.objects.filter(entity=entity, pk=row_id).first()

    ledger_code = _to_int(row.get("ledger_code"))
    if ledger_code:
        existing = account.objects.filter(entity=entity, accountcode=ledger_code).first()
        if existing:
            return existing

    gstno = _normalize_text(row.get("gstno"))
    if gstno and not _is_placeholder(gstno):
        existing = account.objects.filter(entity=entity, compliance_profile__gstno__iexact=gstno).first()
        if existing:
            return existing

    accountname = _normalize_text(row.get("accountname"))
    if accountname and not _is_placeholder(accountname):
        return account.objects.filter(entity=entity, accountname__iexact=accountname).first()
    return None


def _build_payload(entity: Entity, row: dict[str, Any]) -> dict[str, Any]:
    compliance: dict[str, Any] = {}
    commercial: dict[str, Any] = {}
    primary_address: dict[str, Any] = {}

    def set_if_value(target: dict[str, Any], key: str, value: Any):
        if value not in (None, "", "-", "--"):
            target[key] = value

    payload: dict[str, Any] = {
        "entity": entity.id,
    }

    set_if_value(payload, "accountname", _normalize_text(row.get("accountname")))
    set_if_value(payload, "legalname", _normalize_text(row.get("legalname")))
    set_if_value(payload, "emailid", _normalize_text(row.get("emailid")))
    set_if_value(payload, "contactno", _normalize_text(row.get("contactno")))
    payload["isactive"] = _to_bool(row.get("isactive"), default=True)

    set_if_value(payload, "ledger_code", _to_int(row.get("ledger_code")))
    set_if_value(payload, "accounthead", _to_int(row.get("accounthead")))
    set_if_value(payload, "creditaccounthead", _to_int(row.get("creditaccounthead")))
    set_if_value(payload, "accounttype", _to_int(row.get("accounttype")))
    set_if_value(payload, "openingbdr", _to_decimal(row.get("openingbdr"), default=None))
    set_if_value(payload, "openingbcr", _to_decimal(row.get("openingbcr"), default=None))
    if row.get("canbedeleted") not in (None, ""):
        payload["canbedeleted"] = _to_bool(row.get("canbedeleted"), default=True)

    set_if_value(compliance, "gstno", _normalize_text(row.get("gstno")))
    set_if_value(compliance, "pan", _normalize_text(row.get("pan")))
    set_if_value(compliance, "gstintype", _normalize_text(row.get("gstintype")))
    set_if_value(compliance, "gstregtype", _normalize_text(row.get("gstregtype")))
    if row.get("is_sez") not in (None, ""):
        compliance["is_sez"] = _to_bool(row.get("is_sez"), default=False)

    set_if_value(commercial, "partytype", _normalize_text(row.get("partytype")))
    set_if_value(commercial, "creditlimit", _to_decimal(row.get("creditlimit"), default=None))
    set_if_value(commercial, "creditdays", _to_int(row.get("creditdays")))
    set_if_value(commercial, "paymentterms", _normalize_text(row.get("paymentterms")))
    set_if_value(commercial, "currency", _normalize_text(row.get("currency")))

    set_if_value(primary_address, "line1", _normalize_text(row.get("line1")))
    set_if_value(primary_address, "line2", _normalize_text(row.get("line2")))
    set_if_value(primary_address, "floor_no", _normalize_text(row.get("floor_no")))
    set_if_value(primary_address, "street", _normalize_text(row.get("street")))
    set_if_value(primary_address, "country_id", _to_int(row.get("country")))
    set_if_value(primary_address, "state_id", _to_int(row.get("state")))
    set_if_value(primary_address, "district_id", _to_int(row.get("district")))
    set_if_value(primary_address, "city_id", _to_int(row.get("city")))
    set_if_value(primary_address, "pincode", _normalize_text(row.get("pincode")))

    set_if_value(payload, "dateofreg", _to_datetime_string(row.get("dateofreg")))
    set_if_value(payload, "dateofdreg", _to_datetime_string(row.get("dateofdreg")))

    if compliance:
        payload["compliance_profile"] = compliance
    if commercial:
        payload["commercial_profile"] = commercial
    if primary_address:
        payload["primary_address"] = primary_address

    return payload


def validate_payload(
    payload: dict[str, list[dict[str, Any]]],
    entity: Entity,
    *,
    upsert_mode: str = ProductBulkJob.UpsertMode.UPSERT,
) -> ImportResult:
    errors: list[dict[str, Any]] = []
    rows = payload.get(SHEET, [])
    if not rows:
        errors.append({"sheet": SHEET, "row": 0, "field": "file", "message": "No account rows found in upload."})
        return ImportResult(summary={"rows": {SHEET: 0}, "error_count": 1}, errors=errors)

    for idx, row in enumerate(rows, start=2):
        existing = _resolve_existing_account(entity, row)
        if existing and upsert_mode == ProductBulkJob.UpsertMode.CREATE_ONLY:
            continue
        if not existing and upsert_mode == ProductBulkJob.UpsertMode.UPDATE_ONLY:
            continue

        try:
            serializer_payload = _build_payload(entity, row)
        except Exception as exc:
            errors.append({"sheet": SHEET, "row": idx, "field": "row", "message": str(exc)})
            continue

        serializer = AccountProfileV2WriteSerializer(
            existing,
            data=serializer_payload,
            partial=bool(existing),
            context={"request": None},
        )
        if not serializer.is_valid():
            for field, message in _flatten_errors(serializer.errors):
                errors.append({"sheet": SHEET, "row": idx, "field": field, "message": message})

    return ImportResult(summary={"rows": {SHEET: len(rows)}, "error_count": len(errors)}, errors=errors)


def commit_payload(
    payload: dict[str, list[dict[str, Any]]],
    entity: Entity,
    *,
    upsert_mode: str = ProductBulkJob.UpsertMode.UPSERT,
    duplicate_strategy: str = ProductBulkJob.DuplicateStrategy.FAIL,
    request=None,
) -> ImportResult:
    rows = payload.get(SHEET, [])
    errors: list[dict[str, Any]] = []
    summary = {"created": 0, "updated": 0, "skipped": 0}

    for idx, row in enumerate(rows, start=2):
        existing = _resolve_existing_account(entity, row)
        if existing and upsert_mode == ProductBulkJob.UpsertMode.CREATE_ONLY:
            summary["skipped"] += 1
            continue
        if not existing and upsert_mode == ProductBulkJob.UpsertMode.UPDATE_ONLY:
            summary["skipped"] += 1
            continue

        try:
            serializer_payload = _build_payload(entity, row)
            serializer = AccountProfileV2WriteSerializer(
                existing,
                data=serializer_payload,
                partial=bool(existing),
                context={"request": request},
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
            if existing:
                summary["updated"] += 1
            else:
                summary["created"] += 1
        except Exception as exc:
            errors.append({"sheet": SHEET, "row": idx, "field": "row", "message": str(exc)})
            if duplicate_strategy == ProductBulkJob.DuplicateStrategy.FAIL:
                break

    summary["error_count"] = len(errors)
    return ImportResult(summary=summary, errors=errors)
