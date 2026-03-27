from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook

from catalog.models import ProductBulkJob
from withholding.models import EntityWithholdingConfig, WithholdingSection
from withholding.serializers import EntityWithholdingConfigSerializer, WithholdingSectionSerializer

SECTIONS_SHEET = "tcs_sections"
CONFIGS_SHEET = "tcs_configs"


def _normalize_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _to_bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return _normalize_text(value).lower() in {"1", "true", "yes", "y"}


def _to_int(value: Any) -> int | None:
    text = _normalize_text(value)
    if text in {"", "-", "--"}:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid integer value: {value}") from exc


def _to_decimal(value: Any) -> Decimal | None:
    text = _normalize_text(value)
    if text in {"", "-", "--"}:
        return None
    text = text.replace(",", "")
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def _to_date_string(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = _normalize_text(value)
    parsed = parse_date(text)
    if parsed:
        return parsed.isoformat()
    if "-" in text and len(text.split("-")) == 3:
        # dd-mm-yyyy fallback
        parts = text.split("-")
        if len(parts[0]) == 2:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return text


def _read_xlsx(content: bytes, sheet_name: str) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    lookup = {str(name).strip().lower(): name for name in wb.sheetnames}
    actual = lookup.get(sheet_name.lower())
    if not actual:
        return {sheet_name: []}
    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {sheet_name: []}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return {sheet_name: data}


def _read_csv_zip(content: bytes, sheet_name: str) -> dict[str, list[dict[str, Any]]]:
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        lookup = {str(name).strip().lower(): name for name in zf.namelist()}
        fname = lookup.get(f"{sheet_name}.csv")
        if not fname:
            return {sheet_name: []}
        raw = zf.read(fname).decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw))
        return {sheet_name: [dict(r) for r in reader if any((v or "").strip() for v in r.values())]}


def _write_xlsx(data: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = Workbook()
    for idx, (sheet, rows) in enumerate(data.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet[:31]
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h) for h in headers])
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _write_csv_zip(data: dict[str, list[dict[str, Any]]]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        for sheet, rows in data.items():
            stream = io.StringIO()
            headers = list(rows[0].keys()) if rows else []
            writer = csv.DictWriter(stream, fieldnames=headers)
            if headers:
                writer.writeheader()
                writer.writerows(rows)
            zf.writestr(f"{sheet}.csv", stream.getvalue())
    return buff.getvalue()


def parse_payload(file_bytes: bytes, fmt: str, sheet_name: str) -> dict[str, list[dict[str, Any]]]:
    if fmt == "xlsx":
        return _read_xlsx(file_bytes, sheet_name)
    return _read_csv_zip(file_bytes, sheet_name)


def render_payload(payload: dict[str, list[dict[str, Any]]], fmt: str) -> bytes:
    if fmt == "xlsx":
        return _write_xlsx(payload)
    return _write_csv_zip(payload)


def sections_template_payload() -> dict[str, list[dict[str, Any]]]:
    return {
        SECTIONS_SHEET: [
            {
                "id": "",
                "section_code": "206C(1)",
                "description": "TCS on specified goods",
                "law_type": "INCOME_TAX",
                "sub_type": "206C_1",
                "base_rule": 1,
                "rate_default": "1.0000",
                "threshold_default": "0.00",
                "requires_pan": False,
                "higher_rate_no_pan": "",
                "effective_from": "2026-04-01",
                "effective_to": "",
                "is_active": True,
            }
        ]
    }


def sections_export_payload(search: str = "") -> dict[str, list[dict[str, Any]]]:
    qs = WithholdingSection.objects.filter(tax_type=2).order_by("section_code", "-effective_from")
    if search:
        qs = qs.filter(section_code__icontains=search) | qs.filter(description__icontains=search)
    rows = [
        {
            "id": row.id,
            "section_code": row.section_code,
            "description": row.description,
            "law_type": row.law_type,
            "sub_type": row.sub_type,
            "base_rule": row.base_rule,
            "rate_default": str(row.rate_default),
            "threshold_default": str(row.threshold_default or ""),
            "requires_pan": row.requires_pan,
            "higher_rate_no_pan": str(row.higher_rate_no_pan or ""),
            "effective_from": row.effective_from.isoformat() if row.effective_from else "",
            "effective_to": row.effective_to.isoformat() if row.effective_to else "",
            "is_active": row.is_active,
        }
        for row in qs
    ]
    return {SECTIONS_SHEET: rows}


def configs_template_payload(default_entity: int | None, default_entityfin: int | None) -> dict[str, list[dict[str, Any]]]:
    return {
        CONFIGS_SHEET: [
            {
                "id": "",
                "entity": default_entity or "",
                "entityfin": default_entityfin or "",
                "subentity": "",
                "enable_tds": True,
                "enable_tcs": True,
                "default_tds_section": "",
                "default_tcs_section": "",
                "apply_194q": False,
                "apply_tcs_206c1h": False,
                "effective_from": "2026-04-01",
                "rounding_places": 2,
            }
        ]
    }


def configs_export_payload(entity_id: int | None = None, entityfin_id: int | None = None, search: str = "") -> dict[str, list[dict[str, Any]]]:
    qs = EntityWithholdingConfig.objects.all().order_by("-effective_from", "-id")
    if entity_id:
        qs = qs.filter(entity_id=entity_id)
    if entityfin_id:
        qs = qs.filter(entityfin_id=entityfin_id)
    if search:
        qs = qs.filter(entity__entityname__icontains=search)
    rows = [
        {
            "id": row.id,
            "entity": row.entity_id,
            "entityfin": row.entityfin_id,
            "subentity": row.subentity_id or "",
            "enable_tds": row.enable_tds,
            "enable_tcs": row.enable_tcs,
            "default_tds_section": row.default_tds_section_id or "",
            "default_tcs_section": row.default_tcs_section_id or "",
            "apply_194q": row.apply_194q,
            "apply_tcs_206c1h": row.apply_tcs_206c1h,
            "effective_from": row.effective_from.isoformat() if row.effective_from else "",
            "rounding_places": row.rounding_places,
        }
        for row in qs
    ]
    return {CONFIGS_SHEET: rows}


def _flatten_errors(errors: Any, prefix: str = "") -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    if isinstance(errors, dict):
        for key, value in errors.items():
            p = f"{prefix}.{key}" if prefix else str(key)
            out.extend(_flatten_errors(value, p))
    elif isinstance(errors, list):
        for item in errors:
            if isinstance(item, (dict, list)):
                out.extend(_flatten_errors(item, prefix))
            else:
                out.append((prefix or "non_field_errors", str(item)))
    else:
        out.append((prefix or "non_field_errors", str(errors)))
    return out


@dataclass
class ImportResult:
    summary: dict[str, Any]
    errors: list[dict[str, Any]]


def _section_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "tax_type": 2,
        "law_type": _normalize_text(row.get("law_type")) or "INCOME_TAX",
        "sub_type": _normalize_text(row.get("sub_type")) or None,
        "section_code": _normalize_text(row.get("section_code")),
        "description": _normalize_text(row.get("description")),
        "base_rule": _to_int(row.get("base_rule")) or 1,
        "rate_default": _to_decimal(row.get("rate_default")) if row.get("rate_default") not in (None, "") else Decimal("0.0000"),
        "threshold_default": _to_decimal(row.get("threshold_default")),
        "requires_pan": _to_bool(row.get("requires_pan")),
        "higher_rate_no_pan": _to_decimal(row.get("higher_rate_no_pan")),
        "applicability_json": {},
        "effective_from": _to_date_string(row.get("effective_from")),
        "effective_to": _to_date_string(row.get("effective_to")),
        "is_active": _to_bool(row.get("is_active"), default=True),
    }


def _config_payload(row: dict[str, Any], default_entity: int | None, default_entityfin: int | None) -> dict[str, Any]:
    return {
        "entity": _to_int(row.get("entity")) or default_entity,
        "entityfin": _to_int(row.get("entityfin")) or default_entityfin,
        "subentity": _to_int(row.get("subentity")),
        "enable_tds": _to_bool(row.get("enable_tds"), default=True),
        "enable_tcs": _to_bool(row.get("enable_tcs"), default=True),
        "default_tds_section": _to_int(row.get("default_tds_section")),
        "default_tcs_section": _to_int(row.get("default_tcs_section")),
        "apply_194q": _to_bool(row.get("apply_194q"), default=False),
        "apply_tcs_206c1h": _to_bool(row.get("apply_tcs_206c1h"), default=False),
        "effective_from": _to_date_string(row.get("effective_from")),
        "rounding_places": _to_int(row.get("rounding_places")) or 2,
    }


def validate_sections_payload(payload: dict[str, list[dict[str, Any]]], *, upsert_mode: str) -> ImportResult:
    rows = payload.get(SECTIONS_SHEET, [])
    errors: list[dict[str, Any]] = []
    if not rows:
        errors.append({"sheet": SECTIONS_SHEET, "row": 0, "field": "file", "message": "No rows found."})
        return ImportResult(summary={"rows": {SECTIONS_SHEET: 0}, "error_count": 1}, errors=errors)

    for idx, row in enumerate(rows, start=2):
        row_id = _to_int(row.get("id"))
        existing = WithholdingSection.objects.filter(pk=row_id, tax_type=2).first() if row_id else None
        if existing and upsert_mode == ProductBulkJob.UpsertMode.CREATE_ONLY:
            continue
        if not existing and upsert_mode == ProductBulkJob.UpsertMode.UPDATE_ONLY:
            continue
        try:
            data = _section_payload(row)
        except Exception as exc:
            errors.append({"sheet": SECTIONS_SHEET, "row": idx, "field": "row", "message": str(exc)})
            continue
        serializer = WithholdingSectionSerializer(existing, data=data, partial=bool(existing))
        if not serializer.is_valid():
            for field, msg in _flatten_errors(serializer.errors):
                errors.append({"sheet": SECTIONS_SHEET, "row": idx, "field": field, "message": msg})
    return ImportResult(summary={"rows": {SECTIONS_SHEET: len(rows)}, "error_count": len(errors)}, errors=errors)


def validate_configs_payload(
    payload: dict[str, list[dict[str, Any]]],
    *,
    upsert_mode: str,
    default_entity: int | None,
    default_entityfin: int | None,
) -> ImportResult:
    rows = payload.get(CONFIGS_SHEET, [])
    errors: list[dict[str, Any]] = []
    if not rows:
        errors.append({"sheet": CONFIGS_SHEET, "row": 0, "field": "file", "message": "No rows found."})
        return ImportResult(summary={"rows": {CONFIGS_SHEET: 0}, "error_count": 1}, errors=errors)

    for idx, row in enumerate(rows, start=2):
        row_id = _to_int(row.get("id"))
        existing = EntityWithholdingConfig.objects.filter(pk=row_id).first() if row_id else None
        if existing and upsert_mode == ProductBulkJob.UpsertMode.CREATE_ONLY:
            continue
        if not existing and upsert_mode == ProductBulkJob.UpsertMode.UPDATE_ONLY:
            continue
        try:
            data = _config_payload(row, default_entity, default_entityfin)
        except Exception as exc:
            errors.append({"sheet": CONFIGS_SHEET, "row": idx, "field": "row", "message": str(exc)})
            continue
        serializer = EntityWithholdingConfigSerializer(existing, data=data, partial=bool(existing))
        if not serializer.is_valid():
            for field, msg in _flatten_errors(serializer.errors):
                errors.append({"sheet": CONFIGS_SHEET, "row": idx, "field": field, "message": msg})
    return ImportResult(summary={"rows": {CONFIGS_SHEET: len(rows)}, "error_count": len(errors)}, errors=errors)


def commit_sections_payload(
    payload: dict[str, list[dict[str, Any]]],
    *,
    upsert_mode: str,
    duplicate_strategy: str,
) -> ImportResult:
    rows = payload.get(SECTIONS_SHEET, [])
    errors: list[dict[str, Any]] = []
    summary = {"created": 0, "updated": 0, "skipped": 0}
    for idx, row in enumerate(rows, start=2):
        row_id = _to_int(row.get("id"))
        existing = WithholdingSection.objects.filter(pk=row_id, tax_type=2).first() if row_id else None
        if existing and upsert_mode == ProductBulkJob.UpsertMode.CREATE_ONLY:
            summary["skipped"] += 1
            continue
        if not existing and upsert_mode == ProductBulkJob.UpsertMode.UPDATE_ONLY:
            summary["skipped"] += 1
            continue
        try:
            data = _section_payload(row)
            serializer = WithholdingSectionSerializer(existing, data=data, partial=bool(existing))
            serializer.is_valid(raise_exception=True)
            serializer.save()
            summary["updated" if existing else "created"] += 1
        except Exception as exc:
            errors.append({"sheet": SECTIONS_SHEET, "row": idx, "field": "row", "message": str(exc)})
            if duplicate_strategy == ProductBulkJob.DuplicateStrategy.FAIL:
                break
    summary["error_count"] = len(errors)
    return ImportResult(summary=summary, errors=errors)


def commit_configs_payload(
    payload: dict[str, list[dict[str, Any]]],
    *,
    upsert_mode: str,
    duplicate_strategy: str,
    default_entity: int | None,
    default_entityfin: int | None,
) -> ImportResult:
    rows = payload.get(CONFIGS_SHEET, [])
    errors: list[dict[str, Any]] = []
    summary = {"created": 0, "updated": 0, "skipped": 0}
    for idx, row in enumerate(rows, start=2):
        row_id = _to_int(row.get("id"))
        existing = EntityWithholdingConfig.objects.filter(pk=row_id).first() if row_id else None
        if existing and upsert_mode == ProductBulkJob.UpsertMode.CREATE_ONLY:
            summary["skipped"] += 1
            continue
        if not existing and upsert_mode == ProductBulkJob.UpsertMode.UPDATE_ONLY:
            summary["skipped"] += 1
            continue
        try:
            data = _config_payload(row, default_entity, default_entityfin)
            serializer = EntityWithholdingConfigSerializer(existing, data=data, partial=bool(existing))
            serializer.is_valid(raise_exception=True)
            serializer.save()
            summary["updated" if existing else "created"] += 1
        except Exception as exc:
            errors.append({"sheet": CONFIGS_SHEET, "row": idx, "field": "row", "message": str(exc)})
            if duplicate_strategy == ProductBulkJob.DuplicateStrategy.FAIL:
                break
    summary["error_count"] = len(errors)
    return ImportResult(summary=summary, errors=errors)

