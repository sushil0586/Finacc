from __future__ import annotations

from decimal import Decimal
from datetime import date
import csv
import io
import logging
import re
from time import perf_counter
import zipfile

from django.db.models import Count, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.db import transaction
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from rbac.services import EffectivePermissionService
from subscriptions.services import SubscriptionLimitCodes, SubscriptionService
from withholding.models import (
    EntityPartyTaxProfile,
    EntityWithholdingSectionPostingMap,
    EntityTcsThresholdOpening,
    GstTcsComputation,
    GstTcsEcoProfile,
    PartyTaxProfile,
    TcsCollection,
    TcsComputation,
    TcsDeposit,
    TcsDepositAllocation,
    TcsQuarterlyReturn,
    WithholdingSectionPolicyAudit,
    WithholdingSection,
)
from withholding.serializers import (
    EntityPartyTaxProfileSerializer,
    EntityWithholdingSectionPostingMapSerializer,
    EntityTcsThresholdOpeningSerializer,
    EntityWithholdingConfigSerializer,
    GstTcsComputationSerializer,
    GstTcsComputeRequestSerializer,
    GstTcsEcoProfileSerializer,
    PartyTaxProfileSerializer,
    TcsCollectionSerializer,
    TcsComputationSerializer,
    TcsComputeConfirmSerializer,
    TcsComputeRequestSerializer,
    TcsDepositAllocationSerializer,
    TcsDepositAllocationRequestSerializer,
    TcsDepositSerializer,
    TcsQuarterlyReturnSerializer,
    WithholdingSectionPolicyAuditSerializer,
    WithholdingSectionSerializer,
    build_preview_payload,
)
from withholding.services import compute_withholding_preview, q2, upsert_tcs_computation
from financial.profile_access import account_pan
from payments.models.payment_core import PaymentVoucherHeader
from purchase.models import PurchaseInvoiceHeader, PurchaseInvoiceLine
from sales.models import SalesInvoiceHeader, SalesInvoiceLine
from reports.api.receivables_views import _safe_filename


TCS_FEATURE_CODE = SubscriptionLimitCodes.FEATURE_FINANCIAL
TCS_REPORTING_FEATURE_CODE = SubscriptionLimitCodes.FEATURE_REPORTING
TCS_OPERATIONAL_MODE = SubscriptionService.ACCESS_MODE_OPERATIONAL

TCS_CONFIG_VIEW_PERMISSIONS = ("compliance.tcs_config.view", "tcs.config.view")
TCS_CONFIG_MANAGE_PERMISSIONS = ("compliance.tcs_config.update", "tcs.config.update", "tcs.config.edit", "tcs.config.create")
TCS_CONFIG_DELETE_PERMISSIONS = ("compliance.tcs_config.update", "tcs.config.delete")
TCS_SECTION_VIEW_PERMISSIONS = ("compliance.tcs_section.view", "tcs.section.view", "tcs.sections.view")
TCS_SECTION_CREATE_PERMISSIONS = ("compliance.tcs_section.create", "tcs.section.create", "tcs.sections.create")
TCS_SECTION_UPDATE_PERMISSIONS = ("compliance.tcs_section.update", "tcs.section.update", "tcs.sections.update")
TCS_SECTION_DELETE_PERMISSIONS = ("compliance.tcs_section.delete", "tcs.section.delete", "tcs.sections.delete")
TCS_RULE_VIEW_PERMISSIONS = ("compliance.tcs_rule.view", "tcs.rule.view", "tcs.rules.view")
TCS_RULE_CREATE_PERMISSIONS = ("compliance.tcs_rule.create", "tcs.rule.create", "tcs.rules.create")
TCS_RULE_UPDATE_PERMISSIONS = ("compliance.tcs_rule.update", "tcs.rule.update", "tcs.rules.update")
TCS_RULE_DELETE_PERMISSIONS = ("compliance.tcs_rule.delete", "tcs.rule.delete", "tcs.rules.delete")
TCS_PARTY_PROFILE_VIEW_PERMISSIONS = (
    "compliance.tcs_party_profile.view",
    "tcs.party_profile.view",
    "tcs.partyprofile.view",
    "tcs.party_profiles.view",
)
TCS_PARTY_PROFILE_CREATE_PERMISSIONS = (
    "compliance.tcs_party_profile.create",
    "tcs.party_profile.create",
    "tcs.partyprofile.create",
)
TCS_PARTY_PROFILE_UPDATE_PERMISSIONS = (
    "compliance.tcs_party_profile.update",
    "tcs.party_profile.update",
    "tcs.partyprofile.update",
    "tcs.partyprofile.edit",
)
TCS_PARTY_PROFILE_DELETE_PERMISSIONS = (
    "compliance.tcs_party_profile.delete",
    "tcs.party_profile.delete",
    "tcs.partyprofile.delete",
)
TCS_WORKSPACE_VIEW_PERMISSIONS = ("reports.financial_hub.tcs_compliance_center.view", "compliance.tcs_statutory.view", "tcs.menu.access", "tcs.return_27eq.view")
TCS_RETURN_VIEW_PERMISSIONS = ("reports.financial_hub.tcs_compliance_center.view", "compliance.tcs_return_27eq.view", "tcs.return_27eq.view", "compliance.tcs_statutory.view")
TCS_RETURN_FILE_PERMISSIONS = ("compliance.tcs_return_27eq.file", "tcs.return_27eq.view", "compliance.tcs_statutory.view")
TCS_LEDGER_REPORT_VIEW_PERMISSIONS = ("reports.financial_hub.tcs_compliance_center.view", "reports.tcsledgerreport.view", "tcs.ledger_report.view")
TCS_FILING_PACK_VIEW_PERMISSIONS = ("reports.financial_hub.tcs_compliance_center.view", "reports.tcsfilingpack.view", "tcs.filing_pack.view")
WITHHOLDING_READINESS_VIEW_PERMISSIONS = ("purchase.statutory.view", "reports.tds.view")
logger = logging.getLogger(__name__)


def _safe_int(raw):
    if raw in (None, ""):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        raise ValidationError({"detail": "Query parameter must be an integer."})


def _safe_bool(raw) -> bool:
    return str(raw or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _invoice_posting_state(status_value, *, posted_value) -> tuple[bool, str, str]:
    if status_value in (None, ""):
        return False, "unknown", "Posting status unavailable"
    normalized = str(status_value).strip().lower()
    posted_tokens = {
        "posted",
        str(posted_value).strip().lower(),
        str(getattr(posted_value, "value", posted_value)).strip().lower(),
    }
    is_posted = normalized in posted_tokens
    return is_posted, ("posted" if is_posted else "not_posted"), ("Posted" if is_posted else "Invoice not posted")


def _expand_fy_values(raw: str) -> list[str]:
    value = (raw or "").strip()
    if not value:
        return []

    def _token(start_year: int) -> str:
        return f"{start_year}-{str(start_year + 1)[-2:]}"

    out: list[str] = [value]
    m_short = re.match(r"^(\d{4})-(\d{2})$", value)
    if m_short:
        start = int(m_short.group(1))
        end_2 = int(m_short.group(2))
        end_full = (start // 100) * 100 + end_2
        if end_full < start:
            end_full += 100
        if end_full == start + 1:
            token = _token(start)
            if token not in out:
                out.insert(0, token)
            return out
        if end_full > start + 1 and (end_full - start) <= 5:
            candidates = [_token(y) for y in range(start, end_full)]
            return list(dict.fromkeys(candidates + out))
        return out

    m_full = re.match(r"^(\d{4})-(\d{4})$", value)
    if m_full:
        start = int(m_full.group(1))
        end_full = int(m_full.group(2))
        if end_full == start + 1:
            token = _token(start)
            return list(dict.fromkeys([token] + out))
        if end_full > start + 1 and (end_full - start) <= 5:
            candidates = [_token(y) for y in range(start, end_full)]
            return list(dict.fromkeys(candidates + out))
    return out


def _safe_decimal(raw) -> Decimal:
    try:
        return q2(raw or Decimal("0.00"))
    except Exception:
        return Decimal("0.00")


def _non_negative_q2(raw) -> Decimal:
    value = _safe_decimal(raw)
    return value if value > Decimal("0.00") else Decimal("0.00")


def _tcs_search_match(*values, search: str) -> bool:
    token = str(search or "").strip().lower()
    if not token:
        return True
    return any(token in str(value or "").strip().lower() for value in values)


def _request_data(request):
    data = getattr(request, "data", None)
    return data if hasattr(data, "get") else {}


def _entity_id_from_request(request, *, required: bool = True) -> int | None:
    for key in ("entity", "entity_id"):
        raw = request.query_params.get(key)
        if raw in (None, ""):
            raw = _request_data(request).get(key)
        if raw not in (None, ""):
            return _safe_int(raw)
    if required:
        raise ValidationError({"entity": ["entity or entity_id is required."]})
    return None


def _require_tcs_scope_permission(
    *,
    request,
    entity_id: int,
    permission_codes: tuple[str, ...],
    message: str,
    feature_code: str = TCS_FEATURE_CODE,
    access_mode: str = TCS_OPERATIONAL_MODE,
):
    entity = EffectivePermissionService.entity_for_user(request.user, int(entity_id))
    if entity is None:
        raise PermissionDenied("You do not have access to this entity.")

    SubscriptionService.assert_entity_access(
        user=request.user,
        entity=entity,
        access_mode=access_mode,
        feature_code=feature_code,
    )

    current_codes = set(EffectivePermissionService.permission_codes_for_user(request.user, entity.id))
    if permission_codes and not any(code in current_codes for code in permission_codes):
        raise PermissionDenied(message)
    return entity


def _require_tcs_permission_from_request(
    request,
    *,
    permission_codes: tuple[str, ...],
    message: str,
    feature_code: str = TCS_FEATURE_CODE,
    access_mode: str = TCS_OPERATIONAL_MODE,
) -> int:
    entity_id = _entity_id_from_request(request, required=True)
    _require_tcs_scope_permission(
        request=request,
        entity_id=entity_id,
        permission_codes=permission_codes,
        message=message,
        feature_code=feature_code,
        access_mode=access_mode,
    )
    return int(entity_id)


def _tcs_deposit_status_counts_as_deposited(status_value: str | None) -> bool:
    status_token = str(status_value or "").strip().upper()
    return status_token in {TcsDeposit.Status.CONFIRMED, TcsDeposit.Status.FILED}


def _tcs_deposit_status_allows_allocation(status_value: str | None) -> bool:
    status_token = str(status_value or "").strip().upper()
    return status_token == TcsDeposit.Status.CONFIRMED


def _sum_tcs_allocation_rows(allocations, *, deposited_only: bool = False) -> Decimal:
    total = Decimal("0.00")
    for allocation in allocations:
        deposit = getattr(allocation, "deposit", None)
        if deposited_only and not _tcs_deposit_status_counts_as_deposited(getattr(deposit, "status", None)):
            continue
        total += q2(getattr(allocation, "allocated_amount", Decimal("0.00")) or Decimal("0.00"))
    return q2(total)


def _tcs_computation_total_deposited(comp, *, deposited_only: bool = True) -> Decimal:
    total = Decimal("0.00")
    for collection in getattr(comp, "collections", []).all():
        if getattr(collection, "status", None) == TcsCollection.Status.CANCELLED:
            continue
        total += _sum_tcs_allocation_rows(collection.deposit_allocations.all(), deposited_only=deposited_only)
    return q2(total)


def _tcs_return_status_requires_clean_snapshot(status_value: str | None) -> bool:
    status_token = str(status_value or "").strip().upper()
    return status_token in {TcsQuarterlyReturn.Status.VALIDATED, TcsQuarterlyReturn.Status.FILED}


def _is_filed_tcs_return_metadata_update(validated_data) -> bool:
    allowed_fields = {"ack_no", "file_path", "notes", "filed_on", "original_return"}
    return set(validated_data.keys()).issubset(allowed_fields)


def _is_valid_pan(pan: str) -> bool:
    token = (pan or "").strip().upper()
    if not token:
        return False
    return bool(re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", token))


def _fy_quarter_for_doc(doc_date: date | None) -> tuple[str | None, str | None]:
    if not doc_date:
        return None, None
    y = doc_date.year
    if doc_date.month < 4:
        start = y - 1
        end = y
    else:
        start = y
        end = y + 1
    fy = f"{start}-{str(end)[-2:]}"
    if doc_date.month in (4, 5, 6):
        q = "Q1"
    elif doc_date.month in (7, 8, 9):
        q = "Q2"
    elif doc_date.month in (10, 11, 12):
        q = "Q3"
    else:
        q = "Q4"
    return fy, q


def _quarter_boundary_violation(*, doc_date: date | None, fiscal_year: str, quarter: str) -> bool:
    expected_fy, expected_quarter = _fy_quarter_for_doc(doc_date)
    return bool(expected_fy and expected_quarter and (expected_fy != (fiscal_year or "") or expected_quarter != (quarter or "")))


def _zip_csv_payload(named_rows: dict[str, list[dict]]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, rows in named_rows.items():
            stream = io.StringIO()
            headers = list(rows[0].keys()) if rows else []
            writer = csv.DictWriter(stream, fieldnames=headers)
            if headers:
                writer.writeheader()
                writer.writerows(rows)
            zf.writestr(filename, stream.getvalue())
    return buff.getvalue()


def _filing_readiness_errors(snapshot: dict) -> list[str]:
    counts = snapshot.get("counts") if isinstance(snapshot, dict) else {}
    totals = snapshot.get("totals") if isinstance(snapshot, dict) else {}
    errors: list[str] = []

    blocking_count_keys = {
        "missing_pan": "Missing PAN rows exist.",
        "missing_section": "Rows with missing section exist.",
        "not_collected": "Some computations are not collected.",
        "not_deposited": "Some collections are not deposited.",
        "partially_allocated": "Some collections are partially allocated.",
        "deposit_mismatch": "Collection vs deposit allocation mismatch exists.",
    }
    for key, msg in blocking_count_keys.items():
        if int(counts.get(key) or 0) > 0:
            errors.append(msg)

    pending_collection = _safe_decimal(totals.get("pending_collection"))
    pending_deposit = _safe_decimal(totals.get("pending_deposit"))
    if pending_collection > Decimal("0.00"):
        errors.append("Pending collection must be zero before marking FILED.")
    if pending_deposit > Decimal("0.00"):
        errors.append("Pending deposit must be zero before marking FILED.")
    return errors


def _exclude_cancelled_documents(qs):
    """
    Exclude computations whose backing business document is cancelled.
    This keeps ledger reports aligned with active statutory exposure.
    """
    try:
        from sales.models.sales_core import SalesInvoiceHeader
    except Exception:
        return qs

    sales_doc_types = {"invoice", "credit_note", "debit_note"}
    cancelled_sales_ids = list(
        SalesInvoiceHeader.objects.filter(status=SalesInvoiceHeader.Status.CANCELLED).values_list("id", flat=True)
    )
    if cancelled_sales_ids:
        qs = qs.exclude(
            module_name="sales",
            document_type__in=sales_doc_types,
            document_id__in=cancelled_sales_ids,
        )
    return qs


def _resolve_period_bounds(*, fy: str, quarter: str, from_date_raw: str | None, to_date_raw: str | None) -> tuple[date | None, date | None]:
    explicit_from = parse_date(str(from_date_raw or "").strip()) if from_date_raw else None
    explicit_to = parse_date(str(to_date_raw or "").strip()) if to_date_raw else None
    if explicit_from or explicit_to:
        return explicit_from, explicit_to

    fy_tokens = _expand_fy_values(fy)
    token = next((tok for tok in fy_tokens if re.match(r"^\d{4}-\d{2}$", tok or "")), None)
    if not token:
        return None, None
    start_year = int(token[:4])
    fy_start = date(start_year, 4, 1)
    fy_end = date(start_year + 1, 3, 31)
    qtr = (quarter or "").strip().upper()
    quarter_windows = {
        "Q1": (date(start_year, 4, 1), date(start_year, 6, 30)),
        "Q2": (date(start_year, 7, 1), date(start_year, 9, 30)),
        "Q3": (date(start_year, 10, 1), date(start_year, 12, 31)),
        "Q4": (date(start_year + 1, 1, 1), date(start_year + 1, 3, 31)),
    }
    return quarter_windows.get(qtr, (fy_start, fy_end))


def _runtime_quality_flags(*, section_code: str, reason_code: str, pan: str, tax_identifier: str, residency_status: str) -> dict[str, bool]:
    code = (section_code or "").strip().upper()
    reason = (reason_code or "").strip().upper()
    residency = (residency_status or "").strip().lower()
    needs_pan = code in {"194A", "194N"}
    needs_tax_id = code == "195"
    return {
        "missing_pan": bool(needs_pan and not pan),
        "missing_tax_id": bool(needs_tax_id and not tax_identifier),
        "residency_mismatch": bool((code == "195" and residency and residency != "non_resident") or (code in {"194A", "194N"} and residency == "non_resident")),
        "invalid_base_rule": reason == "INVALID_BASE_RULE",
        "missing_section": not code,
    }


def _tcs_runtime_quality_flags(*, section, reason_code: str, pan: str) -> dict[str, bool]:
    reason = (reason_code or "").strip().upper()
    requires_pan = bool(getattr(section, "requires_pan", False)) if section is not None else False
    return {
        "missing_pan": bool(requires_pan and not pan),
        "missing_tax_id": False,
        "residency_mismatch": False,
        "invalid_base_rule": reason == "INVALID_BASE_RULE",
        "missing_section": section is None,
    }


def _tcs_threshold_state(*, section, reason_code: str, computed_tcs: Decimal) -> str:
    section_code = str(getattr(section, "section_code", "") or "").strip().upper()
    if section_code not in {"206C(1H)", "206C1H"}:
        return "not_applicable"

    code = (reason_code or "").strip().upper()
    if code == "BELOW_THRESHOLD_CUMULATIVE":
        return "not_crossed"
    if code == "THRESHOLD_CROSSED_CUMULATIVE":
        return "crossed_in_current_txn"
    if code == "THRESHOLD_ALREADY_CROSSED":
        return "already_crossed"
    if q2(computed_tcs) > Decimal("0.00"):
        return "applicable"
    return "unknown"


def _titleize_token(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    return raw.replace("_", " ").lower().title()


def _tcs_doc_impact_type(*, document_type: str | None, trigger_basis: str | None) -> str:
    trigger = str(trigger_basis or "").strip().upper()
    if trigger == "RECEIPT":
        return "Advance Receipt"
    dtype = str(document_type or "").strip().lower()
    if dtype == "credit_note":
        return "Credit Note"
    if dtype == "debit_note":
        return "Debit Note"
    return "Invoice"


def _tcs_filing_pack_exception_flags(
    *,
    comp_tcs: Decimal,
    comp_collected_total: Decimal,
    comp_alloc_total: Decimal,
    runtime_flags: dict[str, bool],
    invalid_pan_format: bool,
    quarter_boundary_violation: bool,
    is_reversal: bool,
) -> dict[str, bool]:
    effective_exposure = bool(
        q2(comp_tcs) > Decimal("0.00")
        or q2(comp_collected_total) > Decimal("0.00")
        or q2(comp_alloc_total) > Decimal("0.00")
    )
    if not effective_exposure:
        return {
            "missing_pan": False,
            "invalid_pan_format": False,
            "missing_tax_id": False,
            "residency_mismatch": False,
            "missing_section": False,
            "not_collected": False,
            "not_deposited": False,
            "partially_allocated": False,
            "deposit_mismatch": False,
            "quarter_boundary_violation": False,
            "reversal_case": False,
        }

    collected = q2(comp_collected_total)
    allocated = q2(comp_alloc_total)
    return {
        "missing_pan": bool(runtime_flags["missing_pan"]),
        "invalid_pan_format": bool(invalid_pan_format),
        "missing_tax_id": bool(runtime_flags["missing_tax_id"]),
        "residency_mismatch": bool(runtime_flags["residency_mismatch"]),
        "missing_section": bool(runtime_flags["missing_section"]),
        "not_collected": collected <= Decimal("0.00"),
        "not_deposited": collected > Decimal("0.00") and allocated <= Decimal("0.00"),
        "partially_allocated": collected > Decimal("0.00") and allocated > Decimal("0.00") and allocated < collected,
        "deposit_mismatch": collected > Decimal("0.00") and allocated != collected,
        "quarter_boundary_violation": bool(quarter_boundary_violation),
        "reversal_case": bool(is_reversal),
    }


def _row_readiness_status(*, amount: Decimal, flags: dict[str, bool]) -> str:
    if amount <= Decimal("0.00"):
        return "fix_now"
    blocking_keys = {"missing_tax_id", "residency_mismatch", "invalid_base_rule", "missing_section"}
    if any(bool(flags.get(k)) for k in blocking_keys):
        return "blocked"
    if bool(flags.get("missing_pan")):
        return "fix_now"
    return "ready_to_file"


def _tcs_source_route(module_name: str | None, document_type: str | None) -> str | None:
    return _tcs_source_route_for_document(module_name=module_name, document_type=document_type, document_id=None)


def _tcs_source_route_for_document(module_name: str | None, document_type: str | None, document_id: int | None) -> str | None:
    module = str(module_name or "").strip().lower()
    dtype = str(document_type or "").strip().lower()
    if module == "sales" and dtype in {"invoice", "credit_note", "debit_note"}:
        doc_id = _safe_int(document_id)
        if doc_id and SalesInvoiceLine.objects.filter(header_id=doc_id, is_service=True).exists():
            return "/saleserviceinvoice"
        return "/saleinvoice"
    if module == "purchase" and dtype in {"invoice", "credit_note", "debit_note"}:
        doc_id = _safe_int(document_id)
        if doc_id and PurchaseInvoiceLine.objects.filter(header_id=doc_id, is_service=True).exists():
            return "/purchaseserviceinvoice"
        return "/purchaseinvoice"
    return None


def _tcs_posting_lookup_document_type(module_name: str | None, document_type: str | None) -> str | None:
    module = str(module_name or "").strip().lower()
    dtype = str(document_type or "").strip().lower()
    if module == "sales":
        if dtype == "invoice":
            return "sales_invoice"
        if dtype == "credit_note":
            return "sales_credit_note"
        if dtype == "debit_note":
            return "sales_debit_note"
    if module == "purchase":
        if dtype == "invoice":
            return "purchase_invoice"
        if dtype == "credit_note":
            return "purchase_credit_note"
        if dtype == "debit_note":
            return "purchase_debit_note"
    return None


def _party_master_drilldown(*, party_account_id: int | None, entity_id: int | None, subentity_id: int | None, source: str) -> dict | None:
    party_id = _safe_int(party_account_id)
    if not party_id:
        return None
    params = {"source": source}
    if entity_id:
        params["entity"] = int(entity_id)
    if subentity_id:
        params["subentity_id"] = int(subentity_id)
    return {
        "target": "party_master",
        "label": "Open account master",
        "kind": "master_edit",
        "route": f"/financialmaster/accounts/{int(party_id)}/edit",
        "route_name": "financial-master-accounts",
        "params": params,
    }


class TcsSectionListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = WithholdingSectionSerializer

    def get_queryset(self):
        permission_codes = TCS_RULE_VIEW_PERMISSIONS if "rules" in (self.request.path or "") else TCS_SECTION_VIEW_PERMISSIONS
        message = "Missing permission to view TCS rules." if "rules" in (self.request.path or "") else "Missing permission to view TCS sections."
        _require_tcs_permission_from_request(
            self.request,
            permission_codes=permission_codes,
            message=message,
        )
        qs = WithholdingSection.objects.filter(tax_type=2).order_by("section_code", "-effective_from")
        q = (self.request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(section_code__icontains=q) | Q(description__icontains=q))
        law_type = (self.request.query_params.get("law_type") or "").strip().upper()
        if law_type:
            qs = qs.filter(law_type=law_type)
        return qs

    def create(self, request, *args, **kwargs):
        permission_codes = TCS_RULE_CREATE_PERMISSIONS if "rules" in (request.path or "") else TCS_SECTION_CREATE_PERMISSIONS
        message = "Missing permission to create TCS rules." if "rules" in (request.path or "") else "Missing permission to create TCS sections."
        _require_tcs_permission_from_request(request, permission_codes=permission_codes, message=message)
        return super().create(request, *args, **kwargs)


class TcsSectionRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = WithholdingSectionSerializer
    queryset = WithholdingSection.objects.filter(tax_type=2)

    def retrieve(self, request, *args, **kwargs):
        permission_codes = TCS_RULE_VIEW_PERMISSIONS if "rules" in (request.path or "") else TCS_SECTION_VIEW_PERMISSIONS
        message = "Missing permission to view TCS rules." if "rules" in (request.path or "") else "Missing permission to view TCS sections."
        _require_tcs_permission_from_request(request, permission_codes=permission_codes, message=message)
        return super().retrieve(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        permission_codes = TCS_RULE_UPDATE_PERMISSIONS if "rules" in (request.path or "") else TCS_SECTION_UPDATE_PERMISSIONS
        message = "Missing permission to update TCS rules." if "rules" in (request.path or "") else "Missing permission to update TCS sections."
        _require_tcs_permission_from_request(request, permission_codes=permission_codes, message=message)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        permission_codes = TCS_RULE_DELETE_PERMISSIONS if "rules" in (request.path or "") else TCS_SECTION_DELETE_PERMISSIONS
        message = "Missing permission to delete TCS rules." if "rules" in (request.path or "") else "Missing permission to delete TCS sections."
        _require_tcs_permission_from_request(request, permission_codes=permission_codes, message=message)
        return super().destroy(request, *args, **kwargs)

    def perform_destroy(self, instance):
        request = self.request
        changed_by = getattr(request, "user", None) if request else None
        if changed_by is not None and not getattr(changed_by, "is_authenticated", False):
            changed_by = None
        snapshot = WithholdingSectionSerializer._policy_snapshot(instance)
        WithholdingSectionPolicyAudit.objects.create(
            section=instance,
            action=WithholdingSectionPolicyAudit.Action.DELETED,
            changed_by=changed_by,
            changed_fields_json=sorted(snapshot.keys()),
            before_snapshot_json=snapshot,
            after_snapshot_json=None,
            source="api",
        )
        instance.delete()


class TcsEntityConfigListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityWithholdingConfigSerializer

    def get_queryset(self):
        from withholding.models import EntityWithholdingConfig

        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_CONFIG_VIEW_PERMISSIONS,
            message="Missing permission to view TCS configuration.",
        )
        qs = EntityWithholdingConfig.objects.all().order_by("-effective_from", "-id")
        qs = qs.filter(entity_id=entity_id)
        entityfin_id = self.request.query_params.get("entityfin_id")
        if entityfin_id not in (None, ""):
            qs = qs.filter(entityfin_id=int(entityfin_id))
        subentity_id = self.request.query_params.get("subentity_id")
        if subentity_id not in (None, ""):
            qs = qs.filter(subentity_id=int(subentity_id))
        return qs

    def create(self, request, *args, **kwargs):
        entity_id = _entity_id_from_request(request, required=True)
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_CONFIG_MANAGE_PERMISSIONS,
            message="Missing permission to manage TCS configuration.",
        )
        return super().create(request, *args, **kwargs)


class TcsEntityConfigRetrieveUpdateAPIView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityWithholdingConfigSerializer

    def get_queryset(self):
        from withholding.models import EntityWithholdingConfig

        return EntityWithholdingConfig.objects.all()


class TcsEntityConfigRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityWithholdingConfigSerializer

    def get_queryset(self):
        from withholding.models import EntityWithholdingConfig

        return EntityWithholdingConfig.objects.all()

    def get_object(self):
        obj = super().get_object()
        if self.request.method == "GET":
            permission_codes = TCS_CONFIG_VIEW_PERMISSIONS
            message = "Missing permission to view TCS configuration."
        elif self.request.method == "DELETE":
            permission_codes = TCS_CONFIG_DELETE_PERMISSIONS
            message = "Missing permission to delete TCS configuration."
        else:
            permission_codes = TCS_CONFIG_MANAGE_PERMISSIONS
            message = "Missing permission to manage TCS configuration."
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.entity_id,
            permission_codes=permission_codes,
            message=message,
        )
        return obj

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        _require_tcs_scope_permission(
            request=request,
            entity_id=obj.entity_id,
            permission_codes=TCS_CONFIG_DELETE_PERMISSIONS,
            message="Missing permission to delete TCS configuration.",
        )
        return super().destroy(request, *args, **kwargs)


class TcsPartyProfileListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PartyTaxProfileSerializer
    queryset = PartyTaxProfile.objects.all().order_by("-id")

    def get_queryset(self):
        _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_PARTY_PROFILE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS party profiles.",
        )
        return super().get_queryset()

    def create(self, request, *args, **kwargs):
        _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_PARTY_PROFILE_CREATE_PERMISSIONS,
            message="Missing permission to create TCS party profiles.",
        )
        return super().create(request, *args, **kwargs)


class TcsPartyProfileRetrieveUpdateAPIView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PartyTaxProfileSerializer
    queryset = PartyTaxProfile.objects.all()

    def retrieve(self, request, *args, **kwargs):
        _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_PARTY_PROFILE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS party profiles.",
        )
        return super().retrieve(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_PARTY_PROFILE_UPDATE_PERMISSIONS,
            message="Missing permission to update TCS party profiles.",
        )
        return super().update(request, *args, **kwargs)


class TcsPartyProfileRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PartyTaxProfileSerializer
    queryset = PartyTaxProfile.objects.all()

    def retrieve(self, request, *args, **kwargs):
        _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_PARTY_PROFILE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS party profiles.",
        )
        return super().retrieve(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_PARTY_PROFILE_UPDATE_PERMISSIONS,
            message="Missing permission to update TCS party profiles.",
        )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_PARTY_PROFILE_DELETE_PERMISSIONS,
            message="Missing permission to delete TCS party profiles.",
        )
        return super().destroy(request, *args, **kwargs)


class WithholdingEntityPartyProfileListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityPartyTaxProfileSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_PARTY_PROFILE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS party profiles.",
        )
        qs = EntityPartyTaxProfile.objects.all().order_by("-id")
        qs = qs.filter(entity_id=entity_id)
        subentity_id = self.request.query_params.get("subentity_id")
        if subentity_id not in (None, ""):
            parsed_sub = _safe_int(subentity_id)
            if parsed_sub is None:
                qs = qs.filter(subentity__isnull=True)
            else:
                qs = qs.filter(subentity_id=parsed_sub)
        party_account_id = _safe_int(self.request.query_params.get("party_account_id"))
        if party_account_id:
            qs = qs.filter(party_account_id=party_account_id)
        is_active = self.request.query_params.get("is_active")
        if is_active not in (None, ""):
            qs = qs.filter(is_active=_safe_bool(is_active))
        return qs

    def create(self, request, *args, **kwargs):
        entity_id = _entity_id_from_request(request, required=True)
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_PARTY_PROFILE_CREATE_PERMISSIONS,
            message="Missing permission to create TCS party profiles.",
        )
        return super().create(request, *args, **kwargs)


class WithholdingSectionCatalogListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = WithholdingSectionSerializer

    def get_queryset(self):
        qs = WithholdingSection.objects.all().order_by("tax_type", "section_code", "-effective_from")
        q = (self.request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(section_code__icontains=q) | Q(description__icontains=q))
        tax_type = _safe_int(self.request.query_params.get("tax_type"))
        if tax_type is not None:
            qs = qs.filter(tax_type=tax_type)
        law_type = (self.request.query_params.get("law_type") or "").strip().upper()
        if law_type:
            qs = qs.filter(law_type=law_type)
        active_only = self.request.query_params.get("active_only")
        if active_only in (None, ""):
            active_only = "true"
        if _safe_bool(active_only):
            qs = qs.filter(is_active=True)
        return qs


class WithholdingSectionPolicyAuditListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = WithholdingSectionPolicyAuditSerializer

    def get_queryset(self):
        qs = WithholdingSectionPolicyAudit.objects.select_related("section", "changed_by").all().order_by("-created_at", "-id")
        section_id = _safe_int(self.request.query_params.get("section_id"))
        if section_id:
            qs = qs.filter(section_id=section_id)
        tax_type = _safe_int(self.request.query_params.get("tax_type"))
        if tax_type is not None:
            qs = qs.filter(section__tax_type=tax_type)
        action = (self.request.query_params.get("action") or "").strip().upper()
        if action:
            qs = qs.filter(action=action)
        return qs


class WithholdingEntityPartyProfileRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityPartyTaxProfileSerializer
    queryset = EntityPartyTaxProfile.objects.all()

    def get_object(self):
        obj = super().get_object()
        if self.request.method == "GET":
            permission_codes = TCS_PARTY_PROFILE_VIEW_PERMISSIONS
            message = "Missing permission to view TCS party profiles."
        elif self.request.method == "DELETE":
            permission_codes = TCS_PARTY_PROFILE_DELETE_PERMISSIONS
            message = "Missing permission to delete TCS party profiles."
        else:
            permission_codes = TCS_PARTY_PROFILE_UPDATE_PERMISSIONS
            message = "Missing permission to update TCS party profiles."
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.entity_id,
            permission_codes=permission_codes,
            message=message,
        )
        return obj

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        _require_tcs_scope_permission(
            request=request,
            entity_id=obj.entity_id,
            permission_codes=TCS_PARTY_PROFILE_DELETE_PERMISSIONS,
            message="Missing permission to delete TCS party profiles.",
        )
        return super().destroy(request, *args, **kwargs)


class WithholdingSectionPostingMapListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityWithholdingSectionPostingMapSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_CONFIG_VIEW_PERMISSIONS,
            message="Missing permission to view TCS posting maps.",
        )
        qs = EntityWithholdingSectionPostingMap.objects.all().order_by("-effective_from", "-id")
        qs = qs.filter(entity_id=entity_id)
        subentity_id = self.request.query_params.get("subentity_id")
        if subentity_id not in (None, ""):
            parsed_sub = _safe_int(subentity_id)
            if parsed_sub is None:
                qs = qs.filter(subentity__isnull=True)
            else:
                qs = qs.filter(subentity_id=parsed_sub)
        section_id = _safe_int(self.request.query_params.get("section_id"))
        if section_id:
            qs = qs.filter(section_id=section_id)
        is_active = self.request.query_params.get("is_active")
        if is_active not in (None, ""):
            qs = qs.filter(is_active=_safe_bool(is_active))
        return qs

    def create(self, request, *args, **kwargs):
        entity_id = _entity_id_from_request(request, required=True)
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_CONFIG_MANAGE_PERMISSIONS,
            message="Missing permission to manage TCS posting maps.",
        )
        return super().create(request, *args, **kwargs)


class WithholdingSectionPostingMapRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityWithholdingSectionPostingMapSerializer
    queryset = EntityWithholdingSectionPostingMap.objects.all()

    def get_object(self):
        obj = super().get_object()
        if self.request.method == "GET":
            permission_codes = TCS_CONFIG_VIEW_PERMISSIONS
            message = "Missing permission to view TCS posting maps."
        elif self.request.method == "DELETE":
            permission_codes = TCS_CONFIG_DELETE_PERMISSIONS
            message = "Missing permission to delete TCS posting maps."
        else:
            permission_codes = TCS_CONFIG_MANAGE_PERMISSIONS
            message = "Missing permission to manage TCS posting maps."
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.entity_id,
            permission_codes=permission_codes,
            message=message,
        )
        return obj

    def destroy(self, request, *args, **kwargs):
        obj = self.get_object()
        _require_tcs_scope_permission(
            request=request,
            entity_id=obj.entity_id,
            permission_codes=TCS_CONFIG_DELETE_PERMISSIONS,
            message="Missing permission to delete TCS posting maps.",
        )
        return super().destroy(request, *args, **kwargs)


class WithholdingTcsThresholdOpeningListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityTcsThresholdOpeningSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS threshold openings.",
        )
        qs = EntityTcsThresholdOpening.objects.all().order_by("-effective_from", "-id")
        qs = qs.filter(entity_id=entity_id)
        entityfin_id = _safe_int(self.request.query_params.get("entityfin_id"))
        if entityfin_id is not None:
            qs = qs.filter(entityfin_id=entityfin_id)
        subentity_id = self.request.query_params.get("subentity_id")
        if subentity_id not in (None, ""):
            parsed_sub = _safe_int(subentity_id)
            if parsed_sub is None:
                qs = qs.filter(subentity__isnull=True)
            else:
                qs = qs.filter(subentity_id=parsed_sub)
        party_account_id = _safe_int(self.request.query_params.get("party_account_id"))
        if party_account_id is not None:
            qs = qs.filter(party_account_id=party_account_id)
        section_id = _safe_int(self.request.query_params.get("section_id"))
        if section_id is not None:
            qs = qs.filter(section_id=section_id)
        is_active = self.request.query_params.get("is_active")
        if is_active not in (None, ""):
            qs = qs.filter(is_active=_safe_bool(is_active))
        return qs

    def create(self, request, *args, **kwargs):
        entity_id = _entity_id_from_request(request, required=True)
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to create TCS threshold openings.",
        )
        return super().create(request, *args, **kwargs)


class WithholdingTcsThresholdOpeningRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EntityTcsThresholdOpeningSerializer
    queryset = EntityTcsThresholdOpening.objects.all()

    def get_object(self):
        obj = super().get_object()
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to manage TCS threshold openings.",
        )
        return obj


class TcsComputePreviewAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        s = TcsComputeRequestSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        d = s.validated_data
        _require_tcs_scope_permission(
            request=request,
            entity_id=d["entity_id"],
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to access the TCS statutory workspace.",
        )

        req = {
            "entity_id": d["entity_id"],
            "entityfin_id": d["entityfin_id"],
            "subentity_id": d.get("subentity_id"),
            "party_account_id": d.get("party_account_id"),
            "tax_type": d["tax_type"],
            "explicit_section_id": d.get("section_id"),
            "doc_date": d["doc_date"],
            "taxable_total": d.get("taxable_total", Decimal("0.00")),
            "gross_total": d.get("gross_total", Decimal("0.00")),
        }
        return Response(build_preview_payload(req=req, user=request.user))


class TcsComputeConfirmAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        s = TcsComputeConfirmSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        d = s.validated_data
        _require_tcs_scope_permission(
            request=request,
            entity_id=d["entity_id"],
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to manage the TCS statutory workspace.",
        )

        req = {
            "entity_id": d["entity_id"],
            "entityfin_id": d["entityfin_id"],
            "subentity_id": d.get("subentity_id"),
            "party_account_id": d.get("party_account_id"),
            "tax_type": d["tax_type"],
            "explicit_section_id": d.get("section_id"),
            "doc_date": d["doc_date"],
            "taxable_total": d.get("taxable_total", Decimal("0.00")),
            "gross_total": d.get("gross_total", Decimal("0.00")),
        }
        preview = compute_withholding_preview(**req)

        if d["tax_type"] != 2:
            return Response(
                {
                    "preview": build_preview_payload(req=req, user=request.user),
                    "message": "Only TCS persistence is supported in this endpoint.",
                },
                status=status.HTTP_200_OK,
            )

        doc_id = d.get("document_id")
        if not doc_id:
            return Response({"document_id": ["This field is required for confirm."]}, status=status.HTTP_400_BAD_REQUEST)

        row = upsert_tcs_computation(
            module_name=(d.get("module_name") or "sales").strip().lower(),
            document_type=(d.get("document_type") or "invoice").strip().lower(),
            document_id=doc_id,
            document_no=(d.get("document_no") or "").strip(),
            doc_date=d["doc_date"],
            entity_id=d["entity_id"],
            entityfin_id=d["entityfin_id"],
            subentity_id=d.get("subentity_id"),
            party_account_id=d.get("party_account_id"),
            preview=preview,
            status=d.get("status") or TcsComputation.Status.CONFIRMED,
            trigger_basis=d.get("trigger_basis") or "INVOICE",
            override_reason=d.get("override_reason") or "",
            overridden_by=request.user,
        )
        return Response(TcsComputationSerializer(row).data, status=status.HTTP_201_CREATED)


class TcsComputeRecomputeAPIView(TcsComputeConfirmAPIView):
    pass


class TcsComputationListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsComputationSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to access TCS computations.",
        )
        qs = TcsComputation.objects.all().order_by("-doc_date", "-id")
        module = (self.request.query_params.get("module") or "").strip()
        if module:
            qs = qs.filter(module_name=module)
        document_type = (self.request.query_params.get("document_type") or "").strip()
        if document_type:
            qs = qs.filter(document_type=document_type)
        doc_id = self.request.query_params.get("doc_id")
        if doc_id not in (None, ""):
            qs = qs.filter(document_id=int(doc_id))
        qs = qs.filter(entity_id=entity_id)
        fy = (self.request.query_params.get("fy") or "").strip()
        if fy:
            qs = qs.filter(fiscal_year__in=_expand_fy_values(fy))
        quarter = (self.request.query_params.get("quarter") or "").strip().upper()
        if quarter:
            qs = qs.filter(quarter=quarter)
        return qs


class TcsCollectionListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsCollectionSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS collections.",
        )
        qs = TcsCollection.objects.select_related("computation").all().order_by("-collection_date", "-id")
        qs = qs.filter(computation__entity_id=entity_id)
        fy = (self.request.query_params.get("fy") or "").strip()
        if fy:
            qs = qs.filter(computation__fiscal_year__in=_expand_fy_values(fy))
        quarter = (self.request.query_params.get("quarter") or "").strip().upper()
        if quarter:
            qs = qs.filter(computation__quarter=quarter)
        return qs

    def perform_create(self, serializer):
        row = serializer.save()
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=row.computation.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to create TCS collections.",
        )


class TcsCollectionRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsCollectionSerializer

    def get_queryset(self):
        qs = TcsCollection.objects.select_related("computation").all()
        return qs

    def get_object(self):
        obj = super().get_object()
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.computation.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to manage TCS collections.",
        )
        return obj


class TcsDepositListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsDepositSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS deposits.",
        )
        qs = TcsDeposit.objects.all().order_by("-challan_date", "-id")
        qs = qs.filter(entity_id=entity_id)
        fy = (self.request.query_params.get("fy") or "").strip()
        if fy:
            qs = qs.filter(financial_year__in=_expand_fy_values(fy))
        month = _safe_int(self.request.query_params.get("month"))
        if month is not None:
            qs = qs.filter(month=month)
        return qs

    def perform_create(self, serializer):
        row = serializer.save(deposited_by=self.request.user)
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=row.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to create TCS deposits.",
        )


class TcsDepositRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsDepositSerializer

    def get_queryset(self):
        qs = TcsDeposit.objects.all()
        return qs

    def get_object(self):
        obj = super().get_object()
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to manage TCS deposits.",
        )
        return obj

    def perform_update(self, serializer):
        serializer.save(deposited_by=self.request.user)


class TcsDepositConfirmAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int):
        try:
            deposit = TcsDeposit.objects.get(pk=pk)
        except TcsDeposit.DoesNotExist:
            return Response({"detail": "Deposit not found."}, status=status.HTTP_404_NOT_FOUND)
        _require_tcs_scope_permission(
            request=request,
            entity_id=deposit.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to confirm TCS deposits.",
        )

        if deposit.status == TcsDeposit.Status.FILED:
            return Response({"detail": "Filed deposits cannot be reconfirmed."}, status=status.HTTP_400_BAD_REQUEST)
        if deposit.status != TcsDeposit.Status.CONFIRMED:
            deposit.status = TcsDeposit.Status.CONFIRMED
            deposit.deposited_by = request.user
            deposit.save(update_fields=["status", "deposited_by", "updated_at"])

        return Response(TcsDepositSerializer(deposit).data, status=status.HTTP_200_OK)


class TcsDepositAllocateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int):
        try:
            deposit = TcsDeposit.objects.get(pk=pk)
        except TcsDeposit.DoesNotExist:
            return Response({"detail": "Deposit not found."}, status=status.HTTP_404_NOT_FOUND)
        _require_tcs_scope_permission(
            request=request,
            entity_id=deposit.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to allocate TCS deposits.",
        )

        request_serializer = TcsDepositAllocationRequestSerializer(data=request.data)
        request_serializer.is_valid(raise_exception=True)
        collection_id = request_serializer.validated_data["collection_id"]
        allocated_amount = request_serializer.validated_data["allocated_amount"]

        try:
            collection = TcsCollection.objects.get(pk=collection_id)
        except TcsCollection.DoesNotExist:
            return Response({"detail": "Invalid collection_id."}, status=status.HTTP_400_BAD_REQUEST)

        if not _tcs_deposit_status_allows_allocation(deposit.status):
            return Response({"detail": "Allocation is allowed only against confirmed deposits."}, status=status.HTTP_400_BAD_REQUEST)
        if collection.status == TcsCollection.Status.CANCELLED:
            return Response({"detail": "Cannot allocate a cancelled collection."}, status=status.HTTP_400_BAD_REQUEST)
        if int(collection.computation.entity_id) != int(deposit.entity_id):
            return Response({"detail": "Collection and deposit must belong to the same entity."}, status=status.HTTP_400_BAD_REQUEST)
        if (collection.computation.fiscal_year or "").strip() and (deposit.financial_year or "").strip():
            if str(collection.computation.fiscal_year).strip() != str(deposit.financial_year).strip():
                return Response({"detail": "Collection and deposit financial year mismatch."}, status=status.HTTP_400_BAD_REQUEST)

        alloc_amount = q2(allocated_amount)
        if alloc_amount <= Decimal("0.00"):
            return Response({"detail": "allocated_amount must be > 0."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            locked_deposit = TcsDeposit.objects.select_for_update().get(pk=deposit.pk)
            total_alloc = (
                TcsDepositAllocation.objects.filter(deposit=locked_deposit)
                .aggregate(v=Sum("allocated_amount"))
                .get("v")
                or Decimal("0.00")
            )
            if q2(total_alloc + alloc_amount) > q2(locked_deposit.total_deposit_amount):
                return Response({"detail": "Allocation exceeds deposit balance."}, status=status.HTTP_400_BAD_REQUEST)

            collection_alloc_total = (
                TcsDepositAllocation.objects.filter(collection=collection)
                .aggregate(v=Sum("allocated_amount"))
                .get("v")
                or Decimal("0.00")
            )
            if q2(collection_alloc_total + alloc_amount) > q2(collection.tcs_collected_amount):
                return Response({"detail": "Allocation exceeds collection amount."}, status=status.HTTP_400_BAD_REQUEST)

            row = TcsDepositAllocation.objects.create(
                deposit=locked_deposit,
                collection=collection,
                allocated_amount=alloc_amount,
            )
            updated_collection_total = q2(collection_alloc_total + alloc_amount)
            next_status = (
                TcsCollection.Status.ALLOCATED
                if updated_collection_total >= q2(collection.tcs_collected_amount)
                else TcsCollection.Status.OPEN
            )
            if collection.status != next_status:
                collection.status = next_status
                collection.save(update_fields=["status", "updated_at"])
        return Response(TcsDepositAllocationSerializer(row).data, status=status.HTTP_201_CREATED)


class TcsDepositAllocationListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsDepositAllocationSerializer

    def get_queryset(self):
        deposit = TcsDeposit.objects.filter(pk=self.kwargs["pk"]).only("id", "entity_id").first()
        if deposit is None:
            return TcsDepositAllocation.objects.none()
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=deposit.entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to view TCS deposit allocations.",
        )
        return TcsDepositAllocation.objects.filter(deposit_id=deposit.id).order_by("-id")


class TcsReturn27EqListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsQuarterlyReturnSerializer

    def get_queryset(self):
        entity_id = _require_tcs_permission_from_request(
            self.request,
            permission_codes=TCS_RETURN_VIEW_PERMISSIONS,
            message="Missing permission to view TCS Return 27EQ data.",
        )
        qs = TcsQuarterlyReturn.objects.filter(form_name="27EQ").order_by("-id")
        qs = qs.filter(entity_id=entity_id)
        fy = (self.request.query_params.get("fy") or "").strip()
        if fy:
            qs = qs.filter(fy__in=_expand_fy_values(fy))
        quarter = (self.request.query_params.get("quarter") or "").strip().upper()
        if quarter:
            qs = qs.filter(quarter=quarter)
        return qs

    def perform_create(self, serializer):
        entity = serializer.validated_data.get("entity")
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=entity.id,
            permission_codes=TCS_RETURN_FILE_PERMISSIONS,
            message="Missing permission to create or file TCS Return 27EQ.",
        )
        fy = (serializer.validated_data.get("fy") or "").strip()
        quarter = (serializer.validated_data.get("quarter") or "").strip().upper()
        status_value = serializer.validated_data.get("status") or TcsQuarterlyReturn.Status.DRAFT
        snapshot = serializer.validated_data.get("json_snapshot")
        if (_tcs_return_status_requires_clean_snapshot(status_value) and entity and fy and quarter) or (not snapshot and entity and fy and quarter):
            snapshot = _build_tcs_27eq_snapshot(entity_id=int(entity.id), fy=fy, quarter=quarter)
        if _tcs_return_status_requires_clean_snapshot(status_value):
            filing_errors = _filing_readiness_errors(snapshot or {})
            if filing_errors:
                raise ValidationError({"status": filing_errors})
        serializer.save(form_name="27EQ", json_snapshot=snapshot)


class TcsReturn27EqRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TcsQuarterlyReturnSerializer

    def get_queryset(self):
        return TcsQuarterlyReturn.objects.filter(form_name="27EQ")

    def get_object(self):
        obj = super().get_object()
        permission_codes = TCS_RETURN_VIEW_PERMISSIONS if self.request.method == "GET" else TCS_RETURN_FILE_PERMISSIONS
        message = "Missing permission to view TCS Return 27EQ data." if self.request.method == "GET" else "Missing permission to update or file TCS Return 27EQ."
        _require_tcs_scope_permission(
            request=self.request,
            entity_id=obj.entity_id,
            permission_codes=permission_codes,
            message=message,
        )
        return obj

    def perform_update(self, serializer):
        is_filed_metadata_update = (
            serializer.instance.status == TcsQuarterlyReturn.Status.FILED
            and _is_filed_tcs_return_metadata_update(serializer.validated_data)
        )
        if serializer.instance.status == TcsQuarterlyReturn.Status.FILED and not is_filed_metadata_update:
            raise ValidationError({"status": ["Filed returns cannot be edited. Create a correction return instead."]})
        entity = serializer.validated_data.get("entity") or serializer.instance.entity
        fy = (serializer.validated_data.get("fy") or serializer.instance.fy or "").strip()
        quarter = (serializer.validated_data.get("quarter") or serializer.instance.quarter or "").strip().upper()
        status_value = serializer.validated_data.get("status") or serializer.instance.status
        snapshot = serializer.validated_data.get("json_snapshot")
        if is_filed_metadata_update:
            serializer.save(form_name="27EQ")
            return
        if (_tcs_return_status_requires_clean_snapshot(status_value) and entity and fy and quarter) or (not snapshot and entity and fy and quarter):
            snapshot = _build_tcs_27eq_snapshot(entity_id=int(entity.id), fy=fy, quarter=quarter)
        if _tcs_return_status_requires_clean_snapshot(status_value):
            filing_errors = _filing_readiness_errors(snapshot or {})
            if filing_errors:
                raise ValidationError({"status": filing_errors})
        serializer.save(form_name="27EQ", json_snapshot=snapshot)

    def perform_destroy(self, instance):
        if instance.status == TcsQuarterlyReturn.Status.FILED:
            raise ValidationError({"status": ["Filed returns cannot be deleted. Create a correction return instead."]})
        instance.delete()


class TcsReturn27EqPrefillAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        entity_id = _safe_int(request.query_params.get("entity_id"))
        fy = (request.query_params.get("fy") or "").strip()
        quarter = (request.query_params.get("quarter") or "").strip().upper()
        if entity_id is None:
            raise ValidationError({"entity_id": ["This query param is required."]})
        if not fy:
            raise ValidationError({"fy": ["This query param is required."]})
        if quarter not in {"Q1", "Q2", "Q3", "Q4"}:
            raise ValidationError({"quarter": ["quarter must be one of Q1/Q2/Q3/Q4."]})
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_RETURN_VIEW_PERMISSIONS,
            message="Missing permission to view TCS Return 27EQ prefill data.",
        )

        snapshot = _build_tcs_27eq_snapshot(entity_id=entity_id, fy=fy, quarter=quarter)
        existing = (
            TcsQuarterlyReturn.objects.filter(entity_id=entity_id, fy__in=_expand_fy_values(fy), quarter=quarter, form_name="27EQ")
            .order_by("-id")
            .first()
        )
        return Response(
            {
                "entity_id": entity_id,
                "fy": fy,
                "quarter": quarter,
                "snapshot": snapshot,
                "existing_return": TcsQuarterlyReturnSerializer(existing).data if existing else None,
            }
        )


class TcsReportLedgerAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        entity_id = _require_tcs_permission_from_request(
            request,
            permission_codes=TCS_LEDGER_REPORT_VIEW_PERMISSIONS,
            message="Missing permission to view the TCS ledger report.",
            feature_code=TCS_REPORTING_FEATURE_CODE,
        )
        qs = TcsComputation.objects.all()
        qs = qs.filter(entity_id=entity_id)
        fy = (request.query_params.get("fy") or "").strip()
        if fy:
            qs = qs.filter(fiscal_year__in=_expand_fy_values(fy))
        quarter = (request.query_params.get("quarter") or "").strip().upper()
        if quarter:
            qs = qs.filter(quarter=quarter)
        include_reversed = _safe_bool(request.query_params.get("include_reversed"))
        if not include_reversed:
            qs = qs.exclude(status=TcsComputation.Status.REVERSED)
        include_draft = _safe_bool(request.query_params.get("include_draft"))
        if not include_draft:
            qs = qs.exclude(status=TcsComputation.Status.DRAFT)
        include_cancelled = _safe_bool(request.query_params.get("include_cancelled"))
        if not include_cancelled:
            qs = _exclude_cancelled_documents(qs)

        qs = qs.annotate(section_code_norm=Coalesce("section__section_code", Value("UNMAPPED")))
        out = qs.values("section_code_norm").annotate(
            doc_count=Count("id"),
            total_base=Sum("tcs_base_amount"),
            total_tcs=Sum("tcs_amount"),
        ).order_by("section_code_norm")

        return Response(list(out))


class TcsReportLedgerDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        entity_id = _safe_int(request.query_params.get("entity_id"))
        if entity_id is None:
            raise ValidationError({"entity_id": ["This query param is required."]})
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_LEDGER_REPORT_VIEW_PERMISSIONS,
            message="Missing permission to view the TCS ledger report.",
            feature_code=TCS_REPORTING_FEATURE_CODE,
        )

        section_code = (request.query_params.get("section") or "").strip().upper()
        fy = (request.query_params.get("fy") or "").strip()
        quarter = (request.query_params.get("quarter") or "").strip().upper()
        include_reversed = _safe_bool(request.query_params.get("include_reversed"))
        include_draft = _safe_bool(request.query_params.get("include_draft"))
        include_cancelled = _safe_bool(request.query_params.get("include_cancelled"))

        qs = (
            TcsComputation.objects.select_related("party_account", "section")
            .prefetch_related("collections__deposit_allocations__deposit")
            .filter(entity_id=entity_id)
            .order_by("-doc_date", "-id")
        )
        if fy:
            qs = qs.filter(fiscal_year__in=_expand_fy_values(fy))
        if quarter:
            if quarter not in {"Q1", "Q2", "Q3", "Q4"}:
                raise ValidationError({"quarter": ["quarter must be one of Q1/Q2/Q3/Q4."]})
            qs = qs.filter(quarter=quarter)
        if not include_reversed:
            qs = qs.exclude(status=TcsComputation.Status.REVERSED)
        if not include_draft:
            qs = qs.exclude(status=TcsComputation.Status.DRAFT)
        if not include_cancelled:
            qs = _exclude_cancelled_documents(qs)
        if section_code:
            if section_code == "UNMAPPED":
                qs = qs.filter(section__isnull=True)
            else:
                qs = qs.filter(section__section_code__iexact=section_code)

        rows = []
        total_base = Decimal("0.00")
        total_tcs = Decimal("0.00")
        total_collected = Decimal("0.00")
        total_deposited = Decimal("0.00")

        for comp in qs:
            comp_base = q2(comp.tcs_base_amount or Decimal("0.00"))
            comp_tcs = q2(comp.tcs_amount or Decimal("0.00"))
            comp_collected = Decimal("0.00")
            comp_deposited = Decimal("0.00")
            challan_map = {}

            for collection in comp.collections.all():
                if collection.status == TcsCollection.Status.CANCELLED:
                    continue
                comp_collected += q2(collection.tcs_collected_amount or Decimal("0.00"))
                for alloc in collection.deposit_allocations.all():
                    alloc_amt = q2(alloc.allocated_amount or Decimal("0.00"))
                    dep = alloc.deposit
                    counted_alloc_amt = alloc_amt if _tcs_deposit_status_counts_as_deposited(getattr(dep, "status", None)) else Decimal("0.00")
                    comp_deposited += counted_alloc_amt
                    dep = alloc.deposit
                    if dep:
                        key = dep.id
                        if key not in challan_map:
                            challan_map[key] = {
                                "id": dep.id,
                                "challan_no": dep.challan_no,
                                "challan_date": dep.challan_date,
                                "allocated_amount": Decimal("0.00"),
                            }
                        challan_map[key]["allocated_amount"] = q2(challan_map[key]["allocated_amount"] + counted_alloc_amt)

            pending_collection = _non_negative_q2(comp_tcs - comp_collected)
            pending_deposit = _non_negative_q2(comp_collected - comp_deposited)
            challans = [
                {
                    "id": item["id"],
                    "challan_no": item["challan_no"],
                    "challan_date": item["challan_date"],
                    "allocated_amount": q2(item["allocated_amount"]),
                }
                for item in challan_map.values()
            ]

            rows.append(
                {
                    "id": comp.id,
                    "section_code": comp.section.section_code if comp.section else "UNMAPPED",
                    "document_type": comp.document_type,
                    "document_no": comp.document_no,
                    "doc_date": comp.doc_date,
                    "party_account_id": comp.party_account_id,
                    "party_name": getattr(comp.party_account, "accountname", None)
                    or getattr(comp.party_account, "legalname", None)
                    or "",
                    "pan": account_pan(comp.party_account) or getattr(comp.party_account, "pan", None) or "",
                    "status": comp.status,
                    "base_amount": comp_base,
                    "tcs_amount": comp_tcs,
                    "collected_amount": q2(comp_collected),
                    "deposited_amount": q2(comp_deposited),
                    "pending_collection": pending_collection,
                    "pending_deposit": pending_deposit,
                    "challans": challans,
                }
            )

            total_base += comp_base
            total_tcs += comp_tcs
            total_collected += q2(comp_collected)
            total_deposited += q2(comp_deposited)

        return Response(
            {
                "section": section_code or "ALL",
                "rows": rows,
                "summary": {
                    "count": len(rows),
                    "total_base": q2(total_base),
                    "total_tcs": q2(total_tcs),
                    "total_collected": q2(total_collected),
                    "total_deposited": q2(total_deposited),
                    "pending_collection": _non_negative_q2(total_tcs - total_collected),
                    "pending_deposit": _non_negative_q2(total_collected - total_deposited),
                },
            }
        )


class TcsWorkspaceTransactionsAPIView(APIView):
    """
    Operational TCS workspace:
    computation -> collection -> deposit lifecycle at transaction level.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        request_started_at = perf_counter()
        stage_started_at = request_started_at
        stage_timings = {}

        def checkpoint(name):
            nonlocal stage_started_at
            now = perf_counter()
            stage_timings[name] = round((now - stage_started_at) * 1000, 2)
            stage_started_at = now

        entity_id = _safe_int(request.query_params.get("entity_id"))
        if entity_id is None:
            raise ValidationError({"entity_id": ["This query param is required."]})
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to access the TCS statutory workspace.",
        )

        fy = (request.query_params.get("fy") or "").strip()
        quarter = (request.query_params.get("quarter") or "").strip().upper()
        from_date = parse_date((request.query_params.get("from_date") or "").strip()) if request.query_params.get("from_date") else None
        to_date = parse_date((request.query_params.get("to_date") or "").strip()) if request.query_params.get("to_date") else None
        section_code = (request.query_params.get("section") or "").strip().upper()
        customer_id = _safe_int(request.query_params.get("customer_id"))
        customer_q = (request.query_params.get("customer_q") or "").strip()
        search = (request.query_params.get("search") or "").strip()
        include_reversed = _safe_bool(request.query_params.get("include_reversed"))
        include_draft = _safe_bool(request.query_params.get("include_draft"))
        include_cancelled = _safe_bool(request.query_params.get("include_cancelled"))
        checkpoint("scope_parse_ms")

        qs = (
            TcsComputation.objects.select_related("party_account", "section")
            .prefetch_related("collections__deposit_allocations__deposit")
            .filter(entity_id=entity_id)
            .order_by("-doc_date", "-id")
        )

        if fy:
            qs = qs.filter(fiscal_year__in=_expand_fy_values(fy))
        if quarter:
            if quarter not in {"Q1", "Q2", "Q3", "Q4"}:
                raise ValidationError({"quarter": ["quarter must be one of Q1/Q2/Q3/Q4."]})
            qs = qs.filter(quarter=quarter)
        if from_date:
            qs = qs.filter(doc_date__gte=from_date)
        if to_date:
            qs = qs.filter(doc_date__lte=to_date)
        if customer_id is not None:
            qs = qs.filter(party_account_id=customer_id)
        if customer_q:
            qs = qs.filter(
                Q(party_account__accountname__icontains=customer_q)
                | Q(party_account__legalname__icontains=customer_q)
            )
        if section_code:
            if section_code == "UNMAPPED":
                qs = qs.filter(section__isnull=True)
            else:
                qs = qs.filter(section__section_code__iexact=section_code)
        if not include_reversed:
            qs = qs.exclude(status=TcsComputation.Status.REVERSED)
        if not include_draft:
            qs = qs.exclude(status=TcsComputation.Status.DRAFT)
        if not include_cancelled:
            qs = _exclude_cancelled_documents(qs)
        computations = list(qs)
        checkpoint("computations_fetch_ms")

        sales_doc_ids = {
            int(comp.document_id)
            for comp in computations
            if str(comp.module_name or "").strip().lower() == "sales"
            and str(comp.document_type or "").strip().lower() in {"invoice", "credit_note", "debit_note"}
            and _safe_int(comp.document_id)
        }
        purchase_doc_ids = {
            int(comp.document_id)
            for comp in computations
            if str(comp.module_name or "").strip().lower() == "purchase"
            and str(comp.document_type or "").strip().lower() in {"invoice", "credit_note", "debit_note"}
            and _safe_int(comp.document_id)
        }
        sales_status_map = {
            int(row.id): row.status
            for row in SalesInvoiceHeader.objects.filter(id__in=sales_doc_ids).only("id", "status")
        }
        purchase_status_map = {
            int(row.id): row.status
            for row in PurchaseInvoiceHeader.objects.filter(id__in=purchase_doc_ids).only("id", "status")
        }
        checkpoint("document_status_map_ms")

        rows = []
        party_ids = {int(comp.party_account_id) for comp in computations}
        profile_map = {
            row.party_account_id: row
            for row in EntityPartyTaxProfile.objects.filter(entity_id=entity_id, party_account_id__in=party_ids, is_active=True).order_by("-updated_at")
        }
        checkpoint("party_profile_map_ms")
        total_base = Decimal("0.00")
        total_computed = Decimal("0.00")
        total_collected = Decimal("0.00")
        total_deposited = Decimal("0.00")
        quality_counts = {
            "missing_pan": 0,
            "invalid_pan_format": 0,
            "missing_tax_id": 0,
            "residency_mismatch": 0,
            "missing_section": 0,
            "invalid_base_rule": 0,
            "quarter_boundary_violation": 0,
            "incomplete_compliance": 0,
        }
        status_counts = {
            "computed_pending_collection": 0,
            "partially_collected": 0,
            "collected_pending_deposit": 0,
            "deposited": 0,
            "no_computed_tcs": 0,
        }
        pending_row_counts = {
            "pending_collection": 0,
            "pending_deposit": 0,
        }
        threshold_counts = {
            "not_applicable": 0,
            "not_crossed": 0,
            "crossed_in_current_txn": 0,
            "already_crossed": 0,
            "applicable": 0,
            "unknown": 0,
        }
        impact_counts = {
            "invoice": 0,
            "advance_receipt": 0,
            "credit_note": 0,
            "debit_note": 0,
        }
        section_summary = {}

        for comp in computations:
            if not _tcs_search_match(
                comp.document_no,
                getattr(comp.party_account, "legalname", None),
                getattr(comp.party_account, "accountname", None),
                account_pan(comp.party_account),
                getattr(comp.section, "section_code", None),
                comp.document_type,
                getattr(comp, "trigger_basis", None),
                search=search,
            ):
                continue
            comp_base = q2(comp.tcs_base_amount or Decimal("0.00"))
            comp_tcs = q2(comp.tcs_amount or Decimal("0.00"))
            comp_collected = Decimal("0.00")
            comp_deposited = Decimal("0.00")
            collections_payload = []

            for col in comp.collections.all():
                if col.status == TcsCollection.Status.CANCELLED:
                    continue
                col_amt = q2(col.tcs_collected_amount or Decimal("0.00"))
                comp_collected += col_amt
                allocation_rows = []
                col_deposited = Decimal("0.00")
                for alloc in col.deposit_allocations.all():
                    alloc_amt = q2(alloc.allocated_amount or Decimal("0.00"))
                    dep = alloc.deposit
                    counted_alloc_amt = alloc_amt if _tcs_deposit_status_counts_as_deposited(getattr(dep, "status", None)) else Decimal("0.00")
                    col_deposited += counted_alloc_amt
                    comp_deposited += counted_alloc_amt
                    allocation_rows.append(
                        {
                            "id": alloc.id,
                            "allocated_amount": alloc_amt,
                            "deposit_id": dep.id if dep else None,
                            "challan_no": dep.challan_no if dep else "",
                            "challan_date": dep.challan_date if dep else None,
                            "deposit_status": dep.status if dep else "",
                        }
                    )

                collections_payload.append(
                    {
                        "id": col.id,
                        "collection_date": col.collection_date,
                        "receipt_voucher_id": col.receipt_voucher_id,
                        "amount_received": q2(col.amount_received or Decimal("0.00")),
                        "tcs_collected_amount": col_amt,
                        "status": col.status,
                        "collection_reference": col.collection_reference,
                        "deposited_amount": q2(col_deposited),
                        "allocations": allocation_rows,
                    }
                )

            pending_collection = _non_negative_q2(comp_tcs - comp_collected)
            pending_deposit = _non_negative_q2(comp_collected - comp_deposited)
            pan_token = (account_pan(comp.party_account) or getattr(comp.party_account, "pan", None) or "").strip().upper()
            has_missing_pan = not bool(pan_token)
            has_invalid_pan_format = bool(pan_token) and not _is_valid_pan(pan_token)
            has_missing_section = comp.section_id is None
            sec_code = comp.section.section_code if comp.section else "UNMAPPED"
            has_quarter_violation = _quarter_boundary_violation(
                doc_date=comp.doc_date,
                fiscal_year=comp.fiscal_year or "",
                quarter=comp.quarter or "",
            )
            reason_code = str(
                (comp.computation_json or {}).get("reason_code")
                or (comp.rule_snapshot_json or {}).get("reason_code")
                or ""
            ).strip().upper()
            runtime_flags = _tcs_runtime_quality_flags(
                section=comp.section,
                reason_code=reason_code,
                pan=pan_token,
            )
            threshold_state = _tcs_threshold_state(
                section=comp.section,
                reason_code=reason_code,
                computed_tcs=comp_tcs,
            )
            doc_impact_type = _tcs_doc_impact_type(
                document_type=comp.document_type,
                trigger_basis=getattr(comp, "trigger_basis", ""),
            )
            impact_key = str(doc_impact_type or "").strip().lower().replace(" ", "_")
            has_missing_tax_id = bool(runtime_flags["missing_tax_id"])
            has_residency_mismatch = bool(runtime_flags["residency_mismatch"])
            has_invalid_base_rule = bool(runtime_flags["invalid_base_rule"])
            has_incomplete = bool(
                bool(runtime_flags["missing_pan"])
                or has_invalid_pan_format
                or has_missing_tax_id
                or has_residency_mismatch
                or bool(runtime_flags["missing_section"])
                or has_invalid_base_rule
                or has_quarter_violation
            )

            if comp_tcs <= Decimal("0.00"):
                lifecycle_status = "NO_COMPUTED_TCS"
                status_counts["no_computed_tcs"] += 1
            elif q2(comp_deposited) >= comp_tcs:
                lifecycle_status = "DEPOSITED"
                status_counts["deposited"] += 1
            elif q2(comp_collected) >= comp_tcs:
                lifecycle_status = "COLLECTED_PENDING_DEPOSIT"
                status_counts["collected_pending_deposit"] += 1
            elif q2(comp_collected) > Decimal("0.00"):
                lifecycle_status = "PARTIALLY_COLLECTED"
                status_counts["partially_collected"] += 1
            else:
                lifecycle_status = "COMPUTED_PENDING_COLLECTION"
                status_counts["computed_pending_collection"] += 1

            if pending_collection > Decimal("0.00"):
                pending_row_counts["pending_collection"] += 1
            if pending_deposit > Decimal("0.00"):
                pending_row_counts["pending_deposit"] += 1
            threshold_counts[threshold_state] = int(threshold_counts.get(threshold_state) or 0) + 1
            impact_counts[impact_key] = int(impact_counts.get(impact_key) or 0) + 1

            bucket = section_summary.setdefault(
                sec_code,
                {
                    "section_code": sec_code,
                    "document_count": 0,
                    "total_base": Decimal("0.00"),
                    "total_computed_tcs": Decimal("0.00"),
                    "total_collected_tcs": Decimal("0.00"),
                    "total_deposited_tcs": Decimal("0.00"),
                    "pending_collection": Decimal("0.00"),
                    "pending_deposit": Decimal("0.00"),
                },
            )
            bucket["document_count"] += 1
            bucket["total_base"] = q2(bucket["total_base"] + comp_base)
            bucket["total_computed_tcs"] = q2(bucket["total_computed_tcs"] + comp_tcs)
            bucket["total_collected_tcs"] = q2(bucket["total_collected_tcs"] + q2(comp_collected))
            bucket["total_deposited_tcs"] = q2(bucket["total_deposited_tcs"] + q2(comp_deposited))
            bucket["pending_collection"] = q2(bucket["pending_collection"] + pending_collection)
            bucket["pending_deposit"] = q2(bucket["pending_deposit"] + pending_deposit)

            if runtime_flags["missing_pan"]:
                quality_counts["missing_pan"] += 1
            if has_invalid_pan_format:
                quality_counts["invalid_pan_format"] += 1
            if has_missing_tax_id:
                quality_counts["missing_tax_id"] += 1
            if has_residency_mismatch:
                quality_counts["residency_mismatch"] += 1
            if runtime_flags["missing_section"]:
                quality_counts["missing_section"] += 1
            if has_invalid_base_rule:
                quality_counts["invalid_base_rule"] += 1
            if has_quarter_violation:
                quality_counts["quarter_boundary_violation"] += 1
            if has_incomplete:
                quality_counts["incomplete_compliance"] += 1

            total_base += comp_base
            total_computed += comp_tcs
            total_collected += q2(comp_collected)
            total_deposited += q2(comp_deposited)

            source_route = _tcs_source_route_for_document(comp.module_name, comp.document_type, comp.document_id)
            posting_lookup_document_type = _tcs_posting_lookup_document_type(comp.module_name, comp.document_type)
            module_name = str(comp.module_name or "").strip().lower()
            document_type = str(comp.document_type or "").strip().lower()
            doc_id = _safe_int(comp.document_id)
            doc_status = ""
            if module_name == "sales" and document_type in {"invoice", "credit_note", "debit_note"} and doc_id:
                doc_status = sales_status_map.get(int(doc_id), "")
            elif module_name == "purchase" and document_type in {"invoice", "credit_note", "debit_note"} and doc_id:
                doc_status = purchase_status_map.get(int(doc_id), "")
            if module_name == "sales":
                is_posted, posting_state, posting_state_label = _invoice_posting_state(
                    doc_status,
                    posted_value=SalesInvoiceHeader.Status.POSTED,
                )
            elif module_name == "purchase":
                is_posted, posting_state, posting_state_label = _invoice_posting_state(
                    doc_status,
                    posted_value=PurchaseInvoiceHeader.Status.POSTED,
                )
            else:
                is_posted, posting_state, posting_state_label = False, "unknown", "Posting status unavailable"

            rows.append(
                {
                    "id": comp.id,
                    "module_name": comp.module_name,
                    "voucher_type": f"{(comp.module_name or '').upper()}_{(comp.document_type or '').upper()}",
                    "doc_impact_type": doc_impact_type,
                    "trigger_basis": _titleize_token(getattr(comp, "trigger_basis", "")),
                    "voucher_date": comp.doc_date,
                    "voucher_no": comp.document_no,
                    "document_type": comp.document_type,
                    "document_id": comp.document_id,
                    "customer_id": comp.party_account_id,
                    "customer_name": (getattr(comp.party_account, "legalname", None) or getattr(comp.party_account, "accountname", None) or "").strip(),
                    "pan": pan_token,
                    "section_id": comp.section_id,
                    "section_code": sec_code,
                    "base_amount": comp_base,
                    "rate": q2(comp.rate or Decimal("0.00")),
                    "computed_tcs": comp_tcs,
                    "collected_tcs": q2(comp_collected),
                    "deposited_tcs": q2(comp_deposited),
                    "pending_collection": pending_collection,
                    "pending_deposit": pending_deposit,
                    "primary_reason_code": reason_code or None,
                    "threshold_default": q2(getattr(comp.section, "threshold_default", Decimal("0.00")) or Decimal("0.00")) if comp.section else None,
                    "threshold_mode": ((getattr(comp.section, "applicability_json", None) or {}).get("threshold_mode") if comp.section else None),
                    "threshold_state": threshold_state,
                    "flags": {
                        "computed": bool(comp_tcs > Decimal("0.00")),
                        "collected": bool(comp_collected > Decimal("0.00")),
                        "deposited": bool(comp_deposited > Decimal("0.00")),
                        "pending": bool(pending_collection > Decimal("0.00") or pending_deposit > Decimal("0.00")),
                        "missing_pan": bool(runtime_flags["missing_pan"]),
                        "invalid_pan_format": has_invalid_pan_format,
                        "missing_tax_id": has_missing_tax_id,
                        "residency_mismatch": has_residency_mismatch,
                        "missing_section": bool(runtime_flags["missing_section"]),
                        "invalid_base_rule": has_invalid_base_rule,
                        "quarter_boundary_violation": has_quarter_violation,
                        "incomplete_compliance": has_incomplete,
                    },
                    "lifecycle_status": lifecycle_status,
                    "computation_status": comp.status,
                    "collections": collections_payload,
                    "is_posted": is_posted,
                    "posting_state": posting_state,
                    "posting_state_label": posting_state_label,
                    "drilldowns": {
                        "source_document": (
                            {
                                "route": source_route,
                                "params": {
                                    "transactionid": int(doc_id),
                                },
                            }
                            if source_route and doc_id
                            else None
                        ),
                        "posting_lookup": (
                            {
                                "lookup": {
                                    "document_type": posting_lookup_document_type,
                                    "document_id": int(doc_id),
                                    "source_module": module_name,
                                }
                            }
                            if posting_lookup_document_type and doc_id and is_posted
                            else None
                        ),
                        "party_master": _party_master_drilldown(
                            party_account_id=comp.party_account_id,
                            entity_id=entity_id,
                            subentity_id=getattr(comp, "subentity_id", None),
                            source="tcs_workspace",
                        ),
                    },
                }
            )
        checkpoint("row_build_ms")

        unallocated_deposits = []
        deposits_qs = TcsDeposit.objects.filter(entity_id=entity_id)
        if fy:
            deposits_qs = deposits_qs.filter(financial_year__in=_expand_fy_values(fy))
        if quarter in {"Q1", "Q2", "Q3", "Q4"}:
            deposits_qs = deposits_qs.filter(month__in=TcsReportFilingPackAPIView._quarter_months(quarter))
        if from_date:
            deposits_qs = deposits_qs.filter(challan_date__gte=from_date)
        if to_date:
            deposits_qs = deposits_qs.filter(challan_date__lte=to_date)

        for dep in deposits_qs.order_by("-challan_date", "-id"):
            allocated = q2(dep.allocations.aggregate(v=Sum("allocated_amount")).get("v") or Decimal("0.00"))
            remaining = q2(q2(dep.total_deposit_amount or Decimal("0.00")) - allocated)
            if remaining > Decimal("0.00"):
                unallocated_deposits.append(
                    {
                        "id": dep.id,
                        "challan_no": dep.challan_no,
                        "challan_date": dep.challan_date,
                        "status": dep.status,
                        "total_deposit_amount": q2(dep.total_deposit_amount or Decimal("0.00")),
                        "allocated_amount": allocated,
                        "unallocated_amount": remaining,
                    }
                )
        checkpoint("unallocated_deposit_ms")

        payload = {
            "filters": {
                "entity_id": entity_id,
                "fy": fy or None,
                "quarter": quarter or None,
                "from_date": from_date,
                "to_date": to_date,
                "section": section_code or None,
                "customer_id": customer_id,
                "customer_q": customer_q or None,
                "search": search or None,
                "include_reversed": include_reversed,
                "include_draft": include_draft,
                "include_cancelled": include_cancelled,
            },
            "summary": {
                "total_transactions": len(rows),
                "total_base": q2(total_base),
                "total_computed_tcs": q2(total_computed),
                "total_collected_tcs": q2(total_collected),
                "total_deposited_tcs": q2(total_deposited),
                "pending_collection": _non_negative_q2(total_computed - total_collected),
                "pending_deposit": _non_negative_q2(total_collected - total_deposited),
                "status_counts": status_counts,
                "impact_counts": impact_counts,
                "pending_row_counts": pending_row_counts,
                "threshold_counts": threshold_counts,
                "quality_counts": quality_counts,
                "unallocated_deposit_count": len(unallocated_deposits),
                "unallocated_deposit_amount": q2(sum((r["unallocated_amount"] for r in unallocated_deposits), Decimal("0.00"))),
            },
            "section_summary": list(section_summary.values()),
            "rows": rows,
            "unallocated_deposits": unallocated_deposits,
        }
        checkpoint("payload_build_ms")
        logger.info(
            "tcs_workspace_transactions_profile entity=%s fy=%s quarter=%s section=%s customer_id=%s search=%s computations=%s rows=%s section_summary=%s "
            "unallocated_deposits=%s total_ms=%.2f stage_ms=%s",
            entity_id,
            fy or None,
            quarter or None,
            section_code or None,
            customer_id,
            bool(search),
            len(computations),
            len(rows),
            len(payload["section_summary"]),
            len(unallocated_deposits),
            (perf_counter() - request_started_at) * 1000,
            stage_timings,
        )
        return Response(payload)


class TcsReportFilingPackAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _quarter_months(quarter: str):
        mapping = {
            "Q1": [4, 5, 6],
            "Q2": [7, 8, 9],
            "Q3": [10, 11, 12],
            "Q4": [1, 2, 3],
        }
        return mapping[quarter]

    def get(self, request):
        request_started_at = perf_counter()
        stage_started_at = request_started_at
        stage_timings = {}

        def checkpoint(name):
            nonlocal stage_started_at
            now = perf_counter()
            stage_timings[name] = round((now - stage_started_at) * 1000, 2)
            stage_started_at = now

        entity_id = request.query_params.get("entity_id")
        fy = (request.query_params.get("fy") or "").strip()
        quarter = (request.query_params.get("quarter") or "").strip().upper()

        if entity_id in (None, ""):
            raise ValidationError({"entity_id": ["This query param is required."]})
        if not fy:
            raise ValidationError({"fy": ["This query param is required."]})
        if quarter not in {"Q1", "Q2", "Q3", "Q4"}:
            raise ValidationError({"quarter": ["quarter must be one of Q1/Q2/Q3/Q4."]})

        entity_id = int(entity_id)
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_FILING_PACK_VIEW_PERMISSIONS,
            message="Missing permission to view the TCS filing pack.",
            feature_code=TCS_REPORTING_FEATURE_CODE,
        )
        months = self._quarter_months(quarter)
        exceptions_only = _safe_bool(request.query_params.get("exceptions_only"))
        pending_only = _safe_bool(request.query_params.get("pending_only"))
        include_cancelled = _safe_bool(request.query_params.get("include_cancelled"))
        section_code = (request.query_params.get("section") or "").strip().upper()
        customer_id = _safe_int(request.query_params.get("customer_id"))
        customer_q = (request.query_params.get("customer_q") or "").strip()
        search = (request.query_params.get("search") or "").strip()
        checkpoint("scope_parse_ms")

        fy_candidates = _expand_fy_values(fy)
        computations = (
            TcsComputation.objects.select_related("party_account", "section")
            .prefetch_related("collections__deposit_allocations__deposit")
            .filter(entity_id=entity_id, fiscal_year__in=fy_candidates, quarter=quarter)
            .order_by("doc_date", "id")
        )
        if customer_id is not None:
            computations = computations.filter(party_account_id=customer_id)
        if customer_q:
            computations = computations.filter(
                Q(party_account__accountname__icontains=customer_q)
                | Q(party_account__legalname__icontains=customer_q)
            )
        if section_code:
            if section_code == "UNMAPPED":
                computations = computations.filter(section__isnull=True)
            else:
                computations = computations.filter(section__section_code__iexact=section_code)
        if not include_cancelled:
            computations = _exclude_cancelled_documents(computations)
        computations = list(computations)
        checkpoint("computations_fetch_ms")
        deposits = TcsDeposit.objects.filter(entity_id=entity_id, financial_year__in=fy_candidates, month__in=months).order_by("challan_date", "id")
        return_row = TcsQuarterlyReturn.objects.filter(entity_id=entity_id, fy__in=fy_candidates, quarter=quarter, form_name="27EQ").order_by("-id").first()
        checkpoint("deposits_and_return_fetch_ms")

        sales_doc_ids = set()
        purchase_doc_ids = set()
        for comp in computations:
            module_name = str(comp.module_name or "").strip().lower()
            document_type = str(comp.document_type or "").strip().lower()
            document_id = int(comp.document_id or 0)
            if not document_id:
                continue
            if module_name == "sales" and document_type in {"invoice", "credit_note", "debit_note"}:
                sales_doc_ids.add(document_id)
            elif module_name == "purchase" and document_type in {"invoice", "credit_note", "debit_note"}:
                purchase_doc_ids.add(document_id)

        sales_status_map = {
            int(row["id"]): int(row["status"])
            for row in SalesInvoiceHeader.objects.filter(id__in=list(sales_doc_ids)).values("id", "status")
        }
        purchase_status_map = {
            int(row["id"]): int(row["status"])
            for row in PurchaseInvoiceHeader.objects.filter(id__in=list(purchase_doc_ids)).values("id", "status")
        }
        checkpoint("document_status_map_ms")

        rows = []
        section_totals = {}
        total_base = Decimal("0.00")
        total_tcs = Decimal("0.00")
        total_collected = Decimal("0.00")

        for comp in computations:
            if not _tcs_search_match(
                comp.document_no,
                getattr(comp.party_account, "legalname", None),
                getattr(comp.party_account, "accountname", None),
                account_pan(comp.party_account),
                getattr(comp.section, "section_code", None),
                comp.document_type,
                getattr(comp, "trigger_basis", None),
                search=search,
            ):
                continue
            party = comp.party_account
            section = comp.section
            party_name = (getattr(party, "legalname", None) or getattr(party, "accountname", None) or "").strip()
            pan = (account_pan(party) or getattr(party, "pan", None) or "").strip().upper()
            invalid_pan_format = bool(pan) and not _is_valid_pan(pan)
            reason_code = str(
                (comp.computation_json or {}).get("reason_code")
                or (comp.rule_snapshot_json or {}).get("reason_code")
                or ""
            ).strip().upper()
            runtime_flags = _tcs_runtime_quality_flags(
                section=section,
                reason_code=reason_code,
                pan=pan,
            )
            quarter_boundary_violation = _quarter_boundary_violation(
                doc_date=comp.doc_date,
                fiscal_year=comp.fiscal_year or "",
                quarter=comp.quarter or "",
            )

            comp_collections = list(comp.collections.all().order_by("collection_date", "id"))
            if not comp_collections:
                comp_collections = [None]

            comp_alloc_total = Decimal("0.00")
            comp_collected_total = Decimal("0.00")
            for c in comp.collections.all():
                comp_collected_total += q2(c.tcs_collected_amount or Decimal("0.00"))
                comp_alloc_total += _sum_tcs_allocation_rows(c.deposit_allocations.all(), deposited_only=True)

            total_base += q2(comp.tcs_base_amount or Decimal("0.00"))
            total_tcs += q2(comp.tcs_amount or Decimal("0.00"))
            total_collected += q2(comp_collected_total)

            section_code = section.section_code if section else "UNMAPPED"
            if section_code not in section_totals:
                section_totals[section_code] = {"section_code": section_code, "total_base": Decimal("0.00"), "total_tcs": Decimal("0.00")}
            section_totals[section_code]["total_base"] = q2(section_totals[section_code]["total_base"] + q2(comp.tcs_base_amount or Decimal("0.00")))
            section_totals[section_code]["total_tcs"] = q2(section_totals[section_code]["total_tcs"] + q2(comp.tcs_amount or Decimal("0.00")))

            for col in comp_collections:
                allocations = list(col.deposit_allocations.all().select_related("deposit").order_by("id")) if col else [None]
                if not allocations:
                    allocations = [None]

                for alloc in allocations:
                    dep = alloc.deposit if alloc else None
                    row = {
                        "document_type": comp.document_type,
                        "document_id": comp.document_id,
                        "document_no": comp.document_no,
                        "doc_date": comp.doc_date,
                        "doc_impact_type": _tcs_doc_impact_type(
                            document_type=comp.document_type,
                            trigger_basis=getattr(comp, "trigger_basis", ""),
                        ),
                        "trigger_basis": _titleize_token(getattr(comp, "trigger_basis", "")),
                        "party_account": comp.party_account_id,
                        "party_name": party_name,
                        "pan": pan,
                        "section_id": comp.section_id,
                        "section_code": section.section_code if section else None,
                        "section_desc": section.description if section else None,
                        "taxable_base": q2(comp.tcs_base_amount or Decimal("0.00")),
                        "tcs_rate": comp.rate,
                        "tcs_amount": q2(comp.tcs_amount or Decimal("0.00")),
                        "applicability_status": comp.applicability_status,
                        "override_reason": comp.override_reason,
                        "computation_status": comp.status,
                        "is_reversal": bool(comp.status == TcsComputation.Status.REVERSED or q2(comp.tcs_amount or Decimal("0.00")) < Decimal("0.00")),
                        "collection_id": col.id if col else None,
                        "collection_date": col.collection_date if col else None,
                        "receipt_voucher_id": col.receipt_voucher_id if col else None,
                        "amount_received": q2(col.amount_received) if col else Decimal("0.00"),
                        "tcs_collected_amount": q2(col.tcs_collected_amount) if col else Decimal("0.00"),
                        "collection_reference": col.collection_reference if col else "",
                        "collection_status": col.status if col else None,
                        "deposit_id": dep.id if dep else None,
                        "challan_no": dep.challan_no if dep else None,
                        "challan_date": dep.challan_date if dep else None,
                        "bsr_code": dep.bsr_code if dep else None,
                        "cin": dep.cin if dep else None,
                        "bank_name": dep.bank_name if dep else None,
                        "allocated_amount": (
                            q2(alloc.allocated_amount)
                            if alloc and _tcs_deposit_status_counts_as_deposited(getattr(dep, "status", None))
                            else Decimal("0.00")
                        ),
                        "deposit_status": dep.status if dep else None,
                        "return_id": return_row.id if return_row else None,
                        "return_quarter": return_row.quarter if return_row else quarter,
                        "return_type": return_row.return_type if return_row else None,
                        "return_status": return_row.status if return_row else "NOT_CREATED",
                        "ack_no": return_row.ack_no if return_row else "",
                        "filed_on": return_row.filed_on if return_row else None,
                        "original_return": return_row.original_return_id if return_row else None,
                        "return_notes": return_row.notes if return_row else "",
                    }
                    source_route = _tcs_source_route_for_document(comp.module_name, comp.document_type, comp.document_id)
                    posting_lookup_document_type = _tcs_posting_lookup_document_type(comp.module_name, comp.document_type)
                    doc_id = int(comp.document_id or 0)
                    posting_state = "unknown"
                    posting_state_label = "Posting state unknown"
                    is_posted = False
                    module_name = str(comp.module_name or "").strip().lower()
                    document_type = str(comp.document_type or "").strip().lower()
                    if module_name == "sales" and document_type in {"invoice", "credit_note", "debit_note"} and doc_id:
                        status_value = sales_status_map.get(doc_id)
                        is_posted = status_value == SalesInvoiceHeader.Status.POSTED
                        posting_state = "posted" if is_posted else "not_posted"
                        posting_state_label = "Posted" if is_posted else "Invoice not posted"
                    elif module_name == "purchase" and document_type in {"invoice", "credit_note", "debit_note"} and doc_id:
                        status_value = purchase_status_map.get(doc_id)
                        is_posted = status_value == PurchaseInvoiceHeader.Status.POSTED
                        posting_state = "posted" if is_posted else "not_posted"
                        posting_state_label = "Posted" if is_posted else "Invoice not posted"
                    row["posting_state"] = posting_state
                    row["posting_state_label"] = posting_state_label
                    row["is_posted"] = is_posted
                    row["drilldowns"] = {}
                    row["drilldowns"]["party_master"] = _party_master_drilldown(
                        party_account_id=comp.party_account_id,
                        entity_id=entity_id,
                        subentity_id=getattr(comp, "subentity_id", None),
                        source="tcs_filing_pack",
                    )
                    if source_route and doc_id:
                        row["drilldowns"]["source_document"] = {
                            "target": "document_source",
                            "label": "Open source document",
                            "kind": "document",
                            "route": source_route,
                            "params": {"transactionid": doc_id},
                        }
                    if posting_lookup_document_type and doc_id and is_posted:
                        row["drilldowns"]["posting_lookup"] = {
                            "target": "posting_detail_lookup",
                            "label": "Open posted voucher",
                            "kind": "posting_lookup",
                            "lookup": {
                                "document_type": posting_lookup_document_type,
                                "document_id": doc_id,
                                "source_module": module_name,
                            },
                        }
                    row["exceptions"] = _tcs_filing_pack_exception_flags(
                        comp_tcs=q2(comp.tcs_amount or Decimal("0.00")),
                        comp_collected_total=q2(comp_collected_total),
                        comp_alloc_total=q2(comp_alloc_total),
                        runtime_flags=runtime_flags,
                        invalid_pan_format=invalid_pan_format,
                        quarter_boundary_violation=quarter_boundary_violation,
                        is_reversal=bool(row["is_reversal"]),
                    )
                    if exceptions_only and not any(bool(v) for v in row["exceptions"].values()):
                        continue
                    if pending_only and not (
                        row["exceptions"]["not_collected"]
                        or row["exceptions"]["not_deposited"]
                        or row["exceptions"]["partially_allocated"]
                        or row["exceptions"]["deposit_mismatch"]
                    ):
                        continue
                    rows.append(row)
        checkpoint("row_build_ms")

        total_deposited = q2(sum((_tcs_computation_total_deposited(comp, deposited_only=True) for comp in computations), Decimal("0.00")))
        pending_collection = _non_negative_q2(total_tcs - total_collected)
        pending_deposit = _non_negative_q2(total_collected - total_deposited)
        exception_row_count = sum(1 for r in rows if any(bool(v) for v in (r.get("exceptions") or {}).values()))
        checkpoint("summary_finalize_ms")

        payload = {
            "header": {
                "entity": entity_id,
                "fy": fy,
                "quarter": quarter,
                "total_base": q2(total_base),
                "total_tcs": q2(total_tcs),
                "total_collected": q2(total_collected),
                "total_deposited": q2(total_deposited),
                "pending_collection": _non_negative_q2(pending_collection),
                "pending_deposit": _non_negative_q2(pending_deposit),
                "return_status": return_row.status if return_row else "NOT_CREATED",
                "row_count": len(rows),
                "exception_row_count": exception_row_count,
            },
            "rows": rows,
            "section_summary": list(section_totals.values()),
        }
        checkpoint("payload_build_ms")
        logger.info(
            "tcs_filing_pack_profile entity=%s fy=%s quarter=%s section=%s customer_id=%s search=%s computations=%s rows=%s section_summary=%s exceptions_only=%s pending_only=%s total_ms=%.2f stage_ms=%s",
            entity_id,
            fy,
            quarter,
            section_code or None,
            customer_id,
            bool(search),
            len(computations),
            len(rows),
            len(payload["section_summary"]),
            exceptions_only,
            pending_only,
            (perf_counter() - request_started_at) * 1000,
            stage_timings,
        )
        return Response(payload)


class TcsWorkspaceTransactionsExportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        payload = TcsWorkspaceTransactionsAPIView().get(request).data
        rows = payload.get("rows") or []
        section_summary = payload.get("section_summary") or []
        unallocated = payload.get("unallocated_deposits") or []
        summary = payload.get("summary") or {}
        filters = payload.get("filters") or {}

        tx_rows = [
            {
                "voucher_date": row.get("voucher_date"),
                "voucher_type": row.get("voucher_type"),
                "voucher_no": row.get("voucher_no"),
                "customer_name": row.get("customer_name"),
                "pan": row.get("pan"),
                "section_code": row.get("section_code"),
                "doc_impact_type": row.get("doc_impact_type"),
                "trigger_basis": row.get("trigger_basis"),
                "base_amount": row.get("base_amount"),
                "rate": row.get("rate"),
                "computed_tcs": row.get("computed_tcs"),
                "collected_tcs": row.get("collected_tcs"),
                "deposited_tcs": row.get("deposited_tcs"),
                "pending_collection": row.get("pending_collection"),
                "pending_deposit": row.get("pending_deposit"),
                "lifecycle_status": row.get("lifecycle_status"),
                "threshold_state": row.get("threshold_state"),
                "threshold_default": row.get("threshold_default"),
                "threshold_mode": row.get("threshold_mode"),
                "primary_reason_code": row.get("primary_reason_code"),
                "incomplete_compliance": (row.get("flags") or {}).get("incomplete_compliance"),
            }
            for row in rows
        ]
        meta_rows = [{"key": k, "value": v} for k, v in {**filters, **summary}.items()]
        zip_bytes = _zip_csv_payload(
            {
                "workspace_transactions.csv": tx_rows,
                "workspace_section_summary.csv": section_summary,
                "workspace_unallocated_deposits.csv": unallocated,
                "workspace_meta.csv": meta_rows,
            }
        )
        resp = HttpResponse(zip_bytes, content_type="application/zip")
        resp["Content-Disposition"] = "attachment; filename=tcs_workspace_export.zip"
        return resp


class TcsReportFilingPackExportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        payload = TcsReportFilingPackAPIView().get(request).data
        rows = payload.get("rows") or []
        section_summary = payload.get("section_summary") or []
        header = payload.get("header") or {}
        exception_keys = [
            "missing_pan",
            "invalid_pan_format",
            "missing_tax_id",
            "residency_mismatch",
            "missing_section",
            "not_collected",
            "not_deposited",
            "partially_allocated",
            "deposit_mismatch",
            "quarter_boundary_violation",
            "reversal_case",
        ]

        filing_rows = []
        exception_spotlight = {key: 0 for key in exception_keys}
        for row in rows:
            exc = row.get("exceptions") or {}
            for key in exception_keys:
                if exc.get(key):
                    exception_spotlight[key] += 1
            filing_rows.append(
                {
                    "doc_date": row.get("doc_date"),
                    "doc_impact_type": row.get("doc_impact_type"),
                    "trigger_basis": row.get("trigger_basis"),
                    "document_type": row.get("document_type"),
                    "document_no": row.get("document_no"),
                    "party_name": row.get("party_name"),
                    "pan": row.get("pan"),
                    "section_code": row.get("section_code"),
                    "taxable_base": row.get("taxable_base"),
                    "tcs_rate": row.get("tcs_rate"),
                    "tcs_amount": row.get("tcs_amount"),
                    "tcs_collected_amount": row.get("tcs_collected_amount"),
                    "allocated_amount": row.get("allocated_amount"),
                    "return_status": row.get("return_status"),
                    "missing_pan": exc.get("missing_pan"),
                    "invalid_pan_format": exc.get("invalid_pan_format"),
                    "missing_tax_id": exc.get("missing_tax_id"),
                    "residency_mismatch": exc.get("residency_mismatch"),
                    "missing_section": exc.get("missing_section"),
                    "not_collected": exc.get("not_collected"),
                    "not_deposited": exc.get("not_deposited"),
                    "partially_allocated": exc.get("partially_allocated"),
                    "deposit_mismatch": exc.get("deposit_mismatch"),
                    "quarter_boundary_violation": exc.get("quarter_boundary_violation"),
                    "reversal_case": exc.get("reversal_case"),
                }
            )
        header_rows = [{"key": k, "value": v} for k, v in header.items()]
        management_summary = [
            {"metric": "fy", "value": header.get("fy")},
            {"metric": "quarter", "value": header.get("quarter")},
            {"metric": "return_status", "value": header.get("return_status")},
            {"metric": "row_count", "value": header.get("row_count")},
            {"metric": "exception_row_count", "value": header.get("exception_row_count")},
            {"metric": "total_tcs", "value": header.get("total_tcs")},
            {"metric": "total_collected", "value": header.get("total_collected")},
            {"metric": "total_deposited", "value": header.get("total_deposited")},
            {"metric": "pending_collection", "value": header.get("pending_collection")},
            {"metric": "pending_deposit", "value": header.get("pending_deposit")},
        ]
        exception_rows = [
            {"exception": key, "affected_rows": count}
            for key, count in exception_spotlight.items()
            if count > 0
        ]
        return_tracker = []
        seen_returns = set()
        for row in rows:
            return_id = row.get("return_id")
            if not return_id or return_id in seen_returns:
                continue
            seen_returns.add(return_id)
            return_tracker.append(
                {
                    "return_id": return_id,
                    "return_quarter": row.get("return_quarter"),
                    "return_type": row.get("return_type"),
                    "return_status": row.get("return_status"),
                    "original_return": row.get("original_return"),
                    "ack_no": row.get("ack_no"),
                    "filed_on": row.get("filed_on"),
                    "return_notes": row.get("return_notes"),
                }
            )
        zip_bytes = _zip_csv_payload(
            {
                "filing_pack_management_summary.csv": management_summary,
                "filing_pack_transactions.csv": filing_rows,
                "filing_pack_exception_spotlight.csv": exception_rows,
                "filing_pack_return_tracker.csv": return_tracker,
                "filing_pack_section_summary.csv": section_summary,
                "filing_pack_header.csv": header_rows,
            }
        )
        resp = HttpResponse(zip_bytes, content_type="application/zip")
        resp["Content-Disposition"] = "attachment; filename=tcs_filing_pack_export.zip"
        return resp


class TcsComplianceCenterCaPackExportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        entity_id = _safe_int(request.query_params.get("entity_id") or request.query_params.get("entity"))
        fy = (request.query_params.get("fy") or "").strip()
        quarter = (request.query_params.get("quarter") or "").strip().upper()
        if entity_id is None:
            raise ValidationError({"entity_id": ["This query param is required."]})
        if not fy:
            raise ValidationError({"fy": ["This query param is required."]})
        if quarter not in {"Q1", "Q2", "Q3", "Q4"}:
            raise ValidationError({"quarter": ["quarter must be one of Q1/Q2/Q3/Q4 for CA Pack export."]})

        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=TCS_WORKSPACE_VIEW_PERMISSIONS,
            message="Missing permission to export the TCS compliance center CA Pack.",
        )

        workspace_payload = TcsWorkspaceTransactionsAPIView().get(request).data
        filing_payload = TcsReportFilingPackAPIView().get(request).data
        ledger_payload = TcsReportLedgerAPIView().get(request).data
        returns_rows = list(
            TcsQuarterlyReturn.objects.filter(
                entity_id=entity_id,
                fy__in=_expand_fy_values(fy),
                quarter=quarter,
                form_name="27EQ",
            ).order_by("-id")
        )

        wb = Workbook()
        wb.remove(wb.active)

        def create_sheet(title: str):
            return wb.create_sheet(title=title[:31])

        def autosize_sheet(ws):
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 12
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    value = row[0].value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 44)

        def decorate_table(ws, header_row: int = 1):
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            for cell in ws[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            for row_idx in range(header_row + 1, ws.max_row + 1):
                for cell in ws[row_idx]:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top")
            ws.freeze_panes = f"A{header_row + 1}"
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            autosize_sheet(ws)

        def write_rows_sheet(title: str, headers: list[str], rows: list[dict[str, object] | tuple | list]):
            ws = create_sheet(title)
            for index, header in enumerate(headers, start=1):
                ws.cell(row=1, column=index, value=header)
            if rows:
                for row_index, row in enumerate(rows, start=2):
                    if isinstance(row, dict):
                        values = [row.get(header_key) for header_key in headers]
                    else:
                        values = list(row)
                    for col_index, value in enumerate(values, start=1):
                        ws.cell(row=row_index, column=col_index, value=value)
            else:
                ws.cell(row=2, column=1, value="No rows available in the selected scope.")
            decorate_table(ws)

        ws = create_sheet("00_Cover")
        ws["A1"] = "TCS Compliance Center CA Pack"
        ws["A2"] = "Entity ID"
        ws["B2"] = entity_id
        ws["A3"] = "Financial Year"
        ws["B3"] = fy
        ws["A4"] = "Quarter"
        ws["B4"] = quarter
        ws["A5"] = "Section Filter"
        ws["B5"] = workspace_payload.get("filters", {}).get("section") or "All Sections"  # type: ignore[union-attr]
        ws["A6"] = "Customer Filter"
        ws["B6"] = workspace_payload.get("filters", {}).get("customer_q") or "All Customers"  # type: ignore[union-attr]
        ws["A7"] = "Generated At"
        ws["B7"] = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
        ws["A8"] = "Generated By"
        ws["B8"] = getattr(request.user, "username", "") or getattr(request.user, "email", "") or getattr(request.user, "id", "")
        ws["A10"] = "Included Sheets"
        for row_idx, name in enumerate([
            "01_Management_Summary",
            "02_Workspace_Transactions",
            "03_Workspace_Section_Summary",
            "04_Filing_Pack",
            "05_Filing_Section_Summary",
            "06_Ledger_Summary",
            "07_Return_27EQ",
            "08_Unallocated_Deposits",
        ], start=11):
            ws[f"A{row_idx}"] = name
        ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
        for row_idx in range(2, 9):
            ws[f"A{row_idx}"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 42

        management_summary = [
            {"Metric": "Total Transactions", "Value": workspace_payload.get("summary", {}).get("total_transactions")},  # type: ignore[union-attr]
            {"Metric": "Total Computed TCS", "Value": workspace_payload.get("summary", {}).get("total_computed_tcs")},  # type: ignore[union-attr]
            {"Metric": "Pending Collection", "Value": workspace_payload.get("summary", {}).get("pending_collection")},  # type: ignore[union-attr]
            {"Metric": "Pending Deposit", "Value": workspace_payload.get("summary", {}).get("pending_deposit")},  # type: ignore[union-attr]
            {"Metric": "Filing Pack Rows", "Value": filing_payload.get("header", {}).get("row_count")},  # type: ignore[union-attr]
            {"Metric": "Filing Exceptions", "Value": filing_payload.get("header", {}).get("exception_row_count")},  # type: ignore[union-attr]
            {"Metric": "Return Status", "Value": filing_payload.get("header", {}).get("return_status")},  # type: ignore[union-attr]
            {"Metric": "Ledger Buckets", "Value": len(list(ledger_payload or []))},
            {"Metric": "27EQ Returns", "Value": len(returns_rows)},
        ]
        write_rows_sheet("01_Management_Summary", ["Metric", "Value"], management_summary)

        workspace_rows = [
            {
                "voucher_date": row.get("voucher_date"),
                "voucher_no": row.get("voucher_no"),
                "customer_name": row.get("customer_name"),
                "pan": row.get("pan"),
                "section_code": row.get("section_code"),
                "base_amount": row.get("base_amount"),
                "computed_tcs": row.get("computed_tcs"),
                "collected_tcs": row.get("collected_tcs"),
                "deposited_tcs": row.get("deposited_tcs"),
                "pending_collection": row.get("pending_collection"),
                "pending_deposit": row.get("pending_deposit"),
                "lifecycle_status": row.get("lifecycle_status"),
                "readiness_status": row.get("readiness_status"),
            }
            for row in list(workspace_payload.get("rows") or [])  # type: ignore[union-attr]
        ]
        write_rows_sheet(
            "02_Workspace_Transactions",
            ["voucher_date", "voucher_no", "customer_name", "pan", "section_code", "base_amount", "computed_tcs", "collected_tcs", "deposited_tcs", "pending_collection", "pending_deposit", "lifecycle_status", "readiness_status"],
            workspace_rows,
        )
        write_rows_sheet(
            "03_Workspace_Section_Summary",
            ["section_code", "document_count", "total_base", "total_computed_tcs", "total_collected_tcs", "total_deposited_tcs", "pending_collection", "pending_deposit"],
            list(workspace_payload.get("section_summary") or []),  # type: ignore[union-attr]
        )
        write_rows_sheet(
            "04_Filing_Pack",
            ["doc_date", "document_no", "party_name", "pan", "section_code", "taxable_base", "tcs_amount", "collection_status", "deposit_status", "return_status"],
            [
                {
                    "doc_date": row.get("doc_date"),
                    "document_no": row.get("document_no"),
                    "party_name": row.get("party_name"),
                    "pan": row.get("pan"),
                    "section_code": row.get("section_code"),
                    "taxable_base": row.get("taxable_base"),
                    "tcs_amount": row.get("tcs_amount"),
                    "collection_status": row.get("collection_status"),
                    "deposit_status": row.get("deposit_status"),
                    "return_status": row.get("return_status"),
                }
                for row in list(filing_payload.get("rows") or [])  # type: ignore[union-attr]
            ],
        )
        write_rows_sheet(
            "05_Filing_Section_Summary",
            ["section_code", "document_count", "total_base", "total_tcs", "collected_amount", "allocated_amount", "pending_collection", "pending_deposit"],
            list(filing_payload.get("section_summary") or []),  # type: ignore[union-attr]
        )
        write_rows_sheet(
            "06_Ledger_Summary",
            ["section_code_norm", "doc_count", "total_base", "total_tcs"],
            list(ledger_payload or []),
        )
        write_rows_sheet(
            "07_Return_27EQ",
            ["fy", "quarter", "return_type", "status", "ack_no", "filed_on", "original_return", "notes"],
            [
                {
                    "fy": row.fy,
                    "quarter": row.quarter,
                    "return_type": row.return_type,
                    "status": row.status,
                    "ack_no": row.ack_no,
                    "filed_on": row.filed_on,
                    "original_return": row.original_return_id,
                    "notes": row.notes,
                }
                for row in returns_rows
            ],
        )
        write_rows_sheet(
            "08_Unallocated_Deposits",
            ["challan_no", "challan_date", "status", "total_deposit_amount", "allocated_amount", "unallocated_amount"],
            list(workspace_payload.get("unallocated_deposits") or []),  # type: ignore[union-attr]
        )

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{_safe_filename(f"tcs_compliance_ca_pack_{quarter}")}.xlsx"'
        return response


class WithholdingReadinessDashboardAPIView(APIView):
    """
    Unified compliance-readiness summary for payment-triggered withholding (TDS runtime).
    Focus buckets: 194A, 194N, 195.
    """

    permission_classes = [IsAuthenticated]
    TARGET_SECTIONS = {"194A", "194N", "195"}

    def get(self, request):
        entity_id = _safe_int(request.query_params.get("entity_id") or request.query_params.get("entity"))
        if entity_id is None:
            raise ValidationError({"entity_id": ["This query param is required."]})
        _require_tcs_scope_permission(
            request=request,
            entity_id=entity_id,
            permission_codes=WITHHOLDING_READINESS_VIEW_PERMISSIONS,
            message="Missing permission to view withholding readiness.",
        )

        entityfin_id = _safe_int(request.query_params.get("entityfinid") or request.query_params.get("entityfin_id"))
        subentity_id = _safe_int(request.query_params.get("subentity_id") or request.query_params.get("subentity"))
        fy = (request.query_params.get("fy") or "").strip()
        quarter = (request.query_params.get("quarter") or "").strip().upper()
        from_date, to_date = _resolve_period_bounds(
            fy=fy,
            quarter=quarter,
            from_date_raw=request.query_params.get("from_date"),
            to_date_raw=request.query_params.get("to_date"),
        )
        include_all_sections = _safe_bool(request.query_params.get("include_all_sections"))

        vouchers = PaymentVoucherHeader.objects.filter(entity_id=entity_id).exclude(
            status=PaymentVoucherHeader.Status.CANCELLED
        )
        if entityfin_id is not None:
            vouchers = vouchers.filter(entityfinid_id=entityfin_id)
        if subentity_id is not None:
            vouchers = vouchers.filter(subentity_id=subentity_id)
        if from_date:
            vouchers = vouchers.filter(voucher_date__gte=from_date)
        if to_date:
            vouchers = vouchers.filter(voucher_date__lte=to_date)

        voucher_rows = list(
            vouchers.select_related("paid_to").only(
                "id",
                "entity_id",
                "entityfinid_id",
                "subentity_id",
                "voucher_date",
                "doc_code",
                "doc_no",
                "status",
                "paid_to_id",
                "workflow_payload",
            ).order_by("-voucher_date", "-id")
        )

        section_ids = set()
        party_ids = set()
        for voucher in voucher_rows:
            party_ids.add(voucher.paid_to_id)
            payload = voucher.workflow_payload if isinstance(voucher.workflow_payload, dict) else {}
            runtime = payload.get("withholding_runtime_result") if isinstance(payload.get("withholding_runtime_result"), dict) else {}
            section_id = _safe_int(runtime.get("section_id")) or _safe_int((payload.get("withholding") or {}).get("section_id"))
            if section_id:
                section_ids.add(section_id)

        section_map = {
            row.id: row
            for row in WithholdingSection.objects.filter(id__in=section_ids).only("id", "section_code", "description")
        }
        profile_map = {
            row.party_account_id: row
            for row in EntityPartyTaxProfile.objects.filter(entity_id=entity_id, party_account_id__in=party_ids, is_active=True).order_by("-updated_at")
        }

        rows: list[dict] = []
        quality_counts = {
            "missing_pan": 0,
            "missing_tax_id": 0,
            "residency_mismatch": 0,
            "invalid_base_rule": 0,
            "missing_section": 0,
        }
        status_counts = {"ready_to_file": 0, "blocked": 0, "fix_now": 0}
        section_buckets: dict[str, dict] = {}
        total_amount = Decimal("0.00")

        for voucher in voucher_rows:
            payload = voucher.workflow_payload if isinstance(voucher.workflow_payload, dict) else {}
            runtime = payload.get("withholding_runtime_result") if isinstance(payload.get("withholding_runtime_result"), dict) else {}
            withholding_cfg = payload.get("withholding") if isinstance(payload.get("withholding"), dict) else {}
            section_id = _safe_int(runtime.get("section_id")) or _safe_int(withholding_cfg.get("section_id"))
            section_obj = section_map.get(section_id) if section_id else None
            section_code = str(getattr(section_obj, "section_code", "") or "").strip().upper()
            if not include_all_sections and section_code and section_code not in self.TARGET_SECTIONS:
                continue
            if not include_all_sections and not section_code:
                continue

            amount = _safe_decimal(runtime.get("amount"))
            base_amount = _safe_decimal(runtime.get("base_amount"))
            rate = _safe_decimal(runtime.get("rate"))
            reason_code = str(runtime.get("reason_code") or "")
            reason = str(runtime.get("reason") or "")
            mode = str(runtime.get("mode") or withholding_cfg.get("mode") or "AUTO").upper().strip()
            enabled = bool(runtime.get("enabled", withholding_cfg.get("enabled", False)))
            pan = (account_pan(voucher.paid_to) or getattr(voucher.paid_to, "pan", None) or "").strip().upper()
            profile = profile_map.get(voucher.paid_to_id)
            residency_status = str(getattr(profile, "residency_status", "") or "").strip().lower()
            tax_identifier = str(getattr(profile, "tax_identifier", "") or "").strip()

            flags = _runtime_quality_flags(
                section_code=section_code,
                reason_code=reason_code,
                pan=pan,
                tax_identifier=tax_identifier,
                residency_status=residency_status,
            )
            row_status = _row_readiness_status(amount=amount, flags=flags)
            status_counts[row_status] += 1
            for key in quality_counts.keys():
                if flags.get(key):
                    quality_counts[key] += 1

            bucket_key = section_code or "UNMAPPED"
            if bucket_key not in section_buckets:
                section_buckets[bucket_key] = {
                    "section_code": bucket_key,
                    "section_description": getattr(section_obj, "description", "") if section_obj else "",
                    "voucher_count": 0,
                    "total_base": Decimal("0.00"),
                    "total_amount": Decimal("0.00"),
                    "ready_to_file": 0,
                    "blocked": 0,
                    "fix_now": 0,
                }
            bucket = section_buckets[bucket_key]
            bucket["voucher_count"] += 1
            bucket["total_base"] = q2(bucket["total_base"] + base_amount)
            bucket["total_amount"] = q2(bucket["total_amount"] + amount)
            bucket[row_status] += 1
            total_amount = q2(total_amount + amount)

            rows.append(
                {
                    "source": "payment_voucher",
                    "voucher_id": voucher.id,
                    "voucher_date": voucher.voucher_date,
                    "voucher_no": f"{voucher.doc_code}-{voucher.doc_no}" if voucher.doc_no else voucher.doc_code,
                    "status": voucher.status,
                    "party_account_id": voucher.paid_to_id,
                    "section_id": section_id,
                    "section_code": section_code,
                    "mode": mode,
                    "enabled": enabled,
                    "base_amount": base_amount,
                    "rate": rate,
                    "withholding_amount": amount,
                    "reason_code": reason_code,
                    "reason": reason,
                    "pan": pan,
                    "tax_identifier": tax_identifier,
                    "residency_status": residency_status or "unknown",
                    "quality_flags": flags,
                    "readiness_status": row_status,
                    "is_posted": int(voucher.status or 0) == int(PaymentVoucherHeader.Status.POSTED),
                    "posting_state": (
                        "posted"
                        if int(voucher.status or 0) == int(PaymentVoucherHeader.Status.POSTED)
                        else "not_posted"
                    ),
                    "posting_state_label": (
                        "Posted"
                        if int(voucher.status or 0) == int(PaymentVoucherHeader.Status.POSTED)
                        else "Voucher not posted"
                    ),
                    "drilldowns": {
                        "source_document": {
                            "route": "/paymentvoucher",
                            "params": {
                                "transactionid": int(voucher.id),
                            },
                        },
                        "posting_lookup": (
                            {
                                "lookup": {
                                    "document_type": "payment_voucher",
                                    "document_id": int(voucher.id),
                                    "source_module": "payment",
                                }
                            }
                            if int(voucher.status or 0) == int(PaymentVoucherHeader.Status.POSTED)
                            else None
                        ),
                    },
                }
            )

        section_summary = sorted(section_buckets.values(), key=lambda row: row.get("section_code") or "")
        return Response(
            {
                "filters": {
                    "entity_id": entity_id,
                    "entityfinid": entityfin_id,
                    "subentity_id": subentity_id,
                    "fy": fy or None,
                    "quarter": quarter or None,
                    "from_date": from_date,
                    "to_date": to_date,
                    "include_all_sections": include_all_sections,
                },
                "header": {
                    "target_sections": sorted(self.TARGET_SECTIONS),
                    "row_count": len(rows),
                    "total_withholding_amount": total_amount,
                },
                "widgets": {
                    "ready_to_file": status_counts["ready_to_file"],
                    "blocked_items": status_counts["blocked"],
                    "fix_now_items": status_counts["fix_now"],
                },
                "quality_flags": quality_counts,
                "section_summary": section_summary,
                "rows": rows,
            }
        )


def _build_tcs_27eq_snapshot(*, entity_id: int, fy: str, quarter: str) -> dict:
    quarter = (quarter or "").strip().upper()
    months = TcsReportFilingPackAPIView._quarter_months(quarter)
    fy_candidates = _expand_fy_values(fy)
    computations = (
        TcsComputation.objects.select_related("party_account", "section")
        .prefetch_related("collections__deposit_allocations")
        .filter(entity_id=entity_id, fiscal_year__in=fy_candidates, quarter=quarter)
    )
    deposits = TcsDeposit.objects.filter(entity_id=entity_id, financial_year__in=fy_candidates, month__in=months)

    total_base = q2(computations.aggregate(v=Sum("tcs_base_amount")).get("v") or Decimal("0.00"))
    total_tcs = q2(computations.aggregate(v=Sum("tcs_amount")).get("v") or Decimal("0.00"))
    total_collected = q2(
        TcsCollection.objects.filter(computation__in=computations).exclude(status=TcsCollection.Status.CANCELLED).aggregate(v=Sum("tcs_collected_amount")).get("v")
        or Decimal("0.00")
    )
    total_deposited = q2(sum((_tcs_computation_total_deposited(comp, deposited_only=True) for comp in computations), Decimal("0.00")))

    missing_pan_count = 0
    invalid_pan_format_count = 0
    missing_section_count = 0
    missing_tax_id_count = 0
    residency_mismatch_count = 0
    quarter_boundary_violation_count = 0
    not_collected_count = 0
    not_deposited_count = 0
    partially_allocated_count = 0
    deposit_mismatch_count = 0

    for comp in computations:
        pan = (account_pan(comp.party_account) or getattr(comp.party_account, "pan", None) or "").strip().upper()
        reason_code = str(
            (comp.computation_json or {}).get("reason_code")
            or (comp.rule_snapshot_json or {}).get("reason_code")
            or ""
        ).strip().upper()
        runtime_flags = _tcs_runtime_quality_flags(
            section=comp.section,
            reason_code=reason_code,
            pan=pan,
        )
        if runtime_flags["missing_pan"]:
            missing_pan_count += 1
        elif pan and not _is_valid_pan(pan):
            invalid_pan_format_count += 1

        comp_collected_total = Decimal("0.00")
        comp_alloc_total = Decimal("0.00")
        for col in comp.collections.all():
            if col.status == TcsCollection.Status.CANCELLED:
                continue
            comp_collected_total += q2(col.tcs_collected_amount or Decimal("0.00"))
            comp_alloc_total += _sum_tcs_allocation_rows(col.deposit_allocations.all(), deposited_only=True)

        comp_tcs = q2(comp.tcs_amount or Decimal("0.00"))
        quarter_violation = _quarter_boundary_violation(doc_date=comp.doc_date, fiscal_year=comp.fiscal_year or "", quarter=comp.quarter or "")
        exception_flags = _tcs_filing_pack_exception_flags(
            comp_tcs=comp_tcs,
            comp_collected_total=q2(comp_collected_total),
            comp_alloc_total=q2(comp_alloc_total),
            runtime_flags=runtime_flags,
            invalid_pan_format=bool(pan and not _is_valid_pan(pan)),
            quarter_boundary_violation=quarter_violation,
            is_reversal=bool(comp.status == TcsComputation.Status.REVERSED or comp_tcs < Decimal("0.00")),
        )
        if exception_flags["missing_section"]:
            missing_section_count += 1
        if exception_flags["missing_tax_id"]:
            missing_tax_id_count += 1
        if exception_flags["residency_mismatch"]:
            residency_mismatch_count += 1
        if exception_flags["quarter_boundary_violation"]:
            quarter_boundary_violation_count += 1
        if comp_tcs > Decimal("0.00") and comp_collected_total <= Decimal("0.00"):
            not_collected_count += 1
        if comp_collected_total > Decimal("0.00") and comp_alloc_total <= Decimal("0.00"):
            not_deposited_count += 1
        if comp_alloc_total > Decimal("0.00") and comp_alloc_total < comp_collected_total:
            partially_allocated_count += 1
        if q2(comp_alloc_total) != q2(comp_collected_total):
            deposit_mismatch_count += 1

    return {
        "entity_id": entity_id,
        "fy": fy,
        "quarter": quarter,
        "totals": {
            "total_base": total_base,
            "total_tcs": total_tcs,
            "total_collected": total_collected,
            "total_deposited": total_deposited,
            "pending_collection": _non_negative_q2(total_tcs - total_collected),
            "pending_deposit": _non_negative_q2(total_collected - total_deposited),
        },
        "counts": {
            "computations": computations.count(),
            "deposits": deposits.count(),
            "missing_pan": missing_pan_count,
            "invalid_pan_format": invalid_pan_format_count,
            "missing_tax_id": missing_tax_id_count,
            "residency_mismatch": residency_mismatch_count,
            "quarter_boundary_violation": quarter_boundary_violation_count,
            "missing_section": missing_section_count,
            "not_collected": not_collected_count,
            "not_deposited": not_deposited_count,
            "partially_allocated": partially_allocated_count,
            "deposit_mismatch": deposit_mismatch_count,
        },
    }


class GstTcsEcoProfileListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = GstTcsEcoProfileSerializer
    queryset = GstTcsEcoProfile.objects.all().order_by("-effective_from", "-id")


class GstTcsComputationListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = GstTcsComputationSerializer

    def get_queryset(self):
        qs = GstTcsComputation.objects.all().order_by("-doc_date", "-id")
        entity_id = self.request.query_params.get("entity_id")
        if entity_id not in (None, ""):
            qs = qs.filter(entity_id=int(entity_id))
        fy = (self.request.query_params.get("fy") or "").strip()
        if fy:
            qs = qs.filter(fy=fy)
        month = self.request.query_params.get("month")
        if month not in (None, ""):
            qs = qs.filter(month=int(month))
        return qs

    def create(self, request, *args, **kwargs):
        s = GstTcsComputeRequestSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        d = s.validated_data

        try:
            profile = GstTcsEcoProfile.objects.get(pk=d["eco_profile_id"], entity_id=d["entity_id"])
        except GstTcsEcoProfile.DoesNotExist:
            return Response({"detail": "Invalid eco_profile_id for entity."}, status=status.HTTP_400_BAD_REQUEST)

        rate = d.get("gst_tcs_rate")
        if rate is None:
            rate = profile.default_rate

        taxable_value = q2(d["taxable_value"])
        amount = q2((taxable_value * Decimal(rate)) / Decimal("100.0"))

        row, _ = GstTcsComputation.objects.update_or_create(
            entity_id=d["entity_id"],
            document_type=(d.get("document_type") or "invoice").strip().lower(),
            document_id=d.get("document_id") or 0,
            defaults={
                "eco_profile": profile,
                "supplier_account_id": d["supplier_account_id"],
                "doc_date": d["doc_date"],
                "document_no": (d.get("document_no") or "").strip(),
                "taxable_value": taxable_value,
                "gst_tcs_rate": Decimal(rate),
                "gst_tcs_amount": amount,
                "fy": d["fy"],
                "month": d["month"],
                "status": d["status"],
                "snapshot_json": {
                    "section_code": profile.section_code,
                    "gstin": profile.gstin,
                    "computed_rate": str(rate),
                },
            },
        )
        return Response(GstTcsComputationSerializer(row).data, status=status.HTTP_201_CREATED)
