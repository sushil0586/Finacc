from __future__ import annotations

import csv
import re
from io import BytesIO, StringIO
from decimal import Decimal
from datetime import date as date_cls

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from financial.models import account
from financial.profile_access import account_gstno, account_pan, account_partytype
from reports.selectors.financial import resolve_scope_names
from sales.serializers.sales_ar import (
    CustomerAdvanceBalanceSerializer,
    CustomerBillOpenItemSerializer,
    CustomerSettlementSerializer,
)
from sales.services.sales_ar_service import SalesArService
from sales.views.sales_ar import _parse_scope, _require_ar_view_permission


def _filtered_querydict(request, *, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    return params.urlencode()


def _attach_customer_statement_actions(payload, request, *, export_base_path):
    query = _filtered_querydict(request)
    payload["actions"] = {
        "can_view": True,
        "can_export_excel": True,
        "can_export_pdf": True,
        "can_export_csv": True,
        "can_print": True,
        "export_urls": {
            "excel": f"{export_base_path}excel/?{query}",
            "pdf": f"{export_base_path}pdf/?{query}",
            "csv": f"{export_base_path}csv/?{query}",
            "print": f"{export_base_path}print/?{query}",
        },
    }
    payload["available_exports"] = ["excel", "pdf", "csv", "print"]
    return payload


def _safe_filename(value):
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    return text.strip("._-") or "report"


def _format_display_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d-%m-%Y")
    text = str(value).strip()
    if not text:
        return ""
    try:
        return date_cls.fromisoformat(text[:10]).strftime("%d-%m-%Y")
    except Exception:
        return text


def _amount(value, *, display: bool):
    if value in (None, ""):
        return "" if display else None
    try:
        decimal_value = Decimal(str(value))
    except Exception:
        return str(value) if display else value
    return f"{decimal_value:,.2f}" if display else decimal_value


def _build_customer_statement_payload(*, entity_id, entityfinid_id, subentity_id, customer_id, include_closed):
    data = SalesArService.customer_statement(
        entity_id=entity_id,
        entityfinid_id=entityfinid_id,
        subentity_id=subentity_id,
        customer_id=customer_id,
        include_closed=include_closed,
    )
    customer_obj = (
        account.objects.filter(id=customer_id)
        .select_related("ledger", "compliance_profile", "commercial_profile")
        .first()
    )
    customer_block = None
    if customer_obj is not None:
        customer_block = {
            "id": customer_obj.id,
            "accountname": customer_obj.accountname,
            "display_name": customer_obj.effective_accounting_name,
            "accountcode": customer_obj.effective_accounting_code,
            "ledger_id": customer_obj.ledger_id,
            "partytype": account_partytype(customer_obj),
            "gstno": account_gstno(customer_obj),
            "pan": account_pan(customer_obj),
        }

    def _serialize_items(serializer_class, items):
        items = list(items or [])
        if not items:
            return []
        if isinstance(items[0], dict):
            return items
        return serializer_class(items, many=True).data

    payload = {
        "customer": customer_block,
        "totals": data["totals"],
        "open_items": _serialize_items(CustomerBillOpenItemSerializer, data["open_items"]),
        "advances": _serialize_items(CustomerAdvanceBalanceSerializer, data["advances"]),
        "settlements": _serialize_items(CustomerSettlementSerializer, data["settlements"]),
    }
    return payload


def _statement_subtitle(*, scope_names, payload, include_closed: bool):
    customer = payload.get("customer") or {}
    customer_label = customer.get("display_name") or customer.get("accountname") or "Customer"
    fy_label = scope_names.get("entityfin_name") or "Selected financial year"
    subentity_label = scope_names.get("subentity_name") or "All subentities"
    closed_label = "Closed items included" if include_closed else "Open items only"
    return (
        f"Entity: {scope_names.get('entity_name') or 'Selected entity'} | "
        f"Financial Year: {fy_label} | "
        f"Subentity: {subentity_label} | "
        f"Customer: {customer_label} | "
        f"{closed_label}"
    )


def _statement_sheet_data(payload, *, display: bool):
    totals = payload.get("totals") or {}
    customer = payload.get("customer") or {}
    summary_rows = [
        ["Customer", customer.get("display_name") or customer.get("accountname") or ""],
        ["Financial Year", payload.get("financial_year_name") or ""],
        ["Subentity", payload.get("subentity_name") or "All subentities"],
        ["Closed Items", "Included" if payload.get("include_closed") else "Open only"],
        ["Outstanding Total", _amount(totals.get("outstanding_total"), display=display)],
        ["Advances Total", _amount(totals.get("advance_outstanding_total"), display=display)],
        ["Consumed Total", _amount(totals.get("advance_consumed_total"), display=display)],
        ["Net AR Position", _amount(totals.get("net_ar_position", totals.get("net_ap_position")), display=display)],
    ]

    open_items_headers = [
        "Bill Date",
        "Due Date",
        "Invoice No",
        "Ref No",
        "Original",
        "Settled",
        "Outstanding",
        "Status",
    ]
    open_items_rows = [
        [
            _format_display_date(row.get("bill_date")),
            _format_display_date(row.get("due_date")),
            row.get("invoice_number") or row.get("document_no") or row.get("open_item_no") or "-",
            row.get("customer_reference_number") or "-",
            _amount(row.get("original_amount"), display=display),
            _amount(row.get("settled_amount"), display=display),
            _amount(row.get("outstanding_amount"), display=display),
            "Open" if row.get("is_open") else "Closed",
        ]
        for row in payload.get("open_items") or []
    ]

    advances_headers = [
        "Voucher Date",
        "Voucher No",
        "Receipt Type",
        "Reference No",
        "Original",
        "Adjusted",
        "Balance",
        "Status",
    ]
    advances_rows = [
        [
            _format_display_date(row.get("voucher_date")),
            row.get("voucher_code") or row.get("doc_no") or row.get("reference_no") or "-",
            row.get("receipt_type") or "-",
            row.get("reference_no") or "-",
            _amount(row.get("original_amount"), display=display),
            _amount(row.get("adjusted_amount"), display=display),
            _amount(row.get("balance_amount") or row.get("outstanding_amount"), display=display),
            "Open" if row.get("is_open") else "Closed",
        ]
        for row in payload.get("advances") or []
    ]

    settlements_headers = [
        "Settlement Date",
        "Type",
        "Reference No",
        "Total Amount",
        "Status",
        "Advance Ref",
        "Source Voucher",
    ]
    settlements_rows = [
        [
            _format_display_date(row.get("settlement_date")),
            row.get("settlement_type_name") or row.get("settlement_type") or "-",
            row.get("reference_no") or "-",
            _amount(row.get("total_amount"), display=display),
            row.get("status_name") or row.get("status") or "-",
            row.get("advance_reference_no") or "-",
            row.get("source_receipt_voucher_code") or row.get("source_receipt_voucher_id") or "-",
        ]
        for row in payload.get("settlements") or []
    ]

    return {
        "summary": summary_rows,
        "open_items": (open_items_headers, open_items_rows),
        "advances": (advances_headers, advances_rows),
        "settlements": (settlements_headers, settlements_rows),
    }


def _write_summary_sheet(ws, summary_rows):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["Customer Ledger Statement"])
    ws.append([])
    ws.append(["Label", "Value"])
    for row in summary_rows:
        ws.append(row)

    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24


def _write_section_sheet(ws, title, subtitle, headers, rows, numeric_columns):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([title])
    ws.append([subtitle])
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(headers[col_idx - 1]) + 4, 14)

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.row == header_row:
                continue
            if cell.column in numeric_columns and cell.value not in (None, ""):
                cell.alignment = right
                try:
                    cell.value = Decimal(str(cell.value).replace(",", ""))
                    cell.number_format = "#,##0.00"
                except Exception:
                    pass
            else:
                cell.alignment = left


def _write_customer_statement_excel(payload, subtitle):
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    _write_summary_sheet(summary_sheet, _statement_sheet_data(payload, display=False)["summary"])

    sections = _statement_sheet_data(payload, display=False)
    section_specs = [
        ("Open Items", sections["open_items"][0], sections["open_items"][1], [5, 6, 7]),
        ("Advances", sections["advances"][0], sections["advances"][1], [5, 6, 7]),
        ("Settlements", sections["settlements"][0], sections["settlements"][1], [3]),
    ]
    for title, headers, rows, numeric_columns in section_specs:
        ws = wb.create_sheet(title=title[:31])
        _write_section_sheet(ws, title, subtitle, headers, rows, numeric_columns)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_customer_statement_csv(payload, subtitle):
    sections = _statement_sheet_data(payload, display=True)
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow(["Customer Ledger Statement"])
    writer.writerow([subtitle])
    writer.writerow([])
    writer.writerow(["Summary"])
    writer.writerow(["Label", "Value"])
    writer.writerows(sections["summary"])
    writer.writerow([])
    for title, (headers, rows) in [
        ("Open Items", sections["open_items"]),
        ("Advances", sections["advances"]),
        ("Settlements", sections["settlements"]),
    ]:
        writer.writerow([title])
        writer.writerow(headers)
        writer.writerows(rows)
        writer.writerow([])
    return stream.getvalue().encode("utf-8")


def _decorate_pdf(canvas_obj, doc):
    canvas_obj.saveState()
    width, height = doc.pagesize
    top_y = height - 28
    canvas_obj.setFillColor(colors.HexColor("#1f4e79"))
    canvas_obj.rect(18, top_y, width - 36, 10, fill=1, stroke=0)
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.setFillColor(colors.HexColor("#64748b"))
    canvas_obj.drawRightString(width - 18, 12, f"Page {doc.page}")
    canvas_obj.drawString(18, 12, "Finacc ERP")
    canvas_obj.restoreState()


def _pdf_table(headers, rows, *, available_width, numeric_columns=None):
    sample_rows = rows[:20] if rows else []
    widths = []
    for index, header in enumerate(headers):
        lengths = [len(str(header or ""))]
        for row in sample_rows:
            if index < len(row):
                lengths.append(len(str(row[index] or "")))
        widths.append(max(lengths))
    total_weight = sum(widths) or 1
    col_widths = [max(36, min(available_width * (weight / total_weight), 150)) for weight in widths]
    total_width = sum(col_widths)
    if total_width > available_width:
        scale = available_width / total_width
        col_widths = [width * scale for width in col_widths]
    table_data = [headers] + rows
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.6),
        ("LEADING", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d9e2ef")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fbff")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
    ]))
    numeric_columns = set(numeric_columns or [])
    for row_index in range(1, len(table_data)):
        for col_index in range(len(headers)):
            table_style = "RIGHT" if col_index in numeric_columns else "LEFT"
            table.setStyle(TableStyle([("ALIGN", (col_index, row_index), (col_index, row_index), table_style)]))
    return table


def _write_customer_statement_pdf(payload, subtitle):
    sections = _statement_sheet_data(payload, display=True)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
        title="Customer Ledger Statement",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomerLedgerTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=21,
        textColor=colors.HexColor("#163b66"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "CustomerLedgerSubtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#475569"),
    )
    label_style = ParagraphStyle(
        "CustomerLedgerLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#1f2937"),
    )
    value_style = ParagraphStyle(
        "CustomerLedgerValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#334155"),
    )

    summary_data = sections["summary"]
    summary_rows = [[Paragraph(str(label), label_style), Paragraph(str(value or "-"), value_style)] for label, value in summary_data]
    summary_table = Table(summary_rows, colWidths=[150, 330], hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7fbff")),
        ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9e2ef")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    story = [
        Paragraph("Customer Ledger Statement", title_style),
        Paragraph(subtitle, subtitle_style),
        Spacer(1, 8),
        summary_table,
        Spacer(1, 10),
    ]

    available_width = landscape(A4)[0] - 36
    for title, (headers, rows) in [
        ("Open Items", sections["open_items"], [4, 5, 6]),
        ("Advances", sections["advances"], [4, 5, 6]),
        ("Settlements", sections["settlements"], [3]),
    ]:
        story.append(Paragraph(title, ParagraphStyle(
            f"CustomerLedger{title}",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=14,
            textColor=colors.HexColor("#163b66"),
            spaceBefore=4,
            spaceAfter=6,
        )))
        story.append(_pdf_table(headers, rows, available_width=available_width, numeric_columns=numeric_columns))
        story.append(Spacer(1, 10))

    doc.build(story, onFirstPage=_decorate_pdf, onLaterPages=_decorate_pdf)
    buffer.seek(0)
    return buffer.getvalue()


class _BaseCustomerStatementExportAPIView(APIView):
    permission_classes = [IsAuthenticated]
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response


class _CustomerStatementExportMixin(_BaseCustomerStatementExportAPIView):
    def report_data(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_ar_view_permission(user=request.user, entity_id=entity_id)
        customer_value = request.query_params.get("customer")
        if not customer_value:
            raise ValueError("customer query param is required.")
        try:
            customer_id = int(customer_value)
        except (TypeError, ValueError):
            raise ValueError("customer must be an integer.")
        include_closed = str(request.query_params.get("include_closed") or "").lower() in ("1", "true", "yes", "y")

        payload = _build_customer_statement_payload(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            customer_id=customer_id,
            include_closed=include_closed,
        )
        scope_names = resolve_scope_names(entity_id, entityfinid_id, subentity_id)
        payload["financial_year_name"] = scope_names.get("entityfin_name")
        payload["subentity_name"] = scope_names.get("subentity_name")
        payload["include_closed"] = include_closed
        subtitle = _statement_subtitle(scope_names=scope_names, payload=payload, include_closed=include_closed)
        return payload, subtitle


class CustomerStatementExcelAPIView(_CustomerStatementExportMixin):
    def get(self, request):
        payload, subtitle = self.report_data(request)
        content = _write_customer_statement_excel(payload, subtitle)
        return self.export_response(
            filename=f"CustomerLedger_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class CustomerStatementCSVAPIView(_CustomerStatementExportMixin):
    def get(self, request):
        payload, subtitle = self.report_data(request)
        content = _write_customer_statement_csv(payload, subtitle)
        return self.export_response(
            filename=f"CustomerLedger_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class CustomerStatementPDFAPIView(_CustomerStatementExportMixin):
    def get(self, request):
        payload, subtitle = self.report_data(request)
        content = _write_customer_statement_pdf(payload, subtitle)
        return self.export_response(
            filename=f"CustomerLedger_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class CustomerStatementPrintAPIView(CustomerStatementPDFAPIView):
    export_mode = "inline"
