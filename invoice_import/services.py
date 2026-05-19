from __future__ import annotations

import csv
import io
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook

from catalog.models import Product
from entity.models import Entity, EntityFinancialYear, SubEntity
from financial.models import account
from invoice_import.models import ImportJob, ImportProfile, ImportRow
from posting.models import Entry, InventoryMove
from posting.adapters.sales_invoice import SalesInvoicePostingAdapter, SalesInvoicePostingConfig
from posting.adapters.purchase_invoice import PurchaseInvoicePostingAdapter, PurchaseInvoicePostingConfig
from numbering.models import DocumentType
from numbering.services.document_number_service import DocumentNumberService
from purchase.models import PurchaseInvoiceHeader, PurchaseInvoiceLine, PurchaseTaxSummary
from purchase.models.purchase_ap import VendorBillOpenItem
from purchase.services.purchase_ap_service import q2 as ap_q2
from purchase.services.purchase_invoice_service import PurchaseInvoiceService
from purchase.services.purchase_settings_service import PurchaseSettingsService
from sales.models import SalesInvoiceHeader, SalesInvoiceLine, SalesTaxSummary
from sales.models.sales_ar import CustomerBillOpenItem
from sales.services.sales_ar_service import q2 as ar_q2
from sales.services.sales_invoice_service import SalesInvoiceService

ZERO2 = Decimal("0.00")
ZERO4 = Decimal("0.0000")
Q2 = Decimal("0.01")
Q4 = Decimal("0.0001")
SHEET = "invoices"
ERROR_SHEET = "invoice_import_errors"
DOCUMENT_NUMBER_STRATEGY_PRESERVE_LEGACY = "preserve_legacy"
DOCUMENT_NUMBER_STRATEGY_GENERATE_FINACC = "generate_finacc"


def q2(value: Any) -> Decimal:
    return Decimal(value or 0).quantize(Q2)


def q4(value: Any) -> Decimal:
    return Decimal(value or 0).quantize(Q4)


def _normalize_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _to_decimal(value: Any, default: Decimal = ZERO2) -> Decimal:
    if value in (None, "", "-", "--"):
        return default
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def _to_int(value: Any) -> int | None:
    text = _normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid integer value: {value}") from exc


def _to_bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return _normalize_text(value).lower() in {"1", "true", "yes", "y"}


def _to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _normalize_text(value)
    parsed = parse_date(text)
    if parsed is not None:
        return parsed
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date value: {value}")


def _read_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    actual = next((name for name in wb.sheetnames if str(name).strip().lower() == SHEET), wb.sheetnames[0] if wb.sheetnames else None)
    if not actual:
        return []
    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    payload: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        payload.append({headers[i]: row[i] for i in range(len(headers))})
    return payload


def _read_csv_zip(content: bytes) -> list[dict[str, Any]]:
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        lookup = {str(name).strip().lower(): name for name in zf.namelist()}
        actual = lookup.get(f"{SHEET}.csv")
        if not actual:
            return []
        data = zf.read(actual).decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(io.StringIO(data)) if any((v or "").strip() for v in row.values())]


def _write_xlsx(rows: list[dict[str, Any]], *, sheet_name: str = SHEET) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(col) for col in headers])
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _write_csv_zip(rows: list[dict[str, Any]], *, sheet_name: str = SHEET) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        stream = io.StringIO()
        headers = list(rows[0].keys()) if rows else []
        writer = csv.DictWriter(stream, fieldnames=headers)
        if headers:
            writer.writeheader()
            writer.writerows(rows)
        zf.writestr(f"{sheet_name}.csv", stream.getvalue())
    return buff.getvalue()


def _parse_rows(file_bytes: bytes, fmt: str) -> list[dict[str, Any]]:
    if fmt == ImportJob.FileFormat.XLSX:
        return _read_xlsx(file_bytes)
    return _read_csv_zip(file_bytes)


def _render_rows(rows: list[dict[str, Any]], fmt: str, *, sheet_name: str = SHEET) -> bytes:
    if fmt == ImportJob.FileFormat.XLSX:
        return _write_xlsx(rows, sheet_name=sheet_name)
    return _write_csv_zip(rows, sheet_name=sheet_name)


def _normalized_key(value: Any) -> str:
    return _normalize_text(value).lower().replace(" ", "_")


def _profile_snapshot(profile: ImportProfile | None) -> dict[str, Any]:
    if profile is None:
        return {}
    return {
        "id": profile.id,
        "name": profile.name,
        "module": profile.module,
        "source_system": profile.source_system,
        "mapping": profile.mapping or {},
        "options": profile.options or {},
    }


def _apply_import_profile(rows: list[dict[str, Any]], profile: ImportProfile | None, *, source_system: str) -> list[dict[str, Any]]:
    if profile is None:
        return rows

    mapping = profile.mapping or {}
    defaults = dict(mapping.get("defaults") or {})
    raw_source_map = mapping.get("source_to_canonical") or {}
    value_maps = mapping.get("value_maps") or {}

    source_to_canonical: dict[str, str] = {}
    for source_key, canonical in raw_source_map.items():
        canonical_name = _normalize_text(canonical)
        if not canonical_name:
            continue
        source_to_canonical[_normalized_key(source_key)] = canonical_name

    canonical_value_maps: dict[str, dict[str, Any]] = {}
    for field_name, field_map in value_maps.items():
        canonical_value_maps[_normalize_text(field_name)] = {
            _normalized_key(key): value for key, value in dict(field_map or {}).items()
        }

    transformed: list[dict[str, Any]] = []
    for row in rows:
        out = dict(defaults)
        if "legacy_source_system" not in out and source_system:
            out["legacy_source_system"] = source_system
        for source_key, value in row.items():
            normalized_source = _normalized_key(source_key)
            target = source_to_canonical.get(normalized_source)
            if not target:
                direct = _normalize_text(source_key)
                if direct:
                    target = direct
            if not target:
                continue
            field_value = value
            field_map = canonical_value_maps.get(target)
            if field_map:
                mapped = field_map.get(_normalized_key(value))
                if mapped is not None:
                    field_value = mapped
            out[target] = field_value
        transformed.append(out)
    return transformed


def _sales_doc_code(doc_type: int) -> str:
    if int(doc_type) == int(SalesInvoiceHeader.DocType.CREDIT_NOTE):
        return "SCN"
    if int(doc_type) == int(SalesInvoiceHeader.DocType.DEBIT_NOTE):
        return "SDN"
    return "SINV"


def _purchase_doc_code(doc_type: int) -> str:
    if int(doc_type) == int(PurchaseInvoiceHeader.DocType.CREDIT_NOTE):
        return "PCN"
    if int(doc_type) == int(PurchaseInvoiceHeader.DocType.DEBIT_NOTE):
        return "PDN"
    return "PINV"


def _template_row(*, module: str, mode: str, detail_level: str) -> dict[str, Any]:
    common = {
        "entityfinid_id": 1,
        "subentity_id": "",
        "legacy_source_system": "legacy_erp",
        "legacy_source_key": "INV-2025-0001",
        "doc_type": "invoice",
        "status": "posted",
        "source_invoice_number": "INV-2025-0001",
        "bill_date": "2025-04-01",
        "due_date": "2025-04-30",
        "settled_amount": "400.00",
        "outstanding_amount": "600.00",
        "reference": "Carry forward",
        "remarks": "Imported legacy invoice",
    }
    if module == ImportJob.Module.SALES:
        common.update(
            {
                "party_account_code": 5001,
        "party_name": "Alpha Retail",
        "total_discount": "0.00",
        "party_gstin": "27AAAAA9999A1Z5",
                "party_state_code": "27",
                "seller_gstin": "27BBBBB1111B1Z5",
                "seller_state_code": "27",
                "supply_category": int(SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B),
                "taxability": int(SalesInvoiceHeader.Taxability.TAXABLE),
                "tax_regime": int(SalesInvoiceHeader.TaxRegime.INTRA_STATE),
                "total_taxable": "1000.00",
                "total_cgst": "90.00",
                "total_sgst": "90.00",
                "total_igst": "0.00",
                "total_cess": "0.00",
                "round_off": "0.00",
                "grand_total": "1180.00",
                "original_source_key": "",
            }
        )
        if detail_level == ImportJob.DetailLevel.HEADER_PLUS_LINES:
            common.update(
                {
                    "line_no": 1,
                    "product_id": "",
                    "product_code": "WIDGET-001",
                    "product_name": "Widget",
                    "sales_account_id": "",
                    "product_desc": "Imported line item",
                    "is_service": False,
                    "uom_id": "",
                    "hsn_sac_code": "9983",
                    "qty": "10.000",
                    "free_qty": "0.000",
                    "rate": "100.0000",
                    "discount_type": 0,
                    "discount_percent": "0.0000",
                    "discount_amount": "0.00",
                    "gst_rate": "18.00",
                    "cess_percent": "0.00",
                    "taxable_value": "1000.00",
                    "cgst_amount": "90.00",
                    "sgst_amount": "90.00",
                    "igst_amount": "0.00",
                    "cess_amount": "0.00",
                    "line_total": "1180.00",
                }
            )
    else:
        common.update(
            {
                "party_account_code": 6001,
                "party_name": "Vendor One",
                "total_discount": "0.00",
                "party_gstin": "27CCCCC2222C1Z5",
                "supplier_invoice_number": "SUP-2025-0001",
                "supplier_invoice_date": "2025-04-01",
                "supply_category": int(PurchaseInvoiceHeader.SupplyCategory.DOMESTIC),
                "taxability": int(PurchaseInvoiceHeader.Taxability.TAXABLE),
                "tax_regime": int(PurchaseInvoiceHeader.TaxRegime.INTRA),
                "total_taxable": "1000.00",
                "total_cgst": "90.00",
                "total_sgst": "90.00",
                "total_igst": "0.00",
                "total_cess": "0.00",
                "round_off": "0.00",
                "grand_total": "1180.00",
                "tds_amount": "0.00",
                "gst_tds_amount": "0.00",
                "original_source_key": "",
            }
        )
        if detail_level == ImportJob.DetailLevel.HEADER_PLUS_LINES:
            common.update(
                {
                    "line_no": 1,
                    "product_id": "",
                    "product_code": "RM-001",
                    "product_name": "Raw Material One",
                    "purchase_account_id": "",
                    "product_desc": "Imported purchase line",
                    "is_service": False,
                    "purchase_behavior": "inventory",
                    "uom_id": "",
                    "hsn_sac_code": "7203",
                    "qty": "10.0000",
                    "free_qty": "0.0000",
                    "rate": "100.00",
                    "discount_type": "N",
                    "discount_percent": "0.00",
                    "discount_amount": "0.00",
                    "gst_rate": "18.00",
                    "cess_percent": "0.00",
                    "taxable_value": "1000.00",
                    "cgst_amount": "90.00",
                    "sgst_amount": "90.00",
                    "igst_amount": "0.00",
                    "cess_amount": "0.00",
                    "line_total": "1180.00",
                }
            )
    return common


def build_template_content(*, module: str, mode: str, detail_level: str, fmt: str) -> bytes:
    return _render_rows([_template_row(module=module, mode=mode, detail_level=detail_level)], fmt)


@dataclass
class RowValidationResult:
    normalized: dict[str, Any]
    errors: list[dict[str, Any]]
    warnings: list[str]


def _parse_doc_type(module: str, value: Any) -> int:
    text = _normalize_text(value).lower()
    if text in {"", "invoice", "tax_invoice", "1"}:
        return 1
    if text in {"credit_note", "credit", "cn", "2"}:
        return 2
    if text in {"debit_note", "debit", "dn", "3"}:
        return 3
    raise ValueError("Unsupported doc_type.")


def _parse_status(module: str, value: Any, *, default_posted: bool = True) -> int:
    text = _normalize_text(value).lower()
    if not text:
        return 3 if default_posted else 1
    if text in {"draft", "1"}:
        return 1
    if text in {"confirmed", "2"}:
        return 2
    if text in {"posted", "3"}:
        return 3
    if text in {"cancelled", "canceled", "9"}:
        return 9
    raise ValueError("Unsupported status.")


def _resolve_entityfin(entity: Entity, value: Any) -> EntityFinancialYear | None:
    entityfinid = _to_int(value)
    if not entityfinid:
        return None
    return EntityFinancialYear.objects.filter(pk=entityfinid, entity=entity).first()


def _resolve_subentity(entity: Entity, value: Any) -> SubEntity | None:
    subentity_id = _to_int(value)
    if not subentity_id:
        return None
    return SubEntity.objects.filter(pk=subentity_id, entity=entity).first()


def _resolve_party(entity: Entity, code: Any, name: Any) -> account | None:
    account_code = _to_int(code)
    if account_code:
        # Account code now lives on the linked Ledger model.
        party = account.objects.filter(entity=entity, ledger__ledger_code=account_code).first()
        if party:
            return party
    party_name = _normalize_text(name)
    if party_name:
        return account.objects.filter(entity=entity, accountname__iexact=party_name).first()
    return None


def _resolve_product(entity: Entity, product_id: Any, product_code: Any = None, product_name: Any = None) -> Product | None:
    pid = _to_int(product_id)
    if pid:
        product = Product.objects.filter(pk=pid, entity=entity).first()
        if product:
            return product

    code = _normalize_text(product_code)
    if code:
        product = Product.objects.filter(entity=entity, sku__iexact=code).first()
        if product:
            return product

    name = _normalize_text(product_name)
    if name:
        return Product.objects.filter(entity=entity, productname__iexact=name).first()
    return None


def _validate_row(*, job: ImportJob, row: dict[str, Any], row_no: int) -> RowValidationResult:
    errors: list[dict[str, Any]] = []
    warnings: list[str] = []

    try:
        doc_type = _parse_doc_type(job.module, row.get("doc_type"))
    except Exception as exc:
        errors.append({"field": "doc_type", "message": str(exc)})
        doc_type = 1

    try:
        status = _parse_status(job.module, row.get("status"))
    except Exception as exc:
        errors.append({"field": "status", "message": str(exc)})
        status = 1

    entityfin = _resolve_entityfin(job.entity, row.get("entityfinid_id"))
    if entityfin is None:
        errors.append({"field": "entityfinid_id", "message": "Valid entityfinid_id is required."})
    subentity = _resolve_subentity(job.entity, row.get("subentity_id"))
    if _normalize_text(row.get("subentity_id")) and subentity is None:
        errors.append({"field": "subentity_id", "message": "subentity_id does not belong to the entity."})

    source_key = _normalize_text(row.get("legacy_source_key"))
    if not source_key:
        errors.append({"field": "legacy_source_key", "message": "legacy_source_key is required."})

    source_system = _normalize_text(row.get("legacy_source_system") or job.source_system)
    if not source_system:
        errors.append({"field": "legacy_source_system", "message": "legacy_source_system is required."})

    source_invoice_number = _normalize_text(row.get("source_invoice_number"))
    if not source_invoice_number:
        errors.append({"field": "source_invoice_number", "message": "source_invoice_number is required."})

    try:
        bill_date = _to_date(row.get("bill_date"))
    except Exception as exc:
        errors.append({"field": "bill_date", "message": str(exc)})
        bill_date = None
    if bill_date is None:
        errors.append({"field": "bill_date", "message": "bill_date is required."})

    try:
        due_date = _to_date(row.get("due_date"))
    except Exception as exc:
        errors.append({"field": "due_date", "message": str(exc)})
        due_date = None

    party = _resolve_party(job.entity, row.get("party_account_code"), row.get("party_name"))
    if party is None:
        errors.append({"field": "party_account_code", "message": "Unable to resolve party_account_code/party_name."})

    try:
        total_taxable = q2(_to_decimal(row.get("total_taxable")))
        total_discount = q2(_to_decimal(row.get("total_discount")))
        total_cgst = q2(_to_decimal(row.get("total_cgst")))
        total_sgst = q2(_to_decimal(row.get("total_sgst")))
        total_igst = q2(_to_decimal(row.get("total_igst")))
        total_cess = q2(_to_decimal(row.get("total_cess")))
        round_off = q2(_to_decimal(row.get("round_off")))
        grand_total = q2(_to_decimal(row.get("grand_total")))
        settled_amount = q2(_to_decimal(row.get("settled_amount")))
        outstanding_amount = q2(_to_decimal(row.get("outstanding_amount")))
    except Exception as exc:
        errors.append({"field": "totals", "message": str(exc)})
        total_taxable = total_discount = total_cgst = total_sgst = total_igst = total_cess = round_off = grand_total = settled_amount = outstanding_amount = ZERO2

    if total_discount < ZERO2 or outstanding_amount < ZERO2 or settled_amount < ZERO2 or grand_total < ZERO2:
        errors.append({"field": "amounts", "message": "total_discount, grand_total, settled_amount, and outstanding_amount must be non-negative."})

    if job.mode == ImportJob.Mode.FULL_HISTORY and job.detail_level != ImportJob.DetailLevel.HEADER_PLUS_LINES:
        errors.append({"field": "detail_level", "message": "full_history requires header_plus_lines."})

    if job.stock_replay and job.mode != ImportJob.Mode.FULL_HISTORY:
        errors.append({"field": "stock_replay", "message": "stock_replay is allowed only for full_history jobs."})

    line_no = _to_int(row.get("line_no")) or row_no
    is_service = _to_bool(row.get("is_service"), default=False)
    product = _resolve_product(
        job.entity,
        row.get("product_id"),
        row.get("product_code"),
        row.get("product_name"),
    )
    if job.detail_level == ImportJob.DetailLevel.HEADER_PLUS_LINES:
        if not is_service and product is None and job.stock_replay:
            errors.append({"field": "product_id", "message": "product_id is required for stock replay on non-service lines."})
        if job.stock_replay and product is None and not is_service:
            errors.append({"field": "stock_replay", "message": "Historical stock replay requires product lines."})

    original_source_key = _normalize_text(row.get("original_source_key"))
    if doc_type in {2, 3} and not original_source_key:
        errors.append({"field": "original_source_key", "message": "original_source_key is required for credit/debit notes."})

    if job.compliance_mode == ImportJob.ComplianceMode.LIVE and job.module == ImportJob.Module.SALES:
        if not _normalize_text(row.get("seller_gstin")):
            errors.append({"field": "seller_gstin", "message": "seller_gstin is required when compliance_mode=live."})
        if int(row.get("supply_category") or SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B) == int(SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B):
            if not _normalize_text(row.get("party_gstin")):
                errors.append({"field": "party_gstin", "message": "party_gstin is required for live B2B compliance."})

    normalized = {
        "entityfinid_id": getattr(entityfin, "id", None),
        "subentity_id": getattr(subentity, "id", None),
        "legacy_source_system": source_system,
        "legacy_source_key": source_key,
        "doc_type": doc_type,
        "status": status,
        "source_invoice_number": source_invoice_number,
        "bill_date": bill_date.isoformat() if bill_date else None,
        "due_date": due_date.isoformat() if due_date else None,
        "party_id": getattr(party, "id", None),
        "party_name": _normalize_text(row.get("party_name")) or getattr(party, "accountname", ""),
        "party_gstin": _normalize_text(row.get("party_gstin")),
        "party_state_code": _normalize_text(row.get("party_state_code")),
        "seller_gstin": _normalize_text(row.get("seller_gstin")),
        "seller_state_code": _normalize_text(row.get("seller_state_code")),
        "supplier_invoice_number": _normalize_text(row.get("supplier_invoice_number")),
        "supplier_invoice_date": _to_date(row.get("supplier_invoice_date")).isoformat() if _normalize_text(row.get("supplier_invoice_date")) else None,
        "taxability": _to_int(row.get("taxability")),
        "supply_category": _to_int(row.get("supply_category")),
        "tax_regime": _to_int(row.get("tax_regime")),
        "reference": _normalize_text(row.get("reference")),
        "remarks": _normalize_text(row.get("remarks")),
        "original_source_key": original_source_key,
        "total_taxable": str(total_taxable),
        "total_discount": str(total_discount),
        "total_cgst": str(total_cgst),
        "total_sgst": str(total_sgst),
        "total_igst": str(total_igst),
        "total_cess": str(total_cess),
        "round_off": str(round_off),
        "grand_total": str(grand_total),
        "settled_amount": str(settled_amount),
        "outstanding_amount": str(outstanding_amount),
        "tds_amount": str(q2(_to_decimal(row.get("tds_amount")))),
        "gst_tds_amount": str(q2(_to_decimal(row.get("gst_tds_amount")))),
        "line_no": line_no,
        "product_id": getattr(product, "id", None),
        "product_code": _normalize_text(row.get("product_code")) or getattr(product, "sku", ""),
        "product_name": _normalize_text(row.get("product_name")) or getattr(product, "productname", ""),
        "sales_account_id": _to_int(row.get("sales_account_id")),
        "purchase_account_id": _to_int(row.get("purchase_account_id")),
        "product_desc": _normalize_text(row.get("product_desc")),
        "is_service": is_service,
        "purchase_behavior": _normalize_text(row.get("purchase_behavior")) or "inventory",
        "uom_id": _to_int(row.get("uom_id")),
        "hsn_sac_code": _normalize_text(row.get("hsn_sac_code")),
        "qty": str(q4(_to_decimal(row.get("qty"), default=ZERO4))),
        "free_qty": str(q4(_to_decimal(row.get("free_qty"), default=ZERO4))),
        "rate": str(_to_decimal(row.get("rate"), default=ZERO2)),
        "discount_type": row.get("discount_type"),
        "discount_percent": str(_to_decimal(row.get("discount_percent"), default=ZERO2)),
        "discount_amount": str(_to_decimal(row.get("discount_amount"), default=ZERO2)),
        "gst_rate": str(_to_decimal(row.get("gst_rate"), default=ZERO2)),
        "cess_percent": str(_to_decimal(row.get("cess_percent"), default=ZERO2)),
        "taxable_value": str(_to_decimal(row.get("taxable_value"), default=ZERO2)),
        "cgst_amount": str(_to_decimal(row.get("cgst_amount"), default=ZERO2)),
        "sgst_amount": str(_to_decimal(row.get("sgst_amount"), default=ZERO2)),
        "igst_amount": str(_to_decimal(row.get("igst_amount"), default=ZERO2)),
        "cess_amount": str(_to_decimal(row.get("cess_amount"), default=ZERO2)),
        "line_total": str(_to_decimal(row.get("line_total"), default=ZERO2)),
    }
    return RowValidationResult(normalized=normalized, errors=errors, warnings=warnings)


def _job_summary(job: ImportJob) -> dict[str, Any]:
    rows = job.rows.all()
    return {
        "rows_total": rows.count(),
        "rows_valid": rows.filter(status=ImportRow.Status.VALID).count(),
        "rows_error": rows.filter(status=ImportRow.Status.ERROR).count(),
        "groups_total": rows.values("group_key").distinct().count(),
    }


def _resolve_purchase_review_required(*, entity: Entity, rows: list[dict[str, Any]]) -> bool:
    subentity_ids: set[int] = set()
    for row in rows:
        raw = row.get("subentity_id")
        try:
            subentity_id = _to_int(raw)
        except ValueError:
            continue
        if subentity_id:
            subentity_ids.add(subentity_id)
    subentity_id = next(iter(subentity_ids)) if len(subentity_ids) == 1 else None
    policy = PurchaseSettingsService.get_policy(entity.id, subentity_id)
    return str(policy.controls.get("legacy_import_review_required", "off")).lower().strip() == "on"


def _purchase_policy_control_for_rows(rows: list[ImportRow], key: str, default: str) -> str:
    if not rows:
        return default
    subentity_ids: set[int] = set()
    for row in rows:
        raw = (row.normalized_payload or {}).get("subentity_id")
        try:
            subentity_id = _to_int(raw)
        except ValueError:
            continue
        if subentity_id:
            subentity_ids.add(subentity_id)
    subentity_id = next(iter(subentity_ids)) if len(subentity_ids) == 1 else None
    policy = PurchaseSettingsService.get_policy(rows[0].job.entity_id, subentity_id)
    return str(policy.controls.get(key, default)).lower().strip()


def _purchase_policy_decimal_for_rows(rows: list[ImportRow], key: str, default: str) -> Decimal:
    raw = _purchase_policy_control_for_rows(rows, key, default)
    try:
        return q2(_to_decimal(raw))
    except ValueError:
        return q2(_to_decimal(default))


def _append_group_error(rows: list[ImportRow], *, field: str, message: str) -> None:
    for row in rows:
        if any(err.get("field") == field and err.get("message") == message for err in (row.errors or [])):
            continue
        row.status = ImportRow.Status.ERROR
        row.errors = list(row.errors or []) + [{"field": field, "message": message}]
        row.save(update_fields=["status", "errors", "updated_at"])


def _append_group_warning(rows: list[ImportRow], message: str) -> None:
    for row in rows:
        if message in (row.warnings or []):
            continue
        row.warnings = list(row.warnings or []) + [message]
        row.save(update_fields=["warnings", "updated_at"])


def _format_note_review_preview(*, field: str, message: str, is_warning: bool) -> str:
    if not message:
        return ""
    normalized_field = _normalize_text(field).lower()
    normalized_message = message.lower()
    if "bill_date" in normalized_message or "earlier than original invoice" in normalized_message:
        return f"Date warning: {message}" if is_warning else f"Date issue: {message}"
    if "outstanding" in normalized_message and "original invoice" in normalized_message:
        return f"Outstanding warning: {message}" if is_warning else f"Outstanding issue: {message}"
    if normalized_field in {"original_source_key", "party_account_code", "grand_total", "total_taxable"}:
        return f"Original mismatch: {message}"
    if "referenced original invoice" in normalized_message or "original invoice" in normalized_message:
        return f"Original mismatch: {message}"
    return f"Review note: {message}" if is_warning else message


def _group_review_preview(rows: list[ImportRow]) -> str:
    first = rows[0].normalized_payload or {}
    is_purchase_note = int(first.get("doc_type") or 1) in {
        PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
        PurchaseInvoiceHeader.DocType.DEBIT_NOTE,
    }
    for row in rows:
        first_error = next(iter(row.errors or []), None)
        if isinstance(first_error, dict):
            field = _normalize_text(first_error.get("field"))
            message = _normalize_text(first_error.get("message"))
            if message:
                if is_purchase_note:
                    return _format_note_review_preview(field=field, message=message, is_warning=False)
                return message
    for row in rows:
        first_warning = next(iter(row.warnings or []), None)
        message = _normalize_text(first_warning)
        if message:
            if is_purchase_note:
                return _format_note_review_preview(field="", message=message, is_warning=True)
            return message
    return ""


def _compare_purchase_note_against_original(*, rows: list[ImportRow], note: dict[str, Any], original: dict[str, Any], source_label: str) -> None:
    amount_tolerance = _purchase_policy_decimal_for_rows(rows, "legacy_import_note_amount_tolerance", "0.00")
    note_party_id = int(note.get("party_id") or 0)
    original_party_id = int(original.get("party_id") or 0)
    if note_party_id and original_party_id and note_party_id != original_party_id:
        _append_group_error(rows, field="party_account_code", message=f"Credit/debit note vendor must match the referenced original invoice vendor from {source_label}.")

    note_grand_total = q2(abs(_to_decimal(note.get("grand_total"))))
    original_grand_total = q2(abs(_to_decimal(original.get("grand_total"))))
    if original_grand_total > ZERO2 and note_grand_total > (original_grand_total + amount_tolerance):
        _append_group_error(rows, field="grand_total", message=f"Credit/debit note grand_total {note_grand_total} cannot exceed original invoice grand_total {original_grand_total} from {source_label}.")

    note_taxable = q2(abs(_to_decimal(note.get("total_taxable"))))
    original_taxable = q2(abs(_to_decimal(original.get("total_taxable"))))
    if original_taxable > ZERO2 and note_taxable > (original_taxable + amount_tolerance):
        _append_group_error(rows, field="total_taxable", message=f"Credit/debit note taxable total {note_taxable} cannot exceed original invoice taxable total {original_taxable} from {source_label}.")

    note_bill_date = _to_date(note.get("bill_date"))
    original_bill_date = _to_date(original.get("bill_date"))
    if note_bill_date and original_bill_date and note_bill_date < original_bill_date:
        message = f"Credit/debit note bill_date {note_bill_date.isoformat()} is earlier than original invoice bill_date {original_bill_date.isoformat()} from {source_label}."
        date_rule = _purchase_policy_control_for_rows(rows, "legacy_import_note_date_rule", "warn")
        if date_rule == "hard":
            _append_group_error(rows, field="bill_date", message=message)
        elif date_rule == "warn":
            _append_group_warning(rows, message)

    note_outstanding = q2(abs(_to_decimal(note.get("outstanding_amount"))))
    original_outstanding = q2(abs(_to_decimal(original.get("outstanding_amount"))))
    if original_outstanding > ZERO2 and note_outstanding > original_outstanding:
        message = f"Credit/debit note outstanding {note_outstanding} exceeds original invoice outstanding {original_outstanding} from {source_label}."
        outstanding_rule = _purchase_policy_control_for_rows(rows, "legacy_import_note_outstanding_rule", "warn")
        if outstanding_rule == "hard":
            _append_group_error(rows, field="outstanding_amount", message=message)
        elif outstanding_rule == "warn":
            _append_group_warning(rows, message)


def _group_document_summary(group_key: str, rows: list[ImportRow]) -> dict[str, Any]:
    first = rows[0].normalized_payload or {}
    totals = {
        "line_discount": ZERO2,
        "line_taxable": ZERO2,
        "line_cgst": ZERO2,
        "line_sgst": ZERO2,
        "line_igst": ZERO2,
        "line_cess": ZERO2,
        "line_total": ZERO2,
    }
    for row in rows:
        data = row.normalized_payload or {}
        totals["line_discount"] += q2(_to_decimal(data.get("discount_amount")))
        totals["line_taxable"] += q2(_to_decimal(data.get("taxable_value")))
        totals["line_cgst"] += q2(_to_decimal(data.get("cgst_amount")))
        totals["line_sgst"] += q2(_to_decimal(data.get("sgst_amount")))
        totals["line_igst"] += q2(_to_decimal(data.get("igst_amount")))
        totals["line_cess"] += q2(_to_decimal(data.get("cess_amount")))
        totals["line_total"] += q2(_to_decimal(data.get("line_total")))
    original_source_key = _normalize_text(first.get("original_source_key"))
    original_summary: dict[str, Any] | None = None
    if original_source_key:
        original_header = PurchaseInvoiceHeader.objects.filter(
            entity_id=rows[0].job.entity_id,
            legacy_source_key=original_source_key,
        ).only(
            "purchase_number",
            "supplier_invoice_number",
            "bill_date",
            "grand_total",
            "total_taxable",
            "vendor_name",
        ).first() if rows and rows[0].job.module == ImportJob.Module.PURCHASE else None
        if original_header is not None:
            original_summary = {
                "legacy_source_key": original_source_key,
                "document_number": getattr(original_header, "purchase_number", "") or "",
                "supplier_invoice_number": getattr(original_header, "supplier_invoice_number", "") or "",
                "bill_date": getattr(original_header, "bill_date", None).isoformat() if getattr(original_header, "bill_date", None) else None,
                "party_name": getattr(original_header, "vendor_name", "") or "",
                "grand_total": str(q2(getattr(original_header, "grand_total", ZERO2) or ZERO2)),
                "total_taxable": str(q2(getattr(original_header, "total_taxable", ZERO2) or ZERO2)),
            }
        elif rows and rows[0].job.module == ImportJob.Module.PURCHASE:
            grouped = _all_rows_grouped(rows[0].job)
            original_rows = grouped.get(original_source_key) or []
            if original_rows:
                original_first = original_rows[0].normalized_payload or {}
                original_summary = {
                    "legacy_source_key": original_source_key,
                    "document_number": original_first.get("source_invoice_number") or "",
                    "supplier_invoice_number": original_first.get("supplier_invoice_number") or "",
                    "bill_date": original_first.get("bill_date"),
                    "party_name": original_first.get("party_name") or "",
                    "grand_total": original_first.get("grand_total") or "0.00",
                    "total_taxable": original_first.get("total_taxable") or "0.00",
                }

    return {
        "legacy_source_key": group_key,
        "document_number": first.get("source_invoice_number") or "",
        "doc_type": int(first.get("doc_type") or 1),
        "original_source_key": original_source_key,
        "original_document_summary": original_summary,
        "party_name": first.get("party_name") or "",
        "party_id": first.get("party_id"),
        "party_gstin": first.get("party_gstin") or "",
        "supplier_invoice_number": first.get("supplier_invoice_number") or "",
        "bill_date": first.get("bill_date"),
        "status": rows[0].status,
        "row_count": len(rows),
        "header_totals": {
            "total_taxable": first.get("total_taxable") or "0.00",
            "total_discount": first.get("total_discount") or "0.00",
            "total_cgst": first.get("total_cgst") or "0.00",
            "total_sgst": first.get("total_sgst") or "0.00",
            "total_igst": first.get("total_igst") or "0.00",
            "total_cess": first.get("total_cess") or "0.00",
            "tds_amount": first.get("tds_amount") or "0.00",
            "gst_tds_amount": first.get("gst_tds_amount") or "0.00",
            "round_off": first.get("round_off") or "0.00",
            "grand_total": first.get("grand_total") or "0.00",
            "settled_amount": first.get("settled_amount") or "0.00",
            "outstanding_amount": first.get("outstanding_amount") or "0.00",
        },
        "line_totals": {key: str(q2(value)) for key, value in totals.items()},
        "error_count": sum(len(row.errors or []) for row in rows),
        "warning_count": sum(len(row.warnings or []) for row in rows),
        "review_state": "needs_review" if any(row.errors or row.warnings for row in rows) else "ready",
        "review_preview": _group_review_preview(rows),
    }


def _apply_group_validation_rules(job: ImportJob) -> None:
    grouped = _rows_grouped(job)
    grouped_payloads = {group_key: (rows[0].normalized_payload or {}) for group_key, rows in grouped.items()}
    for group_key, rows in grouped.items():
        first = rows[0].normalized_payload or {}
        if any((row.normalized_payload or {}).get("doc_type") != first.get("doc_type") for row in rows):
            _append_group_error(rows, field="doc_type", message="All rows in a legacy document must use the same doc_type.")
        if any((row.normalized_payload or {}).get("source_invoice_number") != first.get("source_invoice_number") for row in rows):
            _append_group_error(rows, field="source_invoice_number", message="All rows in a legacy document must use the same source_invoice_number.")
        if any((row.normalized_payload or {}).get("party_id") != first.get("party_id") for row in rows):
            _append_group_error(rows, field="party_account_code", message="All rows in a legacy document must resolve to the same party.")

        if job.module != ImportJob.Module.PURCHASE:
            continue

        header_taxable = q2(_to_decimal(first.get("total_taxable")))
        header_discount = q2(_to_decimal(first.get("total_discount")))
        header_cgst = q2(_to_decimal(first.get("total_cgst")))
        header_sgst = q2(_to_decimal(first.get("total_sgst")))
        header_igst = q2(_to_decimal(first.get("total_igst")))
        header_cess = q2(_to_decimal(first.get("total_cess")))
        round_off = q2(_to_decimal(first.get("round_off")))
        grand_total = q2(_to_decimal(first.get("grand_total")))
        settled_amount = q2(_to_decimal(first.get("settled_amount")))
        outstanding_amount = q2(_to_decimal(first.get("outstanding_amount")))
        tds_amount = q2(_to_decimal(first.get("tds_amount")))
        gst_tds_amount = q2(_to_decimal(first.get("gst_tds_amount")))

        if settled_amount + outstanding_amount > grand_total:
            _append_group_error(rows, field="outstanding_amount", message="settled_amount + outstanding_amount cannot exceed grand_total.")
        if tds_amount < ZERO2 or gst_tds_amount < ZERO2:
            _append_group_error(rows, field="withholding_amounts", message="tds_amount and gst_tds_amount must be non-negative.")

        if int(first.get("doc_type") or 1) in {2, 3}:
            original_source_key = _normalize_text(first.get("original_source_key"))
            if not original_source_key:
                _append_group_error(rows, field="original_source_key", message="original_source_key is required for credit/debit notes.")
            else:
                original_in_job = grouped_payloads.get(original_source_key)
                if original_in_job:
                    _compare_purchase_note_against_original(
                        rows=rows,
                        note=first,
                        original=original_in_job,
                        source_label="this import job",
                    )
                else:
                    original_header = PurchaseInvoiceHeader.objects.filter(
                        entity=job.entity,
                        legacy_source_key=original_source_key,
                    ).only(
                        "vendor_id",
                        "bill_date",
                        "grand_total",
                        "total_taxable",
                    ).first()
                    if original_header:
                        _compare_purchase_note_against_original(
                            rows=rows,
                            note=first,
                            original={
                                "party_id": getattr(original_header, "vendor_id", None),
                                "bill_date": getattr(original_header, "bill_date", None),
                                "grand_total": getattr(original_header, "grand_total", ZERO2),
                                "total_taxable": getattr(original_header, "total_taxable", ZERO2),
                                "outstanding_amount": (getattr(original_header, "match_notes", {}) or {}).get("legacy_settlement", {}).get("outstanding_amount", "0.00"),
                            },
                            source_label="existing imported purchase invoice",
                        )
                    else:
                        _append_group_warning(rows, "Referenced original invoice will be validated again during commit.")

        if job.detail_level != ImportJob.DetailLevel.HEADER_PLUS_LINES:
            continue

        sum_discount = ZERO2
        sum_taxable = ZERO2
        sum_cgst = ZERO2
        sum_sgst = ZERO2
        sum_igst = ZERO2
        sum_cess = ZERO2
        sum_line_total = ZERO2
        for row in rows:
            data = row.normalized_payload or {}
            sum_discount += q2(_to_decimal(data.get("discount_amount")))
            sum_taxable += q2(_to_decimal(data.get("taxable_value")))
            sum_cgst += q2(_to_decimal(data.get("cgst_amount")))
            sum_sgst += q2(_to_decimal(data.get("sgst_amount")))
            sum_igst += q2(_to_decimal(data.get("igst_amount")))
            sum_cess += q2(_to_decimal(data.get("cess_amount")))
            sum_line_total += q2(_to_decimal(data.get("line_total")))

        if sum_taxable != header_taxable:
            _append_group_error(rows, field="total_taxable", message=f"Header total_taxable {header_taxable} does not match line taxable total {sum_taxable}.")
        if sum_cgst != header_cgst:
            _append_group_error(rows, field="total_cgst", message=f"Header total_cgst {header_cgst} does not match line CGST total {sum_cgst}.")
        if sum_sgst != header_sgst:
            _append_group_error(rows, field="total_sgst", message=f"Header total_sgst {header_sgst} does not match line SGST total {sum_sgst}.")
        if sum_igst != header_igst:
            _append_group_error(rows, field="total_igst", message=f"Header total_igst {header_igst} does not match line IGST total {sum_igst}.")
        if sum_cess != header_cess:
            _append_group_error(rows, field="total_cess", message=f"Header total_cess {header_cess} does not match line cess total {sum_cess}.")
        if header_discount and sum_discount != header_discount:
            _append_group_error(rows, field="total_discount", message=f"Header total_discount {header_discount} does not match line discount total {sum_discount}.")

        expected_grand_total = q2(sum_taxable + sum_cgst + sum_sgst + sum_igst + sum_cess + round_off)
        if expected_grand_total != grand_total:
            _append_group_error(rows, field="grand_total", message=f"grand_total {grand_total} does not match computed line total {expected_grand_total}.")
        if sum_line_total != grand_total:
            _append_group_warning(rows, f"Line total sum {sum_line_total} differs from header grand_total {grand_total}.")

@transaction.atomic
def create_validated_job(
    *,
    entity: Entity,
    user,
    module: str,
    mode: str,
    detail_level: str,
    stock_replay: bool,
    compliance_mode: str,
    withholding_mode: str,
    document_number_strategy: str = DOCUMENT_NUMBER_STRATEGY_PRESERVE_LEGACY,
    source_system: str,
    filename: str,
    fmt: str,
    file_bytes: bytes,
    profile: ImportProfile | None = None,
) -> ImportJob:
    effective_source_system = source_system or getattr(profile, "source_system", "")
    rows = _apply_import_profile(_parse_rows(file_bytes, fmt), profile, source_system=effective_source_system)
    review_required = module == ImportJob.Module.PURCHASE and _resolve_purchase_review_required(entity=entity, rows=rows)
    job = ImportJob.objects.create(
        entity=entity,
        created_by=user,
        profile=profile,
        module=module,
        mode=mode,
        detail_level=detail_level,
        stock_replay=stock_replay,
        compliance_mode=compliance_mode,
        withholding_mode=withholding_mode,
        source_system=effective_source_system,
        input_filename=filename,
        file_format=fmt,
        profile_snapshot=_profile_snapshot(profile),
        review_required=review_required,
        options={
            "stock_replay": stock_replay,
            "compliance_mode": compliance_mode,
            "withholding_mode": withholding_mode,
            "document_number_strategy": document_number_strategy,
        },
    )
    group_counts: dict[str, int] = defaultdict(int)
    for idx, row in enumerate(rows, start=2):
        result = _validate_row(job=job, row=row, row_no=idx)
        group_key = result.normalized.get("legacy_source_key") or f"row-{idx}"
        group_counts[group_key] += 1
        ImportRow.objects.create(
            job=job,
            row_no=idx,
            group_key=group_key,
            status=ImportRow.Status.ERROR if result.errors else ImportRow.Status.VALID,
            raw_payload=row,
            normalized_payload=result.normalized,
            errors=result.errors,
            warnings=result.warnings,
        )

    if detail_level == ImportJob.DetailLevel.HEADER_ONLY:
        for row in job.rows.all():
            if group_counts[row.group_key] > 1 and row.status != ImportRow.Status.ERROR:
                row.status = ImportRow.Status.ERROR
                row.errors = list(row.errors or []) + [{"field": "legacy_source_key", "message": "Duplicate legacy_source_key in header-only import."}]
                row.save(update_fields=["status", "errors", "updated_at"])

    if mode == ImportJob.Mode.OUTSTANDING_ONLY:
        for row in job.rows.all():
            outstanding = Decimal(str((row.normalized_payload or {}).get("outstanding_amount") or "0"))
            if outstanding <= ZERO2 and row.status != ImportRow.Status.ERROR:
                row.status = ImportRow.Status.ERROR
                row.errors = list(row.errors or []) + [{"field": "outstanding_amount", "message": "Outstanding-only imports require outstanding_amount > 0."}]
                row.save(update_fields=["status", "errors", "updated_at"])

    _apply_existing_collision_rules(job)
    _apply_group_validation_rules(job)
    summary = _job_summary(job)
    summary["document_summaries"] = [
        _group_document_summary(group_key, rows)
        for group_key, rows in sorted(_all_rows_grouped(job).items(), key=lambda item: item[0])
    ]
    job.summary = summary
    job.status = ImportJob.Status.VALIDATED if summary["rows_error"] == 0 else ImportJob.Status.FAILED
    job.save(update_fields=["summary", "status", "updated_at"])
    return job


def _apply_existing_collision_rules(job: ImportJob) -> None:
    document_number_strategy = str((job.options or {}).get("document_number_strategy") or DOCUMENT_NUMBER_STRATEGY_PRESERVE_LEGACY)
    for row in job.rows.all():
        normalized = row.normalized_payload or {}
        source_key = normalized.get("legacy_source_key")
        source_system = normalized.get("legacy_source_system")
        invoice_number = normalized.get("source_invoice_number")
        if not source_key or row.status == ImportRow.Status.ERROR:
            continue
        if job.module == ImportJob.Module.SALES:
            if SalesInvoiceHeader.objects.filter(
                entity=job.entity,
                legacy_source_system=source_system,
                legacy_source_key=source_key,
            ).exists():
                row.status = ImportRow.Status.ERROR
                row.errors = list(row.errors or []) + [{"field": "legacy_source_key", "message": "Legacy source key already imported for sales."}]
            elif SalesInvoiceHeader.objects.filter(entity=job.entity, invoice_number=invoice_number, is_legacy_imported=False).exists():
                row.status = ImportRow.Status.ERROR
                row.errors = list(row.errors or []) + [{"field": "source_invoice_number", "message": "Invoice number collides with an existing live sales invoice."}]
        else:
            if PurchaseInvoiceHeader.objects.filter(
                entity=job.entity,
                legacy_source_system=source_system,
                legacy_source_key=source_key,
            ).exists():
                row.status = ImportRow.Status.ERROR
                row.errors = list(row.errors or []) + [{"field": "legacy_source_key", "message": "Legacy source key already imported for purchase."}]
            elif (
                document_number_strategy == DOCUMENT_NUMBER_STRATEGY_PRESERVE_LEGACY
                and PurchaseInvoiceHeader.objects.filter(entity=job.entity, purchase_number=invoice_number, is_legacy_imported=False).exists()
            ):
                row.status = ImportRow.Status.ERROR
                row.errors = list(row.errors or []) + [{"field": "source_invoice_number", "message": "Invoice number collides with an existing live purchase invoice."}]
        if row.status == ImportRow.Status.ERROR:
            row.save(update_fields=["status", "errors", "updated_at"])


def _rows_grouped(job: ImportJob) -> dict[str, list[ImportRow]]:
    grouped: dict[str, list[ImportRow]] = defaultdict(list)
    for row in job.rows.exclude(status=ImportRow.Status.IMPORTED).order_by("group_key", "row_no"):
        grouped[row.group_key].append(row)
    return grouped


def _all_rows_grouped(job: ImportJob) -> dict[str, list[ImportRow]]:
    grouped: dict[str, list[ImportRow]] = defaultdict(list)
    for row in job.rows.order_by("group_key", "row_no"):
        grouped[row.group_key].append(row)
    return grouped


def _sales_status_token(status: int) -> int:
    return {
        1: SalesInvoiceHeader.Status.DRAFT,
        2: SalesInvoiceHeader.Status.CONFIRMED,
        3: SalesInvoiceHeader.Status.POSTED,
        9: SalesInvoiceHeader.Status.CANCELLED,
    }[int(status)]


def _purchase_status_token(status: int) -> int:
    return {
        1: PurchaseInvoiceHeader.Status.DRAFT,
        2: PurchaseInvoiceHeader.Status.CONFIRMED,
        3: PurchaseInvoiceHeader.Status.POSTED,
        9: PurchaseInvoiceHeader.Status.CANCELLED,
    }[int(status)]


def _sales_settlement_status(outstanding_amount: Decimal, settled_amount: Decimal) -> int:
    if outstanding_amount <= ZERO2:
        return SalesInvoiceHeader.SettlementStatus.SETTLED
    if settled_amount > ZERO2:
        return SalesInvoiceHeader.SettlementStatus.PARTIAL
    return SalesInvoiceHeader.SettlementStatus.OPEN


def _purchase_match_status(outstanding_amount: Decimal, settled_amount: Decimal) -> str:
    if outstanding_amount <= ZERO2:
        return "settled"
    if settled_amount > ZERO2:
        return "partial"
    return "open"


def _create_sales_header(job: ImportJob, normalized: dict[str, Any], *, original_invoice_id: int | None, user) -> SalesInvoiceHeader:
    grand_total = Decimal(normalized["grand_total"])
    settled_amount = Decimal(normalized["settled_amount"])
    outstanding_amount = Decimal(normalized["outstanding_amount"])
    header = SalesInvoiceHeader.objects.create(
        entity=job.entity,
        entityfinid_id=normalized["entityfinid_id"],
        subentity_id=normalized.get("subentity_id"),
        doc_type=int(normalized["doc_type"]),
        status=_sales_status_token(int(normalized["status"])),
        bill_date=_to_date(normalized["bill_date"]),
        posting_date=_to_date(normalized["bill_date"]),
        due_date=_to_date(normalized.get("due_date")),
        doc_code=_sales_doc_code(int(normalized["doc_type"])),
        invoice_number=normalized["source_invoice_number"],
        original_invoice_id=original_invoice_id,
        note_reason=SalesInvoiceHeader.NoteReason.OTHER if int(normalized["doc_type"]) in {2, 3} else None,
        affects_inventory=False,
        customer_id=normalized["party_id"],
        customer_ledger_id=getattr(account.objects.filter(pk=normalized["party_id"]).only("ledger_id").first(), "ledger_id", None),
        customer_name=normalized["party_name"],
        customer_gstin=normalized["party_gstin"],
        customer_state_code=normalized["party_state_code"],
        seller_gstin=normalized["seller_gstin"],
        seller_state_code=normalized["seller_state_code"],
        supply_category=int(normalized.get("supply_category") or SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B),
        taxability=int(normalized.get("taxability") or SalesInvoiceHeader.Taxability.TAXABLE),
        tax_regime=int(normalized.get("tax_regime") or SalesInvoiceHeader.TaxRegime.INTRA_STATE),
        is_igst=int(normalized.get("tax_regime") or SalesInvoiceHeader.TaxRegime.INTRA_STATE) == int(SalesInvoiceHeader.TaxRegime.INTER_STATE),
        total_taxable_value=Decimal(normalized["total_taxable"]),
        total_cgst=Decimal(normalized["total_cgst"]),
        total_sgst=Decimal(normalized["total_sgst"]),
        total_igst=Decimal(normalized["total_igst"]),
        total_cess=Decimal(normalized["total_cess"]),
        round_off=Decimal(normalized["round_off"]),
        grand_total=grand_total,
        settled_amount=settled_amount,
        outstanding_amount=outstanding_amount,
        settlement_status=_sales_settlement_status(outstanding_amount, settled_amount),
        reference=normalized.get("reference", ""),
        remarks=normalized.get("remarks", ""),
        gst_compliance_mode=SalesInvoiceHeader.GstComplianceMode.NONE,
        is_einvoice_applicable=False,
        is_eway_applicable=False,
        is_legacy_imported=True,
        legacy_import_job_id=job.id,
        legacy_source_system=normalized["legacy_source_system"],
        legacy_source_key=normalized["legacy_source_key"],
        legacy_import_mode=job.mode,
        legacy_behavior_flags={
            "detail_level": job.detail_level,
            "stock_replay": job.stock_replay,
            "compliance_mode": job.compliance_mode,
        },
        confirmed_at=timezone.now() if int(normalized["status"]) >= 2 else None,
        posted_at=timezone.now() if int(normalized["status"]) >= 3 else None,
        cancelled_at=timezone.now() if int(normalized["status"]) == 9 else None,
        created_by=user,
        updated_by=user,
    )
    return header


def _create_purchase_header(job: ImportJob, normalized: dict[str, Any], *, original_invoice_id: int | None, user) -> PurchaseInvoiceHeader:
    grand_total = Decimal(normalized["grand_total"])
    settled_amount = Decimal(normalized["settled_amount"])
    outstanding_amount = Decimal(normalized["outstanding_amount"])
    document_number_strategy = str((job.options or {}).get("document_number_strategy") or DOCUMENT_NUMBER_STRATEGY_PRESERVE_LEGACY)
    purchase_number = normalized["source_invoice_number"]
    doc_no = None
    if document_number_strategy == DOCUMENT_NUMBER_STRATEGY_GENERATE_FINACC:
        dt = DocumentType.objects.filter(
            module="purchase",
            default_code=_purchase_doc_code(int(normalized["doc_type"])),
            is_active=True,
        ).first()
        if dt is None:
            raise ValueError("DocumentType not found for purchase document number allocation.")
        allocated = DocumentNumberService.allocate_final(
            entity_id=job.entity_id,
            entityfinid_id=normalized["entityfinid_id"],
            subentity_id=normalized.get("subentity_id"),
            doc_type_id=dt.id,
            doc_code=_purchase_doc_code(int(normalized["doc_type"])),
            on_date=_to_date(normalized["bill_date"]),
        )
        doc_no = allocated.doc_no
        purchase_number = allocated.display_no

    header = PurchaseInvoiceHeader.objects.create(
        entity=job.entity,
        entityfinid_id=normalized["entityfinid_id"],
        subentity_id=normalized.get("subentity_id"),
        doc_type=int(normalized["doc_type"]),
        status=_purchase_status_token(int(normalized["status"])),
        bill_date=_to_date(normalized["bill_date"]),
        posting_date=_to_date(normalized["bill_date"]),
        due_date=_to_date(normalized.get("due_date")),
        doc_code=_purchase_doc_code(int(normalized["doc_type"])),
        doc_no=doc_no,
        purchase_number=purchase_number,
        supplier_invoice_number=normalized.get("supplier_invoice_number") or normalized["source_invoice_number"],
        supplier_invoice_date=_to_date(normalized.get("supplier_invoice_date")),
        ref_document_id=original_invoice_id,
        note_reason=PurchaseInvoiceHeader.NoteReason.OTHER if int(normalized["doc_type"]) in {2, 3} else None,
        affects_inventory=False,
        vendor_id=normalized["party_id"],
        vendor_ledger_id=getattr(account.objects.filter(pk=normalized["party_id"]).only("ledger_id").first(), "ledger_id", None),
        vendor_name=normalized["party_name"],
        vendor_gstin=normalized["party_gstin"],
        supply_category=int(normalized.get("supply_category") or PurchaseInvoiceHeader.SupplyCategory.DOMESTIC),
        default_taxability=int(normalized.get("taxability") or PurchaseInvoiceHeader.Taxability.TAXABLE),
        tax_regime=int(normalized.get("tax_regime") or PurchaseInvoiceHeader.TaxRegime.INTRA),
        is_igst=int(normalized.get("tax_regime") or PurchaseInvoiceHeader.TaxRegime.INTRA) == int(PurchaseInvoiceHeader.TaxRegime.INTER),
        total_taxable=Decimal(normalized["total_taxable"]),
        total_cgst=Decimal(normalized["total_cgst"]),
        total_sgst=Decimal(normalized["total_sgst"]),
        total_igst=Decimal(normalized["total_igst"]),
        total_cess=Decimal(normalized["total_cess"]),
        total_gst=q2(Decimal(normalized["total_cgst"]) + Decimal(normalized["total_sgst"]) + Decimal(normalized["total_igst"]) + Decimal(normalized["total_cess"])),
        round_off=Decimal(normalized["round_off"]),
        grand_total=grand_total,
        grand_total_base_currency=grand_total,
        tds_amount=Decimal(normalized.get("tds_amount") or "0.00"),
        gst_tds_amount=Decimal(normalized.get("gst_tds_amount") or "0.00"),
        custom_fields_json={},
        match_notes={
            "legacy_settlement": {
                "settled_amount": str(settled_amount),
                "outstanding_amount": str(outstanding_amount),
            }
        },
        is_legacy_imported=True,
        legacy_import_job_id=job.id,
        legacy_source_system=normalized["legacy_source_system"],
        legacy_source_key=normalized["legacy_source_key"],
        legacy_import_mode=job.mode,
        legacy_behavior_flags={
            "detail_level": job.detail_level,
            "stock_replay": job.stock_replay,
            "withholding_mode": job.withholding_mode,
            "document_number_strategy": document_number_strategy,
        },
        confirmed_at=timezone.now() if int(normalized["status"]) >= 2 else None,
        posted_at=timezone.now() if int(normalized["status"]) >= 3 else None,
        cancelled_at=timezone.now() if int(normalized["status"]) == 9 else None,
        created_by=user,
    )
    header.updated_at = timezone.now()
    header.save(update_fields=["updated_at"])
    return header


def _create_sales_lines(header: SalesInvoiceHeader, rows: list[ImportRow]) -> list[SalesInvoiceLine]:
    lines: list[SalesInvoiceLine] = []
    for idx, row in enumerate(rows, start=1):
        n = row.normalized_payload
        line = SalesInvoiceLine.objects.create(
            header=header,
            entity=header.entity,
            entityfinid_id=header.entityfinid_id,
            subentity_id=header.subentity_id,
            line_no=int(n.get("line_no") or idx),
            product_id=n.get("product_id"),
            productDesc=n.get("product_desc") or "",
            uom_id=n.get("uom_id"),
            hsn_sac_code=n.get("hsn_sac_code") or "",
            is_service=bool(n.get("is_service")),
            qty=Decimal(n.get("qty") or "0"),
            free_qty=Decimal(n.get("free_qty") or "0"),
            rate=Decimal(n.get("rate") or "0"),
            discount_type=int(n.get("discount_type") or SalesInvoiceLine.DiscountType.NONE),
            discount_percent=Decimal(n.get("discount_percent") or "0"),
            discount_amount=Decimal(n.get("discount_amount") or "0"),
            gst_rate=Decimal(n.get("gst_rate") or "0"),
            cess_percent=Decimal(n.get("cess_percent") or "0"),
            taxable_value=Decimal(n.get("taxable_value") or "0"),
            cgst_amount=Decimal(n.get("cgst_amount") or "0"),
            sgst_amount=Decimal(n.get("sgst_amount") or "0"),
            igst_amount=Decimal(n.get("igst_amount") or "0"),
            cess_amount=Decimal(n.get("cess_amount") or "0"),
            line_total=Decimal(n.get("line_total") or "0"),
            sales_account_id=n.get("sales_account_id"),
            created_by=header.created_by,
            updated_by=header.updated_by,
        )
        lines.append(line)
    return lines


def _create_purchase_lines(header: PurchaseInvoiceHeader, rows: list[ImportRow]) -> list[PurchaseInvoiceLine]:
    lines: list[PurchaseInvoiceLine] = []
    for idx, row in enumerate(rows, start=1):
        n = row.normalized_payload
        line = PurchaseInvoiceLine.objects.create(
            header=header,
            line_no=int(n.get("line_no") or idx),
            product_id=n.get("product_id"),
            purchase_account_id=n.get("purchase_account_id"),
            product_desc=n.get("product_desc") or "",
            is_service=bool(n.get("is_service")),
            purchase_behavior=n.get("purchase_behavior") or "inventory",
            hsn_sac=n.get("hsn_sac_code") or "",
            uom_id=n.get("uom_id"),
            qty=Decimal(n.get("qty") or "0"),
            free_qty=Decimal(n.get("free_qty") or "0"),
            rate=Decimal(n.get("rate") or "0"),
            discount_type=n.get("discount_type") or PurchaseInvoiceLine.DiscountType.NONE,
            discount_percent=Decimal(n.get("discount_percent") or "0"),
            discount_amount=Decimal(n.get("discount_amount") or "0"),
            gst_rate=Decimal(n.get("gst_rate") or "0"),
            cess_percent=Decimal(n.get("cess_percent") or "0"),
            taxable_value=Decimal(n.get("taxable_value") or "0"),
            cgst_amount=Decimal(n.get("cgst_amount") or "0"),
            sgst_amount=Decimal(n.get("sgst_amount") or "0"),
            igst_amount=Decimal(n.get("igst_amount") or "0"),
            cess_amount=Decimal(n.get("cess_amount") or "0"),
            line_total=Decimal(n.get("line_total") or "0"),
            taxability=int(n.get("taxability") or PurchaseInvoiceHeader.Taxability.TAXABLE),
            created_at=timezone.now(),
        )
        lines.append(line)
    return lines


def _rebuild_sales_tax_summary(header: SalesInvoiceHeader) -> None:
    SalesTaxSummary.objects.filter(header=header).delete()
    buckets: dict[tuple[Any, ...], dict[str, Decimal]] = defaultdict(lambda: {
        "taxable_value": ZERO2,
        "cgst_amount": ZERO2,
        "sgst_amount": ZERO2,
        "igst_amount": ZERO2,
        "cess_amount": ZERO2,
    })
    for line in header.lines.all():
        key = (
            int(header.taxability or SalesInvoiceHeader.Taxability.TAXABLE),
            line.hsn_sac_code or "",
            bool(line.is_service),
            Decimal(line.gst_rate or 0),
            bool(header.is_reverse_charge),
        )
        buckets[key]["taxable_value"] += Decimal(line.taxable_value or 0)
        buckets[key]["cgst_amount"] += Decimal(line.cgst_amount or 0)
        buckets[key]["sgst_amount"] += Decimal(line.sgst_amount or 0)
        buckets[key]["igst_amount"] += Decimal(line.igst_amount or 0)
        buckets[key]["cess_amount"] += Decimal(line.cess_amount or 0)
    objs = []
    for (taxability, hsn, is_service, gst_rate, is_reverse_charge), values in buckets.items():
        objs.append(
            SalesTaxSummary(
                header=header,
                entity=header.entity,
                entityfinid_id=header.entityfinid_id,
                subentity_id=header.subentity_id,
                taxability=taxability,
                hsn_sac_code=hsn,
                is_service=is_service,
                gst_rate=gst_rate,
                is_reverse_charge=is_reverse_charge,
                taxable_value=q2(values["taxable_value"]),
                cgst_amount=q2(values["cgst_amount"]),
                sgst_amount=q2(values["sgst_amount"]),
                igst_amount=q2(values["igst_amount"]),
                cess_amount=q2(values["cess_amount"]),
            )
        )
    if objs:
        SalesTaxSummary.objects.bulk_create(objs)


def _rebuild_purchase_tax_summary(header: PurchaseInvoiceHeader) -> None:
    PurchaseTaxSummary.objects.filter(header=header).delete()
    PurchaseInvoiceService.rebuild_tax_summary(header)


def _maybe_create_sales_open_item(header: SalesInvoiceHeader) -> None:
    abs_outstanding = Decimal(header.outstanding_amount or 0)
    abs_settled = Decimal(header.settled_amount or 0)
    if header.status == SalesInvoiceHeader.Status.CANCELLED or abs_outstanding <= ZERO2:
        return
    sign = Decimal("-1") if int(header.doc_type) == int(SalesInvoiceHeader.DocType.CREDIT_NOTE) else Decimal("1")
    original = ar_q2(sign * q2(abs_outstanding + abs_settled))
    settled = ar_q2(sign * q2(abs_settled))
    outstanding = ar_q2(sign * q2(abs_outstanding))
    CustomerBillOpenItem.objects.update_or_create(
        header=header,
        defaults={
            "entity": header.entity,
            "entityfinid_id": header.entityfinid_id,
            "subentity_id": header.subentity_id,
            "customer_id": header.customer_id,
            "customer_ledger_id": header.customer_ledger_id,
            "doc_type": int(header.doc_type),
            "bill_date": header.bill_date,
            "due_date": header.due_date,
            "invoice_number": header.invoice_number,
            "customer_reference_number": header.reference,
            "original_amount": original,
            "gross_amount": original,
            "tds_collected": ZERO2,
            "gst_tds_collected": ar_q2(sign * q2(getattr(header, "tcs_amount", ZERO2))),
            "net_receivable_amount": original,
            "settled_amount": settled,
            "outstanding_amount": outstanding,
            "is_open": abs(outstanding) > ZERO2,
        },
    )


def _maybe_create_purchase_open_item(header: PurchaseInvoiceHeader) -> None:
    legacy_settlement = (header.match_notes or {}).get("legacy_settlement", {})
    abs_outstanding = Decimal(str(legacy_settlement.get("outstanding_amount") or "0"))
    abs_settled = Decimal(str(legacy_settlement.get("settled_amount") or "0"))
    if int(header.status) == int(PurchaseInvoiceHeader.Status.CANCELLED) or abs_outstanding <= ZERO2:
        return
    sign = Decimal("-1") if int(header.doc_type) == int(PurchaseInvoiceHeader.DocType.CREDIT_NOTE) else Decimal("1")
    original = ap_q2(sign * q2(abs_outstanding + abs_settled))
    settled = ap_q2(sign * q2(abs_settled))
    outstanding = ap_q2(sign * q2(abs_outstanding))
    VendorBillOpenItem.objects.update_or_create(
        header=header,
        defaults={
            "entity": header.entity,
            "entityfinid_id": header.entityfinid_id,
            "subentity_id": header.subentity_id,
            "vendor_id": header.vendor_id,
            "vendor_ledger_id": header.vendor_ledger_id,
            "doc_type": int(header.doc_type),
            "bill_date": header.bill_date,
            "due_date": header.due_date,
            "purchase_number": header.purchase_number,
            "supplier_invoice_number": header.supplier_invoice_number,
            "original_amount": original,
            "gross_amount": original,
            "tds_deducted": ap_q2(sign * q2(getattr(header, "tds_amount", ZERO2))),
            "gst_tds_deducted": ap_q2(sign * q2(getattr(header, "gst_tds_amount", ZERO2))),
            "net_payable_amount": original,
            "settled_amount": settled,
            "outstanding_amount": outstanding,
            "is_open": abs(outstanding) > ZERO2,
        },
    )


def _maybe_replay_sales_side_effects(job: ImportJob, header: SalesInvoiceHeader, rows: list[ImportRow], user) -> list[str]:
    warnings: list[str] = []
    if job.stock_replay and int(header.status) == int(SalesInvoiceHeader.Status.POSTED) and rows:
        SalesInvoicePostingAdapter.post_sales_invoice(
            header=header,
            lines=list(header.lines.all()),
            user_id=getattr(user, "id", None),
            config=SalesInvoicePostingConfig(post_inventory=True),
        )
        warnings.append("Historical stock replay used posting adapter, which also reconstructed ledger/inventory postings.")
    if job.compliance_mode == ImportJob.ComplianceMode.LIVE and int(header.status) in {int(SalesInvoiceHeader.Status.CONFIRMED), int(SalesInvoiceHeader.Status.POSTED)}:
        SalesInvoiceService._run_auto_compliance(header=header, user=user, stage="post" if int(header.status) == int(SalesInvoiceHeader.Status.POSTED) else "confirm")
    return warnings


def _maybe_replay_purchase_side_effects(job: ImportJob, header: PurchaseInvoiceHeader, rows: list[ImportRow], user) -> list[str]:
    warnings: list[str] = []
    if job.withholding_mode == ImportJob.WithholdingMode.RECOMPUTE_FINACC:
        before = {
            "tds_amount": str(header.tds_amount or ZERO2),
            "gst_tds_amount": str(header.gst_tds_amount or ZERO2),
        }
        PurchaseInvoiceService._apply_tds(header=header)
        PurchaseInvoiceService._apply_gst_tds(header=header)
        header.save(update_fields=[
            "tds_section", "tds_is_manual", "tds_rate", "tds_base_amount", "tds_amount", "tds_reason",
            "vendor_tds_declared", "vendor_tds_rate", "vendor_tds_base_amount", "vendor_tds_amount", "vendor_tds_notes",
            "gst_tds_enabled", "gst_tds_is_manual", "gst_tds_contract_ref", "gst_tds_reason",
            "gst_tds_rate", "gst_tds_base_amount", "gst_tds_cgst_amount", "gst_tds_sgst_amount",
            "gst_tds_igst_amount", "gst_tds_amount", "gst_tds_status",
            "vendor_gst_tds_declared", "vendor_gst_tds_rate", "vendor_gst_tds_base_amount",
            "vendor_gst_tds_cgst_amount", "vendor_gst_tds_sgst_amount", "vendor_gst_tds_igst_amount",
            "vendor_gst_tds_amount", "vendor_gst_tds_notes",
        ])
        after = {
            "tds_amount": str(header.tds_amount or ZERO2),
            "gst_tds_amount": str(header.gst_tds_amount or ZERO2),
        }
        if before != after:
            warnings.append(f"Withholding recomputed from legacy snapshot {before} to {after}.")
    if job.stock_replay and int(header.status) == int(PurchaseInvoiceHeader.Status.POSTED) and rows:
        PurchaseInvoicePostingAdapter.post_purchase_invoice(
            header=header,
            lines=list(header.lines.all()),
            user_id=getattr(user, "id", None),
            config=PurchaseInvoicePostingConfig(),
        )
        warnings.append("Historical stock replay used posting adapter, which also reconstructed ledger/inventory postings.")
    return warnings


@transaction.atomic
def mark_job_reviewed(*, job: ImportJob, user, note: str = "") -> ImportJob:
    if job.status != ImportJob.Status.VALIDATED:
        raise ValueError("Only validated jobs can be marked as reviewed.")
    job.reviewed_by = user
    job.reviewed_at = timezone.now()
    job.review_note = (note or "").strip()[:255]
    job.save(update_fields=["reviewed_by", "reviewed_at", "review_note", "updated_at"])
    return job


@transaction.atomic
def commit_job(*, job: ImportJob, user) -> ImportJob:
    if job.status not in {ImportJob.Status.VALIDATED, ImportJob.Status.PARTIAL, ImportJob.Status.COMMITTED}:
        raise ValueError("Only validated jobs can be committed.")
    if job.review_required and not job.reviewed_at:
        raise ValueError("This import job must be reviewed before commit.")
    groups = _rows_grouped(job)
    created = skipped = 0
    warnings: list[str] = []
    reconciliation_rows: list[dict[str, Any]] = []

    def _sort_key(items: tuple[str, list[ImportRow]]) -> tuple[int, str]:
        first = items[1][0].normalized_payload
        doc_type = int(first.get("doc_type") or 1)
        return (0 if doc_type == 1 else 1, items[0])

    for group_key, rows in sorted(groups.items(), key=_sort_key):
        if any(row.status == ImportRow.Status.ERROR for row in rows):
            skipped += 1
            continue
        if all(row.status == ImportRow.Status.IMPORTED for row in rows):
            skipped += 1
            continue
        first = rows[0].normalized_payload
        original_source_key = first.get("original_source_key")
        original_id = None
        if original_source_key:
            if job.module == ImportJob.Module.SALES:
                original = SalesInvoiceHeader.objects.filter(entity=job.entity, legacy_source_key=original_source_key).first()
            else:
                original = PurchaseInvoiceHeader.objects.filter(entity=job.entity, legacy_source_key=original_source_key).first()
            if original is None:
                for row in rows:
                    row.status = ImportRow.Status.ERROR
                    row.errors = list(row.errors or []) + [{"field": "original_source_key", "message": "Referenced original invoice was not found."}]
                    row.save(update_fields=["status", "errors", "updated_at"])
                continue
            if job.module == ImportJob.Module.PURCHASE and int(first.get("party_id") or 0) != int(getattr(original, "vendor_id", 0) or 0):
                for row in rows:
                    row.status = ImportRow.Status.ERROR
                    row.errors = list(row.errors or []) + [{"field": "party_account_code", "message": "Credit/debit note vendor must match the referenced original invoice vendor."}]
                    row.save(update_fields=["status", "errors", "updated_at"])
                continue
            if job.module == ImportJob.Module.PURCHASE:
                original_grand_total = q2(abs(getattr(original, "grand_total", ZERO2) or ZERO2))
                note_grand_total = q2(abs(_to_decimal(first.get("grand_total"))))
                if original_grand_total > ZERO2 and note_grand_total > original_grand_total:
                    for row in rows:
                        row.status = ImportRow.Status.ERROR
                        row.errors = list(row.errors or []) + [{"field": "grand_total", "message": "Credit/debit note grand_total cannot exceed the referenced original invoice grand_total."}]
                        row.save(update_fields=["status", "errors", "updated_at"])
                    continue

                original_taxable = q2(abs(getattr(original, "total_taxable", ZERO2) or ZERO2))
                note_taxable = q2(abs(_to_decimal(first.get("total_taxable"))))
                if original_taxable > ZERO2 and note_taxable > original_taxable:
                    for row in rows:
                        row.status = ImportRow.Status.ERROR
                        row.errors = list(row.errors or []) + [{"field": "total_taxable", "message": "Credit/debit note taxable total cannot exceed the referenced original invoice taxable total."}]
                        row.save(update_fields=["status", "errors", "updated_at"])
                    continue
            original_id = original.id

        if job.module == ImportJob.Module.SALES:
            header = _create_sales_header(job, first, original_invoice_id=original_id, user=user)
            if job.detail_level == ImportJob.DetailLevel.HEADER_PLUS_LINES:
                _create_sales_lines(header, rows)
                _rebuild_sales_tax_summary(header)
            _maybe_create_sales_open_item(header)
            row_warnings = _maybe_replay_sales_side_effects(job, header, rows, user)
            display_no = header.invoice_number
        else:
            header = _create_purchase_header(job, first, original_invoice_id=original_id, user=user)
            if job.detail_level == ImportJob.DetailLevel.HEADER_PLUS_LINES:
                _create_purchase_lines(header, rows)
                _rebuild_purchase_tax_summary(header)
            _maybe_create_purchase_open_item(header)
            row_warnings = _maybe_replay_purchase_side_effects(job, header, rows, user)
            display_no = header.purchase_number

        for row in rows:
            row.status = ImportRow.Status.IMPORTED
            row.committed_object_id = header.id
            row.warnings = list(row.warnings or []) + row_warnings
            row.save(update_fields=["status", "committed_object_id", "warnings", "updated_at"])
        warnings.extend(row_warnings)
        reconciliation_rows.append(
            {
                "legacy_source_key": group_key,
                "header_id": header.id,
                "document_number": display_no,
                "status": header.status,
                "outstanding_amount": str(first.get("outstanding_amount") or "0.00"),
            }
        )
        created += 1

    job.summary = {
        **(job.summary or {}),
        "groups_imported": created,
        "groups_skipped": skipped,
        "document_summaries": [
            _group_document_summary(group_key, rows)
            for group_key, rows in sorted(_all_rows_grouped(job).items(), key=lambda item: item[0])
        ],
    }
    job.reconciliation_summary = {
        "imported_documents": reconciliation_rows,
        "warning_count": len(warnings),
        "warnings": warnings,
        "ledger_postings_created": Entry.objects.filter(
            entity=job.entity,
            txn_id__in=[row.get("header_id") for row in reconciliation_rows],
        ).count(),
        "inventory_moves_created": InventoryMove.objects.filter(
            entity=job.entity,
            txn_id__in=[row.get("header_id") for row in reconciliation_rows],
        ).count(),
    }
    if job.rows.filter(status=ImportRow.Status.ERROR).exists():
        job.status = ImportJob.Status.PARTIAL if created else ImportJob.Status.FAILED
    else:
        job.status = ImportJob.Status.COMMITTED
    job.save(update_fields=["summary", "reconciliation_summary", "status", "updated_at"])
    return job


def export_job_errors(*, job: ImportJob, fmt: str) -> tuple[bytes, str, str]:
    rows = []
    for row in job.rows.filter(status=ImportRow.Status.ERROR).order_by("row_no"):
        for error in row.errors or []:
            rows.append(
                {
                    "row_no": row.row_no,
                    "group_key": row.group_key,
                    "field": error.get("field"),
                    "message": error.get("message"),
                }
            )
    if fmt == ImportJob.FileFormat.CSV:
        return _render_rows(rows, ImportJob.FileFormat.CSV, sheet_name=ERROR_SHEET), "application/zip", f"invoice_import_errors_{job.id}.zip"
    return _render_rows(rows, ImportJob.FileFormat.XLSX, sheet_name=ERROR_SHEET), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"invoice_import_errors_{job.id}.xlsx"
