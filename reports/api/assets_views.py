import csv
from io import BytesIO, StringIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from reports.schemas.assets_reports import (
    AssetEventReportScopeSerializer,
    AssetHistoryScopeSerializer,
    DepreciationScheduleScopeSerializer,
    FixedAssetRegisterScopeSerializer,
)
from reports.schemas.common import build_report_envelope
from reports.services.assets import build_asset_event_report, build_asset_history, build_depreciation_schedule, build_fixed_asset_register

ASSET_REPORT_DEFAULTS = {
    "default_page_size": 100,
    "decimal_places": 2,
    "show_zero_balances_default": False,
    "show_opening_balance_default": False,
    "enable_drilldown": True,
}


def _filtered_querydict(request, *, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    return params.urlencode()


def _attach_asset_actions(payload, request, *, export_base_path):
    query = _filtered_querydict(request, exclude=["page", "page_size"])
    payload["actions"]["can_print"] = True
    payload["actions"]["export_urls"] = {
        "excel": f"{export_base_path}excel/?{query}",
        "pdf": f"{export_base_path}pdf/?{query}",
        "csv": f"{export_base_path}csv/?{query}",
        "print": f"{export_base_path}print/?{query}",
    }
    return payload


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, right, border


def _write_excel(title, subtitle, headers, rows, *, numeric_columns):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font, header_fill, center, left, right, border = _workbook_styles()
    ws.append([title])
    ws.append([subtitle])
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(headers[col_idx - 1]) + 4, 16)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.row == header_row:
                continue
            cell.alignment = right if cell.column in numeric_columns else left
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_csv(headers, rows):
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8")


def _write_pdf(title, subtitle, headers, rows):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18, title=title)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 6), Paragraph(subtitle, styles["Normal"]), Spacer(1, 10)]
    table = Table([headers] + rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


class _BaseAssetReportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def build_envelope(self, *, report_code, report_name, payload, scope, request, export_base_path):
        response = build_report_envelope(
            report_code=report_code,
            report_name=report_name,
            payload=payload,
            filters=scope,
            defaults=ASSET_REPORT_DEFAULTS,
        )
        return _attach_asset_actions(response, request, export_base_path=export_base_path)


class FixedAssetRegisterAPIView(_BaseAssetReportAPIView):

    def get(self, request):
        serializer = FixedAssetRegisterScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_fixed_asset_register(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            as_of_date=scope.get("as_of_date"),
            category_id=scope.get("category"),
            status=scope.get("status"),
            search=scope.get("search"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", 100),
        )
        return Response(self.build_envelope(report_code="fixed_asset_register", report_name="Fixed Asset Register", payload=payload, scope=scope, request=request, export_base_path="/api/reports/fixed-assets/register/"))


class DepreciationScheduleAPIView(_BaseAssetReportAPIView):

    def get(self, request):
        serializer = DepreciationScheduleScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_depreciation_schedule(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            category_id=scope.get("category"),
            asset_id=scope.get("asset"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", 100),
        )
        return Response(self.build_envelope(report_code="depreciation_schedule", report_name="Depreciation Schedule", payload=payload, scope=scope, request=request, export_base_path="/api/reports/fixed-assets/depreciation-schedule/"))


class AssetEventReportAPIView(_BaseAssetReportAPIView):
    def get(self, request):
        serializer = AssetEventReportScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_asset_event_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            event_type=scope.get("event_type"),
            asset_id=scope.get("asset"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", 100),
        )
        return Response(self.build_envelope(report_code="fixed_asset_events", report_name="Fixed Asset Events", payload=payload, scope=scope, request=request, export_base_path="/api/reports/fixed-assets/events/"))


class AssetHistoryAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        serializer = AssetHistoryScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_asset_history(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            asset_id=scope["asset"],
        )
        return Response(payload)


class _BaseAssetExportAPIView(_BaseAssetReportAPIView):
    filename_prefix = "asset-report"
    media_type = "application/octet-stream"
    extension = "bin"
    inline = False

    def response(self, content):
        resp = HttpResponse(content, content_type=self.media_type)
        disposition = "inline" if self.inline else "attachment"
        resp["Content-Disposition"] = f'{disposition}; filename="{self.filename_prefix}.{self.extension}"'
        return resp


class _FixedAssetRegisterExportMixin(_BaseAssetExportAPIView):
    def get_payload(self, request):
        serializer = FixedAssetRegisterScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_fixed_asset_register(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            as_of_date=scope.get("as_of_date"),
            category_id=scope.get("category"),
            status=scope.get("status"),
            search=scope.get("search"),
            page=1,
            page_size=100000,
        )
        return scope, payload


class FixedAssetRegisterExcelAPIView(_FixedAssetRegisterExportMixin):
    filename_prefix = "fixed-asset-register"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Asset Code", "Asset Name", "Category", "Status", "Gross Block", "Accumulated Depreciation", "Impairment", "NBV", "Location", "Department", "Custodian"]
        rows = [
            [r["asset_code"], r["asset_name"], r["category_name"], r["status"], r["gross_block"], r["accumulated_depreciation"], r["impairment_amount"], r["net_book_value"], r["location_name"], r["department_name"], r["custodian_name"]]
            for r in payload["rows"]
        ]
        subtitle = f"Entity {scope['entity']} | As Of {scope.get('as_of_date') or ''}"
        return self.response(_write_excel("Fixed Asset Register", subtitle, headers, rows, numeric_columns={5, 6, 7, 8}))


class FixedAssetRegisterCSVAPIView(_FixedAssetRegisterExportMixin):
    filename_prefix = "fixed-asset-register"
    media_type = "text/csv"
    extension = "csv"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Asset Code", "Asset Name", "Category", "Status", "Gross Block", "Accumulated Depreciation", "Impairment", "NBV", "Location", "Department", "Custodian"]
        rows = [
            [r["asset_code"], r["asset_name"], r["category_name"], r["status"], r["gross_block"], r["accumulated_depreciation"], r["impairment_amount"], r["net_book_value"], r["location_name"], r["department_name"], r["custodian_name"]]
            for r in payload["rows"]
        ]
        return self.response(_write_csv(headers, rows))


class FixedAssetRegisterPDFAPIView(_FixedAssetRegisterExportMixin):
    filename_prefix = "fixed-asset-register"
    media_type = "application/pdf"
    extension = "pdf"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Asset Code", "Asset Name", "Category", "Status", "Gross", "Acc Dep", "Impair", "NBV"]
        rows = [[r["asset_code"], r["asset_name"], r["category_name"], r["status"], r["gross_block"], r["accumulated_depreciation"], r["impairment_amount"], r["net_book_value"]] for r in payload["rows"]]
        subtitle = f"Entity {scope['entity']} | As Of {scope.get('as_of_date') or ''}"
        return self.response(_write_pdf("Fixed Asset Register", subtitle, headers, rows))


class FixedAssetRegisterPrintAPIView(FixedAssetRegisterPDFAPIView):
    inline = True


class _DepreciationScheduleExportMixin(_BaseAssetExportAPIView):
    def get_payload(self, request):
        serializer = DepreciationScheduleScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_depreciation_schedule(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            category_id=scope.get("category"),
            asset_id=scope.get("asset"),
            page=1,
            page_size=100000,
        )
        return scope, payload


class DepreciationScheduleExcelAPIView(_DepreciationScheduleExportMixin):
    filename_prefix = "depreciation-schedule"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Run Code", "Asset Code", "Asset Name", "Category", "From", "To", "Depreciation", "Closing NBV", "Status"]
        rows = [[r["run_code"], r["asset_code"], r["asset_name"], r["category_name"], r["period_from"], r["period_to"], r["depreciation_amount"], r["closing_net_book_value"], r["run_status"]] for r in payload["rows"]]
        subtitle = f"Entity {scope['entity']} | From {scope.get('from_date') or ''} To {scope.get('to_date') or ''}"
        return self.response(_write_excel("Depreciation Schedule", subtitle, headers, rows, numeric_columns={7, 8}))


class DepreciationScheduleCSVAPIView(_DepreciationScheduleExportMixin):
    filename_prefix = "depreciation-schedule"
    media_type = "text/csv"
    extension = "csv"

    def get(self, request):
        _scope, payload = self.get_payload(request)
        headers = ["Run Code", "Asset Code", "Asset Name", "Category", "From", "To", "Depreciation", "Closing NBV", "Status"]
        rows = [[r["run_code"], r["asset_code"], r["asset_name"], r["category_name"], r["period_from"], r["period_to"], r["depreciation_amount"], r["closing_net_book_value"], r["run_status"]] for r in payload["rows"]]
        return self.response(_write_csv(headers, rows))


class DepreciationSchedulePDFAPIView(_DepreciationScheduleExportMixin):
    filename_prefix = "depreciation-schedule"
    media_type = "application/pdf"
    extension = "pdf"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Run", "Asset", "Category", "From", "To", "Dep", "NBV", "Status"]
        rows = [[r["run_code"], r["asset_code"], r["category_name"], r["period_from"], r["period_to"], r["depreciation_amount"], r["closing_net_book_value"], r["run_status"]] for r in payload["rows"]]
        subtitle = f"Entity {scope['entity']} | From {scope.get('from_date') or ''} To {scope.get('to_date') or ''}"
        return self.response(_write_pdf("Depreciation Schedule", subtitle, headers, rows))


class DepreciationSchedulePrintAPIView(DepreciationSchedulePDFAPIView):
    inline = True


class _AssetEventExportMixin(_BaseAssetExportAPIView):
    def get_payload(self, request):
        serializer = AssetEventReportScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        payload = build_asset_event_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            event_type=scope.get("event_type"),
            asset_id=scope.get("asset"),
            page=1,
            page_size=100000,
        )
        return scope, payload


class AssetEventExcelAPIView(_AssetEventExportMixin):
    filename_prefix = "fixed-asset-events"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Asset Code", "Asset Name", "Category", "Event Type", "Event Date", "Amount", "Subentity"]
        rows = [[r["asset_code"], r["asset_name"], r["category_name"], r["event_type"], r["event_date"], r["amount"], r["subentity_name"]] for r in payload["rows"]]
        subtitle = f"Entity {scope['entity']} | From {scope.get('from_date') or ''} To {scope.get('to_date') or ''}"
        return self.response(_write_excel("Fixed Asset Events", subtitle, headers, rows, numeric_columns={6}))


class AssetEventCSVAPIView(_AssetEventExportMixin):
    filename_prefix = "fixed-asset-events"
    media_type = "text/csv"
    extension = "csv"

    def get(self, request):
        _scope, payload = self.get_payload(request)
        headers = ["Asset Code", "Asset Name", "Category", "Event Type", "Event Date", "Amount", "Subentity"]
        rows = [[r["asset_code"], r["asset_name"], r["category_name"], r["event_type"], r["event_date"], r["amount"], r["subentity_name"]] for r in payload["rows"]]
        return self.response(_write_csv(headers, rows))


class AssetEventPDFAPIView(_AssetEventExportMixin):
    filename_prefix = "fixed-asset-events"
    media_type = "application/pdf"
    extension = "pdf"

    def get(self, request):
        scope, payload = self.get_payload(request)
        headers = ["Asset", "Category", "Event", "Date", "Amount"]
        rows = [[r["asset_code"], r["category_name"], r["event_type"], r["event_date"], r["amount"]] for r in payload["rows"]]
        subtitle = f"Entity {scope['entity']} | From {scope.get('from_date') or ''} To {scope.get('to_date') or ''}"
        return self.response(_write_pdf("Fixed Asset Events", subtitle, headers, rows))


class AssetEventPrintAPIView(AssetEventPDFAPIView):
    inline = True
