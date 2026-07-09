from __future__ import annotations

import csv
from datetime import date as date_cls, datetime as datetime_cls
from decimal import Decimal
from io import BytesIO, StringIO
from xml.sax.saxutils import escape

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape as reportlab_landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from core.entitlements import ScopedEntitlementMixin
from assets.models import DepreciationRun, FixedAsset
from payments.models.payment_core import PaymentVoucherHeader
from posting.models import Entry
from purchase.models.purchase_core import PurchaseInvoiceHeader
from receipts.models.receipt_core import ReceiptVoucherHeader
from reports.schemas.book_reports import CashbookScopeSerializer, DaybookScopeSerializer
from reports.schemas.common import build_report_envelope
from reports.api.financial.export_utils import ExportSection, write_sectioned_csv, write_sectioned_excel, write_sectioned_pdf
from reports.api.report_permissions import assert_any_report_permission
from reports.services.financial_hub_settings import (
    apply_amount_display_unit_override,
    financial_hub_amount_unit_label,
    format_financial_hub_amount,
    get_effective_cashbook_settings,
    get_effective_daybook_settings,
    get_financial_hub_settings_payload,
    get_visible_cashbook_columns,
    get_visible_daybook_columns,
)
from reports.services.financial.books import (
    BOOK_REPORT_DEFAULTS,
    _money,
    build_cashbook,
    build_daybook,
    build_daybook_entry_detail,
    resolve_posting_entry_for_document,
)
from sales.models.sales_core import SalesInvoiceHeader
from subscriptions.services import SubscriptionLimitCodes, SubscriptionService
from vouchers.models.voucher_core import VoucherHeader


class BookReportPermissionMixin:
    required_permission_codes: tuple[str, ...] = ()
    permission_denied_message = "You do not have permission to access this report."

    def enforce_report_permission(self, request, *, entity_id: int):
        assert_any_report_permission(
            user=request.user,
            entity_id=entity_id,
            required_permissions=self.required_permission_codes,
            message=self.permission_denied_message,
        )


class _BaseBookReportAPIView(BookReportPermissionMixin, ScopedEntitlementMixin, APIView):
    """Common utilities for thin report views that delegate accounting logic to services."""

    permission_classes = [permissions.IsAuthenticated]
    serializer_class = None
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get_scope(self, request):
        serializer = self.serializer_class(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        self.enforce_scope(
            request,
            entity_id=scope["entity"],
            entityfinid_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
        )
        self.enforce_report_permission(request, entity_id=scope["entity"])
        return scope

    def attach_pagination_links(self, request, payload):
        """Populate stable `next` and `previous` URLs for paginated report responses."""
        page = payload.get("page") or 1
        pages = payload.get("pages") or 0
        if pages <= 0:
            payload["next"] = None
            payload["previous"] = None
            return payload

        def build_url(target_page):
            params = request.GET.copy()
            params["page"] = str(target_page)
            return request.build_absolute_uri(f"{request.path}?{params.urlencode()}")

        payload["next"] = build_url(page + 1) if page < pages else None
        payload["previous"] = build_url(page - 1) if page > 1 else None
        return payload


def _safe_filename(value):
    text = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in str(value or "").strip())
    text = text.strip("._-")
    return text or "report"


def _format_scope_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime_cls):
        return value.date().strftime("%d %b %Y")
    if isinstance(value, date_cls):
        return value.strftime("%d %b %Y")
    text = str(value).strip()
    if not text:
        return ""
    try:
        return date_cls.fromisoformat(text[:10]).strftime("%d %b %Y")
    except ValueError:
        return text


def _filtered_querydict(request, *, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    return params.urlencode()


def _attach_export_actions(payload, request, *, export_base_path):
    query = _filtered_querydict(request, exclude=["page", "page_size", "orientation"])
    payload.setdefault("actions", {})
    payload["actions"]["can_print"] = True
    payload["actions"]["export_urls"] = {
        "excel": f"{export_base_path}excel/?{query}",
        "pdf": f"{export_base_path}pdf/?{query}",
        "csv": f"{export_base_path}csv/?{query}",
        "print": f"{export_base_path}print/?{query}",
        "excel_landscape": f"{export_base_path}excel/landscape/?{query}",
        "excel_portrait": f"{export_base_path}excel/portrait/?{query}",
        "pdf_landscape": f"{export_base_path}pdf/landscape/?{query}",
        "pdf_portrait": f"{export_base_path}pdf/portrait/?{query}",
    }
    payload["available_exports"] = ["excel", "pdf", "csv", "print", "excel_landscape", "excel_portrait", "pdf_landscape", "pdf_portrait"]
    return payload


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, right, border


def _write_csv(headers, rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def _write_excel(title, subtitle, headers, rows, *, numeric_columns=None, orientation="landscape"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    orientation = (orientation or "landscape").strip().lower()
    if orientation not in {"landscape", "portrait"}:
        orientation = "landscape"
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    header_font, header_fill, center, left, right, border = _workbook_styles()
    numeric_columns = set(numeric_columns or [])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = left
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).alignment = left

    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(rows, start=header_row + 1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = right if col_index in numeric_columns else left

    for col_index, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 14)
        for row in rows[:100]:
            if col_index - 1 < len(row):
                width = max(width, len(str(row[col_index - 1])) + 2)
        ws.column_dimensions[get_column_letter(col_index)].width = min(width, 40)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_pdf(title, subtitle, headers, rows, *, col_widths=None, meta_items=None, numeric_columns=None, pagesize=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize or reportlab_landscape(A4), rightMargin=18, leftMargin=18, topMargin=24, bottomMargin=18)
    styles = getSampleStyleSheet()
    title_style = styles["Title"].clone("FinancialPdfTitle")
    title_style.textColor = colors.HexColor("#FFFFFF")
    title_style.alignment = 1
    title_style.leading = 18
    title_style.fontSize = 16
    title_style.spaceAfter = 0

    subtitle_style = styles["BodyText"].clone("FinancialPdfSubtitle")
    subtitle_style.fontSize = 9
    subtitle_style.leading = 11
    subtitle_style.textColor = colors.HexColor("#4A5568")
    subtitle_style.spaceAfter = 0

    meta_label_style = styles["BodyText"].clone("FinancialPdfMetaLabel")
    meta_label_style.fontSize = 7.5
    meta_label_style.leading = 9
    meta_label_style.textColor = colors.HexColor("#5B6573")
    meta_label_style.spaceAfter = 0

    meta_value_style = styles["BodyText"].clone("FinancialPdfMetaValue")
    meta_value_style.fontSize = 9
    meta_value_style.leading = 11
    meta_value_style.textColor = colors.HexColor("#1F2937")
    meta_value_style.spaceAfter = 0

    story = []
    story.append(Table([[Paragraph(f"<b>{title}</b>", title_style)]], colWidths=[doc.width]))
    story[-1].setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#3851A6")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(Spacer(1, 10))
    story.append(Paragraph(subtitle, subtitle_style))
    story.append(Spacer(1, 10))

    if meta_items:
        meta_rows = []
        for label, value in meta_items:
            meta_rows.append([
                Paragraph(f"<b>{escape(str(label))}:</b>", meta_label_style),
                Paragraph(escape(str(value or "-")), meta_value_style),
            ])
        meta_table = Table(meta_rows, colWidths=[doc.width * 0.28, doc.width * 0.72])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F7FAFF")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E1F2")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 10))

    def wrap_text(value):
        return Paragraph(escape(str(value if value is not None else "-")), styles["BodyText"])

    table_data = [headers] + [[wrap_text(cell) for cell in row] for row in rows]
    if not col_widths:
        col_widths = [doc.width / len(headers)] * len(headers)
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3851A6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E1F2")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def _report_pagesize(orientation):
    return reportlab_landscape(A4) if str(orientation).strip().lower() == "landscape" else A4


def _daybook_export_table(report):
    headers = ["Transaction Date", "Voucher Date", "Voucher No", "Voucher Type", "Narration", "Reference", "Debit", "Credit", "Status", "Posted", "Source"]
    rows = []
    for row in report.get("results") or []:
        rows.append([
            row.get("transaction_date") or "",
            row.get("voucher_date") or "",
            row.get("voucher_number") or "",
            row.get("voucher_type_name") or row.get("voucher_type") or "",
            row.get("narration") or "",
            row.get("reference_number") or "",
            row.get("debit_total") or "0.00",
            row.get("credit_total") or "0.00",
            row.get("status_name") or row.get("status") or "",
            "Yes" if row.get("posted") else "No",
            row.get("source_module") or "",
        ])
    return headers, rows


DAYBOOK_COLUMN_DEFS = {
    "transaction_date": ("Transaction Date", lambda row, _settings: _format_scope_date(row.get("transaction_date")) or ""),
    "voucher_date": ("Voucher Date", lambda row, _settings: _format_scope_date(row.get("voucher_date")) or ""),
    "voucher_no": ("Voucher No", lambda row, _settings: row.get("voucher_number") or ""),
    "voucher_type": ("Voucher Type", lambda row, _settings: row.get("voucher_type_name") or row.get("voucher_type") or ""),
    "narration": ("Narration", lambda row, _settings: row.get("narration") or ""),
    "reference": ("Reference", lambda row, _settings: row.get("reference_number") or ""),
    "debit": ("Debit", lambda row, settings: format_financial_hub_amount(row.get("debit_total"), settings=settings)),
    "credit": ("Credit", lambda row, settings: format_financial_hub_amount(row.get("credit_total"), settings=settings)),
    "status": ("Status", lambda row, _settings: row.get("status_name") or row.get("status") or ""),
    "posted": ("Posted", lambda row, _settings: "Yes" if row.get("posted") else "No"),
    "source": ("Source", lambda row, _settings: row.get("source_module") or ""),
}


def _daybook_export_section(report, *, settings):
    visible_columns = get_visible_daybook_columns(settings)
    headers = [DAYBOOK_COLUMN_DEFS[key][0] for key in visible_columns]
    width_map = {
        "transaction_date": 78,
        "voucher_date": 78,
        "voucher_no": 98,
        "voucher_type": 98,
        "narration": 220,
        "reference": 120,
        "debit": 84,
        "credit": 84,
        "status": 84,
        "posted": 68,
        "source": 88,
    }
    rows = [
        [DAYBOOK_COLUMN_DEFS[key][1](row, settings) for key in visible_columns]
        for row in (report.get("results") or [])
    ]
    row_kinds = ["detail"] * len(rows)
    if report.get("results"):
        totals = report.get("totals") or {}
        total_values = {
            "transaction_date": "",
            "voucher_date": "",
            "voucher_no": "",
            "voucher_type": "",
            "narration": "Report Total",
            "reference": "",
            "debit": format_financial_hub_amount(totals.get("debit_total"), settings=settings),
            "credit": format_financial_hub_amount(totals.get("credit_total"), settings=settings),
            "status": "",
            "posted": "",
            "source": "",
        }
        rows.append([total_values.get(key, "") for key in visible_columns])
        row_kinds.append("final_total")
    numeric_columns = {
        index
        for index, key in enumerate(visible_columns)
        if key in {"debit", "credit"}
    }
    return ExportSection(
        title="Daybook Entries",
        headers=headers,
        rows=rows,
        row_kinds=row_kinds,
        numeric_columns=numeric_columns,
        col_widths=[width_map.get(key, 96) for key in visible_columns],
        empty_message="No vouchers found for the selected scope.",
    )


def _daybook_pdf_table(report, *, settings):
    visible_columns = get_visible_daybook_columns(settings)
    preferred_columns = [
        "transaction_date",
        "voucher_no",
        "voucher_type",
        "narration",
        "debit",
        "credit",
        "status",
    ]
    selected_columns = [key for key in preferred_columns if key in visible_columns] or visible_columns[:7]
    headers = [DAYBOOK_COLUMN_DEFS[key][0] for key in selected_columns]
    rows = [
        [DAYBOOK_COLUMN_DEFS[key][1](row, settings) for key in selected_columns]
        for row in (report.get("results") or [])
    ]
    numeric_columns = {
        index + 1
        for index, key in enumerate(selected_columns)
        if key in {"debit", "credit"}
    }
    width_map = {
        "transaction_date": 0.12,
        "voucher_no": 0.17,
        "voucher_type": 0.12,
        "narration": 0.31,
        "debit": 0.10,
        "credit": 0.10,
        "status": 0.08,
        "voucher_date": 0.12,
        "reference": 0.16,
        "posted": 0.08,
        "source": 0.10,
    }
    col_widths = [width_map.get(key, 0.12) for key in selected_columns]
    return headers, rows, numeric_columns, col_widths



def _posting_detail_subtitle(detail):
    return (
        f"Entry: {detail.get('entry_id') or '-'} | "
        f"Voucher: {detail.get('voucher_number') or '-'} | "
        f"Type: {detail.get('voucher_type_name') or detail.get('voucher_type') or '-'} | "
        f"Posting Date: {detail.get('posting_date') or '-'}"
    )


def _posting_detail_meta_items(detail, *, entity_label, entityfin_label, subentity_label):
    journal_lines = detail.get("lines") or []
    inventory_moves = detail.get("inventory_moves") or []
    debit_total = sum(Decimal(str(row.get("debit") or 0)) for row in journal_lines)
    credit_total = sum(Decimal(str(row.get("credit") or 0)) for row in journal_lines)
    return [
        ("Posting Entry ID", detail.get("entry_id") or "-"),
        ("Posting Date", detail.get("posting_date") or "-"),
        ("Voucher Date", detail.get("voucher_date") or "-"),
        ("Voucher No", detail.get("voucher_number") or "-"),
        ("Voucher Type", detail.get("voucher_type_name") or detail.get("voucher_type") or "-"),
        ("Status", detail.get("status_name") or detail.get("status") or "-"),
        ("Entity", entity_label or "Selected entity"),
        ("Financial Year", entityfin_label or "Current FY"),
        ("Subentity", subentity_label or "All subentities"),
        ("Created By", detail.get("created_by") or "-"),
        ("Narration", detail.get("narration") or "-"),
        ("Journal Lines", len(journal_lines)),
        ("Inventory Moves", len(inventory_moves)),
        ("Debit Total", _money(debit_total)),
        ("Credit Total", _money(credit_total)),
        ("Difference", _money(debit_total - credit_total)),
    ]


def _posting_detail_export_tables(detail):
    journal_headers = ["Account", "Ledger", "Description", "Debit", "Credit"]
    journal_rows = [
        [
            row.get("account_name") or "",
            row.get("ledger_name") or "",
            row.get("description") or "",
            row.get("debit") or "0.00",
            row.get("credit") or "0.00",
        ]
        for row in (detail.get("lines") or [])
    ]

    inventory_headers = [
        "Product", "Batch", "Move Type", "Nature", "Location", "Source", "Destination",
        "Qty", "UOM", "Base Qty", "Base UOM", "Unit Cost", "Base Unit Cost", "Inventory Value", "Reason"
    ]
    inventory_rows = [
        [
            row.get("product_name") or "",
            row.get("batch_number") or "",
            row.get("move_type_name") or row.get("move_type") or "",
            row.get("movement_nature_name") or row.get("movement_nature") or "",
            row.get("location_name") or "",
            row.get("source_location_name") or "",
            row.get("destination_location_name") or "",
            row.get("qty") or "0.0000",
            row.get("uom_name") or "",
            row.get("base_qty") or "0.0000",
            row.get("base_uom_name") or "",
            row.get("unit_cost") or "0.0000",
            row.get("base_unit_cost") or "",
            row.get("ext_cost") or "0.00",
            row.get("movement_reason") or "",
        ]
        for row in (detail.get("inventory_moves") or [])
    ]
    return (journal_headers, journal_rows), (inventory_headers, inventory_rows)


def _append_excel_meta_section(ws, meta_items, *, start_row, title, header_font, header_fill, center, left, border):
    column_count = 2
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=column_count)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = left
    title_cell.border = border

    row_index = start_row + 1
    for label, value in meta_items:
        label_cell = ws.cell(row=row_index, column=1, value=label)
        value_cell = ws.cell(row=row_index, column=2, value=value)
        label_cell.border = border
        value_cell.border = border
        label_cell.alignment = left
        value_cell.alignment = left
        row_index += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    return row_index + 1


def _append_excel_table_section(
    ws,
    *,
    start_row,
    title,
    headers,
    rows,
    numeric_columns=None,
    header_font=None,
    header_fill=None,
    center=None,
    left=None,
    right=None,
    border=None,
):
    numeric_columns = set(numeric_columns or [])
    column_count = max(len(headers), 1)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=column_count)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = left
    title_cell.border = border

    header_row = start_row + 1
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_index = header_row + 1
    for row in rows:
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = right if col_index in numeric_columns else left
        row_index += 1

    for col_index, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 14)
        for row in rows[:100]:
            if col_index - 1 < len(row):
                width = max(width, len(str(row[col_index - 1])) + 2)
        ws.column_dimensions[get_column_letter(col_index)].width = min(width, 34)

    return row_index + 1


def _write_posting_detail_excel(detail, subtitle, meta_items, *, orientation="landscape"):
    wb = Workbook()
    header_font, header_fill, center, left, right, border = _workbook_styles()

    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    orientation = (orientation or "landscape").strip().lower()
    if orientation not in {"landscape", "portrait"}:
        orientation = "landscape"
    summary_sheet.page_setup.orientation = orientation
    summary_sheet.page_setup.fitToWidth = 1
    summary_sheet.page_setup.fitToHeight = 0
    summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    summary_sheet.cell(row=1, column=1, value="Posting Detail")
    summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary_sheet.cell(row=1, column=1).alignment = left
    summary_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    summary_sheet.cell(row=2, column=1, value=subtitle)
    summary_sheet.cell(row=2, column=1).alignment = left
    _append_excel_meta_section(
        summary_sheet,
        meta_items,
        start_row=4,
        title="Posting Overview",
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        border=border,
    )

    (journal_headers, journal_rows), (inventory_headers, inventory_rows) = _posting_detail_export_tables(detail)

    journal_sheet = wb.create_sheet("Journal Lines")
    journal_sheet.page_setup.orientation = orientation
    journal_sheet.page_setup.fitToWidth = 1
    journal_sheet.page_setup.fitToHeight = 0
    journal_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(journal_headers))
    journal_sheet.cell(row=1, column=1, value="Posting Detail - Journal Lines")
    journal_sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
    journal_sheet.cell(row=1, column=1).alignment = left
    journal_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(journal_headers))
    journal_sheet.cell(row=2, column=1, value=subtitle)
    journal_sheet.cell(row=2, column=1).alignment = left
    _append_excel_table_section(
        journal_sheet,
        start_row=4,
        title="Journal Lines",
        headers=journal_headers,
        rows=journal_rows or [["", "", "No journal lines found", "", ""]],
        numeric_columns={4, 5},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
    )

    inventory_sheet = wb.create_sheet("Inventory Moves")
    inventory_sheet.page_setup.orientation = orientation
    inventory_sheet.page_setup.fitToWidth = 1
    inventory_sheet.page_setup.fitToHeight = 0
    inventory_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(inventory_headers))
    inventory_sheet.cell(row=1, column=1, value="Posting Detail - Inventory Moves")
    inventory_sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
    inventory_sheet.cell(row=1, column=1).alignment = left
    inventory_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(inventory_headers))
    inventory_sheet.cell(row=2, column=1, value=subtitle)
    inventory_sheet.cell(row=2, column=1).alignment = left
    _append_excel_table_section(
        inventory_sheet,
        start_row=4,
        title="Inventory Moves",
        headers=inventory_headers,
        rows=inventory_rows or [["", "", "", "", "", "", "", "", "", "", "", "", "", "No inventory moves found", ""]],
        numeric_columns={8, 10, 12, 13, 14},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_posting_detail_csv(detail, meta_items):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Posting Detail"])
    for label, value in meta_items:
        writer.writerow([label, value])
    writer.writerow([])

    (journal_headers, journal_rows), (inventory_headers, inventory_rows) = _posting_detail_export_tables(detail)

    writer.writerow(["Journal Lines"])
    writer.writerow(journal_headers)
    if journal_rows:
        writer.writerows(journal_rows)
    else:
        writer.writerow(["", "", "No journal lines found", "", ""])

    writer.writerow([])
    writer.writerow(["Inventory Moves"])
    writer.writerow(inventory_headers)
    if inventory_rows:
        writer.writerows(inventory_rows)
    else:
        writer.writerow(["", "", "", "", "", "", "", "", "", "", "", "", "", "No inventory moves found", ""])

    return buffer.getvalue().encode("utf-8-sig")


def _write_posting_detail_pdf(detail, subtitle, meta_items, *, pagesize=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize or reportlab_landscape(A4), rightMargin=18, leftMargin=18, topMargin=24, bottomMargin=18)
    styles = getSampleStyleSheet()

    title_style = styles["Title"].clone("PostingDetailPdfTitle")
    title_style.textColor = colors.white
    title_style.alignment = 1
    title_style.leading = 18
    title_style.fontSize = 16

    subtitle_style = styles["BodyText"].clone("PostingDetailPdfSubtitle")
    subtitle_style.fontSize = 8
    subtitle_style.leading = 9
    subtitle_style.textColor = colors.HexColor("#4A5568")

    section_style = styles["Heading4"].clone("PostingDetailPdfSection")
    section_style.fontSize = 9
    section_style.leading = 10
    section_style.textColor = colors.HexColor("#1F2937")
    section_style.spaceAfter = 4

    meta_label_style = styles["BodyText"].clone("PostingDetailPdfMetaLabel")
    meta_label_style.fontSize = 7
    meta_label_style.leading = 8
    meta_label_style.textColor = colors.HexColor("#5B6573")

    meta_value_style = styles["BodyText"].clone("PostingDetailPdfMetaValue")
    meta_value_style.fontSize = 8
    meta_value_style.leading = 9
    meta_value_style.textColor = colors.HexColor("#1F2937")

    body_style = styles["BodyText"].clone("PostingDetailPdfBody")
    body_style.fontSize = 7.5
    body_style.leading = 8.5

    story = []
    story.append(Table([[Paragraph("<b>Posting Detail</b>", title_style)]], colWidths=[doc.width]))
    story[-1].setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#3851A6")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(Spacer(1, 5))
    story.append(Paragraph(escape(subtitle), subtitle_style))
    story.append(Spacer(1, 5))

    meta_map = {str(label): value for label, value in meta_items}
    compact_meta_rows = [
        [
            Paragraph(
                f"<b>Entry</b>: {escape(str(meta_map.get('Posting Entry ID') or '-'))}<br/>"
                f"<b>Posting</b>: {escape(str(meta_map.get('Posting Date') or '-'))}<br/>"
                f"<b>Voucher</b>: {escape(str(meta_map.get('Voucher No') or '-'))}",
                meta_value_style,
            ),
            Paragraph(
                f"<b>Type</b>: {escape(str(meta_map.get('Voucher Type') or '-'))}<br/>"
                f"<b>Status</b>: {escape(str(meta_map.get('Status') or '-'))}<br/>"
                f"<b>Voucher Date</b>: {escape(str(meta_map.get('Voucher Date') or '-'))}",
                meta_value_style,
            ),
            Paragraph(
                f"<b>Entity</b>: {escape(str(meta_map.get('Entity') or '-'))}<br/>"
                f"<b>FY</b>: {escape(str(meta_map.get('Financial Year') or '-'))}<br/>"
                f"<b>Subentity</b>: {escape(str(meta_map.get('Subentity') or '-'))}",
                meta_value_style,
            ),
        ],
        [
            Paragraph(
                f"<b>Debit</b>: {escape(str(meta_map.get('Debit Total') or '-'))}<br/>"
                f"<b>Credit</b>: {escape(str(meta_map.get('Credit Total') or '-'))}<br/>"
                f"<b>Difference</b>: {escape(str(meta_map.get('Difference') or '-'))}",
                meta_value_style,
            ),
            Paragraph(
                f"<b>Journal Lines</b>: {escape(str(meta_map.get('Journal Lines') or '-'))}<br/>"
                f"<b>Inventory Moves</b>: {escape(str(meta_map.get('Inventory Moves') or '-'))}<br/>"
                f"<b>Created By</b>: {escape(str(meta_map.get('Created By') or '-'))}",
                meta_value_style,
            ),
            Paragraph(
                f"<b>Narration</b>: {escape(str(meta_map.get('Narration') or '-'))}",
                meta_value_style,
            ),
        ],
    ]
    meta_table = Table(compact_meta_rows, colWidths=[doc.width * 0.24, doc.width * 0.24, doc.width * 0.52])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F7FAFF")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E1F2")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 7))

    def build_section_table(title, headers, rows, col_widths):
        story.append(Paragraph(title, section_style))
        table_rows = [headers]
        if rows:
            table_rows.extend([
                [Paragraph(escape(str(cell if cell is not None else "-")), body_style) for cell in row]
                for row in rows
            ])
        else:
            table_rows.append([Paragraph("No data found.", body_style)] + [""] * (len(headers) - 1))
        section_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
        section_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3851A6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E1F2")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.2),
            ("LEADING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(section_table)
        story.append(Spacer(1, 6))

    (journal_headers, journal_rows), (inventory_headers, inventory_rows) = _posting_detail_export_tables(detail)
    build_section_table("Journal Lines", journal_headers, journal_rows, [doc.width * 0.16, doc.width * 0.16, doc.width * 0.46, doc.width * 0.11, doc.width * 0.11])
    build_section_table(
        "Inventory Moves",
        inventory_headers,
        inventory_rows,
        [
            doc.width * 0.10,
            doc.width * 0.05,
            doc.width * 0.06,
            doc.width * 0.07,
            doc.width * 0.07,
            doc.width * 0.07,
            doc.width * 0.07,
            doc.width * 0.05,
            doc.width * 0.05,
            doc.width * 0.05,
            doc.width * 0.05,
            doc.width * 0.06,
            doc.width * 0.07,
            doc.width * 0.08,
            doc.width * 0.10,
        ],
    )
    doc.build(story)
    return buffer.getvalue()


class DaybookAPIView(_BaseBookReportAPIView):
    """Return Daybook rows derived from posting entries and journal totals."""

    serializer_class = DaybookScopeSerializer
    required_permission_codes = (
        "reports.financial_hub.daybook.view",
        "reports.daybook.view",
    )

    def get(self, request):
        scope = self.get_scope(request)
        try:
            data = build_daybook(
                entity_id=scope["entity"],
                entityfin_id=scope.get("entityfinid"),
                subentity_id=scope.get("subentity"),
                from_date=scope.get("from_date"),
                to_date=scope.get("to_date"),
                voucher_types=scope.get("voucher_types"),
                account_ids=scope.get("account_ids"),
                statuses=scope.get("statuses"),
                posted=scope.get("posted"),
                search=scope.get("search"),
                page=scope.get("page", BOOK_REPORT_DEFAULTS["default_page_size_page"]),
                page_size=scope.get("page_size", BOOK_REPORT_DEFAULTS["default_page_size"]),
            )
        except ValueError as exc:
            return Response(exc.args[0], status=400)
        self.attach_pagination_links(request, data)
        response = build_report_envelope(
            report_code="daybook",
            report_name="Daybook",
            payload=data,
            filters={
                "entity": scope["entity"],
                "entityfinid": scope.get("entityfinid"),
                "subentity": scope.get("subentity"),
                "scope_mode": scope.get("scope_mode"),
                "from_date": scope.get("from_date"),
                "to_date": scope.get("to_date"),
                "voucher_type": scope.get("voucher_types", []),
                "account": scope.get("account_ids", []),
                "status": scope.get("statuses", []),
                "posted": scope.get("posted"),
                "search": scope.get("search"),
                "page": scope.get("page", BOOK_REPORT_DEFAULTS["default_page_size_page"]),
                "page_size": scope.get("page_size", BOOK_REPORT_DEFAULTS["default_page_size"]),
            },
            defaults=BOOK_REPORT_DEFAULTS,
        )
        response = _attach_export_actions(response, request, export_base_path="/api/reports/financial/daybook/")
        return Response(response)


class DaybookEntryDetailAPIView(BookReportPermissionMixin, ScopedEntitlementMixin, APIView):
    """Return a journal-line drill-down payload for a single posting entry."""

    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL
    required_permission_codes = (
        "reports.financial_hub.daybook.view",
        "reports.daybook.view",
    )

    def get(self, request, entry_id: int):
        entity_id = request.query_params.get("entity")
        if not entity_id:
            return Response({"detail": "entity is required."}, status=400)
        entityfin_id = request.query_params.get("entityfinid")
        subentity_id = request.query_params.get("subentity")
        self.enforce_scope(
            request,
            entity_id=int(entity_id),
            entityfinid_id=int(entityfin_id) if entityfin_id else None,
            subentity_id=int(subentity_id) if subentity_id else None,
        )
        self.enforce_report_permission(request, entity_id=int(entity_id))
        try:
            data = build_daybook_entry_detail(
                entry_id=entry_id,
                entity_id=entity_id,
                entityfin_id=entityfin_id,
                subentity_id=subentity_id,
            )
        except Entry.DoesNotExist:
            return Response({"detail": "Entry not found."}, status=404)
        return Response(data)


class PostingDocumentLookupAPIView(BookReportPermissionMixin, ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL
    required_permission_codes = (
        "reports.financial_hub.ledger_book.view",
        "reports.ledger_book.view",
        "reports.financial_hub.daybook.view",
        "reports.daybook.view",
    )

    def get(self, request):
        entity_id = request.query_params.get("entity")
        document_type = request.query_params.get("document_type")
        document_id = request.query_params.get("document_id")
        entityfin_id = request.query_params.get("entityfinid")
        subentity_id = request.query_params.get("subentity")
        source_module = request.query_params.get("source_module")

        if not entity_id:
            return Response({"entity": ["This query parameter is required."]}, status=400)
        if not document_type:
            return Response({"document_type": ["This query parameter is required."]}, status=400)
        if not document_id:
            return Response({"document_id": ["This query parameter is required."]}, status=400)

        self.enforce_scope(
            request,
            entity_id=int(entity_id),
            entityfinid_id=int(entityfin_id) if entityfin_id else None,
            subentity_id=int(subentity_id) if subentity_id else None,
        )
        self.enforce_report_permission(request, entity_id=int(entity_id))
        try:
            data = resolve_posting_entry_for_document(
                entity_id=entity_id,
                entityfin_id=entityfin_id,
                subentity_id=subentity_id,
                document_type=document_type,
                document_id=document_id,
                source_module=source_module,
            )
        except ValueError as exc:
            return Response(exc.args[0] if exc.args else {"detail": "Invalid request."}, status=400)
        except Entry.DoesNotExist:
            return Response({"detail": "Posting entry not found for this document."}, status=404)
        except (
            FixedAsset.DoesNotExist,
            DepreciationRun.DoesNotExist,
            PurchaseInvoiceHeader.DoesNotExist,
            SalesInvoiceHeader.DoesNotExist,
            PaymentVoucherHeader.DoesNotExist,
            ReceiptVoucherHeader.DoesNotExist,
            VoucherHeader.DoesNotExist,
        ):
            return Response({"detail": "Document not found."}, status=404)
        return Response(data)


class _BasePostingDetailExportAPIView(BookReportPermissionMixin, ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL
    export_mode = "attachment"
    export_orientation = "landscape"
    required_permission_codes = (
        "reports.financial_hub.ledger_book.view",
        "reports.ledger_book.view",
        "reports.financial_hub.daybook.view",
        "reports.daybook.view",
    )

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def build_orientation(self, request):
        orientation = str(
            getattr(self, "export_orientation", None)
            or request.query_params.get("orientation")
            or "landscape"
        ).strip().lower()
        return orientation if orientation in {"landscape", "portrait"} else "landscape"

    def detail_data(self, request, entry_id: int):
        entity_id = request.query_params.get("entity")
        if not entity_id:
            raise ValueError({"detail": "entity is required."})
        entityfin_id = request.query_params.get("entityfinid")
        subentity_id = request.query_params.get("subentity")
        self.enforce_scope(
            request,
            entity_id=int(entity_id),
            entityfinid_id=int(entityfin_id) if entityfin_id else None,
            subentity_id=int(subentity_id) if subentity_id else None,
        )
        self.enforce_report_permission(request, entity_id=int(entity_id))
        detail = build_daybook_entry_detail(
            entry_id=entry_id,
            entity_id=entity_id,
            entityfin_id=entityfin_id,
            subentity_id=subentity_id,
        )
        entity_label = request.query_params.get("entity_name") or "Selected entity"
        entityfin_label = request.query_params.get("entityfin_name") or "Current FY"
        subentity_label = request.query_params.get("subentity_name") or ("All subentities" if not subentity_id else f"Subentity {subentity_id}")
        subtitle = _posting_detail_subtitle(detail)
        meta_items = _posting_detail_meta_items(
            detail,
            entity_label=entity_label,
            entityfin_label=entityfin_label,
            subentity_label=subentity_label,
        )
        return detail, subtitle, meta_items


class PostingDetailExcelAPIView(_BasePostingDetailExportAPIView):
    def get(self, request, entry_id: int):
        detail, subtitle, meta_items = self.detail_data(request, entry_id)
        content = _write_posting_detail_excel(
            detail,
            subtitle,
            meta_items,
            orientation=self.build_orientation(request),
        )
        return self.export_response(
            filename=f"Posting_Detail_{detail.get('entry_id')}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class PostingDetailCSVAPIView(_BasePostingDetailExportAPIView):
    def get(self, request, entry_id: int):
        detail, _subtitle, meta_items = self.detail_data(request, entry_id)
        content = _write_posting_detail_csv(detail, meta_items)
        return self.export_response(
            filename=f"Posting_Detail_{detail.get('entry_id')}.csv",
            content=content,
            content_type="text/csv",
        )


class PostingDetailPDFAPIView(_BasePostingDetailExportAPIView):
    def get(self, request, entry_id: int):
        detail, subtitle, meta_items = self.detail_data(request, entry_id)
        orientation = self.build_orientation(request)
        pagesize = _report_pagesize(orientation)
        content = _write_posting_detail_pdf(
            detail,
            subtitle,
            meta_items,
            pagesize=pagesize,
        )
        return self.export_response(
            filename=f"Posting_Detail_{detail.get('entry_id')}.pdf",
            content=content,
            content_type="application/pdf",
        )


class PostingDetailPrintAPIView(PostingDetailPDFAPIView):
    export_mode = "inline"


def _daybook_subtitle(scope, scope_names, report):
    posted = scope.get("posted")
    posted_label = "Posted only" if posted is True else "Non-posted only" if posted is False else "All entries"
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Period: {_format_scope_date(scope.get('from_date')) or '-'} to {_format_scope_date(scope.get('to_date')) or '-'} | "
        f"{posted_label} | "
        f"Transactions: {report.get('totals', {}).get('transaction_count', 0)}"
    )


class _BaseDaybookExportAPIView(_BaseBookReportAPIView):
    serializer_class = DaybookScopeSerializer
    export_mode = "attachment"
    export_orientation = "landscape"
    required_permission_codes = (
        "reports.financial_hub.daybook.view",
        "reports.daybook.view",
    )

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def build_orientation(self, request):
        orientation = str(
            getattr(self, "export_orientation", None)
            or request.query_params.get("orientation")
            or "landscape"
        ).strip().lower()
        return orientation if orientation in {"landscape", "portrait"} else "landscape"

    def report_data(self, request):
        scope = self.get_scope(request)
        try:
            data = build_daybook(
                entity_id=scope["entity"],
                entityfin_id=scope.get("entityfinid"),
                subentity_id=scope.get("subentity"),
                from_date=scope.get("from_date"),
                to_date=scope.get("to_date"),
                voucher_types=scope.get("voucher_types"),
                account_ids=scope.get("account_ids"),
                statuses=scope.get("statuses"),
                posted=scope.get("posted"),
                search=scope.get("search"),
                page=scope.get("page", BOOK_REPORT_DEFAULTS["default_page_size_page"]),
                page_size=scope.get("page_size", BOOK_REPORT_DEFAULTS["default_page_size"]),
            )
        except ValueError as exc:
            raise ValueError(exc.args[0]) from exc
        scope_names = {
            "entity_name": data.get("entity_name"),
            "entityfin_name": data.get("entityfin_name"),
            "subentity_name": data.get("subentity_name"),
        }
        subtitle = _daybook_subtitle(scope, scope_names, data)
        return scope, data, subtitle


class DaybookExcelAPIView(_BaseDaybookExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        settings_payload = get_financial_hub_settings_payload(user=request.user, entity_id=scope["entity"])
        effective = apply_amount_display_unit_override(
            get_effective_daybook_settings(settings_payload),
            scope.get("amount_display_unit"),
        )
        section = _daybook_export_section(data, settings=effective)
        summary_items = [
            ("Transactions", data.get("totals", {}).get("transaction_count", 0)),
            ("Debit Total", format_financial_hub_amount(data.get("totals", {}).get("debit_total"), settings=effective)),
            ("Credit Total", format_financial_hub_amount(data.get("totals", {}).get("credit_total"), settings=effective)),
        ]
        export_subtitle = f"{subtitle} | Display Unit: {financial_hub_amount_unit_label(effective)}"
        content = write_sectioned_excel(
            title="Daybook",
            subtitle=export_subtitle,
            summary_items=summary_items,
            sections=[section],
            orientation=self.build_orientation(request),
            freeze_header=(effective.get("export_layout") or {}).get("freeze_excel_header", True),
        )
        return self.export_response(
            filename=f"Daybook_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class DaybookCSVAPIView(_BaseDaybookExportAPIView):
    def get(self, request):
        scope, data, _subtitle = self.report_data(request)
        settings_payload = get_financial_hub_settings_payload(user=request.user, entity_id=scope["entity"])
        effective = apply_amount_display_unit_override(
            get_effective_daybook_settings(settings_payload),
            scope.get("amount_display_unit"),
        )
        section = _daybook_export_section(data, settings=effective)
        meta_items = [
            ("Entity", data.get("entity_name") or "Selected entity"),
            ("Financial Year", data.get("entityfin_name") or "Current FY"),
            ("Subentity", data.get("subentity_name") or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Period", f"{_format_scope_date(scope.get('from_date')) or '-'} to {_format_scope_date(scope.get('to_date')) or '-'}"),
            ("Posted", "Posted only" if scope.get("posted") is True else "Non-posted only" if scope.get("posted") is False else "All entries"),
            ("Display Unit", financial_hub_amount_unit_label(effective)),
            ("Transactions", data.get("totals", {}).get("transaction_count", 0)),
        ]
        content = write_sectioned_csv(title="Daybook", meta_items=meta_items, sections=[section])
        return self.export_response(
            filename=f"Daybook_{_safe_filename(_daybook_subtitle(scope, {'entity_name': data.get('entity_name'),'entityfin_name': data.get('entityfin_name'),'subentity_name': data.get('subentity_name')}, data))}.csv",
            content=content,
            content_type="text/csv",
        )


class DaybookPDFAPIView(_BaseDaybookExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        settings_payload = get_financial_hub_settings_payload(user=request.user, entity_id=scope["entity"])
        effective = apply_amount_display_unit_override(
            get_effective_daybook_settings(settings_payload),
            scope.get("amount_display_unit"),
        )
        meta_items = [
            ("Entity", data.get("entity_name") or "Selected entity"),
            ("Financial Year", data.get("entityfin_name") or "Current FY"),
            ("Subentity", data.get("subentity_name") or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Period", f"{_format_scope_date(scope.get('from_date')) or '-'} to {_format_scope_date(scope.get('to_date')) or '-'}"),
            ("Posted", "Posted only" if scope.get("posted") is True else "Non-posted only" if scope.get("posted") is False else "All entries"),
            ("Display Unit", financial_hub_amount_unit_label(effective)),
            ("Transactions", data.get("totals", {}).get("transaction_count", 0)),
            ("Debit Total", format_financial_hub_amount(data.get("totals", {}).get("debit_total"), settings=effective)),
            ("Credit Total", format_financial_hub_amount(data.get("totals", {}).get("credit_total"), settings=effective)),
        ]
        export_subtitle = f"{subtitle} | Display Unit: {financial_hub_amount_unit_label(effective)}"
        content = write_sectioned_pdf(
            title="Daybook",
            subtitle=export_subtitle,
            meta_items=meta_items,
            sections=[_daybook_export_section(data, settings=effective)],
            pagesize=_report_pagesize(self.build_orientation(request)),
            header_density=(effective.get("export_layout") or {}).get("header_density", "compact"),
            metadata_visibility=(effective.get("export_layout") or {}).get("metadata_visibility", "compact"),
        )
        return self.export_response(
            filename=f"Daybook_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class DaybookPrintAPIView(DaybookPDFAPIView):
    export_mode = "inline"


class DaybookExcelLandscapeAPIView(DaybookExcelAPIView):
    export_orientation = "landscape"


class DaybookExcelPortraitAPIView(DaybookExcelAPIView):
    export_orientation = "portrait"


class DaybookPDFLandscapeAPIView(DaybookPDFAPIView):
    export_orientation = "landscape"


class DaybookPDFPortraitAPIView(DaybookPDFAPIView):
    export_orientation = "portrait"


class CashbookAPIView(_BaseBookReportAPIView):
    """Return audit-safe Cashbook detail or summary output depending on account scope."""

    serializer_class = CashbookScopeSerializer
    required_permission_codes = (
        "reports.financial_hub.cashbook.view",
        "reports.cash_book.view",
        "reports.cashbook.view",
    )

    def get(self, request):
        scope = self.get_scope(request)
        try:
            data = build_cashbook(
                entity_id=scope["entity"],
                entityfin_id=scope.get("entityfinid"),
                subentity_id=scope.get("subentity"),
                from_date=scope.get("from_date"),
                to_date=scope.get("to_date"),
                mode=scope.get("mode", "both"),
                cash_account_ids=scope.get("cash_account_ids"),
                bank_account_ids=scope.get("bank_account_ids"),
                counter_account_ids=scope.get("counter_account_ids"),
                voucher_types=scope.get("voucher_types"),
                search=scope.get("search"),
                page=scope.get("page", BOOK_REPORT_DEFAULTS["default_page_size_page"]),
                page_size=scope.get("page_size", BOOK_REPORT_DEFAULTS["default_page_size"]),
            )
        except ValueError as exc:
            return Response(exc.args[0], status=400)
        self.attach_pagination_links(request, data)
        response = build_report_envelope(
            report_code="cashbook",
            report_name="Cashbook",
            payload=data,
            filters={
                "entity": scope["entity"],
                "entityfinid": scope.get("entityfinid"),
                "subentity": scope.get("subentity"),
                "scope_mode": scope.get("scope_mode"),
                "from_date": scope.get("from_date"),
                "to_date": scope.get("to_date"),
                "mode": scope.get("mode", "both"),
                "cash_account": scope.get("cash_account_ids", []),
                "bank_account": scope.get("bank_account_ids", []),
                "account": scope.get("counter_account_ids", []),
                "voucher_type": scope.get("voucher_types", []),
                "search": scope.get("search"),
                "page": scope.get("page", BOOK_REPORT_DEFAULTS["default_page_size_page"]),
                "page_size": scope.get("page_size", BOOK_REPORT_DEFAULTS["default_page_size"]),
            },
            defaults=BOOK_REPORT_DEFAULTS,
        )
        response = _attach_export_actions(response, request, export_base_path="/api/reports/financial/cashbook/")
        return Response(response)


def _cashbook_subtitle(scope, scope_names, report):
    mode = scope.get("mode", "both")
    mode_label = "Cash only" if mode == "cash" else "Bank only" if mode == "bank" else "Cash and bank"
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Period: {scope.get('from_date') or '-'} to {scope.get('to_date') or '-'} | "
        f"{mode_label} | "
        f"Transactions: {report.get('totals', {}).get('transaction_count', 0)}"
    )


def _cashbook_export_table(report):
    headers = [
        "Date",
        "Voucher No",
        "Voucher Type",
        "Account Impacted",
        "Counter Account",
        "Particulars",
        "Receipt",
        "Payment",
        "Running Balance",
        "Source",
    ]
    rows = []
    for row in report.get("results") or []:
        account_impacted = row.get("account_impacted") or {}
        rows.append([
            row.get("date") or "",
            row.get("voucher_number") or "",
            row.get("voucher_type_name") or row.get("voucher_type") or "",
            account_impacted.get("name") or "",
            row.get("counter_account") or "",
            row.get("particulars") or "",
            row.get("receipt_amount") or "0.00",
            row.get("payment_amount") or "0.00",
            row.get("running_balance") if row.get("running_balance") is not None else "",
            row.get("source_module") or "",
        ])
    return headers, rows


CASHBOOK_COLUMN_DEFS = {
    "date": ("Date", lambda row, _settings: row.get("date") or ""),
    "voucher_no": ("Voucher No", lambda row, _settings: row.get("voucher_number") or ""),
    "voucher_type": ("Voucher Type", lambda row, _settings: row.get("voucher_type_name") or row.get("voucher_type") or ""),
    "account_impacted": ("Account Impacted", lambda row, _settings: (row.get("account_impacted") or {}).get("name") or ""),
    "counter_account": ("Counter Account", lambda row, _settings: row.get("counter_account") or ""),
    "particulars": ("Particulars", lambda row, _settings: row.get("particulars") or ""),
    "receipt": ("Receipt", lambda row, settings: format_financial_hub_amount(row.get("receipt_amount"), settings=settings)),
    "payment": ("Payment", lambda row, settings: format_financial_hub_amount(row.get("payment_amount"), settings=settings)),
    "running_balance": ("Running Balance", lambda row, settings: "" if row.get("running_balance") is None else format_financial_hub_amount(row.get("running_balance"), settings=settings)),
    "source": ("Source", lambda row, _settings: row.get("source_module") or ""),
}


def _cashbook_export_section(report, *, settings):
    visible_columns = get_visible_cashbook_columns(settings)
    headers = [CASHBOOK_COLUMN_DEFS[key][0] for key in visible_columns]
    width_map = {
        "date": 72,
        "voucher_no": 98,
        "voucher_type": 92,
        "account_impacted": 112,
        "counter_account": 112,
        "particulars": 260,
        "receipt": 84,
        "payment": 84,
        "running_balance": 102,
        "source": 78,
    }
    rows = [
        [CASHBOOK_COLUMN_DEFS[key][1](row, settings) for key in visible_columns]
        for row in (report.get("results") or [])
    ]
    row_kinds = ["detail"] * len(rows)
    if report.get("results"):
        totals = report.get("totals") or {}
        total_values = {
            "date": "",
            "voucher_no": "",
            "voucher_type": "",
            "account_impacted": "",
            "counter_account": "",
            "particulars": "Report Total",
            "receipt": format_financial_hub_amount(totals.get("receipt_total"), settings=settings),
            "payment": format_financial_hub_amount(totals.get("payment_total"), settings=settings),
            "running_balance": format_financial_hub_amount(report.get("closing_balance"), settings=settings),
            "source": "",
        }
        rows.append([total_values.get(key, "") for key in visible_columns])
        row_kinds.append("final_total")
    numeric_columns = {
        index
        for index, key in enumerate(visible_columns)
        if key in {"receipt", "payment", "running_balance"}
    }
    return ExportSection(
        title="Cashbook Entries",
        headers=headers,
        rows=rows,
        row_kinds=row_kinds,
        numeric_columns=numeric_columns,
        col_widths=[width_map.get(key, 84) for key in visible_columns],
        empty_message="No cashbook rows found for the selected scope.",
    )


def _cashbook_pdf_table(report, *, settings):
    visible_columns = get_visible_cashbook_columns(settings)
    preferred_columns = [
        "date",
        "voucher_no",
        "voucher_type",
        "particulars",
        "receipt",
        "payment",
        "running_balance",
    ]
    selected_columns = [key for key in preferred_columns if key in visible_columns] or visible_columns[:7]
    headers = [CASHBOOK_COLUMN_DEFS[key][0] for key in selected_columns]
    rows = [
        [CASHBOOK_COLUMN_DEFS[key][1](row, settings) for key in selected_columns]
        for row in (report.get("results") or [])
    ]
    numeric_columns = {
        index + 1
        for index, key in enumerate(selected_columns)
        if key in {"receipt", "payment", "running_balance"}
    }
    width_map = {
        "date": 0.11,
        "voucher_no": 0.16,
        "voucher_type": 0.12,
        "particulars": 0.32,
        "receipt": 0.10,
        "payment": 0.10,
        "running_balance": 0.11,
        "account_impacted": 0.16,
        "counter_account": 0.16,
        "source": 0.09,
    }
    col_widths = [width_map.get(key, 0.12) for key in selected_columns]
    return headers, rows, numeric_columns, col_widths



class _BaseCashbookExportAPIView(_BaseBookReportAPIView):
    serializer_class = CashbookScopeSerializer
    export_mode = "attachment"
    export_orientation = "landscape"
    required_permission_codes = (
        "reports.financial_hub.cashbook.view",
        "reports.cash_book.view",
        "reports.cashbook.view",
    )

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def build_orientation(self, request):
        orientation = str(
            getattr(self, "export_orientation", None)
            or request.query_params.get("orientation")
            or "landscape"
        ).strip().lower()
        return orientation if orientation in {"landscape", "portrait"} else "landscape"

    def report_data(self, request):
        scope = self.get_scope(request)
        try:
            data = build_cashbook(
                entity_id=scope["entity"],
                entityfin_id=scope.get("entityfinid"),
                subentity_id=scope.get("subentity"),
                from_date=scope.get("from_date"),
                to_date=scope.get("to_date"),
                mode=scope.get("mode", "both"),
                cash_account_ids=scope.get("cash_account_ids"),
                bank_account_ids=scope.get("bank_account_ids"),
                counter_account_ids=scope.get("counter_account_ids"),
                voucher_types=scope.get("voucher_types"),
                search=scope.get("search"),
                page=scope.get("page", BOOK_REPORT_DEFAULTS["default_page_size_page"]),
                page_size=scope.get("page_size", BOOK_REPORT_DEFAULTS["default_page_size"]),
            )
        except ValueError as exc:
            raise ValueError(exc.args[0]) from exc
        scope_names = {
            "entity_name": data.get("entity_name"),
            "entityfin_name": data.get("entityfin_name"),
            "subentity_name": data.get("subentity_name"),
        }
        subtitle = _cashbook_subtitle(scope, scope_names, data)
        return scope, data, subtitle


class CashbookExcelAPIView(_BaseCashbookExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        settings_payload = get_financial_hub_settings_payload(user=request.user, entity_id=scope["entity"])
        effective = apply_amount_display_unit_override(
            get_effective_cashbook_settings(settings_payload),
            scope.get("amount_display_unit"),
        )
        section = _cashbook_export_section(data, settings=effective)
        summary_items = [
            ("Opening Balance", format_financial_hub_amount(data.get("opening_balance"), settings=effective)),
            ("Receipt Total", format_financial_hub_amount(data.get("totals", {}).get("receipt_total"), settings=effective)),
            ("Payment Total", format_financial_hub_amount(data.get("totals", {}).get("payment_total"), settings=effective)),
            ("Closing Balance", format_financial_hub_amount(data.get("closing_balance"), settings=effective)),
            ("Running Balance Scope", data.get("running_balance_scope") or "-"),
            ("Balance Basis", data.get("balance_basis") or "-"),
            ("Balance Integrity", "Verified" if data.get("balance_integrity") else "Review required"),
        ]
        export_subtitle = f"{subtitle} | Display Unit: {financial_hub_amount_unit_label(effective)}"
        content = write_sectioned_excel(
            title="Cashbook",
            subtitle=export_subtitle,
            summary_items=summary_items,
            sections=[section],
            orientation=self.build_orientation(request),
            freeze_header=(effective.get("export_layout") or {}).get("freeze_excel_header", True),
        )
        return self.export_response(
            filename=f"Cashbook_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class CashbookCSVAPIView(_BaseCashbookExportAPIView):
    def get(self, request):
        scope, data, _subtitle = self.report_data(request)
        settings_payload = get_financial_hub_settings_payload(user=request.user, entity_id=scope["entity"])
        effective = apply_amount_display_unit_override(
            get_effective_cashbook_settings(settings_payload),
            scope.get("amount_display_unit"),
        )
        section = _cashbook_export_section(data, settings=effective)
        meta_items = [
            ("Entity", data.get("entity_name") or "Selected entity"),
            ("Financial Year", data.get("entityfin_name") or "Current FY"),
            ("Subentity", data.get("subentity_name") or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Period", f"{_format_scope_date(scope.get('from_date')) or '-'} to {_format_scope_date(scope.get('to_date')) or '-'}"),
            ("Mode", "Cash only" if scope.get("mode") == "cash" else "Bank only" if scope.get("mode") == "bank" else "Cash and bank"),
            ("Display Unit", financial_hub_amount_unit_label(effective)),
            ("Running Balance Scope", data.get("running_balance_scope") or "-"),
            ("Balance Basis", data.get("balance_basis") or "-"),
            ("Balance Integrity", "Verified" if data.get("balance_integrity") else "Review required"),
            ("Transactions", data.get("totals", {}).get("transaction_count", 0)),
        ]
        content = write_sectioned_csv(title="Cashbook", meta_items=meta_items, sections=[section])
        return self.export_response(
            filename=f"Cashbook_{_safe_filename(_cashbook_subtitle(scope, {'entity_name': data.get('entity_name'),'entityfin_name': data.get('entityfin_name'),'subentity_name': data.get('subentity_name')}, data))}.csv",
            content=content,
            content_type="text/csv",
        )


class CashbookPDFAPIView(_BaseCashbookExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        settings_payload = get_financial_hub_settings_payload(user=request.user, entity_id=scope["entity"])
        effective = apply_amount_display_unit_override(
            get_effective_cashbook_settings(settings_payload),
            scope.get("amount_display_unit"),
        )
        meta_items = [
            ("Entity", data.get("entity_name") or "Selected entity"),
            ("Financial Year", data.get("entityfin_name") or "Current FY"),
            ("Subentity", data.get("subentity_name") or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Period", f"{_format_scope_date(scope.get('from_date')) or '-'} to {_format_scope_date(scope.get('to_date')) or '-'}"),
            ("Mode", "Cash only" if scope.get("mode") == "cash" else "Bank only" if scope.get("mode") == "bank" else "Cash and bank"),
            ("Display Unit", financial_hub_amount_unit_label(effective)),
            ("Running Balance Scope", data.get("running_balance_scope") or "-"),
            ("Balance Basis", data.get("balance_basis") or "-"),
            ("Balance Integrity", "Verified" if data.get("balance_integrity") else "Review required"),
            ("Transactions", data.get("totals", {}).get("transaction_count", 0)),
            ("Opening Balance", format_financial_hub_amount(data.get("opening_balance"), settings=effective)),
            ("Closing Balance", format_financial_hub_amount(data.get("closing_balance"), settings=effective)),
        ]
        export_subtitle = f"{subtitle} | Display Unit: {financial_hub_amount_unit_label(effective)}"
        content = write_sectioned_pdf(
            title="Cashbook",
            subtitle=export_subtitle,
            meta_items=meta_items,
            sections=[_cashbook_export_section(data, settings=effective)],
            pagesize=_report_pagesize(self.build_orientation(request)),
            header_density=(effective.get("export_layout") or {}).get("header_density", "compact"),
            metadata_visibility=(effective.get("export_layout") or {}).get("metadata_visibility", "compact"),
        )
        return self.export_response(
            filename=f"Cashbook_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class CashbookPrintAPIView(CashbookPDFAPIView):
    export_mode = "inline"


class CashbookExcelLandscapeAPIView(CashbookExcelAPIView):
    export_orientation = "landscape"


class CashbookExcelPortraitAPIView(CashbookExcelAPIView):
    export_orientation = "portrait"


class CashbookPDFLandscapeAPIView(CashbookPDFAPIView):
    export_orientation = "landscape"


class CashbookPDFPortraitAPIView(CashbookPDFAPIView):
    export_orientation = "portrait"
