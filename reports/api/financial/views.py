from __future__ import annotations

import csv
from decimal import Decimal
from io import BytesIO, StringIO
from xml.sax.saxutils import escape

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from core.entitlements import ScopedEntitlementMixin
from reports.schemas.common import build_report_envelope
from reports.schemas.financial_reports import FinancialReportScopeSerializer, LedgerBookScopeSerializer
from reports.services.financial.meta import REPORT_DEFAULTS, build_financial_report_meta
from reports.services.financial.ledger_book import build_ledger_book
from reports.services.financial.reporting_policy import resolve_financial_reporting_policy
from reports.services.financial.statements import build_balance_sheet, build_profit_and_loss
from reports.services.financial.trial_balance import build_trial_balance
from reports.services.trading_account import build_trading_account_dynamic
from reports.selectors.financial import resolve_date_window, resolve_scope_names
from subscriptions.services import SubscriptionLimitCodes, SubscriptionService


class _BaseFinancialReportAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = FinancialReportScopeSerializer
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get_scope(self, request):
        serializer = self.serializer_class(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        self.enforce_scope(
            request,
            entity_id=scope["entity"],
            entityfinid_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
        )
        return scope

    def build_filters(self, scope):
        return {
            "entity": scope["entity"],
            "entityfinid": scope.get("entityfinid"),
            "subentity": scope.get("subentity"),
            "scope_mode": scope.get("scope_mode"),
            "as_on_date": scope.get("as_on_date"),
            "from_date": scope.get("from_date"),
            "to_date": scope.get("to_date"),
            "as_of_date": scope.get("as_of_date"),
            "view_type": scope.get("view_type"),
            "account_group": scope.get("account_group"),
            "ledger_ids": scope.get("ledger_ids"),
            "include_zero_balance": scope.get("include_zero_balance", REPORT_DEFAULTS["show_zero_balances_default"]),
            "hide_zero_rows": scope.get("hide_zero_rows", not REPORT_DEFAULTS["show_zero_balances_default"]),
            "group_by": scope.get("group_by"),
            "period_by": scope.get("period_by"),
            "include_opening": scope.get("include_opening"),
            "include_movement": scope.get("include_movement"),
            "include_closing": scope.get("include_closing"),
            "posted_only": scope.get("posted_only", True),
            "stock_valuation_mode": scope.get(
                "stock_valuation_mode",
                REPORT_DEFAULTS["balance_sheet_stock_valuation_mode"],
            ),
            "stock_valuation_method": scope.get(
                "stock_valuation_method",
                REPORT_DEFAULTS["balance_sheet_stock_valuation_method"],
            ),
            "include_zero_balances": scope.get(
                "include_zero_balances",
                REPORT_DEFAULTS["show_zero_balances_default"],
            ),
            "include_inactive_ledgers": scope.get("include_inactive_ledgers", False),
            "search": scope.get("search"),
            "sort_by": scope.get("sort_by"),
            "sort_order": scope.get("sort_order", "asc"),
            "page": scope.get("page", 1),
            "page_size": scope.get("page_size", REPORT_DEFAULTS["default_page_size"]),
            "export": scope.get("export"),
        }


def _safe_filename(value):
    text = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in str(value or "").strip())
    text = text.strip("._-")
    return text or "report"


def _format_scope_date(value):
    if not value:
        return "-"
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return str(value)


def _filtered_querydict(request, *, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    return params.urlencode()


def _attach_financial_actions(payload, request, *, export_base_path):
    query = _filtered_querydict(request, exclude=["page", "page_size"])
    payload["actions"]["can_print"] = True
    payload["actions"]["export_urls"] = {
        "excel": f"{export_base_path}excel/?{query}",
        "pdf": f"{export_base_path}pdf/?{query}",
        "csv": f"{export_base_path}csv/?{query}",
        "print": f"{export_base_path}print/?{query}",
    }
    payload["available_exports"] = ["excel", "pdf", "csv", "print"]
    return payload


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, right, border


def _write_csv(headers, rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def _write_excel(title, subtitle, headers, rows, *, numeric_columns=None, orientation="landscape"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    orientation = (orientation or "landscape").strip().lower()
    if orientation not in {"landscape", "portrait"}:
        orientation = "landscape"
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    header_font, header_fill, center, left, right, border = _workbook_styles()
    numeric_columns = set(numeric_columns or [])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = left
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).alignment = left

    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(rows, start=header_row + 1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = right if col_index in numeric_columns else left

    for col_index, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 14)
        for row in rows[:100]:
            width = max(width, len(str(row[col_index - 1])) + 2 if col_index - 1 < len(row) else width)
        ws.column_dimensions[get_column_letter(col_index)].width = min(width, 40)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_pdf(title, subtitle, headers, rows, *, col_widths=None, meta_items=None, numeric_columns=None, pagesize=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize or landscape(A4), rightMargin=18, leftMargin=18, topMargin=24, bottomMargin=18)
    styles = getSampleStyleSheet()
    title_style = styles["Title"].clone("FinancialPdfTitle")
    title_style.textColor = colors.HexColor("#FFFFFF")
    title_style.alignment = 1
    title_style.leading = 18
    title_style.fontSize = 16
    title_style.spaceAfter = 0

    subtitle_style = styles["BodyText"].clone("FinancialPdfSubtitle")
    subtitle_style.fontSize = 9
    subtitle_style.leading = 11
    subtitle_style.textColor = colors.HexColor("#4A5568")
    subtitle_style.spaceAfter = 0

    meta_label_style = styles["BodyText"].clone("FinancialPdfMetaLabel")
    meta_label_style.fontSize = 7.5
    meta_label_style.leading = 9
    meta_label_style.textColor = colors.HexColor("#5B6573")
    meta_label_style.spaceAfter = 0

    meta_value_style = styles["BodyText"].clone("FinancialPdfMetaValue")
    meta_value_style.fontSize = 9
    meta_value_style.leading = 11
    meta_value_style.textColor = colors.HexColor("#1F2937")
    meta_value_style.spaceAfter = 0

    def _page_decorator(canvas, _doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D7E2F1"))
        canvas.setLineWidth(0.8)
        canvas.line(doc.leftMargin, doc.bottomMargin - 8, doc.pagesize[0] - doc.rightMargin, doc.bottomMargin - 8)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, doc.bottomMargin - 18, f"Page {_doc.page}")
        canvas.restoreState()

    header_rows = [
        [
            Paragraph(escape(title), title_style)
        ]
    ]
    header_table = Table(header_rows, colWidths=[doc.width])
    header_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2F5597")),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#2F5597")),
            ]
        )
    )

    story = [header_table, Spacer(1, 8), Paragraph(escape(subtitle), subtitle_style), Spacer(1, 10)]

    if meta_items:
        meta_cards = []
        for label, value in meta_items:
            meta_cards.append(
                Paragraph(
                    f"<font color='#5B6573'><b>{escape(str(label))}:</b></font><br/><font color='#1F2937'>{escape(str(value or '-'))}</font>",
                    meta_value_style,
                )
            )
        while len(meta_cards) % 3 != 0:
            meta_cards.append(Paragraph("&nbsp;", meta_label_style))

        meta_rows = [meta_cards[index:index + 3] for index in range(0, len(meta_cards), 3)]
        meta_table = Table(meta_rows, colWidths=[doc.width / 3.0] * 3)
        meta_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F7FB")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D7E2F1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7E2F1")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.extend([meta_table, Spacer(1, 10)])

    table_data = [headers, *rows]
    if col_widths:
        table_data = [
            [
                _truncate_text(value, col_widths[index] if index < len(col_widths) else 72)
                for index, value in enumerate(row)
            ]
            for row in table_data
        ]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    numeric_columns = set(numeric_columns or [])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8.5),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CFCFCF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FC")]),
            ]
        )
    )
    for col_index in numeric_columns:
        table.setStyle(TableStyle([("ALIGN", (col_index, 1), (col_index, -1), "RIGHT")]))
    story.append(table)
    doc.build(story, onFirstPage=_page_decorator, onLaterPages=_page_decorator)
    return buffer.getvalue()


def _truncate_text(value, width_points, *, min_chars=8):
    text = "" if value is None else str(value)
    max_chars = max(min_chars, int((width_points or 72) / 5))
    if len(text) <= max_chars:
        return text
    if max_chars <= 3:
        return text[:max_chars]
    return f"{text[: max_chars - 3].rstrip()}..."


def _format_balance_amount(value, *, decimals=2):
    amount = Decimal(str(value or 0))
    if amount == 0:
        return f"{amount:.{decimals}f}"
    display = f"{abs(amount):.{decimals}f}"
    return f"{display} {'Dr' if amount >= 0 else 'Cr'}"


def _trial_balance_subtitle(scope_names, scope):
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Scope: {scope.get('scope_mode') or 'financial_year'} | "
        f"Group by: {scope.get('account_group') or scope.get('group_by') or 'ledger'} | "
        f"View: {scope.get('view_type') or 'summary'} | "
        f"Posted only: {scope.get('posted_only', True)}"
    )


def _ledger_book_subtitle(scope_names, scope, ledger):
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Ledger: {ledger.get('name') or 'Selected ledger'} ({ledger.get('ledger_code') or '-'}) | "
        f"Account head: {ledger.get('accounthead_name') or '-'} | "
        f"Account type: {ledger.get('accounttype_name') or '-'} | "
        f"Scope: {scope.get('scope_mode') or 'financial_year'} | "
        f"Sort: {scope.get('sort_by') or 'posting_date'} {scope.get('sort_order') or 'asc'}"
    )


def _trial_balance_flatten_rows(rows, *, depth=0):
    flattened = []
    for row in rows:
        flattened.append((row, depth))
        children = row.get("children") or []
        if children:
            flattened.extend(_trial_balance_flatten_rows(children, depth=depth + 1))
    return flattened


def _trial_balance_export_table(report):
    rows = report.get("rows") or []
    periods = report.get("periods") or []
    flattened = _trial_balance_flatten_rows(rows)

    headers = [
        "Level",
        "Type",
        "Code",
        "Name",
        "Account Head",
        "Account Type",
        "Opening",
        "Debit",
        "Credit",
        "Closing",
        "Abnormal",
    ]
    for period in periods:
        label = period.get("period_label") or period.get("label") or period.get("name") or period.get("title") or period.get("key")
        headers.extend([
            f"{label} Opening",
            f"{label} Debit",
            f"{label} Credit",
            f"{label} Closing",
        ])

    table_rows = []
    for row, depth in flattened:
        label = row.get("label") or row.get("ledger_name") or row.get("accounthead_name") or row.get("accounttype_name") or "-"
        prefix = ("  " * depth) + label
        row_periods = {str(item.get("key") or item.get("code") or item.get("label") or item.get("name") or ""): item for item in row.get("periods") or []}
        values = [
            depth,
            "Group" if row.get("children") else "Row",
            row.get("ledger_code") or row.get("code") or "",
            prefix,
            row.get("accounthead_name") or "",
            row.get("accounttype_name") or "",
            _format_balance_amount(row.get("opening", "0.00")),
            row.get("debit", "0.00"),
            row.get("credit", "0.00"),
            _format_balance_amount(row.get("closing", "0.00")),
            "Yes" if row.get("is_abnormal_balance") else "No",
        ]
        for period in periods:
            key = str(period.get("period_key") or period.get("key") or period.get("code") or period.get("label") or period.get("name") or "")
            period_row = row_periods.get(key) or {}
            values.extend([
                _format_balance_amount(period_row.get("opening", "0.00")),
                period_row.get("debit", "0.00"),
                period_row.get("credit", "0.00"),
                _format_balance_amount(period_row.get("closing", "0.00")),
            ])
        table_rows.append(values)

    totals = report.get("totals") or {}
    total_row = [
        "",
        "",
        "",
        "Totals",
        "",
        "",
        totals.get("opening", "0.00"),
        totals.get("debit", "0.00"),
        totals.get("credit", "0.00"),
        totals.get("closing", "0.00"),
        "",
    ]
    for _period in periods:
        total_row.extend(["", "", "", ""])
    table_rows.append(total_row)

    return headers, table_rows


def _profit_loss_subtitle(scope_names, scope, report):
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    view_type = scope.get("view_type") or "summary"
    group_by = scope.get("account_group") or scope.get("group_by") or "ledger"
    stock_mode = report.get("stock_valuation", {}).get("effective_mode") or scope.get("stock_valuation_mode") or "trading_account"
    stock_method = report.get("stock_valuation", {}).get("valuation_method") or scope.get("stock_valuation_method") or "fifo"
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Scope: {scope.get('scope_mode') or 'financial_year'} | "
        f"View: {view_type} | "
        f"Group by: {group_by} | "
        f"Posted only: {scope.get('posted_only', True)} | "
        f"Stock: {stock_mode}/{stock_method}"
    )


def _profit_loss_period_lookup(row):
    lookup = {}
    for item in row.get("periods") or []:
        key = str(item.get("key") or item.get("code") or item.get("label") or item.get("name") or item.get("title") or "")
        if key:
            lookup[key] = item
    return lookup


def _profit_loss_export_table(report, *, include_periods=True):
    headers = ["Section", "Particulars", "Account Head", "Account Type"]
    periods = report.get("periods") or []
    empty_period_cells = [""] * len(periods) if include_periods else []
    if include_periods:
        for period in periods:
            label = period.get("period_label") or period.get("label") or period.get("name") or period.get("title") or period.get("key")
            headers.extend([f"{label} Amount"])
    headers.append("Amount")

    def build_rows(section_name, rows, subtotal):
        section_rows = []
        for row in rows:
            label = row.get("label") or row.get("name") or row.get("title") or row.get("ledger_name") or row.get("accounthead_name") or row.get("accounttype_name") or "-"
            value_row = [
                section_name,
                label,
                row.get("accounthead_name") or "",
                row.get("accounttype_name") or "",
            ]
            period_lookup = _profit_loss_period_lookup(row)
            if include_periods:
                for period in periods:
                    key = str(period.get("period_key") or period.get("key") or period.get("code") or period.get("label") or period.get("name") or period.get("title") or "")
                    period_item = period_lookup.get(key) or {}
                    amount = period_item.get("amount", period_item.get("value", "0.00"))
                    value_row.append(amount)
            value_row.append(row.get("amount", "0.00"))
            section_rows.append(value_row)

        section_rows.append([
            section_name,
            f"{section_name} subtotal",
            "",
            "",
            *empty_period_cells,
            subtotal,
        ])
        return section_rows

    rows = []
    rows.extend(build_rows("Income", report.get("income") or [], report.get("totals", {}).get("income", "0.00")))
    rows.append(["", "", "", "", *empty_period_cells, ""])
    rows.extend(build_rows("Expense", report.get("expenses") or [], report.get("totals", {}).get("expense", "0.00")))
    rows.append(["", "", "", "", *empty_period_cells, ""])
    rows.append([
        "Summary",
        "Net Profit",
        "",
        "",
        *empty_period_cells,
        report.get("totals", {}).get("net_profit", "0.00"),
    ])

    return headers, rows


def _trading_account_subtitle(scope_names, scope, report):
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Scope: {scope.get('scope_mode') or 'financial_year'} | "
        f"View: {scope.get('view_type') or 'summary'} | "
        f"Group by: {scope.get('account_group') or scope.get('group_by') or 'accounthead'} | "
        f"Valuation: {scope.get('valuation_method') or report.get('params', {}).get('valuation_method') or 'fifo'}"
    )


def _trading_account_flatten_rows(rows, side, out, depth=0):
    for row in rows or []:
        out.append([
            side,
            "  " * depth + str(row.get("label") or "-"),
            row.get("qty", ""),
            row.get("amount", "0.00"),
        ])
        if row.get("children"):
            _trading_account_flatten_rows(row["children"], side, out, depth + 1)


def _trading_account_export_table(report):
    headers = ["Side", "Particulars", "Qty", "Amount"]
    rows = []
    _trading_account_flatten_rows(report.get("debit_rows") or [], "Debit", rows)
    rows.append(["", "", "", ""])
    rows.append(["Debit Total", "", "", report.get("debit_total", "0.00")])
    rows.append(["", "", "", ""])
    _trading_account_flatten_rows(report.get("credit_rows") or [], "Credit", rows)
    rows.append(["", "", "", ""])
    rows.append(["Credit Total", "", "", report.get("credit_total", "0.00")])
    rows.append(["", "", "", ""])
    rows.append(["Gross Profit", "", "", report.get("gross_profit", "0.00") or "0.00"])
    rows.append(["Gross Loss", "", "", report.get("gross_loss", "0.00") or "0.00"])
    return headers, rows


def _balance_sheet_row_label(row, depth=0):
    label = (
        row.get("label")
        or row.get("name")
        or row.get("title")
        or row.get("ledger_name")
        or row.get("accounthead_name")
        or row.get("accounttype_name")
        or "-"
    )
    return f"{'  ' * depth}{label}"


def _balance_sheet_row_period_value(row, period_key):
    period_values = row.get("period_values") or {}
    if period_key in period_values:
        return period_values.get(period_key)

    amounts_by_period = row.get("amounts_by_period") or {}
    if period_key in amounts_by_period:
        return amounts_by_period.get(period_key)

    for item in row.get("periods") or []:
        item_key = str(item.get("key") or item.get("code") or item.get("label") or item.get("name") or "")
        if item_key == period_key:
            return item.get("amount") or item.get("value") or item.get("closing")

    return None


def _balance_sheet_flatten_rows(rows, *, section_label, period_keys, depth=0):
    flattened = []
    for row in rows or []:
        flattened.append(
            {
                "section": section_label if depth == 0 else "",
                "particulars": _balance_sheet_row_label(row, depth=depth),
                "account_head": row.get("accounthead_name") or "-",
                "account_type": row.get("accounttype_name") or "-",
                "period_values": [_balance_sheet_row_period_value(row, period_key) for period_key in period_keys],
                "amount": row.get("amount"),
            }
        )
        children = row.get("children") or []
        if children:
            flattened.extend(
                _balance_sheet_flatten_rows(
                    children,
                    section_label=section_label,
                    period_keys=period_keys,
                    depth=depth + 1,
                )
            )
    return flattened


def _balance_sheet_export_table(report):
    periods = report.get("periods") or []
    period_keys = [str(period.get("period_key") or period.get("key") or period.get("code") or period.get("label") or f"period_{index}") for index, period in enumerate(periods, start=1)]
    period_labels = [str(period.get("period_label") or period.get("label") or period.get("name") or period.get("code") or period.get("period_key") or f"Period {index}") for index, period in enumerate(periods, start=1)]

    headers = ["Section", "Particulars", "Account Head", "Account Type", *period_labels, "Amount"]
    rows = []

    for section_label, section_rows in (
        ("Assets", report.get("assets") or []),
        ("Liabilities & Equity", report.get("liabilities_and_equity") or []),
    ):
        for item in _balance_sheet_flatten_rows(section_rows, section_label=section_label, period_keys=period_keys):
            rows.append(
                [
                    item["section"],
                    item["particulars"],
                    item["account_head"],
                    item["account_type"],
                    *item["period_values"],
                    item["amount"],
                ]
            )
        rows.append(
            [
                section_label,
                f"{section_label} Total",
                "",
                "",
                *["" for _ in period_keys],
                report.get("totals", {}).get("assets" if section_label == "Assets" else "liabilities_and_equity") or "0.00",
            ]
        )

    assets_total = Decimal(str(report.get("totals", {}).get("assets") or "0"))
    liabilities_total = Decimal(str(report.get("totals", {}).get("liabilities_and_equity") or "0"))
    balance_diff = assets_total - liabilities_total
    rows.append(["", "Balance Difference", "", "", *["" for _ in period_keys], f"{balance_diff:.2f}"])

    return headers, rows


class FinancialReportsMetaAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get(self, request):
        entity_id = request.query_params.get("entity")
        if not entity_id:
            return Response({"detail": "entity is required."}, status=400)
        entity_id = int(entity_id)
        self.enforce_scope(request, entity_id=entity_id)
        payload = build_financial_report_meta(entity_id)
        payload["reporting_policy"] = resolve_financial_reporting_policy(entity_id)
        return Response(payload)


class TrialBalanceAPIView(_BaseFinancialReportAPIView):
    def get(self, request):
        scope = self.get_scope(request)
        data = build_trial_balance(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            ledger_ids=scope.get("ledger_ids"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            account_group=scope.get("account_group") or scope.get("group_by"),
            include_zero_balances=scope.get(
                "include_zero_balances",
                REPORT_DEFAULTS["show_zero_balances_default"],
            ),
            posted_only=scope.get("posted_only", True),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", REPORT_DEFAULTS["default_page_size"]),
            period_by=scope.get("period_by"),
            view_type=scope.get("view_type"),
            include_opening=scope.get("include_opening", REPORT_DEFAULTS["show_opening_balance_default"]),
            include_movement=scope.get("include_movement", True),
            include_closing=scope.get("include_closing", True),
        )
        response = build_report_envelope(
            report_code="trial_balance",
            report_name="Trial Balance",
            payload=data,
            filters=self.build_filters(scope),
            defaults=REPORT_DEFAULTS,
        )
        return Response(_attach_financial_actions(response, request, export_base_path="/api/reports/financial/trial-balance/"))


class _BaseTrialBalanceExportAPIView(_BaseFinancialReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        data = build_trial_balance(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            ledger_ids=scope.get("ledger_ids"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            account_group=scope.get("account_group") or scope.get("group_by"),
            include_zero_balances=scope.get("include_zero_balances", REPORT_DEFAULTS["show_zero_balances_default"]),
            posted_only=scope.get("posted_only", True),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=1,
            page_size=100000,
            period_by=scope.get("period_by"),
            view_type=scope.get("view_type"),
            include_opening=scope.get("include_opening", REPORT_DEFAULTS["show_opening_balance_default"]),
            include_movement=scope.get("include_movement", True),
            include_closing=scope.get("include_closing", True),
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        headers, rows = _trial_balance_export_table(data)
        subtitle = _trial_balance_subtitle(scope_names, scope)
        return scope, data, headers, rows, subtitle


class TrialBalanceExcelAPIView(_BaseTrialBalanceExportAPIView):
    def get(self, request):
        _scope, _data, headers, rows, subtitle = self.report_data(request)
        content = _write_excel("Trial Balance", subtitle, headers, rows, numeric_columns={0, 6, 7, 8, 9})
        return self.export_response(
            filename=f"TrialBalance_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class TrialBalanceCSVAPIView(_BaseTrialBalanceExportAPIView):
    def get(self, request):
        _scope, _data, headers, rows, subtitle = self.report_data(request)
        content = _write_csv(headers, rows)
        return self.export_response(
            filename=f"TrialBalance_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class TrialBalancePDFAPIView(_BaseTrialBalanceExportAPIView):
    def get(self, request):
        scope, data, headers, rows, subtitle = self.report_data(request)
        col_widths = [28, 40, 55, 130, 95, 78, 78, 78, 78, 78, 40]
        if len(headers) > len(col_widths):
            col_widths.extend([58] * (len(headers) - len(col_widths)))
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        meta_items = [
            ("Entity", scope_names["entity_name"] or "Selected entity"),
            ("Financial Year", scope_names["entityfin_name"] or "Current FY"),
            ("Subentity", scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Scope", scope.get("scope_mode") or "financial_year"),
            ("Group by", scope.get("account_group") or scope.get("group_by") or "ledger"),
            ("View", scope.get("view_type") or "summary"),
        ]
        content = _write_pdf(
            "Trial Balance",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            meta_items=meta_items,
            numeric_columns={6, 7, 8, 9},
        )
        return self.export_response(
            filename=f"TrialBalance_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class TrialBalancePrintAPIView(TrialBalancePDFAPIView):
    export_mode = "inline"


class LedgerBookAPIView(_BaseFinancialReportAPIView):
    serializer_class = LedgerBookScopeSerializer

    def get(self, request):
        scope = self.get_scope(request)
        data = build_ledger_book(
            entity_id=scope["entity"],
            ledger_id=scope["ledger"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            search=scope.get("search"),
            voucher_types=[scope.get("voucher_type")] if scope.get("voucher_type") else None,
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", REPORT_DEFAULTS["default_page_size"]),
        )
        return Response(
            _attach_financial_actions(
                build_report_envelope(
                report_code="ledger_book",
                report_name="Ledger Book",
                payload=data,
                filters={**self.build_filters(scope), "ledger": scope["ledger"], "voucher_type": scope.get("voucher_type")},
                defaults=REPORT_DEFAULTS,
                ),
                request,
                export_base_path="/api/reports/financial/ledger-book/",
            )
        )


class _BaseLedgerBookExportAPIView(_BaseFinancialReportAPIView):
    serializer_class = LedgerBookScopeSerializer
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        data = build_ledger_book(
            entity_id=scope["entity"],
            ledger_id=scope["ledger"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            search=scope.get("search"),
            voucher_types=[scope.get("voucher_type")] if scope.get("voucher_type") else None,
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=1,
            page_size=100000,
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        ledger = data.get("ledger") or {}
        subtitle = _ledger_book_subtitle(scope_names, scope, ledger)
        return scope, data, subtitle

    def build_rows(self, data):
        rows = [
            [
                _format_scope_date(row.get("posting_date")),
                row.get("voucher_number") or "-",
                row.get("voucher_type_name") or row.get("voucher_type") or "-",
                row.get("description") or "-",
                row.get("debit", "0.00"),
                row.get("credit", "0.00"),
                _format_balance_amount(row.get("running_balance", "0.00")),
            ]
            for row in data.get("rows", [])
        ]
        totals = data.get("totals") or {}
        rows.append([
            "",
            "",
            "",
            "Totals",
            totals.get("debit", "0.00"),
            totals.get("credit", "0.00"),
            _format_balance_amount(totals.get("closing_balance", "0.00")),
        ])
        return rows


class LedgerBookExcelAPIView(_BaseLedgerBookExportAPIView):
    def get(self, request):
        _scope, data, subtitle = self.report_data(request)
        headers = ["Date", "Voucher No", "Voucher Type", "Description", "Debit", "Credit", "Running Balance"]
        rows = self.build_rows(data)
        content = _write_excel("Ledger Book", subtitle, headers, rows, numeric_columns={4, 5, 6})
        return self.export_response(
            filename=f"LedgerBook_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class LedgerBookCSVAPIView(_BaseLedgerBookExportAPIView):
    def get(self, request):
        _scope, data, subtitle = self.report_data(request)
        headers = ["Date", "Voucher No", "Voucher Type", "Description", "Debit", "Credit", "Running Balance"]
        rows = self.build_rows(data)
        content = _write_csv(headers, rows)
        return self.export_response(
            filename=f"LedgerBook_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class LedgerBookPDFAPIView(_BaseLedgerBookExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        headers = ["Date", "Voucher No", "Voucher Type", "Description", "Debit", "Credit", "Running Balance"]
        rows = self.build_rows(data)
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        ledger = data.get("ledger") or {}
        meta_items = [
            ("Entity", scope_names["entity_name"] or "Selected entity"),
            ("Financial Year", scope_names["entityfin_name"] or "Current FY"),
            ("Subentity", scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Ledger", ledger.get("name") or "Selected ledger"),
            ("Account head", ledger.get("accounthead_name") or "-"),
            ("Account type", ledger.get("accounttype_name") or "-"),
        ]
        content = _write_pdf(
            "Ledger Book",
            subtitle,
            headers,
            rows,
            col_widths=[58, 76, 100, 220, 74, 74, 88],
            meta_items=meta_items,
            numeric_columns={4, 5, 6},
        )
        return self.export_response(
            filename=f"LedgerBook_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class LedgerBookPrintAPIView(LedgerBookPDFAPIView):
    export_mode = "inline"


class ProfitAndLossAPIView(_BaseFinancialReportAPIView):
    def get(self, request):
        scope = self.get_scope(request)
        reporting_policy = resolve_financial_reporting_policy(scope["entity"])
        view_type = (scope.get("view_type") or "summary").lower()
        account_group = scope.get("account_group") or scope.get("group_by")
        if not account_group:
            account_group = "ledger" if view_type == "detailed" else "accounthead"
        stock_valuation_mode = scope.get(
            "stock_valuation_mode",
            REPORT_DEFAULTS["profit_loss_stock_valuation_mode"],
        )
        stock_valuation_method = scope.get(
            "stock_valuation_method",
            REPORT_DEFAULTS["profit_loss_stock_valuation_method"],
        )
        data = build_profit_and_loss(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            group_by=account_group,
            include_zero_balances=scope.get(
                "include_zero_balances",
                REPORT_DEFAULTS["show_zero_balances_default"],
            ),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", REPORT_DEFAULTS["default_page_size"]),
            period_by=scope.get("period_by"),
            view_type=view_type,
            posted_only=scope.get("posted_only", True),
            hide_zero_rows=not scope.get("include_zero_balance", False),
            account_group=account_group,
            ledger_ids=scope.get("ledger_ids") or None,
            stock_valuation_mode=stock_valuation_mode,
            stock_valuation_method=stock_valuation_method,
            reporting_policy=reporting_policy,
        )
        response = build_report_envelope(
            report_code="profit_loss",
            report_name="Profit and Loss",
            payload=data,
            filters=self.build_filters(scope),
            defaults=REPORT_DEFAULTS,
        )
        response = _attach_financial_actions(
            response,
            request,
            export_base_path="/api/reports/financial/profit-loss/",
        )
        query = _filtered_querydict(request, exclude=["page", "page_size", "orientation"])
        base_url = "/api/reports/financial/profit-loss/"
        query_suffix = f"?{query}" if query else ""
        response["actions"]["export_urls"]["pdf_landscape"] = f"{base_url}pdf/landscape/{query_suffix}"
        response["actions"]["export_urls"]["pdf_portrait"] = f"{base_url}pdf/portrait/{query_suffix}"
        response["actions"]["export_urls"]["excel_landscape"] = f"{base_url}excel/landscape/{query_suffix}"
        response["actions"]["export_urls"]["excel_portrait"] = f"{base_url}excel/portrait/{query_suffix}"
        response["available_exports"] = ["excel", "pdf", "csv", "print", "pdf_landscape", "pdf_portrait", "excel_landscape", "excel_portrait"]
        return Response(response)


class _BaseProfitAndLossExportAPIView(_BaseFinancialReportAPIView):
    export_mode = "attachment"
    export_orientation = "landscape"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        reporting_policy = resolve_financial_reporting_policy(scope["entity"])
        view_type = (scope.get("view_type") or "summary").lower()
        account_group = scope.get("account_group") or scope.get("group_by")
        if not account_group:
            account_group = "ledger" if view_type == "detailed" else "accounthead"
        stock_valuation_mode = scope.get(
            "stock_valuation_mode",
            REPORT_DEFAULTS["profit_loss_stock_valuation_mode"],
        )
        stock_valuation_method = scope.get(
            "stock_valuation_method",
            REPORT_DEFAULTS["profit_loss_stock_valuation_method"],
        )
        data = build_profit_and_loss(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            group_by=account_group,
            include_zero_balances=scope.get("include_zero_balances", REPORT_DEFAULTS["show_zero_balances_default"]),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=1,
            page_size=100000,
            period_by=scope.get("period_by"),
            view_type=view_type,
            posted_only=scope.get("posted_only", True),
            hide_zero_rows=not scope.get("include_zero_balance", False),
            account_group=account_group,
            ledger_ids=scope.get("ledger_ids") or None,
            stock_valuation_mode=stock_valuation_mode,
            stock_valuation_method=stock_valuation_method,
            reporting_policy=reporting_policy,
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        subtitle = _profit_loss_subtitle(scope_names, scope, data)
        return scope, data, subtitle

    def build_orientation(self, request):
        orientation = str(
            getattr(self, "export_orientation", None)
            or request.query_params.get("orientation")
            or "landscape"
        ).strip().lower()
        return orientation if orientation in {"landscape", "portrait"} else "landscape"


class ProfitAndLossExcelAPIView(_BaseProfitAndLossExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        headers, rows = _profit_loss_export_table(data, include_periods=True)
        content = _write_excel(
            "Profit & Loss",
            subtitle,
            headers,
            rows,
            numeric_columns=set(range(4, len(headers))),
            orientation=self.build_orientation(request),
        )
        return self.export_response(
            filename=f"ProfitLoss_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class ProfitAndLossCSVAPIView(_BaseProfitAndLossExportAPIView):
    def get(self, request):
        _scope, data, subtitle = self.report_data(request)
        headers, rows = _profit_loss_export_table(data, include_periods=True)
        content = _write_csv(headers, rows)
        return self.export_response(
            filename=f"ProfitLoss_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class ProfitAndLossPDFAPIView(_BaseProfitAndLossExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        orientation = self.build_orientation(request)
        portrait_mode = orientation == "portrait"
        headers, rows = _profit_loss_export_table(data, include_periods=not portrait_mode)
        periods = data.get("periods") or []
        if portrait_mode:
            col_widths = [54, 180, 92, 86, 78]
            pagesize = A4
        else:
            col_widths = [58, 130, 88, 78]
            col_widths.extend([72] * len(periods))
            col_widths.append(78)
            pagesize = landscape(A4)
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        stock = data.get("stock_valuation") or {}
        summary = data.get("summary") or {}
        meta_items = [
            ("Entity", scope_names["entity_name"] or "Selected entity"),
            ("Financial Year", scope_names["entityfin_name"] or "Current FY"),
            ("Subentity", scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("View", scope.get("view_type") or "summary"),
            ("Group by", scope.get("account_group") or scope.get("group_by") or "ledger"),
            ("Stock", f"{stock.get('effective_mode') or 'trading_account'} / {stock.get('valuation_method') or 'fifo'}"),
            ("Gross result", summary.get("gross_result") or "0.00"),
            ("Net margin", f"{summary.get('net_margin_percent') or '0.00'}%"),
        ]
        content = _write_pdf(
            "Profit & Loss",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            meta_items=meta_items,
            numeric_columns=set(range(4, len(headers))),
            pagesize=pagesize,
        )
        return self.export_response(
            filename=f"ProfitLoss_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class ProfitAndLossPrintAPIView(ProfitAndLossPDFAPIView):
    export_mode = "inline"


class ProfitAndLossExcelLandscapeAPIView(ProfitAndLossExcelAPIView):
    export_orientation = "landscape"


class ProfitAndLossExcelPortraitAPIView(ProfitAndLossExcelAPIView):
    export_orientation = "portrait"


class ProfitAndLossPDFLandscapeAPIView(ProfitAndLossPDFAPIView):
    export_orientation = "landscape"


class ProfitAndLossPDFPortraitAPIView(ProfitAndLossPDFAPIView):
    export_orientation = "portrait"


class BalanceSheetAPIView(_BaseFinancialReportAPIView):
    def get(self, request):
        scope = self.get_scope(request)
        reporting_policy = resolve_financial_reporting_policy(scope["entity"])
        data = build_balance_sheet(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            group_by=scope.get("account_group") or scope.get("group_by"),
            view_type=scope.get("view_type"),
            posted_only=scope.get("posted_only", REPORT_DEFAULTS["balance_sheet_posted_only_default"]),
            hide_zero_rows=scope.get("hide_zero_rows", REPORT_DEFAULTS["balance_sheet_hide_zero_rows_default"]),
            include_zero_balances=scope.get(
                "include_zero_balances",
                REPORT_DEFAULTS["show_zero_balances_default"],
            ),
            account_group=scope.get("account_group"),
            ledger_ids=scope.get("ledger_ids"),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", REPORT_DEFAULTS["default_page_size"]),
            period_by=scope.get("period_by"),
            reporting_policy=reporting_policy,
        )
        response = build_report_envelope(
            report_code="balance_sheet",
            report_name="Balance Sheet",
            payload=data,
            filters=self.build_filters(scope),
            defaults=REPORT_DEFAULTS,
        )
        response = _attach_financial_actions(
            response,
            request,
            export_base_path="/api/reports/financial/balance-sheet/",
        )
        query = _filtered_querydict(request, exclude=["page", "page_size", "orientation"])
        base_url = "/api/reports/financial/balance-sheet/"
        query_suffix = f"?{query}" if query else ""
        response["actions"]["export_urls"]["pdf_landscape"] = f"{base_url}pdf/landscape/{query_suffix}"
        response["actions"]["export_urls"]["pdf_portrait"] = f"{base_url}pdf/portrait/{query_suffix}"
        response["actions"]["export_urls"]["excel_landscape"] = f"{base_url}excel/landscape/{query_suffix}"
        response["actions"]["export_urls"]["excel_portrait"] = f"{base_url}excel/portrait/{query_suffix}"
        response["available_exports"] = ["excel", "pdf", "csv", "print", "pdf_landscape", "pdf_portrait", "excel_landscape", "excel_portrait"]
        return Response(response)


def _balance_sheet_subtitle(scope_names, scope, report):
    view_type = scope.get("view_type") or "summary"
    group_by = scope.get("account_group") or scope.get("group_by") or "accounthead"
    stock_mode = report.get("reporting", {}).get("stock_valuation_mode") or scope.get("stock_valuation_mode") or "auto"
    stock_method = report.get("reporting", {}).get("stock_valuation_method") or scope.get("stock_valuation_method") or "fifo"
    balance_diff = Decimal(str(report.get("totals", {}).get("assets") or "0")) - Decimal(str(report.get("totals", {}).get("liabilities_and_equity") or "0"))
    subentity_label = scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {subentity_label} | "
        f"Scope: {scope.get('scope_mode') or 'financial_year'} | "
        f"View: {view_type} | "
        f"Group by: {group_by} | "
        f"Stock: {stock_mode}/{stock_method} | "
        f"Delta: {balance_diff:.2f}"
    )


class _BaseBalanceSheetExportAPIView(_BaseFinancialReportAPIView):
    export_mode = "attachment"
    export_orientation = "landscape"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def build_orientation(self, request):
        orientation = str(
            getattr(self, "export_orientation", None)
            or request.query_params.get("orientation")
            or "landscape"
        ).strip().lower()
        return orientation if orientation in {"landscape", "portrait"} else "landscape"

    def report_data(self, request):
        scope = self.get_scope(request)
        reporting_policy = resolve_financial_reporting_policy(scope["entity"])
        data = build_balance_sheet(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            group_by=scope.get("account_group") or scope.get("group_by"),
            view_type=scope.get("view_type"),
            posted_only=scope.get("posted_only", REPORT_DEFAULTS["balance_sheet_posted_only_default"]),
            hide_zero_rows=scope.get("hide_zero_rows", REPORT_DEFAULTS["balance_sheet_hide_zero_rows_default"]),
            include_zero_balances=scope.get("include_zero_balances", REPORT_DEFAULTS["show_zero_balances_default"]),
            account_group=scope.get("account_group"),
            ledger_ids=scope.get("ledger_ids"),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", REPORT_DEFAULTS["default_page_size"]),
            period_by=scope.get("period_by"),
            reporting_policy=reporting_policy,
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        subtitle = _balance_sheet_subtitle(scope_names, scope, data)
        return scope, data, subtitle


class BalanceSheetExcelAPIView(_BaseBalanceSheetExportAPIView):
    def get(self, request):
        _scope, data, subtitle = self.report_data(request)
        headers, rows = _balance_sheet_export_table(data)
        content = _write_excel(
            "Balance Sheet",
            subtitle,
            headers,
            rows,
            numeric_columns=set(range(5, len(headers) + 1)),
            orientation=self.build_orientation(request),
        )
        return self.export_response(
            filename=f"BalanceSheet_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class BalanceSheetCSVAPIView(_BaseBalanceSheetExportAPIView):
    def get(self, request):
        _scope, data, subtitle = self.report_data(request)
        headers, rows = _balance_sheet_export_table(data)
        content = _write_csv(headers, rows)
        return self.export_response(
            filename=f"BalanceSheet_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class BalanceSheetPDFAPIView(_BaseBalanceSheetExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        orientation = self.build_orientation(request)
        pagesize = landscape(A4) if orientation == "landscape" else A4
        headers, rows = _balance_sheet_export_table(data)
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        balance_diff = Decimal(str(data.get("totals", {}).get("assets") or "0")) - Decimal(str(data.get("totals", {}).get("liabilities_and_equity") or "0"))
        periods = data.get("periods") or []
        meta_items = [
            ("Entity", scope_names["entity_name"] or "Selected entity"),
            ("Financial Year", scope_names["entityfin_name"] or "Current FY"),
            ("Subentity", scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("Scope", scope.get("scope_mode") or "financial_year"),
            ("View", scope.get("view_type") or "summary"),
            ("Group by", scope.get("account_group") or scope.get("group_by") or "accounthead"),
            ("Stock", f"{scope.get('stock_valuation_mode') or data.get('reporting', {}).get('stock_valuation_mode') or 'auto'}/{scope.get('stock_valuation_method') or data.get('reporting', {}).get('stock_valuation_method') or 'fifo'}"),
            ("Balance Status", "Balanced" if balance_diff == 0 else "Mismatch"),
            ("Difference", f"{balance_diff:.2f}"),
        ]
        if orientation == "portrait":
            col_widths = [54, 160, 82, 82] + [42] * len(periods) + [60]
        else:
            col_widths = [60, 175, 90, 90] + [52] * len(periods) + [72]
        content = _write_pdf(
            "Balance Sheet",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            meta_items=meta_items,
            numeric_columns=set(range(5, len(headers) + 1)),
            pagesize=pagesize,
        )
        return self.export_response(
            filename=f"BalanceSheet_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class BalanceSheetPrintAPIView(BalanceSheetPDFAPIView):
    export_mode = "inline"


class BalanceSheetExcelLandscapeAPIView(BalanceSheetExcelAPIView):
    export_orientation = "landscape"


class BalanceSheetExcelPortraitAPIView(BalanceSheetExcelAPIView):
    export_orientation = "portrait"


class BalanceSheetPDFLandscapeAPIView(BalanceSheetPDFAPIView):
    export_orientation = "landscape"


class BalanceSheetPDFPortraitAPIView(BalanceSheetPDFAPIView):
    export_orientation = "portrait"


class TradingAccountAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get(self, request):
        serializer = FinancialReportScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        valuation_method = (request.query_params.get("valuation_method") or "fifo").lower()
        level = (request.query_params.get("level") or ("account" if scope.get("view_type") == "detailed" else "head")).lower()
        period_by = (scope.get("period_by") or request.query_params.get("period_by") or "").strip().lower() or None

        self.enforce_scope(
            request,
            entity_id=scope["entity"],
            entityfinid_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
        )

        from_date = scope.get("from_date")
        to_date = scope.get("to_date")
        as_of_date = scope.get("as_of_date") or scope.get("as_on_date")
        if as_of_date and not to_date:
            to_date = as_of_date
        start, end = resolve_date_window(
            scope.get("entityfinid"),
            from_date,
            to_date,
        )
        data = build_trading_account_dynamic(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            startdate=start.isoformat(),
            enddate=end.isoformat(),
            period_by=period_by,
            posted_only=scope.get("posted_only", True),
            hide_zero_rows=not scope.get("include_zero_balance", False),
            view_type=scope.get("view_type") or ("detailed" if level == "account" else "summary"),
            account_group=scope.get("account_group") or scope.get("group_by"),
            ledger_ids=scope.get("ledger_ids") or None,
            valuation_method=valuation_method,
            level=level,
        )
        response = build_report_envelope(
            report_code="trading_account",
            report_name="Trading Account",
            payload=data,
            filters=self.build_filters(scope, level=level, valuation_method=valuation_method, start=start, end=end, period_by=period_by),
            defaults=REPORT_DEFAULTS,
        )
        response = _attach_financial_actions(
            response,
            request,
            export_base_path="/api/reports/financial/trading-account/",
        )
        query = _filtered_querydict(request, exclude=["page", "page_size", "orientation"])
        base_url = "/api/reports/financial/trading-account/"
        query_suffix = f"?{query}" if query else ""
        response["actions"]["export_urls"]["pdf_landscape"] = f"{base_url}pdf/landscape/{query_suffix}"
        response["actions"]["export_urls"]["pdf_portrait"] = f"{base_url}pdf/portrait/{query_suffix}"
        response["actions"]["export_urls"]["excel_landscape"] = f"{base_url}excel/landscape/{query_suffix}"
        response["actions"]["export_urls"]["excel_portrait"] = f"{base_url}excel/portrait/{query_suffix}"
        response["available_exports"] = ["excel", "pdf", "csv", "print", "pdf_landscape", "pdf_portrait", "excel_landscape", "excel_portrait"]
        return Response(response)

    def build_filters(self, scope, *, level: str, valuation_method: str, start, end, period_by: str | None):
        return {
            "entity": scope["entity"],
            "entityfinid": scope.get("entityfinid"),
            "financial_year": scope.get("financial_year") or scope.get("entityfinid"),
            "subentity": scope.get("subentity"),
            "scope_mode": scope.get("scope_mode"),
            "as_on_date": scope.get("as_on_date"),
            "from_date": start.isoformat() if start else None,
            "to_date": end.isoformat() if end else None,
            "period_by": period_by,
            "view_type": scope.get("view_type") or ("detailed" if level == "account" else "summary"),
            "account_group": scope.get("account_group") or scope.get("group_by"),
            "ledger_ids": scope.get("ledger_ids"),
            "posted_only": scope.get("posted_only", True),
            "include_zero_balance": scope.get("include_zero_balance", False),
            "hide_zero_rows": not scope.get("include_zero_balance", False),
            "valuation_method": valuation_method,
            "level": level,
        }


class _BaseTradingAccountExportAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL
    export_mode = "attachment"
    export_orientation = "landscape"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def build_orientation(self, request):
        orientation = str(
            getattr(self, "export_orientation", None)
            or request.query_params.get("orientation")
            or "landscape"
        ).strip().lower()
        return orientation if orientation in {"landscape", "portrait"} else "landscape"

    def report_data(self, request):
        serializer = FinancialReportScopeSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        valuation_method = (request.query_params.get("valuation_method") or "fifo").lower()
        level = (request.query_params.get("level") or ("account" if scope.get("view_type") == "detailed" else "head")).lower()
        period_by = (scope.get("period_by") or request.query_params.get("period_by") or "").strip().lower() or None
        self.enforce_scope(
            request,
            entity_id=scope["entity"],
            entityfinid_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
        )
        from_date = scope.get("from_date")
        to_date = scope.get("to_date")
        as_of_date = scope.get("as_of_date") or scope.get("as_on_date")
        if as_of_date and not to_date:
            to_date = as_of_date
        start, end = resolve_date_window(
            scope.get("entityfinid"),
            from_date,
            to_date,
        )
        data = build_trading_account_dynamic(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            startdate=start.isoformat(),
            enddate=end.isoformat(),
            period_by=period_by,
            posted_only=scope.get("posted_only", True),
            hide_zero_rows=not scope.get("include_zero_balance", False),
            view_type=scope.get("view_type") or ("detailed" if level == "account" else "summary"),
            account_group=scope.get("account_group") or scope.get("group_by"),
            ledger_ids=scope.get("ledger_ids") or None,
            valuation_method=valuation_method,
            level=level,
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        subtitle = _trading_account_subtitle(scope_names, scope, data)
        return scope, data, subtitle


class TradingAccountExcelAPIView(_BaseTradingAccountExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        headers, rows = _trading_account_export_table(data)
        content = _write_excel(
            "Trading Account",
            subtitle,
            headers,
            rows,
            numeric_columns={2, 3},
            orientation=self.build_orientation(request),
        )
        return self.export_response(
            filename=f"TradingAccount_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class TradingAccountCSVAPIView(_BaseTradingAccountExportAPIView):
    def get(self, request):
        _scope, data, subtitle = self.report_data(request)
        headers, rows = _trading_account_export_table(data)
        content = _write_csv(headers, rows)
        return self.export_response(
            filename=f"TradingAccount_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class TradingAccountPDFAPIView(_BaseTradingAccountExportAPIView):
    def get(self, request):
        scope, data, subtitle = self.report_data(request)
        orientation = self.build_orientation(request)
        pagesize = landscape(A4) if orientation == "landscape" else A4
        headers, rows = _trading_account_export_table(data)
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        meta_items = [
            ("Entity", scope_names["entity_name"] or "Selected entity"),
            ("Financial Year", scope_names["entityfin_name"] or "Current FY"),
            ("Subentity", scope_names["subentity_name"] or (f"Subentity {scope.get('subentity')}" if scope.get("subentity") else "All subentities")),
            ("View", scope.get("view_type") or "summary"),
            ("Group by", scope.get("account_group") or scope.get("group_by") or "accounthead"),
            ("Valuation", scope.get("valuation_method") or data.get("params", {}).get("valuation_method") or "fifo"),
            ("Gross Profit", data.get("gross_profit", "0.00") or "0.00"),
            ("Gross Loss", data.get("gross_loss", "0.00") or "0.00"),
        ]
        content = _write_pdf(
            "Trading Account",
            subtitle,
            headers,
            rows,
            col_widths=[46, 230, 70, 80] if orientation == "portrait" else [50, 250, 70, 90],
            meta_items=meta_items,
            numeric_columns={2, 3},
            pagesize=pagesize,
        )
        return self.export_response(
            filename=f"TradingAccount_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class TradingAccountPrintAPIView(TradingAccountPDFAPIView):
    export_mode = "inline"


class TradingAccountExcelLandscapeAPIView(TradingAccountExcelAPIView):
    export_orientation = "landscape"


class TradingAccountExcelPortraitAPIView(TradingAccountExcelAPIView):
    export_orientation = "portrait"


class TradingAccountPDFLandscapeAPIView(TradingAccountPDFAPIView):
    export_orientation = "landscape"


class TradingAccountPDFPortraitAPIView(TradingAccountPDFAPIView):
    export_orientation = "portrait"
