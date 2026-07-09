from __future__ import annotations

import csv
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from typing import Iterable, Sequence
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

EXECUTIVE_HEADER_BG = "233F73"
EXECUTIVE_SECTION_BG = "F3F7FD"
EXECUTIVE_SECTION_BORDER = "C8D6EA"
EXECUTIVE_ZEBRA_BG = "F8FBFF"
EXECUTIVE_TEXT_MUTED = "#5E6F89"
EXECUTIVE_TEXT_STRONG = "#1E2A3A"
EXECUTIVE_ROW_RULE = "#D6E0EE"
EXECUTIVE_TOTAL_BG = "#E4EEF9"
EXECUTIVE_GROUP_BG = "#EEF4FC"
EXECUTIVE_FINAL_BG = "#D6E4F6"
EXECUTIVE_SHEET_HEADER_BG = "EEF4FC"
EXECUTIVE_SHEET_HEADER_BORDER = "D2DDED"
EXECUTIVE_META_BG = "#FAFCFF"
EXECUTIVE_META_ALT_BG = "#F3F7FD"
EXECUTIVE_TITLE_RULE = "#C4D2E6"


def _pdf_page_decorator_factory(doc):
    def _page_decorator(canvas, page_doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D2DDED"))
        canvas.setLineWidth(0.8)
        canvas.line(doc.leftMargin, doc.bottomMargin - 8, doc.pagesize[0] - doc.rightMargin, doc.bottomMargin - 8)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor(EXECUTIVE_TEXT_MUTED))
        canvas.drawString(doc.leftMargin, doc.bottomMargin - 18, "Finacc ERP")
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, doc.bottomMargin - 18, f"Page {page_doc.page}")
        canvas.restoreState()

    return _page_decorator


def _build_pdf_header_table(*, title, width, title_style, top_padding, bottom_padding):
    header_table = Table([[Paragraph(escape(title), title_style)]], colWidths=[width])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{EXECUTIVE_HEADER_BG}")),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING", (0, 0), (-1, -1), top_padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), bottom_padding),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#CFD9E9")),
    ]))
    return header_table


def _build_pdf_meta_table(*, meta_items, width, text_style, background_color):
    column_count = 2 if len(meta_items or []) <= 4 else 3
    cell_width = width / float(column_count or 1)
    meta_cards = [
        Paragraph(
            f"<font color='#667085'><b>{escape(str(label))}</b></font><br/><font color='#1F2937'>{escape(str(value or '-'))}</font>",
            text_style,
        )
        for label, value in meta_items
    ]
    while len(meta_cards) % column_count != 0:
        meta_cards.append(Paragraph("&nbsp;", text_style))
    meta_rows = [meta_cards[index:index + column_count] for index in range(0, len(meta_cards), column_count)]
    meta_table = Table(meta_rows, colWidths=[cell_width] * column_count)
    meta_table.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor(EXECUTIVE_META_BG), colors.HexColor(EXECUTIVE_META_ALT_BG)]),
        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{EXECUTIVE_SECTION_BORDER}")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{EXECUTIVE_SECTION_BORDER}")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return meta_table


def _apply_pdf_document_metadata(doc, *, title, subtitle):
    doc.title = str(title or "Financial Report")
    doc.author = "Finacc ERP"
    doc.subject = str(subtitle or "")
    doc.creator = "Finacc ERP"
    doc.producer = "Finacc ERP"


@dataclass(frozen=True)
class ExportSection:
    title: str
    headers: list[str]
    rows: list[list[object]]
    row_kinds: list[str] | None = None
    numeric_columns: set[int] = field(default_factory=set)
    center_columns: set[int] = field(default_factory=set)
    col_widths: list[int | float] | None = None
    empty_message: str = "No data found."


def safe_filename(value):
    text = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in str(value or "").strip())
    text = text.strip("._-")
    return text or "report"


def build_report_filename(report_slug, *, entity_name=None, scope_label=None, extension):
    parts = [report_slug]
    if entity_name:
        parts.append(str(entity_name))
    if scope_label:
        parts.append(str(scope_label))
    safe_parts = [safe_filename(part) for part in parts if part]
    filename = "_".join(part for part in safe_parts if part) or safe_filename(report_slug)
    return f"{filename}.{extension.lstrip('.')}"


def filtered_querydict(request, *, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    return params.urlencode()


def attach_export_actions(payload, request, *, export_base_path, exclude=None, include_orientation=False):
    exclusions = ["page", "page_size", *(exclude or [])]
    if include_orientation:
        exclusions.append("orientation")
    query = filtered_querydict(request, exclude=exclusions)
    payload["actions"]["can_print"] = True
    payload["actions"]["export_urls"] = {
        "excel": f"{export_base_path}excel/?{query}",
        "pdf": f"{export_base_path}pdf/?{query}",
        "csv": f"{export_base_path}csv/?{query}",
        "print": f"{export_base_path}print/?{query}",
    }
    payload["available_exports"] = ["excel", "pdf", "csv", "print"]
    return payload


def workbook_styles():
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=EXECUTIVE_HEADER_BG)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, right, border


def write_csv(headers, rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def write_sectioned_csv(*, title, meta_items=None, sections: Sequence[ExportSection]):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([title])
    for label, value in meta_items or []:
        writer.writerow([label, value])
    if meta_items:
        writer.writerow([])

    for section_index, section in enumerate(sections):
        writer.writerow([section.title])
        writer.writerow(section.headers)
        if section.rows:
            writer.writerows(section.rows)
        else:
            writer.writerow([section.empty_message, *([""] * (len(section.headers) - 1))])
        if section_index != len(sections) - 1:
            writer.writerow([])

    return buffer.getvalue().encode("utf-8-sig")


def write_excel(title, subtitle, headers, rows, *, numeric_columns=None, orientation="landscape"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    orientation = (orientation or "landscape").strip().lower()
    if orientation not in {"landscape", "portrait"}:
        orientation = "landscape"
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    header_font, header_fill, center, left, right, border = workbook_styles()
    numeric_columns = set(numeric_columns or [])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = left
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).alignment = left

    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    summary_labels = {"totals", "summary", "balance difference", "debit total", "credit total"}
    summary_fill = PatternFill("solid", fgColor="E8F1FB")
    zebra_fill = PatternFill("solid", fgColor="F7FAFE")
    for row_index, row in enumerate(rows, start=header_row + 1):
        row_key = str(row[0] if row else "").strip().lower()
        is_summary_row = row_key in summary_labels or row_key.endswith(" total")
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            if is_summary_row:
                cell.fill = summary_fill
                cell.font = Font(bold=True)
            elif row_index % 2 == 1:
                cell.fill = zebra_fill
            cell.alignment = right if col_index in numeric_columns else left

    for col_index, header in enumerate(headers, start=1):
        width = max(len(str(header)) + 2, 14)
        for row in rows[:100]:
            width = max(width, len(str(row[col_index - 1])) + 2 if col_index - 1 < len(row) else width)
        ws.column_dimensions[get_column_letter(col_index)].width = min(width, 40)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _apply_sheet_header(ws, *, title, subtitle, column_count, left_alignment):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Calibri", size=9.5, color="667085")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    title_cell.fill = PatternFill("solid", fgColor=EXECUTIVE_HEADER_BG)
    title_cell.border = Border(
        left=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
        right=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
        top=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
        bottom=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
    )
    subtitle_cell.fill = PatternFill("solid", fgColor="FFFFFF")
    subtitle_cell.border = Border(
        left=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
        right=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
        top=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
        bottom=Side(style="thin", color=EXECUTIVE_SHEET_HEADER_BORDER),
    )

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def _prepare_sheet(ws, *, orientation):
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = None


def _append_section_to_sheet(ws, section: ExportSection, *, start_row, header_font, header_fill, center, left, right, border):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(section.headers))
    title_cell = ws.cell(row=start_row, column=1, value=section.title)
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = left
    title_cell.border = border
    ws.row_dimensions[start_row].height = 22

    header_row = start_row + 1
    for col_index, header in enumerate(section.headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 20

    rows = section.rows or [[section.empty_message, *([""] * (len(section.headers) - 1))]]
    zebra_fill = PatternFill("solid", fgColor="F8FBFF")
    group_fill = PatternFill("solid", fgColor="EEF4FC")
    subtotal_fill = PatternFill("solid", fgColor="E4EEF9")
    final_fill = PatternFill("solid", fgColor="D6E4F6")
    row_kinds = list(section.row_kinds or [])
    for row_index, row in enumerate(rows, start=header_row + 1):
        row_kind = row_kinds[row_index - header_row - 1] if row_index - header_row - 1 < len(row_kinds) else "detail"
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            if row_kind == "group":
                cell.fill = group_fill
                cell.font = Font(name="Calibri", bold=True, color="1F2937")
            elif row_kind == "subtotal":
                cell.fill = subtotal_fill
                cell.font = Font(name="Calibri", bold=True, color="1F2937")
            elif row_kind in {"final_total", "difference"}:
                cell.fill = final_fill
                cell.font = Font(name="Calibri", bold=True, color="1F2937")
            elif row_index % 2 == 1:
                cell.fill = zebra_fill
            if (col_index - 1) in section.center_columns:
                cell.alignment = center
            elif (col_index - 1) in section.numeric_columns:
                cell.alignment = right
            else:
                cell.alignment = left
        ws.row_dimensions[row_index].height = 18

    for col_index, header in enumerate(section.headers, start=1):
        width = max(len(str(header)) + 2, 12)
        for row in rows[:100]:
            width = max(width, len(str(row[col_index - 1])) + 2 if col_index - 1 < len(row) else width)
        ws.column_dimensions[get_column_letter(col_index)].width = min(width, 32)

    return header_row + len(rows) + 2


def write_sectioned_excel(*, title, subtitle, summary_items=None, sections: Sequence[ExportSection], orientation="landscape", freeze_header=True):
    wb = Workbook()
    header_font, _header_fill, center, left, right, border = workbook_styles()
    executive_fill = PatternFill("solid", fgColor=EXECUTIVE_HEADER_BG)
    orientation = (orientation or "landscape").strip().lower()
    if orientation not in {"landscape", "portrait"}:
        orientation = "landscape"

    if summary_items:
        ws = wb.active
        ws.title = "Summary"
        _prepare_sheet(ws, orientation=orientation)
        _apply_sheet_header(ws, title=title, subtitle=subtitle, column_count=2, left_alignment=left)
        current_row = 4
        for label, value in summary_items:
            label_cell = ws.cell(row=current_row, column=1, value=label)
            value_cell = ws.cell(row=current_row, column=2, value=value)
            label_cell.font = Font(name="Calibri", bold=True, color="1F2937")
            value_cell.font = Font(name="Calibri", bold=False, color="1F2937")
            label_cell.border = border
            value_cell.border = border
            if current_row % 2 == 0:
                label_cell.fill = PatternFill("solid", fgColor=EXECUTIVE_ZEBRA_BG)
                value_cell.fill = PatternFill("solid", fgColor=EXECUTIVE_ZEBRA_BG)
            else:
                label_cell.fill = PatternFill("solid", fgColor="FFFFFF")
                value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            label_cell.alignment = left
            value_cell.alignment = left
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 38
        ws.freeze_panes = "A4"
    else:
        ws = wb.active
        ws.title = (sections[0].title if sections else title)[:31]
        _prepare_sheet(ws, orientation=orientation)

    for index, section in enumerate(sections):
        sheet = wb.active if index == 0 and not summary_items else wb.create_sheet(section.title[:31])
        _prepare_sheet(sheet, orientation=orientation)
        _apply_sheet_header(sheet, title=title, subtitle=subtitle, column_count=len(section.headers), left_alignment=left)
        _append_section_to_sheet(
            sheet,
            section,
            start_row=4,
            header_font=header_font,
            header_fill=executive_fill,
            center=center,
            left=left,
            right=right,
            border=border,
        )
        if freeze_header:
            sheet.freeze_panes = "A6"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def write_pdf(
    title,
    subtitle,
    headers,
    rows,
    *,
    col_widths=None,
    meta_items=None,
    numeric_columns=None,
    center_columns=None,
    pagesize=None,
    statement_layout=False,
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize or landscape(A4), rightMargin=18, leftMargin=18, topMargin=24, bottomMargin=18)
    _apply_pdf_document_metadata(doc, title=title, subtitle=subtitle)
    styles = getSampleStyleSheet()
    title_style = styles["Title"].clone("FinancialPdfTitle")
    title_style.textColor = colors.HexColor("#FFFFFF")
    title_style.alignment = 1
    title_style.leading = 18
    title_style.fontSize = 16
    title_style.spaceAfter = 0

    subtitle_style = styles["BodyText"].clone("FinancialPdfSubtitle")
    subtitle_style.fontSize = 9
    subtitle_style.leading = 11
    subtitle_style.textColor = colors.HexColor("#4A5568")
    subtitle_style.spaceAfter = 0

    meta_label_style = styles["BodyText"].clone("FinancialPdfMetaLabel")
    meta_label_style.fontSize = 7.5
    meta_label_style.leading = 9
    meta_label_style.textColor = colors.HexColor("#5B6573")
    meta_label_style.spaceAfter = 0

    meta_value_style = styles["BodyText"].clone("FinancialPdfMetaValue")
    meta_value_style.fontSize = 9
    meta_value_style.leading = 11
    meta_value_style.textColor = colors.HexColor("#1F2937")
    meta_value_style.spaceAfter = 0

    def _page_decorator(canvas, _doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D7E2F1"))
        canvas.setLineWidth(0.8)
        canvas.line(doc.leftMargin, doc.bottomMargin - 8, doc.pagesize[0] - doc.rightMargin, doc.bottomMargin - 8)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, doc.bottomMargin - 18, f"Page {_doc.page}")
        canvas.restoreState()

    header_rows = [[Paragraph(escape(title), title_style)]]
    header_table = Table(header_rows, colWidths=[doc.width])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2F5597")),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#2F5597")),
    ]))

    story = [header_table, Spacer(1, 8), Paragraph(escape(subtitle), subtitle_style), Spacer(1, 10)]

    if meta_items:
        meta_cards = []
        for label, value in meta_items:
            meta_cards.append(
                Paragraph(
                    f"<font color='#5B6573'><b>{escape(str(label))}:</b></font><br/><font color='#1F2937'>{escape(str(value or '-'))}</font>",
                    meta_value_style,
                )
            )
        while len(meta_cards) % 3 != 0:
            meta_cards.append(Paragraph("&nbsp;", meta_label_style))
        meta_rows = [meta_cards[index:index + 3] for index in range(0, len(meta_cards), 3)]
        meta_table = Table(meta_rows, colWidths=[doc.width / 3.0] * 3)
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F7FB")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D7E2F1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7E2F1")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.extend([meta_table, Spacer(1, 10)])

    numeric_columns = set(numeric_columns or [])
    center_columns = set(center_columns or [])
    resolved_col_widths = normalize_col_widths(col_widths, doc.width, default_columns=len(headers))
    if not col_widths:
        resolved_col_widths = infer_pdf_col_widths(
            headers,
            rows,
            doc.width,
            numeric_columns=numeric_columns,
            center_columns=center_columns,
        )
    table_data = [headers, *rows]
    if resolved_col_widths:
        table_data = [
            [
                value
                if index in numeric_columns
                else truncate_text(value, resolved_col_widths[index] if index < len(resolved_col_widths) else 72)
                for index, value in enumerate(row)
            ]
            for row in table_data
        ]
    table = Table(table_data, colWidths=resolved_col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D6DDE8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FC")]),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
    ]))
    for col_index in numeric_columns:
        table.setStyle(TableStyle([("ALIGN", (col_index, 0), (col_index, -1), "RIGHT")]))
    for col_index in center_columns:
        table.setStyle(TableStyle([("ALIGN", (col_index, 0), (col_index, -1), "CENTER")]))

    if statement_layout:
        if numeric_columns:
            for col_index in numeric_columns:
                table.setStyle(TableStyle([("RIGHTPADDING", (col_index, 1), (col_index, -1), 12)]))
        for row_index, row in enumerate(rows, start=1):
            first_value = str(row[0] if row else "").strip()
            remaining = [str(value or "").strip() for value in row[1:]]
            is_blank_row = not first_value and not any(remaining)
            is_section_row = bool(first_value) and not any(remaining)
            normalized = first_value.lower()
            is_total_row = normalized.startswith("total ") or normalized in {
                "net profit / loss",
                "gross profit",
                "gross loss",
                "balance difference",
            }
            if is_blank_row:
                table.setStyle(TableStyle([
                    ("LINEBELOW", (0, row_index), (-1, row_index), 0, colors.white),
                    ("LINEABOVE", (0, row_index), (-1, row_index), 0, colors.white),
                    ("TOPPADDING", (0, row_index), (-1, row_index), 2),
                    ("BOTTOMPADDING", (0, row_index), (-1, row_index), 2),
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.white),
                ]))
            elif is_section_row:
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#EAF1FB")),
                    ("TEXTCOLOR", (0, row_index), (-1, row_index), colors.HexColor("#1E3A5F")),
                    ("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold"),
                    ("LINEABOVE", (0, row_index), (-1, row_index), 0.6, colors.HexColor("#B9C7DB")),
                ]))
            elif is_total_row:
                total_row_style = [
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#EEF3F9")),
                    ("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold"),
                    ("LINEABOVE", (0, row_index), (-1, row_index), 0.75, colors.HexColor("#8FA5C6")),
                    ("LINEBELOW", (0, row_index), (-1, row_index), 0.4, colors.HexColor("#C9D4E5")),
                ]
                if numeric_columns:
                    for col_index in numeric_columns:
                        total_row_style.append(("RIGHTPADDING", (col_index, row_index), (col_index, row_index), 4))
                table.setStyle(TableStyle(total_row_style))
            else:
                if numeric_columns:
                    table.setStyle(TableStyle([("RIGHTPADDING", (col_index, row_index), (col_index, row_index), 14) for col_index in numeric_columns]))

    story.append(table)
    doc.build(story, onFirstPage=_page_decorator, onLaterPages=_page_decorator)
    return buffer.getvalue()


def write_sectioned_pdf(*, title, subtitle, meta_items=None, sections: Sequence[ExportSection], pagesize=None, header_density="compact", metadata_visibility="compact"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize or landscape(A4), rightMargin=20, leftMargin=20, topMargin=18, bottomMargin=16)
    _apply_pdf_document_metadata(doc, title=title, subtitle=subtitle)
    styles = getSampleStyleSheet()

    title_style = styles["Title"].clone("FinancialSectionPdfTitle")
    title_style.textColor = colors.HexColor("#FFFFFF")
    title_style.alignment = 1
    title_style.leading = 17
    title_style.fontSize = 15

    subtitle_style = styles["BodyText"].clone("FinancialSectionPdfSubtitle")
    subtitle_style.fontSize = 8.8
    subtitle_style.leading = 10.6
    subtitle_style.textColor = colors.HexColor(EXECUTIVE_TEXT_MUTED)
    subtitle_style.alignment = 1

    section_style = styles["Heading4"].clone("FinancialSectionPdfHeading")
    section_style.fontSize = 10
    section_style.leading = 11
    section_style.textColor = colors.HexColor(EXECUTIVE_TEXT_STRONG)
    section_style.spaceAfter = 4

    body_style = styles["BodyText"].clone("FinancialSectionPdfBody")
    body_style.fontSize = 7.8
    body_style.leading = 9.5
    body_style.textColor = colors.HexColor(EXECUTIVE_TEXT_STRONG)
    body_bold_style = body_style.clone("FinancialSectionPdfBodyBold")
    body_bold_style.fontName = "Helvetica-Bold"
    body_bold_inverse_style = body_bold_style.clone("FinancialSectionPdfBodyBoldInverse")
    body_bold_inverse_style.textColor = colors.HexColor("#1E3A5F")
    body_right_style = body_style.clone("FinancialSectionPdfBodyRight")
    body_right_style.alignment = 2
    body_right_bold_style = body_right_style.clone("FinancialSectionPdfBodyRightBold")
    body_right_bold_style.fontName = "Helvetica-Bold"
    body_center_style = body_style.clone("FinancialSectionPdfBodyCenter")
    body_center_style.alignment = 1
    body_center_bold_style = body_center_style.clone("FinancialSectionPdfBodyCenterBold")
    body_center_bold_style.fontName = "Helvetica-Bold"

    density = (header_density or "compact").strip().lower()
    if density not in {"full", "compact", "minimal"}:
        density = "compact"
    metadata_mode = (metadata_visibility or "compact").strip().lower()
    if metadata_mode not in {"full", "compact", "hide"}:
        metadata_mode = "compact"
    title_font_size = 17 if density == "full" else 15 if density == "compact" else 12
    title_leading = 19 if density == "full" else 17 if density == "compact" else 14
    subtitle_font_size = 9.2 if density == "full" else 8.5 if density == "compact" else 7.4
    subtitle_leading = 11.4 if density == "full" else 10 if density == "compact" else 8.4
    header_top = 10 if density == "full" else 8 if density == "compact" else 5
    header_bottom = 11 if density == "full" else 8 if density == "compact" else 5
    spacer_after_header = 7 if density == "full" else 5 if density == "compact" else 3
    spacer_after_subtitle = 10 if density == "full" else 7 if density == "compact" else 4

    title_style.fontSize = title_font_size
    title_style.leading = title_leading
    subtitle_style.fontSize = subtitle_font_size
    subtitle_style.leading = subtitle_leading

    story = [_build_pdf_header_table(
        title=title,
        width=doc.width,
        title_style=title_style,
        top_padding=header_top,
        bottom_padding=header_bottom,
    )]
    subtitle_table = Table([[Paragraph(escape(subtitle), subtitle_style)]], colWidths=[doc.width])
    subtitle_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(EXECUTIVE_TITLE_RULE)),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([Spacer(1, spacer_after_header), subtitle_table, Spacer(1, spacer_after_subtitle)])

    if meta_items and metadata_mode != "hide":
        visible_meta_items = meta_items
        meta_table = _build_pdf_meta_table(
            meta_items=visible_meta_items,
            width=doc.width,
            text_style=body_style,
            background_color=colors.HexColor(f"#{EXECUTIVE_SECTION_BG}"),
        )
        story.extend([meta_table, Spacer(1, 6 if density == "minimal" else 9)])

    for section in sections:
        section_banner = Table([[Paragraph(escape(section.title), body_bold_inverse_style)]], colWidths=[doc.width])
        section_banner.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E3EBF7")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1E3A5F")),
            ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor("#C9D6E8")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.extend([section_banner, Spacer(1, 5)])
        rows = section.rows or [[section.empty_message, *([""] * (len(section.headers) - 1))]]
        row_kinds = list(section.row_kinds or [])
        resolved_col_widths = normalize_col_widths(section.col_widths, doc.width, default_columns=len(section.headers))
        if not section.col_widths:
            resolved_col_widths = infer_pdf_col_widths(
                section.headers,
                rows,
                doc.width,
                numeric_columns=section.numeric_columns,
                center_columns=section.center_columns,
            )
        table_rows = [section.headers]
        for row_index, row in enumerate(rows):
            row_kind = row_kinds[row_index] if row_index < len(row_kinds) else "detail"
            table_rows.append([
                value
                if isinstance(value, Paragraph)
                else Paragraph(
                    escape(str(value if value is not None else "")),
                    (
                        body_right_bold_style
                        if index in section.numeric_columns and row_kind in {"group", "subtotal", "final_total", "difference"}
                        else body_right_style
                        if index in section.numeric_columns
                        else body_center_bold_style
                        if index in section.center_columns and row_kind in {"group", "subtotal", "final_total", "difference"}
                        else body_center_style
                        if index in section.center_columns
                        else body_bold_style
                        if row_kind in {"group", "subtotal", "final_total", "difference"}
                        else body_style
                    ),
                )
                for index, value in enumerate(row)
            ])
        table = Table(table_rows, colWidths=resolved_col_widths, repeatRows=1)
        base_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{EXECUTIVE_HEADER_BG}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.6),
            ("FONTSIZE", (0, 1), (-1, -1), 7.9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor(EXECUTIVE_ROW_RULE)),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4.8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{EXECUTIVE_ZEBRA_BG}")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.55, colors.HexColor("#2D3748")),
        ]
        for row_index, row_kind in enumerate(row_kinds, start=1):
            if row_kind == "group":
                base_style.extend([
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#EDF3FB")),
                    ("LINEABOVE", (0, row_index), (-1, row_index), 0.5, colors.HexColor("#B9C7DB")),
                ])
            elif row_kind == "subtotal":
                base_style.extend([
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#E8EEF8")),
                    ("LINEABOVE", (0, row_index), (-1, row_index), 0.65, colors.HexColor("#A5B4CA")),
                    ("LINEBELOW", (0, row_index), (-1, row_index), 0.35, colors.HexColor("#C8D2E1")),
                ])
            elif row_kind in {"final_total", "difference"}:
                base_style.extend([
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#D7E4F5")),
                    ("LINEABOVE", (0, row_index), (-1, row_index), 0.9, colors.HexColor("#879AB8")),
                    ("LINEBELOW", (0, row_index), (-1, row_index), 0.45, colors.HexColor("#A7B7CD")),
                ])
        table.setStyle(TableStyle(base_style))
        for col_index in section.numeric_columns:
            table.setStyle(TableStyle([("ALIGN", (col_index, 0), (col_index, -1), "RIGHT")]))
        for col_index in section.center_columns:
            table.setStyle(TableStyle([("ALIGN", (col_index, 0), (col_index, -1), "CENTER")]))
        story.extend([table, Spacer(1, 12)])

    page_decorator = _pdf_page_decorator_factory(doc)
    doc.build(story, onFirstPage=page_decorator, onLaterPages=page_decorator)
    return buffer.getvalue()


def write_balance_sheet_statement_pdf(
    *,
    title,
    subtitle,
    meta_items=None,
    amount_headers: Sequence[str],
    sections: Sequence[dict],
    header_density="compact",
    metadata_visibility="compact",
    amount_col_widths: Sequence[int | float] | None = None,
    particulars_min_width: int | float | None = None,
):
    amount_headers = list(amount_headers or ["Current"])
    amount_count = len(amount_headers)
    resolved_amount_widths = [max(float(width or 0), 42.0) for width in (amount_col_widths or [])]
    if len(resolved_amount_widths) != amount_count:
        resolved_amount_widths = []

    if amount_count <= 2:
        pagesize = A4
        default_amount_width = 86
    elif amount_count <= 5:
        pagesize = landscape(A4)
        default_amount_width = 74
    else:
        pagesize = landscape(A3)
        default_amount_width = 64

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, rightMargin=20, leftMargin=20, topMargin=18, bottomMargin=16)
    _apply_pdf_document_metadata(doc, title=title, subtitle=subtitle)
    styles = getSampleStyleSheet()

    title_style = styles["Title"].clone("BalanceStatementPdfTitle")
    title_style.fontName = "Helvetica-Bold"
    title_style.textColor = colors.HexColor("#FFFFFF")
    title_style.alignment = 1
    title_style.leading = 17
    title_style.fontSize = 15

    subtitle_style = styles["BodyText"].clone("BalanceStatementPdfSubtitle")
    subtitle_style.fontSize = 8.5
    subtitle_style.leading = 10
    subtitle_style.textColor = colors.HexColor(EXECUTIVE_TEXT_MUTED)
    subtitle_style.alignment = 1

    body_style = styles["BodyText"].clone("BalanceStatementPdfBody")
    body_style.fontSize = 8
    body_style.leading = 9.8
    body_style.textColor = colors.HexColor(EXECUTIVE_TEXT_STRONG)
    body_bold_style = body_style.clone("BalanceStatementPdfBodyBold")
    body_bold_style.fontName = "Helvetica-Bold"
    body_bold_inverse_style = body_bold_style.clone("BalanceStatementPdfBodyBoldInverse")
    body_bold_inverse_style.textColor = colors.white
    body_italic_style = body_style.clone("BalanceStatementPdfBodyItalic")
    body_italic_style.fontName = "Helvetica-Oblique"
    body_right_style = body_style.clone("BalanceStatementPdfBodyRight")
    body_right_style.alignment = 2
    body_right_bold_style = body_right_style.clone("BalanceStatementPdfBodyRightBold")
    body_right_bold_style.fontName = "Helvetica-Bold"
    body_right_bold_inverse_style = body_right_bold_style.clone("BalanceStatementPdfBodyRightBoldInverse")
    body_right_bold_inverse_style.textColor = colors.white
    body_muted_style = body_style.clone("BalanceStatementPdfBodyMuted")
    body_muted_style.textColor = colors.HexColor(EXECUTIVE_TEXT_MUTED)
    body_muted_style.fontSize = max(7.0, body_style.fontSize - 0.3)
    section_label_style = body_bold_inverse_style.clone("BalanceStatementPdfSectionLabel")
    section_label_style.fontSize = max(8.3, body_style.fontSize + 0.5)
    section_label_style.leading = max(9.5, body_style.leading + 0.2)
    column_header_style = body_bold_style.clone("BalanceStatementPdfColumnHeader")
    column_header_style.fontSize = max(7.2, body_style.fontSize - 0.2)
    column_header_style.textColor = colors.HexColor(EXECUTIVE_TEXT_MUTED)
    column_header_right_style = body_right_bold_style.clone("BalanceStatementPdfColumnHeaderRight")
    column_header_right_style.fontSize = column_header_style.fontSize
    column_header_right_style.textColor = colors.HexColor(EXECUTIVE_TEXT_MUTED)
    group_total_style = body_bold_style.clone("BalanceStatementPdfGroupTotal")
    group_total_right_style = body_right_bold_style.clone("BalanceStatementPdfGroupTotalRight")
    compact_group_label_style = body_bold_style.clone("BalanceStatementPdfCompactGroupLabel")
    compact_group_label_style.fontSize = body_style.fontSize + 0.1
    compact_group_right_style = body_right_bold_style.clone("BalanceStatementPdfCompactGroupRight")
    compact_group_right_style.fontSize = body_right_style.fontSize + 0.1
    final_total_style = body_bold_style.clone("BalanceStatementPdfFinalTotal")
    final_total_style.fontSize = body_style.fontSize + 0.2
    final_total_right_style = body_right_bold_style.clone("BalanceStatementPdfFinalTotalRight")
    final_total_right_style.fontSize = body_right_bold_style.fontSize + 0.2

    density = (header_density or "compact").strip().lower()
    if density not in {"full", "compact", "minimal"}:
        density = "compact"
    metadata_mode = (metadata_visibility or "compact").strip().lower()
    if metadata_mode not in {"full", "compact", "hide"}:
        metadata_mode = "compact"
    title_font_size = 17 if density == "full" else 15 if density == "compact" else 12
    title_leading = 19 if density == "full" else 17 if density == "compact" else 14
    subtitle_font_size = 9.2 if density == "full" else 8.5 if density == "compact" else 7.4
    subtitle_leading = 11.4 if density == "full" else 10 if density == "compact" else 8.4
    header_top = 10 if density == "full" else 8 if density == "compact" else 5
    header_bottom = 11 if density == "full" else 8 if density == "compact" else 5
    spacer_after_header = 7 if density == "full" else 5 if density == "compact" else 3
    spacer_after_subtitle = 10 if density == "full" else 7 if density == "compact" else 4

    title_style.fontSize = title_font_size
    title_style.leading = title_leading
    subtitle_style.fontSize = subtitle_font_size
    subtitle_style.leading = subtitle_leading

    story = [_build_pdf_header_table(
        title=title,
        width=doc.width,
        title_style=title_style,
        top_padding=header_top,
        bottom_padding=header_bottom,
    )]
    subtitle_table = Table([[Paragraph(escape(subtitle), subtitle_style)]], colWidths=[doc.width])
    subtitle_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(EXECUTIVE_TITLE_RULE)),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([Spacer(1, spacer_after_header), subtitle_table, Spacer(1, spacer_after_subtitle)])

    if meta_items and metadata_mode != "hide":
        visible_meta_items = meta_items
        meta_table = _build_pdf_meta_table(
            meta_items=visible_meta_items,
            width=doc.width,
            text_style=body_style,
            background_color=colors.HexColor(f"#{EXECUTIVE_SECTION_BG}"),
        )
        story.extend([meta_table, Spacer(1, 10)])

    effective_amount_widths = resolved_amount_widths or [default_amount_width] * amount_count
    minimum_particulars_width = max(float(particulars_min_width or 250), 180.0)
    particulars_width = max(minimum_particulars_width, doc.width - sum(effective_amount_widths))
    col_widths = normalize_col_widths([particulars_width] + effective_amount_widths, doc.width, default_columns=amount_count + 1)

    for section in sections:
        section_title = str(section.get("title") or "")
        section_table = Table([
            [Paragraph(escape(section_title), section_label_style)] + [""] * amount_count,
            [Paragraph("Particulars", column_header_style)] + [
                Paragraph(escape(str(label)), column_header_right_style) for label in amount_headers
            ],
        ], colWidths=col_widths, repeatRows=2)
        section_table.setStyle(TableStyle([
            ("SPAN", (0, 0), (-1, 0)),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{EXECUTIVE_HEADER_BG}")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F7F9FC")),
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(EXECUTIVE_ROW_RULE)),
            ("LINEBELOW", (0, 1), (-1, 1), 0.45, colors.HexColor(EXECUTIVE_ROW_RULE)),
            ("LEFTPADDING", (0, 0), (-1, 0), 12),
            ("RIGHTPADDING", (0, 0), (-1, 0), 12),
            ("TOPPADDING", (0, 0), (-1, 0), 7),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
            ("LEFTPADDING", (0, 1), (-1, 1), 10),
            ("RIGHTPADDING", (0, 1), (-1, 1), 10),
            ("TOPPADDING", (0, 1), (-1, 1), 5),
            ("BOTTOMPADDING", (0, 1), (-1, 1), 5),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ]))
        story.extend([KeepTogether([section_table]), Spacer(1, 5)])

        for group in section.get("groups") or []:
            compact_row = group.get("compact_row") or None
            if compact_row:
                compact_table = Table([[
                    Paragraph(escape(str(compact_row.get("label") or "")), compact_group_label_style),
                    *[
                        Paragraph(escape(str(value or "")), compact_group_right_style)
                        for value in compact_row.get("amounts") or []
                    ],
                ]], colWidths=col_widths)
                compact_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.28, colors.HexColor(EXECUTIVE_ROW_RULE)),
                    ("LEFTPADDING", (0, 0), (0, -1), 12),
                    ("RIGHTPADDING", (0, 0), (0, -1), 10),
                    ("LEFTPADDING", (1, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (1, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                story.extend([compact_table, Spacer(1, 4)])
                continue

            group_title = str(group.get("title") or "")
            group_table = Table([[Paragraph(escape(group_title), body_italic_style)] + [""] * amount_count], colWidths=col_widths)
            group_table.setStyle(TableStyle([
                ("SPAN", (0, 0), (-1, 0)),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1FB")),
                ("LINEBEFORE", (0, 0), (0, 0), 2.2, colors.HexColor(f"#{EXECUTIVE_HEADER_BG}")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 11),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.extend([KeepTogether([group_table]), Spacer(1, 2)])

            detail_rows = []
            for line in group.get("lines") or []:
                detail_rows.append(
                    [Paragraph(escape(str(line.get("label") or "")), body_style)]
                    + [Paragraph(escape(str(value or "")), body_right_style) for value in line.get("amounts") or []]
                )
            if detail_rows:
                detail_table = Table(detail_rows, colWidths=col_widths)
                detail_table.setStyle(TableStyle([
                    ("LEFTPADDING", (0, 0), (-1, -1), 19),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 4.4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4.4),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor(f"#{EXECUTIVE_ZEBRA_BG}")]),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor(EXECUTIVE_ROW_RULE)),
                ]))
                story.extend([detail_table, Spacer(1, 2)])

            if group.get("total_label"):
                total_row = [Paragraph(escape(str(group.get("total_label") or "")), group_total_style)] + [
                    Paragraph(escape(str(value or "")), group_total_right_style) for value in group.get("total_amounts") or []
                ]
                total_table = Table([total_row], colWidths=col_widths)
                total_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF3F9")),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.45, colors.HexColor(EXECUTIVE_ROW_RULE)),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ]))
                story.extend([KeepTogether([total_table]), Spacer(1, 8)])

        if section.get("total_label"):
            section_total_row = [Paragraph(escape(str(section.get("total_label") or "")), final_total_style)] + [
                Paragraph(escape(str(value or "")), final_total_right_style) for value in section.get("total_amounts") or []
            ]
            section_total_table = Table([section_total_row], colWidths=col_widths)
            section_total_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#D7E4F5")),
                ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor(EXECUTIVE_ROW_RULE)),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]))
            story.extend([KeepTogether([section_total_table]), Spacer(1, 14)])

    page_decorator = _pdf_page_decorator_factory(doc)
    doc.build(story, onFirstPage=page_decorator, onLaterPages=page_decorator)
    return buffer.getvalue()


def truncate_text(value, width_points, *, min_chars=8):
    text = "" if value is None else str(value)
    max_chars = max(min_chars, int((width_points or 72) / 5))
    if len(text) <= max_chars:
        return text
    if max_chars <= 3:
        return text[:max_chars]
    return f"{text[: max_chars - 3].rstrip()}..."


def infer_pdf_col_widths(
    headers: Sequence[object],
    rows: Sequence[Sequence[object]],
    available_width: float,
    *,
    numeric_columns: set[int] | None = None,
    center_columns: set[int] | None = None,
) -> list[float] | None:
    headers = list(headers or [])
    if not headers:
        return None

    numeric_columns = set(numeric_columns or [])
    center_columns = set(center_columns or [])
    sample_rows = list(rows[:80]) if rows else []
    widths: list[float] = []

    for index, header in enumerate(headers):
        header_text = str(header or "").strip()
        data_lengths = [len(str(row[index] or "").strip()) for row in sample_rows if index < len(row)]
        max_length = max([len(header_text), *data_lengths, 6])

        if index in numeric_columns:
            preferred = min(max(60.0, max_length * 5.2), 92.0)
        elif index in center_columns:
            preferred = min(max(64.0, max_length * 5.8), 110.0)
        else:
            preferred = min(max(88.0, max_length * 6.3), 260.0)

        widths.append(preferred)

    total = sum(widths)
    if total <= 0:
        return normalize_col_widths(None, available_width, default_columns=len(headers))

    scale = available_width / total
    scaled = [round(width * scale, 2) for width in widths]
    drift = round(available_width - sum(scaled), 2)
    scaled[-1] = round(scaled[-1] + drift, 2)
    return scaled


def normalize_col_widths(col_widths: Sequence[int | float] | None, available_width: float, *, default_columns: int) -> list[float] | None:
    if col_widths:
        widths = [max(float(width or 0), 24.0) for width in col_widths]
        total = sum(widths)
        if total > 0:
            scale = available_width / total
            widths = [round(width * scale, 2) for width in widths]
            drift = round(available_width - sum(widths), 2)
            if widths:
                widths[-1] = round(widths[-1] + drift, 2)
        return widths
    if default_columns <= 0:
        return None
    width = round(available_width / default_columns, 2)
    widths = [width] * default_columns
    drift = round(available_width - sum(widths), 2)
    widths[-1] = round(widths[-1] + drift, 2)
    return widths
