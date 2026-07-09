from __future__ import annotations

import csv
from io import BytesIO, StringIO

from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils import timezone
from entity.models import Entity, EntityFinancialYear, SubEntity
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from core.entitlements import ScopedEntitlementMixin
from reports.api.report_permissions import assert_any_report_permission
from reports.gstr1.services.report import Gstr1ReportService
from reports.gstr3b.services import Gstr3bSummaryService
from reports.schemas.common import build_report_envelope
from reports.services.gst_exception_dashboard import build_gst_exception_dashboard
from reports.services.gst_reconciliation import build_gstr1_vs_gstr3b_reconciliation
from subscriptions.services import SubscriptionLimitCodes, SubscriptionService


class GstExceptionDashboardAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get(self, request):
        gstr1_service = Gstr1ReportService()
        gstr3b_service = Gstr3bSummaryService()
        try:
            gstr1_scope = gstr1_service.build_scope(request.query_params)
            gstr3b_scope = gstr3b_service.build_scope(request.query_params)
        except ValidationError as exc:
            return Response(exc.message_dict, status=400)

        self.enforce_scope(
            request,
            entity_id=gstr1_scope.entity_id,
            entityfinid_id=gstr1_scope.entityfinid_id,
            subentity_id=gstr1_scope.subentity_id,
        )
        assert_any_report_permission(
            user=request.user,
            entity_id=gstr1_scope.entity_id,
            required_permissions=("reports.gst_exception_dashboard.view",),
            message="You do not have permission to access the GST exception dashboard.",
        )
        gstr1_warnings = gstr1_service.validations(gstr1_scope)
        gstr3b_warnings = gstr3b_service.validations(gstr3b_scope)
        reconciliation = build_gstr1_vs_gstr3b_reconciliation(
            gstr1_summary=gstr1_service.summary(gstr1_scope),
            gstr3b_summary=gstr3b_service.build(gstr3b_scope),
            scope_params={
                "entityfinid": gstr1_scope.entityfinid_id,
                "subentity": gstr1_scope.subentity_id,
                "from_date": gstr1_scope.from_date,
                "to_date": gstr1_scope.to_date,
            },
            gstr1_scope=gstr1_scope,
            include_contributors=False,
        )
        payload = build_gst_exception_dashboard(
            gstr1_warnings=gstr1_warnings,
            gstr3b_warnings=gstr3b_warnings,
            reconciliation_payload=reconciliation,
            scope_params={
                "entityfinid": gstr1_scope.entityfinid_id,
                "subentity": gstr1_scope.subentity_id,
                "from_date": gstr1_scope.from_date,
                "to_date": gstr1_scope.to_date,
            },
        )
        response = build_report_envelope(
            report_code="gst-exception-dashboard",
            report_name="GST Exception Dashboard",
            payload=payload,
            filters={
                "entity": gstr1_scope.entity_id,
                "entityfinid": gstr1_scope.entityfinid_id,
                "subentity": gstr1_scope.subentity_id,
                "month": gstr1_scope.month,
                "year": gstr1_scope.year,
                "from_date": gstr1_scope.from_date,
                "to_date": gstr1_scope.to_date,
            },
            defaults={
                "decimal_places": 2,
                "show_zero_balances_default": True,
                "show_opening_balance_default": False,
                "enable_drilldown": False,
            },
        )
        query = request.GET.copy()
        query.pop("page", None)
        query.pop("page_size", None)
        encoded = query.urlencode()
        response["actions"]["export_urls"] = {
            "excel": f"/api/reports/gst-exception-dashboard/export/?format=xlsx&{encoded}",
            "csv": f"/api/reports/gst-exception-dashboard/export/?format=csv&{encoded}",
            "json": f"/api/reports/gst-exception-dashboard/export/?format=json&{encoded}",
        }
        response["available_exports"] = ["excel", "csv", "json"]
        return Response(response)


class GstExceptionDashboardExportAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get(self, request):
        export_format = (request.query_params.get("format") or "json").lower()
        gstr1_service = Gstr1ReportService()
        gstr3b_service = Gstr3bSummaryService()
        params = request.query_params.copy()
        params.pop("format", None)
        try:
            gstr1_scope = gstr1_service.build_scope(params)
            gstr3b_scope = gstr3b_service.build_scope(params)
        except ValidationError as exc:
            return Response(exc.message_dict, status=400)

        self.enforce_scope(
            request,
            entity_id=gstr1_scope.entity_id,
            entityfinid_id=gstr1_scope.entityfinid_id,
            subentity_id=gstr1_scope.subentity_id,
        )
        assert_any_report_permission(
            user=request.user,
            entity_id=gstr1_scope.entity_id,
            required_permissions=("reports.gst_exception_dashboard.view",),
            message="You do not have permission to access the GST exception dashboard.",
        )
        gstr1_warnings = gstr1_service.validations(gstr1_scope)
        gstr3b_warnings = gstr3b_service.validations(gstr3b_scope)
        reconciliation = build_gstr1_vs_gstr3b_reconciliation(
            gstr1_summary=gstr1_service.summary(gstr1_scope),
            gstr3b_summary=gstr3b_service.build(gstr3b_scope),
            scope_params={
                "entityfinid": gstr1_scope.entityfinid_id,
                "subentity": gstr1_scope.subentity_id,
                "from_date": gstr1_scope.from_date,
                "to_date": gstr1_scope.to_date,
            },
            gstr1_scope=gstr1_scope,
            include_contributors=False,
        )
        payload = build_gst_exception_dashboard(
            gstr1_warnings=gstr1_warnings,
            gstr3b_warnings=gstr3b_warnings,
            reconciliation_payload=reconciliation,
            scope_params={
                "entityfinid": gstr1_scope.entityfinid_id,
                "subentity": gstr1_scope.subentity_id,
                "from_date": gstr1_scope.from_date,
                "to_date": gstr1_scope.to_date,
            },
        )
        if export_format == "json":
            return Response(payload)
        if export_format == "csv":
            content = self._export_csv(payload, scope=gstr1_scope)
            return _file_response("GST_Exception_Dashboard.csv", content, "text/csv")
        if export_format == "xlsx":
            content = self._export_xlsx(payload, scope=gstr1_scope)
            return _file_response(
                "GST_Exception_Dashboard.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return Response({"detail": "Unsupported export format."}, status=400)

    def _export_csv(self, payload: dict, *, scope) -> bytes:
        stream = StringIO()
        writer = csv.writer(stream)
        audit = _audit_context(scope)
        overview = payload.get("overview") or {}
        warnings = payload.get("warnings") or []

        writer.writerow(["Report", audit["report_title"]])
        writer.writerow(["Period", audit["period_label"]])
        writer.writerow(["Financial Year", audit["financial_year"]])
        writer.writerow(["Subentity", audit["subentity"]])
        writer.writerow(["Generated On", audit["generated_on"]])
        writer.writerow([])

        writer.writerow(["Overview"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Exceptions", overview.get("total_exception_count", 0)])
        writer.writerow(["Blocking Exceptions", overview.get("blocking_exception_count", 0)])
        writer.writerow(["GSTR-1 Warnings", overview.get("gstr1_warning_count", 0)])
        writer.writerow(["GSTR-3B Warnings", overview.get("gstr3b_warning_count", 0)])
        writer.writerow(["Reconciliation Mismatches", overview.get("reconciliation_mismatch_count", 0)])
        writer.writerow(["Reconciliation Advisories", overview.get("reconciliation_advisory_count", 0)])
        writer.writerow(["Max Reconciliation Tax Gap", overview.get("max_reconciliation_tax_gap", "0.00")])
        writer.writerow([])

        writer.writerow(["Source Summary"])
        writer.writerow(["Source", "Total", "Errors", "Warnings", "Infos"])
        totals = {"total": 0, "errors": 0, "warnings": 0, "infos": 0}
        for row in payload.get("source_summary") or []:
            totals["total"] += int(row.get("total") or 0)
            totals["errors"] += int(row.get("errors") or 0)
            totals["warnings"] += int(row.get("warnings") or 0)
            totals["infos"] += int(row.get("infos") or 0)
            writer.writerow([row.get("source"), row.get("total"), row.get("errors"), row.get("warnings"), row.get("infos")])
        writer.writerow(["Report Total", totals["total"], totals["errors"], totals["warnings"], totals["infos"]])
        writer.writerow([])

        writer.writerow(["Category Spotlight"])
        writer.writerow(["Category", "Count"])
        for row in payload.get("category_spotlight") or []:
            writer.writerow([row.get("category"), row.get("count")])
        writer.writerow([])

        writer.writerow(["Exception Rows"])
        writer.writerow(["Source", "Severity", "Category", "Code / Label", "Affected Rows", "Reference", "Message / Note"])
        for row in payload.get("gstr1_exception_rows") or []:
            writer.writerow(["GSTR-1", row.get("severity"), row.get("category"), row.get("code"), row.get("affected_rows"), row.get("sample_references"), row.get("message")])
        for row in payload.get("gstr3b_exception_rows") or []:
            writer.writerow(["GSTR-3B", row.get("severity"), row.get("category"), row.get("code"), row.get("affected_rows"), row.get("sample_references"), row.get("message")])
        for row in payload.get("reconciliation_rows") or []:
            writer.writerow([
                "Reconciliation",
                "error",
                "Reconciliation Gaps",
                row.get("label"),
                1,
                "-",
                f"Taxable diff {row.get('difference_taxable_value')}; Tax diff {row.get('difference_total_tax')}. {row.get('note') or ''}".strip(),
            ])
        writer.writerow([])

        writer.writerow(["Warnings"])
        writer.writerow(["Severity", "Code", "Message"])
        for warning in warnings:
            writer.writerow([warning.get("severity"), warning.get("code"), warning.get("message")])

        return stream.getvalue().encode("utf-8")

    def _export_xlsx(self, payload: dict, *, scope) -> bytes:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "GST Exception Dashboard"
        styles = _workbook_styles()
        audit = _audit_context(scope)
        overview = payload.get("overview") or {}

        _write_title_block(
            ws,
            title=audit["report_title"],
            subtitle=f"{audit['period_label']} | {audit['financial_year']} | {audit['subentity']}",
            generated_on=audit["generated_on"],
            width=5,
            styles=styles,
        )

        ws.append(["Overview"])
        _style_section_title(ws, ws.max_row, 1, 2, styles)
        ws.append(["Metric", "Value"])
        _style_header_row(ws, ws.max_row, 2, styles)
        overview_start = ws.max_row + 1
        overview_rows = [
            ["Total Exceptions", overview.get("total_exception_count", 0)],
            ["Blocking Exceptions", overview.get("blocking_exception_count", 0)],
            ["GSTR-1 Warnings", overview.get("gstr1_warning_count", 0)],
            ["GSTR-3B Warnings", overview.get("gstr3b_warning_count", 0)],
            ["Reconciliation Mismatches", overview.get("reconciliation_mismatch_count", 0)],
            ["Reconciliation Advisories", overview.get("reconciliation_advisory_count", 0)],
            ["Max Reconciliation Tax Gap", overview.get("max_reconciliation_tax_gap", "0.00")],
        ]
        for row in overview_rows:
            ws.append(row)
        _style_body_rows(ws, overview_start, ws.max_row, 2, numeric_columns={2}, styles=styles)

        ws.append([])
        ws.append(["Source Summary"])
        _style_section_title(ws, ws.max_row, 1, 5, styles)
        ws.append(["Source", "Total", "Errors", "Warnings", "Infos"])
        _style_header_row(ws, ws.max_row, 5, styles)
        source_start = ws.max_row + 1
        totals = {"total": 0, "errors": 0, "warnings": 0, "infos": 0}
        for row in payload.get("source_summary") or []:
            totals["total"] += int(row.get("total") or 0)
            totals["errors"] += int(row.get("errors") or 0)
            totals["warnings"] += int(row.get("warnings") or 0)
            totals["infos"] += int(row.get("infos") or 0)
            ws.append([row.get("source"), row.get("total"), row.get("errors"), row.get("warnings"), row.get("infos")])
        ws.append(["Report Total", totals["total"], totals["errors"], totals["warnings"], totals["infos"]])
        _style_body_rows(ws, source_start, ws.max_row, 5, numeric_columns={2, 3, 4, 5}, styles=styles)
        total_row = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.font = styles["section_font"]
            cell.fill = styles["total_fill"]

        ws.append([])
        ws.append(["Category Spotlight"])
        _style_section_title(ws, ws.max_row, 1, 2, styles)
        ws.append(["Category", "Count"])
        _style_header_row(ws, ws.max_row, 2, styles)
        category_start = ws.max_row + 1
        for row in payload.get("category_spotlight") or []:
            ws.append([row.get("category"), row.get("count")])
        _style_body_rows(ws, category_start, ws.max_row, 2, numeric_columns={2}, styles=styles)

        ws.append([])
        ws.append(["Exception Rows"])
        _style_section_title(ws, ws.max_row, 1, 7, styles)
        ws.append(["Source", "Severity", "Category", "Code / Label", "Affected Rows", "Reference", "Message / Note"])
        _style_header_row(ws, ws.max_row, 7, styles)
        detail_start = ws.max_row + 1
        for row in payload.get("gstr1_exception_rows") or []:
            ws.append(["GSTR-1", row.get("severity"), row.get("category"), row.get("code"), row.get("affected_rows"), row.get("sample_references"), row.get("message")])
        for row in payload.get("gstr3b_exception_rows") or []:
            ws.append(["GSTR-3B", row.get("severity"), row.get("category"), row.get("code"), row.get("affected_rows"), row.get("sample_references"), row.get("message")])
        for row in payload.get("reconciliation_rows") or []:
            ws.append([
                "Reconciliation",
                "error",
                "Reconciliation Gaps",
                row.get("label"),
                1,
                "-",
                f"Taxable diff {row.get('difference_taxable_value')}; Tax diff {row.get('difference_total_tax')}. {row.get('note') or ''}".strip(),
            ])
        _style_body_rows(ws, detail_start, ws.max_row, 7, numeric_columns={5}, styles=styles)

        warnings = payload.get("warnings") or []
        if warnings:
            warning_ws = workbook.create_sheet("Warnings")
            _write_title_block(
                warning_ws,
                title="GST Exception Dashboard Warnings",
                subtitle=f"{audit['period_label']} | {audit['subentity']}",
                generated_on=audit["generated_on"],
                width=3,
                styles=styles,
            )
            warning_ws.append(["Severity", "Code", "Message"])
            _style_header_row(warning_ws, warning_ws.max_row, 3, styles)
            warning_start = warning_ws.max_row + 1
            for warning in warnings:
                warning_ws.append([warning.get("severity"), warning.get("code"), warning.get("message")])
            _style_body_rows(warning_ws, warning_start, warning_ws.max_row, 3, numeric_columns=set(), styles=styles)
            _apply_column_widths(warning_ws, {1: 16, 2: 28, 3: 88})

        _apply_column_widths(
            ws,
            {1: 18, 2: 14, 3: 28, 4: 24, 5: 16, 6: 32, 7: 84}
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def _file_response(filename, content, content_type):
    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _audit_context(scope):
    entity_name = (
        Entity.objects.filter(id=scope.entity_id).values_list("entityname", flat=True).first()
        or f"Entity {scope.entity_id}"
    )
    financial_year = "All financial years"
    if scope.entityfinid_id:
        financial_year = (
            EntityFinancialYear.objects.filter(id=scope.entityfinid_id, entity_id=scope.entity_id)
            .values_list("desc", flat=True)
            .first()
            or f"FY {scope.entityfinid_id}"
        )
    subentity = "All subentities"
    if scope.subentity_id:
        subentity = (
            SubEntity.objects.filter(id=scope.subentity_id, entity_id=scope.entity_id)
            .values_list("subentityname", flat=True)
            .first()
            or f"Subentity {scope.subentity_id}"
        )
    return {
        "report_title": f"{entity_name} - GST Exception Dashboard",
        "period_label": f"{scope.from_date} to {scope.to_date}",
        "financial_year": financial_year,
        "subentity": subentity,
        "generated_on": timezone.localtime().strftime("%d %b %Y %I:%M %p"),
    }


def _workbook_styles():
    header_fill = PatternFill("solid", fgColor="2D4F83")
    section_fill = PatternFill("solid", fgColor="EEF4FB")
    total_fill = PatternFill("solid", fgColor="EAF2FB")
    border_color = "D6DEEA"
    thin = Side(style="thin", color=border_color)
    return {
        "title_font": Font(name="Calibri", bold=True, size=16, color="20324D"),
        "subtitle_font": Font(name="Calibri", size=10, color="5E6B80"),
        "header_font": Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
        "body_font": Font(name="Calibri", size=10, color="233248"),
        "section_font": Font(name="Calibri", bold=True, size=11, color="20324D"),
        "section_fill": section_fill,
        "header_fill": header_fill,
        "total_fill": total_fill,
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "left": Alignment(horizontal="left", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "wrap_left": Alignment(horizontal="left", vertical="top", wrap_text=True),
    }


def _write_title_block(ws, *, title, subtitle, generated_on, width, styles):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws["A1"] = title
    ws["A1"].font = styles["title_font"]
    ws["A1"].alignment = styles["left"]

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws["A2"] = subtitle
    ws["A2"].font = styles["subtitle_font"]
    ws["A2"].alignment = styles["left"]

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=width)
    ws["A3"] = f"Generated on {generated_on}"
    ws["A3"].font = styles["subtitle_font"]
    ws["A3"].alignment = styles["left"]
    ws.append([])


def _style_section_title(ws, row_number, start_column, end_column, styles):
    for col in range(start_column, end_column + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.fill = styles["section_fill"]
        cell.border = styles["border"]
        cell.font = styles["section_font"]
        cell.alignment = styles["left"]


def _style_header_row(ws, row_number, width, styles):
    for col in range(1, width + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["left"]


def _style_body_rows(ws, start_row, end_row, width, *, numeric_columns, styles):
    for row in range(start_row, end_row + 1):
        for col in range(1, width + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = styles["body_font"]
            cell.border = styles["border"]
            cell.alignment = styles["right"] if col in numeric_columns else styles["wrap_left"]


def _apply_column_widths(ws, widths):
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width
