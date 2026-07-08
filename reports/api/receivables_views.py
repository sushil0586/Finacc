from __future__ import annotations

import csv
import json
from html import escape
from io import BytesIO, StringIO

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, LEGAL, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas as pdf_canvas
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from financial.models import account
from sales.models.sales_ar import CustomerSettlement
from reports.schemas.common import build_report_envelope
from reports.selectors.financial import resolve_scope_names
from reports.schemas.receivables_reports import CollectionsHistoryScopeSerializer, ReceivableAgingScopeSerializer, ReceivableReportScopeSerializer
from reports.services.receivables import build_collections_history_report, build_customer_outstanding_report, build_open_items_report, build_receivable_aging_report


RECEIVABLE_DEFAULTS = {
    "default_page_size": 100,
    "decimal_places": 2,
    "show_zero_balances_default": False,
    "show_opening_balance_default": True,
    "enable_drilldown": True,
}

CUSTOMER_OUTSTANDING_PDF_LAYOUT = {
    "page_size": landscape(LEGAL),
    "col_widths": [110, 40, 48, 54, 54, 54, 68, 54, 48, 48, 40, 48, 48, 128, 38, 62],
    "header_font_size": 7.2,
    "body_font_size": 6.6,
    "leading": 7.6,
}

RECEIVABLE_AGING_SUMMARY_PDF_LAYOUT = {
    "page_size": landscape(LEGAL),
    "col_widths": [150, 42, 62, 62, 54, 54, 46, 46, 42, 56, 62, 44, 60, 50],
    "header_font_size": 7.2,
    "body_font_size": 6.6,
    "leading": 7.6,
}

RECEIVABLE_AGING_INVOICE_PDF_LAYOUT = {
    "page_size": landscape(LEGAL),
    "col_widths": [82, 36, 74, 56, 56, 42, 58, 56, 56, 46, 46, 46, 46, 42, 48],
    "header_font_size": 7.0,
    "body_font_size": 6.4,
    "leading": 7.4,
}

OPEN_ITEMS_PDF_LAYOUT = {
    "page_size": landscape(LEGAL),
    "col_widths": [88, 52, 56, 56, 68, 44, 62, 56, 56, 64, 40, 66, 48, 80],
    "header_font_size": 7.1,
    "body_font_size": 6.5,
    "leading": 7.4,
}

COLLECTIONS_HISTORY_PDF_LAYOUT = {
    "page_size": landscape(LEGAL),
    "col_widths": [84, 40, 54, 70, 66, 60, 52, 44, 36, 62, 66, 48, 78, 56],
    "header_font_size": 7.0,
    "body_font_size": 6.4,
    "leading": 7.3,
}


def _receivable_scope_filters(scope):
    return {
        "entity": scope["entity"],
        "entityfinid": scope.get("entityfinid"),
        "subentity": scope.get("subentity"),
        "from_date": scope.get("from_date"),
        "to_date": scope.get("to_date"),
        "as_of_date": scope.get("as_of_date"),
        "customer": scope.get("customer"),
        "customer_group": scope.get("customer_group"),
        "region": scope.get("region"),
        "territory": scope.get("territory"),
        "salesperson": scope.get("salesperson"),
        "currency": scope.get("currency"),
        "overdue_only": scope.get("overdue_only", False),
        "credit_limit_exceeded": scope.get("credit_limit_exceeded", False),
        "exception_only": scope.get("exception_only", False),
        "settlement_type": scope.get("settlement_type"),
        "status": scope.get("status"),
        "outstanding_gt": scope.get("outstanding_gt"),
        "search": scope.get("search"),
        "sort_by": scope.get("sort_by"),
        "sort_order": scope.get("sort_order", "desc"),
        "page": scope.get("page", 1),
        "page_size": scope.get("page_size", RECEIVABLE_DEFAULTS["default_page_size"]),
        "view": scope.get("view"),
    }


def _filtered_querydict(request, *, overrides=None, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    for key, value in (overrides or {}).items():
        if value is None:
            params.pop(key, None)
        else:
            params[key] = str(value)
    return params.urlencode()


def _attach_receivable_actions(payload, request, *, export_base_path):
    query = _filtered_querydict(request, exclude=["page", "page_size"])
    payload["actions"]["can_print"] = True
    payload["actions"]["export_urls"] = {
        "excel": f"{export_base_path}excel/?{query}",
        "pdf": f"{export_base_path}pdf/?{query}",
        "csv": f"{export_base_path}csv/?{query}",
        "print": f"{export_base_path}print/?{query}",
    }
    return payload


def _format_scope_date(value):
    if not value:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    return str(value)


def _format_display_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d %b %Y")
    text = str(value).strip()
    if not text:
        return ""
    try:
        return value.date().strftime("%d %b %Y")  # type: ignore[union-attr]
    except Exception:
        pass
    try:
        from datetime import date as date_cls
        return date_cls.fromisoformat(text[:10]).strftime("%d %b %Y")
    except Exception:
        pass
    normalized = text.replace("/", "-")
    for pattern in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            from datetime import datetime as dt_cls
            return dt_cls.strptime(normalized[:10], pattern).strftime("%d %b %Y")
        except Exception:
            continue
    return text


def _safe_filename(value):
    text = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in str(value or "").strip())
    text = text.strip("._-")
    return text or "report"


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, right, border


def _estimate_excel_col_widths(headers, rows, *, min_width=12, max_width=28):
    sample_rows = rows[:50] if rows else []
    widths = []
    for index, header in enumerate(headers):
        values = [str(header or "")]
        for row in sample_rows:
            if index < len(row):
                values.append(str(row[index] or ""))
        widths.append(max(min_width, min(max_width, max(len(value) for value in values) + 2)))
    return widths


def _normalize_export_rows(rows, *, date_columns=None):
    date_columns = set(date_columns or [])
    normalized = []
    for row in rows or []:
        normalized_row = []
        for index, value in enumerate(row, start=1):
            if index in date_columns:
                if value in (None, ""):
                    normalized_row.append("")
                else:
                    normalized_row.append(_excel_safe_value(_format_display_date(value)))
            else:
                normalized_row.append(_excel_safe_value(value))
        normalized.append(normalized_row)
    return normalized


def _write_excel(title, subtitle, headers, rows, *, numeric_columns, date_columns=None, total_row=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font, header_fill, center, left, right, border = _workbook_styles()
    numeric_columns = set(numeric_columns or [])
    date_columns = set(date_columns or [])

    ws.append([_excel_safe_value(title)])
    ws.append([_excel_safe_value(subtitle)])
    ws.append([])
    ws.append([_excel_safe_value(header) for header in headers])
    header_row = ws.max_row

    normalized_rows = _normalize_export_rows(rows, date_columns=date_columns)
    normalized_total_row = None
    if total_row:
        normalized_total_row = _normalize_export_rows([total_row], date_columns=date_columns)[0]

    estimated_widths = _estimate_excel_col_widths(headers, normalized_rows + ([normalized_total_row] if normalized_total_row else []))
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = right if col_idx in numeric_columns else left
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = estimated_widths[col_idx - 1]

    for row in normalized_rows:
        ws.append(row)

    if normalized_total_row:
        ws.append(normalized_total_row)
        total_row_index = ws.max_row
    else:
        total_row_index = None

    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.row == header_row:
                continue
            cell.alignment = right if cell.column in numeric_columns else left
            if cell.column in numeric_columns and cell.value not in (None, ""):
                try:
                    cell.value = float(str(cell.value).replace(",", ""))
                    cell.number_format = "#,##0.00"
                except Exception:
                    pass
            if total_row_index and cell.row == total_row_index:
                cell.font = Font(bold=True, color="1F2937")
                cell.fill = PatternFill("solid", fgColor="EAF2FB")

    if total_row_index:
        ws.freeze_panes = f"A{header_row + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_csv(headers, rows, *, date_columns=None, total_row=None):
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    for row in _normalize_export_rows(rows, date_columns=date_columns):
        writer.writerow(row)
    if total_row:
        writer.writerow(_normalize_export_rows([total_row], date_columns=date_columns)[0])
    return stream.getvalue().encode("utf-8")


def _excel_safe_value(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, default=str, sort_keys=True)
    return str(value)


def _pdf_wrap_text(value):
    text = "" if value is None else str(value)
    return escape(text).replace(", ", ",<br/>").replace("-", "- ")


def _estimate_pdf_col_widths(headers, rows, *, available_width, min_width=42, max_width=150):
    sample_rows = rows[:20] if rows else []
    weights = []
    for index, header in enumerate(headers):
        cell_lengths = [len(str(header or ''))]
        for row in sample_rows:
            if index < len(row):
                cell_lengths.append(len(str(row[index] or '')))
        weights.append(max(cell_lengths))

    total_weight = sum(weights) or 1
    widths = []
    for weight in weights:
        share = available_width * (weight / total_weight)
        widths.append(max(min_width, min(max_width, share)))

    total_width = sum(widths)
    if total_width > available_width:
        scale = available_width / total_width
        widths = [width * scale for width in widths]

    return widths


def _select_pdf_pagesize(headers, rows):
    portrait_width = portrait(A4)[0] - 36
    estimated_portrait = _estimate_pdf_col_widths(headers, rows, available_width=portrait_width)
    if sum(estimated_portrait) <= portrait_width and len(headers) <= 8:
        return portrait(A4)
    return landscape(A4)


def _decorate_pdf(canvas, doc, title):
    canvas.saveState()
    width, height = doc.pagesize
    top_y = height - 28
    canvas.setFillColor(colors.HexColor("#1f4e79"))
    canvas.rect(18, top_y, width - 36, 10, fill=1, stroke=0)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawRightString(width - 18, 12, f"Page {doc.page}")
    canvas.drawString(18, 12, "Finacc ERP")
    canvas.restoreState()


def _write_pdf(
    title,
    subtitle,
    headers,
    rows,
    *,
    numeric_columns=None,
    date_columns=None,
    total_row=None,
    page_size=None,
    col_widths=None,
    header_font_size=8.5,
    body_font_size=7.4,
    leading=8.5,
):
    headers = list(headers or [])
    rows = list(rows or [])
    numeric_columns = set(numeric_columns or [])
    date_columns = set(date_columns or [])
    if not headers:
        if rows:
            headers = [f"Column {index + 1}" for index in range(len(rows[0]))]
        else:
            headers = ["Details"]
            rows = [["No records found"]]

    buffer = BytesIO()
    normalized_rows = _normalize_export_rows(rows, date_columns=date_columns)
    normalized_total_row = None
    if total_row:
        normalized_total_row = _normalize_export_rows([total_row], date_columns=date_columns)[0]
    pdf_rows = normalized_rows + ([normalized_total_row] if normalized_total_row else [])

    page_size = page_size or _select_pdf_pagesize(headers, pdf_rows)
    page_width, _page_height = page_size
    available_width = page_width - 36
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PayablesPdfTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=21,
        textColor=colors.HexColor("#163b66"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "PayablesPdfSubtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#475569"),
    )
    chip_label_style = ParagraphStyle(
        "PayablesPdfChipLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#1f2937"),
    )
    chip_value_style = ParagraphStyle(
        "PayablesPdfChipValue",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#334155"),
    )
    table_header_style = ParagraphStyle(
        "PayablesPdfTableHeader",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=header_font_size,
        leading=max(leading, header_font_size + 1),
        textColor=colors.white,
        alignment=0,
    )
    table_text_style = ParagraphStyle(
        "PayablesPdfTableText",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=body_font_size,
        leading=leading,
        textColor=colors.HexColor("#111827"),
        alignment=0,
    )
    story = [
        Paragraph(title, title_style),
        Spacer(1, 3),
    ]

    subtitle_parts = [part.strip() for part in str(subtitle or '').split('|') if part.strip()]
    if subtitle_parts:
        chip_rows = []
        current_row = []
        for part in subtitle_parts:
            label, value = (part.split(':', 1) + [''])[:2] if ':' in part else ('', part)
            if label:
                chip = Paragraph(f"<b>{label.strip()}:</b> {value.strip()}", chip_value_style)
            else:
                chip = Paragraph(part, chip_value_style)
            current_row.append(chip)
            if len(current_row) == 4:
                chip_rows.append(current_row)
                current_row = []
        if current_row:
            chip_rows.append(current_row)

        for chip_row in chip_rows:
            chip_table = Table([chip_row], colWidths=[available_width / len(chip_row)] * len(chip_row))
            chip_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fbff")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7d4e2")),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d7e2ec")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(chip_table)
            story.append(Spacer(1, 4))
    else:
        story.append(Paragraph("", subtitle_style))
        story.append(Spacer(1, 4))

    col_widths = col_widths or _estimate_pdf_col_widths(headers, pdf_rows, available_width=available_width)
    table_headers = [Paragraph(_pdf_wrap_text(header), table_header_style) for header in headers]
    table_rows = []
    for row in pdf_rows:
        rendered_row = []
        for col_index, value in enumerate(row):
            if (col_index + 1) in numeric_columns:
                rendered_row.append(value)
            else:
                rendered_row.append(Paragraph(_pdf_wrap_text(value), table_text_style))
        table_rows.append(rendered_row)
    table = Table([table_headers] + table_rows, repeatRows=1, colWidths=col_widths)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), header_font_size),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7e2ec")),
        ("FONTSIZE", (0, 1), (-1, -1), body_font_size),
        ("LEADING", (0, 0), (-1, -1), leading),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f2f6fb")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for col_index in range(len(headers)):
        align = "RIGHT" if (col_index + 1) in numeric_columns else "LEFT"
        style_commands.append(("ALIGN", (col_index, 0), (col_index, 0), align))
        style_commands.append(("ALIGN", (col_index, 1), (col_index, len(pdf_rows)), align))
    if normalized_total_row:
        total_row_index = len(pdf_rows)
        style_commands.extend([
            ("BACKGROUND", (0, total_row_index), (-1, total_row_index), colors.HexColor("#eaf2fb")),
            ("FONTNAME", (0, total_row_index), (-1, total_row_index), "Helvetica-Bold"),
        ])
    table.setStyle(TableStyle(style_commands))
    story.append(table)
    doc.build(story, onFirstPage=lambda canvas, doc_: _decorate_pdf(canvas, doc_, title), onLaterPages=lambda canvas, doc_: _decorate_pdf(canvas, doc_, title))
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


class _BaseReceivableAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ReceivableReportScopeSerializer

    def get_scope(self, request):
        serializer = self.serializer_class(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        return serializer.validated_data

    def build_envelope(self, *, report_code, report_name, payload, scope, request, export_base_path):
        response = build_report_envelope(
            report_code=report_code,
            report_name=report_name,
            payload=payload,
            filters=_receivable_scope_filters(scope),
            defaults=RECEIVABLE_DEFAULTS,
        )
        return _attach_receivable_actions(response, request, export_base_path=export_base_path)


class CustomerOutstandingReportAPIView(_BaseReceivableAPIView):
    serializer_class = ReceivableReportScopeSerializer

    def get(self, request):
        scope = self.get_scope(request)
        exception_only = scope.get("exception_only", False)
        data = build_customer_outstanding_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            customer_id=scope.get("customer"),
            customer_group=scope.get("customer_group"),
            region_id=scope.get("region") or scope.get("territory"),
            currency=scope.get("currency"),
            overdue_only=scope.get("overdue_only", False),
            outstanding_gt=scope.get("outstanding_gt"),
            credit_limit_exceeded=scope.get("credit_limit_exceeded", False),
            exception_only=exception_only,
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", RECEIVABLE_DEFAULTS["default_page_size"]),
        )
        return Response(
            self.build_envelope(
                report_code="receivables_exceptions" if exception_only else "customer_outstanding",
                report_name="Receivables Exceptions Report" if exception_only else "Customer Outstanding Report",
                payload=data,
                scope=scope,
                request=request,
                export_base_path="/api/reports/receivables/customer-outstanding/",
            )
        )


class ReceivableAgingReportAPIView(_BaseReceivableAPIView):
    serializer_class = ReceivableAgingScopeSerializer

    def get(self, request):
        scope = self.get_scope(request)
        data = build_receivable_aging_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            as_of_date=scope.get("as_of_date") or scope.get("to_date"),
            customer_id=scope.get("customer"),
            customer_group=scope.get("customer_group"),
            region_id=scope.get("region") or scope.get("territory"),
            currency=scope.get("currency"),
            overdue_only=scope.get("overdue_only", False),
            credit_limit_exceeded=scope.get("credit_limit_exceeded", False),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", RECEIVABLE_DEFAULTS["default_page_size"]),
            view=scope.get("view") or "summary",
        )
        return Response(
            self.build_envelope(
                report_code="receivable_aging",
                report_name="Receivable Aging Report",
                payload=data,
                scope=scope,
                request=request,
                export_base_path="/api/reports/receivables/aging/",
            )
        )


class _BaseReceivableExportAPIView(_BaseReceivableAPIView):
    file_type = None
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response


class _CustomerOutstandingExportMixin(_BaseReceivableExportAPIView):
    serializer_class = ReceivableReportScopeSerializer

    @staticmethod
    def _variant_titles(scope):
        if scope.get("exception_only", False):
            return "Receivables Exceptions", "Receivables Exceptions Report"
        if scope.get("credit_limit_exceeded", False) and not scope.get("overdue_only", False):
            return "Credit Exposure", "Credit Exposure Report"
        if scope.get("overdue_only", False) and not scope.get("credit_limit_exceeded", False):
            return "Overdue Customers", "Overdue Customers Report"
        return "Customer Outstanding", "Customer Outstanding Report"

    def report_data(self, request):
        scope = self.get_scope(request)
        exception_only = scope.get("exception_only", False)
        excel_title, pdf_title = self._variant_titles(scope)
        data = build_customer_outstanding_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            customer_id=scope.get("customer"),
            customer_group=scope.get("customer_group"),
            region_id=scope.get("region") or scope.get("territory"),
            currency=scope.get("currency"),
            overdue_only=scope.get("overdue_only", False),
            outstanding_gt=scope.get("outstanding_gt"),
            credit_limit_exceeded=scope.get("credit_limit_exceeded", False),
            exception_only=exception_only,
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
        )
        headers = [
            "Customer",
            "Code",
            "Opening Bal",
            "Invoice Amt",
            "Receipt Amt",
            "Credit Note",
            "Net Outstanding",
            "Overdue",
            "Unapplied",
            "Credit Limit",
            "Credit Days",
            "Last Inv",
            "Last Pay",
            "Exceptions",
            "Currency",
            "GSTIN",
        ]
        rows = [
            [
                row["customer_name"],
                row["customer_code"],
                row["opening_balance"],
                row["invoice_amount"],
                row["receipt_amount"],
                row["credit_note"],
                row["net_outstanding"],
                row["overdue_amount"],
                row["unapplied_receipt"],
                row["credit_limit"],
                row["credit_days"],
                row["last_invoice_date"],
                row["last_payment_date"],
                ", ".join(row.get("exception_reasons") or []) or "N/A",
                row["currency"],
                row["gstin"],
            ]
            for row in data["rows"]
        ]
        total_row = [
            "Report Total",
            "",
            data["totals"].get("opening_balance"),
            data["totals"].get("invoice_amount"),
            data["totals"].get("receipt_amount"),
            data["totals"].get("credit_note"),
            data["totals"].get("net_outstanding"),
            data["totals"].get("overdue_amount"),
            data["totals"].get("unapplied_receipt"),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date')) or 'Selected date'}"
        )
        return scope, data, headers, rows, subtitle, exception_only, total_row, excel_title, pdf_title


class CustomerOutstandingExcelAPIView(_CustomerOutstandingExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, _exception_only, total_row, title, _pdf_title = self.report_data(request)
        content = _write_excel(
            title,
            subtitle,
            headers,
            rows,
            numeric_columns={3, 4, 5, 6, 7, 8, 9, 10},
            date_columns={12, 13},
            total_row=total_row,
        )
        return self.export_response(
            filename=f"{_safe_filename(title)}_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class CustomerOutstandingCSVAPIView(_CustomerOutstandingExportMixin):
    def get(self, request):
        scope, _data, headers, rows, _subtitle, _exception_only, total_row, title, _pdf_title = self.report_data(request)
        content = _write_csv(headers, rows, date_columns={12, 13}, total_row=total_row)
        return self.export_response(
            filename=f"{_safe_filename(title)}_{_safe_filename(_subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class CustomerOutstandingPDFAPIView(_CustomerOutstandingExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, _exception_only, total_row, _excel_title, title = self.report_data(request)
        content = _write_pdf(
            title,
            subtitle,
            headers,
            rows,
            numeric_columns={3, 4, 5, 6, 7, 8, 9, 10},
            date_columns={12, 13},
            total_row=total_row,
            **CUSTOMER_OUTSTANDING_PDF_LAYOUT,
        )
        return self.export_response(
            filename=f"{_safe_filename(title)}_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class CustomerOutstandingPrintAPIView(CustomerOutstandingPDFAPIView):
    export_mode = "inline"


class OpenItemsReportAPIView(_BaseReceivableAPIView):
    serializer_class = ReceivableReportScopeSerializer

    def get(self, request):
        scope = self.get_scope(request)
        data = build_open_items_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            as_of_date=scope.get("as_of_date"),
            customer_id=scope.get("customer"),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", RECEIVABLE_DEFAULTS["default_page_size"]),
        )
        return Response(
            self.build_envelope(
                report_code="open_items",
                report_name="Open Items Report",
                payload=data,
                scope=scope,
                request=request,
                export_base_path="/api/reports/receivables/open-items/",
            )
        )


class _OpenItemsExportMixin(_BaseReceivableExportAPIView):
    serializer_class = ReceivableReportScopeSerializer

    def report_data(self, request):
        scope = self.get_scope(request)
        data = build_open_items_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            as_of_date=scope.get("as_of_date"),
            customer_id=scope.get("customer"),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        customer_label = "All customers"
        if scope.get("customer"):
            customer_obj = account.objects.filter(id=scope.get("customer")).only("id", "accountname", "legalname").first()
            customer_label = (customer_obj.accountname or customer_obj.legalname or f"Customer {scope.get('customer')}") if customer_obj else f"Customer {scope.get('customer')}"
        headers = [
            "Customer",
            "Customer Code",
            "Bill Date",
            "Due Date",
            "Invoice No",
            "Ref No",
            "Doc Type",
            "Original",
            "Settled",
            "Outstanding",
            "Status",
            "Last Settled",
            "Currency",
            "GSTIN",
        ]
        rows = [
            [
                row["customer_name"],
                row["customer_code"],
                _format_display_date(row.get("bill_date")),
                _format_display_date(row.get("due_date")),
                row["invoice_number"],
                row.get("customer_reference_number") or "-",
                row.get("doc_type_name") or "-",
                row["original_amount"],
                row["settled_amount"],
                row["outstanding_amount"],
                row["status"],
                _format_display_date(row.get("last_settled_at")),
                row["currency"],
                row["gstin"],
            ]
            for row in data["rows"]
        ]
        total_row = [
            "Report Total",
            "",
            "",
            "",
            "",
            "",
            "",
            data["totals"].get("original_amount"),
            data["totals"].get("settled_amount"),
            data["totals"].get("outstanding_amount"),
            "",
            "",
            "",
            "",
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"Customer: {customer_label} | "
            f"As of: {_format_scope_date(scope.get('as_of_date')) or 'Selected date'}"
        )
        return scope, data, headers, rows, subtitle, total_row


class OpenItemsExcelAPIView(_OpenItemsExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, total_row = self.report_data(request)
        content = _write_excel(
            "Open Items Report",
            subtitle,
            headers,
            rows,
            numeric_columns={8, 9, 10},
            date_columns={3, 4, 12},
            total_row=total_row,
        )
        return self.export_response(
            filename=f"OpenItems_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class OpenItemsCSVAPIView(_OpenItemsExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, total_row = self.report_data(request)
        content = _write_csv(headers, rows, date_columns={3, 4, 12}, total_row=total_row)
        return self.export_response(
            filename=f"OpenItems_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class OpenItemsPDFAPIView(_OpenItemsExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, total_row = self.report_data(request)
        content = _write_pdf(
            "Open Items Report",
            subtitle,
            headers,
            rows,
            numeric_columns={8, 9, 10},
            date_columns={3, 4, 12},
            total_row=total_row,
            **OPEN_ITEMS_PDF_LAYOUT,
        )
        return self.export_response(
            filename=f"OpenItems_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class OpenItemsPrintAPIView(OpenItemsPDFAPIView):
    export_mode = "inline"


def _settlement_type_label(value):
    if not value:
        return "All types"
    try:
        return CustomerSettlement.SettlementType(value).label
    except Exception:
        return str(value).replace("_", " ").title()


def _settlement_status_label(value):
    if value in (None, "", "null"):
        return "All statuses"
    try:
        return CustomerSettlement.Status(int(value)).label
    except Exception:
        return str(value).replace("_", " ").title()


class CollectionsHistoryReportAPIView(_BaseReceivableAPIView):
    serializer_class = CollectionsHistoryScopeSerializer

    def get(self, request):
        scope = self.get_scope(request)
        data = build_collections_history_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            customer_id=scope.get("customer"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            settlement_type=scope.get("settlement_type"),
            status=scope.get("status"),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", RECEIVABLE_DEFAULTS["default_page_size"]),
        )
        return Response(
            self.build_envelope(
                report_code="collections_history",
                report_name="Collections History Report",
                payload=data,
                scope=scope,
                request=request,
                export_base_path="/api/reports/receivables/collections-history/",
            )
        )


class _CollectionsHistoryExportMixin(_BaseReceivableExportAPIView):
    serializer_class = CollectionsHistoryScopeSerializer

    def report_data(self, request):
        scope = self.get_scope(request)
        data = build_collections_history_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            customer_id=scope.get("customer"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            settlement_type=scope.get("settlement_type"),
            status=scope.get("status"),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
        )
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        customer_label = "All customers"
        if scope.get("customer"):
            customer_obj = account.objects.filter(id=scope.get("customer")).only("id", "accountname", "legalname").first()
            customer_label = (customer_obj.accountname or customer_obj.legalname or f"Customer {scope.get('customer')}") if customer_obj else f"Customer {scope.get('customer')}"
        headers = [
            "Customer",
            "Code",
            "Settle Dt",
            "Type",
            "Ref No",
            "Ext Voucher",
            "Amount",
            "Status",
            "Lines",
            "Advance Ref",
            "Source Vch",
            "Currency",
            "GSTIN",
            "Remarks",
        ]
        rows = [
            [
                row["customer_name"],
                row["customer_code"],
                _format_display_date(row.get("settlement_date")),
                row.get("settlement_type_name") or row.get("settlement_type") or "-",
                row.get("reference_no") or "-",
                row.get("external_voucher_no") or "-",
                row["total_amount"],
                row.get("status_name") or row.get("status") or "-",
                row.get("line_count") or 0,
                row.get("advance_reference_no") or "-",
                row.get("source_receipt_voucher_code") or row.get("source_receipt_voucher_id") or "-",
                row.get("currency") or "-",
                row.get("gstin") or "-",
                row.get("remarks") or "-",
            ]
            for row in data["rows"]
        ]
        total_row = [
            "Report Total",
            "",
            "",
            "",
            "",
            "",
            data["totals"].get("total_amount"),
            "",
            data["totals"].get("line_count"),
            "",
            "",
            "",
            "",
            "",
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"Customer: {customer_label} | "
            f"From: {_format_scope_date(scope.get('from_date')) or 'Start date'} | "
            f"To: {_format_scope_date(scope.get('to_date')) or 'End date'} | "
            f"Type: {_settlement_type_label(scope.get('settlement_type'))} | "
            f"Status: {_settlement_status_label(scope.get('status'))}"
        )
        return scope, data, headers, rows, subtitle, total_row


class CollectionsHistoryExcelAPIView(_CollectionsHistoryExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, total_row = self.report_data(request)
        content = _write_excel(
            "Collections History Report",
            subtitle,
            headers,
            rows,
            numeric_columns={7, 9},
            date_columns={3},
            total_row=total_row,
        )
        return self.export_response(
            filename=f"CollectionsHistory_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class CollectionsHistoryCSVAPIView(_CollectionsHistoryExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, total_row = self.report_data(request)
        content = _write_csv(headers, rows, date_columns={3}, total_row=total_row)
        return self.export_response(
            filename=f"CollectionsHistory_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class CollectionsHistoryPDFAPIView(_CollectionsHistoryExportMixin):
    def get(self, request):
        scope, _data, headers, rows, subtitle, total_row = self.report_data(request)
        content = _write_pdf(
            "Collections History Report",
            subtitle,
            headers,
            rows,
            numeric_columns={7, 9},
            date_columns={3},
            total_row=total_row,
            **COLLECTIONS_HISTORY_PDF_LAYOUT,
        )
        return self.export_response(
            filename=f"CollectionsHistory_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class CollectionsHistoryPrintAPIView(CollectionsHistoryPDFAPIView):
    export_mode = "inline"


class _ReceivableAgingExportMixin(_BaseReceivableExportAPIView):
    serializer_class = ReceivableAgingScopeSerializer

    def report_data(self, request):
        scope = self.get_scope(request)
        data = build_receivable_aging_report(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            as_of_date=scope.get("as_of_date") or scope.get("to_date"),
            customer_id=scope.get("customer"),
            customer_group=scope.get("customer_group"),
            region_id=scope.get("region") or scope.get("territory"),
            currency=scope.get("currency"),
            overdue_only=scope.get("overdue_only", False),
            credit_limit_exceeded=scope.get("credit_limit_exceeded", False),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
            view=scope.get("view") or "summary",
        )
        customer_label = "All customers"
        if scope.get("customer"):
            customer_obj = account.objects.filter(id=scope.get("customer")).only("id", "accountname", "legalname").first()
            customer_label = (customer_obj.accountname or customer_obj.legalname or f"Customer {scope.get('customer')}") if customer_obj else f"Customer {scope.get('customer')}"
        if (scope.get("view") or "summary") == "invoice":
            headers = [
                "Customer",
                "Code",
                "Invoice No",
                "Inv Date",
                "Due Date",
                "Cr Days",
                "Invoice Amt",
                "Received",
                "Balance",
                "Current",
                "1-30",
                "31-60",
                "61-90",
                "90+",
                "Currency",
            ]
            rows = [
                [
                    row["customer_name"],
                    row["customer_code"],
                    row["invoice_number"],
                    row["invoice_date"],
                    row["due_date"],
                    row["credit_days"],
                    row["invoice_amount"],
                    row["received_amount"],
                    row["balance"],
                    row["current"],
                    row["bucket_1_30"],
                    row["bucket_31_60"],
                    row["bucket_61_90"],
                    row["bucket_90_plus"],
                    row["currency"],
                ]
                for row in data["rows"]
            ]
            title = "Receivable Aging Invoice Report"
            numeric_columns = set(range(6, 15))
            total_row = [
                "Report Total",
                "",
                "",
                "",
                "",
                "",
                f"{sum((float(str(row.get('invoice_amount') or 0)) for row in data['rows'])):.2f}",
                f"{sum((float(str(row.get('received_amount') or 0)) for row in data['rows'])):.2f}",
                data["totals"].get("balance"),
                data["totals"].get("current"),
                data["totals"].get("bucket_1_30"),
                data["totals"].get("bucket_31_60"),
                data["totals"].get("bucket_61_90"),
                data["totals"].get("bucket_90_plus"),
                "",
            ]
        else:
            headers = [
                "Customer",
                "Code",
                "Outstanding",
                "Overdue",
                "Current",
                "1-30",
                "31-60",
                "61-90",
                "90+",
                "Unapplied",
                "Credit Limit",
                "Cr Days",
                "Last Pay",
                "Currency",
            ]
            rows = [
                [
                    row["customer_name"],
                    row["customer_code"],
                    row["outstanding"],
                    row["overdue_amount"],
                    row["current"],
                    row["bucket_1_30"],
                    row["bucket_31_60"],
                    row["bucket_61_90"],
                    row["bucket_90_plus"],
                    row["unapplied_receipt"],
                    row["credit_limit"],
                    row["credit_days"],
                    row["last_payment_date"],
                    row["currency"],
                ]
                for row in data["rows"]
            ]
            title = "Receivable Aging Summary Report"
            numeric_columns = set(range(3, 13))
            total_row = [
                "Report Total",
                "",
                data["totals"].get("outstanding"),
                data["totals"].get("overdue_amount"),
                data["totals"].get("current"),
                data["totals"].get("bucket_1_30"),
                data["totals"].get("bucket_31_60"),
                data["totals"].get("bucket_61_90"),
                data["totals"].get("bucket_90_plus"),
                data["totals"].get("unapplied_receipt"),
                "",
                "",
                "",
                "",
            ]
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date')) or 'Selected date'} | "
            f"Customer: {customer_label} | "
            f"View: {(scope.get('view') or 'summary').title()}"
        )
        return scope, headers, rows, subtitle, title, numeric_columns, total_row


class ReceivableAgingExcelAPIView(_ReceivableAgingExportMixin):
    def get(self, request):
        scope, headers, rows, subtitle, title, numeric_columns, total_row = self.report_data(request)
        content = _write_excel(title, subtitle, headers, rows, numeric_columns=numeric_columns, date_columns={4, 5, 13}, total_row=total_row)
        return self.export_response(
            filename=f"ReceivableAging_Entity{scope['entity']}_{scope.get('view') or 'summary'}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class ReceivableAgingCSVAPIView(_ReceivableAgingExportMixin):
    def get(self, request):
        scope, headers, rows, _subtitle, _title, _numeric_columns, total_row = self.report_data(request)
        content = _write_csv(headers, rows, date_columns={4, 5, 13}, total_row=total_row)
        return self.export_response(
            filename=f"ReceivableAging_Entity{scope['entity']}_{scope.get('view') or 'summary'}.csv",
            content=content,
            content_type="text/csv",
        )


class ReceivableAgingPDFAPIView(_ReceivableAgingExportMixin):
    def get(self, request):
        scope, headers, rows, subtitle, title, numeric_columns, total_row = self.report_data(request)
        pdf_layout = RECEIVABLE_AGING_INVOICE_PDF_LAYOUT if (scope.get("view") or "summary") == "invoice" else RECEIVABLE_AGING_SUMMARY_PDF_LAYOUT
        content = _write_pdf(
            title,
            subtitle,
            headers,
            rows,
            numeric_columns=numeric_columns,
            date_columns={4, 5, 13},
            total_row=total_row,
            **pdf_layout,
        )
        return self.export_response(
            filename=f"ReceivableAging_Entity{scope['entity']}_{scope.get('view') or 'summary'}.pdf",
            content=content,
            content_type="application/pdf",
        )


class ReceivableAgingPrintAPIView(ReceivableAgingPDFAPIView):
    export_mode = "inline"
