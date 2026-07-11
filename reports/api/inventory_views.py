from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import permissions
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView
from xml.sax.saxutils import escape

from core.entitlements import ScopedEntitlementMixin
from reports.schemas.common import build_report_envelope
from reports.schemas.inventory_reports import InventoryReportScopeSerializer
from reports.selectors.financial import resolve_scope_names
from reports.services.inventory.meta import INVENTORY_REPORT_DEFAULTS, build_inventory_report_meta
from reports.services.inventory.control import (
    build_inventory_non_moving_stock,
    build_inventory_reorder_status,
    build_inventory_slow_moving_dead_stock,
)
from reports.services.inventory.location_stock import build_inventory_location_stock
from reports.services.inventory.operational import (
    build_inventory_stock_book_detail,
    build_inventory_stock_book_summary,
    build_inventory_stock_day_book,
    build_inventory_stock_movement,
)
from reports.services.inventory.stock_aging import build_inventory_stock_aging
from reports.services.inventory.stock_ledger import build_inventory_stock_ledger
from reports.services.inventory.stock_summary import build_inventory_stock_summary
from rbac.services import EffectivePermissionService
from subscriptions.services import SubscriptionLimitCodes, SubscriptionService


def _format_scope_date(value):
    if not value:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%b-%Y")
    text = str(value).strip()
    if not text:
        return "-"
    try:
        normalized = text.replace("Z", "+00:00")
        if "T" in normalized or "+" in normalized[10:]:
            return datetime.fromisoformat(normalized).strftime("%d-%b-%Y")
        return date.fromisoformat(normalized).strftime("%d-%b-%Y")
    except ValueError:
        return text


def _safe_filename(value):
    text = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in str(value or "").strip())
    text = text.strip("._-")
    return text or "report"


def _build_total_row(length, *, label="Report Total", values=None):
    row = [""] * length
    if length:
        row[0] = label
    for index, value in (values or {}).items():
        if 0 <= index < length and value not in (None, ""):
            row[index] = value
    return row


def _inventory_querydict(request, *, exclude=None):
    params = request.GET.copy()
    for key in exclude or []:
        params.pop(key, None)
    return params.urlencode()


def _attach_inventory_actions(payload, request, *, export_base_path):
    query = _inventory_querydict(request, exclude=["page", "page_size"])
    payload["actions"]["can_print"] = True
    payload["actions"]["export_urls"] = {
        "excel": f"{export_base_path}excel/?{query}",
        "pdf": f"{export_base_path}pdf/?{query}",
        "csv": f"{export_base_path}csv/?{query}",
        "print": f"{export_base_path}print/?{query}",
    }
    payload["available_exports"] = ["excel", "pdf", "csv", "print"]
    return payload


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, right, border


def _write_excel(title, subtitle, headers, rows, *, numeric_columns, totals_row=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font, header_fill, center, left, right, border = _workbook_styles()

    ws.append([title])
    ws.append([subtitle])
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(headers[col_idx - 1]) + 4, 16)
    for row in rows:
        ws.append(row)
    if totals_row:
        ws.append(totals_row)
        total_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.font = Font(bold=True, color="1F1F1F")
            cell.fill = PatternFill("solid", fgColor="E8F0FE")
            cell.border = border
            cell.alignment = right if cell.column in numeric_columns else left
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.row == header_row:
                continue
            cell.alignment = right if cell.column in numeric_columns else left
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_csv(headers, rows, *, totals_row=None):
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    writer.writerows(rows)
    if totals_row:
        writer.writerow(totals_row)
    return stream.getvalue().encode("utf-8")


def _write_pdf(title, subtitle, headers, rows, *, col_widths=None, numeric_columns=None, totals_row=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        leftMargin=14,
        rightMargin=14,
        topMargin=14,
        bottomMargin=14,
        title=title,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
        Spacer(1, 6),
        Paragraph(subtitle, styles["Normal"]),
        Spacer(1, 10),
    ]
    numeric_columns = {idx - 1 for idx in (numeric_columns or set()) if idx > 0}
    cell_style = ParagraphStyle(
        "InventoryPdfCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=6.5,
        leading=7.5,
        leftIndent=0,
        rightIndent=0,
        spaceBefore=0,
        spaceAfter=0,
        wordWrap="LTR",
        splitLongWords=True,
    )
    total_cell_style = ParagraphStyle(
        "InventoryPdfTotalCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
    )

    def _pdf_cell(value, *, total=False):
        if value in (None, ""):
            return ""
        style = total_cell_style if total else cell_style
        return Paragraph(escape(str(value)).replace("\n", "<br/>"), style)

    if col_widths:
        max_width = doc.width
        total_width = sum(col_widths)
        if total_width > max_width and total_width > 0:
            scale = max_width / total_width
            col_widths = [width * scale for width in col_widths]

    table_rows = [headers]
    for row in rows:
        table_rows.append([
            str(value) if idx in numeric_columns else _pdf_cell(value)
            for idx, value in enumerate(row)
        ])
    if totals_row:
        table_rows.append([
            str(value) if idx in numeric_columns else _pdf_cell(value, total=True)
            for idx, value in enumerate(totals_row)
        ])
    table = Table(table_rows, repeatRows=1, colWidths=col_widths)
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]
    for col_idx in numeric_columns:
        table_style.append(("ALIGN", (col_idx, 1), (col_idx, -1), "RIGHT"))
        table_style.append(("ALIGN", (col_idx, 0), (col_idx, 0), "CENTER"))
    table.setStyle(
        TableStyle(table_style)
    )
    if totals_row:
        total_row_idx = len(table_rows) - 1
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), colors.HexColor("#E8F0FE")),
                    ("FONTNAME", (0, total_row_idx), (-1, total_row_idx), "Helvetica-Bold"),
                    ("LINEABOVE", (0, total_row_idx), (-1, total_row_idx), 0.5, colors.HexColor("#2F5597")),
                ]
            )
        )
    story.append(table)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


class _BaseInventoryReportAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = InventoryReportScopeSerializer
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get_scope(self, request):
        serializer = self.serializer_class(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        scope = serializer.validated_data
        self.enforce_scope(
            request,
            entity_id=scope["entity"],
            entityfinid_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
        )
        return scope

    def get_permission_codes(self, request, scope):
        return EffectivePermissionService.permission_codes_for_user(request.user, scope["entity"])

    def assert_report_permission(self, request, scope, permission_code: str):
        permission_codes = self.get_permission_codes(request, scope)
        if permission_code not in permission_codes:
            raise PermissionDenied(f"Missing permission: {permission_code}")

    def build_filters(self, scope):
        return {
            "entity": scope["entity"],
            "entityfinid": scope.get("entityfinid"),
            "financial_year": scope.get("financial_year") or scope.get("entityfinid"),
            "subentity": scope.get("subentity"),
            "scope_mode": scope.get("scope_mode"),
            "as_on_date": scope.get("as_on_date"),
            "from_date": scope.get("from_date"),
            "to_date": scope.get("to_date"),
            "as_of_date": scope.get("as_of_date"),
            "valuation_method": scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            "product_ids": scope.get("product_ids", []),
            "category_ids": scope.get("category_ids", []),
            "hsn_ids": scope.get("hsn_ids", []),
            "location_ids": scope.get("location_ids", []),
            "include_zero": scope.get("include_zero", INVENTORY_REPORT_DEFAULTS["show_zero_balances_default"]),
            "include_negative": scope.get("include_negative", True),
            "search": scope.get("search"),
            "sort_by": scope.get("sort_by"),
            "sort_order": scope.get("sort_order", "desc"),
            "group_by_location": scope.get("group_by_location", True),
            "bucket_ends": scope.get("bucket_ends", []),
            "non_moving_days": scope.get("non_moving_days"),
            "page": scope.get("page", 1),
            "page_size": scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
            "export": scope.get("export"),
        }

    def decorate_payload(self, payload):
        payload.setdefault("actions", {})
        payload["actions"].update(
            {
                "can_view": True,
                "can_export_excel": True,
                "can_export_pdf": True,
                "can_export_csv": True,
                "can_drilldown": False,
            }
        )
        payload.setdefault("available_drilldowns", [])
        return payload


class InventoryReportsMetaAPIView(ScopedEntitlementMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    subscription_feature_code = SubscriptionLimitCodes.FEATURE_REPORTING
    subscription_access_mode = SubscriptionService.ACCESS_MODE_OPERATIONAL

    def get(self, request):
        entity_id = request.query_params.get("entity")
        if not entity_id:
            return Response({"detail": "entity is required."}, status=400)
        entity_id = int(entity_id)
        self.enforce_scope(request, entity_id=entity_id)
        meta = build_inventory_report_meta(entity_id)
        meta["required_permission_codes"] = [
            "reports.inventory.view",
            "reports.inventory.stock_summary.view",
            "reports.inventory.stock_ledger.view",
            "reports.inventory.stock_aging.view",
            "reports.inventory.location_stock.view",
            "reports.inventory.stock_movement.view",
            "reports.inventory.stock_day_book.view",
            "reports.inventory.stock_book_summary.view",
            "reports.inventory.stock_book_detail.view",
            "reports.inventory.non_moving_stock.view",
            "reports.inventory.reorder_status.view",
        ]
        return Response(meta)


class InventoryStockSummaryAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-summary/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_summary.view")
        data = build_inventory_stock_summary(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_stock_summary",
            report_name="Stock Summary",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_inventory_drilldowns(payload, "inventory_stock_summary"))


class _BaseInventoryStockSummaryExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"
    export_format = "xlsx"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_summary.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_summary(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            search=scope.get("search"),
            sort_by=scope.get("sort_by"),
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "HSN",
            "UOM",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Movement Count",
            "Last Movement Date",
            "Reorder Level",
            "Min Stock",
            "Max Stock",
            "Stock Status",
            "Stock Gap",
        ]
        rows = [
            [
                row.get("product_name"),
                row.get("sku"),
                row.get("category_name"),
                row.get("hsn_code"),
                row.get("uom_name"),
                row.get("closing_qty"),
                row.get("closing_value"),
                row.get("rate"),
                row.get("movement_count"),
                _format_scope_date(row.get("last_movement_date")),
                row.get("reorder_level"),
                row.get("min_stock"),
                row.get("max_stock"),
                row.get("stock_status"),
                row.get("stock_gap"),
            ]
            for row in data["rows"]
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date') or scope.get('as_on_date'))} | "
            f"Valuation: {scope.get('valuation_method') or 'fifo'}"
        )
        totals_row = _build_total_row(len(headers), values={
            5: totals.get("closing_qty"),
            6: totals.get("closing_value"),
            8: data.get("summary", {}).get("movement_count"),
        })
        return scope, headers, rows, subtitle, totals_row


class InventoryStockSummaryExcelAPIView(_BaseInventoryStockSummaryExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Summary", subtitle, headers, rows, numeric_columns={6, 7, 8, 9, 11, 12, 13, 15}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockSummary_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockSummaryCSVAPIView(_BaseInventoryStockSummaryExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockSummary_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockSummaryPDFAPIView(_BaseInventoryStockSummaryExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [
            174,
            70,
            120,
            56,
            58,
            68,
            78,
            62,
            68,
            78,
            70,
            64,
            64,
            72,
            64,
        ]
        content = _write_pdf(
            "Inventory Stock Summary",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={6, 7, 8, 9, 11, 12, 13, 15},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockSummary_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockSummaryPrintAPIView(InventoryStockSummaryPDFAPIView):
    export_mode = "inline"


class InventoryStockLedgerAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-ledger/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_ledger.view")
        data = build_inventory_stock_ledger(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", True),
            include_negative=scope.get("include_negative", True),
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_stock_ledger",
            report_name="Stock Ledger",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_inventory_drilldowns(payload, "inventory_stock_ledger"))


class _BaseInventoryStockLedgerExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_ledger.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_ledger(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", True),
            include_negative=scope.get("include_negative", True),
            sort_order=scope.get("sort_order", "asc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Posting Date",
            "Voucher No",
            "Product",
            "SKU",
            "Category",
            "Location",
            "Move Type",
            "Qty In",
            "Qty Out",
            "Unit Cost",
            "Line Value",
            "Opening Qty",
            "Opening Value",
            "Running Qty",
            "Running Value",
        ]
        rows = [
            [
                _format_scope_date(row.get("posting_date")),
                row.get("voucher_no"),
                row.get("product_name"),
                row.get("sku"),
                row.get("category_name"),
                row.get("location_name"),
                row.get("move_type"),
                row.get("qty_in"),
                row.get("qty_out"),
                row.get("unit_cost"),
                row.get("line_value"),
                row.get("opening_qty"),
                row.get("opening_value"),
                row.get("running_qty"),
                row.get("running_value"),
            ]
            for row in data["rows"]
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"From: {_format_scope_date(scope.get('from_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
            f"To: {_format_scope_date(scope.get('to_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
            f"Valuation: {scope.get('valuation_method') or 'fifo'}"
        )
        totals_row = _build_total_row(len(headers), values={
            7: totals.get("inward_qty"),
            8: totals.get("outward_qty"),
            11: totals.get("opening_qty"),
            12: totals.get("opening_value"),
            13: totals.get("closing_qty"),
            14: totals.get("closing_value"),
        })
        return scope, headers, rows, subtitle, totals_row


class InventoryStockLedgerExcelAPIView(_BaseInventoryStockLedgerExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Ledger", subtitle, headers, rows, numeric_columns={8, 9, 10, 11, 12, 13, 14, 15}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockLedger_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockLedgerCSVAPIView(_BaseInventoryStockLedgerExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockLedger_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockLedgerPDFAPIView(_BaseInventoryStockLedgerExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [78, 84, 150, 58, 88, 76, 52, 56, 56, 60, 66, 62, 68, 62, 68]
        content = _write_pdf(
            "Inventory Stock Ledger",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={8, 9, 10, 11, 12, 13, 14, 15},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockLedger_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockLedgerPrintAPIView(InventoryStockLedgerPDFAPIView):
    export_mode = "inline"


class InventoryStockAgingAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-aging/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_aging.view")
        data = build_inventory_stock_aging(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            bucket_ends=scope.get("bucket_ends") or None,
            group_by_location=scope.get("group_by_location", True),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "age_days",
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_stock_aging",
            report_name="Stock Aging",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_inventory_drilldowns(payload, "inventory_stock_aging"))


class _BaseInventoryStockAgingExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_aging.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_aging(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            bucket_ends=scope.get("bucket_ends") or None,
            group_by_location=scope.get("group_by_location", True),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "age_days",
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "Location",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Last Movement Date",
            "Age Days",
            "Age Bucket",
            "Status",
        ]
        rows = [
            [
                row.get("product_name"),
                row.get("sku"),
                row.get("category_name"),
                row.get("location_name"),
                row.get("closing_qty"),
                row.get("closing_value"),
                row.get("rate"),
                _format_scope_date(row.get("last_movement_date")),
                row.get("age_days"),
                row.get("age_bucket"),
                row.get("stock_status"),
            ]
            for row in data["rows"]
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date') or scope.get('as_on_date'))} | "
            f"Valuation: {scope.get('valuation_method') or 'fifo'} | "
            f"Buckets: {', '.join(str(end) for end in (scope.get('bucket_ends') or [30, 60, 90, 120, 150]))}"
        )
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("closing_qty"),
            5: totals.get("closing_value"),
        })
        return scope, headers, rows, subtitle, totals_row


class InventoryStockAgingExcelAPIView(_BaseInventoryStockAgingExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Aging", subtitle, headers, rows, numeric_columns={5, 6, 7, 9}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockAging_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockAgingCSVAPIView(_BaseInventoryStockAgingExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockAging_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockAgingPDFAPIView(_BaseInventoryStockAgingExportAPIView):
    def get(self, request):
        scope, headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [140, 70, 110, 90, 60, 74, 60, 84, 56, 84, 64]
        content = _write_pdf(
            "Inventory Stock Aging",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 9},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockAging_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockAgingPrintAPIView(InventoryStockAgingPDFAPIView):
    export_mode = "inline"


class InventoryLocationStockAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/location-stock/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.location_stock.view")
        data = build_inventory_location_stock(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            include_zero=scope.get("include_zero", True),
            include_negative=scope.get("include_negative", True),
            search=scope.get("search"),
            sort_by=scope.get("sort_by") or "value",
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_location_stock",
            report_name="Location Stock",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_location_drilldowns(payload, "inventory_location_stock"))


class _BaseInventoryLocationStockExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.location_stock.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_location_stock(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", "fifo"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            include_zero=scope.get("include_zero", True),
            include_negative=scope.get("include_negative", True),
            search=scope.get("search"),
            sort_by=scope.get("sort_by") or "value",
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Location",
            "Code",
            "City",
            "State",
            "Products",
            "Movements",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Low Stock",
            "Negative Stock",
            "Zero Stock",
            "First Movement",
            "Last Movement",
            "Status",
        ]
        rows = [_location_stock_row_export(row) for row in data["rows"]]
        subtitle = _location_stock_subtitle(scope, scope_names)
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("product_count"),
            5: totals.get("movement_count"),
            6: totals.get("closing_qty"),
            7: totals.get("closing_value"),
            9: data.get("summary", {}).get("low_stock_count"),
            10: data.get("summary", {}).get("negative_stock_count"),
            11: data.get("summary", {}).get("zero_stock_count"),
        })
        return headers, rows, subtitle, totals_row


class InventoryLocationStockExcelAPIView(_BaseInventoryLocationStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Location Stock", subtitle, headers, rows, numeric_columns={5, 6, 7, 8, 9, 10, 11, 12}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryLocationStock_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryLocationStockCSVAPIView(_BaseInventoryLocationStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryLocationStock_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryLocationStockPDFAPIView(_BaseInventoryLocationStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [90, 52, 54, 54, 48, 52, 58, 68, 52, 52, 56, 52, 62, 62, 52]
        content = _write_pdf(
            "Inventory Location Stock",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 8, 9, 10, 11, 12},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryLocationStock_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryLocationStockPrintAPIView(InventoryLocationStockPDFAPIView):
    export_mode = "inline"


class InventoryNonMovingStockAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/non-moving-stock/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.non_moving_stock.view")
        data = build_inventory_non_moving_stock(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            non_moving_days=scope.get("non_moving_days"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "last_movement_date",
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_non_moving_stock",
            report_name="Non-Moving Stock",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_control_drilldowns(payload, "inventory_non_moving_stock"))


class _BaseInventoryNonMovingStockExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.non_moving_stock.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_non_moving_stock(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            non_moving_days=scope.get("non_moving_days"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "last_movement_date",
            sort_order=scope.get("sort_order", "asc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "Location",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Last Movement Date",
            "Age Days",
            "Non Moving Days",
            "Status",
        ]
        rows = [
            [
                row.get("product_name"),
                row.get("sku"),
                row.get("category_name"),
                row.get("location_name"),
                row.get("closing_qty"),
                row.get("closing_value"),
                row.get("rate"),
                _format_scope_date(row.get("last_movement_date")),
                row.get("age_days"),
                row.get("non_moving_days"),
                row.get("stock_status"),
            ]
            for row in data["rows"]
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date') or scope.get('as_on_date'))} | "
            f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']} | "
            f"Days: {scope.get('non_moving_days') or 90}"
        )
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("closing_qty"),
            5: totals.get("closing_value"),
        })
        return headers, rows, subtitle, totals_row


class InventoryNonMovingStockExcelAPIView(_BaseInventoryNonMovingStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Non-Moving Stock", subtitle, headers, rows, numeric_columns={5, 6, 7, 9, 10}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryNonMovingStock_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryNonMovingStockCSVAPIView(_BaseInventoryNonMovingStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryNonMovingStock_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryNonMovingStockPDFAPIView(_BaseInventoryNonMovingStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [138, 72, 110, 88, 60, 72, 60, 82, 54, 66, 64]
        content = _write_pdf(
            "Inventory Non-Moving Stock",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 9, 10},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryNonMovingStock_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryNonMovingStockPrintAPIView(InventoryNonMovingStockPDFAPIView):
    export_mode = "inline"


class InventoryReorderStatusAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/reorder-status/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.reorder_status.view")
        data = build_inventory_reorder_status(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", True),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "reorder_gap",
            sort_order=scope.get("sort_order", "asc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_reorder_status",
            report_name="Reorder Status",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_control_drilldowns(payload, "inventory_reorder_status"))


class _BaseInventoryReorderStatusExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.reorder_status.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_reorder_status(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", True),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "reorder_gap",
            sort_order=scope.get("sort_order", "asc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "Location",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Reorder Level",
            "Min Stock",
            "Max Stock",
            "Reorder Qty",
            "Gap",
            "Last Movement Date",
            "Status",
        ]
        rows = [
            [
                row.get("product_name"),
                row.get("sku"),
                row.get("category_name"),
                row.get("location_name"),
                row.get("closing_qty"),
                row.get("closing_value"),
                row.get("rate"),
                row.get("reorder_level"),
                row.get("min_stock"),
                row.get("max_stock"),
                row.get("reorder_qty"),
                row.get("reorder_gap"),
                _format_scope_date(row.get("last_movement_date")),
                row.get("stock_status"),
            ]
            for row in data["rows"]
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date') or scope.get('as_on_date'))} | "
            f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']}"
        )
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("closing_qty"),
            5: totals.get("closing_value"),
        })
        return headers, rows, subtitle, totals_row


class InventoryReorderStatusExcelAPIView(_BaseInventoryReorderStatusExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Reorder Status", subtitle, headers, rows, numeric_columns={5, 6, 7, 8, 9, 10, 11, 12}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryReorderStatus_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryReorderStatusCSVAPIView(_BaseInventoryReorderStatusExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryReorderStatus_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryReorderStatusPDFAPIView(_BaseInventoryReorderStatusExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [126, 66, 100, 84, 56, 70, 56, 64, 58, 58, 62, 54, 82, 60]
        content = _write_pdf(
            "Inventory Reorder Status",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 8, 9, 10, 11, 12},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryReorderStatus_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryReorderStatusPrintAPIView(InventoryReorderStatusPDFAPIView):
    export_mode = "inline"


class InventorySlowMovingDeadStockAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/slow-moving-dead-stock/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.slow_moving_dead_stock.view")
        data = build_inventory_slow_moving_dead_stock(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            non_moving_days=scope.get("non_moving_days"),
            dead_stock_days=scope.get("dead_stock_days"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "age_days",
            sort_order=scope.get("sort_order", "desc"),
            page=scope.get("page", 1),
            page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
        )
        response = build_report_envelope(
            report_code="inventory_slow_moving_dead_stock",
            report_name="Slow Moving vs Dead Stock",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_control_drilldowns(payload, "inventory_slow_moving_dead_stock"))


class _BaseInventorySlowMovingDeadStockExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.slow_moving_dead_stock.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_slow_moving_dead_stock(
            entity_id=scope["entity"],
            entityfin_id=scope.get("entityfinid"),
            subentity_id=scope.get("subentity"),
            from_date=scope.get("from_date"),
            to_date=scope.get("to_date"),
            as_of_date=scope.get("as_of_date"),
            valuation_method=scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
            non_moving_days=scope.get("non_moving_days"),
            dead_stock_days=scope.get("dead_stock_days"),
            product_ids=scope.get("product_ids") or None,
            category_ids=scope.get("category_ids") or None,
            hsn_ids=scope.get("hsn_ids") or None,
            location_ids=scope.get("location_ids") or None,
            search=scope.get("search"),
            include_zero=scope.get("include_zero", False),
            include_negative=scope.get("include_negative", True),
            sort_by=scope.get("sort_by") or "age_days",
            sort_order=scope.get("sort_order", "desc"),
            page=1,
            page_size=100000,
            paginate=False,
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "Location",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Last Movement Date",
            "Age Days",
            "Slow Days",
            "Dead Days",
            "Class",
            "Status",
        ]
        rows = [
            [
                row.get("product_name"),
                row.get("sku"),
                row.get("category_name"),
                row.get("location_name"),
                row.get("closing_qty"),
                row.get("closing_value"),
                row.get("rate"),
                _format_scope_date(row.get("last_movement_date")),
                row.get("age_days"),
                row.get("non_moving_days"),
                row.get("dead_stock_days"),
                row.get("movement_class_label"),
                row.get("stock_status"),
            ]
            for row in data["rows"]
        ]
        subtitle = (
            f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
            f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
            f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
            f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date') or scope.get('as_on_date'))} | "
            f"Slow Days: {scope.get('non_moving_days') or 90} | "
            f"Dead Days: {scope.get('dead_stock_days') or 180} | "
            f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']}"
        )
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("closing_qty"),
            5: totals.get("closing_value"),
        })
        return headers, rows, subtitle, totals_row


class InventorySlowMovingDeadStockExcelAPIView(_BaseInventorySlowMovingDeadStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Slow Moving vs Dead Stock", subtitle, headers, rows, numeric_columns={5, 6, 7, 9, 10, 11}, totals_row=totals_row)
        return self.export_response(
            filename=f"SlowMovingDeadStock_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventorySlowMovingDeadStockCSVAPIView(_BaseInventorySlowMovingDeadStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"SlowMovingDeadStock_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventorySlowMovingDeadStockPDFAPIView(_BaseInventorySlowMovingDeadStockExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [140, 70, 110, 90, 60, 74, 60, 84, 56, 56, 56, 74, 64]
        content = _write_pdf(
            "Slow Moving vs Dead Stock",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 9, 10, 11},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"SlowMovingDeadStock_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventorySlowMovingDeadStockPrintAPIView(InventorySlowMovingDeadStockPDFAPIView):
    export_mode = "inline"


def _inventory_operational_base_kwargs(scope, *, paginate: bool, page: int, page_size: int, group_by_location: bool, sort_by: str, sort_order: str):
    return {
        "entity_id": scope["entity"],
        "entityfin_id": scope.get("entityfinid"),
        "subentity_id": scope.get("subentity"),
        "from_date": scope.get("from_date"),
        "to_date": scope.get("to_date"),
        "as_of_date": scope.get("as_of_date"),
        "valuation_method": scope.get("valuation_method", INVENTORY_REPORT_DEFAULTS["default_valuation_method"]),
        "product_ids": scope.get("product_ids") or None,
        "category_ids": scope.get("category_ids") or None,
        "hsn_ids": scope.get("hsn_ids") or None,
        "location_ids": scope.get("location_ids") or None,
        "search": scope.get("search"),
        "include_zero": scope.get("include_zero", True),
        "include_negative": scope.get("include_negative", True),
        "sort_by": sort_by,
        "sort_order": sort_order,
        "page": page,
        "page_size": page_size,
        "paginate": paginate,
        "group_by_location": group_by_location,
    }


def _movement_row_export(row: dict) -> list:
    return [
        row.get("product_name"),
        row.get("sku"),
        row.get("category_name"),
        row.get("location_name"),
        row.get("opening_qty"),
        row.get("opening_value"),
        row.get("inward_qty"),
        row.get("inward_value"),
        row.get("outward_qty"),
        row.get("outward_value"),
        row.get("net_qty"),
        row.get("net_value"),
        row.get("closing_qty"),
        row.get("closing_value"),
        row.get("rate"),
        row.get("movement_count"),
        _format_scope_date(row.get("first_movement_date")),
        _format_scope_date(row.get("last_movement_date")),
        row.get("stock_status"),
    ]


def _daybook_row_export(row: dict) -> list:
    return [
        _format_scope_date(row.get("posting_date")),
        row.get("opening_qty"),
        row.get("opening_value"),
        row.get("inward_qty"),
        row.get("inward_value"),
        row.get("outward_qty"),
        row.get("outward_value"),
        row.get("closing_qty"),
        row.get("closing_value"),
        row.get("movement_count"),
        row.get("product_count"),
    ]


def _book_detail_row_export(row: dict) -> list:
    return [
        _format_scope_date(row.get("posting_date")),
        row.get("voucher_no"),
        row.get("product_name"),
        row.get("sku"),
        row.get("category_name"),
        row.get("location_name"),
        row.get("move_type"),
        row.get("source_location_name"),
        row.get("destination_location_name"),
        row.get("movement_nature"),
        row.get("movement_reason"),
        row.get("qty_in"),
        row.get("qty_out"),
        row.get("unit_cost"),
        row.get("line_value"),
        row.get("opening_qty"),
        row.get("opening_value"),
        row.get("running_qty"),
        row.get("running_value"),
    ]


def _location_stock_row_export(row: dict) -> list:
    return [
        row.get("location_name"),
        row.get("location_code"),
        row.get("city"),
        row.get("state"),
        row.get("product_count"),
        row.get("movement_count"),
        row.get("closing_qty"),
        row.get("closing_value"),
        row.get("rate"),
        row.get("low_stock_count"),
        row.get("negative_stock_count"),
        row.get("zero_stock_count"),
        _format_scope_date(row.get("first_movement_date")),
        _format_scope_date(row.get("last_movement_date")),
        row.get("stock_status"),
    ]


def _movement_subtitle(scope, scope_names):
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
        f"From: {_format_scope_date(scope.get('from_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
        f"To: {_format_scope_date(scope.get('to_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
        f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']}"
    )


def _daybook_subtitle(scope, scope_names):
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
        f"From: {_format_scope_date(scope.get('from_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
        f"To: {_format_scope_date(scope.get('to_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
        f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']}"
    )


def _book_detail_subtitle(scope, scope_names):
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
        f"From: {_format_scope_date(scope.get('from_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
        f"To: {_format_scope_date(scope.get('to_date') or scope.get('as_of_date') or scope.get('as_on_date'))} | "
        f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']}"
    )


def _location_stock_subtitle(scope, scope_names):
    return (
        f"Entity: {scope_names['entity_name'] or 'Selected entity'} | "
        f"FY: {scope_names['entityfin_name'] or 'Current FY'} | "
        f"Subentity: {scope_names['subentity_name'] or 'All subentities'} | "
        f"As of: {_format_scope_date(scope.get('as_of_date') or scope.get('to_date') or scope.get('as_on_date'))} | "
        f"Valuation: {scope.get('valuation_method') or INVENTORY_REPORT_DEFAULTS['default_valuation_method']}"
    )


_OPERATIONAL_DRILLDOWNS = {
    "inventory_stock_movement": [
        {"code": "inventory_stock_book_detail", "label": "Book Detail"},
        {"code": "inventory_stock_book_summary", "label": "Book Summary"},
    ],
    "inventory_stock_day_book": [
        {"code": "inventory_stock_movement", "label": "Movement"},
    ],
    "inventory_stock_book_summary": [
        {"code": "inventory_stock_book_detail", "label": "Book Detail"},
        {"code": "inventory_stock_movement", "label": "Movement"},
    ],
    "inventory_stock_book_detail": [
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
        {"code": "inventory_stock_summary", "label": "Stock Summary"},
    ],
}


def _attach_operational_drilldowns(payload, report_code: str):
    payload = payload.copy()
    payload.setdefault("actions", {})
    payload["actions"]["can_drilldown"] = True
    payload["available_drilldowns"] = _OPERATIONAL_DRILLDOWNS.get(report_code, [])
    return payload


_INVENTORY_DRILLDOWNS = {
    "inventory_stock_summary": [
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
        {"code": "inventory_stock_aging", "label": "Stock Aging"},
    ],
    "inventory_stock_ledger": [
        {"code": "inventory_stock_movement", "label": "Stock Movement"},
        {"code": "inventory_stock_book_detail", "label": "Book Detail"},
    ],
    "inventory_stock_aging": [
        {"code": "inventory_stock_summary", "label": "Stock Summary"},
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
    ],
    "inventory_location_stock": [
        {"code": "inventory_stock_summary", "label": "Stock Summary"},
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
    ],
}


def _attach_inventory_drilldowns(payload, report_code: str):
    payload = payload.copy()
    payload.setdefault("actions", {})
    payload["actions"]["can_drilldown"] = True
    payload["available_drilldowns"] = _INVENTORY_DRILLDOWNS.get(report_code, [])
    return payload


_CONTROL_DRILLDOWNS = {
    "inventory_non_moving_stock": [
        {"code": "inventory_stock_summary", "label": "Stock Summary"},
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
    ],
    "inventory_slow_moving_dead_stock": [
        {"code": "inventory_stock_summary", "label": "Stock Summary"},
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
    ],
    "inventory_reorder_status": [
        {"code": "inventory_stock_summary", "label": "Stock Summary"},
        {"code": "inventory_stock_ledger", "label": "Stock Ledger"},
    ],
}


def _attach_control_drilldowns(payload, report_code: str):
    payload = payload.copy()
    payload.setdefault("actions", {})
    payload["actions"]["can_drilldown"] = True
    payload["available_drilldowns"] = _CONTROL_DRILLDOWNS.get(report_code, [])
    return payload


def _attach_location_drilldowns(payload, report_code: str):
    payload = payload.copy()
    payload.setdefault("actions", {})
    payload["actions"]["can_drilldown"] = True
    payload["available_drilldowns"] = _INVENTORY_DRILLDOWNS.get(report_code, [])
    return payload


class InventoryStockMovementAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-movement/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_movement.view")
        data = build_inventory_stock_movement(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=True,
                page=scope.get("page", 1),
                page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
                group_by_location=scope.get("group_by_location", True),
                sort_by=scope.get("sort_by") or "value",
                sort_order=scope.get("sort_order", "desc"),
            )
        )
        response = build_report_envelope(
            report_code="inventory_stock_movement",
            report_name="Stock Movement",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_operational_drilldowns(payload, "inventory_stock_movement"))


class _BaseInventoryStockMovementExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_movement.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_movement(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=False,
                page=1,
                page_size=100000,
                group_by_location=scope.get("group_by_location", True),
                sort_by=scope.get("sort_by") or "value",
                sort_order=scope.get("sort_order", "desc"),
            )
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "Location",
            "Opening Qty",
            "Opening Value",
            "Inward Qty",
            "Inward Value",
            "Outward Qty",
            "Outward Value",
            "Net Qty",
            "Net Value",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Movement Count",
            "First Movement",
            "Last Movement",
            "Status",
        ]
        rows = [_movement_row_export(row) for row in data["rows"]]
        subtitle = _movement_subtitle(scope, scope_names)
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("opening_qty"),
            5: totals.get("opening_value"),
            6: totals.get("inward_qty"),
            7: totals.get("inward_value"),
            8: totals.get("outward_qty"),
            9: totals.get("outward_value"),
            10: totals.get("net_qty"),
            11: totals.get("net_value"),
            12: totals.get("closing_qty"),
            13: totals.get("closing_value"),
            15: data.get("summary", {}).get("movement_count"),
        })
        return headers, rows, subtitle, totals_row


class InventoryStockMovementExcelAPIView(_BaseInventoryStockMovementExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Movement", subtitle, headers, rows, numeric_columns={5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockMovement_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockMovementCSVAPIView(_BaseInventoryStockMovementExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockMovement_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockMovementPDFAPIView(_BaseInventoryStockMovementExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [148, 58, 86, 76, 52, 64, 52, 62, 52, 62, 52, 62, 52, 64, 46, 54, 64, 64, 58]
        content = _write_pdf(
            "Inventory Stock Movement",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockMovement_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockMovementPrintAPIView(InventoryStockMovementPDFAPIView):
    export_mode = "inline"


class InventoryStockDayBookAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-day-book/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_day_book.view")
        data = build_inventory_stock_day_book(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=True,
                page=scope.get("page", 1),
                page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
                group_by_location=False,
                sort_by=scope.get("sort_by") or "posting_date",
                sort_order=scope.get("sort_order", "asc"),
            )
        )
        response = build_report_envelope(
            report_code="inventory_stock_day_book",
            report_name="Stock Day Book",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_operational_drilldowns(payload, "inventory_stock_day_book"))


class _BaseInventoryStockDayBookExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_day_book.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_day_book(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=False,
                page=1,
                page_size=100000,
                group_by_location=False,
                sort_by=scope.get("sort_by") or "posting_date",
                sort_order=scope.get("sort_order", "asc"),
            )
        )
        totals = data.get("totals", {})
        headers = [
            "Date",
            "Opening Qty",
            "Opening Value",
            "Inward Qty",
            "Inward Value",
            "Outward Qty",
            "Outward Value",
            "Closing Qty",
            "Closing Value",
            "Movements",
            "Products",
        ]
        rows = [_daybook_row_export(row) for row in data["rows"]]
        subtitle = _daybook_subtitle(scope, scope_names)
        totals_row = _build_total_row(len(headers), values={
            1: totals.get("opening_qty"),
            2: totals.get("opening_value"),
            3: totals.get("inward_qty"),
            4: totals.get("inward_value"),
            5: totals.get("outward_qty"),
            6: totals.get("outward_value"),
            7: totals.get("closing_qty"),
            8: totals.get("closing_value"),
            9: data.get("summary", {}).get("movement_count"),
        })
        return headers, rows, subtitle, totals_row


class InventoryStockDayBookExcelAPIView(_BaseInventoryStockDayBookExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Day Book", subtitle, headers, rows, numeric_columns={2, 3, 4, 5, 6, 7, 8, 9, 10, 11}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockDayBook_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockDayBookCSVAPIView(_BaseInventoryStockDayBookExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockDayBook_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockDayBookPDFAPIView(_BaseInventoryStockDayBookExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [78, 72, 72, 72, 72, 72, 72, 72, 72, 58, 58]
        content = _write_pdf(
            "Inventory Stock Day Book",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={2, 3, 4, 5, 6, 7, 8, 9, 10, 11},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockDayBook_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockDayBookPrintAPIView(InventoryStockDayBookPDFAPIView):
    export_mode = "inline"


class InventoryStockBookSummaryAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-book-summary/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_book_summary.view")
        data = build_inventory_stock_book_summary(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=True,
                page=scope.get("page", 1),
                page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
                group_by_location=False,
                sort_by=scope.get("sort_by") or "value",
                sort_order=scope.get("sort_order", "desc"),
            )
        )
        response = build_report_envelope(
            report_code="inventory_stock_book_summary",
            report_name="Stock Book Summary",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_operational_drilldowns(payload, "inventory_stock_book_summary"))


class _BaseInventoryStockBookSummaryExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_book_summary.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_book_summary(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=False,
                page=1,
                page_size=100000,
                group_by_location=False,
                sort_by=scope.get("sort_by") or "value",
                sort_order=scope.get("sort_order", "desc"),
            )
        )
        totals = data.get("totals", {})
        headers = [
            "Product",
            "SKU",
            "Category",
            "Location",
            "Opening Qty",
            "Opening Value",
            "Inward Qty",
            "Inward Value",
            "Outward Qty",
            "Outward Value",
            "Net Qty",
            "Net Value",
            "Closing Qty",
            "Closing Value",
            "Rate",
            "Movement Count",
            "First Movement",
            "Last Movement",
            "Status",
        ]
        rows = [_movement_row_export(row) for row in data["rows"]]
        subtitle = _movement_subtitle(scope, scope_names)
        totals_row = _build_total_row(len(headers), values={
            4: totals.get("opening_qty"),
            5: totals.get("opening_value"),
            6: totals.get("inward_qty"),
            7: totals.get("inward_value"),
            8: totals.get("outward_qty"),
            9: totals.get("outward_value"),
            10: totals.get("net_qty"),
            11: totals.get("net_value"),
            12: totals.get("closing_qty"),
            13: totals.get("closing_value"),
            15: data.get("summary", {}).get("movement_count"),
        })
        return headers, rows, subtitle, totals_row


class InventoryStockBookSummaryExcelAPIView(_BaseInventoryStockBookSummaryExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Book Summary", subtitle, headers, rows, numeric_columns={5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockBookSummary_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockBookSummaryCSVAPIView(_BaseInventoryStockBookSummaryExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockBookSummary_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockBookSummaryPDFAPIView(_BaseInventoryStockBookSummaryExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [148, 58, 86, 76, 52, 64, 52, 62, 52, 62, 52, 62, 52, 64, 46, 54, 64, 64, 58]
        content = _write_pdf(
            "Inventory Stock Book Summary",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockBookSummary_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockBookSummaryPrintAPIView(InventoryStockBookSummaryPDFAPIView):
    export_mode = "inline"


class InventoryStockBookDetailAPIView(_BaseInventoryReportAPIView):
    export_base_path = "/api/reports/inventory/stock-book-detail/"

    def get(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_book_detail.view")
        data = build_inventory_stock_book_detail(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=True,
                page=scope.get("page", 1),
                page_size=scope.get("page_size", INVENTORY_REPORT_DEFAULTS["default_page_size"]),
                group_by_location=scope.get("group_by_location", True),
                sort_by=scope.get("sort_by") or "posting_date",
                sort_order=scope.get("sort_order", "asc"),
            )
        )
        response = build_report_envelope(
            report_code="inventory_stock_book_detail",
            report_name="Stock Book Detail",
            payload=data,
            filters=self.build_filters(scope),
            defaults=INVENTORY_REPORT_DEFAULTS,
        )
        payload = self.decorate_payload(_attach_inventory_actions(response, request, export_base_path=self.export_base_path))
        return Response(_attach_operational_drilldowns(payload, "inventory_stock_book_detail"))


class _BaseInventoryStockBookDetailExportAPIView(_BaseInventoryReportAPIView):
    export_mode = "attachment"

    def export_response(self, *, filename, content, content_type):
        response = HttpResponse(content=content, content_type=content_type)
        disposition = "inline" if self.export_mode == "inline" else "attachment"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response

    def report_data(self, request):
        scope = self.get_scope(request)
        self.assert_report_permission(request, scope, "reports.inventory.stock_book_detail.view")
        scope_names = resolve_scope_names(scope["entity"], scope.get("entityfinid"), scope.get("subentity"))
        data = build_inventory_stock_book_detail(
            **_inventory_operational_base_kwargs(
                scope,
                paginate=False,
                page=1,
                page_size=100000,
                group_by_location=scope.get("group_by_location", True),
                sort_by=scope.get("sort_by") or "posting_date",
                sort_order=scope.get("sort_order", "asc"),
            )
        )
        totals = data.get("totals", {})
        headers = [
            "Posting Date",
            "Voucher No",
            "Product",
            "SKU",
            "Category",
            "Location",
            "Move Type",
            "Source Location",
            "Destination Location",
            "Nature",
            "Reason",
            "Qty In",
            "Qty Out",
            "Unit Cost",
            "Line Value",
            "Opening Qty",
            "Opening Value",
            "Running Qty",
            "Running Value",
        ]
        rows = [_book_detail_row_export(row) for row in data["rows"]]
        subtitle = _book_detail_subtitle(scope, scope_names)
        totals_row = _build_total_row(len(headers), values={
            11: totals.get("inward_qty"),
            12: totals.get("outward_qty"),
            15: totals.get("opening_qty"),
            16: totals.get("opening_value"),
            17: totals.get("closing_qty"),
            18: totals.get("closing_value"),
        })
        return headers, rows, subtitle, totals_row


class InventoryStockBookDetailExcelAPIView(_BaseInventoryStockBookDetailExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_excel("Inventory Stock Book Detail", subtitle, headers, rows, numeric_columns={12, 13, 14, 15, 16, 17, 18, 19}, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockBookDetail_{_safe_filename(subtitle)}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class InventoryStockBookDetailCSVAPIView(_BaseInventoryStockBookDetailExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        content = _write_csv(headers, rows, totals_row=totals_row)
        return self.export_response(
            filename=f"InventoryStockBookDetail_{_safe_filename(subtitle)}.csv",
            content=content,
            content_type="text/csv",
        )


class InventoryStockBookDetailPDFAPIView(_BaseInventoryStockBookDetailExportAPIView):
    def get(self, request):
        headers, rows, subtitle, totals_row = self.report_data(request)
        col_widths = [74, 80, 150, 58, 88, 82, 54, 74, 74, 56, 80, 52, 52, 60, 68, 60, 68, 60, 68]
        content = _write_pdf(
            "Inventory Stock Book Detail",
            subtitle,
            headers,
            rows,
            col_widths=col_widths,
            numeric_columns={12, 13, 14, 15, 16, 17, 18, 19},
            totals_row=totals_row,
        )
        return self.export_response(
            filename=f"InventoryStockBookDetail_{_safe_filename(subtitle)}.pdf",
            content=content,
            content_type="application/pdf",
        )


class InventoryStockBookDetailPrintAPIView(InventoryStockBookDetailPDFAPIView):
    export_mode = "inline"
