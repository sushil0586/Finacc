from __future__ import annotations

"""
Export helpers for GSTR-1. Keeps accountant-friendly column ordering, labels, and totals.
"""

import json
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports.api.receivables_views import _write_csv, _write_excel
from reports.gstr1.services.table_views import ALL_GSTR1_TABLES


class Gstr1ExportService:
    def export_section_csv(self, headers, rows):
        return _write_csv(headers, rows)

    def export_section_excel(self, title, subtitle, headers, rows, *, numeric_columns):
        return _write_excel(title, subtitle, headers, rows, numeric_columns=numeric_columns)

    def export_summary_excel(self, *, sections, hsn_summary, document_summary, nil_exempt_summary):
        wb = Workbook()
        wb.remove(wb.active)
        header_font, header_fill, center, left, right, border = _workbook_styles()

        _write_sheet(
            wb,
            "Section Summary",
            ["Section", "Documents", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
            [
                [
                    row.get("section"),
                    row.get("document_count"),
                    row.get("taxable_amount"),
                    row.get("cgst_amount"),
                    row.get("sgst_amount"),
                    row.get("igst_amount"),
                    row.get("cess_amount"),
                    row.get("grand_total"),
                ]
                for row in sections
            ],
            header_font=header_font,
            header_fill=header_fill,
            center=center,
            left=left,
            right=right,
            border=border,
            numeric_columns={2, 3, 4, 5, 6, 7},
            totals_row=_totals_row(sections),
        )
        _write_sheet(
            wb,
            "HSN Summary",
            ["HSN/SAC", "Service", "GST Rate", "Qty", "Taxable", "CGST", "SGST", "IGST", "Cess", "Docs"],
            [
                [
                    row.get("hsn_sac_code"),
                    "Y" if row.get("is_service") else "N",
                    row.get("gst_rate"),
                    row.get("total_qty"),
                    row.get("taxable_value"),
                    row.get("cgst_amount"),
                    row.get("sgst_amount"),
                    row.get("igst_amount"),
                    row.get("cess_amount"),
                    row.get("document_count"),
                ]
                for row in hsn_summary
            ],
            header_font=header_font,
            header_fill=header_fill,
            center=center,
            left=left,
            right=right,
            border=border,
            numeric_columns={2, 3, 4, 5, 6, 7, 8, 9},
            totals_row=None,
        )
        _write_sheet(
            wb,
            "Document Summary",
            ["Doc Type", "Series", "Min No", "Max No", "Total", "Cancelled"],
            [
                [
                    row.get("doc_type"),
                    row.get("doc_code"),
                    row.get("min_doc_no"),
                    row.get("max_doc_no"),
                    row.get("document_count"),
                    row.get("cancelled_count"),
                ]
                for row in document_summary
            ],
            header_font=header_font,
            header_fill=header_fill,
            center=center,
            left=left,
            right=right,
            border=border,
            numeric_columns={2, 3, 4, 5},
            totals_row=None,
        )
        _write_sheet(
            wb,
            "Nil Exempt",
            ["Taxability", "Taxable", "CGST", "SGST", "IGST", "Cess"],
            [
                [
                    row.get("taxability"),
                    row.get("taxable_value"),
                    row.get("cgst_amount"),
                    row.get("sgst_amount"),
                    row.get("igst_amount"),
                    row.get("cess_amount"),
                ]
                for row in nil_exempt_summary
            ],
            header_font=header_font,
            header_fill=header_fill,
            center=center,
            left=left,
            right=right,
            border=border,
            numeric_columns={1, 2, 3, 4, 5},
            totals_row=None,
        )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def export_comprehensive_excel(
        self,
        *,
        sections,
        hsn_summary,
        document_summary,
        nil_exempt_summary,
        table_payloads,
        warnings,
        export_meta=None,
    ):
        wb = Workbook()
        wb.remove(wb.active)
        styles = _workbook_styles()

        # Keep current validation-friendly summary sheets.
        _write_sheet(
            wb,
            "Section Summary",
            ["Section", "Documents", "Taxable", "CGST", "SGST", "IGST", "Cess", "Total"],
            [
                [
                    row.get("section"),
                    row.get("document_count"),
                    row.get("taxable_amount"),
                    row.get("cgst_amount"),
                    row.get("sgst_amount"),
                    row.get("igst_amount"),
                    row.get("cess_amount"),
                    row.get("grand_total"),
                ]
                for row in sections
            ],
            styles=styles,
            numeric_columns={2, 3, 4, 5, 6, 7},
            totals_row=_totals_row(sections),
            report_title="GSTR-1 Readiness and Review",
            subtitle="Section summary",
            meta_rows=_build_export_meta_rows(export_meta),
        )
        _write_sheet(
            wb,
            "HSN Summary",
            ["HSN/SAC", "Service", "GST Rate", "Qty", "Taxable", "CGST", "SGST", "IGST", "Cess", "Docs"],
            [
                [
                    row.get("hsn_sac_code"),
                    "Y" if row.get("is_service") else "N",
                    row.get("gst_rate"),
                    row.get("total_qty"),
                    row.get("taxable_value"),
                    row.get("cgst_amount"),
                    row.get("sgst_amount"),
                    row.get("igst_amount"),
                    row.get("cess_amount"),
                    row.get("document_count"),
                ]
                for row in hsn_summary
            ],
            styles=styles,
            numeric_columns={2, 3, 4, 5, 6, 7, 8, 9},
            totals_row=None,
            report_title="GSTR-1 Readiness and Review",
            subtitle="HSN summary",
            meta_rows=_build_export_meta_rows(export_meta),
        )
        _write_sheet(
            wb,
            "Document Summary",
            ["Doc Type", "Series", "Min No", "Max No", "Total", "Cancelled"],
            [
                [
                    row.get("doc_type"),
                    row.get("doc_code"),
                    row.get("min_doc_no"),
                    row.get("max_doc_no"),
                    row.get("document_count"),
                    row.get("cancelled_count"),
                ]
                for row in document_summary
            ],
            styles=styles,
            numeric_columns={2, 3, 4, 5},
            totals_row=None,
            report_title="GSTR-1 Readiness and Review",
            subtitle="Document summary",
            meta_rows=_build_export_meta_rows(export_meta),
        )
        _write_sheet(
            wb,
            "Nil Exempt",
            ["Taxability", "Taxable", "CGST", "SGST", "IGST", "Cess"],
            [
                [
                    row.get("taxability"),
                    row.get("taxable_value"),
                    row.get("cgst_amount"),
                    row.get("sgst_amount"),
                    row.get("igst_amount"),
                    row.get("cess_amount"),
                ]
                for row in nil_exempt_summary
            ],
            styles=styles,
            numeric_columns={1, 2, 3, 4, 5},
            totals_row=None,
            report_title="GSTR-1 Readiness and Review",
            subtitle="Nil / exempt summary",
            meta_rows=_build_export_meta_rows(export_meta),
        )

        warning_rows = [
            [
                row.get("code"),
                row.get("severity"),
                row.get("message"),
                row.get("invoice_id"),
                row.get("invoice_number"),
                row.get("field"),
            ]
            for row in warnings
        ]
        _write_sheet(
            wb,
            "Validations",
            ["Code", "Severity", "Message", "Invoice ID", "Invoice Number", "Field"],
            warning_rows,
            styles=styles,
            numeric_columns={3},
            totals_row=None,
            report_title="GSTR-1 Readiness and Review",
            subtitle="Validation warnings",
            meta_rows=_build_export_meta_rows(export_meta),
        )

        # Add each table sheet in statutory order.
        for definition in ALL_GSTR1_TABLES:
            payload = table_payloads.get(definition.code) or {}
            rows = payload.get("rows") or []
            headers = list(rows[0].keys()) if rows else ["info"]
            excel_rows = [[row.get(header) for header in headers] for row in rows] if rows else [[payload.get("coverage", {}).get("message") or "No rows for selected scope."]]
            numeric_columns = {
                idx + 1
                for idx, header in enumerate(headers)
                if any(token in header for token in ("amount", "value", "qty", "count", "total"))
            }
            raw_title = f"{definition.code.replace('TABLE_', '').replace('TAXPAYER_', '1_3_')} {definition.label}"
            sheet_title = _safe_sheet_title(raw_title)
            _write_sheet(
                wb,
                sheet_title,
                [header.replace("_", " ").title() for header in headers],
                excel_rows,
                styles=styles,
                numeric_columns=numeric_columns,
                totals_row=None,
                report_title="GSTR-1 Readiness and Review",
                subtitle=definition.label,
                meta_rows=_build_export_meta_rows(export_meta),
            )

            if definition.code == "TABLE_11":
                groups = payload.get("groups") or {}
                for group_code in ("11A", "11B"):
                    group_rows = ((groups.get(group_code) or {}).get("rows")) or []
                    group_headers = list(group_rows[0].keys()) if group_rows else ["info"]
                    group_excel_rows = [[row.get(header) for header in group_headers] for row in group_rows] if group_rows else [["No rows for selected scope."]]
                    group_numeric = {
                        idx + 1
                        for idx, header in enumerate(group_headers)
                        if any(token in header for token in ("amount", "value", "qty", "count", "total"))
                    }
                    _write_sheet(
                        wb,
                        _safe_sheet_title(f"{group_code} Advances"),
                        [header.replace("_", " ").title() for header in group_headers],
                        group_excel_rows,
                        styles=styles,
                        numeric_columns=group_numeric,
                        totals_row=None,
                        report_title="GSTR-1 Readiness and Review",
                        subtitle=f"{group_code} advances",
                        meta_rows=_build_export_meta_rows(export_meta),
                    )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    title_font = Font(bold=True, color="1F2937", size=14)
    subtitle_font = Font(color="475569", size=10, italic=True)
    meta_label_font = Font(bold=True, color="334155", size=9)
    meta_value_font = Font(color="475569", size=9)
    total_font = Font(bold=True, color="1F2937")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="F8FBFF")
    meta_fill = PatternFill("solid", fgColor="F8FAFC")
    total_fill = PatternFill("solid", fgColor="EAF2FB")
    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "title_font": title_font,
        "subtitle_font": subtitle_font,
        "meta_label_font": meta_label_font,
        "meta_value_font": meta_value_font,
        "total_font": total_font,
        "center": center,
        "left": left,
        "right": right,
        "border": border,
        "title_fill": title_fill,
        "meta_fill": meta_fill,
        "total_fill": total_fill,
    }


def _estimate_excel_col_widths(headers, rows, *, min_width=12, max_width=28):
    sample_rows = rows[:50] if rows else []
    widths = []
    for index, header in enumerate(headers):
        values = [str(header or "")]
        for row in sample_rows:
            if index < len(row):
                values.append(str(row[index] or ""))
        widths.append(max(min_width, min(max_width, max(len(value) for value in values) + 2)))
    return widths


def _write_sheet(
    wb,
    title,
    headers,
    rows,
    *,
    styles,
    numeric_columns,
    totals_row,
    report_title=None,
    subtitle=None,
    meta_rows=None,
):
    ws = wb.create_sheet(title=_safe_sheet_title(title))
    if report_title:
        ws.append([_excel_safe_value(report_title)])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = styles["title_font"]
        title_cell.fill = styles["title_fill"]
        title_cell.alignment = styles["left"]
        title_cell.border = styles["border"]
    if subtitle:
        ws.append([_excel_safe_value(subtitle)])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        subtitle_cell = ws.cell(row=2, column=1)
        subtitle_cell.font = styles["subtitle_font"]
        subtitle_cell.alignment = styles["left"]
    for label, value in meta_rows or []:
        ws.append([_excel_safe_value(label), _excel_safe_value(value)])
        row_index = ws.max_row
        ws.cell(row=row_index, column=1).font = styles["meta_label_font"]
        ws.cell(row=row_index, column=2).font = styles["meta_value_font"]
        ws.cell(row=row_index, column=1).fill = styles["meta_fill"]
        ws.cell(row=row_index, column=2).fill = styles["meta_fill"]
    if report_title or subtitle or meta_rows:
        ws.append([])
    ws.append([_excel_safe_value(header) for header in headers])
    header_row = ws.max_row
    estimated_widths = _estimate_excel_col_widths(headers, rows + ([totals_row] if totals_row else []))
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["right"] if col_idx in numeric_columns else styles["left"]
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = estimated_widths[col_idx - 1]
    for row in rows:
        ws.append([_excel_safe_value(value) for value in row])
    if totals_row:
        ws.append([_excel_safe_value(value) for value in totals_row])
        total_row_idx = ws.max_row
    else:
        total_row_idx = None
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = styles["border"]
            if cell.row == header_row:
                continue
            cell.alignment = styles["right"] if cell.column in numeric_columns else styles["left"]
            if cell.column in numeric_columns and cell.value not in (None, ""):
                try:
                    cell.number_format = "#,##0.00"
                except Exception:
                    pass
            if total_row_idx and cell.row == total_row_idx:
                cell.font = styles["total_font"]
                cell.fill = styles["total_fill"]
    ws.freeze_panes = f"A{header_row + 1}"


def _totals_row(sections):
    if not sections:
        return None
    total_documents = sum(int(row.get("document_count") or 0) for row in sections)
    total_taxable = sum(Decimal(row.get("taxable_amount") or 0) for row in sections)
    total_cgst = sum(Decimal(row.get("cgst_amount") or 0) for row in sections)
    total_sgst = sum(Decimal(row.get("sgst_amount") or 0) for row in sections)
    total_igst = sum(Decimal(row.get("igst_amount") or 0) for row in sections)
    total_cess = sum(Decimal(row.get("cess_amount") or 0) for row in sections)
    total_grand = sum(Decimal(row.get("grand_total") or 0) for row in sections)
    return ["TOTAL", total_documents, total_taxable, total_cgst, total_sgst, total_igst, total_cess, total_grand]


def _build_export_meta_rows(export_meta):
    if not export_meta:
        return []
    return [
        ("Generated On", export_meta.get("generated_on") or "-"),
        ("Scope", export_meta.get("scope") or "-"),
        ("Context", export_meta.get("entity") or "-"),
    ]


def _safe_sheet_title(value: str) -> str:
    title = str(value or "Sheet")
    for ch in ('\\', '/', '*', '[', ']', ':', '?'):
        title = title.replace(ch, " ")
    title = " ".join(title.split()).strip()
    return (title or "Sheet")[:31]


def _excel_safe_value(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool, Decimal)):
        return value
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, default=str, sort_keys=True)
    return str(value)
