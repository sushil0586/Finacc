from io import BytesIO
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from reports.serializers import StockDayBookRequestSerializer
from reports.services.stock_daybook_service import compute_stock_daybook


def _safe_decimal_str(x, default="0"):
    return str(x) if x is not None else default


class StockDayBookExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockDayBookRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        data = compute_stock_daybook(
            entity_id=p["entity"],
            from_date=p["from_date"],
            to_date=p["to_date"],
            product_id=p.get("product"),
            location_id=p.get("location"),
            group_by_location=p.get("group_by_location", True),
            include_details=p.get("include_details", False),
        )

        rows = data.get("rows", [])
        include_details = p.get("include_details", False)

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Day Book"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title
        title = f"Stock Day Book (Entity: {p['entity']})"
        period = f"Period: {p['from_date']} to {p['to_date']} | Group By Location: {p.get('group_by_location', True)}"
        ws.append([title])
        ws.append([period])
        ws.append([])

        # Headers
        headers = [
            "Date", "Product", "Location",
            "Opening Qty", "In Qty", "Out Qty", "Closing Qty",
            "In Value", "Out Value",
        ]
        ws.append(headers)

        # Style header row
        header_row_idx = ws.max_row
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row_idx, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        # Data rows
        for r in rows:
            ws.append([
                r.get("date"),
                r.get("product_name") or r.get("product_id"),
                r.get("location"),
                r.get("opening_qty"),
                r.get("in_qty"),
                r.get("out_qty"),
                r.get("closing_qty"),
                r.get("in_value"),
                r.get("out_value"),
            ])

        # Column widths + alignment + borders
        col_widths = [12, 35, 10, 14, 12, 12, 14, 12, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows(min_row=header_row_idx, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                if cell.row == header_row_idx:
                    continue
                # alignment rules
                if cell.column in (1, 2):  # date, product
                    cell.alignment = left
                elif cell.column == 3:      # location
                    cell.alignment = center
                else:
                    cell.alignment = right

        # Optional details sheet
        if include_details and data.get("details"):
            ws2 = wb.create_sheet("Details")
            d_headers = [
                "Date", "Txn Type", "Txn Id", "Detail Id", "Voucher No",
                "Product", "Location", "Qty", "Unit Cost", "Ext Cost", "Move Type"
            ]
            ws2.append(d_headers)
            d_header_row = ws2.max_row
            for col_idx, h in enumerate(d_headers, start=1):
                c = ws2.cell(row=d_header_row, column=col_idx)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border

            for d in data["details"]:
                ws2.append([
                    d.get("entrydate"),
                    d.get("transactiontype"),
                    d.get("transactionid"),
                    d.get("detailid"),
                    d.get("voucherno"),
                    d.get("product__productname") or d.get("product_id"),
                    d.get("location"),
                    d.get("qty"),
                    d.get("unit_cost"),
                    d.get("ext_cost"),
                    d.get("move_type"),
                ])

            widths2 = [12, 12, 10, 10, 16, 30, 10, 12, 12, 12, 10]
            for i, w in enumerate(widths2, start=1):
                ws2.column_dimensions[get_column_letter(i)].width = w

            for row in ws2.iter_rows(min_row=d_header_row, max_row=ws2.max_row, min_col=1, max_col=len(d_headers)):
                for cell in row:
                    cell.border = border
                    if cell.row == d_header_row:
                        continue
                    if cell.column in (1, 2, 6):
                        cell.alignment = left
                    elif cell.column in (7, 11):
                        cell.alignment = center
                    else:
                        cell.alignment = right

        # Output
        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)

        filename = f"StockDayBook_Entity{p['entity']}_{p['from_date']}_to_{p['to_date']}.xlsx"
        resp = HttpResponse(
            buff.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class StockDayBookPDFAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockDayBookRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        data = compute_stock_daybook(
            entity_id=p["entity"],
            from_date=p["from_date"],
            to_date=p["to_date"],
            product_id=p.get("product"),
            location_id=p.get("location"),
            group_by_location=p.get("group_by_location", True),
            include_details=False,  # PDF: summary only (keeps it readable)
        )

        rows = data.get("rows", [])

        buff = BytesIO()
        filename = f"StockDayBook_Entity{p['entity']}_{p['from_date']}_to_{p['to_date']}.pdf"

        doc = SimpleDocTemplate(
            buff,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
            title=filename,
        )

        styles = getSampleStyleSheet()
        story = []

        title = Paragraph(f"<b>Stock Day Book</b>", styles["Title"])
        meta = Paragraph(
            f"Entity: <b>{p['entity']}</b> | Period: <b>{p['from_date']}</b> to <b>{p['to_date']}</b> | "
            f"Group By Location: <b>{p.get('group_by_location', True)}</b>",
            styles["Normal"]
        )

        story.extend([title, Spacer(1, 8), meta, Spacer(1, 12)])

        table_data = [[
            "Date", "Product", "Loc",
            "Opening", "In", "Out", "Closing",
            "In Value", "Out Value"
        ]]

        for r in rows:
            table_data.append([
                r.get("date") or "",
                (r.get("product_name") or str(r.get("product_id") or ""))[:40],
                "" if r.get("location") is None else str(r.get("location")),
                r.get("opening_qty") or "0.0000",
                r.get("in_qty") or "0.0000",
                r.get("out_qty") or "0.0000",
                r.get("closing_qty") or "0.0000",
                r.get("in_value") or "0.00",
                r.get("out_value") or "0.00",
            ])

        tbl = Table(table_data, repeatRows=1, colWidths=[70, 230, 40, 70, 55, 55, 70, 70, 70])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),

            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("FONTSIZE", (0, 1), (-1, -1), 8),

            ("ALIGN", (0, 1), (1, -1), "LEFT"),     # date, product left
            ("ALIGN", (2, 1), (2, -1), "CENTER"),   # loc center
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),   # numbers right

            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        story.append(tbl)
        doc.build(story)

        pdf = buff.getvalue()
        buff.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
