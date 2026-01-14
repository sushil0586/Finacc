from io import BytesIO

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from reports.serializers_stock_movement import StockMovementRequestSerializer
from reports.services.stock_movement_service import compute_stock_movement


class StockMovementExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockMovementRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        # export should include all rows
        p["page"] = 1
        p["page_size"] = 10_000_000

        data = compute_stock_movement(p=p)
        rows = data["summary"]
        totals = data["totals"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Movement"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append([f"Stock Movement Report (Entity: {p['entity']})"])
        ws.append([f"Period: {p['from_date']} to {p['to_date']}"])
        ws.append([])

        headers = [
            "Product", "Location",
            "Opening Qty", "Opening Value",
            "In Qty", "In Value",
            "Out Qty", "Out Value",
            "Net Qty",
            "Closing Qty", "Closing Value",
        ]
        ws.append(headers)
        hr = ws.max_row

        for cidx in range(1, len(headers) + 1):
            c = ws.cell(row=hr, column=cidx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        for r in rows:
            ws.append([
                r.get("product_name") or str(r["product_id"]),
                "" if r.get("location") is None else str(r.get("location")),
                str(r["opening_qty"]), str(r["opening_value"]),
                str(r["in_qty"]), str(r["in_value"]),
                str(r["out_qty"]), str(r["out_value"]),
                str(r["net_qty"]),
                str(r["closing_qty"]), str(r["closing_value"]),
            ])

        # Totals row
        ws.append([])
        ws.append([
            "TOTALS", "",
            str(totals["opening_qty"]), str(totals["opening_value"]),
            str(totals["in_qty"]), str(totals["in_value"]),
            str(totals["out_qty"]), str(totals["out_value"]),
            "",
            str(totals["closing_qty"]), str(totals["closing_value"]),
        ])

        widths = [35, 10, 14, 14, 12, 12, 12, 12, 12, 14, 14]
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

        for row in ws.iter_rows(min_row=hr, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                if cell.row == hr:
                    continue
                if cell.column == 1:
                    cell.alignment = left
                elif cell.column == 2:
                    cell.alignment = center
                else:
                    cell.alignment = right

        # optional details sheet
        if p.get("include_details") and data.get("details"):
            ws2 = wb.create_sheet("Details")
            d_headers = [
                "Date", "Txn Type", "Txn Id", "Detail Id", "Voucher",
                "Product", "Location", "Qty", "Unit Cost", "Ext Cost", "Move Type"
            ]
            ws2.append(d_headers)
            dhr = ws2.max_row
            for cidx in range(1, len(d_headers) + 1):
                c = ws2.cell(row=dhr, column=cidx)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border

            for d in data["details"]:
                ws2.append([
                    str(d["entrydate"]),
                    d.get("transactiontype"),
                    d.get("transactionid"),
                    d.get("detailid"),
                    d.get("voucherno"),
                    d.get("product__productname") or str(d["product_id"]),
                    "" if d.get("location") is None else str(d.get("location")),
                    str(d["qty"]),
                    str(d["unit_cost"]),
                    str(d["ext_cost"]),
                    d.get("move_type"),
                ])

        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)

        filename = f"StockMovement_Entity{p['entity']}_{p['from_date']}_to_{p['to_date']}.xlsx"
        resp = HttpResponse(
            buff.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class StockMovementPDFAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockMovementRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        # PDF should be summary only (readable)
        p["include_details"] = False

        data = compute_stock_movement(p=p)
        rows = data["summary"]
        totals = data["totals"]

        buff = BytesIO()
        filename = f"StockMovement_Entity{p['entity']}_{p['from_date']}_to_{p['to_date']}.pdf"

        doc = SimpleDocTemplate(
            buff,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
            title=filename,
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph("<b>Stock Movement Report</b>", styles["Title"]),
            Spacer(1, 6),
            Paragraph(
                f"Entity: <b>{p['entity']}</b> | Period: <b>{p['from_date']}</b> to <b>{p['to_date']}</b>",
                styles["Normal"]
            ),
            Spacer(1, 10),
        ]

        table_data = [[
            "Product", "Loc",
            "Opening Qty", "Opening Val",
            "In Qty", "In Val",
            "Out Qty", "Out Val",
            "Closing Qty", "Closing Val"
        ]]

        for r in rows:
            table_data.append([
                (r.get("product_name") or str(r["product_id"]))[:40],
                "" if r.get("location") is None else str(r.get("location")),
                str(r["opening_qty"]), str(r["opening_value"]),
                str(r["in_qty"]), str(r["in_value"]),
                str(r["out_qty"]), str(r["out_value"]),
                str(r["closing_qty"]), str(r["closing_value"]),
            ])

        table_data.append([
            "TOTALS", "",
            str(totals["opening_qty"]), str(totals["opening_value"]),
            str(totals["in_qty"]), str(totals["in_value"]),
            str(totals["out_qty"]), str(totals["out_value"]),
            str(totals["closing_qty"]), str(totals["closing_value"]),
        ])

        tbl = Table(table_data, repeatRows=1, colWidths=[220, 40, 70, 70, 60, 70, 60, 70, 70, 80])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        story.append(tbl)
        doc.build(story)

        pdf = buff.getvalue()
        buff.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
