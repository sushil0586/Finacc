from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports.api.receivables_views import _write_csv


def export_gstr3b_excel(*, summary: dict, warnings: list[dict], audit_context: dict | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    styles = _workbook_styles()
    header_font = styles["header_font"]
    header_fill = styles["header_fill"]
    center = styles["center"]
    left = styles["left"]
    right = styles["right"]
    border = styles["border"]

    _write_sheet(
        wb,
        "3.1 Outward-RCM",
        ["Nature of Supplies", "Taxable Value", "CGST", "SGST", "IGST", "Cess", "Total Tax"],
        [
            ["Outward taxable supplies", *_bucket_values(summary["section_3_1"]["outward_taxable_supplies"])],
            ["Outward zero-rated supplies", *_bucket_values(summary["section_3_1"]["outward_zero_rated_supplies"])],
            ["Inward supplies liable to reverse charge", *_bucket_values(summary["section_3_1"]["inward_supplies_reverse_charge"])],
            [
                "Outward nil/exempt/non-GST",
                summary["section_3_1"]["outward_nil_exempt_non_gst"]["taxable_value"],
                0,
                0,
                0,
                0,
                0,
            ],
            [
                "Non-GST outward supplies",
                summary["section_3_1"]["non_gst_outward_supplies"]["taxable_value"],
                0,
                0,
                0,
                0,
                0,
            ],
            _section_total_row(
                [
                    summary["section_3_1"]["outward_taxable_supplies"],
                    summary["section_3_1"]["outward_zero_rated_supplies"],
                    summary["section_3_1"]["inward_supplies_reverse_charge"],
                    {
                        "taxable_value": summary["section_3_1"]["outward_nil_exempt_non_gst"]["taxable_value"],
                        "cgst": 0,
                        "sgst": 0,
                        "igst": 0,
                        "cess": 0,
                        "total_tax": 0,
                    },
                    {
                        "taxable_value": summary["section_3_1"]["non_gst_outward_supplies"]["taxable_value"],
                        "cgst": 0,
                        "sgst": 0,
                        "igst": 0,
                        "cess": 0,
                        "total_tax": 0,
                    },
                ]
            ),
        ],
        numeric_columns={2, 3, 4, 5, 6, 7},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
        audit_rows=_audit_rows(audit_context),
        report_title=(audit_context or {}).get("report_title") or "GSTR-3B Summary",
    )
    _write_sheet(
        wb,
        "3.2 Inter-state",
        ["Supply Category", "Taxable Value", "CGST", "SGST", "IGST", "Cess", "Total Tax"],
        [
            ["Inter-state to unregistered", *_bucket_values(summary["section_3_2"]["interstate_supplies_to_unregistered"])],
            ["Inter-state to composition", *_bucket_values(summary["section_3_2"]["interstate_supplies_to_composition"])],
            ["Inter-state to UIN holders", *_bucket_values(summary["section_3_2"]["interstate_supplies_to_uin_holders"])],
            _section_total_row(
                [
                    summary["section_3_2"]["interstate_supplies_to_unregistered"],
                    summary["section_3_2"]["interstate_supplies_to_composition"],
                    summary["section_3_2"]["interstate_supplies_to_uin_holders"],
                ]
            ),
        ],
        numeric_columns={2, 3, 4, 5, 6, 7},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
        audit_rows=_audit_rows(audit_context),
        report_title=(audit_context or {}).get("report_title") or "GSTR-3B Summary",
    )
    _write_sheet(
        wb,
        "4 ITC",
        ["Details", "Taxable Value", "CGST", "SGST", "IGST", "Cess", "Total Tax"],
        [
            ["ITC available", *_bucket_values(summary["section_4"]["itc_available"])],
            ["ITC reversed", *_bucket_values(summary["section_4"]["itc_reversed"])],
            ["Net ITC", *_bucket_values(summary["section_4"]["net_itc"])],
            _section_total_row(
                [
                    summary["section_4"]["itc_available"],
                    summary["section_4"]["itc_reversed"],
                    summary["section_4"]["net_itc"],
                ]
            ),
        ],
        numeric_columns={2, 3, 4, 5, 6, 7},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
        audit_rows=_audit_rows(audit_context),
        report_title=(audit_context or {}).get("report_title") or "GSTR-3B Summary",
    )
    _write_sheet(
        wb,
        "5.1 Inward Exempt",
        ["Details", "Taxable Value"],
        [
            ["Inward exempt / nil / non-GST", summary["section_5_1"]["inward_exempt_nil_non_gst"]["taxable_value"]],
            _single_value_total_row(summary["section_5_1"]["inward_exempt_nil_non_gst"]["taxable_value"]),
        ],
        numeric_columns={2},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
        audit_rows=_audit_rows(audit_context),
        report_title=(audit_context or {}).get("report_title") or "GSTR-3B Summary",
    )
    _write_sheet(
        wb,
        "6.1 Tax Payment",
        ["Description", "Taxable Value", "CGST", "SGST", "IGST", "Cess", "Total Tax"],
        [
            ["Tax payable", *_bucket_values(summary["section_6_1"]["tax_payable"])],
            ["Paid through ITC", *_bucket_values(summary["section_6_1"]["tax_paid_itc"])],
            ["Paid in cash", *_bucket_values(summary["section_6_1"]["tax_paid_cash"])],
            ["Balance payable", *_bucket_values(summary["section_6_1"]["balance_payable"])],
            _section_total_row(
                [
                    summary["section_6_1"]["tax_payable"],
                    summary["section_6_1"]["tax_paid_itc"],
                    summary["section_6_1"]["tax_paid_cash"],
                    summary["section_6_1"]["balance_payable"],
                ]
            ),
        ],
        numeric_columns={2, 3, 4, 5, 6, 7},
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
        audit_rows=_audit_rows(audit_context),
        report_title=(audit_context or {}).get("report_title") or "GSTR-3B Summary",
    )
    _write_sheet(
        wb,
        "Warnings",
        ["Severity", "Code", "Message", "Section Route", "Section", "Related Report Route"],
        [
            [
                w.get("severity"),
                w.get("code"),
                w.get("message"),
                (((w.get("drilldowns") or {}).get("section_view") or {}).get("route")),
                ((((w.get("drilldowns") or {}).get("section_view") or {}).get("params") or {}).get("section")),
                (((w.get("drilldowns") or {}).get("related_report") or {}).get("route")),
            ]
            for w in warnings
        ],
        numeric_columns=set(),
        header_font=header_font,
        header_fill=header_fill,
        center=center,
        left=left,
        right=right,
        border=border,
        audit_rows=_audit_rows(audit_context),
        report_title=(audit_context or {}).get("report_title") or "GSTR-3B Summary",
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_gstr3b_csv_rows(*, summary: dict, warnings: list[dict]) -> bytes:
    headers = [
        "Section",
        "Row",
        "Taxable Value",
        "CGST",
        "SGST",
        "IGST",
        "Cess",
        "Total Tax",
        "Severity",
        "Code",
        "Message",
        "Section Route",
        "Section Key",
        "Related Report Route",
    ]
    rows = []
    rows.extend(_csv_section_rows("3.1", summary.get("section_3_1", {}).get("rows", [])))
    rows.append(_csv_total_row("3.1", summary.get("section_3_1", {}).get("rows", [])))
    rows.extend(_csv_section_rows("3.2", summary.get("section_3_2", {}).get("rows", [])))
    rows.append(_csv_total_row("3.2", summary.get("section_3_2", {}).get("rows", [])))
    rows.extend(_csv_section_rows("4", summary.get("section_4", {}).get("rows", [])))
    rows.append(_csv_total_row("4", summary.get("section_4", {}).get("rows", [])))
    rows.extend(_csv_section_rows("5.1", summary.get("section_5_1", {}).get("rows", [])))
    rows.append(_csv_total_row("5.1", summary.get("section_5_1", {}).get("rows", [])))
    rows.extend(_csv_section_rows("6.1", summary.get("section_6_1", {}).get("rows", [])))
    rows.append(_csv_total_row("6.1", summary.get("section_6_1", {}).get("rows", [])))
    for warning in warnings:
        drilldowns = warning.get("drilldowns") or {}
        section_view = drilldowns.get("section_view") or {}
        related_report = drilldowns.get("related_report") or {}
        rows.append(
            [
                "WARN",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                warning.get("severity"),
                warning.get("code"),
                warning.get("message"),
                section_view.get("route"),
                (section_view.get("params") or {}).get("section"),
                related_report.get("route"),
            ]
        )
    return _write_csv(headers, rows)


def _csv_section_rows(section: str, section_rows: list[dict]) -> list[list]:
    rows = []
    for row in section_rows:
        rows.append(
            [
                section,
                row.get("label", ""),
                _format_csv_number(row.get("taxable_value", 0)),
                _format_csv_number(row.get("cgst", 0)),
                _format_csv_number(row.get("sgst", 0)),
                _format_csv_number(row.get("igst", 0)),
                _format_csv_number(row.get("cess", 0)),
                _format_csv_number(row.get("total_tax", 0)),
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
    return rows


def _bucket_values(bucket: dict) -> list:
    return [
        bucket.get("taxable_value", 0),
        bucket.get("cgst", 0),
        bucket.get("sgst", 0),
        bucket.get("igst", 0),
        bucket.get("cess", 0),
        bucket.get("total_tax", 0),
    ]


def _section_total_row(buckets: list[dict]) -> list:
    totals = {
        "taxable_value": 0,
        "cgst": 0,
        "sgst": 0,
        "igst": 0,
        "cess": 0,
        "total_tax": 0,
    }
    for bucket in buckets:
        totals["taxable_value"] += bucket.get("taxable_value", 0) or 0
        totals["cgst"] += bucket.get("cgst", 0) or 0
        totals["sgst"] += bucket.get("sgst", 0) or 0
        totals["igst"] += bucket.get("igst", 0) or 0
        totals["cess"] += bucket.get("cess", 0) or 0
        totals["total_tax"] += bucket.get("total_tax", 0) or 0
    return ["Report Total", totals["taxable_value"], totals["cgst"], totals["sgst"], totals["igst"], totals["cess"], totals["total_tax"]]


def _single_value_total_row(taxable_value: float | int) -> list:
    return ["Report Total", taxable_value or 0]


def _csv_total_row(section: str, section_rows: list[dict]) -> list:
    total_taxable = sum((row.get("taxable_value", 0) or 0) for row in section_rows)
    total_cgst = sum((row.get("cgst", 0) or 0) for row in section_rows)
    total_sgst = sum((row.get("sgst", 0) or 0) for row in section_rows)
    total_igst = sum((row.get("igst", 0) or 0) for row in section_rows)
    total_cess = sum((row.get("cess", 0) or 0) for row in section_rows)
    total_tax = sum((row.get("total_tax", 0) or 0) for row in section_rows)
    return [
        section,
        "Report Total",
        _format_csv_number(total_taxable),
        _format_csv_number(total_cgst),
        _format_csv_number(total_sgst),
        _format_csv_number(total_igst),
        _format_csv_number(total_cess),
        _format_csv_number(total_tax),
        "",
        "",
        "",
        "",
        "",
        "",
    ]


def _format_csv_number(value) -> str:
    try:
        return f"{float(value or 0):.2f}"
    except Exception:
        return "0.00"


def _workbook_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    title_font = Font(bold=True, color="1F2937", size=14)
    subtitle_font = Font(color="475569", size=10, italic=True)
    meta_label_font = Font(bold=True, color="334155", size=9)
    meta_value_font = Font(color="475569", size=9)
    total_font = Font(bold=True, color="1F2937")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="F8FBFF")
    meta_fill = PatternFill("solid", fgColor="F8FAFC")
    total_fill = PatternFill("solid", fgColor="EAF2FB")
    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "title_font": title_font,
        "subtitle_font": subtitle_font,
        "meta_label_font": meta_label_font,
        "meta_value_font": meta_value_font,
        "total_font": total_font,
        "center": center,
        "left": left,
        "right": right,
        "border": border,
        "title_fill": title_fill,
        "meta_fill": meta_fill,
        "total_fill": total_fill,
    }


def _estimate_excel_col_widths(headers, rows, *, min_width=12, max_width=28):
    sample_rows = list(rows[:50] if isinstance(rows, list) else rows)
    widths = []
    for index, header in enumerate(headers):
        values = [str(header or "")]
        for row in sample_rows:
            if index < len(row):
                values.append(str(row[index] or ""))
        widths.append(max(min_width, min(max_width, max(len(value) for value in values) + 2)))
    return widths


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: Iterable[list],
    *,
    numeric_columns: set[int],
    header_font,
    header_fill,
    center,
    left,
    right,
    border,
    audit_rows: list[list] | None = None,
    report_title: str | None = None,
) -> None:
    ws = wb.create_sheet(title=title[:31])
    audit_rows = list(audit_rows or [])
    if report_title:
        ws.append([report_title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, color="1F2937", size=14)
        title_cell.alignment = left
        title_cell.fill = PatternFill("solid", fgColor="F8FBFF")
        for column in range(1, len(headers) + 1):
            ws.cell(row=1, column=column).border = border
    for row in audit_rows:
        ws.append(row)
        row_index = ws.max_row
        ws.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(row=row_index, column=2).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(row=row_index, column=1).font = Font(bold=True, color="334155", size=9)
        ws.cell(row=row_index, column=2).font = Font(color="475569", size=9)
        ws.cell(row=row_index, column=1).alignment = left
        ws.cell(row=row_index, column=2).alignment = left
    if audit_rows:
        ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    materialized_rows = list(rows)
    estimated_widths = _estimate_excel_col_widths(headers, materialized_rows)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = right if col_idx in numeric_columns else left
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = estimated_widths[col_idx - 1]
    for row in materialized_rows:
        ws.append(row)
    total_row_index = None
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.row == header_row:
                continue
            cell.alignment = right if cell.column in numeric_columns else left
            if cell.column in numeric_columns and cell.value not in (None, ""):
                try:
                    cell.number_format = "#,##0.00"
                except Exception:
                    pass
        first_cell = ws.cell(row=row[0].row, column=1)
        if first_cell.value == "Report Total":
            total_row_index = first_cell.row
            for total_cell in ws.iter_rows(min_row=total_row_index, max_row=total_row_index, min_col=1, max_col=len(headers)):
                for item in total_cell:
                    item.font = Font(bold=True, color="1F2937")
                    item.fill = PatternFill("solid", fgColor="EAF2FB")
    ws.freeze_panes = f"A{header_row + 1}"


def _audit_rows(audit_context: dict | None) -> list[list]:
    if not audit_context:
        return []
    generated_on = audit_context.get("generated_on") or "-"
    scope = audit_context.get("scope") or "-"
    report_period = audit_context.get("period") or "-"
    return [
        ["Period", report_period],
        ["Generated On", generated_on],
        ["Scope", scope],
    ]
