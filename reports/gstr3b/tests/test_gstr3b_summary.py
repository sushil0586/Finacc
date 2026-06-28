from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.exceptions import PermissionDenied
from rest_framework.test import APIClient, APITestCase

from Authentication.models import User
from catalog.models import Product, ProductCategory, ProductPurchaseBehavior, UnitOfMeasure
from entity.models import Entity, EntityFinancialYear, GstRegistrationType, SubEntity
from financial.models import AccountComplianceProfile, Ledger, account, accountHead, accounttype
from financial.services import create_account_with_synced_ledger
from posting.models import Entry, EntityStaticAccountMap, JournalLine, PostingBatch, StaticAccount, TxnType
from purchase.models import PurchaseInvoiceHeader, PurchaseInvoiceLine, PurchaseTaxSummary
from sales.models import SalesInvoiceHeader


@override_settings(ROOT_URLCONF="FA.urls", AUTH_PASSWORD_VALIDATORS=[])
class Gstr3bSummaryAPITests(APITestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username="gstr3b-user",
            email="gstr3b@example.com",
            password="pass123",
        )
        self.permission_codes_patch = patch(
            "reports.api.report_permissions.EffectivePermissionService.permission_codes_for_user",
            return_value=["reports.gstr3b.view"],
        )
        self.permission_codes_patch.start()
        self.addCleanup(self.permission_codes_patch.stop)
        self.client.force_authenticate(user=self.user)

        self.summary_url = reverse("reports_api:gstr3b-summary")
        self.meta_url = reverse("reports_api:gstr3b-meta")
        self.validation_url = reverse("reports_api:gstr3b-validations")
        self.export_url = reverse("reports_api:gstr3b-export")

        gst_type = GstRegistrationType.objects.create(Name="Regular", Description="Regular")
        self.entity = Entity.objects.create(
            entityname="Finacc Entity",
            legalname="Finacc Entity Pvt Ltd",
            GstRegitrationType=gst_type,
            createdby=self.user,
        )
        self.subentity = SubEntity.objects.create(entity=self.entity, subentityname="Main Branch")

        fy_start = timezone.make_aware(datetime(2025, 4, 1))
        fy_end = timezone.make_aware(datetime(2026, 3, 31))
        self.entityfin = EntityFinancialYear.objects.create(
            entity=self.entity,
            desc="FY 2025-26",
            finstartyear=fy_start,
            finendyear=fy_end,
            createdby=self.user,
        )

        self.params = {
            "entity": self.entity.id,
            "entityfinid": self.entityfin.id,
            "subentity": self.subentity.id,
            "from_date": "2025-04-01",
            "to_date": "2025-04-30",
        }
        self.acc_type = accounttype.objects.create(
            entity=self.entity,
            accounttypename="Control",
            accounttypecode="CTRL",
            createdby=self.user,
        )
        self.acc_head = accountHead.objects.create(
            entity=self.entity,
            name="Control Head",
            code=999,
            drcreffect="Debit",
            balanceType="Debit",
            accounttype=self.acc_type,
            createdby=self.user,
        )

    def _create_sales_doc(
        self,
        *,
        doc_no,
        taxable,
        cgst,
        sgst,
        igst,
        taxability,
        supply_category,
        doc_type=1,
        status=SalesInvoiceHeader.Status.POSTED,
        pos_state_code="27",
    ):
        SalesInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            doc_type=doc_type,
            status=status,
            bill_date="2025-04-05",
            posting_date="2025-04-05",
            doc_code="SINV",
            doc_no=doc_no,
            invoice_number=f"S-{doc_no}",
            customer_name="Customer",
            seller_gstin="27AAAAA1111A1Z1",
            seller_state_code="27",
            place_of_supply_state_code=pos_state_code,
            taxability=taxability,
            supply_category=supply_category,
            tax_regime=(
                SalesInvoiceHeader.TaxRegime.INTER_STATE
                if Decimal(igst) > 0
                else SalesInvoiceHeader.TaxRegime.INTRA_STATE
            ),
            is_igst=Decimal(igst) > 0,
            total_taxable_value=Decimal(taxable),
            total_cgst=Decimal(cgst),
            total_sgst=Decimal(sgst),
            total_igst=Decimal(igst),
            total_cess=Decimal("0.00"),
            grand_total=Decimal(taxable) + Decimal(cgst) + Decimal(sgst) + Decimal(igst),
        )

    def _d(self, value):
        return Decimal(str(value))

    def _create_purchase_doc(
        self,
        *,
        doc_no,
        taxable,
        cgst,
        sgst,
        igst,
        default_taxability,
        is_itc_eligible=True,
        is_reverse_charge=False,
        doc_type=1,
        vendor=None,
        supply_category=PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
        bill_date="2025-04-10",
        posting_date="2025-04-10",
        ref_document=None,
    ):
        return PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            doc_type=doc_type,
            status=PurchaseInvoiceHeader.Status.POSTED,
            bill_date=bill_date,
            posting_date=posting_date,
            doc_code="PINV",
            doc_no=doc_no,
            vendor=vendor,
            vendor_name="Vendor",
            vendor_gstin=getattr(getattr(vendor, "compliance_profile", None), "gstno", None),
            supply_category=supply_category,
            default_taxability=default_taxability,
            is_itc_eligible=is_itc_eligible,
            is_reverse_charge=is_reverse_charge,
            ref_document=ref_document,
            total_taxable=Decimal(taxable),
            total_cgst=Decimal(cgst),
            total_sgst=Decimal(sgst),
            total_igst=Decimal(igst),
            total_cess=Decimal("0.00"),
            total_gst=Decimal(cgst) + Decimal(sgst) + Decimal(igst),
            grand_total=Decimal(taxable) + Decimal(cgst) + Decimal(sgst) + Decimal(igst),
        )

    def _create_output_tax_mapping(self, code: str, ledger_code: int, account_code: int, name: str):
        static, _ = StaticAccount.objects.get_or_create(
            code=code,
            defaults={"name": code, "group": "GST_OUTPUT", "is_required": False},
        )
        ledger = Ledger.objects.create(
            entity=self.entity,
            ledger_code=ledger_code,
            name=name,
            accounthead=self.acc_head,
            createdby=self.user,
        )
        acc = create_account_with_synced_ledger(
            account_data={
                "entity": self.entity,
                "ledger": ledger,
                "accountname": name,
                "createdby": self.user,
            },
            ledger_overrides={"ledger_code": account_code, "accounthead": self.acc_head, "is_party": True},
        )
        EntityStaticAccountMap.objects.create(
            entity=self.entity,
            sub_entity=self.subentity,
            static_account=static,
            account=acc,
            ledger=ledger,
            is_active=True,
            createdby=self.user,
        )
        return acc, ledger

    def test_meta_endpoint(self):
        response = self.client.get(self.meta_url, {"entity": self.entity.id})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["phase"], 2)

    @patch("reports.gstr3b.views.assert_any_report_permission", side_effect=PermissionDenied("forbidden"))
    def test_summary_denies_when_report_permission_is_missing(self, _assert_permission):
        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 403)

    def test_summary_exposes_browser_contract_metadata(self):
        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "gstr3b-summary")
        self.assertEqual(payload["report_name"], "GSTR-3B Summary")
        self.assertEqual(set(payload["available_exports"]), {"json", "xlsx", "csv"})
        self.assertTrue(payload["actions"]["can_export_excel"])
        self.assertFalse(payload["actions"]["can_export_pdf"])
        self.assertTrue(payload["actions"]["can_export_csv"])
        self.assertTrue(payload["actions"]["can_drilldown"])
        self.assertIn("excel", payload["actions"]["export_urls"])
        self.assertIn("csv", payload["actions"]["export_urls"])
        self.assertIn("json", payload["actions"]["export_urls"])
        self.assertIn("format=xlsx", payload["actions"]["export_urls"]["excel"])
        self.assertIn("format=csv", payload["actions"]["export_urls"]["csv"])
        self.assertIn("format=json", payload["actions"]["export_urls"]["json"])

    def test_summary_computes_phase1_sections(self):
        self._create_sales_doc(
            doc_no=1,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
        )
        self._create_sales_doc(
            doc_no=2,
            doc_type=SalesInvoiceHeader.DocType.CREDIT_NOTE,
            taxable="200.00",
            cgst="18.00",
            sgst="18.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
        )
        self._create_sales_doc(
            doc_no=3,
            taxable="500.00",
            cgst="0.00",
            sgst="0.00",
            igst="90.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.EXPORT_WITH_IGST,
        )
        self._create_sales_doc(
            doc_no=4,
            taxable="300.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.EXEMPT,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
        )

        self._create_purchase_doc(
            doc_no=11,
            taxable="400.00",
            cgst="0.00",
            sgst="0.00",
            igst="72.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=True,
        )
        self._create_purchase_doc(
            doc_no=12,
            taxable="500.00",
            cgst="45.00",
            sgst="45.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
        )
        self._create_purchase_doc(
            doc_no=13,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=False,
            is_reverse_charge=False,
        )
        self._create_purchase_doc(
            doc_no=14,
            taxable="250.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.EXEMPT,
            is_itc_eligible=False,
            is_reverse_charge=False,
        )

        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("actions", payload)
        self.assertEqual(set(payload["available_exports"]), {"json", "xlsx", "csv"})
        self.assertEqual(payload["actions"]["can_export_excel"], True)
        self.assertEqual(payload["actions"]["can_export_pdf"], False)
        self.assertEqual(payload["actions"]["can_export_csv"], True)
        self.assertEqual(payload["actions"]["can_drilldown"], True)
        data = payload["summary"]

        self.assertEqual(self._d(data["section_3_1"]["outward_taxable_supplies"]["taxable_value"]), Decimal("800.00"))
        self.assertEqual(self._d(data["section_3_1"]["outward_taxable_supplies"]["cgst"]), Decimal("72.00"))
        self.assertEqual(self._d(data["section_3_1"]["outward_taxable_supplies"]["sgst"]), Decimal("72.00"))
        self.assertEqual(self._d(data["section_3_1"]["outward_zero_rated_supplies"]["taxable_value"]), Decimal("500.00"))
        self.assertEqual(self._d(data["section_3_1"]["outward_zero_rated_supplies"]["igst"]), Decimal("90.00"))
        self.assertEqual(self._d(data["section_3_1"]["outward_nil_exempt_non_gst"]["taxable_value"]), Decimal("300.00"))
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["igst"]), Decimal("72.00"))
        section_31_rows = data["section_3_1"]["rows"]
        self.assertEqual(len(section_31_rows), 5)
        self.assertEqual(section_31_rows[0]["label"], "Outward taxable supplies")
        self.assertEqual(self._d(section_31_rows[0]["total_tax"]), Decimal("144.00"))
        self.assertEqual(section_31_rows[3]["label"], "Outward nil/exempt/non-GST")
        self.assertEqual(self._d(section_31_rows[3]["taxable_value"]), Decimal("300.00"))
        self.assertEqual(self._d(section_31_rows[3]["total_tax"]), Decimal("0.00"))
        self.assertEqual(section_31_rows[4]["label"], "Non-GST outward supplies")
        self.assertEqual(self._d(section_31_rows[4]["taxable_value"]), Decimal("0.00"))
        self.assertEqual(
            self._d(data["section_3_2"]["interstate_supplies_to_unregistered"]["taxable_value"]),
            Decimal("0.00"),
        )
        self.assertEqual(
            self._d(data["section_3_2"]["interstate_supplies_to_composition"]["taxable_value"]),
            Decimal("0.00"),
        )
        section_32_rows = data["section_3_2"]["rows"]
        self.assertEqual(len(section_32_rows), 3)
        self.assertEqual(section_32_rows[0]["label"], "Inter-state to unregistered")
        self.assertEqual(self._d(section_32_rows[0]["taxable_value"]), Decimal("0.00"))

        self.assertEqual(self._d(data["section_4"]["itc_available"]["cgst"]), Decimal("45.00"))
        self.assertEqual(self._d(data["section_4"]["itc_reversed"]["cgst"]), Decimal("9.00"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["cgst"]), Decimal("36.00"))
        section_4_rows = data["section_4"]["rows"]
        self.assertEqual(len(section_4_rows), 3)
        self.assertEqual(section_4_rows[2]["label"], "Net ITC")
        self.assertEqual(self._d(section_4_rows[2]["cgst"]), Decimal("36.00"))
        self.assertEqual(self._d(data["section_5_1"]["inward_exempt_nil_non_gst"]["taxable_value"]), Decimal("250.00"))
        section_51_rows = data["section_5_1"]["rows"]
        self.assertEqual(len(section_51_rows), 1)
        self.assertEqual(section_51_rows[0]["label"], "Inward exempt / nil / non-GST")
        self.assertEqual(self._d(section_51_rows[0]["taxable_value"]), Decimal("250.00"))
        self.assertEqual(self._d(data["section_6_1"]["tax_payable"]["igst"]), Decimal("162.00"))
        self.assertEqual(self._d(data["section_6_1"]["tax_paid_cash"]["igst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_6_1"]["tax_paid_itc"]["igst"]), Decimal("72.00"))
        self.assertEqual(self._d(data["section_6_1"]["balance_payable"]["igst"]), Decimal("90.00"))
        section_61_rows = data["section_6_1"]["rows"]
        self.assertEqual(len(section_61_rows), 4)
        self.assertEqual(section_61_rows[1]["label"], "Paid through ITC")
        self.assertEqual(self._d(section_61_rows[1]["igst"]), Decimal("72.00"))
        self.assertEqual(section_61_rows[3]["label"], "Balance payable")
        self.assertEqual(self._d(section_61_rows[3]["igst"]), Decimal("90.00"))

        self.assertEqual(self._d(data["totals"]["tax_payable"]["cgst"]), Decimal("72.00"))
        self.assertEqual(self._d(data["totals"]["tax_payable"]["igst"]), Decimal("162.00"))
        self.assertEqual(self._d(data["totals"]["net_cash_tax_payable"]["cgst"]), Decimal("36.00"))
        self.assertEqual(self._d(data["totals"]["net_cash_tax_payable"]["igst"]), Decimal("90.00"))

    def test_csv_export_preserves_all_summary_rows_and_warning_metadata(self):
        self._create_sales_doc(
            doc_no=1,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
        )
        self._create_sales_doc(
            doc_no=2,
            taxable="300.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.EXEMPT,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
        )
        self._create_sales_doc(
            doc_no=3,
            taxable="500.00",
            cgst="0.00",
            sgst="0.00",
            igst="90.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
            pos_state_code="29",
        )
        self._create_purchase_doc(
            doc_no=11,
            taxable="400.00",
            cgst="36.00",
            sgst="36.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
        )
        self._create_purchase_doc(
            doc_no=12,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=False,
            is_reverse_charge=False,
        )
        self._create_purchase_doc(
            doc_no=13,
            taxable="250.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.EXEMPT,
            is_itc_eligible=False,
            is_reverse_charge=False,
        )

        response = self.client.get(self.export_url, {**self.params, "format": "csv"})

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Section,Row,Taxable Value,CGST,SGST,IGST,Cess,Total Tax,Severity,Code,Message,Section Route,Section Key,Related Report Route", content)
        self.assertIn("3.1,Outward nil/exempt/non-GST,300.00,0.00,0.00,0.00,0.00,0.00", content)
        self.assertIn("3.1,Non-GST outward supplies,0.00,0.00,0.00,0.00,0.00,0.00", content)
        self.assertIn("3.2,Inter-state to unregistered,500.00,0.00,0.00,90.00,0.00,90.00", content)
        self.assertIn("3.2,Inter-state to composition,0.00,0.00,0.00,0.00,0.00,0.00", content)
        self.assertIn("3.2,Inter-state to UIN holders,0.00,0.00,0.00,0.00,0.00,0.00", content)
        self.assertIn("4,ITC available,400.00,36.00,36.00,0.00,0.00,72.00", content)
        self.assertIn("4,ITC reversed,100.00,9.00,9.00,0.00,0.00,18.00", content)
        self.assertIn("4,Net ITC,300.00,27.00,27.00,0.00,0.00,54.00", content)
        self.assertIn("5.1,Inward exempt / nil / non-GST,250.00,0,0,0,0,0", content)

    def test_xlsx_export_preserves_summary_rows_and_warnings_sheet(self):
        self._create_sales_doc(
            doc_no=1,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
        )
        self._create_sales_doc(
            doc_no=2,
            taxable="300.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.EXEMPT,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
        )
        self._create_sales_doc(
            doc_no=3,
            taxable="500.00",
            cgst="0.00",
            sgst="0.00",
            igst="90.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
            pos_state_code="29",
        )
        self._create_purchase_doc(
            doc_no=11,
            taxable="400.00",
            cgst="36.00",
            sgst="36.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
        )
        self._create_purchase_doc(
            doc_no=12,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=False,
            is_reverse_charge=False,
        )
        self._create_purchase_doc(
            doc_no=13,
            taxable="250.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.EXEMPT,
            is_itc_eligible=False,
            is_reverse_charge=False,
        )

        response = self.client.get(self.export_url, {**self.params, "format": "xlsx"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["3.1 Outward-RCM", "3.2 Inter-state", "4 ITC", "5.1 Inward Exempt", "6.1 Tax Payment", "Warnings"],
        )

        sheet31 = workbook["3.1 Outward-RCM"]
        self.assertEqual(sheet31["A5"].value, "Outward taxable supplies")
        self.assertEqual(Decimal(str(sheet31["B5"].value)), Decimal("1500.00"))
        self.assertEqual(sheet31["A8"].value, "Outward nil/exempt/non-GST")
        self.assertEqual(Decimal(str(sheet31["B8"].value)), Decimal("300.00"))
        self.assertEqual(sheet31["A9"].value, "Non-GST outward supplies")

        sheet32 = workbook["3.2 Inter-state"]
        self.assertEqual(sheet32["A5"].value, "Inter-state to unregistered")
        self.assertEqual(Decimal(str(sheet32["B5"].value)), Decimal("500.00"))
        self.assertEqual(sheet32["A6"].value, "Inter-state to composition")
        self.assertEqual(sheet32["A7"].value, "Inter-state to UIN holders")

        sheet4 = workbook["4 ITC"]
        self.assertEqual(sheet4["A5"].value, "ITC available")
        self.assertEqual(Decimal(str(sheet4["B5"].value)), Decimal("400.00"))
        self.assertEqual(sheet4["A6"].value, "ITC reversed")
        self.assertEqual(Decimal(str(sheet4["B6"].value)), Decimal("100.00"))
        self.assertEqual(sheet4["A7"].value, "Net ITC")
        self.assertEqual(Decimal(str(sheet4["B7"].value)), Decimal("300.00"))

        sheet51 = workbook["5.1 Inward Exempt"]
        self.assertEqual(sheet51["A5"].value, "Inward exempt / nil / non-GST")
        self.assertEqual(Decimal(str(sheet51["B5"].value)), Decimal("250.00"))

        warnings = workbook["Warnings"]
        self.assertEqual(warnings["A5"].value, "info")
        self.assertEqual(warnings["B5"].value, "GSTR3B_CASH_TAX_SOURCE_PENDING")

    def test_xlsx_export_includes_audit_context_rows(self):
        self._create_sales_doc(
            doc_no=1,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
        )

        response = self.client.get(self.export_url, {**self.params, "format": "xlsx"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        sheet31 = workbook["3.1 Outward-RCM"]
        self.assertEqual(sheet31["A1"].value, "Generated On")
        self.assertTrue(bool(sheet31["B1"].value))
        self.assertEqual(sheet31["A2"].value, "Scope")
        self.assertIn("entity=", str(sheet31["B2"].value))
        self.assertIn("from_date=2025-04-01", str(sheet31["B2"].value))
        self.assertIn("to_date=2025-04-30", str(sheet31["B2"].value))
        self.assertEqual(sheet31["A4"].value, "Nature of Supplies")

    def test_exports_preserve_warning_drilldown_context(self):
        self._create_sales_doc(
            doc_no=1,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
            pos_state_code="",
        )

        csv_response = self.client.get(self.export_url, {**self.params, "format": "csv"})
        xlsx_response = self.client.get(self.export_url, {**self.params, "format": "xlsx"})

        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(xlsx_response.status_code, 200)

        csv_content = csv_response.content.decode("utf-8")
        self.assertIn("WARN,,,,,,,,warning,GSTR3B_POS_MISSING,1 in-scope sales invoices have missing place of supply.,/gstr3breport,3.1,/gstreport", csv_content)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content), data_only=True)
        warnings = workbook["Warnings"]
        self.assertEqual(warnings["A1"].value, "Generated On")
        self.assertTrue(bool(warnings["B1"].value))
        self.assertEqual(warnings["A2"].value, "Scope")
        self.assertIn("entity=", str(warnings["B2"].value))
        self.assertEqual(warnings["A4"].value, "Severity")
        self.assertEqual(warnings["B4"].value, "Code")
        self.assertEqual(warnings["C4"].value, "Message")
        self.assertEqual(warnings["D4"].value, "Section Route")
        self.assertEqual(warnings["E4"].value, "Section")
        self.assertEqual(warnings["F4"].value, "Related Report Route")
        self.assertEqual(warnings["A5"].value, "warning")
        self.assertEqual(warnings["B5"].value, "GSTR3B_POS_MISSING")
        self.assertEqual(warnings["D5"].value, "/gstr3breport")
        self.assertEqual(warnings["E5"].value, "3.1")
        self.assertEqual(warnings["F5"].value, "/gstreport")

    def test_summary_uses_posted_bank_cash_vouchers_for_tax_paid_cash(self):
        self._create_sales_doc(
            doc_no=1,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
        )
        self._create_purchase_doc(
            doc_no=12,
            taxable="500.00",
            cgst="45.00",
            sgst="45.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
        )
        cgst_account, cgst_ledger = self._create_output_tax_mapping("OUTPUT_CGST", 8101, 9101, "Output CGST")
        sgst_account, sgst_ledger = self._create_output_tax_mapping("OUTPUT_SGST", 8102, 9102, "Output SGST")

        batch = PostingBatch.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL_BANK,
            txn_id=5001,
            voucher_no="GST-PAY-001",
            created_by=self.user,
        )
        entry = Entry.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL_BANK,
            txn_id=5001,
            voucher_no="GST-PAY-001",
            posting_date=timezone.datetime(2025, 4, 21).date(),
            voucher_date=timezone.datetime(2025, 4, 21).date(),
            status=2,
            posting_batch=batch,
            created_by=self.user,
        )
        JournalLine.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL_BANK,
            txn_id=5001,
            voucher_no="GST-PAY-001",
            account=cgst_account,
            ledger=cgst_ledger,
            drcr=True,
            amount=Decimal("50.00"),
            posting_date=timezone.datetime(2025, 4, 21).date(),
            created_by=self.user,
        )
        JournalLine.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL_BANK,
            txn_id=5001,
            voucher_no="GST-PAY-001",
            account=sgst_account,
            ledger=sgst_ledger,
            drcr=True,
            amount=Decimal("40.00"),
            posting_date=timezone.datetime(2025, 4, 21).date(),
            created_by=self.user,
        )

        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_6_1"]["tax_paid_cash"]["cgst"]), Decimal("50.00"))
        self.assertEqual(self._d(data["section_6_1"]["tax_paid_cash"]["sgst"]), Decimal("40.00"))
        self.assertEqual(self._d(data["section_6_1"]["balance_payable"]["cgst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_6_1"]["balance_payable"]["sgst"]), Decimal("5.00"))

    def test_summary_excludes_unregistered_non_rcm_purchase_from_itc_available(self):
        self._create_purchase_doc(
            doc_no=21,
            taxable="500.00",
            cgst="45.00",
            sgst="45.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
        )
        PurchaseInvoiceHeader.objects.filter(doc_no=21).update(vendor_gstin="")

        response = self.client.get(self.summary_url, self.params)

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["taxable_value"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["itc_available"]["cgst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["itc_available"]["sgst"]), Decimal("0.00"))

    def test_summary_excludes_composition_vendor_purchase_from_itc_available(self):
        vendor = account.objects.create(entity=self.entity, accountname="Composition Vendor")
        AccountComplianceProfile.objects.create(
            entity=self.entity,
            account=vendor,
            gstno="27ABCDE1234F1Z5",
            gstregtype="Composition",
            createdby=self.user,
        )
        self._create_purchase_doc(
            doc_no=22,
            taxable="500.00",
            cgst="45.00",
            sgst="45.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
            vendor=vendor,
        )

        response = self.client.get(self.summary_url, self.params)

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["taxable_value"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["itc_available"]["cgst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["itc_available"]["sgst"]), Decimal("0.00"))

    def test_summary_excludes_import_goods_from_normal_itc_available(self):
        self._create_purchase_doc(
            doc_no=23,
            taxable="1000.00",
            cgst="0.00",
            sgst="0.00",
            igst="180.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=False,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.IMPORT_GOODS,
        )

        response = self.client.get(self.summary_url, self.params)

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["taxable_value"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["itc_available"]["igst"]), Decimal("0.00"))

    def test_summary_places_purchase_credit_note_itc_adjustment_in_current_period(self):
        original = self._create_purchase_doc(
            doc_no=24,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-04-10",
            posting_date="2025-04-10",
        )
        self._create_purchase_doc(
            doc_no=25,
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-05",
            posting_date="2025-05-05",
            ref_document=original,
        )

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["cgst"]), Decimal("-9.00"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["cgst"]), Decimal("-9.00"))

    def test_summary_places_reverse_charge_credit_note_adjustment_in_current_period(self):
        original = self._create_purchase_doc(
            doc_no=26,
            taxable="300.00",
            cgst="0.00",
            sgst="0.00",
            igst="54.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=True,
            bill_date="2025-04-12",
            posting_date="2025-04-12",
        )
        self._create_purchase_doc(
            doc_no=27,
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            taxable="300.00",
            cgst="0.00",
            sgst="0.00",
            igst="54.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=True,
            bill_date="2025-05-06",
            posting_date="2025-05-06",
            ref_document=original,
        )

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["igst"]), Decimal("-54.00"))

    def test_summary_itc_lifecycle_keeps_blocked_inputs_out_and_applies_current_period_adjustment(self):
        self._create_purchase_doc(
            doc_no=28,
            taxable="200.00",
            cgst="18.00",
            sgst="18.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-08",
            posting_date="2025-05-08",
        )
        self._create_purchase_doc(
            doc_no=29,
            taxable="150.00",
            cgst="13.50",
            sgst="13.50",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=False,
            bill_date="2025-05-09",
            posting_date="2025-05-09",
        )
        original = self._create_purchase_doc(
            doc_no=30,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-04-10",
            posting_date="2025-04-10",
        )
        self._create_purchase_doc(
            doc_no=31,
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            taxable="100.00",
            cgst="9.00",
            sgst="9.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-11",
            posting_date="2025-05-11",
            ref_document=original,
        )
        self._create_purchase_doc(
            doc_no=32,
            taxable="300.00",
            cgst="0.00",
            sgst="0.00",
            igst="54.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=True,
            bill_date="2025-05-12",
            posting_date="2025-05-12",
        )

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["cgst"]), Decimal("9.00"))
        self.assertEqual(self._d(data["section_4"]["itc_available"]["sgst"]), Decimal("9.00"))
        self.assertEqual(self._d(data["section_4"]["itc_reversed"]["cgst"]), Decimal("13.50"))
        self.assertEqual(self._d(data["section_4"]["itc_reversed"]["sgst"]), Decimal("13.50"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["cgst"]), Decimal("-4.50"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["sgst"]), Decimal("-4.50"))
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["igst"]), Decimal("54.00"))

    def test_summary_uses_tax_summary_split_for_partial_itc_invoice(self):
        header = self._create_purchase_doc(
            doc_no=33,
            taxable="1500.00",
            cgst="0.00",
            sgst="0.00",
            igst="270.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-13",
            posting_date="2025-05-13",
        )
        PurchaseTaxSummary.objects.create(
            header=header,
            taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            hsn_sac="1001",
            is_service=False,
            gst_rate=Decimal("18.00"),
            is_reverse_charge=False,
            taxable_value=Decimal("1500.00"),
            igst_amount=Decimal("270.00"),
            total_value=Decimal("1770.00"),
            itc_eligible_tax=Decimal("180.00"),
            itc_ineligible_tax=Decimal("90.00"),
        )

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["igst"]), Decimal("180.00"))
        self.assertEqual(self._d(data["section_4"]["itc_reversed"]["igst"]), Decimal("90.00"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["igst"]), Decimal("90.00"))

    def test_summary_keeps_itc_buckets_unchanged_when_purchase_has_tds_deduction(self):
        baseline = self._create_purchase_doc(
            doc_no=330,
            taxable="100.00",
            cgst="0.00",
            sgst="0.00",
            igst="18.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-13",
            posting_date="2025-05-13",
        )
        tds_doc = self._create_purchase_doc(
            doc_no=334,
            taxable="100.00",
            cgst="0.00",
            sgst="0.00",
            igst="18.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-13",
            posting_date="2025-05-13",
        )
        baseline.tds_amount = Decimal("0.00")
        baseline.save(update_fields=["tds_amount", "updated_at"])
        tds_doc.tds_amount = Decimal("10.00")
        tds_doc.save(update_fields=["tds_amount", "updated_at"])

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["igst"]), Decimal("36.00"))
        self.assertEqual(self._d(data["section_4"]["itc_reversed"]["igst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["igst"]), Decimal("36.00"))

    def test_summary_uses_tax_summaries_for_reverse_charge_liability_when_header_gst_is_zero(self):
        same_state = self._create_purchase_doc(
            doc_no=331,
            taxable="500.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=True,
            bill_date="2025-05-13",
            posting_date="2025-05-13",
        )
        PurchaseTaxSummary.objects.create(
            header=same_state,
            taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            hsn_sac="9965",
            is_service=True,
            gst_rate=Decimal("18.00"),
            is_reverse_charge=True,
            taxable_value=Decimal("500.00"),
            cgst_amount=Decimal("45.00"),
            sgst_amount=Decimal("45.00"),
            igst_amount=Decimal("0.00"),
            total_value=Decimal("590.00"),
            itc_eligible_tax=Decimal("90.00"),
            itc_ineligible_tax=Decimal("0.00"),
        )
        interstate = self._create_purchase_doc(
            doc_no=332,
            taxable="500.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            is_reverse_charge=True,
            bill_date="2025-05-13",
            posting_date="2025-05-13",
        )
        PurchaseTaxSummary.objects.create(
            header=interstate,
            taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            hsn_sac="9982",
            is_service=True,
            gst_rate=Decimal("18.00"),
            is_reverse_charge=True,
            taxable_value=Decimal("500.00"),
            cgst_amount=Decimal("0.00"),
            sgst_amount=Decimal("0.00"),
            igst_amount=Decimal("90.00"),
            total_value=Decimal("590.00"),
            itc_eligible_tax=Decimal("90.00"),
            itc_ineligible_tax=Decimal("0.00"),
        )

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["taxable_value"]), Decimal("1000.00"))
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["cgst"]), Decimal("45.00"))
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["sgst"]), Decimal("45.00"))
        self.assertEqual(self._d(data["section_3_1"]["inward_supplies_reverse_charge"]["igst"]), Decimal("90.00"))

    def test_summary_treats_asset_purchase_itc_in_normal_available_bucket(self):
        uom = UnitOfMeasure.objects.create(entity=self.entity, code="PCS", description="Pieces")
        category = ProductCategory.objects.create(entity=self.entity, pcategoryname="Capital Goods")
        product = Product.objects.create(
            entity=self.entity,
            productname="Laptop",
            sku="AST-LAP-01",
            productdesc="Capital asset",
            productcategory=category,
            base_uom=uom,
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.ASSET,
        )
        header = self._create_purchase_doc(
            doc_no=34,
            taxable="100000.00",
            cgst="0.00",
            sgst="0.00",
            igst="18000.00",
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            is_itc_eligible=True,
            bill_date="2025-05-14",
            posting_date="2025-05-14",
        )
        PurchaseInvoiceLine.objects.create(
            header=header,
            line_no=1,
            product=product,
            uom=uom,
            qty=Decimal("1.0000"),
            rate=Decimal("100000.00"),
            product_desc="Capital asset",
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.ASSET,
            taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            taxable_value=Decimal("100000.00"),
            gst_rate=Decimal("18.00"),
            cgst_percent=Decimal("0.00"),
            sgst_percent=Decimal("0.00"),
            igst_percent=Decimal("18.00"),
            cgst_amount=Decimal("0.00"),
            sgst_amount=Decimal("0.00"),
            igst_amount=Decimal("18000.00"),
            cess_amount=Decimal("0.00"),
            line_total=Decimal("118000.00"),
            is_itc_eligible=True,
        )
        PurchaseTaxSummary.objects.create(
            header=header,
            taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            hsn_sac="8471",
            is_service=False,
            gst_rate=Decimal("18.00"),
            is_reverse_charge=False,
            taxable_value=Decimal("100000.00"),
            igst_amount=Decimal("18000.00"),
            total_value=Decimal("118000.00"),
            itc_eligible_tax=Decimal("18000.00"),
            itc_ineligible_tax=Decimal("0.00"),
        )

        response = self.client.get(
            self.summary_url,
            {**self.params, "from_date": "2025-05-01", "to_date": "2025-05-31"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]
        self.assertEqual(self._d(data["section_4"]["itc_available"]["igst"]), Decimal("18000.00"))
        self.assertEqual(self._d(data["section_4"]["itc_reversed"]["igst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["section_4"]["net_itc"]["igst"]), Decimal("18000.00"))

    def test_summary_classifies_sez_and_deemed_export_under_outward_taxable(self):
        self._create_sales_doc(
            doc_no=31,
            taxable="600.00",
            cgst="0.00",
            sgst="0.00",
            igst="108.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.SEZ_WITH_IGST,
        )
        self._create_sales_doc(
            doc_no=32,
            taxable="400.00",
            cgst="36.00",
            sgst="36.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DEEMED_EXPORT,
        )
        self._create_sales_doc(
            doc_no=33,
            taxable="500.00",
            cgst="0.00",
            sgst="0.00",
            igst="90.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.EXPORT_WITH_IGST,
        )

        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]["section_3_1"]

        self.assertEqual(self._d(data["outward_taxable_supplies"]["taxable_value"]), Decimal("1000.00"))
        self.assertEqual(self._d(data["outward_taxable_supplies"]["igst"]), Decimal("108.00"))
        self.assertEqual(self._d(data["outward_taxable_supplies"]["cgst"]), Decimal("36.00"))
        self.assertEqual(self._d(data["outward_taxable_supplies"]["sgst"]), Decimal("36.00"))
        self.assertEqual(self._d(data["outward_zero_rated_supplies"]["taxable_value"]), Decimal("500.00"))
        self.assertEqual(self._d(data["outward_zero_rated_supplies"]["igst"]), Decimal("90.00"))

    def test_summary_classifies_export_without_igst_as_zero_rated_and_sez_without_igst_as_outward_taxable(self):
        self._create_sales_doc(
            doc_no=34,
            taxable="700.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.EXPORT_WITHOUT_IGST,
            pos_state_code="96",
        )
        self._create_sales_doc(
            doc_no=35,
            taxable="450.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.SEZ_WITHOUT_IGST,
            pos_state_code="29",
        )

        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]["section_3_1"]

        self.assertEqual(self._d(data["outward_zero_rated_supplies"]["taxable_value"]), Decimal("700.00"))
        self.assertEqual(self._d(data["outward_zero_rated_supplies"]["igst"]), Decimal("0.00"))
        self.assertEqual(self._d(data["outward_taxable_supplies"]["taxable_value"]), Decimal("450.00"))
        self.assertEqual(self._d(data["outward_taxable_supplies"]["igst"]), Decimal("0.00"))

    def test_summary_groups_nil_rated_and_non_gst_sales_under_nil_exempt_non_gst(self):
        self._create_sales_doc(
            doc_no=36,
            taxable="250.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.NIL_RATED,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
        )
        self._create_sales_doc(
            doc_no=37,
            taxable="400.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.NON_GST,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2C,
        )

        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        data = response.json()["summary"]["section_3_1"]

        self.assertEqual(self._d(data["outward_nil_exempt_non_gst"]["taxable_value"]), Decimal("650.00"))

    def test_summary_includes_confirmed_sales_invoices_for_outward_taxable(self):
        self._create_sales_doc(
            doc_no=41,
            taxable="1000.00",
            cgst="90.00",
            sgst="90.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
            status=SalesInvoiceHeader.Status.POSTED,
        )
        self._create_sales_doc(
            doc_no=42,
            taxable="847.46",
            cgst="76.27",
            sgst="76.27",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
            status=SalesInvoiceHeader.Status.CONFIRMED,
        )

        response = self.client.get(self.summary_url, self.params)
        self.assertEqual(response.status_code, 200)
        section = response.json()["summary"]["section_3_1"]["outward_taxable_supplies"]

        self.assertEqual(self._d(section["taxable_value"]), Decimal("1847.46"))
        self.assertEqual(self._d(section["cgst"]), Decimal("166.27"))
        self.assertEqual(self._d(section["sgst"]), Decimal("166.27"))

    def test_validations_endpoint(self):
        self._create_sales_doc(
            doc_no=50,
            taxable="100.00",
            cgst="0.00",
            sgst="0.00",
            igst="0.00",
            taxability=SalesInvoiceHeader.Taxability.TAXABLE,
            supply_category=SalesInvoiceHeader.SupplyCategory.DOMESTIC_B2B,
            pos_state_code="",
        )
        response = self.client.get(self.validation_url, self.params)
        self.assertEqual(response.status_code, 200)
        warnings = response.json()["warnings"]
        self.assertTrue(any(w["code"] == "GSTR3B_POS_MISSING" for w in warnings))
        self.assertTrue(any(w["code"] == "GSTR3B_TAX_BREAKUP_MISSING" for w in warnings))
        self.assertTrue(any(w["code"] == "GSTR3B_CASH_TAX_SOURCE_PENDING" for w in warnings))

        pos_warning = next(w for w in warnings if w["code"] == "GSTR3B_POS_MISSING")
        self.assertEqual(pos_warning["drilldowns"]["section_view"]["route"], "/gstr3breport")
        self.assertEqual(pos_warning["drilldowns"]["section_view"]["params"]["section"], "3.1")
        self.assertEqual(pos_warning["drilldowns"]["related_report"]["route"], "/gstreport")
        self.assertEqual(pos_warning["drilldowns"]["related_report"]["params"]["entityfinid"], self.entityfin.id)

        cash_warning = next(w for w in warnings if w["code"] == "GSTR3B_CASH_TAX_SOURCE_PENDING")
        self.assertEqual(cash_warning["drilldowns"]["section_view"]["params"]["section"], "6.1")
        self.assertNotIn("related_report", cash_warning["drilldowns"])
