# reports/views_stock_ledger_export.py
# ✅ Corrected exports to use the SAME common compute_stock_ledger() output (including fixed OUT unit_cost logic)
# ✅ Safer number conversions (no float() on Decimal/""/None issues)
# ✅ Better totals formatting + alignment
# ✅ Common helpers reused by both Excel + PDF

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus.tables import LongTable
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reports.serializers import StockLedgerRequestSerializer
from reports.services.stock_ledger_service import compute_stock_ledger

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# -----------------------------
# Common helpers
# -----------------------------
def _safe_decimal(v, default=Decimal("0")) -> Decimal:
    if v is None:
        return default
    if isinstance(v, Decimal):
        return v
    try:
        s = str(v).strip()
        if s == "":
            return default
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return default


def _safe_float(v, default=0.0) -> float:
    # Excel likes native numbers; keep it robust
    d = _safe_decimal(v, default=Decimal(str(default)))
    try:
        return float(d)
    except Exception:
        return default


def _fmt_date(d):
    # rows["entrydate"] might be date object or string depending on your service
    if not d:
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _set_cell(ws, row, col, value, bold=False, size=None, align="left", fill=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size) if (bold or size) else Font()
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        c.fill = fill
    if border:
        c.border = BORDER
    return c


def _build_filename(ext: str, data: dict) -> str:
    return f"stock_ledger_{data['product']}_{data['from_date']}_to_{data['to_date']}.{ext}"


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.grey)
    # right side in landscape A4
    canvas.drawRightString(285 * mm, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


# -----------------------------
# Excel Export
# -----------------------------
class StockLedgerExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockLedgerRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        data = compute_stock_ledger(p)  # ✅ must match the same service as API (includes OUT unit_cost fix)
        rows = data.get("results", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Ledger"

        header_fill = PatternFill("solid", fgColor="EDEDED")

        # Title
        _set_cell(ws, 1, 1, "Stock Ledger", bold=True, size=14, border=False)
        ws.merge_cells("A1:K1")

        # Meta
        _set_cell(ws, 2, 1, f"Entity: {data.get('entity_name','')} (#{data.get('entity')})", bold=True, border=False)
        ws.merge_cells("A2:K2")
        _set_cell(ws, 3, 1, f"Product: {data.get('product_name','')} (#{data.get('product')})", bold=True, border=False)
        ws.merge_cells("A3:K3")

        loc_txt = "ALL" if data.get("location") is None else str(data.get("location"))
        _set_cell(
            ws, 4, 1,
            f"Location: {loc_txt}    Period: {data.get('from_date')} to {data.get('to_date')}    Method: {data.get('valuation_method')}",
            border=False
        )
        ws.merge_cells("A4:K4")

        # Opening/Closing
        opening = data.get("opening", {"qty": "0", "value": "0"})
        closing = data.get("closing", {"qty": "0", "value": "0"})

        _set_cell(ws, 5, 1, "Opening Qty", bold=True)
        _set_cell(ws, 5, 2, _safe_float(opening.get("qty")))
        _set_cell(ws, 5, 4, "Opening Value", bold=True)
        _set_cell(ws, 5, 5, _safe_float(opening.get("value")))

        _set_cell(ws, 6, 1, "Closing Qty", bold=True)
        _set_cell(ws, 6, 2, _safe_float(closing.get("qty")))
        _set_cell(ws, 6, 4, "Closing Value", bold=True)
        _set_cell(ws, 6, 5, _safe_float(closing.get("value")))

        # Table header
        headers = [
            "Date", "Txn Type", "Txn ID", "Detail ID", "Voucher No",
            "Qty In", "Qty Out", "Unit Cost", "Amount", "Balance Qty", "Balance Value"
        ]
        start_row = 8
        for col, h in enumerate(headers, 1):
            _set_cell(ws, start_row, col, h, bold=True, align="center", fill=header_fill)

        # Freeze panes below header
        ws.freeze_panes = ws["A9"]

        # Data rows
        r = start_row + 1
        for row in rows:
            values = [
                _fmt_date(row.get("entrydate")),
                row.get("transactiontype") or "",
                row.get("transactionid") or "",
                row.get("detailid") or "",
                row.get("voucherno") or "",
                _safe_float(row.get("qty_in")),
                _safe_float(row.get("qty_out")),
                _safe_float(row.get("unit_cost")),
                _safe_float(row.get("amount")),
                _safe_float(row.get("balance_qty")),
                _safe_float(row.get("balance_value")),
            ]

            for col, v in enumerate(values, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = BORDER
                c.alignment = Alignment(horizontal="right" if col >= 6 else "left", vertical="center")
            r += 1

        last_data_row = r - 1

        # Auto filter
        ws.auto_filter.ref = f"A{start_row}:K{last_data_row}"

        # Totals row (after one blank line)
        r += 1
        _set_cell(ws, r, 1, "TOTALS", bold=True)
        ws.merge_cells(f"A{r}:E{r}")

        totals = data.get("totals", {})
        ws.cell(r, 6, _safe_float(totals.get("total_in_qty"))).font = Font(bold=True)
        ws.cell(r, 7, _safe_float(totals.get("total_out_qty"))).font = Font(bold=True)
        ws.cell(r, 9, _safe_float(totals.get("total_amount"))).font = Font(bold=True)

        for col in range(1, 12):
            ws.cell(r, col).border = BORDER
            ws.cell(r, col).alignment = Alignment(horizontal="right" if col >= 6 else "left", vertical="center")

        # Column widths
        col_widths = [12, 10, 10, 10, 18, 10, 10, 12, 12, 12, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Number formats (apply to data rows + totals row)
        for rr in range(start_row + 1, r + 1):
            ws.cell(rr, 6).number_format = "0.0000"
            ws.cell(rr, 7).number_format = "0.0000"
            ws.cell(rr, 8).number_format = "0.0000"
            ws.cell(rr, 9).number_format = "#,##0.00"
            ws.cell(rr, 10).number_format = "0.0000"
            ws.cell(rr, 11).number_format = "#,##0.00"

        # Signature block
        r += 3
        _set_cell(ws, r, 1, "Prepared By:", bold=True, border=False)
        _set_cell(ws, r, 4, "Checked By:", bold=True, border=False)
        _set_cell(ws, r, 7, "Approved By:", bold=True, border=False)
        r += 2
        ws.merge_cells(f"A{r}:C{r}")
        ws.merge_cells(f"D{r}:F{r}")
        ws.merge_cells(f"G{r}:K{r}")
        ws.cell(r, 1, "_______________________")
        ws.cell(r, 4, "_______________________")
        ws.cell(r, 7, "_______________________")

        r += 2
        _set_cell(ws, r, 1, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", border=False)
        ws.merge_cells(f"A{r}:K{r}")

        # Output
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = _build_filename("xlsx", data)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# -----------------------------
# PDF Export
# -----------------------------
class StockLedgerPDFAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockLedgerRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        data = compute_stock_ledger(p)  # ✅ same service, same OUT unit_cost fix
        rows = data.get("results", [])

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=14 * mm,
        )

        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Stock Ledger", styles["Title"]))
        story.append(Paragraph(f"Entity: {data.get('entity_name','')} (#{data.get('entity')})", styles["Normal"]))
        story.append(Paragraph(f"Product: {data.get('product_name','')} (#{data.get('product')})", styles["Normal"]))
        loc_txt = "ALL" if data.get("location") is None else str(data.get("location"))
        story.append(Paragraph(f"Location: {loc_txt}", styles["Normal"]))
        story.append(
            Paragraph(
                f"Period: {data.get('from_date')} to {data.get('to_date')} &nbsp;&nbsp;&nbsp; Method: {data.get('valuation_method')}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 6))

        opening = data.get("opening", {"qty": "0", "value": "0"})
        closing = data.get("closing", {"qty": "0", "value": "0"})
        story.append(
            Paragraph(
                f"Opening Qty: {opening.get('qty')} &nbsp;&nbsp; Opening Value: {opening.get('value')}",
                styles["Normal"],
            )
        )
        story.append(
            Paragraph(
                f"Closing Qty: {closing.get('qty')} &nbsp;&nbsp; Closing Value: {closing.get('value')}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 8))

        header = [
            "Date", "Txn", "TxnID", "DtlID", "Voucher",
            "Qty In", "Qty Out", "Unit Cost", "Amount", "Bal Qty", "Bal Value",
        ]
        table_data = [header]

        for rr in rows:
            table_data.append([
                _fmt_date(rr.get("entrydate")),
                rr.get("transactiontype") or "",
                str(rr.get("transactionid") or ""),
                str(rr.get("detailid") or ""),
                rr.get("voucherno") or "",
                str(rr.get("qty_in") or "0"),
                str(rr.get("qty_out") or "0"),
                str(rr.get("unit_cost") or "0"),
                str(rr.get("amount") or "0"),
                str(rr.get("balance_qty") or "0"),
                str(rr.get("balance_value") or "0"),
            ])

        totals = data.get("totals", {})
        table_data.append([
            "", "", "", "", "TOTALS",
            str(totals.get("total_in_qty") or "0"),
            str(totals.get("total_out_qty") or "0"),
            "",
            str(totals.get("total_amount") or "0"),
            "",
            "",
        ])

        # Landscape widths tuned
        col_widths = [20*mm, 12*mm, 16*mm, 14*mm, 30*mm, 18*mm, 18*mm, 20*mm, 22*mm, 20*mm, 24*mm]

        tbl = LongTable(table_data, repeatRows=1, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (5, 1), (-1, -2), "RIGHT"),  # numeric columns, exclude header and totals style override below

            # Totals row (last row)
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
            ("ALIGN", (5, -1), (-1, -1), "RIGHT"),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 12))

        # Signature section
        sig_tbl = Table([
            ["Prepared By", "Checked By", "Approved By"],
            ["__________________________", "__________________________", "__________________________"],
            ["Date", "Date", "Date"],
            ["__________________________", "__________________________", "__________________________"],
        ], colWidths=[90*mm, 90*mm, 90*mm])

        sig_tbl.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(sig_tbl)

        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))

        doc.build(story, onFirstPage=_footer, onLaterPages=_footer)

        buf.seek(0)
        filename = _build_filename("pdf", data)
        response = HttpResponse(buf.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
