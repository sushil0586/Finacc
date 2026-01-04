from django.shortcuts import render
from decimal import Decimal, ROUND_HALF_UP
from rest_framework.exceptions import ValidationError

# Create your views here.
from collections import defaultdict
from decimal import Decimal as D
from typing import Any, Dict, List
from math import isfinite
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.units import inch
from django.http import HttpResponse
from rest_framework.generics import GenericAPIView
from rest_framework.request import Request
from rest_framework.test import APIRequestFactory
import re
from django.db.models import Min, Max
from django.db.models.expressions import Window
from invoice.models import JournalLine  # <-- your GL table
from typing import Tuple, List,Optional
from invoice.serializers import stocktransconstant
from .pagination import SmallPageNumberPagination,SimpleNumberPagination
from django.db.models.functions import TruncDay, TruncMonth, TruncQuarter, TruncYear

from reports.serializers import StockSummaryRequestSerializer, StockSummaryRowSerializer
from invoice.models import InventoryMove  # adjust path


from itertools import product
from django.http import request,JsonResponse
from django.shortcuts import render
import json
from pandas.tseries.offsets import MonthEnd,QuarterEnd
from decimal import Decimal,InvalidOperation
from django.db.models import Sum, Q, Value as V,DecimalField,ExpressionWrapper,Count
from django.utils.timezone import make_aware, is_aware
from .serializers import CashbookUnifiedSerializer,TrialBalanceAccountRowSerializer,TrialBalanceAccountLedgerRowSerializer
from django.utils.dateparse import parse_date
from collections import deque
from django.utils import timezone  
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side,NamedStyle,numbers
from openpyxl.utils import get_column_letter
from rest_framework.generics import CreateAPIView,ListAPIView,ListCreateAPIView,RetrieveUpdateDestroyAPIView,GenericAPIView,RetrieveAPIView,UpdateAPIView
from invoice.models import StockTransactions,closingstock,salesOrderdetails,entry,SalesOderHeader,PurchaseReturn,purchaseorder,salereturn,journalmain,salereturnDetails,Purchasereturndetails
# from invoice.serializers import SalesOderHeaderSerializer,salesOrderdetailsSerializer,purchaseorderSerializer,PurchaseOrderDetailsSerializer,POSerializer,SOSerializer,journalSerializer,SRSerializer,salesreturnSerializer,salesreturnDetailsSerializer,JournalVSerializer,PurchasereturnSerializer,\
# purchasereturndetailsSerializer,PRSerializer,TrialbalanceSerializer,TrialbalanceSerializerbyaccounthead,TrialbalanceSerializerbyaccount,accountheadserializer,accountHead,accountserializer,accounthserializer, stocktranserilaizer,cashserializer,journalmainSerializer,stockdetailsSerializer,stockmainSerializer,\
# PRSerializer,SRSerializer,stockVSerializer,stockserializer,Purchasebyaccountserializer,Salebyaccountserializer,entitySerializer1,cbserializer,ledgerserializer,ledgersummaryserializer,stockledgersummaryserializer,stockledgerbookserializer,balancesheetserializer,gstr1b2bserializer,gstr1hsnserializer,\
# purchasetaxtypeserializer,tdsmainSerializer,tdsVSerializer,tdstypeSerializer,tdsmaincancelSerializer,salesordercancelSerializer,purchaseordercancelSerializer,purchasereturncancelSerializer,salesreturncancelSerializer,journalcancelSerializer,stockcancelSerializer,SalesOderHeaderpdfSerializer,productionmainSerializer,productionVSerializer,productioncancelSerializer,tdsreturnSerializer,gstorderservicesSerializer,SSSerializer,gstorderservicecancelSerializer,jobworkchallancancelSerializer,JwvoucherSerializer,jobworkchallanSerializer,debitcreditnoteSerializer,dcnoSerializer,debitcreditcancelSerializer,closingstockSerializer

from reports.serializers import closingstockSerializer,stockledgerbookserializer,stockledgersummaryserializer,ledgerserializer,cbserializer,stockserializer,cashserializer,accountListSerializer2,ledgerdetailsSerializer,ledgersummarySerializer,stockledgerdetailSerializer,stockledgersummarySerializer,TrialBalanceSerializer,StockDayBookSerializer,StockSummarySerializerList,SalesGSTSummarySerializer,LedgerSummaryRequestSerializer,LedgerSummaryRowSerializer,NegativeValuationPolicy
from rest_framework import permissions,status
from django_filters.rest_framework import DjangoFilterBackend
from django.db import DatabaseError, transaction
from rest_framework.response import Response
from django.db.models import Sum,OuterRef,Subquery,F,Sum, F, Value, DecimalField,CharField
from django.db.models import Prefetch
from financial.models import account,accountHead
from inventory.models import Product
from django.db import connection
from django.core import serializers
from rest_framework.renderers import JSONRenderer
from drf_excel.mixins import XLSXFileMixin
from drf_excel.renderers import XLSXRenderer
from rest_framework.viewsets import ReadOnlyModelViewSet
from entity.models import Entity,entityconstitution,entityfinancialyear,entityfinancialyear as FinancialYear
from django_pandas.io import read_frame
from django.db.models import Q
import numpy as np
import pandas as pd
from decimal import Decimal
from datetime import timedelta,date,datetime
import pytz
import json
from django.db.models import Case, When
from django.db.models import Q, Sum, F
from django.db.models.functions import Coalesce
from django_pandas.io import read_frame
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from entity.models import Entity, entityfinancialyear
from financial.models import account
from inventory.models import Product
import pandas as pd
from datetime import datetime
from reports.models import TransactionType
from .serializers import TransactionTypeSerializer
from .serializers import EMICalculatorSerializer
from helpers.utils.emi import calculate_emi
from reports.services.trading_account import build_trading_account_dynamic
from reports.services.profit_and_loss import build_profit_and_loss_statement
from reports.services.balance_sheet import build_balance_sheet_statement
from reports.serializers import TrialBalanceHeadRowSerializer, LedgerFilterSerializer, LedgerAccountSerializer,DaybookUnifiedSerializer
ZERO = Decimal("0.00")

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import A4,landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,KeepTogether,PageTemplate,BaseDocTemplate,Frame,LongTable
from io import BytesIO
from django.utils.timezone import now



class closingstockBalance(ListAPIView):

    #serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if (len(dfg) > 0):
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg


        puchases = StockTransactions.objects.filter(Q(isactive =1,stockttype__in = ['P','O','R'],accounttype = 'DD',entity = entity1,entrydatetime__lte = enddate)).values('stock','stockttype','quantity','entrydatetime','stock__id','id')
        sales = StockTransactions.objects.filter(isactive =1,stockttype__in = ['S','I'],accounttype = 'DD',entity = entity1,entrydatetime__lte = enddate).values('stock','stockttype','quantity','entrydatetime','stock__id','id')
        inventory = puchases.union(sales).order_by('entrydatetime')
        closingprice = closingstock.objects.filter(entity = entity1).values('stock__id','closingrate')
        #idf1 = read_frame(puchases)
       # print(idf1)
        idf = read_frame(inventory)
        #print(idf)
        cdf = read_frame(closingprice)
        cdf['closingrate'] = cdf['closingrate'].astype(float).fillna(0)
      #  print(cdf)
        

        idf = pd.merge(idf,cdf,on='stock__id',how='outer',indicator=True)
       # print(idf)

        idf['closingrate'] = idf['closingrate'].astype(float).fillna(0)

        #print



       # print(idf)
        idf['stockttype'] = np.where(idf['stockttype'].isin(['P','R','O']),'P','S')
        idf['quantity'] = np.where(idf['stockttype'].isin(['P','O','R']), idf['quantity'],-1 * (idf['quantity']))
        #print(idf)
        idf['quantity'] = idf['quantity'].astype(float)
        idf['CS'] = idf.groupby(['stock','stockttype'])['quantity'].cumsum()

      #  dfR['closingrate'] = dfR['closingrate'].astype(float).fillna(0)
       # dfR['rate'] = dfR['rate'].astype(float).fillna(0)
       # print(idf)
        dfR = idf.groupby(['stock'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)
        #print(dfR)

       # dfR['rate'] = dfR['rate'].fillna(0)

        #dfR['closingrate'] = np.where(dfR['closingrate'] >0,dfR['closingrate'],dfR['rate'])
        dfR['balance'] = dfR['quantity'].astype(float)  * dfR['closingrate'].astype(float)
       # print(dfR)
        dfR = dfR.drop(['_merge','stockttype','entrydatetime','id'],axis=1) 

        #print(dfR)

        #dfi = dfi.drop(['account__id','transactiontype','entrydatetime','account__accountname'],axis=1) 

        dfR.rename(columns = {'stock':'stockname'}, inplace = True)

        dfR.rename(columns = {'stock__id':'stock', 'closingrate':'rate'}, inplace = True)



       


        print(dfR)

        dfR = dfR.groupby(['stockname','stock','rate'])[['quantity','balance']].sum().abs().reset_index()

        return Response(dfR.T.to_dict().values())



class closingstocknew(ListAPIView):

    #serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate =   self.request.query_params.get('enddate')

        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if (len(dfg) > 0):
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg


        puchases = StockTransactions.objects.filter(Q(isactive =1,stockttype__in = ['P','O','R'],accounttype = 'DD',entity = entity1,entrydatetime__lte = enddate)).values('stock','rate','stockttype','quantity','entrydatetime','stock__id','id')
        sales = StockTransactions.objects.filter(isactive =1,stockttype__in = ['S','I'],accounttype = 'DD',entity = entity1,entrydatetime__lte = enddate).values('stock','rate','stockttype','quantity','entrydatetime','stock__id','id')
        inventory = puchases.union(sales).order_by('entrydatetime')
        closingprice = closingstock.objects.filter(entity = entity1).values('stock__id','closingrate')
        #idf1 = read_frame(puchases)
       # print(idf1)
        idf = read_frame(inventory)
      #  print(idf)
        cdf = read_frame(closingprice)
        cdf['closingrate'] = cdf['closingrate'].astype(float).fillna(0)
        print(cdf)
        

        idf = pd.merge(idf,cdf,on='stock__id',how='outer',indicator=True)
        print(idf)

        idf['closingrate'] = idf['closingrate'].astype(float).fillna(0)

        #print



       # print(idf)
        idf['stockttype'] = np.where(idf['stockttype'].isin(['P','R','O']),'P','S')
        idf['quantity'] = np.where(idf['stockttype'].isin(['P','O','R']), idf['quantity'],-1 * (idf['quantity']))
        #print(idf)
        idf['quantity'] = idf['quantity'].astype(float)
        idf['CS'] = idf.groupby(['stock','stockttype'])['quantity'].cumsum()

      #  dfR['closingrate'] = dfR['closingrate'].astype(float).fillna(0)
       # dfR['rate'] = dfR['rate'].astype(float).fillna(0)
       # print(idf)
        dfR = idf.groupby(['stock'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)
       # print(dfR)

        dfR['rate'] = dfR['rate'].fillna(0)

        #dfR['closingrate'] = np.where(dfR['closingrate'] >0,dfR['closingrate'],dfR['rate'])
        dfR['balance'] = dfR['quantity'].astype(float)  * dfR['closingrate'].astype(float)
       # print(dfR)
        dfR = dfR.drop(['rate','_merge','stockttype','entrydatetime','id'],axis=1) 

        #dfi = dfi.drop(['account__id','transactiontype','entrydatetime','account__accountname'],axis=1) 

        dfR.rename(columns = {'stock__id':'stockid', 'stock':'stockname'}, inplace = True)


        dfR = dfR.groupby(['stockname','stockid','closingrate'])[['quantity','balance']].sum().abs().reset_index()

       


        print(dfR)

        return Response(dfR.T.to_dict().values())
    


class closingstockView(ListCreateAPIView):  
         
        serializer_class = closingstockSerializer  


        # def perform_create(self, serializer):
        #     entity = self.request.query_params.get('entity')
        #     closingstock.objects.filter(entity = entity).delete()


        #     return serializer.save(createdby = self.request.user)
      
        def create(self, request, *args, **kwargs):  
            entity = self.request.query_params.get('entity')
            closingstock.objects.filter(entity = entity).delete()

            print(request.data)
          # print(entity)
            serializer = self.get_serializer(data=request.data, many=True)  
            serializer.is_valid(raise_exception=True)  
      
            try:  
                self.perform_create(serializer)  
                return Response(serializer.data, status=status.HTTP_201_CREATED)  
            except:  
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST) 
            
        def get_queryset(self):
            entity = self.request.query_params.get('entity')
            return   closingstock.objects.filter(entity = entity)
        
class dashboardkpis(ListAPIView):

   # serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        stk =StockTransactions.objects.filter(Q(isactive = 1),Q(entity = entity1)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__gte = 0)
        stk2 =StockTransactions.objects.filter(Q(isactive = 1),Q(entity = entity1)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__lte = 0)



        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if (len(dfg) > 0):
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg


        puchases = StockTransactions.objects.filter(Q(isactive =1),Q(transactiontype__in = ['P','OS']),Q(accounttype = 'DD'),Q(entity = entity1)).values('account__accounthead__name','account__accounthead','account__id','account__accountname','stock','rate','transactiontype','quantity','entrydatetime')
        sales = StockTransactions.objects.filter(isactive =1,transactiontype = 'S',accounttype = 'DD',entity = entity1).values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname','stock','rate','transactiontype','quantity','entrydatetime')

        inventory = puchases.union(sales).order_by('entrydatetime')

        #idf1 = read_frame(puchases)

       # print(idf1)

        idf = read_frame(inventory)


        idf['transactiontype'] = np.where(idf['transactiontype'] == 'OS','P',idf['transactiontype'])


        

        idf['quantity'] = np.where(idf['transactiontype'].isin(['P','OS']), idf['quantity'],-1 * (idf['quantity']))

        #print(idf)

        idf['quantity'] = idf['quantity'].astype(float)
        idf['CS'] = idf.groupby(['stock','transactiontype'])['quantity'].cumsum()

        #print(idf)



        dfR = idf.groupby(['stock'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)

       # print(dfR)


        dfR['balance'] = dfR['quantity'].astype(float) * -1 * dfR['rate'].astype(float)

        dfR = dfR.drop(['stock','transactiontype','entrydatetime'],axis=1) 

        dfR['account__accounthead__name'] = 'Closing Stock'
        dfR['account__accounthead'] = -1

       # print(dfR)




    #     print(stk2.query.__str__())
    #    # q = stk.filter(balance__gt=0)



        stkunion = stk.union(stk2)

        df = read_frame(stkunion)
        df = df.drop(['debit','credit'],axis=1)

       # print(df)

        

        
        #print(df)

        #print(df)

        frames = [df, dfR]

        df = pd.concat(frames)

       # print(df)

        df['balance'] = df['balance'].astype(float)

        pl1 =StockTransactions.objects.filter(Q(isactive = 1)).filter(account__accounthead__detailsingroup = 2,entity = entity1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__gte = 0)
        pl2 =StockTransactions.objects.filter(Q(isactive = 1)).filter(account__accounthead__detailsingroup = 2,entity = entity1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__lte = 0)

        plunion = pl1.union(pl2)

        pldf = read_frame(plunion)
        pldf = pldf.drop(['debit','credit'],axis=1)


        if df['balance'].sum() < 0:
            pldf.loc[len(pldf.index)] = ['Gross Profit', -1, -1, 'Gross Profit',df['balance'].sum()]
        else:
            pldf.loc[len(pldf.index)] = ['Gross Loss', -1, -1, 'Gross Loss',df['balance'].sum()]


        #print(pldf)

        pldf['balance'] = pldf['balance'].astype(float)

        if pldf['balance'].sum() < 0:
            pldf.loc[len(pldf.index)] = ['Net Profit', -2, -2, 'Net Profit',-pldf['balance'].sum()]
        else:
            pldf.loc[len(pldf.index)] = ['Net Loss', -2, -2, 'Net Loss',-pldf['balance'].sum()]



        pldf['drcr'] = pldf['balance'].apply(lambda x: 0 if x > 0 else 1)
        #df = df.loc['Column_Total']= df.sum(numeric_only=True, axis=0)


        frames = [df, pldf]

        df = pd.concat(frames)

        #print(df)

        

       

        

        

        df.rename(columns = {'account__accounthead__name':'accountheadname', 'account__accounthead':'accounthead','account__accountname':'accountname','account__id':'accountid'}, inplace = True)


       # print(df)

        df2=df[df['accountheadname'].astype(str).str.contains("Purchase|Sale|Gross Profit|Net Profit")]


        #print(df.groupby(['accounthead','accountheadname','drcr','accountname','accountid'])[['balance']].sum().abs())

        #return Response(df2)

        df1 = pd.DataFrame(
       {
        "month": ["May", "June", "July", "August", "Sep", "Oct"],
        "sales": [15000, 7000, 19000,25000,12000,4000],
         "purchases": [18000, 14000, 35000,38000,20000,6000],
         "NetProfit": [3000, 7000, 14000,13000,8000,2000],
            },
            
        )

        return Response(df2.groupby(['accounthead','accountheadname','accountname','accountid'])[['balance']].sum().abs().reset_index().T.to_dict().values())
    

class dashboardgraphkpis(ListAPIView):

    #serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        stk =StockTransactions.objects.filter(Q(isactive = 1),Q(entity = entity1)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__gte = 0)
        stk2 =StockTransactions.objects.filter(Q(isactive = 1),Q(entity = entity1)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__lte = 0)



        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if (len(dfg) > 0):
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg


        puchases = StockTransactions.objects.filter(Q(isactive =1),Q(transactiontype__in = ['P','OS']),Q(accounttype = 'DD'),Q(entity = entity1)).values('account__accounthead__name','account__accounthead','account__id','account__accountname','stock','rate','transactiontype','quantity','entrydatetime')
        sales = StockTransactions.objects.filter(isactive =1,transactiontype = 'S',accounttype = 'DD',entity = entity1).values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname','stock','rate','transactiontype','quantity','entrydatetime')

        inventory = puchases.union(sales).order_by('entrydatetime')

        #idf1 = read_frame(puchases)

       # print(idf1)

        idf = read_frame(inventory)


        idf['transactiontype'] = np.where(idf['transactiontype'] == 'OS','P',idf['transactiontype'])


        

        idf['quantity'] = np.where(idf['transactiontype'].isin(['P','OS']), idf['quantity'],-1 * (idf['quantity']))

        #print(idf)

        idf['quantity'] = idf['quantity'].astype(float)
        idf['CS'] = idf.groupby(['stock','transactiontype'])['quantity'].cumsum()

        #print(idf)



        dfR = idf.groupby(['stock'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)

       # print(dfR)


        dfR['balance'] = dfR['quantity'].astype(float) * -1 * dfR['rate'].astype(float)

        dfR = dfR.drop(['stock','transactiontype','entrydatetime'],axis=1) 

        dfR['account__accounthead__name'] = 'Closing Stock'
        dfR['account__accounthead'] = -1

       # print(dfR)




    #     print(stk2.query.__str__())
    #    # q = stk.filter(balance__gt=0)



        stkunion = stk.union(stk2)

        df = read_frame(stkunion)
        df = df.drop(['debit','credit'],axis=1)

       # print(df)

        

        
        #print(df)

        #print(df)

        frames = [df, dfR]

        df = pd.concat(frames)

       # print(df)

        df['balance'] = df['balance'].astype(float)

        pl1 =StockTransactions.objects.filter(Q(isactive = 1)).filter(account__accounthead__detailsingroup = 2,entity = entity1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__gte = 0)
        pl2 =StockTransactions.objects.filter(Q(isactive = 1)).filter(account__accounthead__detailsingroup = 2,entity = entity1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__lte = 0)

        plunion = pl1.union(pl2)

        pldf = read_frame(plunion)
        pldf = pldf.drop(['debit','credit'],axis=1)


        if df['balance'].sum() < 0:
            pldf.loc[len(pldf.index)] = ['Gross Profit', -1, -1, 'Gross Profit',df['balance'].sum()]
        else:
            pldf.loc[len(pldf.index)] = ['Gross Loss', -1, -1, 'Gross Loss',df['balance'].sum()]


        #print(pldf)

        pldf['balance'] = pldf['balance'].astype(float)

        if pldf['balance'].sum() < 0:
            pldf.loc[len(pldf.index)] = ['Net Profit', -2, -2, 'Net Profit',-pldf['balance'].sum()]
        else:
            pldf.loc[len(pldf.index)] = ['Net Loss', -2, -2, 'Net Loss',-pldf['balance'].sum()]



        pldf['drcr'] = pldf['balance'].apply(lambda x: 0 if x > 0 else 1)
        #df = df.loc['Column_Total']= df.sum(numeric_only=True, axis=0)


        frames = [df, pldf]

        df = pd.concat(frames)

        #print(df)

        

       

        

        

        df.rename(columns = {'account__accounthead__name':'accountheadname', 'account__accounthead':'accounthead','account__accountname':'accountname','account__id':'accountid'}, inplace = True)


       # print(df)

        df2=df[df['accountheadname'].astype(str).str.contains("Purchase|Sale|Gross Profit|Net Profit")]


        #print(df.groupby(['accounthead','accountheadname','drcr','accountname','accountid'])[['balance']].sum().abs())

        #return Response(df2)

        df1 = pd.DataFrame(
       {
        "month": ["May", "June", "July", "August", "Sep", "Oct"],
        "sales": [15000, 7000, 19000,25000,12000,4000],
         "purchases": [18000, 14000, 35000,38000,20000,6000],
         "netprofit": [3000, 7000, 14000,13000,8000,2000],
            },
            
        )

        return Response(df1)


class gstr3b1(ListAPIView):

    #serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        stk =StockTransactions.objects.filter(Q(isactive = 1),Q(entity = entity1)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__gte = 0)
        stk2 =StockTransactions.objects.filter(Q(isactive = 1),Q(entity = entity1)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0) , balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__lte = 0)



        


        #print(df.groupby(['accounthead','accountheadname','drcr','accountname','accountid'])[['balance']].sum().abs())

        #return Response(df2)

        df1 = pd.DataFrame(
       {
        "NatureofSupplies": ["(a) outward taxable supplies(Other then Zero rated ,Nil rated and Execpted)", "(b) Outward taxable supplies (Zero rated)", "(C) Other outward supplies", "(d) Inward supplies", "(e) Non-GST outward supplies"],
        "Totaltaxablesupplies": [15000, 7000, 19000,25000,12000],
         "IntegartedTax": [18000, 14000, 35000,38000,20000],
         "CentralTax": [3000, 7000, 14000,13000,8000],
         "StateUTTax": [3000, 7000, 14000,13000,8000],
         "Cess": [3000, 7000, 14000,13000,8000],
            },
            
        )


        df2 = pd.DataFrame(
       {
        "NatureofSupplies": [" Supplies made to unregistered persons", "Supplies made to composition taxable persons", "Supplies made to UIN holders"],
        "placeofsupplies": ["Punjab", "Haryana", "chandigarh"],
         "Totaltaxablevalue": [18000, 14000, 35000],
         "integratedtax": [3000, 7000, 14000],
        
            },
            
        )


        df3 = pd.DataFrame(
       {
        "Details": ["(A) ITC Available(whether in full or part)", "(1) Import of goods", "(2) Import of services","(3) Import supplies liable to reverse chnarge(Other than 1 & 2 above)", "(4) Inwrad supplies from ISD","(5) All other ITC","(B) ITC Reversed","(1) As per rules 42 & 43 of CGST rules","(2) Others","(C) Net ITC available(A) - (B)","(D) Ineligible ITC","(1) AS Per section 17(5)","(2) Others"],
         "integratedtax": [18000, 14000, 35000,18000, 14000, 35000,18000, 14000, 35000,35000,1200,1200,1200],
         "CentralTax": [18000, 14000, 35000,18000, 14000, 35000,18000, 14000, 35000,3500,1200,1200,1200],
         "StateUTTax": [18000, 14000, 35000,18000, 14000, 35000,18000, 14000, 35000,3500,1200,1200,1200],
         "Cess": [18000, 14000, 35000,18000, 14000, 35000,18000, 14000, 35000,3500,1200,1200,1200],
         "Isbold": [True, False, False,False, False, False,True,False, False,True,True,False,False],
        
            },
            
        )


        df4 = pd.DataFrame(
       {
        "natureofsupplies": ["From a supplier under composition scheme, Exempt and Nil rated supplies", "Non GST Supply"],
         "interstatesupplies": [18000, 14000],
         "intrastatesupplies": [18000, 14000],
        },
            
        )

        df5 = pd.DataFrame(
       {
        "description": ["Integrated Tax", "Central Tax","State/UT Tax","Cess"],
        "Taxpayable": [18000, 14000,18000, 14000],
        "integratedtax": [18000, 14000,18000, 14000],
        "centraltax": [18000, 14000,18000, 14000],
        "stateuttax": [18000, 14000,18000, 14000],
        "cess": [18000, 14000,18000, 14000],
        "integratedtax": [18000, 14000,18000, 14000],
        "taxpaidtdstcs" : [18000, 14000,18000, 14000],
        "taxcesspaidincash": [18000, 14000,18000, 14000],
        "interest": [18000, 14000,18000, 14000],
        "latefee": [18000, 14000,18000, 14000],
        },
            
        )

        df6 = pd.DataFrame(
       {
        "details": ["TDS", "TCS"],
        "integratedtax": [18000, 14000],
        "centraltax": [18000, 14000],
        "statetax": [18000, 14000],
        },
            
        )

        #print(json.loads(df1.T.to_json()))


        headers = json.dumps({ 

                              'Year'  : 2023,
                              'Month' : 12,
                              'entity' : 'Reliance industries',
                              'GSTIN' : '03APXPB5894F',
                              
                              'table1':json.loads(df1.to_json(orient='records')),
                              'table2':json.loads(df2.to_json(orient='records')),
                              'table3':json.loads(df3.to_json(orient='records')),
                              'table4':json.loads(df4.to_json(orient='records')),
                              'table5':json.loads(df5.to_json(orient='records')),
                              'table6':json.loads(df6.to_json(orient='records')),
                               })

        # allDays=[]
        # allDays.append(df1.T.to_dict().values())
        # allDays.append(df2.T.to_dict().values())

        return JsonResponse(json.loads(headers), safe = False)

        #return JsonResponse(json.loads(df1.T.to_json()), safe = False)

       # return Response(df1.T.to_dict().values())
    

    

class generalfunctions:

    def __init__(self, entityid,startdate,enddate):
        self.entityid = entityid
        self.startdate = startdate
        self.enddate = enddate

    
        

    def getstockdetails(self):
        stk =StockTransactions.objects.filter(isactive = 1,entity = self.entityid,entrydatetime__range=(self.startdate, self.enddate)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__gte = 0)
        stk2 =StockTransactions.objects.filter(isactive = 1,entity = self.entityid,entrydatetime__range=(self.startdate, self.enddate)).filter(account__accounthead__detailsingroup = 1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0)).filter(balance__lte = 0)
        stkunion = stk.union(stk2)
        df = read_frame(stkunion)
      #  df = df.drop(['debit','credit'],axis=1)
        return df

    def getinventorydetails(self):
        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if len(dfg) > 0:
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg

        utc = pytz.UTC

        currentdates = entityfinancialyear.objects.get(
            entity=self.entityid,
            finendyear__gte=self.startdate,
            finstartyear__lte=self.startdate
        )

        # Optimize filtering logic
        is_active = currentdates.isactive == 1
        enddate_filtered = utc.localize(datetime.strptime(self.enddate, '%Y-%m-%d')) >= currentdates.finendyear

        purchase_filters = {
            'isactive': 1,
            'stockttype__in': ['P', 'R', 'O'],
            'accounttype': 'DD',
            'entity': self.entityid,
            'entrydatetime__lte': self.enddate
        }

        sales_filters = {
            'isactive': 1,
            'stockttype__in': ['S', 'I'],
            'accounttype': 'DD',
            'entity': self.entityid,
            'entrydatetime__lte': self.enddate
        }

        # Apply filters
        purchases = StockTransactions.objects.filter(**purchase_filters)
        sales = StockTransactions.objects.filter(**sales_filters)

        if is_active:
            purchases = purchases.values('stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate', 'id')
            sales = sales.values('stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate', 'id')
        elif enddate_filtered:
            purchases = purchases.exclude(account__accountcode=9000).values('stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate', 'id')
            sales = sales.exclude(isbalancesheet=0).values('stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate', 'id')
        else:
            purchases = purchases.values('stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate', 'id')
            sales = sales.exclude(isbalancesheet=0).values('stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate', 'id')

        inventory = purchases.union(sales).order_by('entrydatetime')

        closingprice = closingstock.objects.filter(entity=self.entityid).values('stock__id', 'closingrate')

        idf = read_frame(inventory)
        cdf = read_frame(closingprice)

        idf = pd.merge(idf, cdf, on='stock__id', how='outer', indicator=True)
        idf['stockttype'] = np.where(idf['stockttype'].isin(['P', 'R', 'O']), 'P', 'S')
        idf['quantity'] = np.where(idf['stockttype'].isin(['P', 'R', 'O']), idf['quantity'], -1 * idf['quantity'])
        idf['quantity'] = idf['quantity'].astype(float).fillna(0)
        idf['CS'] = idf.groupby(['stock__id', 'stockttype'])['quantity'].cumsum()

        dfR = idf.groupby(['stock__id'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)

        return dfR


    
    



    def getinventorydetails_1(self, dfR_1):
        # Convert necessary columns to float once
        dfR_1['closingrate'] = np.where(dfR_1['closingrate'] > 0, dfR_1['closingrate'], dfR_1['rate'])
        
        # Calculate 'balance' efficiently
        dfR_1['quantity'] = dfR_1['quantity'].astype(float)  # Convert once to float for reuse
        dfR_1['closingrate'] = dfR_1['closingrate'].astype(float)  # Ensure it's float
        dfR_1['balance'] = dfR_1['quantity'] * -1 * dfR_1['closingrate']
        
        # Drop unnecessary columns
        dfR_1.drop(['stockttype', 'entrydatetime', '_merge', 'rate', 'id'], axis=1, inplace=True)
        
        # Rename columns
        dfR_1.rename(columns={'stock__id': 'account__id', 'stock': 'account__accountname'}, inplace=True)
        
        # Add static columns
        dfR_1['accounthead__name'] = 'Closing Stock'
        
        # Fetch the account ID only once
        account_id = accountHead.objects.get(code=200, entity=self.entityid).id
        dfR_1['account__accounthead'] = account_id
        
        return dfR_1

    def getinventorydetails_2(self,dfi):
        dfi['balance'] = dfi['quantity'].astype(float) * 1 * dfi['closingrate'].astype(float)
        dfi = dfi.drop(['stockttype','entrydatetime','_merge'],axis=1) 
        dfi.rename(columns = {'stock__id':'account__id', 'stock':'account__accountname'}, inplace = True)
        dfi['account__accounthead__name'] = 'Closing Stock'

        account_id = accountHead.objects.get(code = 200).id
        dfi['account__accounthead'] = account_id
        return dfi

    def openinginventorydetails(self):
        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if (len(dfg) > 0):
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg
        

        opuchases = StockTransactions.objects.filter(isactive =1,stockttype__in = ['P','R','O'],accounttype = 'DD',entity = self.entityid,entrydatetime__lt = self.startdate).values('stock','rate','stockttype','quantity','entrydatetime','stock__id')
        osales = StockTransactions.objects.filter(isactive =1,stockttype__in = ['S','I'],accounttype = 'DD',entity = self.entityid,entrydatetime__lt = self.startdate).values('stock','rate','stockttype','quantity','entrydatetime','stock__id')
        oinventory = opuchases.union(osales).order_by('entrydatetime')
        idf = read_frame(oinventory)


       
        closingprice = closingstock.objects.filter(entity = self.entityid).values('stock__id','closingrate')

        cdf = read_frame(closingprice)

        idf = pd.merge(idf,cdf,on='stock__id',how='outer',indicator=True)
        idf['stockttype'] = np.where(idf['stockttype'].isin(['P','R','O']),'P','S')
        idf['quantity'] = np.where(idf['stockttype'].isin(['P','O','R']), idf['quantity'],-1 * (idf['quantity']))
        idf['quantity'] = idf['quantity'].astype(float)
        idf['CS'] = idf.groupby(['stock','stockttype'])['quantity'].cumsum()
        odfR = idf.groupby(['stock'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)
        return odfR

    def openinginventorydetails_1(self,odfi):
        odfi['balance'] = odfi['quantity'].astype(float) * -1 * odfi['rate'].astype(float)
        odfi = odfi.drop(['stockttype','entrydatetime','closingrate','_merge'],axis=1) 
        odfi.rename(columns = {'stock__id':'account__id', 'stock__productname':'account__accountname'}, inplace = True)
        odfi['account__accounthead__name'] = 'Opening Stock'
        account_id = accountHead.objects.get(code = 9000,entity = self.entityid).id
        odfi['account__accounthead'] = account_id
        return odfi

    def openinginventorydetails_2(self,odfR):
        odfR['balance'] = odfR['quantity'].astype(float)  * odfR['rate'].astype(float)
        odfR = odfR.drop(['stock','stockttype','entrydatetime','closingrate','_merge'],axis=1) 
        odfR.rename(columns = {'stock__id':'account__id', 'stock__productname':'account__accountname'}, inplace = True)
        odfR['account__accounthead__name'] = 'Opening Stock'
        account_id = accountHead.objects.get(code = 9000,entity = self.entityid).id
        #dfR_1['account__accounthead'] = account_id


        odfR['account__accounthead'] = account_id
        return odfR


    def getprofitandloss(self):

        utc=pytz.UTC
        
        currentdates = entityfinancialyear.objects.get(entity = self.entityid,finendyear__gte = self.startdate  ,finstartyear__lte =  self.startdate)

        if currentdates.isactive == 1 or utc.localize(datetime.strptime(self.enddate, '%Y-%m-%d')):
            pl1 =StockTransactions.objects.filter(isactive = 1,entity = self.entityid,entrydatetime__range=(self.startdate, self.enddate),account__accounthead__detailsingroup = 2).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(balance = Sum('creditamount',default = 0) - Sum('debitamount',default = 0)).filter(balance__gte = 0)
            pl2 =StockTransactions.objects.filter(isactive = 1,entity = self.entityid,entrydatetime__range=(self.startdate, self.enddate),account__accounthead__detailsingroup = 2).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(balance = Sum('creditamount',default = 0) - Sum('debitamount',default = 0)).filter(balance__lte = 0)
        else:
            pl1 =StockTransactions.objects.filter(isactive = 1,entity = self.entityid,entrydatetime__range=(self.startdate, self.enddate),account__accounthead__detailsingroup = 2,isbalancesheet = 1).exclude(accounttype = 'MD').values('account__accounthead__name','account__accounthead','account__id','account__accountname').annotate(balance = Sum('creditamount',default = 0) - Sum('debitamount',default = 0)).filter(balance__gte = 0)
            pl2 =StockTransactions.objects.filter(isactive = 1,entity = self.entityid,entrydatetime__range=(self.startdate, self.enddate),account__accounthead__detailsingroup = 2,isbalancesheet = 1).exclude(accounttype = 'MD').values('account__creditaccounthead__name','account__creditaccounthead','account__id','account__accountname').annotate(balance = Sum('creditamount',default = 0) - Sum('debitamount',default = 0)).filter(balance__lte = 0)


        plunion = pl1.union(pl2)

        pldf = read_frame(plunion)
       # pldf = pldf.drop(['debit','credit'],axis=1)




        return pldf
    

    def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if (len(dfg) > 0):
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg
    

    def getopeningstockdetails(self):
        def FiFo(dfg):
            if dfg[dfg['CS'] < 0]['quantity'].count():
                subT = dfg[dfg['CS'] < 0]['CS'].iloc[-1]
                dfg['quantity'] = np.where((dfg['CS'] + subT) <= 0, 0, dfg['quantity'])
                dfg = dfg[dfg['quantity'] > 0]
                if len(dfg) > 0:
                    dfg['quantity'].iloc[0] = dfg['CS'].iloc[0] + subT
            return dfg

        # Combined filter for transactions
        filters = {
            'isactive': 1,
            'accounttype': 'DD',
            'entity': self.entityid,
            'entrydatetime__lt': self.startdate
        }

        # Single query for both purchases and sales
        transactions = StockTransactions.objects.filter(**filters).values(
            'stock', 'stockttype', 'quantity', 'entrydatetime', 'stock__id', 'rate'
        )

        # Fetch closing prices in one go
        closingprice = closingstock.objects.filter(entity=self.entityid).values('stock__id', 'closingrate')

        # Union and order the inventory (no need to filter separately for purchases and sales)
        oinventory = transactions.order_by('entrydatetime')

        # Read frames
        cdf = read_frame(closingprice)
        idf = read_frame(oinventory)

        # Merge dataframes
        idf = pd.merge(idf, cdf, on='stock__id', how='outer', indicator=True)

        # Vectorized operations for stock type and quantity adjustments
        idf['stockttype'] = np.where(idf['stockttype'].isin(['P', 'R', 'O']), 'P', 'S')
        idf['quantity'] = np.where(idf['stockttype'].isin(['P', 'O', 'R']), idf['quantity'], -1 * idf['quantity'])
        idf['quantity'] = idf['quantity'].astype(float)
        idf['CS'] = idf.groupby(['stock', 'stockttype'])['quantity'].cumsum()

        # Apply FiFo function and reset index
        odfR = idf.groupby(['stock'], as_index=False).apply(FiFo).drop(['CS'], axis=1).reset_index(drop=True)

        # Update closing rate and calculate balance
        odfR['closingrate'] = np.where(odfR['rate'] > 0, odfR['rate'], odfR['closingrate'])
        odfR['balance'] = odfR['quantity'] * odfR['closingrate']

        # Drop unnecessary columns
        odfR.drop(['stockttype', 'entrydatetime', '_merge', 'rate'], axis=1, inplace=True)

        # Rename columns for clarity
        odfR.rename(columns={'stock__id': 'account__id', 'stock': 'account__accountname'}, inplace=True)
        odfR['accounthead__name'] = 'Opening Stock'

        # Fetch account ID once and use it directly
        account_id = accountHead.objects.get(code=9000, entity=self.entityid).id
        odfR['account__accounthead'] = account_id

        return odfR



      #  return odfR
    

    def gettradingdetails(self):

       # currentdates = entityfinancialyear.objects.get(entity = self.entityid,finendyear__gte = self.startdate  ,finstartyear__lte =  self.startdate)

       # currentdates.finstartyear

        # Common filters
        common_filters = {
            'isactive': 1,
            'entity': self.entityid,
            'account__accounthead__detailsingroup': 1,
        }

        exclude_filters = {
            'accounttype': 'MD',
            'transactiontype': 'PC',
        }

        annotate_fields = {
            'balance': Sum('debitamount', default=0) - Sum('creditamount', default=0),
            'quantity': Sum('quantity', default=0),
        }

        values_fields = ['accounthead__name', 'account__accounthead', 'account__id', 'account__accountname']

        # stk0 Query
        stk0 = StockTransactions.objects.filter(
            **common_filters,
            entrydatetime=self.startdate,
            account__accountcode=9000
        ).exclude(**exclude_filters).values(*values_fields).annotate(**annotate_fields).filter(balance__gt=0)

        # stk Query
        stk = StockTransactions.objects.filter(
            **common_filters,
            entrydatetime__range=(self.startdate, self.enddate)
        ).exclude(
            **exclude_filters
        ).exclude(
            account__accountcode__in=[200, 9000]
        ).values(*values_fields).annotate(**annotate_fields).filter(balance__gt=0)

        # stk2 Query (similar to stk with balance__lt=0 and a different value field)
        stk2 = StockTransactions.objects.filter(
            **common_filters,
            entrydatetime__range=(self.startdate, self.enddate)
        ).exclude(
            **exclude_filters
        ).exclude(
            account__accountcode__in=[200, 9000]
        ).values('accounthead__name', 'account__creditaccounthead', 'account__id', 'account__accountname').annotate(**annotate_fields).filter(balance__lt=0)

        # Union of queries
        plunion = stk.union(stk2, stk0)

        # Convert to DataFrame
        pldf = read_frame(plunion)

       # print(pldf)

        return pldf
    

    def getgrossprofit(self,odfR,df, dfR):

        df = pd.concat([odfR, df, dfR], ignore_index=True)

        # Convert and fill missing values in one step
        df[['balance', 'quantity', 'closingrate']] = df[['balance', 'quantity', 'closingrate']].fillna(0).astype(float)

        # Determine the gross result type and value
        balance_sum = df['balance'].sum()
        if balance_sum < 0:
            result_row = ['Gross Profit', 0.00, -1, 0.00, -balance_sum, 'Gross Profit', -1]
        else:
            result_row = ['Gross Loss', 0.00, -1, 0.00, -balance_sum, 'Gross Loss', -1]

        # Append the result row
        df.loc[len(df.index)] = result_row

        return df

    def get_gpp_and_l(self, odfR, df, dfR, pldf):
        # Combine all frames into one DataFrame
        combined_df = pd.concat([odfR, df, dfR])

        # Ensure 'balance' column is of float type
        combined_df['balance'] = combined_df['balance'].astype(float)

        # Calculate the gross profit/loss
        total_balance = combined_df['balance'].sum()

        if total_balance <= 0:
            pldf.loc[len(pldf.index)] = ['Gross Profit', -1, -1, 'Gross Profit', -total_balance]
        else:
            pldf.loc[len(pldf.index)] = ['Gross Loss', -1, -1, 'Gross Loss', -total_balance]

        # Ensure 'balance' column in pldf is of float type
        pldf['balance'] = pldf['balance'].astype(float)

        # Calculate net profit/loss based on updated pldf balance
        net_balance = pldf['balance'].sum()

        if net_balance > 0:
            pldf.loc[len(pldf.index)] = ['Net Profit', -2, -2, 'Net Profit', -net_balance]
        else:
            pldf.loc[len(pldf.index)] = ['Net Loss', -2, -2, 'Net Loss', -net_balance]

        # Optional: Uncomment the following line if you need a 'drcr' column
        # pldf['drcr'] = pldf['balance'].apply(lambda x: 0 if x > 0 else 1)

        return pldf
    
    def getgppandl(self,odfR,df, dfR,pldf):
        frames = [odfR,df, dfR]

        df = pd.concat(frames)

       # print(df)

        df['balance'] = df['balance'].astype(float)
       # print(df)

        if df['balance'].sum() <= 0:
            pldf.loc[len(pldf.index)] = ['Gross Profit', -1, -1, 'Gross Profit',-df['balance'].sum()]
        else:
            pldf.loc[len(pldf.index)] = ['Gross Loss', -1, -1, 'Gross Loss',-df['balance'].sum()]

        #print(pldf)

        pldf['balance'] = pldf['balance'].astype(float)
        if df['balance'].sum() > 0:
            pldf.loc[len(pldf.index)] = ['Net Profit', -2, -2, 'Net Profit',-pldf['balance'].sum()]
            
        else:
            pldf.loc[len(pldf.index)] = ['Net Loss ', -2, -2, 'Net Loss',-pldf['balance'].sum()]
           

        #pldf['drcr'] = pldf['balance'].apply(lambda x: 0 if x > 0 else 1)

        return pldf

    


class BalanceStatement(ListAPIView):

    #serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']
    

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        # Initialize the stk object with necessary parameters
        stk = generalfunctions(entityid=entity1, startdate=startdate, enddate=enddate)

        # Retrieve required data frames in one go
        dfR_initial = stk.getinventorydetails()
        dfR = stk.getinventorydetails_1(dfR_1=dfR_initial)

        # Opening stock details
        odfR = stk.getopeningstockdetails()

        # Other details
        df = stk.gettradingdetails()
        pldf = stk.getprofitandloss()

        # Final data frame
        dfi = dfR  # or you can keep it as dfR if needed in later code

        # Step 1: Optimize `pldf` initialization
        pldf = stk.get_gpp_and_l(odfR=odfR, df=df, dfR=dfR, pldf=pldf)
        pldf = pldf.loc[pldf.account__id == -2]  # Filter directly after initialization

        # Step 2: Pre-fetch related data for `account` and `entityfinancialyear` queries
        acc = account.objects.select_related().get(accountcode=9000, entity=entity1)
        currentdates = entityfinancialyear.objects.get(
            entity=entity1, finendyear__gte=startdate, finstartyear__lte=startdate
        )

        utc = pytz.UTC
        is_financial_year_active = (
            currentdates.isactive == 1
            or utc.localize(datetime.strptime(enddate, '%Y-%m-%d')) > currentdates.finendyear
        )

        # Step 3: Consolidate common query parameters for StockTransactions
        common_filters = {
            "isactive": 1,
            "entity": entity1,
            "entrydatetime__range": (currentdates.finstartyear, enddate),
            "account__accounthead__detailsingroup": 3,
        }
        exclude_md = {"accounttype": "MD"}

        # Step 4: Use condition to construct queries
        if is_financial_year_active:
            bs1 = (
                StockTransactions.objects.filter(**common_filters)
                .exclude(**exclude_md)
                .values(
                    "accounthead__name",
                    "account__accounthead",
                    "account__id",
                    "account__accountname",
                )
                .annotate(balance=Sum("debitamount", default=0) - Sum("creditamount", default=0))
                .filter(balance__gt=0)
            )
            bs2 = (
                StockTransactions.objects.filter(**common_filters)
                .exclude(**exclude_md)
                .values(
                    "account__creditaccounthead__name",
                    "account__creditaccounthead",
                    "account__id",
                    "account__accountname",
                )
                .annotate(balance=Sum("debitamount", default=0) - Sum("creditamount", default=0))
                .filter(balance__lt=0)
            )
        else:
            bs1 = (
                StockTransactions.objects.filter(**common_filters, isbalancesheet=1)
                .exclude(**exclude_md)
                .values(
                    "accounthead__name",
                    "account__accounthead",
                    "account__id",
                    "account__accountname",
                )
                .annotate(balance=Sum("debitamount", default=0) - Sum("creditamount", default=0))
                .filter(balance__gt=0)
            )
            bs2 = (
                StockTransactions.objects.filter(**common_filters, isbalancesheet=1)
                .exclude(**exclude_md)
                .values(
                    "account__creditaccounthead__name",
                    "account__creditaccounthead",
                    "account__id",
                    "account__accountname",
                )
                .annotate(balance=Sum("debitamount", default=0) - Sum("creditamount", default=0))
                .filter(balance__lt=0)
            )

        # Step 5: Union queries and convert to DataFrame
        bsunion = bs1.union(bs2)
        bsdf = read_frame(bsunion)

        # Step 6: Rename columns in `pldf`
        pldf.rename(columns={"account__accounthead__name": "accounthead__name"}, inplace=True)


       # print(pldf)

        
        bsdf = pd.concat([bsdf, dfi, pldf], ignore_index=True)

        # Convert 'balance' column to float in one step
        bsdf['balance'] = bsdf['balance'].astype(float)

        # Determine 'drcr' column using vectorized operations
        bsdf['drcr'] = np.where(
            bsdf['accounthead__name'] == 'Closing Stock', 0,
            np.where(
                bsdf['accounthead__name'] == 'Opening Stock', 1,
                np.where(bsdf['balance'] > 0, 0, 1)
            )
        )

        # Rename columns in one operation
        bsdf.rename(
            columns={
                'accounthead__name': 'accountheadname',
                'account__accounthead': 'accounthead',
                'account__accountname': 'accountname',
                'account__id': 'accountid'
            },
            inplace=True
        )

        # Sort the dataframe by 'accounthead'
        bsdf.sort_values(by=['accounthead'], ascending=False, inplace=True)

        # Update 'balance' column based on 'drcr'
        bsdf['balance'] = np.where(bsdf['drcr'] == 1, bsdf['balance'].abs(), -bsdf['balance'].abs())

        # Group and aggregate the data
        grouped_df = (
            bsdf.groupby(['accounthead', 'accountheadname', 'drcr', 'accountname', 'accountid'])
            [['balance', 'quantity']]
            .sum()
            .abs()
            .reset_index()
            .sort_values(by=['accounthead'])
        )
        

        # Return the transformed data as a dictionary of values
        return Response(grouped_df.T.to_dict().values())
    


class tradingaccountstatement(ListAPIView):

    #serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        # Fetch query parameters
        entity_id = self.request.query_params.get('entity')
        start_date = self.request.query_params.get('startdate')
        end_date = self.request.query_params.get('enddate')

        # Initialize general functions with parameters
        gf = generalfunctions(entityid=entity_id, startdate=start_date, enddate=end_date)

        # Get inventory details
        dfR_initial = gf.getinventorydetails()

        # Process inventory details with custom function
        dfR = gf.getinventorydetails_1(dfR_1=dfR_initial)

        # Get opening stock details
        odfR = gf.getopeningstockdetails()

        # Get trading details
        trading_df = gf.gettradingdetails()

        # Calculate gross profit
        df = gf.getgrossprofit(odfR=odfR, df=trading_df, dfR=dfR)

        # Apply 'drcr' logic in a vectorized manner for performance
        df['drcr'] = np.where(df['accounthead__name'] == 'Closing Stock', 1, np.where(df['balance'] >= 0, 0, 1))

        # Rename columns for better readability
        df.rename(columns={
            'accounthead__name': 'accountheadname',
            'account__accounthead': 'accounthead',
            'account__accountname': 'accountname',
            'account__id': 'accountid'
        }, inplace=True)

        # Group by relevant fields, aggregate, and sort the data
        grouped_df = (df.groupby(
                        ['accounthead', 'accountheadname', 'drcr', 'accountname', 'accountid', 'closingrate']
                    )[['balance', 'quantity']].sum()
                    .abs().reset_index()
                    .sort_values(by='accounthead', ascending=False))

        # Convert the DataFrame to a list of dictionaries and return the response
        return Response(grouped_df.T.to_dict().values())
    

class incomeandexpensesstatement(ListAPIView):

   # serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        # Initialize stk object
        stk = generalfunctions(entityid=entity1, startdate=startdate, enddate=enddate)

        # Fetch initial inventory details
        dfR_initial = stk.getinventorydetails()
        
        # Process inventory details (avoid unnecessary copying)
        dfR = stk.getinventorydetails_1(dfR_1=dfR_initial)
        
        # Get opening stock details
        odfR = stk.getopeningstockdetails()

        # Fetch trading details and profit & loss details
        df = stk.gettradingdetails()
        pldf = stk.getprofitandloss()

        # Combine dataframes using optimized processing
        pldf = stk.get_gpp_and_l(odfR=odfR, df=df, dfR=dfR, pldf=pldf)

        # Replace lambda with vectorized operations
        pldf['drcr'] = (pldf['balance'] > 0).astype(int)

        # Rename columns (inplace is already optimized)
        pldf.rename(columns={
            'account__accounthead__name': 'accountheadname', 
            'account__accounthead': 'accounthead',
            'account__accountname': 'accountname',
            'account__id': 'accountid'
        }, inplace=True)

        # Final aggregation (ensure efficient groupby)
        result = pldf.groupby(
            ['accounthead', 'accountheadname', 'drcr', 'accountname', 'accountid']
        )[['balance']].sum().abs().reset_index().sort_values(by=['accounthead'])

        # Return the result as a dictionary of values
        return Response(result.T.to_dict().values())

    

class netprofitbalance(ListAPIView):

   # serializer_class = balancesheetserializer
    permission_classes = (permissions.IsAuthenticated,)
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['entity']

    def get(self, request, format=None):
        entity1 = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
        stk = generalfunctions(entityid = entity1,startdate= startdate,enddate=enddate)
       
        dfR_initial = stk.getinventorydetails()
        dfR = stk.getinventorydetails_1(dfR_1 = dfR_initial)

        ##################################################################
        odfR = stk.getopeningstockdetails()
        ##################################################################
        df = stk.getstockdetails()
        pldf = stk.getprofitandloss()

        pldf = stk.getgppandl(odfR = odfR,df = df, dfR = dfR,pldf=pldf)

        pldf['drcr'] = pldf['balance'].apply(lambda x: 0 if x > 0 else 1)
        
          

        

        pldf.rename(columns = {'account__accounthead__name':'accountheadname', 'account__accounthead':'accounthead','account__accountname':'accountname','account__id':'accountid'}, inplace = True)


       # print(pldf)


        pldf = pldf[((pldf.accountid == -2))]

        constitution = Entity.objects.get(id = entity1)

        print(constitution.const.constcode)

        if constitution.const.constcode == 1001:
            aqs = account.objects.filter(entity = entity1,accounthead__code = 6200).values('id','accountname','accounthead__id', 'accounthead__name','sharepercentage')
            
            


        if constitution.const.constcode == 1000:
            aqs = account.objects.filter(entity = entity1,accounthead__code = 6300).values('id','accountname','accounthead__id','sharepercentage')

        adf = read_frame(aqs)
        print(adf)

        adf['key'] = 1
        pldf['key'] = 1

        print(adf)
        print(pldf)


        result = pd.merge(adf, pldf, on ='key').drop(['key','accounthead','accountid','accountname_y'],axis = 1)


        result['sharepercentage'] = result['sharepercentage'].astype(float)
        result['balance'] = result['balance'].astype(float)

        

        result['balance'] = (result['sharepercentage'] * result['balance'])/100

        

        result.rename(columns = {'accounthead__id':'accounthead','accountname_x':'accountname','id':'account'}, inplace = True)

        print(result)








        


        


        



        


     


     




      
    
     
        return Response(result.groupby(['accounthead','accountheadname','drcr','accountname','account'])[['balance']].sum().abs().reset_index().sort_values(by=['accounthead']).T.to_dict().values())
    
    

class gstrhsnapi(ListAPIView):

    #serializer_class = gstr1b2bserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'id':["in", "exact"]
    
    # }
    #filterset_fields = ['id']
    def get(self, request, format=None):
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')
        stardate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')


        stk =salesOrderdetails.objects.filter(Q(isactive = 1),Q(entity = entity),Q(salesorderheader__sorderdate__range = (stardate,enddate))).values('product__hsn','product__productname','product__unitofmeasurement__unitname','orderqty','linetotal','amount','cgst','sgst','igst','cess','product__totalgst')

        df = read_frame(stk)

        print(df)

        df.rename(columns = {'product__hsn':'hsn', 'product__productname':'description','product__unitofmeasurement__unitname':'uqc','orderqty':'totalquantity','linetotal':'totalvalue','product__totalgst':'rate','amount':'taxablevalue'}, inplace = True)


       
        return  Response(df.groupby(['hsn','description','uqc','rate'])[['totalquantity','totalvalue','taxablevalue','cgst','sgst','igst','cess']].sum().abs().reset_index().T.to_dict().values())
    

class gstr1b2csmallapi(ListAPIView):

    #serializer_class = gstr1b2bserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'id':["in", "exact"]
    
    # }
    #filterset_fields = ['id']
    def get(self, request, format=None):
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')
        stardate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')


        stk =salesOrderdetails.objects.filter(Q(isactive = 1),Q(entity = entity),Q(salesorderheader__sorderdate__range = (stardate,enddate))).values('salesorderheader__billno','salesorderheader__sorderdate','linetotal','amount','cgst','sgst','igst','cess','product__totalgst')

        df = read_frame(stk)

        print(df)

        df.rename(columns = {'salesorderheader__billno':'invoiceno','salesorderheader__sorderdate':'invoicedate1','salesorderheader__sorderdate':'invoicedate1','product__totalgst':'rate','linetotal':'invoicevalue','amount':'taxablevalue'}, inplace = True)




        df['type'] = 'OE'
        df['placeofsupply'] = '03-Punjab'
        df['applicableoftaxrate'] = ''
        df['ecomgstin'] = ''
     #   df['reversecharge'] = 'N'
       
        df['invoicedate'] = pd.to_datetime(df['invoicedate1']).dt.strftime('%d-%B-%Y')
       # df['account__accounthead'] = -1

       # return Response(df)
        return  Response(df.groupby(['type','applicableoftaxrate','rate','placeofsupply','ecomgstin'])[['invoicevalue','taxablevalue','cgst','sgst','igst','cess']].sum().abs().reset_index().T.to_dict().values())
    

class gstr1b2clargeapi(ListAPIView):

   # serializer_class = gstr1b2bserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'id':["in", "exact"]
    
    # }
    #filterset_fields = ['id']
    def get(self, request, format=None):
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')
        stardate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')


        stk =salesOrderdetails.objects.filter(Q(isactive = 1),Q(entity = entity),Q(salesorderheader__sorderdate__range = (stardate,enddate))).values('salesorderheader__billno','salesorderheader__sorderdate','linetotal','amount','cgst','sgst','igst','cess','product__totalgst')

        df = read_frame(stk)

        print(df)

        df.rename(columns = {'salesorderheader__billno':'invoiceno','salesorderheader__sorderdate':'invoicedate1','salesorderheader__sorderdate':'invoicedate1','product__totalgst':'rate','linetotal':'invoicevalue','amount':'taxablevalue'}, inplace = True)


        df['placeofsupply'] = '03-Punjab'
        df['applicableoftaxrate'] = ''
        df['ecomgstin'] = ''
     #   df['reversecharge'] = 'N'
       
        df['invoicedate'] = pd.to_datetime(df['invoicedate1']).dt.strftime('%d-%B-%Y')
       # df['account__accounthead'] = -1

       # return Response(df)
        return  Response(df.groupby(['invoiceno','invoicedate','applicableoftaxrate','rate','placeofsupply','ecomgstin'])[['invoicevalue','taxablevalue','cgst','sgst','igst','cess']].sum().abs().reset_index().T.to_dict().values())
    

class gstr1b2baapi(ListAPIView):

    #serializer_class = gstr1b2bserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'id':["in", "exact"]
    
    # }
    #filterset_fields = ['id']
    def get(self, request, format=None):
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')
        #stardate = self.request.query_params.get('stardate')

        stardate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')


        stk =salesOrderdetails.objects.filter(Q(isactive = 1),Q(entity = entity),Q(salesorderheader__sorderdate__range = (stardate,enddate))).values('salesorderheader__accountid__gstno','salesorderheader__accountid__accountname','salesorderheader__billno','salesorderheader__sorderdate','linetotal','amount','cgst','sgst','igst','cess','product__totalgst')

        df = read_frame(stk)

        print(df)

        df.rename(columns = {'salesorderheader__accountid__gstno':'gstin', 'salesorderheader__accountid__accountname':'receivername','salesorderheader__billno':'originalinvoiceno5','salesorderheader__billno':'revisedinvoiceno','salesorderheader__sorderdate':'originalinvoicedate1','salesorderheader__sorderdate':'revisedinvoicedate1','product__totalgst':'rate','linetotal': 'invoicevalue','amount':'taxablevalue'}, inplace = True)


        df['placeofsupply'] = '03-Punjab'
        df['reversecharge'] = 'N'
        df['invoicetype'] = 'RegularB2B'
        df['originalinvoicedate'] = pd.to_datetime(df['revisedinvoicedate1']).dt.strftime('%d-%B-%Y')
        df['revisedinvoicedate'] = pd.to_datetime(df['revisedinvoicedate1']).dt.strftime('%d-%B-%Y')
        df['applicableoftaxrate'] = ''
        df['ecomgstin'] = 'N'
        df['originalinvoiceno'] = 1
       # df['account__accounthead'] = -1

        print(df)

       # return Response(df)
        return  Response(df.groupby(['gstin','receivername','originalinvoiceno','revisedinvoiceno','originalinvoicedate','revisedinvoicedate','rate','placeofsupply','reversecharge','invoicetype','applicableoftaxrate','ecomgstin'])[['invoicevalue','taxablevalue','cgst','sgst','igst','cess']].sum().abs().reset_index().T.to_dict().values())
    

class gstr1b2bapi(ListAPIView):

    #serializer_class = gstr1b2bserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'id':["in", "exact"]
    
    # }
    #filterset_fields = ['id']
    def get(self, request, format=None):
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')
        #stardate = self.request.query_params.get('stardate')

        stardate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')


        stk =salesOrderdetails.objects.filter(Q(isactive = 1),Q(entity = entity),Q(salesorderheader__sorderdate__range = (stardate,enddate))).values('salesorderheader__accountid__gstno','salesorderheader__accountid__accountname','salesorderheader__billno','salesorderheader__sorderdate','linetotal','amount','cgst','sgst','igst','cess','product__totalgst')

        df = read_frame(stk)

        print(df)

        df.rename(columns = {'salesorderheader__accountid__gstno':'gstin', 'salesorderheader__accountid__accountname':'receivername','salesorderheader__billno':'invoiceno','salesorderheader__sorderdate':'invoicedate1','salesorderheader__sorderdate':'invoicedate1','product__totalgst':'rate','linetotal': 'invoicevalue','amount':'taxablevalue'}, inplace = True)


        df['placeofsupply'] = '03-Punjab'
        df['reversecharge'] = 'N'
        df['invoicetype'] = 'RegularB2B'
        df['invoicedate'] = pd.to_datetime(df['invoicedate1']).dt.strftime('%d-%B-%Y')
        df['applicableoftaxrate'] = ''
        df['ecomgstin'] = 'N'
       # df['account__accounthead'] = -1

       # return Response(df)
        return  Response(df.groupby(['gstin','receivername','invoiceno','invoicedate','rate','placeofsupply','reversecharge','invoicetype','applicableoftaxrate','ecomgstin'])[['invoicevalue','taxablevalue','cgst','sgst','igst','cess']].sum().abs().reset_index().T.to_dict().values())
    



class stockledgerbookapi(ListAPIView):

    serializer_class = stockledgerbookserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = {'id':["in", "exact"]
    
    }
    #filterset_fields = ['id']
    def get_queryset(self):
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')

        queryset=Product.objects.filter(entity=entity).prefetch_related('stocktrans').order_by('productname')

        return queryset
    

class stockledgersummaryapi(ListAPIView):


    serializer_class = stockledgersummaryserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = {'id':["in", "exact"]
    
    }
    #filterset_fields = ['id']
    def get(self, request, format=None):
        
    
       # acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')
        # accounthead = self.request.query_params.get('accounthead')
        # drcrgroup = self.request.query_params.get('drcr')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate),accounttype = 'DD').values('stock__id','stockttype').annotate(quantity = Sum('quantity',default = 0))

        prd = Product.objects.filter(entity = entity,isactive = 1).values('id','productname')




        
        df = read_frame(stk)

        print(df)

        p_table = pd.pivot_table( data=df, 
                        index=['stock__id'], 
                        columns=['stockttype'], 
                        values='quantity')
        
        print(p_table)


        df = p_table.rename_axis(None)

        print(df)

        df['stockttype'] = df.index


        df['I'] = df['I'].fillna(0).astype(float)
        df['P'] = df['P'].fillna(0).astype(float)
        df['S'] = df['S'].fillna(0).astype(float)
        df['R'] = df['R'].fillna(0).astype(float)

        df['balancetotal'] = df['P'] + df['R'] - df['I'] - df['S']

        df.rename(columns = {'I':'issued','P':'purchase','S': 'sale', 'R': 'recieved','stockttype' : 'id'}, inplace = True)



  
       

     
        
     
        return  Response(df.T.to_dict().values())


class ledgersummarylatestget(ListAPIView):

   # serializer_class = TrialbalanceSerializerbyaccounthead
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['account__accounthead']


    
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        # accounthead = self.request.query_params.get('accounthead')
        # drcrgroup = self.request.query_params.get('drcr')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
       
        ob =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__lt = startdate).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).values('account__accounthead','account__accountname','account__id').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),balance1 = Sum('debitamount',default = 0) - Sum('creditamount',default = 0))
        stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).values('account__accounthead','account__accountname','account__id').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))




        
        df = read_frame(stk)

        if len(df.index) > 0:

            df['drcr'] = 'CR'

            df['drcr'] = df['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')
            df['credit'] = np.where(df['balance'] < 0, df['balance'],0)
            df['debit'] = np.where(df['balance'] > 0, df['balance'],0)

            

            df.rename(columns = {'account__accountname':'accountname','account__id':'account','account__accounthead': 'accounthead__id'}, inplace = True)


            obdf = read_frame(ob)


            if 'balance1' in df.columns:
                df['balance1'] = df['balance1']
            else:
                df['balance1'] = 0

            
            if 'debit' in df.columns:
                df['debit'] = df['debit']
            else:
                df['debit'] = 0

            
            if 'credit' in df.columns:
                df['credit'] = df['credit']
            else:
                df['credit'] = 0

        # obdf['drcr'] = obdf['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')

            obdf.rename(columns = {'account__accountname':'accountname','account__id':'account'}, inplace = True)

            #obdf['drcr'] = obdf['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')

            obdf = obdf.groupby(['accountname','account'])[['balance1']].sum().reset_index()

        #  print(obdf)



            df = df.groupby(['accounthead__id','accountname','drcr','account'])[['debit','credit','balance','quantity']].sum().abs().reset_index()

            df = pd.merge(df,obdf,on='account',how='outer',indicator=True)


            if 'balance1' in df.columns:
                df['balance1'] = df['balance1']
            else:
                df['balance1'] = 0

            
            if 'debit' in df.columns:
                df['debit'] = df['debit']
            else:
                df['debit'] = 0

            
            if 'credit' in df.columns:
                df['credit'] = df['credit']
            else:
                df['credit'] = 0

            df['debit'] = df['debit'].fillna(0).astype(float)
            df['credit'] = df['credit'].fillna(0).astype(float)
            df['quantity'] = df['quantity'].fillna(0).astype(float)
            df['openingbalance'] = np.where(df['_merge'] == 'left_only', 0,df['balance1'].astype(float))
            df['openingbalance'] = df['openingbalance'].astype(float)
            df['balance'] = df['debit'] - df['credit'] + df['openingbalance']
            df['balance'] = df['balance'].astype(float)
            # df['openingbalance'] = np.where(df['_merge'] == 'both',df['balance1'],df['openingbalance'])
            # df['openingbalance'] = np.where(df['_merge'] == 'right_only', 0,df['balance'])
            df['accountname'] = np.where(df['_merge'] == 'right_only', df['accountname_y'],df['accountname_x'])
            df['drcr'] = df['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')
            df['obdrcr'] = df['openingbalance'].apply(lambda x: 'CR' if x < 0 else 'DR')
            #df['accounthead'] = accounthead

            df = df.drop(['accountname_y', 'accountname_x','_merge','balance1','accounthead__id','obdrcr','drcr'],axis = 1)

            print(df)

            df.rename(columns = {'account__accountname':'accountname','account':'id','account__accounthead': 'accounthead__id','balance':'balancetotal'}, inplace = True)

            finaldf = df.sort_values(by=['accountname']).T.to_dict().values()

            return Response(finaldf)
         
        return Response(df)


class ledgersummarylatest(ListAPIView):

    serializer_class = ledgersummarySerializer
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['account__accounthead']


    
    def post(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = request.data.get('entity',None)
        # accounthead = self.request.query_params.get('accounthead')
        # drcrgroup = self.request.query_params.get('drcr')
        startdate = request.data.get('startdate',None)
        enddate = request.data.get('enddate',None)
       
        ob =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__lt = startdate).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).values('account__accounthead','account__accountname','account__id').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),balance1 = Sum('debitamount',default = 0) - Sum('creditamount',default = 0))
        stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).values('account__accounthead','account__accountname','account__id').annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),balance = Sum('debitamount',default = 0) - Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))

        if request.data.get('accounthead'):
            accountheads =  [int(x) for x in request.data.get('accounthead', '').split(',')]
            ob = ob.filter(accounthead__in=accountheads)
            stk = stk.filter(accounthead__in=accountheads)

        if request.data.get('account'):
            accounts =  [int(x) for x in request.data.get('account', '').split(',')]
            ob = ob.filter(account__in=accounts)
            stk = stk.filter(account__in=accounts)

        if request.data.get('drcr') == '1':

            stk = stk.filter(balance__gt=0)
          #  print(stk.query.__str__())
        
        if request.data.get('drcr') == '0':

            stk = stk.filter(balance__lt=0)

        if request.data.get('drcr') == 'O':

            stk = stk.filter(transactiontype = 'O')
           # print(stk.query.__str__())

        if request.data.get('amountstart') and request.data.get('amountend'):
            stk = stk.filter((Q(balance__gte=Decimal(request.data.get('amountstart'))) & Q(balance__lte=Decimal(request.data.get('amountend')))))




        
        df = read_frame(stk)

        print(df)

        if len(df.index) > 0:

            df['drcr'] = 'CR'

            df['drcr'] = df['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')
            df['credit'] = np.where(df['balance'] < 0, df['balance'],0)
            df['debit'] = np.where(df['balance'] >= 0, df['balance'],0)

            

            df.rename(columns = {'account__accountname':'accountname','account__id':'account','account__accounthead': 'accounthead__id'}, inplace = True)


            obdf = read_frame(ob)


            if 'balance1' in df.columns:
                df['balance1'] = df['balance1']
            else:
                df['balance1'] = 0

            
            if 'debit' in df.columns:
                df['debit'] = df['debit']
            else:
                df['debit'] = 0

            
            if 'credit' in df.columns:
                df['credit'] = df['credit']
            else:
                df['credit'] = 0

        # obdf['drcr'] = obdf['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')

            obdf.rename(columns = {'account__accountname':'accountname','account__id':'account'}, inplace = True)

            #obdf['drcr'] = obdf['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')

            obdf = obdf.groupby(['accountname','account'])[['balance1']].sum().reset_index()

        #  print(obdf)



            df = df.groupby(['accounthead__id','accountname','drcr','account'])[['debit','credit','balance','quantity']].sum().abs().reset_index()

            df = pd.merge(df,obdf,on='account',how='outer',indicator=True)


            if 'balance1' in df.columns:
                df['balance1'] = df['balance1']
            else:
                df['balance1'] = 0

            
            if 'debit' in df.columns:
                df['debit'] = df['debit']
            else:
                df['debit'] = 0

            
            if 'credit' in df.columns:
                df['credit'] = df['credit']
            else:
                df['credit'] = 0

            df['debit'] = df['debit'].fillna(0).astype(float)
            df['credit'] = df['credit'].fillna(0).astype(float)
            df['quantity'] = df['quantity'].fillna(0).astype(float)
            df['openingbalance'] = np.where(df['_merge'] == 'left_only', 0,df['balance1'].astype(float))
            df['openingbalance'] = df['openingbalance'].astype(float)
            df['balance'] = df['debit'] - df['credit'] + df['openingbalance']
            df['balance'] = df['balance'].astype(float)
            # df['openingbalance'] = np.where(df['_merge'] == 'both',df['balance1'],df['openingbalance'])
            # df['openingbalance'] = np.where(df['_merge'] == 'right_only', 0,df['balance'])
            df['accountname'] = np.where(df['_merge'] == 'right_only', df['accountname_y'],df['accountname_x'])
            df['drcr'] = df['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')
            df['obdrcr'] = df['openingbalance'].apply(lambda x: 'CR' if x < 0 else 'DR')
            #df['accounthead'] = accounthead

            df = df.drop(['accountname_y', 'accountname_x','_merge','balance1','accounthead__id','obdrcr','drcr'],axis = 1)

            print(df)

            df.rename(columns = {'account__accountname':'accountname','account':'id','account__accounthead': 'accounthead__id','balance':'balancetotal'}, inplace = True)

            finaldf = df.sort_values(by=['accountname']).T.to_dict().values()

            return Response(finaldf)
         
        return Response(df)
    


class accountbalance(ListAPIView):

   # serializer_class = TrialbalanceSerializerbyaccounthead
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['account__accounthead']


    
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        # accounthead = self.request.query_params.get('accounthead')
        # drcrgroup = self.request.query_params.get('drcr')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
        stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).exclude(account__accountcode__in = ['200','9000']).values('account__accounthead','account__accountname','account__id','account__accounthead__name','account__accounthead__detailsingroup',).annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))
        df = read_frame(stk)

        df['balance'] = df['debit'] - df['credit']

        df['debit'] = np.where(df['balance'] >= 0,df['balance'],0)
        df['credit'] = np.where(df['balance'] <= 0,-df['balance'],0)



        df['drcr'] = np.where(df['credit'] > df['debit'],0,1)

        df = df[~((df.credit == 0.0) & (df.debit == 0.0))]

        #df = df.groupby(['accounthead__id','accountname','drcr','account'])[['debit','credit','quantity']].sum().abs().reset_index()
        df.rename(columns = {'account__accountname':'accountname','account__id':'account','account__accounthead': 'accounthead','account__accounthead__name':'accountheadname','account__accounthead__detailsingroup':'accounttype','credit':'creditamount','debit':'debitamount'}, inplace = True)
        return Response(df.sort_values(by=['accountname']).T.to_dict().values())



class ledgerapiApiView(GenericAPIView):
    permission_classes = (permissions.AllowAny,)
    authentication_classes = []

    #serializer_class = LoginSerializer


    def post(self,request):
        email = request.data.get('email',None)
        password = request.data.get('password',None)

        print(email)
        print(password)


        return Response(email)

       # user = authenticate(username = email,password = password)

        # if user:
        #     serializer = self.serializer_class(user)
        #     return response.Response(serializer.data,status = status.HTTP_200_OK)
        # return response.Response({'message': "Invalid credentials"},status = status.HTTP_401_UNAUTHORIZED)
    


class ledgerdetails(ListAPIView):

    serializer_class = ledgerdetailsSerializer
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'account':["in", "exact"],'accounthead':["in", "exact"]}


    
    def post(self, request, format=None):
        entity = request.data.get('entity', None)
        startdate = request.data.get('startdate', None)
        enddate = request.data.get('enddate')
        print(account)
        currentdates = entityfinancialyear.objects.filter(
                entity=entity,
                finstartyear__lte=enddate,  
                finendyear__gte=startdate   
            ).first()
        utc = pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        stk = StockTransactions.objects.filter(
            entry__entrydate1__range=(currentdates.finstartyear, enddate),
            isactive=1,
            entity=entity
        ).exclude(accounttype='MD').exclude(transactiontype__in=['PC']).values(
            'account__accountname',
            'entry__entrydate1',
            'transactiontype',
            'transactionid',
            'drcr',
            'desc',
            'account__id'
        ).annotate(
            debitamount=Sum('debitamount'),
            creditamount=Sum('creditamount'),
            quantity=Sum('quantity')
        ).order_by('entry__entrydate1')

        # Define transformations for filters
        filters = {
            'accounthead': lambda x: list(map(int, x.split(','))),
            'account': lambda x: list(map(int, x.split(','))),
            'transactiontype': lambda x: x.split(','),  # No need for explicit `str` casting
        }

        # Apply filters based on request data
        for key, transform in filters.items():
            value = request.data.get(key)
            if value:
                stk = stk.filter(**{f'{key}__in': transform(value)})

        # Handle 'drcr' filters
        drcr = request.data.get('drcr')
        if drcr == '1':
            stk = stk.filter(debitamount__gt=0)
        elif drcr == '0':
            stk = stk.filter(creditamount__gt=0)

        # Apply description filter
        desc = request.data.get('desc')
        if desc:
            stk = stk.filter(desc__icontains=desc)

        # Apply amount range filter
        amountstart = request.data.get('amountstart')
        amountend = request.data.get('amountend')
        if amountstart and amountend:
            amountstart = Decimal(amountstart)
            amountend = Decimal(amountend)
            stk = stk.filter(
                Q(debitamount__range=(amountstart, amountend)) |
                Q(creditamount__range=(amountstart, amountend))
            )

        df = read_frame(stk)

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        openingbalance = df[(df['entry__entrydate1'] >= datetime.date(currentdates.finstartyear)) & (df['entry__entrydate1'] < startdate.date())]

        details = df[(df['entry__entrydate1'] >= startdate.date()) & (df['entry__entrydate1'] < enddate.date())]

        details['debitamount'] = details['debitamount'].astype(float).fillna(0)
        details['creditamount'] = details['creditamount'].astype(float).fillna(0)
        details['quantity'] = details['quantity'].astype(float).fillna(0)

        if request.data.get('aggby'):
            details['entry__entrydate1'] = pd.to_datetime(details['entry__entrydate1'], errors='coerce')
            details['entry__entrydate1'] = details['entry__entrydate1'].dt.to_period(request.data.get('aggby')).dt.end_time
            if request.data.get('aggby') != 'D':
                details['transactiontype'] = request.data.get('aggby')
                details['transactionid'] = -1
                details['drcr'] = True
                details['desc'] = pd.to_datetime(details['entry__entrydate1'], format='%m').dt.strftime('%b')
                details = details.groupby(['account__accountname', 'account__id', 'entry__entrydate1', 'transactiontype', 'transactionid', 'drcr', 'desc'])[['debitamount', 'creditamount', 'quantity']].sum().abs().reset_index()

        openingbalance = openingbalance[['account__accountname', 'debitamount', 'creditamount', 'quantity', 'account__id']].copy()

        openingbalance['debitamount'] = openingbalance['debitamount'].astype(float).fillna(0)
        openingbalance['creditamount'] = openingbalance['creditamount'].astype(float).fillna(0)
        openingbalance['quantity'] = openingbalance['quantity'].astype(float).fillna(0)

        openingbalance = openingbalance.groupby(['account__accountname', 'account__id'])[['debitamount', 'creditamount', 'quantity']].sum().abs().reset_index()

        openingbalance['balance'] = openingbalance['debitamount'] - openingbalance['creditamount']

        openingbalance['debitamount'] = np.where(openingbalance['balance'] >= 0, openingbalance['balance'], 0)
        openingbalance['creditamount'] = np.where(openingbalance['balance'] <= 0, -openingbalance['balance'], 0)

        openingbalance['drcr'] = np.where(openingbalance['balance'] > 0, True, False)

        openingbalance['entry__entrydate1'] = startdate
        openingbalance['transactiontype'] = 'O'
        openingbalance['transactionid'] = -1
        openingbalance['desc'] = 'Opening'

        openingbalance.drop(['balance'], axis=1)

        df = pd.concat([openingbalance, details]).reset_index()

        df['debitamount'] = df['debitamount'].astype(float).fillna(0)
        df['creditamount'] = df['creditamount'].astype(float).fillna(0)
        df['quantity'] = df['quantity'].astype(float).fillna(0)

        df = df.groupby(['account__accountname', 'account__id'])[['debitamount', 'creditamount', 'quantity']].sum().abs().reset_index()

        df['balance'] = df['debitamount'] - df['creditamount']

        df['drcr'] = np.where(df['balance'] > 0, True, False)

        df['entry__entrydate1'] = enddate
        df['transactiontype'] = 'T'
        df['transactionid'] = -1
        df['desc'] = 'Total'

        df.drop(['balance'], axis=1)

        print(df)

        df2 = pd.concat([openingbalance, details])

        bsdf = pd.concat([df2, df]).reset_index()

        bsdf = bsdf[['account__accountname', 'account__id', 'creditamount', 'debitamount', 'desc', 'entry__entrydate1', 'transactiontype', 'transactionid', 'drcr', 'quantity']]

        bsdf.rename(columns={'account__accountname': 'accountname', 'account__id': 'accountid', 'entry__entrydate1': 'entrydate'}, inplace=True)

        bsdf['displaydate'] = pd.to_datetime(bsdf['entrydate']).dt.strftime('%d-%m-%Y')

        pd.set_option('display.max_columns', None)
        bsdf.head()

        print(bsdf)

        account_summaries = pd.DataFrame()

        if len(bsdf.index) > 0:
            account_summaries = (bsdf.groupby(['accountname', 'accountid'])
                         .apply(lambda x: x[['creditamount', 'debitamount', 'desc', 'entrydate', 'transactiontype', 'transactionid', 'displaydate', 'drcr', 'quantity']].to_dict('records'))
                         .reset_index()
                         .rename(columns={0: 'accounts'})).T.to_dict().values()

        print(account_summaries)




        
        
        
        return Response(account_summaries)




class interestdetails(ListAPIView):

    serializer_class = ledgerdetailsSerializer
    permission_classes = (permissions.IsAuthenticated,)

    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = {'account':["in", "exact"],'accounthead':["in", "exact"]}


    
    def post(self, request, format=None):
        #entity = request.data.get('entity')
        entity = request.data.get('entity',None)
        startdate = request.data.get('startdate',None)
        enddate = request.data.get('enddate')
        # param_list = request.GET.getlist('account')
        # print(param_list)
        # if request.data.get('account'):
        #     accounts =  [int(x) for x in request.GET.get('account', '').split(',')]
        print(account)
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).values('account__accountname','entry__entrydate1','account__id').annotate(debitamount = Sum('debitamount',default=0),creditamount = Sum('creditamount',default=0)).order_by('account__id','entry__entrydate1')

        # if request.data.get('accounthead'):
        #     accountheads =  [int(x) for x in request.data.get('accounthead', '').split(',')]
        #     stk = stk.filter(accounthead__in=accountheads)
        
        if request.data.get('account'):
            accounts =  [int(x) for x in request.data.get('account', '').split(',')]
            stk = stk.filter(account__in=accounts)
        
        # if request.data.get('transactiontype'):
        #     transactiontype =  [str(x) for x in request.data.get('transactiontype', '').split(',')]
        #     stk = stk.filter(transactiontype__in=transactiontype)
        #    # print(stk.query.__str__())
        #    # details = details[(details['transactiontype'].isin(transactiontype))]

        # if request.data.get('drcr') == '1':

        #     stk = stk.filter(debitamount__gt=0)
        #   #  print(stk.query.__str__())
        
        # if request.data.get('drcr') == '0':

        #     stk = stk.filter(creditamount__gt=0)
        #    # print(stk.query.__str__())

        # if request.data.get('desc'):
        #     stk = stk.filter(desc__icontains=request.data.get('desc'))

        # if request.data.get('amountstart') and request.data.get('amountend'):
        #     stk = stk.filter((Q(debitamount__gte=Decimal(request.data.get('amountstart'))) & Q(debitamount__lte=Decimal(request.data.get('amountend')))) | (Q(creditamount__gte=Decimal(request.data.get('amountstart'))) & Q(creditamount__lte=Decimal(request.data.get('amountend')))))
           


            

         
            
        df = read_frame(stk)
        df['balance'] = df['debitamount'] - df['creditamount']


        df['balance'] = df['balance'].astype(float)



        #df['Closingbalance'] = df.groupby(['account__id','account__accountname','entry__entrydate1'])['balance'].cumsum()

        
        df['Closingbalance'] = df.groupby(['account__id','account__accountname','entry__entrydate1'])['balance'].cumsum()
       # df['entry__entrydate1'] = df.groupby(['account__id','account__accountname'])['entry__entrydate1'].apply(lambda x: x.sort_values())
        df['diff'] = df.groupby(['account__id','account__accountname'])['entry__entrydate1'].diff().astype('timedelta64[D]')

       # print(np.timedelta64(1,'D'))

       # print(df)
        df['diff'] = df['diff'].shift(-1)
        df['diff'] = df['diff'].fillna(0)   
        df['Closingbalance'] = df['Closingbalance'].astype(float).abs()

      #  print(df.dtypes)


        df['int'] = ((df['Closingbalance']  * df['diff'] * 6.00/365.00)/100.00).abs().round(2)


        df['Total'] = df['balance'].astype(float).abs() +  df['int'].abs()

        df['Total'] = df['Total'].round(2)

        df.rename(columns = {'account__accountname':'accountname','account__id':'accountid','entry__entrydate1':'entrydate'}, inplace = True)


        print(df)

        


    #     df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

    #     openingbalance = df[(df['entry__entrydate1'] >= datetime.date(currentdates.finstartyear)) & (df['entry__entrydate1'] < startdate.date())]

    #     details = df[(df['entry__entrydate1'] >= startdate.date()) & (df['entry__entrydate1'] < enddate.date())]

    #     details['debitamount'] = details['debitamount'].astype(float).fillna(0)
    #     details['creditamount'] = details['creditamount'].astype(float).fillna(0)
    #     details['quantity'] = details['quantity'].astype(float).fillna(0)



    #     ################################### TransactionType ###################################


    #     # if request.data.get('transactiontype'):
    #     #     transactiontype =  [str(x) for x in request.data.get('transactiontype', '').split(',')]
    #     #     details = details[(details['transactiontype'].isin(transactiontype))]

       
    #    ################################### Debit/Credit ###################################
        
    #     # if request.data.get('drcr') == '1':

            

    #     #     details = details[(details['debitamount'] > 0)]

        

    #     # if request.data.get('drcr') == '0':

    #     #     details = details[(details['creditamount'] > 0)]

        
    #     ################################### desc ###################################

        
    #     # if request.data.get('desc'):
    #     #     details = details[details['desc'].str.contains(request.data.get('desc'), regex=True, na=True)]


    #     ################################### Range between 2 values ###################################

        
    #     # if request.data.get('amountstart') and request.data.get('amountend') :
    #     #     details = details[((details['debitamount'] >= int(request.data.get('amountstart'))) & (details['debitamount'] <= int(request.data.get('amountend')))) | ((details['creditamount'] >= int(request.data.get('amountstart'))) & (details['creditamount'] <= int(request.data.get('amountend'))))]

        
    #   #  print(details)

    #     # if request.data.get('aggby') == 'M':
    #     #     details['entry__entrydate1'] = pd.to_datetime(details['entry__entrydate1'], format="%Y-%m") + MonthEnd(0)
    #     #     details['transactiontype'] = 'A'
    #     #     details['transactionid'] = -1
    #     #     details['drcr'] = True
    #     #    # details['desc'] = 'Agg'
    #     #     details['desc'] = pd.to_datetime(details['entry__entrydate1'], format='%q').dt.strftime('%b')
    #     #     details = details.groupby(['account__accountname','account__id','entry__entrydate1','transactiontype','transactionid','drcr','desc'])[['debitamount','creditamount','quantity']].sum().abs().reset_index()

    #     if request.data.get('aggby'):
    #         details['entry__entrydate1'] = pd.to_datetime(details['entry__entrydate1'], errors='coerce')
    #         details['entry__entrydate1'] = details['entry__entrydate1'].dt.to_period(request.data.get('aggby')).dt.end_time
    #         if request.data.get('aggby') != 'D':
    #             details['transactiontype'] = request.data.get('aggby')
    #             details['transactionid'] = -1
            
    #         details['drcr'] = True
    #        # details['desc'] = 'Agg'
    #         details['desc'] = pd.to_datetime(details['entry__entrydate1'], format='%m').dt.strftime('%b')
    #         details = details.groupby(['account__accountname','account__id','entry__entrydate1','transactiontype','transactionid','drcr','desc'])[['debitamount','creditamount','quantity']].sum().abs().reset_index()
        
    #     # if request.data.get('aggby') == 'W':
    #     #     details['entry__entrydate1'] = pd.to_datetime(details['entry__entrydate1'], errors='coerce')
    #     #     details['entry__entrydate1'] = details['entry__entrydate1'].dt.to_period("W").dt.end_time
    #     #     details['transactiontype'] = 'A'
    #     #     details['transactionid'] = -1
    #     #     details['drcr'] = True
    #     #    # details['desc'] = 'Agg'
    #     #     details['desc'] = pd.to_datetime(details['entry__entrydate1'], format='%m').dt.strftime('%b')
    #     #     details = details.groupby(['account__accountname','account__id','entry__entrydate1','transactiontype','transactionid','drcr','desc'])[['debitamount','creditamount','quantity']].sum().abs().reset_index()
            


        

    #     openingbalance = openingbalance[['account__accountname','debitamount','creditamount','quantity','account__id']].copy()

    #     openingbalance['debitamount'] = openingbalance['debitamount'].astype(float).fillna(0)
    #     openingbalance['creditamount'] = openingbalance['creditamount'].astype(float).fillna(0)
    #     openingbalance['quantity'] = openingbalance['quantity'].astype(float).fillna(0)
    #    # openingbalance['balance'] = openingbalance['balance'].astype(float).fillna(0)

    #     openingbalance = openingbalance.groupby(['account__accountname','account__id'])[['debitamount','creditamount','quantity']].sum().abs().reset_index()
    #    # print(openingbalance)

        
    #     openingbalance['balance'] = openingbalance['debitamount'] - openingbalance['creditamount']
        
    #     openingbalance['debitamount'] = np.where(openingbalance['balance'] >= 0,openingbalance['balance'],0)
    #     openingbalance['creditamount'] = np.where(openingbalance['balance'] <= 0,-openingbalance['balance'],0)

    #     openingbalance['drcr'] = np.where(openingbalance['balance'] > 0,True,False)

    #     openingbalance['entry__entrydate1'] = startdate
    #     openingbalance['transactiontype'] = 'O'
    #     openingbalance['transactionid'] = -1
    #     openingbalance['desc'] = 'Opening'

    #     openingbalance.drop(['balance'],axis = 1)


    #     df = pd.concat([openingbalance,details]).reset_index()

    #     ##############################################################

    #     df['debitamount'] = df['debitamount'].astype(float).fillna(0)
    #     df['creditamount'] = df['creditamount'].astype(float).fillna(0)
    #     df['quantity'] = df['quantity'].astype(float).fillna(0)

    #     df = df.groupby(['account__accountname','account__id'])[['debitamount','creditamount','quantity']].sum().abs().reset_index()
    #    # print(openingbalance)

        

        
      #  df['balance'] = df['debitamount'] - df['creditamount']
        
        # df['debitamount'] = np.where(df['balance'] >= 0,df['balance'],0)
        # df['creditamount'] = np.where(df['balance'] <= 0,-df['balance'],0)

      #  df['drcr'] = np.where(df['balance'] > 0,True,False)

    #     df['entry__entrydate1'] = enddate
    #     df['transactiontype'] = 'T'
    #     df['transactionid'] = -1
    #     df['desc'] = 'Total'

    #     df.drop(['balance'],axis = 1)

    #     print(df)


    #     #frames = [openingbalance,details]

    #     df2 = pd.concat([openingbalance,details])

    #     bsdf = pd.concat([df2,df]).reset_index()


    #     bsdf =bsdf[['account__accountname','account__id','creditamount','debitamount','desc','entry__entrydate1','transactiontype','transactionid','drcr','quantity']]

    #   #  bsdf['entrydatetime'] = pd.to_datetime(bsdf['entrydatetime']).dt.strftime('%d-%m-%Y')

    #     bsdf.rename(columns = {'account__accountname':'accountname','account__id':'accountid','entry__entrydate1':'entrydate'}, inplace = True)

    #     bsdf['displaydate'] = pd.to_datetime(bsdf['entrydate']).dt.strftime('%d-%m-%Y')

    #     pd.set_option('display.max_columns', None)
    #     bsdf.head()

    #     print(bsdf)

       # bsdf['entrydate'] = pd.to_datetime(bsdf['entrydate'], format="%Y-%m") + MonthEnd(0)

        # bsdf['entrydate'] = pd.to_datetime(bsdf['entrydate'],format="%Y-%m-%d %H:%M:%S")



        # newdf = bsdf.groupby(['accountname','accountid',bsdf['entrydate'].dt.month])['creditamount','debitamount'].sum().reset_index()

        # newdf['entrydate'] = pd.to_datetime(newdf['entrydate'], format='%m').dt.strftime('%b')

        # print(newdf)

        



        

        j = pd.DataFrame()

        if len(df.index) > 0:
            j = (df.groupby(['accountname','accountid'])
            .apply(lambda x: x[['entrydate','balance','Closingbalance','diff','int','Total']].to_dict('records'))
            .reset_index()
            .rename(columns={0:'accounts'})).T.to_dict().values()




        
        
        
        return Response(j)

class stockledgersummary(ListAPIView):

   
    permission_classes = (permissions.IsAuthenticated,)

      
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
            
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        # stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).exclude(account__accountcode__in = ['200','9000']).values('account__accounthead','account__accountname','account__id','account__accounthead__name','account__accounthead__detailsingroup',).annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))
        # if self.request.query_params.get('stock'):
        #         stocks =  [int(x) for x in request.GET.get('stock', '').split(',')]
        #         stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,stock__in=stocks,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
        # else:
        stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')

        if self.request.query_params.get('stockcategory'):
                stockcategories =  [int(x) for x in request.GET.get('stockcategory', '').split(',')]
                stk = stk.filter(stock__productcategory__in = stockcategories)

        if self.request.query_params.get('stock'):
                stocks =  [int(x) for x in request.GET.get('stock', '').split(',')]
                stk = stk.filter(stock__in = stocks)
        
        if self.request.query_params.get('stocktype'):
                stocktypes =  [int(x) for x in request.GET.get('stocktype', '').split(',')]
                stk = stk.filter(stocktype__in = stocktypes)
        
            
        df = read_frame(stk)

       

        

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        df.rename(columns = {'stock__productname':'productname','stock__id':'productid','entry__entrydate1':'entrydate'}, inplace = True)

        

        df['salequantity'] = np.where(df['stockttype'] == "S",df['quantity'],0)
        df['purchasequantity'] = np.where(df['stockttype'] == "P",df['quantity'],0)
        df['iquantity'] = np.where(df['stockttype'] == "I",df['quantity'],0)
        df['rquantity'] = np.where(df['stockttype'] == "R",df['quantity'],0)

        dfd =df[(df['entrydate'] >= startdate.date()) & (df['entrydate'] < enddate.date())]
      
        openingbalance = df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & (df['entrydate'] < startdate.date())]
        openingbalance['salequantity'] = openingbalance['salequantity'].astype(float).fillna(0)
        openingbalance['purchasequantity'] = openingbalance['purchasequantity'].astype(float).fillna(0)
        openingbalance['rquantity'] = openingbalance['rquantity'].astype(float).fillna(0)
        openingbalance['iquantity'] = openingbalance['iquantity'].astype(float).fillna(0)
        df = dfd.groupby(['productname','productid'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()
        openingbalance = openingbalance.groupby(['productname','productid'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()   
        openingbalance['openingbalance'] = openingbalance['purchasequantity'] + openingbalance['rquantity'] - openingbalance['iquantity'] - openingbalance['salequantity']

        

        openingbalance = openingbalance.drop(['salequantity','purchasequantity','rquantity','iquantity','productname'],axis = 1)
        df1 = pd.merge(df,openingbalance,on='productid',how='outer',indicator=True).reset_index()
        df1['openingbalance'] = df1['openingbalance'].astype(float).fillna(0)
        print(df1)
      #  bsdf = pd.concat([df]).reset_index()




       




        
        
        
        return Response(df1.T.to_dict().values())
    

class stockledgersummarypost(ListAPIView):

    serializer_class = stockledgersummarySerializer

    

   
    permission_classes = (permissions.IsAuthenticated,)

      
    def post(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = request.data.get('entity',None)
        startdate = request.data.get('startdate')
        enddate = request.data.get('enddate')
            
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        # stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).exclude(account__accountcode__in = ['200','9000']).values('account__accounthead','account__accountname','account__id','account__accounthead__name','account__accounthead__detailsingroup',).annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))
        # if self.request.query_params.get('stock'):
        #         stocks =  [int(x) for x in request.GET.get('stock', '').split(',')]
        #         stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,stock__in=stocks,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
        # else:
        stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')

        if request.data.get('stockcategory'):
                stockcategories =  [int(x) for x in request.data.get('stockcategory', '').split(',')]
                stk = stk.filter(stock__productcategory__in = stockcategories)

        if request.data.get('stock'):
                stocks =  [int(x) for x in request.data.get('stock', '').split(',')]
                stk = stk.filter(stock__in = stocks)
        
        if request.data.get('stocktype'):
                stocktypes =  [str(x) for x in request.data.get('stocktype', '').split(',')]
                stk = stk.filter(stocktype__in = stocktypes)
        
            
        df = read_frame(stk)

       

        

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        df.rename(columns = {'stock__productname':'productname','stock__id':'productid','entry__entrydate1':'entrydate'}, inplace = True)

        

        df['salequantity'] = np.where(df['stockttype'] == "S",df['quantity'],0)
        df['purchasequantity'] = np.where(df['stockttype'] == "P",df['quantity'],0)
        df['iquantity'] = np.where(df['stockttype'] == "I",df['quantity'],0)
        df['rquantity'] = np.where(df['stockttype'] == "R",df['quantity'],0)

        dfd =df[(df['entrydate'] >= startdate.date()) & (df['entrydate'] < enddate.date())]
      
        openingbalance = df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & (df['entrydate'] < startdate.date())]
        openingbalance['salequantity'] = openingbalance['salequantity'].astype(float).fillna(0)
        openingbalance['purchasequantity'] = openingbalance['purchasequantity'].astype(float).fillna(0)
        openingbalance['rquantity'] = openingbalance['rquantity'].astype(float).fillna(0)
        openingbalance['iquantity'] = openingbalance['iquantity'].astype(float).fillna(0)
        df = dfd.groupby(['productname','productid'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()
        openingbalance = openingbalance.groupby(['productname','productid'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()   
        openingbalance['openingbalance'] = openingbalance['purchasequantity'] + openingbalance['rquantity'] - openingbalance['iquantity'] - openingbalance['salequantity']

        

        openingbalance = openingbalance.drop(['salequantity','purchasequantity','rquantity','iquantity','productname'],axis = 1)
        df1 = pd.merge(df,openingbalance,on='productid',how='outer',indicator=True).reset_index()
        df1['openingbalance'] = df1['openingbalance'].astype(float).fillna(0)
        print(df1)
      #  bsdf = pd.concat([df]).reset_index()




       




        
        
        
        return Response(df1.T.to_dict().values())


class daybookdetails(ListAPIView):

   
    permission_classes = (permissions.IsAuthenticated,)

      
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
            
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')

        
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        # stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).exclude(account__accountcode__in = ['200','9000']).values('account__accounthead','account__accountname','account__id','account__accounthead__name','account__accounthead__detailsingroup',).annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))
        # if self.request.query_params.get('stock'):
        #         stocks =  [int(x) for x in request.GET.get('stock', '').split(',')]
        #         stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,stock__in=stocks,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
        # else:
        stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity).exclude(accounttype__in = ['MD']).values('account__id','account__accountname','transactiontype','transactionid','desc','entry__entrydate1','debitamount','creditamount','drcr','accounttype','iscashtransaction').order_by('entry__entrydate1')
            
        df = read_frame(stk)

        datestk = entry.objects.filter(entrydate1__range = (startdate,enddate),isactive = 1,entity = entity).values('entrydate1').order_by('entrydate1')

        dfdate = read_frame(datestk)

        dfdate.rename(columns = {'entrydate1':'entrydate'}, inplace = True)

       

        

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        df.rename(columns = {'account__accountname':'accountname','account__id':'accountid','entry__entrydate1':'entrydate'}, inplace = True)

        dfd =df[(df['entrydate'] >= startdate.date()) & (df['entrydate'] < enddate.date()) & (df['accounttype'] == 'CIH')]

        accdetails =df[(df['entrydate'] >= startdate.date()) & (df['entrydate'] <= enddate.date()) & (df['accounttype'] != 'CIH')]
      
        openingbalance = df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & (df['entrydate'] < startdate.date()) & (df['accounttype'] == 'CIH') & (df['iscashtransaction'] == True) ] 

        openingbalance['entrydate'] = startdate

        openingbalance['entrydate'] = pd.to_datetime(openingbalance['entrydate'], format='%Y-%m-%d').dt.date

        openingbalance = openingbalance.groupby(['entrydate'])[['debitamount','creditamount']].sum().abs().reset_index()
        


        dfd = dfd.groupby(['entrydate'])[['debitamount','creditamount']].sum().abs().reset_index()
        dfd['debitamount'] = dfd['debitamount'].astype(float).fillna(0)
        dfd['creditamount'] = dfd['creditamount'].astype(float).fillna(0)


        print(dfd)

        dfdnew = pd.merge(dfdate,dfd,on='entrydate',how='outer',indicator=True).reset_index()


        print(dfdnew)

       

        dfdnew['debitamount'] = dfdnew['debitamount'].astype(float).fillna(0)
        dfdnew['creditamount'] = dfdnew['creditamount'].astype(float).fillna(0)

        

        dfdnew = dfdnew.drop(['_merge','index'],axis = 1)

       # dfd = dfdnew


        # dfd['debitamount'] = dfd['debitamount'].astype(float).fillna(0)
        # dfd['creditamount'] = dfd['creditamount'].astype(float).fillna(0)

        print(openingbalance)

        bsdf = pd.concat([openingbalance,dfdnew]).reset_index()

        print(bsdf)

        bsdf['balance'] = bsdf['debitamount'] - bsdf['creditamount']

        print(bsdf.dtypes)

        bsdf['balance'] = bsdf['balance'].astype(float).fillna(0)

        bsdf['Closingbalance'] = bsdf['balance'].cumsum()

        #

        


        bsdf['Openingbalance'] = bsdf['Closingbalance'] - bsdf['balance']

        bsdf['Openingbalance'] = bsdf['Openingbalance'].astype(float).fillna(0)
        bsdf['debitamount'] = bsdf['debitamount'].astype(float).fillna(0)

        bsdf['Closingbalance'] = bsdf['Closingbalance'].astype(float).fillna(0)
        bsdf['creditamount'] = bsdf['creditamount'].astype(float).fillna(0)

        bsdf['receipttotal'] = bsdf['Openingbalance'] +  bsdf['debitamount']

        bsdf['paymenttotal'] = bsdf['Closingbalance'] +  bsdf['creditamount']

        bsdf.rename(columns = {'debitamount':'receipt','creditamount':'payment'}, inplace = True)

        print(bsdf)


        print('--------------------------------------------------')


        #pr


        accdetails = accdetails.groupby(['accountid','accountname','transactiontype','transactionid','desc','entrydate','drcr'])[['debitamount','creditamount']].sum().abs().reset_index()


        bsdf = pd.merge(bsdf,accdetails,on='entrydate',how='outer',indicator=True).reset_index()

      #  print(bsdf['balance'].cumsum(axis = 0, skipna = True))

        #bsdf = bsdf.groupby(['entrydate'])[['debitamount','creditamount']].sum().abs().reset_index()


        print('abdnsbdnsdbndbn')


        pd.set_option('display.max_columns', None)
        bsdf.head()

        bsdf = bsdf.fillna(0)


     #   print(bsdf)


        j = pd.DataFrame()
        if len(bsdf.index) > 0:
            j = (bsdf.groupby(['entrydate','receipt','payment','Closingbalance','Openingbalance','receipttotal','paymenttotal'])
            .apply(lambda x: x[['accountid','accountname','desc','debitamount','creditamount','drcr','transactiontype','transactionid']].to_dict('records'))
            .reset_index()
            .rename(columns={0:'accounts'})).T.to_dict().values()

        
          
        
        
        return Response(j)


class DayDetails(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, format=None):
        entity = request.data.get('entity', None)
        startdate = request.data.get('startdate')
        enddate = request.data.get('enddate')

        if not (entity and startdate and enddate):
            return Response({"error": "Missing required parameters."}, status=400)

        # Convert startdate and enddate to datetime
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')

        # Get financial year for the entity
        try:
            currentdates = entityfinancialyear.objects.get(
                entity=entity,
                finendyear__gte=startdate,
                finstartyear__lte=startdate
            )
        except entityfinancialyear.DoesNotExist:
            return Response({"error": "Financial year not found."}, status=404)

        # Combine filters to reduce redundant queries
        filter_conditions = Q(entry__entrydate1__range=(currentdates.finstartyear, enddate)) & Q(isactive=1) & Q(entity=entity)

        # First Query: Stock transactions (filtered)
        stk = StockTransactions.objects.filter(
            filter_conditions & Q(accounttype__in=['M', 'CIH']) & Q(iscashtransaction=1)
        ).values(
            'account__id', 'account__accountname', 'transactiontype', 'transactionid', 'desc', 
            'entry__entrydate1', 'drcr', 'accounttype', 'entry'
        ).annotate(
            debitamount=Sum('debitamount'),
            creditamount=Sum('creditamount')
        ).order_by('entry__entrydate1')

        # Load transactions to dataframe
        df = read_frame(stk)

        # Second Query: Stock transactions details (more detailed filtering)
        stkdetails = StockTransactions.objects.filter(
            filter_conditions
        ).exclude(
            accounttype__in=['MD']
        ).values(
            'account__id', 'account__accountname', 'transactiontype', 'transactionid', 'desc', 
            'entry__entrydate1', 'drcr', 'accounttype', 'entry'
        ).annotate(
            debitamount=Sum('debitamount'),
            creditamount=Sum('creditamount')
        ).order_by('entry__entrydate1')

                # Apply transaction type filter if provided
        transactiontype = request.data.get('transactiontype')
        if transactiontype:
            transactiontype_list = [str(x) for x in transactiontype.split(',')]
            stkdetails = stkdetails.filter(transactiontype__in=transactiontype_list)

        # Load detailed transactions to dataframe
        dfdetails = read_frame(stkdetails)

        # Process df and dfdetails
        def process_dataframe(df):
            df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date
            df.rename(columns={
                'account__accountname': 'accountname',
                'account__id': 'accountid',
                'entry__entrydate1': 'entrydate'
            }, inplace=True)
            return df

        df = process_dataframe(df)
        dfdetails = process_dataframe(dfdetails)

        # Filter CIH and non-CIH account types
        dfd = df[(df['entrydate'] >= currentdates.finstartyear.date()) &
                 (df['entrydate'] <= enddate.date()) &
                 (df['accounttype'] == 'CIH')]

        accdetails = dfdetails[(dfdetails['entrydate'] >= startdate.date()) &
                               (dfdetails['entrydate'] <= enddate.date()) &
                               (dfdetails['accounttype'] != 'CIH')]

        # Sum debit and credit amounts
        accdetails['debitamount'] = accdetails['debitamount'].astype(float).fillna(0)
        accdetails['creditamount'] = accdetails['creditamount'].astype(float).fillna(0)
        dfddetails = accdetails.groupby(['entrydate', 'entry'])[['debitamount', 'creditamount']].sum().abs().reset_index()

        # Rename columns
        dfddetails.rename(columns={'debitamount': 'payment', 'creditamount': 'receipt'}, inplace=True)

        dfd['debitamount'] = dfd['debitamount'].astype(float).fillna(0)
        dfd['creditamount'] = dfd['creditamount'].astype(float).fillna(0)
        dfd = dfd.groupby(['entrydate', 'entry'])[['debitamount', 'creditamount']].sum().abs().reset_index()

        # Merge dataframes
        bsdf = pd.merge(dfddetails, dfd, on='entrydate', how='outer').reset_index(drop=True)
        bsdf.fillna(0, inplace=True)

        # Calculate balances
        bsdf['balance'] = bsdf['debitamount'] - bsdf['creditamount']
        bsdf['Closingbalance'] = bsdf['balance'].cumsum()
        bsdf['Openingbalance'] = bsdf['Closingbalance'] - bsdf['balance']
        bsdf['receipttotal'] = bsdf['Openingbalance'] + bsdf['receipt']
        bsdf['paymenttotal'] = bsdf['Closingbalance'] + bsdf['payment']

        # Drop unnecessary columns
        bsdf = bsdf.drop(['debitamount', 'creditamount'], axis=1)

        # Merge with account details
        bsdfnew = accdetails.merge(bsdf, on='entrydate')
        bsdfnew = bsdfnew[(bsdfnew['entrydate'] >= startdate.date()) &
                          (bsdfnew['entrydate'] <= enddate.date())]
        

        print(bsdfnew.columns.tolist())

        # Prepare final JSON response
        if not bsdfnew.empty:
            grouped = bsdfnew.groupby(['entrydate', 'receipt', 'payment',
                                       'Closingbalance', 'Openingbalance',
                                       'receipttotal', 'paymenttotal'])
            j = (grouped.apply(lambda x: x[['accountid', 'accountname', 'desc',
                                            'debitamount', 'creditamount',
                                            'drcr']].to_dict('records'))
                 .reset_index().rename(columns={0: 'accounts'}).to_dict('records'))
            
            #    j = (grouped.apply(
            #         lambda x: x.groupby(['accountid', 'accountname'])
            #                 .apply(lambda y: y[['desc', 'debitamount', 'creditamount', 'drcr']].to_dict('records'))
            #                 .reset_index(name='transactions')
            #                 .to_dict('records')
            #     )
            #     .reset_index(name='accounts')
            #     .to_dict('records'))
        else:
            j = []

        return Response(j)



class cashbooksummary(ListAPIView):

   
    permission_classes = (permissions.IsAuthenticated,)

      
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')


            
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')

        
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,accounttype__in = ['M','CIH'],iscashtransaction= 1).values('account__id','account__accountname','transactiontype','transactionid','desc','entry__entrydate1','debitamount','creditamount','drcr','accounttype','entry').order_by('entry__entrydate1')
            
        df = read_frame(stk)

      #  print(df)

       

        

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        df.rename(columns = {'account__accountname':'accountname','account__id':'accountid','entry__entrydate1':'entrydate'}, inplace = True)

        dfd =df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & (df['entrydate'] <= enddate.date()) & (df['accounttype'] == 'CIH')]

        accdetails =df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & (df['entrydate'] <=   enddate.date()) & (df['accounttype'] == 'M')]

        

      
        

        dfd['debitamount'] = dfd['debitamount'].astype(float).fillna(0)
        dfd['creditamount'] = dfd['creditamount'].astype(float).fillna(0)

        if self.request.query_params.get('aggby'):
            dfd['entrydate'] = pd.to_datetime(dfd['entrydate'], errors='coerce')
            dfd['entrydate'] = dfd['entrydate'].dt.to_period(self.request.query_params.get('aggby')).dt.end_time
            dfd['startdate'] = dfd['entrydate'].dt.to_period(self.request.query_params.get('aggby')).dt.start_time
            # details['transactiontype'] = request.data.get('aggby')
            # details['transactionid'] = -1
            # details['drcr'] = True
           # details['desc'] = 'Agg'
            dfd['entrydate'] = pd.to_datetime(dfd['entrydate']).dt.date
            dfd['startdate'] = pd.to_datetime(dfd['startdate']).dt.date
            #dfd['entrydate'] = pd.to_datetime(dfd['entrydate'], format='%m').dt.strftime('%b')
            dfd = dfd.groupby(['startdate','entrydate'])[['debitamount','creditamount']].sum().abs().reset_index()

        else:
            dfd = dfd.groupby(['entrydate'])[['debitamount','creditamount']].sum().abs().reset_index()
        bsdf = dfd
        print(bsdf)

        bsdf['balance'] = bsdf['debitamount'] - bsdf['creditamount']

        bsdf['Closingbalance'] = bsdf['balance'].cumsum()

        bsdf['Openingbalance'] = bsdf['Closingbalance'] - bsdf['balance']

       # bsdf['receipttotal'] = bsdf['Openingbalance'] +  bsdf['debitamount']

       # bsdf['paymenttotal'] = bsdf['Closingbalance'] +  bsdf['creditamount']

        bsdf.rename(columns = {'debitamount':'receipt','creditamount':'payment'}, inplace = True)

        bsdfnew = bsdf
       # bsdfnew = accdetails.merge(bsdf,on='entrydate')

        bsdfnew =bsdfnew[(bsdfnew['entrydate'] >= startdate.date()) & (bsdfnew['entrydate'] <=   enddate.date())]
        # j = pd.DataFrame()
        # if len(bsdfnew.index) > 0:
        #     j = (bsdfnew.groupby(['entrydate','receipt','payment','Closingbalance','Openingbalance','receipttotal','paymenttotal'])
        #     .apply(lambda x: x[['accountid','accountname','desc','debitamount','creditamount','drcr']].to_dict('records'))
        #     .reset_index()
        #     .rename(columns={0:'accounts'})).T.to_dict().values()

        

       

        
        
        
        return Response(bsdfnew.T.to_dict().values())


class cashbookdetails(ListAPIView):

   
    permission_classes = (permissions.IsAuthenticated,)

      
    def get(self, request, format=None):

        
        entity = request.query_params.get('entity')
        start_date = request.query_params.get('startdate')
        end_date = request.query_params.get('enddate')

        if not all([entity, start_date, end_date]):
            return Response({"error": "Missing required query parameters"}, status=400)

        # Convert start and end dates to datetime
        utc = pytz.UTC
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        print(start_date)

        # Fetch the current financial year for the entity
        # current_fin_year = entityfinancialyear.objects.filter(
        #     entity=entity,
        #     finendyear__gte=start_date,
        #     finstartyear__lte=start_date,
        # ).first()

        current_fin_year = entityfinancialyear.objects.filter(
            entity=entity,
            finstartyear__lte=start_date,  
            finendyear__gte=end_date   
        ).first()

        # If no exact match is found, extend to the earliest and latest available financial years
        if not current_fin_year:
            min_finstartyear = entityfinancialyear.objects.filter(entity=entity).aggregate(Min('finstartyear'))['finstartyear__min']
            max_finendyear = entityfinancialyear.objects.filter(entity=entity).aggregate(Max('finendyear'))['finendyear__max']

            if min_finstartyear and max_finendyear:
                current_fin_year = entityfinancialyear.objects.filter(
                    entity=entity,
                    finstartyear=min_finstartyear,
                    finendyear=max_finendyear
                )
        
        #print(min_finstartyear)

        print(current_fin_year.finstartyear)

        # Fetch stock transactions
        transactions = StockTransactions.objects.filter(
            entry__entrydate1__range=(current_fin_year.finstartyear, end_date),
            isactive=1,
            entity=entity,
            accounttype__in=['M', 'CIH'],
            iscashtransaction=1,
        ).values(
            'account__id', 'account__accountname', 'transactiontype', 'transactionid',
            'desc', 'entry__entrydate1', 'debitamount', 'creditamount', 'drcr',
            'accounttype', 'entry'
        ).order_by('entry__entrydate1')

        if not transactions.exists():
            return Response([])

        # Convert transactions to a DataFrame
        df = read_frame(transactions)
        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1']).dt.date
        df.rename(columns={
            'account__accountname': 'accountname',
            'account__id': 'accountid',
            'entry__entrydate1': 'entrydate',
        }, inplace=True)

        # Filter data for cash-in-hand (CIH) and accounts (M)
        cash_in_hand = df[(df['entrydate'] >= current_fin_year.finstartyear.date()) &
                          (df['entrydate'] <= end_date.date()) &
                          (df['accounttype'] == 'CIH')]

        accounts = df[(df['entrydate'] >= current_fin_year.finstartyear.date()) &
                      (df['entrydate'] <= end_date.date()) &
                      (df['accounttype'] == 'M')]

        # Process cash-in-hand data
        cash_in_hand[['debitamount', 'creditamount']] = cash_in_hand[['debitamount', 'creditamount']].astype(float).fillna(0)
        grouped = cash_in_hand.groupby(['entrydate', 'entry'])[['debitamount', 'creditamount']].sum().abs().reset_index()

        grouped['balance'] = grouped['debitamount'] - grouped['creditamount']
        grouped['Closingbalance'] = grouped['balance'].cumsum()
        grouped['Openingbalance'] = grouped['Closingbalance'] - grouped['balance']
        grouped['receipttotal'] = grouped['Openingbalance'] + grouped['debitamount']
        grouped['paymenttotal'] = grouped['Closingbalance'] + grouped['creditamount']
        grouped.rename(columns={'debitamount': 'receipt', 'creditamount': 'payment'}, inplace=True)

        # Merge account details with grouped data
        merged_data = accounts.merge(grouped, on='entrydate', how='inner')
        merged_data = merged_data[(merged_data['entrydate'] >= start_date.date()) &
                                  (merged_data['entrydate'] <= end_date.date())]

        # Generate response
        if merged_data.empty:
            return Response([])
            

        grouped_data = (
            merged_data
            .groupby(['entrydate', 'receipt', 'payment', 'Closingbalance', 'Openingbalance', 'receipttotal', 'paymenttotal'])
            .apply(lambda x: x[['accountid', 'accountname', 'desc', 'debitamount', 'creditamount', 'drcr']].to_dict('records'))
        )


        # grouped_data = (
        #     merged_data
        #     .groupby(['entrydate', 'receipt', 'payment', 'Closingbalance', 'Openingbalance', 'receipttotal', 'paymenttotal'])
        #     .apply(lambda group: group.groupby('drcr').apply(
        #         lambda drcr_group: drcr_group.groupby('desc').apply(
        #             lambda desc_group: desc_group[['accountid', 'accountname', 'debitamount', 'creditamount']].to_dict('records')
        #         ).to_dict()
        #     ).to_dict())
        # )

        # Reset the index and format the output
        grouped_data = grouped_data.reset_index(name='accounts')

        # Convert to dictionary values
        result = grouped_data.to_dict(orient='records')

        return Response(result)


class stockledgerdetails(ListAPIView):

    serializer_class = stockledgerdetailSerializer
    permission_classes = (permissions.IsAuthenticated,)

      
    def post(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        #entity = request.data.get('entity',None)
        entity = request.data.get('entity',None)
        startdate = request.data.get('startdate',None)
        enddate = request.data.get('enddate',None)
            
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d') if startdate else None
        enddate = datetime.strptime(enddate, '%Y-%m-%d') if enddate else None
        # stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).exclude(account__accountcode__in = ['200','9000']).values('account__accounthead','account__accountname','account__id','account__accounthead__name','account__accounthead__detailsingroup',).annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))
        # if self.request.query_params.get('stock'):
        #         stocks =  [int(x) for x in request.GET.get('stock', '').split(',')]
        #         stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,stock__in=stocks,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
        # else:

        stk = StockTransactions.objects.filter(
        entry__entrydate1__range=(currentdates.finstartyear, enddate),
        isactive=1,
        entity=entity,
        accounttype='DD'
        ).values(
            'stock__id', 'stock__productname', 'entry', 'transactiontype', 
            'transactionid', 'stockttype', 'desc', 'quantity', 'entry__entrydate1'
        ).order_by('entry__entrydate1')
        #stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
        filters = {
        'stockcategory': ('stock__productcategory__id__in', int),
        'stock': ('stock__in', int),
        'stocktype': ('stockttype__in', str),
        'transactiontype': ('transactiontype__in', str)
        }

        for key, (filter_field, cast_type) in filters.items():
            if request.data.get(key):
                values = [cast_type(x) for x in request.data.get(key, '').split(',')]
                stk = stk.filter(**{filter_field: values})
        
            
        df = read_frame(stk)

       

        

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        df.rename(columns = {'stock__productname':'productname','stock__id':'productid','entry__entrydate1':'entrydate'}, inplace = True)

        for col, stock_type in [('salequantity', 'S'), ('purchasequantity', 'P'),
                            ('iquantity', 'I'), ('rquantity', 'R')]:
            df[col] = np.where(df['stockttype'] == stock_type, df['quantity'], 0)

        
        openingbalance = df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & 
                         (df['entrydate'] < startdate.date())]
        details = df[(df['entrydate'] >= startdate.date()) & 
                    (df['entrydate'] < enddate.date())]
      
        


        details = details[['productname','productid','transactiontype','transactionid','desc','entrydate','salequantity','purchasequantity','rquantity','iquantity']]

       #details = details.groupby(['productname','productid','transactiontype','transactionid','desc','entrydate'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()

       # print(details[['productname','productid','transactiontype','transactionid','desc','entrydate']])

        aggby = request.data.get('aggby')
        if aggby:
            details['entrydate'] = pd.to_datetime(details['entrydate']).dt.to_period(aggby).dt.end_time
            if aggby != 'D':
                details['transactiontype'] = aggby
                details['transactionid'] = -1
            details['desc'] = pd.to_datetime(details['entrydate']).dt.strftime('%b')
            details = details.groupby(
                ['productname', 'productid', 'transactiontype', 'transactionid', 'desc', 'entrydate']
            ).agg(
                {'salequantity': 'sum', 'purchasequantity': 'sum', 
                'rquantity': 'sum', 'iquantity': 'sum'}
            ).reset_index()

       # print

        


        def process_group(df, transactiontype, desc, entrydate):
            return df.groupby(['productname', 'productid']).agg(
            {'salequantity': 'sum', 'purchasequantity': 'sum', 
             'rquantity': 'sum', 'iquantity': 'sum'}
            ).reset_index().assign(
                transactiontype=transactiontype, transactionid='-1', 
                desc=desc, entrydate=entrydate
            )

        openingbalance = process_group(openingbalance, 'O', 'Opening Balance', startdate)
        total = process_group(df, 'ST', 'Total', enddate)

        # Combine results
        bsdf = pd.concat([openingbalance, details, total]).reset_index(drop=True)
        bsdf['entrydate'] = bsdf['entrydate'].dt.strftime('%d-%m-%Y')

       
        grouped_data = (
            bsdf.groupby(['productname', 'productid'])
            .apply(lambda x: x[['salequantity', 'purchasequantity', 
                                'rquantity', 'iquantity', 'desc', 
                                'entrydate', 'transactiontype', 'transactionid']].to_dict('records'))
            .reset_index().rename(columns={0: 'accounts'})
        )
        
        return Response(grouped_data.to_dict('records'))


class stockledgerdetailsget(ListAPIView):

   
    permission_classes = (permissions.IsAuthenticated,)

      
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        #entity = request.data.get('entity',None)
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
            
        currentdates = entityfinancialyear.objects.get(entity = entity,finendyear__gte = startdate  ,finstartyear__lte =  startdate)
        utc=pytz.UTC
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        # stk =StockTransactions.objects.filter(entity = entity,isactive = 1,entrydatetime__range=(startdate, enddate)).exclude(accounttype = 'MD').exclude(transactiontype__in = ['PC']).exclude(account__accountcode__in = ['200','9000']).values('account__accounthead','account__accountname','account__id','account__accounthead__name','account__accounthead__detailsingroup',).annotate(debit = Sum('debitamount',default = 0),credit = Sum('creditamount',default = 0),quantity = Sum('quantity',default = 0))
        if self.request.query_params.get('stock'):
                stocks =  [int(x) for x in request.GET.get('stock', '').split(',')]
                stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,stock__in=stocks,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
        else:
            stk = StockTransactions.objects.filter(entry__entrydate1__range = (currentdates.finstartyear,enddate),isactive = 1,entity = entity,accounttype = 'DD').values('stock__id','stock__productname','entry','transactiontype','transactionid','stockttype','desc','quantity','entry__entrydate1').order_by('entry__entrydate1')
            
        df = read_frame(stk)

       

        

        df['entry__entrydate1'] = pd.to_datetime(df['entry__entrydate1'], format='%Y-%m-%d').dt.date

        df.rename(columns = {'stock__productname':'productname','stock__id':'productid','entry__entrydate1':'entrydate'}, inplace = True)

        df['salequantity'] = np.where(df['stockttype'] == "S",df['quantity'],0)
        df['purchasequantity'] = np.where(df['stockttype'] == "P",df['quantity'],0)
        df['iquantity'] = np.where(df['stockttype'] == "I",df['quantity'],0)
        df['rquantity'] = np.where(df['stockttype'] == "R",df['quantity'],0)
      
        openingbalance = df[(df['entrydate'] >= datetime.date(currentdates.finstartyear)) & (df['entrydate'] < startdate.date())]

        details = df[(df['entrydate'] >= startdate.date()) & (df['entrydate'] < enddate.date())]

        details['salequantity'] = details['salequantity'].astype(float).fillna(0)
        details['purchasequantity'] = details['purchasequantity'].astype(float).fillna(0)
        details['rquantity'] = details['rquantity'].astype(float).fillna(0)
        details['iquantity'] = details['iquantity'].astype(float).fillna(0)


        openingbalance['salequantity'] = openingbalance['salequantity'].astype(float).fillna(0)
        openingbalance['purchasequantity'] = openingbalance['purchasequantity'].astype(float).fillna(0)
        openingbalance['rquantity'] = openingbalance['rquantity'].astype(float).fillna(0)
        openingbalance['iquantity'] = openingbalance['iquantity'].astype(float).fillna(0)



        df = df.groupby(['productname','productid'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()

        openingbalance = openingbalance.groupby(['productname','productid'])[['salequantity','purchasequantity','rquantity','iquantity']].sum().abs().reset_index()


        details = details.drop(['entry','stockttype','quantity'],axis = 1)

        df['transactiontype'] = 'ST'
        df['transactionid'] = '-1'
        df['desc'] = 'Total'
        df['entrydate'] = enddate


        openingbalance['transactiontype'] = 'O'
        openingbalance['transactionid'] = '-1'
        openingbalance['desc'] = 'Opening Balance'
        openingbalance['entrydate'] = startdate

        

        


        bsdf = pd.concat([openingbalance,details,df]).reset_index()

       # bsdf = bsdf.drop(['index'],axis = 1) 

       
        j = pd.DataFrame()
        if len(bsdf.index) > 0:
            j = (bsdf.groupby(['productname','productid'])
            .apply(lambda x: x[['salequantity','purchasequantity','rquantity','iquantity','desc','entrydate','transactiontype','transactionid']].to_dict('records'))
            .reset_index()
            .rename(columns={0:'accounts'})).T.to_dict().values()




        
        
        
        return Response(j)
    

class ledgerviewapi(ListAPIView):

    serializer_class = ledgerserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = {'id':["in", "exact"],'creditaccounthead':["in", "exact"],'accounthead':["in", "exact"]
    
    }
    #filterset_fields = ['id']
    def get_queryset(self):
        acc = self.request.query_params.get('acc')
        entity = self.request.query_params.get('entity')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')

        queryset=account.objects.filter(entity=entity).prefetch_related('accounttrans').order_by('accountname')


        print(queryset.query.__str__())

       

     
        
     
        return queryset


# class ledgerviewapi(ListAPIView):

#     serializer_class = ledgerserializer
#   #  filter_class = accountheadFilter
#     permission_classes = (permissions.IsAuthenticated,)

#     filter_backends = [DjangoFilterBackend]
#     filterset_fields = {'id':["in", "exact"],'creditaccounthead':["in", "exact"],'accounthead':["in", "exact"]
    
#     }
#     #filterset_fields = ['id']
#     def get_queryset(self):
#         acc = self.request.query_params.get('acc')
#         entity = self.request.query_params.get('entity')

        



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')

        queryset=account.objects.filter(entity=entity).prefetch_related('accounttrans').order_by('accountname')


        print(queryset.query.__str__())


class cbviewapi(ListAPIView):

    serializer_class = cbserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    #filterset_fields = ['id']
    def get_queryset(self):
        #account = self.request.query_params.get('account')
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate =  self.request.query_params.get('enddate')



      #  queryset1=StockTransactions.objects.filter(entity=entity,accounttype = 'M').order_by('account').only('account__accountname','transactiontype','drcr','transactionid','desc','debitamount','creditamount')

        queryset=entry.objects.filter(entity=entity,entrydate1__range = (startdate,enddate)).prefetch_related('cashtrans').order_by('entrydate1')

       

     
        
     
        return queryset
    
class salebyaccountapi(ListAPIView):

    #serializer_class = Salebyaccountserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    #filterset_fields = ['id']


    def get(self, request):
        entity = self.request.query_params.get('entity')
        transactiontype = self.request.query_params.get('transactiontype')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        # Validate required parameters
        if not all([entity, transactiontype, startdate, enddate]):
            return Response({"error": "Missing required parameters"}, status=400)

        # Determine the appropriate queryset based on transaction type
        queryset = None
        filter_criteria = {
            "entity": entity,
            "isactive": 1,
            "sorderdate__range": (startdate, enddate),
        }
        selected_fields = [
            "id", "sorderdate", "accountid__accountname", "accountid__accountcode",
            "totalpieces", "totalquanity", "subtotal", "cgst", "sgst", "igst",
            "cess", "gtotal", "accountid__city__cityname",
        ]

        if transactiontype == "S":
            queryset = SalesOderHeader.objects.filter(**filter_criteria).values(*selected_fields).order_by("sorderdate")
        elif transactiontype == "PR":
            queryset = PurchaseReturn.objects.filter(**filter_criteria).values(*selected_fields).order_by("sorderdate")
        else:
            return Response({"error": "Invalid transaction type"}, status=400)

        # Convert queryset to DataFrame and process it
        df = read_frame(queryset)
        column_mapping = {
            "accountid__accountname": "accountname",
            "accountid__accountcode": "accountcode",
            "totalpieces": "pieces",
            "totalquanity": "weightqty",
            "id": "transactionid",
            "sorderdate": "entrydate",
            "accountid__city__cityname": "city",
        }
        df.rename(columns=column_mapping, inplace=True)

        # Add additional columns
        df["transactiontype"] = transactiontype
        df["entrydate"] = pd.to_datetime(df["entrydate"]).dt.strftime("%d-%m-%y")

        # Return the response
        return Response(df.to_dict(orient="records"))


class printvoucherapi(ListAPIView):

    #serializer_class = Salebyaccountserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    #filterset_fields = ['id']


    def get(self, request):
        entity = self.request.query_params.get('entity')
        transactiontype = self.request.query_params.get('transactiontype')
        transactionid = self.request.query_params.get('transactionid')

        # Fetch the stock transactions
        queryset = StockTransactions.objects.filter(
            isactive=1,
            entity=entity,
            transactiontype=transactiontype,
            transactionid=transactionid
        ).exclude(accounttype__in=['MD']).values(
            'account__id', 'account__accountname', 'transactiontype', 'transactionid', 'desc',
            'entry__entrydate1', 'debitamount', 'creditamount', 'drcr', 'id', 'voucherno'
        ).order_by('id')

        # Convert queryset to DataFrame
        df = read_frame(queryset)
        df.rename(columns={
            'account__accountname': 'accountname',
            'account__id': 'account',
            'entry__entrydate1': 'entrydate'
        }, inplace=True)

        # Add columns and format
        df['transactiontype'] = transactiontype
        df['entrydate'] = pd.to_datetime(df['entrydate']).dt.strftime('%d-%m-%y')

        # Group data
        df = df.groupby(
            ['entrydate', 'voucherno', 'account', 'accountname', 'desc', 'drcr']
        )[['debitamount', 'creditamount']].sum().abs().reset_index()

        # Summarize totals
        df_sum = df.groupby(
            ['entrydate', 'voucherno']
        )[['debitamount', 'creditamount']].sum().abs().reset_index()

        df_sum['account'] = -1
        df_sum['desc'] = ''
        df_sum['drcr'] = True
        df_sum['accountname'] = 'Grand Total'

        # Concatenate summarized totals with the grouped data
        df = pd.concat([df, df_sum]).reset_index(drop=True)

        # Add voucher type based on transaction type
        voucher_map = {
            'C': 'Cash Voucher',
            'B': 'Bank Voucher',
            'J': 'Journal Voucher',
            'S': 'Sale Bill Voucher',
            'P': 'Purchase Voucher',
            'T': 'TDS Voucher',
            'PR': 'Purchase Return Voucher',
            'SR': 'Sale Return Voucher',
            'PI': 'Purchase Import Voucher',
            'RV': 'Receipt Voucher'
        }

        df['voucher'] = voucher_map.get(transactiontype, 'Unknown Voucher')

        # Format final result
        result = []
        if len(df) > 0:
            result = (df.groupby(['entrydate', 'voucherno', 'voucher'])
                    .apply(lambda x: x[['account', 'accountname', 'desc', 'debitamount', 'creditamount', 'drcr']].to_dict('records'))
                    .reset_index()
                    .rename(columns={0: 'accounts'})
                    .T.to_dict()
                    .values())

        return Response(result)
    
class purchasebyaccountapi(ListAPIView):

   #serializer_class = Purchasebyaccountserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    #filterset_fields = ['id']
    def get(self, request):
    # Extract query parameters
        params = self.request.query_params
        entity = params.get('entity')
        transactiontype = params.get('transactiontype')
        startdate = params.get('startdate')
        enddate = params.get('enddate')

        # Define a mapping of transaction types to models
        model_mapping = {
            'P': purchaseorder,
            'SR': salereturn
        }

        # Fetch the appropriate model based on transactiontype
        model = model_mapping.get(transactiontype)

        if model:
            queryset = model.objects.filter(
                entity=entity,
                isactive=1,
                voucherdate__range=(startdate, enddate)
            ).values(
                'id', 'voucherdate', 'account__accountname', 'account__accountcode',
                'totalpieces', 'totalquanity', 'subtotal', 'cgst', 'sgst',
                'igst', 'cess', 'gtotal', 'account__city__cityname'
            ).order_by('-voucherdate')

            # Convert queryset to DataFrame and process
            df = read_frame(queryset)
            df.rename(
                columns={
                    'account__accountname': 'accountname',
                    'account__id': 'account',
                    'totalpieces': 'pieces',
                    'totalquanity': 'weightqty',
                    'id': 'transactionid',
                    'accountid__accountcode': 'accountcode',
                    'voucherdate': 'entrydate',
                    'account__city__cityname': 'city'
                },
                inplace=True
            )
            df['transactiontype'] = transactiontype
            df['entrydate'] = pd.to_datetime(df['entrydate']).dt.strftime('%d-%m-%y')

            return Response(df.T.to_dict().values())

        # Handle invalid transaction type
        return Response({"error": "Invalid transaction type"}, status=400)
    

class stockviewapi(ListAPIView):

    serializer_class = stockserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    #filterset_fields = ['id']
    def get_queryset(self):
        #account = self.request.query_params.get('account')
        entity = self.request.query_params.get('entity')



        queryset1=StockTransactions.objects.filter(entity=entity).order_by('entity').only('account__accountname','stock__productname','transactiontype','transactionid','entrydatetime')

        queryset=Product.objects.filter(entity=entity).prefetch_related(Prefetch('stocktrans', queryset=queryset1,to_attr='account_transactions'))

       

     
        
     
        return queryset

class daybookviewapi(ListAPIView):

    serializer_class = cashserializer
  #  filter_class = accountheadFilter
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = {'cashtrans__transactiontype':["in", "exact"]}
    #filterset_fields = ['cashtrans__accounttype']
    #filterset_fields = ['id']
    def get_queryset(self):
        #account = self.request.query_params.get('account')
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate =  self.request.query_params.get('enddate')



        queryset1=StockTransactions.objects.filter(entity=entity,isactive = 1).annotate(debit = Sum('debitamount'),credit =Sum('creditamount')).order_by('account')

      #  print(queryset1)

        queryset=entry.objects.filter(entity=entity,entrydate1__range = (startdate,enddate)).prefetch_related(Prefetch('cashtrans', queryset=queryset1,to_attr='account_transactions')).order_by('entrydate1')
        # for q in queryset.account_transactions:
        #     print(q)

        # print(queryset)
        # print([queryset])
        # for p in queryset:
        #     print(p.account_transactions[0].credit)

       

     
        
     
        return queryset
    


class TrialbalanceApiView(ListAPIView):

    #serializer_class = TrialbalanceSerializer
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    #filterset_fields = ['id','unitType','entityName']
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')
        utc = pytz.UTC

        print("startdate (before conversion):", type(startdate), startdate)
        print("enddate (before conversion):", type(enddate), enddate)

        if isinstance(startdate, str):
            startdate = datetime.strptime(startdate, "%Y-%m-%d")

        if isinstance(enddate, str):
            enddate = datetime.strptime(enddate, "%Y-%m-%d")

        if not is_aware(startdate):
            startdate = make_aware(startdate)

        if not is_aware(enddate):
            enddate = make_aware(enddate)

        print(startdate)
        print(enddate)
        

        # Fetch the financial year details for the given entity and startdate
        try:
            currentdates = entityfinancialyear.objects.filter(
                entity=entity,
                finstartyear__lte=enddate,  
                finendyear__gte=startdate   
            ).first()

            if currentdates:
                finstartyear = currentdates.finstartyear
                finendyear = currentdates.finendyear

                # Ensure finstartyear and finendyear are timezone-aware
                if not is_aware(finstartyear):
                    finstartyear = make_aware(finstartyear)

                if not is_aware(finendyear):
                    finendyear = make_aware(finendyear)

                # Now, max() and min() will work without errors
                startdate = max(startdate, finstartyear)
                enddate = min(enddate, finendyear)
        except entityfinancialyear.DoesNotExist:
            currentdates = None  # Or handle the exception as needed

                # Common filters and configurations
        base_filters = {
            'entity': entity,
            'isactive': 1,
            'entrydatetime__lte': startdate
        }
        exclude_filters = {
            'accounttype__in': ['MD'],
            'transactiontype__in': ['PC']
        }
        annotations = {
            'debit': Sum('debitamount', default=0),
            'credit': Sum('creditamount', default=0),
            'balance1': Sum('debitamount', default=0) - Sum('creditamount', default=0)
        }

        # Determine the entrydatetime filter based on conditions
        if currentdates.finstartyear == startdate:
            obp_filters = {**base_filters, 'entrydatetime__gt': currentdates.finstartyear}
        elif currentdates.isactive == 1:
            obp_filters = {**base_filters, 'entrydatetime__gte': currentdates.finstartyear}
        else:
            obp_filters = {**base_filters, 'entrydatetime__gte': currentdates.finstartyear}

        # Query for obp (positive balance)
        obp = StockTransactions.objects.filter(**obp_filters).exclude(**exclude_filters).values(
            'account__accounthead__name', 'account__accounthead', 'account__id'
        ).annotate(**annotations).filter(balance1__gt=0)

        # Query for obn (negative balance)
        obn = StockTransactions.objects.filter(**obp_filters).exclude(**exclude_filters).values(
            'account__creditaccounthead__name', 'account__creditaccounthead', 'account__id'
        ).annotate(**annotations).filter(balance1__lt=0)



        

            # Union operation
        ob = obp.union(obn)

        # Convert queryset to DataFrame once
        df = read_frame(ob)

        # Rename columns more efficiently
        df.rename(columns={
            'account__accounthead__name': 'accountheadname',
            'account__accounthead': 'accounthead'
        }, inplace=True)

        # Perform the groupby operation in a single step, summing 'balance1'
        dffinal1 = df.groupby(['accounthead', 'accountheadname'], as_index=False)['balance1'].sum()


        print(dffinal1)

        
        if currentdates.finstartyear < startdate:
            base_queryset = StockTransactions.objects.filter(
            entity=entity,
            isactive=1,
            entrydatetime__range=(startdate, enddate),
            istrial=1
            ).exclude(
                accounttype__in=['MD']
            ).exclude(
                transactiontype__in=['PC']
            ).exclude(
                account__accountcode__in=[200, 9000]
            )

            # Query for positive balance (balance__gt=0) 
            stk = base_queryset.values(
                'account__accounthead__name', 
                'account__accounthead', 
                'account__id'
            ).annotate(
                debit=Sum('debitamount', default=0),
                credit=Sum('creditamount', default=0),
                balance=Sum('debitamount', default=0) - Sum('creditamount', default=0),
                quantity=Sum('quantity', default=0)
            ).filter(balance__gt=0)

            # Query for negative balance (balance__lt=0)
            stk2 = base_queryset.values(
                'account__creditaccounthead__name',
                'account__creditaccounthead',
                'account__id'
            ).annotate(
                debit=Sum('debitamount', default=0),
                credit=Sum('creditamount', default=0),
                balance=Sum('debitamount', default=0) - Sum('creditamount', default=0),
                quantity=Sum('quantity', default=0)
            ).filter(balance__lt=0)
            stkunion = stk.union(stk2)
        else:
            # Common filters and annotations
            common_filters = {
                'entity': entity,
                'isactive': 1,
                'istrial': 1
            }

            common_annotations = {
                'debit': Sum('debitamount', default=0),
                'credit': Sum('creditamount', default=0),
                'balance': Sum('debitamount', default=0) - Sum('creditamount', default=0),
                'quantity': Sum('quantity', default=0)
            }

            # Query for stk0
            stk0 = StockTransactions.objects.filter(
                **common_filters,
                entrydatetime=currentdates.finstartyear,
                account__accountcode=9000
            ).exclude(accounttype__in=['MD']).exclude(transactiontype__in=['PC'])\
            .values('account__accounthead__name', 'account__accounthead', 'account__id')\
            .annotate(**common_annotations).filter(balance__gt=0)

            # Query for stk
            stk = StockTransactions.objects.filter(
                **common_filters,
                entrydatetime__range=(startdate, enddate)
            ).exclude(accounttype__in=['MD']).exclude(transactiontype__in=['PC'])\
            .exclude(account__accountcode__in=[200, 9000])\
            .values('account__accounthead__name', 'account__accounthead', 'account__id')\
            .annotate(**common_annotations).filter(balance__gt=0)

            # Query for stk2
            stk2 = StockTransactions.objects.filter(
                **common_filters,
                entrydatetime__range=(startdate, enddate)
            ).exclude(accounttype__in=['MD']).exclude(transactiontype__in=['PC'])\
            .exclude(account__accountcode__in=[200, 9000])\
            .values('account__creditaccounthead__name', 'account__creditaccounthead', 'account__id')\
            .annotate(**common_annotations).filter(balance__lt=0)

            stkunion = stk.union(stk2,stk0)

        

    #  print(stkunion.query.__str__())

        df = read_frame(stkunion)

        print(df)
        
        df['drcr'] = np.where(df['balance'] < 0, 'CR', 'DR')
        df['credit'] = np.where(df['balance'] < 0, df['balance'], 0)
        df['debit'] = np.where(df['balance'] > 0, df['balance'], 0)

        # Rename columns
        df.rename(columns={
            'account__accounthead__name': 'accountheadname',
            'account__accounthead': 'accounthead'
        }, inplace=True)

        # Group by and aggregate
        dffinal = df.groupby(['accounthead', 'accountheadname', 'drcr'])[['debit', 'credit', 'balance', 'quantity']].sum().abs().reset_index()

        # Merge with dffinal1 and handle missing columns
        df = pd.merge(dffinal, dffinal1, on='accounthead', how='outer', indicator=True)

        # Set default values for missing columns
        for col in ['balance1', 'debit', 'credit', 'quantity']:
            if col in df.columns:
                df[col] = df[col].astype(float).fillna(0)
            else:
                df[col] = 0.0  # Create the column with 0.0 as default value

        # Calculate opening balance and final balance
        df['openingbalance'] = np.where(df['_merge'] == 'left_only', 0, df['balance1'])
        df['balance'] = df['debit'] - df['credit'] + df['openingbalance']

        # Update accountheadname and DR/CR flags
        df['accountheadname'] = np.where(df['_merge'] == 'right_only', df['accountheadname_y'], df['accountheadname_x'])
        df['drcr'] = np.where(df['balance'] < 0, 'CR', 'DR')
        df['obdrcr'] = np.where(df['openingbalance'] < 0, 'CR', 'DR')

        # Drop unnecessary columns and sort
        df = df.drop(['accountheadname_y', 'accountheadname_x', '_merge', 'balance1'], axis=1)
        df = df.sort_values(by='accountheadname')
        return Response(df.T.to_dict().values())
    

class TrialbalancebyaccountheadApiView(ListAPIView):

    #serializer_class = TrialbalanceSerializerbyaccounthead
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['account__accounthead']


    
    def get(self, request, format=None):
        #entity = self.request.query_params.get('entity')
        entity = self.request.query_params.get('entity')
        accounthead = self.request.query_params.get('accounthead')
        drcrgroup = self.request.query_params.get('drcr')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        # Constants
        utc = pytz.UTC
        exclude_account_types = ['MD']
        exclude_transaction_types = ['PC']

        # Get financial year details
        currentdates = entityfinancialyear.objects.filter(
                entity=entity,
                finstartyear__lte=enddate,  
                finendyear__gte=startdate   
            ).first()

        # Helper: Base Query Filters
        base_filters = {
            'entity': entity,
            'isactive': 1,
            'entrydatetime__lte': startdate,
            'entrydatetime__gte': currentdates.finstartyear,
        }

        # Adjust filters based on DR/CR group
        balance_filter = Q()
        if drcrgroup == 'DR':
            balance_filter = Q(balance1__gt=0)
        elif drcrgroup == 'CR':
            balance_filter = Q(balance1__lt=0)

        # Handle initial balance case
        if currentdates.finstartyear == utc.localize(datetime.strptime(startdate, '%Y-%m-%d')):
            base_filters['entrydatetime__gt'] = currentdates.finstartyear

        ob = StockTransactions.objects.filter(**base_filters) \
            .exclude(accounttype__in=exclude_account_types) \
            .exclude(transactiontype__in=exclude_transaction_types) \
            .values('accounthead__id', 'account__accountname', 'account__id') \
            .annotate(
                debit=Sum('debitamount', default=0),
                credit=Sum('creditamount', default=0),
                balance1=Sum('debitamount', default=0) - Sum('creditamount', default=0)
            ) \
            .filter(balance_filter)

        # Query for stock transactions within date range
        stk_filters = {
            'entity': entity,
            'isactive': 1,
            'entrydatetime__range': (startdate, enddate),
        }
        if drcrgroup == 'DR':
            stk_filters['account__accounthead'] = accounthead
            balance_condition = Q(balance__gt=0)
        elif drcrgroup == 'CR':
            stk_filters['account__creditaccounthead'] = accounthead
            balance_condition = Q(balance__lt=0)
        else:
            balance_condition = Q()

        stk = StockTransactions.objects.filter(**stk_filters) \
            .exclude(accounttype='MD') \
            .exclude(transactiontype__in=exclude_transaction_types) \
            .values('accounthead__id', 'account__accountname', 'account__id') \
            .annotate(
                debit=Sum('debitamount', default=0),
                credit=Sum('creditamount', default=0),
                balance=Sum('debitamount', default=0) - Sum('creditamount', default=0),
                quantity=Sum('quantity', default=0)
            ) \
            .filter(balance_condition)




        
        df = read_frame(stk)
        df['drcr'] = df['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')
        df['credit'] = np.where(df['balance'] < 0, df['balance'], 0)
        df['debit'] = np.where(df['balance'] > 0, df['balance'], 0)

        # Rename columns for consistency
        df.rename(columns={
            'account__accountname': 'accountname',
            'account__id': 'account',
            'account__accounthead': 'accounthead__id'
        }, inplace=True)

        # Convert `ob` queryset to DataFrame
        obdf = read_frame(ob)
        obdf.rename(columns={
            'account__accountname': 'accountname',
            'account__id': 'account'
        }, inplace=True)

        # Group `obdf` by account and sum `balance1`
        obdf = obdf.groupby(['accountname', 'account'])[['balance1']].sum().reset_index()

        # Group `df` by required fields
        df = df.groupby(['accounthead__id', 'accountname', 'drcr', 'account'])[['debit', 'credit', 'balance', 'quantity']] \
            .sum().abs().reset_index()

        # Merge `df` and `obdf` on account, preserving all entries
        df = pd.merge(df, obdf, on='account', how='outer', indicator=True)

        # Fill missing columns with default values
        for col in ['balance1', 'debit', 'credit', 'quantity']:
            if col in df.columns:
                df[col] = df[col].fillna(0).astype(float)
            else:
                df[col] = 0.0  # Create the column with default value

        # Calculate opening balance and updated balance
        df['openingbalance'] = np.where(df['_merge'] == 'left_only', 0, df['balance1'])
        df['balance'] = df['debit'] - df['credit'] + df['openingbalance']

        # Update account names based on merge results
        df['accountname'] = np.where(df['_merge'] == 'right_only', df['accountname_y'], df['accountname_x'])
        df['drcr'] = df['balance'].apply(lambda x: 'CR' if x < 0 else 'DR')
        df['obdrcr'] = df['openingbalance'].apply(lambda x: 'CR' if x < 0 else 'DR')

        # Add `accounthead` column
        df['accounthead'] = accounthead

        # Drop unnecessary columns
        df.drop(['accountname_y', 'accountname_x', '_merge', 'balance1', 'accounthead__id'], axis=1, inplace=True)

        # Remove rows where both balance and opening balance are zero
        df = df[~((df['balance'] == 0.0) & (df['openingbalance'] == 0.0))]

        # Return sorted response
        return Response(df.sort_values(by=['accountname']).T.to_dict().values())

class TrialbalancebyaccountApiView(ListAPIView):

    #serializer_class = TrialbalanceSerializerbyaccount
    permission_classes = (permissions.IsAuthenticated,)

    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['account']


    
    def get(self, request, format=None):
        # Extract query parameters
        entity = self.request.query_params.get('entity')
        account1 = self.request.query_params.get('account')
        startdate = self.request.query_params.get('startdate')
        enddate = self.request.query_params.get('enddate')

        # Common filtering condition for both queries
        common_filters = {
            'entity': entity,
            'isactive': 1,
            'account': account1
        }
        
        # Query for transactions within the date range
        stk = StockTransactions.objects.filter(
            **common_filters, 
            entrydatetime__range=(startdate, enddate)
        ).exclude(accounttype='MD').exclude(transactiontype__in=['PC']).values(
            'account__accountname', 'transactiontype', 'transactionid', 
            'entrydatetime', 'desc'
        ).annotate(
            debit=Sum('debitamount'),
            credit=Sum('creditamount'),
            quantity=Sum('quantity')
        ).order_by('entrydatetime')
        
        # Query for opening balance (before the start date)
        ob = StockTransactions.objects.filter(
            **common_filters,
            entrydatetime__lt=startdate
        ).exclude(accounttype='MD').exclude(transactiontype__in=['PC']).values(
            'account__accountname'
        ).annotate(
            debit=Sum('debitamount'),
            credit=Sum('creditamount'),
            quantity=Sum('quantity')
        ).order_by('entrydatetime')

        # Convert querysets to DataFrames
        df1 = read_frame(ob)
        df1['desc'] = 'Opening Balance'
        df1['entrydatetime'] = startdate
        
        # Ensure numeric columns are properly formatted
        df1[['quantity', 'debit', 'credit']] = df1[['quantity', 'debit', 'credit']].fillna(0).astype(float)

        # Group and aggregate data for opening balance
        df1 = df1.groupby(['account__accountname', 'entrydatetime', 'desc'])[['debit', 'credit', 'quantity']].sum().abs().reset_index()

        # If df1 has data, calculate the balance and adjust debit/credit
        if not df1.empty:
            df1['transactionid'] = -1
            df1['balance'] = df1['debit'] - df1['credit']
            df1['balance'] = df1['balance'].fillna(0).astype(float)
            
            # Adjust debit and credit based on balance
            df1['debit'] = df1['balance'].apply(lambda x: max(x, 0))
            df1['credit'] = df1['balance'].apply(lambda x: min(x, 0))
            
            # Drop the balance column
            df1.drop('balance', axis=1, inplace=True)

        # Convert main transaction DataFrame to DataFrame
        df = read_frame(stk)
        
        # Ensure numeric columns are properly formatted
        df[['quantity', 'debit', 'credit']] = df[['quantity', 'debit', 'credit']].fillna(0).astype(float)

        # Concatenate opening balance and main transactions DataFrames
        union_dfs = pd.concat([df1, df], ignore_index=True)

        # Fill missing values and preprocess columns
        union_dfs['transactiontype'] = union_dfs['transactiontype'].fillna('')
        union_dfs['transactionid'] = union_dfs['transactionid'].fillna('')
        union_dfs['desc'] = union_dfs['desc'].fillna('')
        union_dfs['sortdatetime'] = pd.to_datetime(union_dfs['entrydatetime'])
        union_dfs['entrydatetime'] = union_dfs['entrydatetime'].apply(lambda x: pd.to_datetime(x).strftime('%d-%m-%Y'))
        
        # Sorting by datetime for final presentation
        union_dfs.sort_values(by='sortdatetime', inplace=True)

        # Grouping final data and summing over specified columns
        union_dfs = union_dfs.groupby(
            ['account__accountname', 'sortdatetime', 'desc', 'transactionid', 'transactiontype', 'entrydatetime']
        )[['debit', 'credit', 'quantity']].sum().reset_index().sort_values(by='sortdatetime', ascending=True)

        # Return the response
        return Response(union_dfs.T.to_dict().values())

    

class accountListapiview(ListAPIView):
    serializer_class = accountListSerializer2
    permission_classes = (permissions.IsAuthenticated,)

    
    
    def get(self, request, format=None):
        entity = self.request.query_params.get('entity')

        if self.request.query_params.get('accounthead'):
            accountheads =  [int(x) for x in request.GET.get('accounthead', '').split(',')]
            queryset =  StockTransactions.objects.filter(entity = entity,accounthead__in = accountheads).values('account__id','account__accountname').distinct().order_by('account')
        else:
            queryset =  StockTransactions.objects.filter(entity = entity).values('account__id','account__accountname').distinct().order_by('account')

        df = read_frame(queryset) 
        df.rename(columns = {'account__accountname':'accountname','account__id':'id'}, inplace = True)

        return Response(df.T.to_dict().values())
    

class accountheadListapiview(ListAPIView):
    serializer_class = accountListSerializer2
    permission_classes = (permissions.IsAuthenticated,)

    
    
    def get(self, request, format=None):
        entity = self.request.query_params.get('entity')
        queryset =  StockTransactions.objects.filter(entity = entity).values('accounthead__id','accounthead__name').distinct().order_by('accounthead__id')

        df = read_frame(queryset) 

        df=df.dropna(subset=['accounthead__id','accounthead__name'])

        

        print(df)
        df.rename(columns = {'accounthead__name':'name','accounthead__id':'id'}, inplace = True)

        return Response(df.T.to_dict().values())
    


class productListapiview(ListAPIView):
    serializer_class = accountListSerializer2
    permission_classes = (permissions.IsAuthenticated,)

    
    
    def get(self, request, format=None):
        entity = self.request.query_params.get('entity')
        queryset =  StockTransactions.objects.filter(entity = entity,accounttype = 'DD',isactive =1).values('stock__id','stock__productname').distinct().order_by('stock')

        df = read_frame(queryset) 
        df.rename(columns = {'stock__id':'id','stock__productname':'productname'}, inplace = True)

        return Response(df.T.to_dict().values())
    
class productcategoryListapiview(ListAPIView):
    serializer_class = accountListSerializer2
    permission_classes = (permissions.IsAuthenticated,)

    
    
    def get(self, request, format=None):
        entity = self.request.query_params.get('entity')
        queryset =  StockTransactions.objects.filter(entity = entity,accounttype = 'DD',isactive =1).values('stock__productcategory__id','stock__productcategory__pcategoryname').distinct().order_by('stock__productcategory__id')

        df = read_frame(queryset) 
        df.rename(columns = {'stock__productcategory__id':'id','stock__productcategory__pcategoryname':'productcategoryname'}, inplace = True)

        return Response(df.T.to_dict().values())




class stocktypeListapiview(ListAPIView):
    serializer_class = accountListSerializer2
    permission_classes = (permissions.IsAuthenticated,)

    
    
    def get(self, request, format=None):
        entity = self.request.query_params.get('entity')
        queryset =  StockTransactions.objects.filter(entity = entity,accounttype = 'DD',isactive =1).values('stockttype').distinct().order_by('stockttype')

        df = read_frame(queryset) 

        df['stocktypename'] = np.where(df['stockttype'] == 'S', 'Sale',
              np.where(df['stockttype'] == 'P', 'Purchase',
              np.where(df['stockttype'] == 'I', 'Issued', 'recieved')))


       # df.rename(columns = {'stock__productcategory__id':'id','stock__productcategory__pcategoryname':'productcategoryname'}, inplace = True)

        return Response(df.T.to_dict().values())




class accountbindapiview(ListAPIView):
    serializer_class = accountListSerializer2
    permission_classes = (permissions.IsAuthenticated,)

    
    
    def get(self, request, format=None):
        entity = self.request.query_params.get('entity')

        if self.request.query_params.get('accounthead'):
            accountheads =  [int(x) for x in request.GET.get('accounthead', '').split(',')]
            queryset =  StockTransactions.objects.filter(entity = entity,accounthead__in = accountheads).values('account__id','account__accountname').distinct().order_by('account')
        else:
            queryset =  StockTransactions.objects.filter(entity = entity).values('account__id','account__accountname','account__accountcode','account__gstno','account__pan','account__city','account__saccode').annotate(balance1 = Sum('debitamount',default = 0) - Sum('creditamount',default = 0))

           # 'accountname','accountcode','city','gstno','pan','saccode'

        df = read_frame(queryset) 
        df.rename(columns = {'account__accountname':'accountname','account__id':'id','account__accountcode':'accountcode','account__gstno':'gstno','account__pan':'pan','account__city':'city','account__saccode':'saccode'}, inplace = True)

        return Response(df.T.to_dict().values())
    

class TrialBalanceViewFinal(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get_opening_balance(self, entity_id, start_date):
        """Fetches opening balances before the start date."""
        opening_transactions = StockTransactions.objects.filter(
            entity_id=entity_id,
            entrydatetime__date__lt=start_date
        ).exclude(accounttype='MD').exclude(transactiontype='PC')

        opening_balance_data = opening_transactions.values('account__accounthead__id', 'account__creditaccounthead__id', 'account__accountname').annotate(
            opening_debit=Coalesce(Sum('debitamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4)),
            opening_credit=Coalesce(Sum('creditamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4))
        ).annotate(
            opening_balance=F('opening_debit') - F('opening_credit'),
            accounthead_id=Case(
                When(opening_balance__lt=0, then=F('account__creditaccounthead__id')),
                default=F('account__accounthead__id'),
                output_field=CharField()
            ),
            accounthead=Case(
                When(opening_balance__lt=0, then=F('account__creditaccounthead__name')),
                default=F('account__accounthead__name'),
                output_field=CharField()
            )
        ).values(
            accounthead_id=F('accounthead_id'),
            accounthead=F('accounthead'),
            opening_balance=F('opening_balance')
        )

        return {entry["accounthead"]: {"id": entry["accounthead_id"], "opening_balance": entry["opening_balance"]} for entry in opening_balance_data}

    def get_transactions(self, entity_id, start_date, end_date):
        """Fetches transactions within the given date range."""
        transactions = StockTransactions.objects.filter(
            entity_id=entity_id,
            entrydatetime__date__range=(start_date, end_date)
        ).exclude(accounttype='MD').exclude(transactiontype='PC')

        transaction_data = transactions.values(
            'account__accounthead__id', 
            'account__creditaccounthead__id', 
            'account__accountname'
        ).annotate(
            total_debit=Coalesce(Sum('debitamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4)),
            total_credit=Coalesce(Sum('creditamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4))
        ).annotate(
            balance=F('total_debit') - F('total_credit'),
            accounthead_id=Case(
                When(balance__lt=0, then=F('account__creditaccounthead__id')),
                default=F('account__accounthead__id'),
                output_field=CharField()
            ),
            accounthead=Case(
                When(balance__lt=0, then=F('account__creditaccounthead__name')),
                default=F('account__accounthead__name'),
                output_field=CharField()
            )
        ).values(
            accounthead_id=F('accounthead_id'),
            accounthead=F('accounthead'),
            balance=F('balance'),
            total_debit=F('total_debit'),
            total_credit=F('total_credit')
        )

        return {
            entry["accounthead"]: {
                "id": entry["accounthead_id"],
                "balance": entry["balance"],
                "debit": entry["balance"] if entry["balance"] > 0 else 0,
                "credit": -entry["balance"] if entry["balance"] < 0 else 0
            } for entry in transaction_data
        }

    def aggregate_data(self, opening_balance_dict, transaction_dict):
        """Aggregates opening balances and transactions into a final result."""
        final_aggregation = {}

        # Add Opening Balances First
        for accounthead, data in opening_balance_dict.items():
            final_aggregation[accounthead] = {
                'accounthead_id': data["id"],  # ✅ Correcting to accounthead_id
                'accounthead': accounthead,
                'debit': 0,
                'credit': 0,
                'balance': data["opening_balance"],
                'opening_balance': data["opening_balance"]
            }

        # Add Transactions to the Final Data
        for accounthead, trans_data in transaction_dict.items():
            if accounthead in final_aggregation:
                final_aggregation[accounthead]['debit'] += trans_data["debit"]
                final_aggregation[accounthead]['credit'] += trans_data["credit"]
                final_aggregation[accounthead]['balance'] += trans_data["balance"]
            else:
                final_aggregation[accounthead] = {
                    'accounthead_id': trans_data["id"],  # ✅ Correcting to accounthead_id
                    'accounthead': accounthead,
                    'debit': trans_data["debit"],
                    'credit': trans_data["credit"],
                    'opening_balance': 0,
                    'balance': trans_data["balance"]
                }

        # Add DR/CR Indicators
        final_result = []
        for accounthead, data in final_aggregation.items():
            data["drcr"] = "DR" if data["balance"] > 0 else "CR"
            data["obdrcr"] = "DR" if data["opening_balance"] > 0 else "CR"

            final_result.append(data)

        return final_result

    def get(self, request, *args, **kwargs):
        entity_id = request.query_params.get('entity')
        start_date = request.query_params.get('startdate')
        end_date = request.query_params.get('enddate')

        if not entity_id or not start_date or not end_date:
            return Response({'error': 'Missing required parameters'}, status=400)

        start_date = parse_date(start_date)
        end_date = parse_date(end_date)

        # Get Data
        opening_balance_dict = self.get_opening_balance(entity_id, start_date)
        transaction_dict = self.get_transactions(entity_id, start_date, end_date)

        # Aggregate & Return
        final_result = self.aggregate_data(opening_balance_dict, transaction_dict)
        return Response(final_result)


class TrialBalanceViewaccountFinal(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get_opening_balance(self, entity_id, start_date, accounthead=None, drcr=None):
        """Fetches opening balances before the start date with filters."""
        opening_transactions = StockTransactions.objects.filter(
            entity_id=entity_id,
            entrydatetime__date__lt=start_date
        ).exclude(accounttype='MD').exclude(transactiontype='PC')

        if accounthead:
            if drcr == 'dr':
                opening_transactions = opening_transactions.filter(account__accounthead=accounthead)
            else:
                opening_transactions = opening_transactions.filter(account__creditaccounthead=accounthead)

        opening_balance_data = opening_transactions.values('account__id', 'account__accountname', 'account__accounthead').annotate(
            opening_debit=Coalesce(Sum('debitamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4)),
            opening_credit=Coalesce(Sum('creditamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4))
        ).annotate(
            opening_balance=F('opening_debit') - F('opening_credit')
        ).values(
            account_id=F('account__id'),
            accountname=F('account__accountname'),
            accounthead=F('account__accounthead'),
            opening_balance=F('opening_balance')
        )

        return {
            entry["account_id"]: {
                "accountname": entry["accountname"],
                "accounthead": entry["accounthead"],
                "opening_balance": entry["opening_balance"]
            } for entry in opening_balance_data
        }

    def get_transactions(self, entity_id, start_date, end_date, accounthead=None, drcr=None):
        """Fetches transactions within the given date range with filters."""
        transactions = StockTransactions.objects.filter(
            entity_id=entity_id,
            entrydatetime__date__range=(start_date, end_date)
        ).exclude(accounttype='MD').exclude(transactiontype='PC')

        if accounthead:
            if drcr == 'dr':
                transactions = transactions.filter(account__accounthead=accounthead)
            else:
                transactions = transactions.filter(account__creditaccounthead=accounthead)

        transaction_data = transactions.values(
            'account__id', 'account__accountname', 'account__accounthead'
        ).annotate(
            total_debit=Coalesce(Sum('debitamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4)),
            total_credit=Coalesce(Sum('creditamount'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4)),
            total_quantity=Coalesce(Sum('quantity'), Value(0), output_field=DecimalField(max_digits=14, decimal_places=4))
        ).annotate(
            balance=F('total_debit') - F('total_credit')
        ).values(
            account_id=F('account__id'),
            accountname=F('account__accountname'),
            accounthead=F('account__accounthead'),
            balance=F('balance'),
            total_debit=F('total_debit'),
            total_credit=F('total_credit'),
            total_quantity=F('total_quantity')
        )

        # Apply final filtering based on drcr parameter
        if drcr == 'dr':
            transaction_data = [entry for entry in transaction_data if entry["balance"] > 0]
        elif drcr == 'cr':
            transaction_data = [entry for entry in transaction_data if entry["balance"] < 0]

        return {
            entry["account_id"]: {
                "accountname": entry["accountname"],
                "accounthead": entry["accounthead"],
                "balance": entry["balance"],
                "debit": entry["balance"] if entry["balance"] > 0 else 0,
                "credit": -entry["balance"] if entry["balance"] < 0 else 0,
                "total_quantity": entry["total_quantity"]
            } for entry in transaction_data
        }

    def aggregate_data(self, opening_balance_dict, transaction_dict):
        """Aggregates opening balances and transactions into a final result."""
        final_aggregation = {}

        for account_id, data in opening_balance_dict.items():
            final_aggregation[account_id] = {
                'account_id': account_id,
                'accountname': data["accountname"],
                'accounthead': data["accounthead"],
                'debit': 0,
                'credit': 0,
                'balance': data["opening_balance"],
                'opening_balance': data["opening_balance"],
                'total_quantity': 0
            }

        for account_id, trans_data in transaction_dict.items():
            if account_id in final_aggregation:
                final_aggregation[account_id]['debit'] += trans_data["debit"]
                final_aggregation[account_id]['credit'] += trans_data["credit"]
                final_aggregation[account_id]['balance'] += trans_data["balance"]
                final_aggregation[account_id]['total_quantity'] += trans_data["total_quantity"]
            else:
                final_aggregation[account_id] = {
                    'account_id': account_id,
                    'accountname': trans_data["accountname"],
                    'accounthead': trans_data["accounthead"],
                    'debit': trans_data["debit"],
                    'credit': trans_data["credit"],
                    'opening_balance': 0,
                    'balance': trans_data["balance"],
                    'total_quantity': trans_data["total_quantity"]
                }

        final_result = []
        for account_id, data in final_aggregation.items():
            data["drcr"] = "DR" if data["balance"] > 0 else "CR"
            data["obdrcr"] = "DR" if data["opening_balance"] > 0 else "CR"
            final_result.append(data)

        return final_result

    def get(self, request, *args, **kwargs):
        entity_id = request.query_params.get('entity')
        start_date = request.query_params.get('startdate')
        end_date = request.query_params.get('enddate')
        accounthead = request.query_params.get('accounthead')
        drcr = request.query_params.get('drcr')

        if not entity_id or not start_date or not end_date:
            return Response({'error': 'Missing required parameters'}, status=400)

        start_date = parse_date(start_date)
        end_date = parse_date(end_date)

        opening_balance_dict = self.get_opening_balance(entity_id, start_date, accounthead, drcr)
        transaction_dict = self.get_transactions(entity_id, start_date, end_date, accounthead, drcr)

        final_result = self.aggregate_data(opening_balance_dict, transaction_dict)
        return Response(final_result)


# class StockSummaryAPIView(APIView):
#     permission_classes = [permissions.IsAuthenticated]

#     def get(self, request, *args, **kwargs):
#         entity_id = request.query_params.get('entity_id')
#         if not entity_id:
#             return Response({"error": "entity_id is required"}, status=400)

#         transactions = StockTransactions.objects.filter(entity_id=entity_id).order_by('entrydate', 'id')
#         products = Product.objects.select_related('productcategory').filter(stocktrans__entity_id=entity_id).distinct()

#         summary = []

#         for product in products:
#             fifo_stack = deque()
#             sale_rate = 0
#             total_inward_qty = 0
#             total_outward_qty = 0
#             total_inward_value = 0
#             last_movement = None

#             product_trans = transactions.filter(stock=product)

#             for tx in product_trans:
#                 qty = tx.quantity or 0
#                 last_movement = tx.entrydate if tx.entrydate else last_movement

#                 if tx.stockttype in ['P', 'R']:
#                     rate = tx.rate or sale_rate
#                     fifo_stack.append((qty, rate))
#                     total_inward_qty += qty
#                     total_inward_value += qty * rate

#                     if tx.stockttype == 'S':
#                         sale_rate = tx.rate or sale_rate

#                 elif tx.stockttype in ['S', 'I']:
#                     qty_out = qty
#                     cost_out = 0

#                     while qty_out > 0 and fifo_stack:
#                         available_qty, rate = fifo_stack[0]
#                         if available_qty <= qty_out:
#                             cost_out += available_qty * rate
#                             qty_out -= available_qty
#                             fifo_stack.popleft()
#                         else:
#                             cost_out += qty_out * rate
#                             fifo_stack[0] = (available_qty - qty_out, rate)
#                             qty_out = 0

#                     total_outward_qty += qty

#             qty_available = total_inward_qty - total_outward_qty
#             unit_rate = (total_inward_value / total_inward_qty) if total_inward_qty else 0
#             total_value = qty_available * unit_rate

#             summary.append({
#                 'Category': product.productcategory.pcategoryname if product.productcategory else '',
#                 'Code': product.productcode,
#                 'Description': product.productname,
#               #  'UOM': product.uom,
#                 'Quantity_Available': round(qty_available, 4),
#                 'Unit_Rate_FIFO': round(unit_rate, 4),
#                 'Total_Value': round(total_value, 2),
#                 'Last_Movement_Date': last_movement,
#             })

#         return Response(summary)

class StockDayBookReportView(ListAPIView):
    serializer_class = StockDayBookSerializer

    def get_queryset(self):
        queryset = StockTransactions.objects.select_related(
            'stock', 'stock__unitofmeasurement', 'account'
        ).order_by('entrydatetime')

        entity_id = self.request.query_params.get('entity')
        if entity_id:
            queryset = queryset.filter(entity_id=entity_id,accounttype='DD',isactive =1)

        return queryset
        
def calculate_fifo_rate(product, entity_id, end_date, total_qty):
    try:
        if total_qty <= 0:
            return Decimal('0.00'), Decimal('0.00')

        qty_remaining = total_qty
        total_value = Decimal('0.00')

        inflow_txns = StockTransactions.objects.filter(
            stock=product,
            entity_id=entity_id,
            accounttype='DD',
            isactive=1,
            stockttype__in=['P'],
            entrydatetime__lte=end_date
        ).order_by('entrydatetime')

        for txn in inflow_txns:
            txn_qty = txn.quantity or 0
            txn_rate = txn.rate or 0

            if txn_qty <= 0:
                continue

            if qty_remaining >= txn_qty:
                total_value += txn_qty * txn_rate
                qty_remaining -= txn_qty
            else:
                total_value += qty_remaining * txn_rate
                break

        fifo_rate = total_value / total_qty if total_qty > 0 else Decimal('0.00')
        return round(fifo_rate, 2), round(total_value, 2)

    except Exception:
        return Decimal('0.00'), Decimal('0.00')

def calculate_lifo_rate(product, entity_id, end_date, total_qty):
    try:
        if total_qty <= 0:
            return Decimal('0.00'), Decimal('0.00')

        qty_remaining = total_qty
        total_value = Decimal('0.00')

        inflow_txns = StockTransactions.objects.filter(
            stock=product,
            entity_id=entity_id,
            accounttype='DD',
            isactive=1,
            stockttype__in=['P'],
            entrydatetime__lte=end_date
        ).order_by('-entrydatetime')

        for txn in inflow_txns:
            txn_qty = txn.quantity or 0
            txn_rate = txn.rate or 0

            if txn_qty <= 0:
                continue

            if qty_remaining >= txn_qty:
                total_value += txn_qty * txn_rate
                qty_remaining -= txn_qty
            else:
                total_value += qty_remaining * txn_rate
                break

        lifo_rate = total_value / total_qty if total_qty > 0 else Decimal('0.00')
        return round(lifo_rate, 2), round(total_value, 2)

    except Exception:
        return Decimal('0.00'), Decimal('0.00')

def calculate_average_rate(product, entity_id, end_date):
    try:
        inflow_txns = StockTransactions.objects.filter(
            stock=product,
            entity_id=entity_id,
            accounttype='DD',
            isactive=1,
            stockttype__in=['P'],
            entrydatetime__lte=end_date
        )

        total_qty = inflow_txns.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
        total_value = sum((txn.quantity or 0) * (txn.rate or 0) for txn in inflow_txns)

        avg_rate = total_value / total_qty if total_qty > 0 else Decimal('0.00')
        return round(avg_rate, 2), round(total_value, 2)

    except Exception:
        return Decimal('0.00'), Decimal('0.00')

def get_last_sale_rate(product, entity_id, end_date):
    last_sale_txn = StockTransactions.objects.filter(
        stock=product,
        entity_id=entity_id,
        accounttype='DD',
        isactive=1,
        stockttype='S',
        entrydatetime__lte=end_date
    ).order_by('-entrydatetime').first()

    return round(last_sale_txn.rate or Decimal('0.00'), 2) if last_sale_txn else Decimal('0.00')

# Stock summary view
class StockSummaryView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            entity_id = request.query_params.get('entity_id')
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            ratemethod = request.query_params.get('method', 'fifo')

            if not (entity_id and start_date and end_date):
                return Response({'error': 'Missing parameters'}, status=status.HTTP_400_BAD_REQUEST)

            start_date = timezone.make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
            end_date = timezone.make_aware(datetime.strptime(end_date, "%Y-%m-%d"))

            summary = []
            products = Product.objects.filter(entity_id=entity_id)

            for product in products:
                opening_in = StockTransactions.objects.filter(
                    stock=product,
                    entity_id=entity_id,
                    accounttype='DD',
                    isactive=1,
                    stockttype__in=['R', 'P'],
                    entry__entrydate1__lt=start_date
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')

                opening_out = StockTransactions.objects.filter(
                    stock=product,
                    entity_id=entity_id,
                    accounttype='DD',
                    isactive=1,
                    stockttype__in=['I', 'S'],
                    entry__entrydate1__lt=start_date
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')

                opening_qty = opening_in - opening_out

                inward_qty = StockTransactions.objects.filter(
                    stock=product,
                    entity_id=entity_id,
                    accounttype='DD',
                    isactive=1,
                    stockttype__in=['R', 'P'],
                    entry__entrydate1__range=[start_date, end_date]
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')

                outward_qty = StockTransactions.objects.filter(
                    stock=product,
                    entity_id=entity_id,
                    accounttype='DD',
                    isactive=1,
                    stockttype__in=['I', 'S'],
                    entry__entrydate1__range=[start_date, end_date]
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')

                if opening_qty == 0 and inward_qty == 0 and outward_qty == 0:
                    continue

                closing_qty = opening_qty + inward_qty - outward_qty

                if ratemethod == 'fifo':
                    rate, value = calculate_lifo_rate(product, entity_id, end_date, closing_qty)
                elif ratemethod == 'lifo':
                    rate, value = calculate_fifo_rate(product, entity_id, end_date, closing_qty)
                elif ratemethod == 'avg':
                    rate, value = calculate_average_rate(product, entity_id, end_date)
                    value = closing_qty * rate
                elif ratemethod == 'lastsale':
                    rate = get_last_sale_rate(product, entity_id, end_date)
                    value = closing_qty * rate
                else:
                    rate, value = Decimal('0.00'), Decimal('0.00')

                last_movement = StockTransactions.objects.filter(
                    stock=product,
                    entity_id=entity_id,
                    accounttype='DD',
                    isactive=1
                ).aggregate(last=Max('entrydatetime'))['last']

                summary.append({
                    'productdesc': product.productdesc or '',
                    'productname': product.productname,
                    'category': product.productcategory.pcategoryname if hasattr(product, 'productcategory') and product.productcategory else '',
                    'uom': product.unitofmeasurement.unitcode if product.unitofmeasurement else '',
                    'opening_qty': float(opening_qty),
                    'inward_qty': float(inward_qty),
                    'outward_qty': float(outward_qty),
                    'closing_qty': float(closing_qty),
                    'rate': float(rate),
                    'value': float(value),
                    'last_movement_date': last_movement.strftime('%d-%m-%Y') if last_movement else None
                })

            serializer = StockSummarySerializerList(summary, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StockLedgerBookView(APIView):
    def get(self, request):
        entity_id = request.GET.get("entity_id")
        product_id = request.GET.get("product")
        method = request.GET.get("method", "fifo").lower()
        date_from = request.GET.get("start_date")
        date_to = request.GET.get("end_date")

        if not (entity_id and product_id):
            return Response({"error": "Entity and Product are required."}, status=400)

        all_transactions = StockTransactions.objects.filter(entity_id=entity_id, stock_id=product_id,accounttype='DD',isactive=1).order_by("entrydatetime")

        # Opening Balance Calculation (before date_from)
        opening_qty = Decimal(0)
        opening_value = Decimal(0)
        rate_stack = deque()
        total_qty = Decimal(0)
        total_value = Decimal(0)

        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                opening_txns = all_transactions.filter(entrydatetime__date__lt=date_from_obj)

                for tx in opening_txns:
                    qty = tx.quantity or Decimal(0)
                    rate = tx.rate or Decimal(0)

                    if tx.stockttype in ['P', 'R']:  # Inward
                        if method in ['fifo', 'lifo']:
                            rate_stack.append({'qty': qty, 'rate': rate})
                        elif method == 'avg':
                            total_qty += qty
                            total_value += qty * rate
                        opening_qty += qty
                        opening_value += qty * rate

                    elif tx.stockttype in ['S', 'I']:  # Outward
                        if method == 'fifo':
                            rate_used = self.calculate_fifo_rate(rate_stack, qty)
                        elif method == 'lifo':
                            rate_used = self.calculate_lifo_rate(rate_stack, qty)
                        elif method == 'avg':
                            rate_used = (total_value / total_qty) if total_qty else Decimal(0)
                            total_qty -= qty
                            total_value -= qty * rate_used
                            rate_used = rate_used
                        opening_qty -= qty
                        opening_value -= qty * rate_used

            except ValueError:
                return Response({"error": "Invalid date_from format. Use YYYY-MM-DD"}, status=400)

        # Filter main transactions
        txns = all_transactions
        if date_from:
            txns = txns.filter(entrydatetime__date__gte=date_from_obj)
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                txns = txns.filter(entrydatetime__date__lte=date_to_obj)
            except ValueError:
                return Response({"error": "Invalid date_to format. Use YYYY-MM-DD"}, status=400)

        ledger = []
        balance_qty = opening_qty
        last_known_rate = (opening_value / opening_qty) if opening_qty > 0 else Decimal(0)
        value_balance = balance_qty * last_known_rate

        # Add Opening Row
        ledger.append({
            "Date": date_from_obj if date_from else None,
            "Trans_No": "OPENING",
            "Type": "Opening Balance",
            "In_Qty": None,
            "Out_Qty": None,
            "Rate": float(last_known_rate),
            "Value": float(opening_value),
            "Balance_Qty": float(balance_qty),
            "Value_Balance": float(value_balance),
            "From": "",
            "To": "",
            "Remarks": ""
        })

        # Process Each Transaction
        for tx in txns:
            row = {}
            row['Date'] = tx.entrydatetime.date() if tx.entrydatetime else None
            row['Trans No'] = tx.voucherno or ''
            row['Type'] = self.get_transaction_type(tx)

            in_qty = out_qty = None
            if tx.stockttype in ['P', 'R']:
                in_qty = tx.quantity or Decimal(0)
                rate = tx.rate or Decimal(0)
                last_known_rate = rate

                if method in ['fifo', 'lifo']:
                    rate_stack.append({'qty': in_qty, 'rate': rate})
                elif method == 'avg':
                    total_qty += in_qty
                    total_value += in_qty * rate
                    last_known_rate = (total_value / total_qty) if total_qty else Decimal(0)

            elif tx.stockttype in ['S', 'I']:
                out_qty = tx.quantity or Decimal(0)
                if method == 'fifo':
                    last_known_rate = self.calculate_fifo_rate(rate_stack, out_qty)
                elif method == 'lifo':
                    last_known_rate = self.calculate_lifo_rate(rate_stack, out_qty)
                elif method == 'avg':
                    last_known_rate = (total_value / total_qty) if total_qty else Decimal(0)
                    total_qty -= out_qty
                    total_value -= out_qty * last_known_rate

            row['In Qty'] = float(in_qty) if in_qty else None
            row['Out Qty'] = float(out_qty) if out_qty else None
            row['Rate'] = float(last_known_rate)
            row['Value'] = float((in_qty or out_qty or Decimal(0)) * last_known_rate)

            # Update balance
            balance_qty += (in_qty or Decimal(0)) - (out_qty or Decimal(0))
            value_balance = balance_qty * last_known_rate
            row['Balance Qty'] = float(balance_qty)
            row['Value Balance'] = float(value_balance)

            row['From'] = tx.account.accountname if tx.drcr else ''
            row['To'] = tx.account.accountname if not tx.drcr else ''
            row['Remarks'] = tx.desc

            ledger.append(row)

        return Response(ledger)

    def calculate_fifo_rate(self, rate_stack, out_qty):
        value = Decimal(0)
        remaining = out_qty
        while remaining > 0 and rate_stack:
            layer = rate_stack[0]
            used = min(remaining, layer['qty'])
            value += used * layer['rate']
            layer['qty'] -= used
            remaining -= used
            if layer['qty'] == 0:
                rate_stack.popleft()
        return (value / out_qty) if out_qty else Decimal(0)

    def calculate_lifo_rate(self, rate_stack, out_qty):
        value = Decimal(0)
        remaining = out_qty
        while remaining > 0 and rate_stack:
            layer = rate_stack[-1]
            used = min(remaining, layer['qty'])
            value += used * layer['rate']
            layer['qty'] -= used
            remaining -= used
            if layer['qty'] == 0:
                rate_stack.pop()
        return (value / out_qty) if out_qty else Decimal(0)

    def get_transaction_type(self, tx):
        if tx.transactiontype == 'PRO':
            return 'Issued to Production' if tx.stockttype == 'I' else 'Received from Production'
        elif tx.stockttype == 'P':
            return 'Purchase'
        elif tx.stockttype == 'S':
            return 'Sale'
        return tx.transactiontype or ''


class TransactionTypeListView(APIView):
    def get(self, request):
        transaction_types = TransactionType.objects.all()
        serializer = TransactionTypeSerializer(transaction_types, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

def get_aging_bucket(days_overdue):
    if days_overdue <= 0:
        return 'current'
    elif days_overdue <= 30:
        return 'bucket_1_30'
    elif days_overdue <= 60:
        return 'bucket_31_60'
    elif days_overdue <= 90:
        return 'bucket_61_90'
    else:
        return 'bucket_90_plus'


class AccountsReceivableAgingReport(APIView):
    def post(self, request):
        try:
            entity = request.data.get("entity")
            startdate_str = request.data.get("startdate")
            enddate_str = request.data.get("enddate")

            if not all([entity,startdate_str, enddate_str]):
                return Response({"error": "Missing required parameters"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                startdate = datetime.strptime(startdate_str, "%Y-%m-%d").date()
                enddate = datetime.strptime(enddate_str, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

            today = date.today()

            transactions = StockTransactions.objects.select_related('account', 'entry').filter(
                entity=entity,
                accounttype='M',
                isactive=True,
                account__isnull=False,
                entry__isnull=False,
                entry__entrydate1__range=(startdate, enddate)
            ).order_by('entry__entrydate1')

            customer_data = defaultdict(lambda: {
                "customer_name": "",
                "total_amount": 0.0,
                "amount_received": 0.0,
                "total_balance": 0.0,
                "current": 0.0,
                "bucket_1_30": 0.0,
                "bucket_31_60": 0.0,
                "bucket_61_90": 0.0,
                "bucket_90_plus": 0.0,
                "last_invoice_date": None,
                "last_due_date": None,
                "last_payment_received_date": None
            })

            account_transactions = defaultdict(list)
            customer_payments = defaultdict(list)

            for txn in transactions:
                account_name = txn.account.accountname if txn.account else "Unknown"
                entry_date = txn.entry.entrydate1
                due_date = entry_date + timedelta(days=15) if entry_date else None

                if not account_name or not entry_date:
                    continue

                customer_data[account_name]["customer_name"] = account_name

                if txn.debitamount and txn.debitamount > 0:
                    account_transactions[account_name].append({
                        "amount": float(txn.debitamount),
                        "entry_date": entry_date,
                        "due_date": due_date,
                    })
                    customer_data[account_name]["total_amount"] += float(txn.debitamount)

                    if not customer_data[account_name]["last_invoice_date"] or entry_date > customer_data[account_name]["last_invoice_date"]:
                        customer_data[account_name]["last_invoice_date"] = entry_date
                        customer_data[account_name]["last_due_date"] = due_date

                elif txn.creditamount and txn.creditamount > 0:
                    customer_data[account_name]["amount_received"] += float(txn.creditamount)
                    customer_payments[account_name].append({
                        "amount": float(txn.creditamount),
                        "payment_date": entry_date
                    })

                    if (not customer_data[account_name]["last_payment_received_date"]
                            or entry_date > customer_data[account_name]["last_payment_received_date"]):
                        customer_data[account_name]["last_payment_received_date"] = entry_date

            final_output = []
            for account_name, invoices in account_transactions.items():
                payments = customer_payments.get(account_name, [])
                invoice_queue = list(invoices)
                payment_queue = list(payments)

                for payment in payment_queue:
                    amount = payment["amount"]
                    while amount > 0 and invoice_queue:
                        invoice = invoice_queue[0]
                        if invoice["amount"] <= amount:
                            amount -= invoice["amount"]
                            invoice_queue.pop(0)
                        else:
                            invoice["amount"] -= amount
                            amount = 0

                total_balance = 0.0
                for invoice in invoice_queue:
                    balance = invoice["amount"]
                    days_due = (today - invoice["due_date"]).days
                    total_balance += balance
                    if days_due <= 0:
                        customer_data[account_name]["current"] += balance
                    elif days_due <= 30:
                        customer_data[account_name]["bucket_1_30"] += balance
                    elif days_due <= 60:
                        customer_data[account_name]["bucket_31_60"] += balance
                    elif days_due <= 90:
                        customer_data[account_name]["bucket_61_90"] += balance
                    else:
                        customer_data[account_name]["bucket_90_plus"] += balance

                customer_data[account_name]["total_balance"] = total_balance

            for data in customer_data.values():
                final_output.append({
                    "customer_name": data["customer_name"],
                    "total_amount": round(data["total_amount"], 2),
                    "amount_received": round(data["amount_received"], 2),
                    "total_balance": round(data["total_balance"], 2),
                    "current": round(data["current"], 2),
                    "bucket_1_30": round(data["bucket_1_30"], 2),
                    "bucket_31_60": round(data["bucket_31_60"], 2),
                    "bucket_61_90": round(data["bucket_61_90"], 2),
                    "bucket_90_plus": round(data["bucket_90_plus"], 2),
                    "last_invoice_date": data["last_invoice_date"],
                    "last_due_date": data["last_due_date"],
                    "last_payment_received_date": data["last_payment_received_date"]
                })

            return Response(final_output)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class AccountsPayableAgingReportView(APIView):
    def post(self, request):
        try:
            entity = request.data.get('entity')
            startdate = request.data.get('startdate')
            enddate = request.data.get('enddate')

            if not (entity and  startdate and enddate):
                return Response({"error": "Missing required parameters"}, status=status.HTTP_400_BAD_REQUEST)

            start_date = datetime.strptime(startdate, "%Y-%m-%d").date()
            end_date = datetime.strptime(enddate, "%Y-%m-%d").date()
            today = date.today()

            transactions = StockTransactions.objects.select_related('account', 'entry').filter(
                accounttype='M',
                isactive=True,
                account__isnull=False,
                entry__entrydate1__range=(start_date, end_date),
                entity=entity
            )

            vendor_data = defaultdict(lambda: {
                "vendor_name": "",
                "total_amount": 0.0,
                "amount_paid": 0.0,
                "total_balance": 0.0,
                "current": 0.0,
                "bucket_1_30": 0.0,
                "bucket_31_60": 0.0,
                "bucket_61_90": 0.0,
                "bucket_90_plus": 0.0,
                "last_invoice_date": None,
                "last_due_date": None,
                "last_payment_date": None
            })

            transactions_by_vendor = defaultdict(list)
            for txn in transactions:
                if txn.account and txn.entry:
                    transactions_by_vendor[txn.account.accountname].append(txn)

            for vendor_name, txns in transactions_by_vendor.items():
                bills = []
                payments = []

                for txn in sorted(txns, key=lambda x: x.entry.entrydate1):
                    if txn.debitamount and txn.debitamount > 0:
                        bills.append({
                            "amount": float(txn.debitamount),
                            "entry_date": txn.entry.entrydate1,
                            "due_date": txn.entry.entrydate1 + timedelta(days=15),
                        })
                        if (not vendor_data[vendor_name]["last_invoice_date"]) or txn.entry.entrydate1 > vendor_data[vendor_name]["last_invoice_date"]:
                            vendor_data[vendor_name]["last_invoice_date"] = txn.entry.entrydate1
                            vendor_data[vendor_name]["last_due_date"] = txn.entry.entrydate1 + timedelta(days=15)

                        vendor_data[vendor_name]["total_amount"] += float(txn.debitamount)

                    elif txn.creditamount and txn.creditamount > 0:
                        payments.append({
                            "amount": float(txn.creditamount),
                            "entry_date": txn.entry.entrydate1
                        })
                        vendor_data[vendor_name]["amount_paid"] += float(txn.creditamount)

                        if (not vendor_data[vendor_name]["last_payment_date"]) or txn.entry.entrydate1 > vendor_data[vendor_name]["last_payment_date"]:
                            vendor_data[vendor_name]["last_payment_date"] = txn.entry.entrydate1

                for payment in payments:
                    amt = payment["amount"]
                    for bill in bills:
                        if bill["amount"] == 0:
                            continue
                        if amt <= 0:
                            break
                        applied = min(amt, bill["amount"])
                        bill["amount"] -= applied
                        amt -= applied

                for bill in bills:
                    if bill["amount"] > 0:
                        days_due = (today - bill["due_date"]).days
                        vendor_data[vendor_name]["total_balance"] += bill["amount"]

                        if days_due <= 0:
                            vendor_data[vendor_name]["current"] += bill["amount"]
                        elif days_due <= 30:
                            vendor_data[vendor_name]["bucket_1_30"] += bill["amount"]
                        elif days_due <= 60:
                            vendor_data[vendor_name]["bucket_31_60"] += bill["amount"]
                        elif days_due <= 90:
                            vendor_data[vendor_name]["bucket_61_90"] += bill["amount"]
                        else:
                            vendor_data[vendor_name]["bucket_90_plus"] += bill["amount"]

                vendor_data[vendor_name]["vendor_name"] = vendor_name

            return Response(list(vendor_data.values()))

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EMICalculatorAPIView(APIView):
    def post(self, request):
        serializer = EMICalculatorSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            emi, schedule = calculate_emi(
                data['principal'],
                data['annual_rate'],
                data['tenure_months'],
                data['start_date'],
                data['interest_type']
            )
            return Response({
                'monthly_emi': emi,
                'amortization_schedule': schedule
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class GSTSummaryView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        def aggregate_queryset(qs):
            return qs.aggregate(
                amount=Coalesce(Sum('amount'), Value(0, output_field=DecimalField())),
                cgst=Coalesce(Sum('cgst'), Value(0, output_field=DecimalField())),
                sgst=Coalesce(Sum('sgst'), Value(0, output_field=DecimalField())),
                igst=Coalesce(Sum('igst'), Value(0, output_field=DecimalField()))
            )

        def fetch_and_aggregate(gst_filter):
            sales_qs = salesOrderdetails.objects.filter(gst_filter)
            return_qs = Purchasereturndetails.objects.filter(gst_filter)
            salereturn_qs = salereturnDetails.objects.filter(gst_filter)
    
            sales_agg = aggregate_queryset(sales_qs)
            return_agg = aggregate_queryset(return_qs)
            salereturn_agg = aggregate_queryset(salereturn_qs)

            return {
                "total_amount": (sales_agg["amount"] + return_agg["amount"]) - salereturn_agg["amount"],
                "cgst": (sales_agg["cgst"] + return_agg["cgst"]) - salereturn_agg["cgst"],
                "sgst": (sales_agg["sgst"] + return_agg["sgst"]) - salereturn_agg["sgst"],
                "igst": (sales_agg["igst"] + return_agg["igst"]) - salereturn_agg["igst"]
            }

        # Outward taxable supplies (> 0%)
        taxable_filter = (
            Q(cgstpercent__gt=0) |
            Q(sgstpercent__gt=0) |
            Q(igstpercent__gt=0)
        )

        # 0% GST supplies (exactly 0)
        zero_filter = (
            Q(cgstpercent=0) &
            Q(sgstpercent=0) &
            Q(igstpercent=0)
        )

        # Nil-rated supplies (null or not provided)
        nil_filter = (
            Q(cgstpercent__isnull=True) &
            Q(sgstpercent__isnull=True) &
            Q(igstpercent__isnull=True)
        )

        data = {
            "Outward taxable supplies": fetch_and_aggregate(taxable_filter),
            "Outward supplies at 0% GST": fetch_and_aggregate(zero_filter),
            "Nil rated outward supplies": fetch_and_aggregate(nil_filter),
        }

        return Response(data)



class TrialBalanceView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    # ----- helpers ---------------------------------------------------------
    def _sum_debit(self, qs):
        return Coalesce(Sum(Case(When(drcr=True, then=F('amount')),
                                 default=Value(0),
                                 output_field=DecimalField(max_digits=14, decimal_places=2))),
                        Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)))

    def _sum_credit(self, qs):
        return Coalesce(Sum(Case(When(drcr=False, then=F('amount')),
                                 default=Value(0),
                                 output_field=DecimalField(max_digits=14, decimal_places=2))),
                        Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)))

    def _opening_by_account(self, entity_id, start_date, accounthead_id=None, drcr_filter=None):
        """
        Opening = sum of all postings strictly before start_date
        """
        qs = (JournalLine.objects
              .filter(entity_id=entity_id, entrydate__lt=start_date)
              .select_related('account', 'accounthead'))

        if accounthead_id:
            qs = qs.filter(accounthead_id=accounthead_id)

        # (optional) filter by side for opening
        if drcr_filter == 'dr':
            qs = qs.filter(drcr=True)
        elif drcr_filter == 'cr':
            qs = qs.filter(drcr=False)

        # aggregate per account
        agg = (qs.values('account_id', 'account__accountname', 'accounthead_id', 'accounthead__accounthead')
                 .annotate(
                     opening_debit=self._sum_debit(qs),
                     opening_credit=self._sum_credit(qs),
                 )
                 .annotate(
                     opening_balance=F('opening_debit') - F('opening_credit')
                 ))

        # return as dict keyed by account_id
        return {
            row['account_id']: {
                'account_id': row['account_id'],
                'accountname': row['account__accountname'],
                'accounthead_id': row['accounthead_id'],
                'accounthead': row['accounthead__accounthead'],
                'opening_debit': row['opening_debit'],
                'opening_credit': row['opening_credit'],
                'opening_balance': row['opening_balance'],
            }
            for row in agg
        }

    def _period_by_account(self, entity_id, start_date, end_date, accounthead_id=None, drcr_filter=None):
        """
        Movements in the date window (inclusive)
        """
        qs = (JournalLine.objects
              .filter(entity_id=entity_id, entrydate__range=(start_date, end_date))
              .select_related('account', 'accounthead'))

        if accounthead_id:
            qs = qs.filter(accounthead_id=accounthead_id)

        # (optional) filter by side for period totals
        if drcr_filter == 'dr':
            qs = qs.filter(drcr=True)
        elif drcr_filter == 'cr':
            qs = qs.filter(drcr=False)

        agg = (qs.values('account_id', 'account__accountname', 'accounthead_id', 'accounthead__accounthead')
                .annotate(
                    period_debit=self._sum_debit(qs),
                    period_credit=self._sum_credit(qs),
                )
                .annotate(
                    period_balance=F('period_debit') - F('period_credit')
                ))

        return {
            row['account_id']: {
                'account_id': row['account_id'],
                'accountname': row['account__accountname'],
                'accounthead_id': row['accounthead_id'],
                'accounthead': row['accounthead__accounthead'],
                'period_debit': row['period_debit'],
                'period_credit': row['period_credit'],
                'period_balance': row['period_balance'],
            }
            for row in agg
        }

    # ----- GET -------------------------------------------------------------
    def get(self, request, *args, **kwargs):
        entity_id = request.query_params.get('entity')
        start_date = request.query_params.get('startdate')
        end_date = request.query_params.get('enddate')
        accounthead_id = request.query_params.get('accounthead')  # accountHead PK (optional)
        drcr = request.query_params.get('drcr')  # 'dr' | 'cr' | None

        if not entity_id or not start_date or not end_date:
            return Response({'error': 'Missing required parameters: entity, startdate, enddate'}, status=400)

        start_date = parse_date(start_date)
        end_date = parse_date(end_date)

        # Opening and period dictionaries keyed by account_id
        opening = self._opening_by_account(entity_id, start_date, accounthead_id, drcr)
        period = self._period_by_account(entity_id, start_date, end_date, accounthead_id, drcr)

        # Merge
        account_ids = set(opening.keys()) | set(period.keys())
        result = []
        for acc_id in account_ids:
            o = opening.get(acc_id)
            p = period.get(acc_id)

            accountname = (o or p)['accountname']
            accounthead = (o or p)['accounthead']
            accounthead_id = (o or p)['accounthead_id']

            opening_debit = (o or {}).get('opening_debit', Decimal('0'))
            opening_credit = (o or {}).get('opening_credit', Decimal('0'))
            opening_balance = (o or {}).get('opening_balance', Decimal('0'))

            period_debit = (p or {}).get('period_debit', Decimal('0'))
            period_credit = (p or {}).get('period_credit', Decimal('0'))
            period_balance = (p or {}).get('period_balance', Decimal('0'))

            closing_balance = opening_balance + period_balance

            result.append({
                'account_id': acc_id,
                'accountname': accountname,
                'accounthead_id': accounthead_id,
                'accounthead': accounthead,

                'opening_debit': opening_debit,
                'opening_credit': opening_credit,
                'opening_balance': opening_balance,
                'obdrcr': 'DR' if opening_balance > 0 else 'CR' if opening_balance < 0 else '',

                'period_debit': period_debit,
                'period_credit': period_credit,
                'period_balance': period_balance,

                'closing_balance': closing_balance,
                'drcr': 'DR' if closing_balance > 0 else 'CR' if closing_balance < 0 else '',
            })

        # Optional post-filter by DR/CR on closing (kept for parity with your old view)
        if drcr == 'dr':
            result = [r for r in result if r['closing_balance'] > 0]
        elif drcr == 'cr':
            result = [r for r in result if r['closing_balance'] < 0]

        # Sort nicely by account head then account name
        result.sort(key=lambda r: (str(r['accounthead'] or ''), str(r['accountname'] or '')))

        return Response(result)



DZERO = Decimal("0.00")

def dec(x):
    # Coerce anything materialized (None/int/str/Decimal) to Decimal; reject expressions early.
    if x is None:
        return DZERO
    if isinstance(x, Decimal):
        return x
    # Strings like '0', ints, Decimals are fine. F/Value/CombinedExpression never reach here if used correctly.
    return Decimal(str(x))

class TrialbalanceApiViewJournal(ListAPIView):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TrialBalanceHeadRowSerializer

    def _parse_ymd(self, s: str) -> date:
        s = str(s or "").strip().replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
        return date.fromisoformat(s)

    def get(self, request, *args, **kwargs):
        entity_id = request.query_params.get("entity")
        start_s   = request.query_params.get("startdate")
        end_s     = request.query_params.get("enddate")

        if entity_id is None or start_s is None or end_s is None:
            return Response({"detail": "Query params required: entity, startdate, enddate (YYYY-MM-DD)"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            entity_id = int(str(entity_id).strip())
            startdate = self._parse_ymd(start_s)
            enddate   = self._parse_ymd(end_s)
        except Exception:
            return Response({"detail": "Dates/entity invalid.", "received": {"entity": entity_id, "startdate": start_s, "enddate": end_s}},
                            status=status.HTTP_400_BAD_REQUEST)

        if startdate > enddate:
            return Response([], status=status.HTTP_200_OK)

        fy = (entityfinancialyear.objects
              .filter(entity_id=entity_id, finstartyear__date__lte=enddate, finendyear__date__gte=startdate)
              .order_by("-finstartyear").first())
        if fy:
            startdate = max(startdate, fy.finstartyear.date())
            enddate   = min(enddate, fy.finendyear.date())
            if startdate > enddate:
                return Response([], status=status.HTTP_200_OK)

        opening_acct = (
            JournalLine.objects
            .filter(entity_id=entity_id, entrydate__lt=startdate)
            .values(
                "account_id",
                "account__accounthead_id", "account__accounthead__name",
                "account__creditaccounthead_id", "account__creditaccounthead__name",
                "accounthead_id", "accounthead__name",
            )
            .annotate(
                opening=Sum(
                    Case(
                        When(drcr=True, then=F("amount")),
                        When(drcr=False, then=-F("amount")),
                        default=V(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2),
                    ),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )
        )

        period_acct = (
            JournalLine.objects
            .filter(entity_id=entity_id, entrydate__gte=startdate, entrydate__lte=enddate)
            .values(
                "account_id",
                "account__accounthead_id", "account__accounthead__name",
                "account__creditaccounthead_id", "account__creditaccounthead__name",
                "accounthead_id", "accounthead__name",
            )
            .annotate(
                debit=Sum(
                    Case(When(drcr=True, then=F("amount")), default=V(0),
                         output_field=DecimalField(max_digits=18, decimal_places=2)),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
                credit=Sum(
                    Case(When(drcr=False, then=F("amount")), default=V(0),
                         output_field=DecimalField(max_digits=18, decimal_places=2)),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
            )
        )

        def _extract_heads(row):
            ah_id = row.get("account__accounthead_id") or row.get("accounthead_id")
            ah_nm = row.get("account__accounthead__name") or row.get("accounthead__name")
            ch_id = row.get("account__creditaccounthead_id") or ah_id
            ch_nm = row.get("account__creditaccounthead__name") or ah_nm
            return ah_id, ah_nm, ch_id, ch_nm

        per_acct = {}

        for r in opening_acct:
            aid = r["account_id"]
            ah_id, ah_nm, ch_id, ch_nm = _extract_heads(r)
            if ah_id is None and ch_id is None:
                continue
            per_acct[aid] = dict(
                opening=dec(r.get("opening")),
                debit=DZERO, credit=DZERO,
                ah_id=ah_id, ah_name=ah_nm or "",
                ch_id=ch_id, ch_name=ch_nm or "",
            )

        for r in period_acct:
            aid = r["account_id"]
            ah_id, ah_nm, ch_id, ch_nm = _extract_heads(r)
            if ah_id is None and ch_id is None:
                continue
            item = per_acct.setdefault(aid, dict(
                opening=DZERO, debit=DZERO, credit=DZERO,
                ah_id=ah_id, ah_name=ah_nm or "",
                ch_id=ch_id, ch_name=ch_nm or "",
            ))
            item["debit"]  = item["debit"]  + dec(r.get("debit"))
            item["credit"] = item["credit"] + dec(r.get("credit"))
            if not item["ah_name"] and ah_nm:
                item["ah_name"] = ah_nm
            if not item["ch_name"] and ch_nm:
                item["ch_name"] = ch_nm

        if not per_acct:
            return Response([], status=status.HTTP_200_OK)

        by_head = {}
        for v in per_acct.values():
            opening = dec(v.get("opening"))
            debit   = dec(v.get("debit"))
            credit  = dec(v.get("credit"))
            closing = opening + debit - credit  # pure Decimal
            hid, hname = (v["ah_id"], v["ah_name"]) if (closing >= DZERO) else (v["ch_id"], v["ch_name"])
            if hid is None:
                continue
            agg = by_head.setdefault(hid, dict(
                accounthead=hid,
                accountheadname=hname or "",
                openingbalance=DZERO, debit=DZERO, credit=DZERO,
            ))
            agg["openingbalance"] = agg["openingbalance"] + opening
            agg["debit"]          = agg["debit"]          + debit
            agg["credit"]         = agg["credit"]         + credit

        out = []
        for hid, v in by_head.items():
            opening = dec(v.get("openingbalance"))
            debit   = dec(v.get("debit"))
            credit  = dec(v.get("credit"))
            closing = opening + debit - credit
            v["closingbalance"] = closing
            v["drcr"]   = "CR" if closing < DZERO else "DR"
            v["obdrcr"] = "CR" if opening < DZERO else "DR"
            out.append(v)

        out.sort(key=lambda x: (x["accountheadname"] or "").lower())
        return Response(self.serializer_class(out, many=True).data, status=status.HTTP_200_OK)


# ---- numeric constants ----
ZERO = Decimal("0.00")
DEC0 = V(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))

# ----------------------
# Safe Decimal caster
# ----------------------
def _D(x) -> Decimal:
    """
    Coerce arbitrary input to Decimal safely:
    None/''/'—'/'NaN'/float NaN/Inf → 0
    '1,234.56' → 1234.56
    Decimal passthrough (guard bad states)
    """
    if x is None:
        return Decimal("0")
    if isinstance(x, Decimal):
        try:
            _ = x + Decimal(0)
            return x
        except Exception:
            return Decimal("0")
    if isinstance(x, int):
        return Decimal(x)
    if isinstance(x, float):
        if not isfinite(x):
            return Decimal("0")
        return Decimal(str(x))
    s = str(x).strip().replace(",", "")
    if not s or s in {"—", "-", "NaN", "nan", "None"}:
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")

# ----------------------
# Helper functions
# ----------------------
def _parse_iso_date(s: str) -> date:
    """Strict YYYY-MM-DD with dash normalization and trimming."""
    if s is None:
        raise ValueError("empty")
    s = str(s).strip().replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
    return date.fromisoformat(s)

def _bool(s: Optional[str], default: bool = False) -> bool:
    if s is None:
        return default
    return str(s).strip().lower() in {"1", "true", "yes", "y", "t"}

def _fy_covering_date(entity_id: int, d: date):
    """Return FY row covering date d (entityfinancialyear has finstartyear/finendyear)."""
    return (
        entityfinancialyear.objects
        .filter(entity_id=entity_id, finstartyear__date__lte=d, finendyear__date__gte=d)
        .order_by("finstartyear")
        .first()
    )

def _fys_overlapping_range(entity_id: int, d1: date, d2: date) -> List[entityfinancialyear]:
    return list(
        entityfinancialyear.objects
        .filter(entity_id=entity_id)
        .filter(finstartyear__date__lte=d2, finendyear__date__gte=d1)
        .order_by("finstartyear")
    )

def _clip_to_fy(d1: date, d2: date, fy: entityfinancialyear) -> Tuple[date, date]:
    fy_start = fy.finstartyear.date()
    fy_end   = fy.finendyear.date()
    return max(d1, fy_start), min(d2, fy_end)

def _fy_name(fy: entityfinancialyear) -> str:
    s = fy.finstartyear.date()
    e = fy.finendyear.date()
    return f"FY {s.year}-{str(e.year)[-2:]}"

def _aggregate_opening(base_qs, before_date: date) -> Decimal:
    """Opening = Σ(Dr − Cr) strictly before 'before_date'."""
    agg = base_qs.filter(entrydate__lt=before_date).aggregate(
        dr=Coalesce(Sum("amount", filter=Q(drcr=True)), DEC0),
        cr=Coalesce(Sum("amount", filter=Q(drcr=False)), DEC0),
    )
    return _D(agg["dr"]) - _D(agg["cr"])

def _fetch_period_rows(base_qs, d1: date, d2: date):
    return list(
        base_qs.filter(entrydate__gte=d1, entrydate__lte=d2)
               .order_by("entrydate", "voucherno", "id")
               .values("entrydate", "voucherno", "desc", "drcr", "amount")
    )

def _build_lines_and_totals(rows, opening: Decimal):
    running = _D(opening)
    out = []
    tot_dr = ZERO
    tot_cr = ZERO

    for r in rows:
        amt = _D(r["amount"])
        if r["drcr"]:
            debit, credit = amt, ZERO
            running += amt
            tot_dr += amt
        else:
            debit, credit = ZERO, amt
            running -= amt
            tot_cr += amt

        out.append({
            "date": r["entrydate"],              # ← serializer wants 'date'
            "voucherno": r.get("voucherno") or "",
            "desc": r.get("desc") or "",
            "debit": _D(debit),                  # ← 'debit'
            "credit": _D(credit),                # ← 'credit'
            "balance": _D(running),
            # Optional: keep txn if you have it upstream; harmless if absent.
            "transactiontype": r.get("transactiontype") or "",
        })

    closing = _D(opening) + _D(tot_dr) - _D(tot_cr)
    return out, _D(tot_dr), _D(tot_cr), _D(closing)

def _build_day_sections(lines, grand_opening, include_empty_days, from_dt, to_dt):
    """
    Builds daily summaries with correct day_opening, day_receipts, day_payments, and day_closing_balance.
    """

    # Group transactions by date
    grouped = defaultdict(list)
    for l in lines:
        grouped[l["date"]].append(l)

    # Sort the days
    all_days = sorted(grouped.keys())

    # If include_empty_days=True, fill missing days with empty items
    if include_empty_days:
        from datetime import timedelta
        day = from_dt
        while day <= to_dt:
            d_str = day.strftime("%d-%m-%Y")
            if d_str not in grouped:
                grouped[d_str] = []
                all_days.append(d_str)
            day += timedelta(days=1)
        all_days.sort()

    results = []
    current_balance = D(grand_opening)

    for day in all_days:
        items = grouped[day]
        day_receipts = D(sum(D(x.get("debit", 0) or 0) for x in items))
        day_payments = D(sum(D(x.get("credit", 0) or 0) for x in items))
        day_closing = current_balance + day_receipts - day_payments

        results.append({
            "date": day,
            "day_opening": D(current_balance),
            "day_receipts": D(day_receipts),
            "day_payments": D(day_payments),
            "day_closing_balance": D(day_closing),
            "items": items,
        })

        current_balance = day_closing

    return results

# ----------------------
# JSON API View
# ----------------------
class CashBookAPIView(APIView):
    """
    Unified Cash Book API (single or multi-FY via `sections[]`)

    GET /api/reports/cashbook?entity=1&from=YYYY-MM-DD&to=YYYY-MM-DD
      Optional:
        - account_id=INT
        - voucherno=JV123
        - txn=Sale,Receipt
        - desc_contains=rent
        - min_amount=100.00
        - max_amount=5000
        - include_empty_days=true|false (default false)
        - posted_only=true|false        (default true)
        - strict_fy=true|false          (default false)
    """
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        # required params
        entity_raw = request.query_params.get("entity")
        from_raw   = request.query_params.get("from")
        to_raw     = request.query_params.get("to")
        if entity_raw is None or from_raw is None or to_raw is None:
            return Response({"detail": "Query params required: entity, from, to (YYYY-MM-DD)"},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            entity_id = int(str(entity_raw).strip())
            from_dt   = _parse_iso_date(from_raw)
            to_dt     = _parse_iso_date(to_raw)
        except Exception:
            return Response({"detail": "Dates must be YYYY-MM-DD.",
                             "received": {"from": from_raw, "to": to_raw}}, status=400)
        if to_dt < from_dt:
            return Response({"detail": "'to' must be >= 'from'."}, status=400)

        # resolve account (default cash)
        account_q = request.query_params.get("account_id")
        if account_q:
            try:
                account_pk = int(account_q)
            except ValueError:
                return Response({"detail": "Invalid account_id."}, status=400)
        else:
            try:
                const = stocktransconstant()
                acct = const.getcashid(entity_id)
                account_pk = int(acct.pk) if hasattr(acct, "pk") else int(acct)
            except Exception as ex:
                return Response({"detail": f"Unable to resolve cash account for entity={entity_id}: {ex}"}, status=400)

        include_empty_days = _bool(request.query_params.get("include_empty_days"), default=False)
        posted_only        = _bool(request.query_params.get("posted_only"), default=True)
        strict_fy          = _bool(request.query_params.get("strict_fy"), default=False)

        # base queryset
        base = JournalLine.objects.filter(entity_id=entity_id, account_id=account_pk)
        # if posted_only: base = base.filter(entry__is_posted=True)

        # optional filters
        voucherno = request.query_params.get("voucherno")
        if voucherno:
            base = base.filter(voucherno=voucherno)

        txn_raw = request.query_params.get("txn")
        if txn_raw:
            txns = [t.strip() for t in txn_raw.split(",") if t.strip()]
            if txns:
                base = base.filter(transactiontype__in=txns)

        desc_contains = request.query_params.get("desc_contains")
        if desc_contains:
            base = base.filter(desc__icontains=desc_contains)

        min_amount = request.query_params.get("min_amount")
        if min_amount:
            try:
                base = base.filter(amount__gte=_D(min_amount))
            except Exception:
                return Response({"detail": "Invalid min_amount."}, status=400)

        max_amount = request.query_params.get("max_amount")
        if max_amount:
            try:
                base = base.filter(amount__lte=_D(max_amount))
            except Exception:
                return Response({"detail": "Invalid max_amount."}, status=400)

        # FY handling
        fy_from = _fy_covering_date(entity_id, from_dt)
        fy_to   = _fy_covering_date(entity_id, to_dt)

        def _single_section_payload() -> dict:
            grand_opening = _aggregate_opening(base, from_dt)
            rows = _fetch_period_rows(base, from_dt, to_dt)
            lines, tot_dr, tot_cr, closing = _build_lines_and_totals(rows, grand_opening)
            day_sections = _build_day_sections(lines, grand_opening, include_empty_days, from_dt, to_dt)
            return {
                "entity": entity_id,
                "account_id": account_pk,
                "from_date": from_dt,
                "to_date": to_dt,
                "spans_multiple_fy": False,
                "grand_opening": _D(grand_opening),
                "grand_total_receipts": _D(tot_dr),
                "grand_total_payments": _D(tot_cr),
                "grand_closing": _D(closing),
                "sections": [{
                    "from_date": from_dt,
                    "to_date": to_dt,
                    "opening_balance": _D(grand_opening),
                    "total_receipts": _D(tot_dr),
                    "total_payments": _D(tot_cr),
                    "closing_balance": _D(closing),
                    "lines": lines,
                    "day_sections": day_sections,
                    "fy_name": (_fy_name(fy_from) if fy_from and fy_to and fy_from.id == fy_to.id else None),
                    "fy_start": fy_from.finstartyear.date() if fy_from else None,
                    "fy_end": fy_from.finendyear.date() if fy_from else None,
                }],
            }

        if not fy_from or not fy_to:
            if strict_fy:
                return Response({"detail": "Requested dates are not fully covered by financial year setup.",
                                 "hint": "Add FY rows for these dates or call without strict_fy."}, status=400)
            payload = _single_section_payload()
            payload = _recompute_grands(payload)  # <— ADD THIS LINE

        elif fy_from.id == fy_to.id:
            payload = _single_section_payload()
            payload = _recompute_grands(payload)  # <— ADD THIS LINE

        else:
            # multi-FY
            fys = _fys_overlapping_range(entity_id, from_dt, to_dt)
            if not fys:
                return Response({"detail": "No financial year overlaps the requested range."}, status=400)

            sections = []
            rolling_opening = _D(_aggregate_opening(base, from_dt))
            grand_dr = _D(0)
            grand_cr = _D(0)
            last_closing = rolling_opening

            for fy in fys:
                sub_from, sub_to = _clip_to_fy(from_dt, to_dt, fy)
                rows = _fetch_period_rows(base, sub_from, sub_to)
                lines, tot_dr, tot_cr, closing = _build_lines_and_totals(rows, rolling_opening)
                day_sections = _build_day_sections(lines, rolling_opening, include_empty_days, sub_from, sub_to)

                sections.append({
                    "fy_name": _fy_name(fy),
                    "fy_start": fy.finstartyear.date(),
                    "fy_end": fy.finendyear.date(),
                    "from_date": sub_from,
                    "to_date": sub_to,
                    "opening_balance": _D(rolling_opening),
                    "total_receipts": _D(tot_dr),
                    "total_payments": _D(tot_cr),
                    "closing_balance": _D(closing),
                    "lines": lines,
                    "day_sections": day_sections,
                })
                grand_dr += _D(tot_dr)
                grand_cr += _D(tot_cr)
                rolling_opening = _D(closing)
                last_closing = _D(closing)

            payload = {
                "entity": entity_id,
                "account_id": account_pk,
                "from_date": from_dt,
                "to_date": to_dt,
                "spans_multiple_fy": True,
                "grand_opening": _D(_aggregate_opening(base, from_dt)),
                "grand_total_receipts": _D(grand_dr),
                "grand_total_payments": _D(grand_cr),
                "grand_closing": _D(last_closing),
                "sections": sections,
            }

            payload = _recompute_grands(payload)  # <— ADD

        return Response(CashbookUnifiedSerializer(payload).data, status=200)

# ----------------------
# Excel helpers & view
# ----------------------
_HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
DATE_FMT = "yyyy-mm-dd"
AMT_FMT = "#,##0.00;[Red]-#,##0.00"
COLS = [
    ("Date", 12),
    ("Voucher No", 16),
    ("Transaction", 16),
    ("Description", 50),
    ("Receipt (Dr)", 16),
    ("Payment (Cr)", 16),
    ("Running Balance", 18),
]

def _auto_width(ws, extra_pad: int = 1):
    for i, (_, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w + extra_pad

def _write_header(ws):
    ws.append([c[0] for c in COLS])
    for cell in ws[1]:
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

def _fmt_amount(cell):
    cell.number_format = AMT_FMT
    cell.alignment = _RIGHT

def _fmt_date(cell):
    cell.number_format = DATE_FMT
    cell.alignment = _CENTER

def _open_row(ws, opening):
    opening = _D(opening)
    ws.append([None, None, None, "Opening Balance",
               opening if opening >= 0 else None,
               None if opening >= 0 else -opening,
               opening])
    row = ws.max_row
    if opening >= 0:
        _fmt_amount(ws.cell(row=row, column=5))
    else:
        _fmt_amount(ws.cell(row=row, column=6))
    _fmt_amount(ws.cell(row=row, column=7))
    ws.cell(row=row, column=4).font = _BOLD

def _subtotal_row(ws, label: str, tot_dr, tot_cr, closing):
    tot_dr = _D(tot_dr)
    tot_cr = _D(tot_cr)
    closing = _D(closing)
    ws.append([None, None, None, label, tot_dr, tot_cr, closing])
    row = ws.max_row
    ws.cell(row=row, column=4).font = _BOLD
    for col in (5, 6, 7):
        c = ws.cell(row=row, column=col)
        _fmt_amount(c)
        c.font = _BOLD

def _write_lines(ws, lines: List[dict], opening):
    bal = _D(opening)
    for ln in (lines or []):
        d  = ln.get("date")                         # ← was 'entrydate'
        dr = _D(ln.get("debit"))                    # ← was 'debitamount'
        cr = _D(ln.get("credit"))                   # ← was 'creditamount'
        dr_cell = None if dr == 0 else dr
        cr_cell = None if cr == 0 else cr
        bal = bal + dr - cr

        ws.append([
            d,
            ln.get("voucherno") or "",
            ln.get("transactiontype") or "",        # safe if missing
            ln.get("desc") or "",
            dr_cell,
            cr_cell,
            bal
        ])
        row = ws.max_row
        if d:
            _fmt_date(ws.cell(row=row, column=1))
        _fmt_amount(ws.cell(row=row, column=5))
        _fmt_amount(ws.cell(row=row, column=6))
        _fmt_amount(ws.cell(row=row, column=7))

def _sheet_title_for_section(idx: int, sec: dict) -> str:
    base = f"{idx:02d}-" + (sec.get("fy_name") or f"{sec['from_date']}..{sec['to_date']}")
    return base[:31]

def _write_section_sheet(wb, idx: int, section: dict):
    ws = wb.create_sheet(title=_sheet_title_for_section(idx, section))
    _write_header(ws)
    ws.freeze_panes = "A2"

    opening = _D(section.get("opening_balance"))
    _open_row(ws, opening)

    _write_lines(ws, section.get("lines") or [], opening)

    tot_dr  = _D(section.get("total_receipts"))
    tot_cr  = _D(section.get("total_payments"))
    closing = _D(section.get("closing_balance"))
    _subtotal_row(ws, "Section Total / Closing", tot_dr, tot_cr, closing)

    _auto_width(ws)

def _recompute_grands(payload: dict) -> dict:
    """
    Ensure payload has correct grand_* fields.
    Uses section totals if present; if a section is missing totals, derives from its lines.
    """
    secs = payload.get("sections") or []
    if not secs:
        # nothing to do
        payload["grand_opening"] = _D(payload.get("grand_opening"))
        payload["grand_total_receipts"] = _D(payload.get("grand_total_receipts"))
        payload["grand_total_payments"] = _D(payload.get("grand_total_payments"))
        payload["grand_closing"] = _D(payload.get("grand_closing"))
        return payload

    grand_opening = _D(secs[0].get("opening_balance"))

    def _sec_totals(sec: dict):
        # Prefer section totals; if missing/zero with non-empty lines, compute from lines
        tr = _D(sec.get("total_receipts"))
        tp = _D(sec.get("total_payments"))

        if (tr == 0 and tp == 0) and (sec.get("lines")):
            # Derive from lines
            ls = sec.get("lines") or []
            tr = sum((_D(x.get("debit")) for x in ls), _D(0))
            tp = sum((_D(x.get("credit")) for x in ls), _D(0))
        return tr, tp

    tot_receipts = _D(0)
    tot_payments = _D(0)
    for s in secs:
        sr, sp = _sec_totals(s)
        tot_receipts += sr
        tot_payments += sp

    # closing: prefer last section's closing_balance; else compute
    last_sec = secs[-1]
    grand_closing = _D(last_sec.get("closing_balance"))
    if grand_closing == 0:
        grand_closing = grand_opening + tot_receipts - tot_payments

    payload["grand_opening"] = grand_opening
    payload["grand_total_receipts"] = tot_receipts
    payload["grand_total_payments"] = tot_payments
    payload["grand_closing"] = grand_closing
    return payload

def _section_numbers(sec: dict):
    """
    Returns (opening, receipts, payments, closing) for a section.
    If section totals are missing/zero but lines exist, compute from lines.
    Also derives closing from last line balance if available.
    """
    open_bal = _D(sec.get("opening_balance"))
    rec = _D(sec.get("total_receipts"))
    pay = _D(sec.get("total_payments"))
    cls = _D(sec.get("closing_balance"))

    lines = sec.get("lines") or []

    # If both receipts & payments are zero while we have lines, compute from lines
    if (rec == 0 and pay == 0) and lines:
        rec = sum((_D(l.get("debit")) for l in lines), _D(0))
        pay = sum((_D(l.get("credit")) for l in lines), _D(0))

    # Closing: prefer given, else last line balance, else opening + rec - pay
    if cls == 0 and lines:
        last_bal = _D(lines[-1].get("balance"))
        cls = last_bal if last_bal != 0 else (open_bal + rec - pay)
    elif cls == 0:
        cls = open_bal + rec - pay

    return open_bal, rec, pay, cls

def _write_summary_sheet(wb, payload: dict):
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Cash Book Summary"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _LEFT

    # ------------------------
    # Header metadata
    # ------------------------
    ws.append([])
    meta_rows = [
        ("Entity", payload.get("entity")),
        ("Account ID", payload.get("account_id")),
        ("From", payload.get("from_date")),
        ("To", payload.get("to_date")),
        ("Spans Multiple FY", "Yes" if payload.get("spans_multiple_fy") else "No"),
    ]
    for k, v in meta_rows:
        ws.append([k, v])

    # ------------------------
    # Grand Totals (top row)
    # ------------------------
    ws.append([])
    ws.append(["Grand Opening", "Total Receipts", "Total Payments", "Grand Closing"])
    for c in ws[ws.max_row]:
        c.font = _BOLD
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    ws.append([
        _D(payload.get("grand_opening")),
        _D(payload.get("grand_total_receipts")),
        _D(payload.get("grand_total_payments")),
        _D(payload.get("grand_closing")),
    ])
    r = ws.max_row
    for c in range(1, 5):
        _fmt_amount(ws.cell(row=r, column=c))

    # ------------------------
    # Per-FY / Section Table
    # ------------------------
    ws.append([])
    ws.append(["#", "FY", "From", "To", "Opening", "Receipts", "Payments", "Closing"])
    for c in ws[ws.max_row]:
        c.font = _BOLD
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    sections = payload.get("sections") or []
    for idx, sec in enumerate(sections, start=1):
        open_bal, receipts, payments, closing = _section_numbers(sec)

        ws.append([
            idx,
            sec.get("fy_name") or "-",
            sec.get("from_date"),
            sec.get("to_date"),
            open_bal,
            receipts,
            payments,
            closing,
        ])
        r = ws.max_row
        ws.cell(row=r, column=1).alignment = _CENTER
        ws.cell(row=r, column=3).number_format = DATE_FMT; ws.cell(row=r, column=3).alignment = _CENTER
        ws.cell(row=r, column=4).number_format = DATE_FMT; ws.cell(row=r, column=4).alignment = _CENTER
        for col in (5, 6, 7, 8):
            _fmt_amount(ws.cell(row=r, column=col))

    # ------------------------
    # Widths
    # ------------------------
    for col, w in zip("ABCDEFGH", [6, 22, 12, 12, 15, 15, 15, 15]):
        ws.column_dimensions[col].width = w


class CashBookExcelAPIView(APIView):
    """
    GET /api/reports/cashbook-xlsx?entity=1&from=YYYY-MM-DD&to=YYYY-MM-DD
      (same optional params as JSON view)
    """
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        # required
        entity_raw = request.query_params.get("entity")
        from_raw   = request.query_params.get("from")
        to_raw     = request.query_params.get("to")
        if entity_raw is None or from_raw is None or to_raw is None:
            return Response({"detail": "Query params required: entity, from, to (YYYY-MM-DD)"},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            entity_id = int(str(entity_raw).strip())
            from_dt   = _parse_iso_date(from_raw)
            to_dt     = _parse_iso_date(to_raw)
        except Exception:
            return Response({"detail": "Dates must be YYYY-MM-DD.",
                             "received": {"from": from_raw, "to": to_raw}}, status=400)
        if to_dt < from_dt:
            return Response({"detail": "'to' must be >= 'from'."}, status=400)

        # resolve account
        account_q = request.query_params.get("account_id")
        if account_q:
            try:
                account_pk = int(account_q)
            except ValueError:
                return Response({"detail": "Invalid account_id."}, status=400)
        else:
            try:
                const = stocktransconstant()
                acct = const.getcashid(entity_id)
                account_pk = int(acct.pk) if hasattr(acct, "pk") else int(acct)
            except Exception as ex:
                return Response({"detail": f"Unable to resolve cash account for entity={entity_id}: {ex}"}, status=400)

        include_empty_days = _bool(request.query_params.get("include_empty_days"), default=False)
        posted_only        = _bool(request.query_params.get("posted_only"), default=True)
        strict_fy          = _bool(request.query_params.get("strict_fy"), default=False)

        # base
        base = JournalLine.objects.filter(entity_id=entity_id, account_id=account_pk)
        # if posted_only: base = base.filter(entry__is_posted=True)

        # filters
        voucherno = request.query_params.get("voucherno")
        if voucherno:
            base = base.filter(voucherno=voucherno)
        txn_raw = request.query_params.get("txn")
        if txn_raw:
            txns = [t.strip() for t in txn_raw.split(",") if t.strip()]
            if txns:
                base = base.filter(transactiontype__in=txns)
        desc_contains = request.query_params.get("desc_contains")
        if desc_contains:
            base = base.filter(desc__icontains=desc_contains)
        min_amount = request.query_params.get("min_amount")
        if min_amount:
            try:
                base = base.filter(amount__gte=_D(min_amount))
            except Exception:
                return Response({"detail": "Invalid min_amount."}, status=400)
        max_amount = request.query_params.get("max_amount")
        if max_amount:
            try:
                base = base.filter(amount__lte=_D(max_amount))
            except Exception:
                return Response({"detail": "Invalid max_amount."}, status=400)

        # FY logic → payload
        fy_from = _fy_covering_date(entity_id, from_dt)
        fy_to   = _fy_covering_date(entity_id, to_dt)

        def _single_section_payload() -> dict:
            grand_opening = _aggregate_opening(base, from_dt)
            rows = _fetch_period_rows(base, from_dt, to_dt)
            lines, tot_dr, tot_cr, closing = _build_lines_and_totals(rows, grand_opening)
            day_sections = _build_day_sections(lines, grand_opening, include_empty_days, from_dt, to_dt)
            return {
                "entity": entity_id,
                "account_id": account_pk,
                "from_date": from_dt,
                "to_date": to_dt,
                "spans_multiple_fy": False,
                "grand_opening": _D(grand_opening),
                "grand_total_receipts": _D(tot_dr),
                "grand_total_payments": _D(tot_cr),
                "grand_closing": _D(closing),
                "sections": [{
                    "from_date": from_dt,
                    "to_date": to_dt,
                    "opening_balance": _D(grand_opening),
                    "total_receipts": _D(tot_dr),
                    "total_payments": _D(tot_cr),
                    "closing_balance": _D(closing),
                    "lines": lines,
                    "day_sections": day_sections,
                    "fy_name": (_fy_name(fy_from) if fy_from and fy_to and fy_from.id == fy_to.id else None),
                    "fy_start": fy_from.finstartyear.date() if fy_from else None,
                    "fy_end": fy_from.finendyear.date() if fy_from else None,
                }],
            }

        if not fy_from or not fy_to:
            if strict_fy:
                return Response({"detail": "Requested dates are not fully covered by financial year setup.",
                                 "hint": "Add FY rows for these dates or call without strict_fy."}, status=400)
            payload = _single_section_payload()
            payload = _recompute_grands(payload)  # <— ADD THIS LINE

        elif fy_from.id == fy_to.id:
            payload = _single_section_payload()
            payload = _recompute_grands(payload)  # <— ADD THIS LINE

        else:
            fys = _fys_overlapping_range(entity_id, from_dt, to_dt)
            if not fys:
                return Response({"detail": "No financial year overlaps the requested range."}, status=400)

            sections = []
            rolling_opening = _D(_aggregate_opening(base, from_dt))
            grand_dr = _D(0); grand_cr = _D(0)
            last_closing = rolling_opening

            for fy in fys:
                sub_from, sub_to = _clip_to_fy(from_dt, to_dt, fy)
                rows = _fetch_period_rows(base, sub_from, sub_to)
                lines, tot_dr, tot_cr, closing = _build_lines_and_totals(rows, rolling_opening)
                day_sections = _build_day_sections(lines, rolling_opening, include_empty_days, sub_from, sub_to)
                sections.append({
                    "fy_name": _fy_name(fy),
                    "fy_start": fy.finstartyear.date(),
                    "fy_end": fy.finendyear.date(),
                    "from_date": sub_from,
                    "to_date": sub_to,
                    "opening_balance": _D(rolling_opening),
                    "total_receipts": _D(tot_dr),
                    "total_payments": _D(tot_cr),
                    "closing_balance": _D(closing),
                    "lines": lines,
                    "day_sections": day_sections,
                })
                grand_dr += _D(tot_dr); grand_cr += _D(tot_cr)
                rolling_opening = _D(closing); last_closing = _D(closing)

            payload = {
                "entity": entity_id,
                "account_id": account_pk,
                "from_date": from_dt,
                "to_date": to_dt,
                "spans_multiple_fy": True,
                "grand_opening": _D(_aggregate_opening(base, from_dt)),
                "grand_total_receipts": _D(grand_dr),
                "grand_total_payments": _D(grand_cr),
                "grand_closing": _D(last_closing),
                "sections": sections,
            }

            payload = _recompute_grands(payload)  # <— ADD

        # ----- Build workbook & return -----
        wb = Workbook()  # normal mode
        _write_summary_sheet(wb, payload)
        for i, sec in enumerate(payload["sections"], start=1):
            _write_section_sheet(wb, i, sec)

        fname = f"CashBook_{entity_id}_{from_dt}_to_{to_dt}.xlsx"
        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(resp)
        return resp


def _section_numbers(sec: dict):
    """
    Returns (opening, receipts, payments, closing) for a section.
    Uses section totals if present; if missing/zero with lines present, derive from lines.
    Closing prefers provided; else last line.balance; else open + receipts - payments.
    """
    def D(x): return _D(x)

    open_bal = D(sec.get("opening_balance"))
    rec = D(sec.get("total_receipts"))
    pay = D(sec.get("total_payments"))
    cls = D(sec.get("closing_balance"))
    lines = sec.get("lines") or []

    if (rec == 0 and pay == 0) and lines:
        rec = sum((D(l.get("debit")) for l in lines), D(0))
        pay = sum((D(l.get("credit")) for l in lines), D(0))

    if cls == 0:
        if lines:
            last_bal = D(lines[-1].get("balance"))
            cls = last_bal if last_bal != 0 else (open_bal + rec - pay)
        else:
            cls = open_bal + rec - pay

    return open_bal, rec, pay, cls


def _fmt2(x: Decimal) -> str:
    """#,##0.00 with minus sign if negative."""
    return f"{_D(x):,.2f}"


def _cashbook_pdf_header_footer(canvas, doc):
    canvas.saveState()
    w, h = landscape(A4)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(40, h - 25, "Cash Book Report")
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(w - 40, 25, f"Page {canvas.getPageNumber()}")
    canvas.restoreState()


class CashBookPDFAPIView(APIView):
    """
    GET /api/reports/cashbook-pdf?entity=1&from=YYYY-MM-DD&to=YYYY-MM-DD[&print=grand|sections|both]

    - print=grand     → only summary page
    - print=sections  → only per-section detail tables
    - print=both      → both (default)
    """
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, *args, **kwargs):
        # 1) Build payload by reusing your JSON view logic
        cb_view = CashBookAPIView()
        resp = cb_view.get(request)
        if resp.status_code != 200:
            return resp
        payload = resp.data

        # 2) Ensure grand_* totals are correct (derive from sections/lines if needed)
        payload = _recompute_grands(payload)

        print_mode = (request.query_params.get("print") or "both").strip().lower()
        if print_mode not in {"grand", "sections", "both"}:
            print_mode = "both"

        # 3) Prepare PDF doc
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=40
        )
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle("title", parent=styles["Heading1"],
                                     alignment=1, fontSize=14, leading=16, spaceAfter=12)
        style_h3 = ParagraphStyle("h3", parent=styles["Heading3"], fontSize=11, spaceAfter=6)
        style_norm = ParagraphStyle("norm", parent=styles["Normal"], fontSize=9, leading=11)
        style_th = ParagraphStyle("th", parent=styles["Normal"], alignment=1, fontSize=9)

        elems = []

        # ---------- Summary (Grand totals) ----------
        if print_mode in {"grand", "both"}:
            title = f"Cash Book - Entity {payload['entity']}"
            elems.append(Paragraph(title, style_title))
            elems.append(Paragraph(f"From {payload['from_date']} to {payload['to_date']}", style_norm))
            elems.append(Spacer(1, 10))

            # Grand totals row — these are guaranteed correct by _recompute_grands
            grand = [
                ["Grand Opening", "Total Receipts", "Total Payments", "Grand Closing"],
                [
                    _fmt2(payload["grand_opening"]),
                    _fmt2(payload["grand_total_receipts"]),
                    _fmt2(payload["grand_total_payments"]),
                    _fmt2(payload["grand_closing"]),
                ],
            ]
            t = Table(grand, colWidths=[2.0*inch]*4)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elems.append(t)
            elems.append(Spacer(1, 14))

            # Per-section summary table
            sections = payload.get("sections") or []
            if sections:
                data = [["#", "FY", "From", "To", "Opening", "Receipts", "Payments", "Closing"]]
                for i, sec in enumerate(sections, start=1):
                    open_bal, rec, pay, cls = _section_numbers(sec)
                    data.append([
                        i,
                        sec.get("fy_name") or "-",
                        str(sec.get("from_date") or ""),
                        str(sec.get("to_date") or ""),
                        _fmt2(open_bal),
                        _fmt2(rec),
                        _fmt2(pay),
                        _fmt2(cls),
                    ])
                st = Table(data, repeatRows=1,
                           colWidths=[0.5*inch, 1.6*inch, 1.1*inch, 1.1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
                st.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ]))
                elems.append(st)

            # If we are also printing sections, new page after summary
            if print_mode == "both":
                elems.append(PageBreak())

        # ---------- Sections (detail tables) ----------
        if print_mode in {"sections", "both"}:
            sections = payload.get("sections") or []
            for idx, sec in enumerate(sections, start=1):
                elems.append(Paragraph(
                    f"Section {idx}: {sec.get('fy_name') or '-'}",
                    style_h3
                ))
                elems.append(Paragraph(
                    f"Period: {sec.get('from_date')} to {sec.get('to_date')}",
                    style_norm
                ))
                elems.append(Spacer(1, 6))

                # Build table rows (repeat header on each page automatically)
                header = ["Date", "Voucher No", "Description", "Debit (₹)", "Credit (₹)", "Running Balance (₹)"]
                rows = [header]

                for ln in sec.get("lines") or []:
                    rows.append([
                        str(ln.get("date") or ""),
                        ln.get("voucherno") or "",
                        Paragraph(ln.get("desc") or "", style_norm),
                        _fmt2(ln.get("debit")),
                        _fmt2(ln.get("credit")),
                        _fmt2(ln.get("balance")),
                    ])

                # Section totals (computed robustly)
                open_bal, rec, pay, cls = _section_numbers(sec)
                rows.append(["", "", Paragraph("<b>Totals:</b>", style_th),
                             _fmt2(rec), _fmt2(pay), _fmt2(cls)])

                tbl = Table(
                    rows, repeatRows=1,
                    colWidths=[1.0*inch, 1.2*inch, 4.0*inch, 1.1*inch, 1.1*inch, 1.3*inch]
                )
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]))
                elems.append(tbl)

                # Page break between sections
                if idx < len(sections):
                    elems.append(PageBreak())

        # 4) Build and return
        doc.build(elems, onFirstPage=_cashbook_pdf_header_footer, onLaterPages=_cashbook_pdf_header_footer)
        buf.seek(0)
        filename = f"cashbook_{payload['entity']}_{payload['from_date']}_{payload['to_date']}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return resp


DZERO = Decimal("0.00")  # Python-side zero ONLY for arithmetic & comparisons


def dec(x):
    """
    Coerce a materialized DB value to Decimal.
    Avoids passing any Expressions (F/Value/CombinedExpression) here.
    """
    if x is None:
        return DZERO
    if isinstance(x, Decimal):
        return x
    return Decimal(str(x))


class TrialbalanceApiViewJournalByAccount(ListAPIView):
    """
    GET /api/reports/trial-balance/accounts/?entity=1[&accounthead=10]&startdate=2025-04-01&enddate=2025-04-30

    Emits Trial Balance rows *per account*. Display head is chosen per-account:
      closing >= 0 → account.accounthead, else → account.creditaccounthead.

    If ?accounthead is passed → filter to that head; else → include all heads.
    """
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TrialBalanceAccountRowSerializer

    def _parse_ymd(self, s: str) -> date:
        s = str(s or "").strip().replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
        return date.fromisoformat(s)  # strict YYYY-MM-DD

    def get(self, request, *args, **kwargs):
        entity_id = request.query_params.get("entity")
        head_id_s = request.query_params.get("accounthead")  # optional now
        start_s   = request.query_params.get("startdate")
        end_s     = request.query_params.get("enddate")

        # --- Validate required params (head is optional) ---
        if not (entity_id and start_s and end_s):
            return Response(
                {"detail": "Required: entity, startdate, enddate (YYYY-MM-DD). Optional: accounthead"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            entity_id = int(str(entity_id).strip())
            head_id   = int(str(head_id_s).strip()) if head_id_s not in (None, "",) else None
            startdate = self._parse_ymd(start_s)
            enddate   = self._parse_ymd(end_s)
        except Exception:
            return Response(
                {"detail": "Invalid inputs.", "received": {
                    "entity": request.query_params.get("entity"),
                    "accounthead": head_id_s, "startdate": start_s, "enddate": end_s}},
                status=status.HTTP_400_BAD_REQUEST
            )

        if startdate > enddate:
            return Response([], status=status.HTTP_200_OK)

        # --- Clamp to FY if present ---
        fy = (
            entityfinancialyear.objects
            .filter(entity_id=entity_id,
                    finstartyear__date__lte=enddate,
                    finendyear__date__gte=startdate)
            .order_by("-finstartyear")
            .first()
        )
        if fy:
            startdate = max(startdate, fy.finstartyear.date())
            enddate   = min(enddate, fy.finendyear.date())
            if startdate > enddate:
                return Response([], status=status.HTTP_200_OK)

        # ---- Opening (< startdate) per account ----
        opening_qs = (
            JournalLine.objects
            .filter(entity_id=entity_id, entrydate__lt=startdate)
            .exclude(account_id__isnull=True)
            .values(
                "account_id", "account__accountname",
                "account__accounthead_id", "account__accounthead__name",
                "account__creditaccounthead_id", "account__creditaccounthead__name",
            )
            .annotate(
                opening=Sum(
                    Case(
                        When(drcr=True, then=F("amount")),
                        When(drcr=False, then=-F("amount")),
                        default=V(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2),
                    ),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )
        )

        # ---- Period ([start,end]) per account ----
        period_qs = (
            JournalLine.objects
            .filter(entity_id=entity_id, entrydate__gte=startdate, entrydate__lte=enddate)
            .exclude(account_id__isnull=True)
            .values(
                "account_id", "account__accountname",
                "account__accounthead_id", "account__accounthead__name",
                "account__creditaccounthead_id", "account__creditaccounthead__name",
            )
            .annotate(
                debit=Sum(
                    Case(
                        When(drcr=True, then=F("amount")),
                        default=V(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2),
                    ),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
                credit=Sum(
                    Case(
                        When(drcr=False, then=F("amount")),
                        default=V(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2),
                    ),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
            )
        )

        # ---- Merge per account ----
        per_account = {}
        for r in opening_qs:
            aid = r["account_id"]
            per_account[aid] = dict(
                account=aid,
                accountname=r.get("account__accountname") or "",
                opening=dec(r.get("opening")),
                debit=DZERO, credit=DZERO,
                ah_id=r.get("account__accounthead_id"),
                ah_name=r.get("account__accounthead__name") or "",
                ch_id=r.get("account__creditaccounthead_id"),
                ch_name=r.get("account__creditaccounthead__name") or "",
            )

        for r in period_qs:
            aid = r["account_id"]
            item = per_account.setdefault(aid, dict(
                account=aid,
                accountname=r.get("account__accountname") or "",
                opening=DZERO, debit=DZERO, credit=DZERO,
                ah_id=r.get("account__accounthead_id"),
                ah_name=r.get("account__accounthead__name") or "",
                ch_id=r.get("account__creditaccounthead_id"),
                ch_name=r.get("account__creditaccounthead__name") or "",
            ))
            item["debit"]  = item["debit"]  + dec(r.get("debit"))
            item["credit"] = item["credit"] + dec(r.get("credit"))

        if not per_account:
            return Response([], status=status.HTTP_200_OK)

        # ---- Choose head by sign; optionally filter by ?accounthead ----
        rows = []
        for v in per_account.values():
            opening = dec(v.get("opening"))
            debit   = dec(v.get("debit"))
            credit  = dec(v.get("credit"))
            closing = opening + debit - credit

            disp_head_id, disp_head_name = (
                (v["ah_id"], v["ah_name"]) if (closing >= DZERO) else (v["ch_id"], v["ch_name"])
            )

            # If caller passed a head → filter; else include all
            if head_id is not None:
                if disp_head_id is None or int(disp_head_id) != head_id:
                    continue

            rows.append(dict(
                account=v["account"],
                accountname=v["accountname"],
                # if head_id was not passed, emit the chosen head for this row
                accounthead=(int(disp_head_id) if disp_head_id is not None else None),
                accountheadname=disp_head_name or "",
                openingbalance=opening,
                debit=debit,
                credit=credit,
                closingbalance=closing,
                drcr=("CR" if closing < DZERO else "DR"),
                obdrcr=("CR" if opening < DZERO else "DR"),
            ))

        rows.sort(key=lambda x: (x["accountname"] or "").lower())
        return Response(self.serializer_class(rows, many=True).data, status=status.HTTP_200_OK)

DEC = DecimalField(max_digits=18, decimal_places=2)               # reuse this
VZ  = lambda: V(ZERO, output_field=DEC)                           # Decimal zero Value()
def to_dec(x):
    """Coerce ORM values (including Decimal, int, str, None) to Python Decimal."""
    if x is None:
        return DZERO
    if isinstance(x, Decimal):
        return x
    # Avoid passing Django expressions here; this function is used on materialized values only.
    return Decimal(str(x))


class TrialbalanceApiViewJournalByAccountLedger(ListAPIView):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TrialBalanceAccountLedgerRowSerializer

    def _parse_ymd(self, s: str) -> date:
        s = str(s or "").strip().replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
        return date.fromisoformat(s)

    def get(self, request, *args, **kwargs):
        entity_id  = request.query_params.get("entity")
        account_id = request.query_params.get("account")
        start_s    = request.query_params.get("startdate")
        end_s      = request.query_params.get("enddate")

        if not (entity_id and account_id and start_s and end_s):
            return Response(
                {"detail": "Required: entity, account, startdate, enddate (YYYY-MM-DD)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            entity_id  = int(str(entity_id).strip())
            account_id = int(str(account_id).strip())
            startdate  = self._parse_ymd(start_s)
            enddate    = self._parse_ymd(end_s)
        except Exception:
            return Response(
                {"detail": "Invalid inputs.", "received": {
                    "entity": request.query_params.get("entity"),
                    "account": request.query_params.get("account"),
                    "startdate": start_s, "enddate": end_s
                }},
                status=status.HTTP_400_BAD_REQUEST
            )

        if startdate > enddate:
            return Response([], status=status.HTTP_200_OK)

        # ---- Clamp to FY, if any ----
        fy = (
            entityfinancialyear.objects
            .filter(entity_id=entity_id,
                    finstartyear__date__lte=enddate,
                    finendyear__date__gte=startdate)
            .order_by("-finstartyear")
            .first()
        )
        if fy:
            startdate = max(startdate, fy.finstartyear.date())
            enddate   = min(enddate, fy.finendyear.date())
            if startdate > enddate:
                return Response([], status=status.HTTP_200_OK)

        # =========================================
        # 1) Opening balance (< startdate) -> Python Decimal
        # =========================================
        opening_sum_raw = (
            JournalLine.objects
            .filter(entity_id=entity_id, account_id=account_id, entrydate__lt=startdate)
            .aggregate(opening=Coalesce(
                Sum(
                    Case(
                        When(drcr=True,  then=F("amount")),   # Debit +
                        When(drcr=False, then=-F("amount")),  # Credit −
                        default=V(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2),
                    ),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
                V(0),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ))
        )["opening"]

        opening_sum = to_dec(opening_sum_raw)

        # =========================================
        # Account name (best-effort)
        # =========================================
        name_row = (
            JournalLine.objects
            .filter(entity_id=entity_id, account_id=account_id)
            .values("account__accountname")
            .order_by("entrydate", "id")
            .first()
        )
        accountname = (name_row or {}).get("account__accountname") or ""

        rows = []

        # Opening synthetic row (include even if zero to show running start)
        obal = opening_sum
        opening_row = dict(
            account=account_id,
            accountname=accountname,
            sortdate=startdate,                             # date object; serializer can handle
            entrydate=startdate.strftime("%d-%m-%Y"),       # display string
            narration="Opening Balance",
            transactiontype="OPENING",
            transactionid=0,                                # sentinel
            debit=(obal if obal > DZERO else DZERO),
            credit=(-obal if obal < DZERO else DZERO),      # positive credit display
            runningbalance=obal,
        )
        rows.append(opening_row)

        # =========================================
        # 2) Period lines (detail-level)
        # =========================================
        values_list = [
            "id", "account_id", "account__accountname", "entrydate",
            "drcr", "amount",
            "transactiontype", "transactionid", "voucherno",
        ]
        for opt in ("desc", "narration", "notes"):
            try:
                JournalLine._meta.get_field(opt)
                values_list.append(opt)
            except Exception:
                pass

        lines_qs = (
            JournalLine.objects
            .filter(entity_id=entity_id, account_id=account_id,
                    entrydate__gte=startdate, entrydate__lte=enddate)
            .values(*values_list)
            .order_by("entrydate", "id")
        )

        def pick_first(row, candidates):
            for c in candidates:
                if c in row and row[c] not in (None, "", " "):
                    return str(row[c])
            return ""

        running = obal  # Decimal
        for r in lines_qs:
            amt = to_dec(r.get("amount"))

            is_dr = bool(r["drcr"])
            debit  = amt if is_dr else DZERO
            credit = amt if not is_dr else DZERO
            running = running + debit - credit  # pure Decimal math

            narration = pick_first(r, ["desc", "narration", "notes"])
            txn_type = (r.get("transactiontype") or "").strip() or "UNKNOWN"

            jl_txn_id = r.get("transactionid")
            try:
                txn_id = int(jl_txn_id) if jl_txn_id is not None and str(jl_txn_id).strip() != "" else int(r["id"])
            except Exception:
                txn_id = int(r["id"])

            rows.append(dict(
                account=r["account_id"],
                accountname=r.get("account__accountname") or accountname,
                sortdate=r["entrydate"],                        # date object
                entrydate=r["entrydate"].strftime("%d-%m-%Y"),  # display string
                narration=narration or "",
                transactiontype=txn_type,
                transactionid=txn_id,
                debit=debit,
                credit=credit,
                runningbalance=running,
            ))

        return Response(self.serializer_class(rows, many=True).data, status=status.HTTP_200_OK)

DEC18_2 = DecimalField(max_digits=18, decimal_places=2)
ZERO_D = Decimal("0.00")
ZERO = V(ZERO_D, output_field=DEC18_2)

class LedgerSummaryJournalline(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    ORDER_MAP = {
        "accountname": "account__accountname",
        "-accountname": "-account__accountname",
        "head_name": "account__accounthead__name",
        "-head_name": "-account__accounthead__name",
        "opening": "openingbalance",
        "-opening": "-openingbalance",
        "debit": "debit",
        "-debit": "-debit",
        "credit": "credit",
        "-credit": "-credit",
        "period_net": "period_net",
        "-period_net": "-period_net",
        "abs_movement": "abs_movement",
        "-abs_movement": "-abs_movement",
        "balancetotal": "balancetotal",
        "-balancetotal": "-balancetotal",
    }

    def post(self, request, *_args, **_kwargs):
        # 1) Validate input
        in_ser = LedgerSummaryRequestSerializer(data=request.data)
        in_ser.is_valid(raise_exception=True)
        p = in_ser.validated_data

        entity_id = p["entity"]
        startdate = p["startdate"]
        enddate = p["enddate"]
        group_by = p["group_by"]

        # 2) Base queryset (up to enddate)
        qs = JournalLine.objects.filter(
            entity_id=entity_id,
            account__isnull=False,
            entrydate__lte=enddate,
        )

        # 3) Optional filters (apply to qs)
        if p.get("txn_in"):
            include_txn = [x.strip() for x in str(p["txn_in"]).split(",") if x.strip()]
            qs = qs.filter(transactiontype__in=include_txn)

        if p.get("voucherno"):
            qs = qs.filter(voucherno=str(p["voucherno"]))
        if p.get("vno_contains"):
            qs = qs.filter(voucherno__icontains=str(p["vno_contains"]))
        if p.get("desc_contains"):
            qs = qs.filter(desc__icontains=str(p["desc_contains"]))

        if p.get("accounthead"):
            ah_ids = [int(x) for x in str(p["accounthead"]).split(",") if x.strip()]
            qs = qs.filter(account__accounthead_id__in=ah_ids)
        else:
            ah_ids = None

        if p.get("account"):
            acc_ids = [int(x) for x in str(p["account"]).split(",") if x.strip()]
            qs = qs.filter(account_id__in=acc_ids)
        else:
            acc_ids = None

        # 4) Aggregations (one grouped query)
        period_f = Q(entrydate__gte=startdate, entrydate__lte=enddate)

        opening_expr = Coalesce(
            Sum(
                Case(
                    When(drcr=True, then=F("amount")),
                    When(drcr=False, then=-F("amount")),
                    default=V(0),
                    output_field=DEC18_2,
                ),
                filter=Q(entrydate__lt=startdate),
                output_field=DEC18_2,
            ),
            ZERO,
        )
        debit_expr = Coalesce(
            Sum(F("amount"), filter=period_f & Q(drcr=True), output_field=DEC18_2),
            ZERO,
        )
        credit_expr = Coalesce(
            Sum(F("amount"), filter=period_f & Q(drcr=False), output_field=DEC18_2),
            ZERO,
        )
        period_net_expr = ExpressionWrapper(F("debit") - F("credit"), output_field=DEC18_2)
        abs_mov_expr = ExpressionWrapper(F("debit") + F("credit"), output_field=DEC18_2)
        closing_expr = ExpressionWrapper(F("openingbalance") + F("debit") - F("credit"), output_field=DEC18_2)

        # Group fields
        if group_by == "head":
            group_vals = ["account__accounthead_id", "account__accounthead__name"]
        else:
            group_vals = [
                "account_id",
                "account__accountname",
                "account__accounthead_id",
                "account__accounthead__name",
            ]

        qs = (
            qs.values(*group_vals)
              .annotate(
                  openingbalance=opening_expr,
                  debit=debit_expr,
                  credit=credit_expr,
              )
              .annotate(
                  period_net=period_net_expr,
                  abs_movement=abs_mov_expr,
                  balancetotal=closing_expr,
                  txn_count=Count("id", filter=period_f),
                  last_txn_date=Max("entrydate", filter=period_f),
              )
              .annotate(
                  drcr=Case(When(balancetotal__lt=0, then=V("CR")), default=V("DR")),
                  obdrcr=Case(When(openingbalance__lt=0, then=V("CR")), default=V("DR")),
              )
        )

        # 5) Summary-focused filters
        sign = p.get("sign", "ALL")
        if sign == "DR":
            qs = qs.filter(balancetotal__gt=0)
        elif sign == "CR":
            qs = qs.filter(balancetotal__lt=0)

        if not p.get("include_zero", False):
            qs = qs.exclude(balancetotal=ZERO_D)

        if p.get("min_activity") is not None:
            qs = qs.filter(abs_movement__gte=p["min_activity"])

        if p.get("amount_min") is not None:
            qs = qs.filter(balancetotal__gte=p["amount_min"])
        if p.get("amount_max") is not None:
            qs = qs.filter(balancetotal__lte=p["amount_max"])

        # New: range_min / range_max on balancetotal (inclusive)
        if p.get("range_min") is not None:
            qs = qs.filter(balancetotal__gte=p["range_min"])
        if p.get("range_max") is not None:
            qs = qs.filter(balancetotal__lte=p["range_max"])

        # 6) Ordering (safe per grouping)
        order_key = (p.get("order_by") or "").strip() or ("head_name" if group_by == "head" else "accountname")
        order_field = self.ORDER_MAP.get(order_key)
        if group_by == "head" and order_field in ("account__accountname", "-account__accountname"):
            order_field = "account__accounthead__name"
        if not order_field:
            order_field = "account__accounthead__name" if group_by == "head" else "account__accountname"
        qs = qs.order_by(order_field)

        # 7) GRAND TOTALS — recompute from JournalLine with same base filters (pre-agg filters)
        totals_qs = JournalLine.objects.filter(
            entity_id=entity_id,
            account__isnull=False,
            entrydate__lte=enddate,
        )
        if p.get("txn_in"):
            totals_qs = totals_qs.filter(transactiontype__in=include_txn)
        if p.get("voucherno"):
            totals_qs = totals_qs.filter(voucherno=str(p["voucherno"]))
        if p.get("vno_contains"):
            totals_qs = totals_qs.filter(voucherno__icontains=str(p["vno_contains"]))
        if p.get("desc_contains"):
            totals_qs = totals_qs.filter(desc__icontains=str(p["desc_contains"]))
        if ah_ids:
            totals_qs = totals_qs.filter(account__accounthead_id__in=ah_ids)
        if acc_ids:
            totals_qs = totals_qs.filter(account_id__in=acc_ids)

        opening_total_expr = Coalesce(
            Sum(
                Case(
                    When(drcr=True, then=F("amount")),
                    When(drcr=False, then=-F("amount")),
                    default=V(0),
                    output_field=DEC18_2,
                ),
                filter=Q(entrydate__lt=startdate),
                output_field=DEC18_2,
            ),
            ZERO,
        )
        debit_total_expr = Coalesce(
            Sum(F("amount"), filter=Q(entrydate__gte=startdate, entrydate__lte=enddate) & Q(drcr=True), output_field=DEC18_2),
            ZERO,
        )
        credit_total_expr = Coalesce(
            Sum(F("amount"), filter=Q(entrydate__gte=startdate, entrydate__lte=enddate) & Q(drcr=False), output_field=DEC18_2),
            ZERO,
        )

        totals_raw = totals_qs.aggregate(
            opening=opening_total_expr,
            debit=debit_total_expr,
            credit=credit_total_expr,
        )
        opening_t = totals_raw["opening"] or ZERO_D
        debit_t = totals_raw["debit"] or ZERO_D
        credit_t = totals_raw["credit"] or ZERO_D
        totals = {
            "opening": opening_t,
            "debit": debit_t,
            "credit": credit_t,
            "closing": opening_t + debit_t - credit_t,
        }

        # 8) Pagination
        paginator = SimpleNumberPagination()
        paginator.page_size = p["pagesize"]
        page = paginator.paginate_queryset(list(qs), request)

        # 9) Shape rows — ensure all keys exist, null when not applicable
        rows = []
        for r in page:
            # head fields
            head_id = r.get("account__accounthead_id") if group_by != "head" else r.get("account__accounthead_id")
            head_name = r.get("account__accounthead__name")

            if group_by == "head":
                # values() only included head fields
                head_id = r.get("account__accounthead_id")
                head_name = r.get("account__accounthead__name")
                account_id = None
                account_name = None
                links = None
            elif group_by == "account":
                head_id = r.get("account__accounthead_id")
                head_name = r.get("account__accounthead__name")
                account_id = r.get("account_id")
                account_name = r.get("account__accountname") or ""
                links = {
                    "detail": f"/api/reports/ledger-detail?entity={entity_id}&account={account_id}&from={startdate}&to={enddate}"
                }
            else:  # head_account
                head_id = r.get("account__accounthead_id")
                head_name = r.get("account__accounthead__name")
                account_id = r.get("account_id")
                account_name = r.get("account__accountname") or ""
                links = {
                    "detail": f"/api/reports/ledger-detail?entity={entity_id}&account={account_id}&from={startdate}&to={enddate}"
                }

            row = dict(
                head_id=head_id,
                head_name=head_name,
                account=account_id,
                accountname=account_name,

                openingbalance=r["openingbalance"],
                debit=r["debit"],
                credit=r["credit"],
                period_net=r["period_net"],
                abs_movement=r["abs_movement"],
                balancetotal=r["balancetotal"],
                drcr=r["drcr"],
                obdrcr=r["obdrcr"],

                txn_count=r["txn_count"],
                last_txn_date=r["last_txn_date"],

                links=links,
            )
            rows.append(row)

        out_ser = LedgerSummaryRowSerializer(rows, many=True)

        # 10) Meta echo & paginated response
        meta_echo = {
            "entity": entity_id,
            "from": str(startdate),
            "to": str(enddate),
            "group_by": group_by,
            "order_by": (p.get("order_by") or ("head_name" if group_by == "head" else "accountname")),
        }
        return paginator.get_paginated_response(out_ser.data, totals=totals, meta_echo=meta_echo)

class LedgerSummaryExcelAPIView(APIView):
    """
    POST /api/reports/ledger-summary.xlsx

    Request body must match LedgerSummaryRequestSerializer (same as JSON API):
    {
      "entity": 1,
      "startdate": "2025-04-01",
      "enddate":   "2025-04-30",
      "group_by": "head|account|head_account",
      "order_by": "...",                 # optional (same keys as JSON API)
      "txn_in": "SALE,PURCHASE",         # optional CSV
      "voucherno": "SO-0001",            # optional
      "vno_contains": "SO-",             # optional
      "desc_contains": "freight",        # optional
      "accounthead": "1,2,3",            # optional CSV
      "account": "10,12",                # optional CSV
      "sign": "ALL|DR|CR",               # optional
      "include_zero": false,             # optional
      "min_activity": 0,                 # optional
      "amount_min": null,                # optional
      "amount_max": null,                # optional
      "range_min": null,                 # optional
      "range_max": null,                 # optional
      "pagesize": 50                     # ignored here (no pagination in Excel)
    }
    """
    permission_classes = (permissions.IsAuthenticated,)

    ORDER_MAP = {
        "accountname": "account__accountname",
        "-accountname": "-account__accountname",
        "head_name": "account__accounthead__name",
        "-head_name": "-account__accounthead__name",
        "opening": "openingbalance",
        "-opening": "-openingbalance",
        "debit": "debit",
        "-debit": "-debit",
        "credit": "credit",
        "-credit": "-credit",
        "period_net": "period_net",
        "-period_net": "-period_net",
        "abs_movement": "abs_movement",
        "-abs_movement": "-abs_movement",
        "balancetotal": "balancetotal",
        "-balancetotal": "-balancetotal",
    }

    def post(self, request, *_args, **_kwargs):
        # -----------------------------
        # 1) Validate input (reuse same serializer)
        # -----------------------------
        in_ser = LedgerSummaryRequestSerializer(data=request.data)
        in_ser.is_valid(raise_exception=True)
        p = in_ser.validated_data

        entity_id = p["entity"]
        startdate = p["startdate"]
        enddate   = p["enddate"]
        group_by  = p["group_by"]

        # -----------------------------
        # 2) Base queryset up to enddate + common filters
        # -----------------------------
        qs = JournalLine.objects.filter(
            entity_id=entity_id,
            account__isnull=False,
            entrydate__lte=enddate,
        )

        include_txn = None
        if p.get("txn_in"):
            include_txn = [x.strip() for x in str(p["txn_in"]).split(",") if x.strip()]
            qs = qs.filter(transactiontype__in=include_txn)

        if p.get("voucherno"):
            qs = qs.filter(voucherno=str(p["voucherno"]))
        if p.get("vno_contains"):
            qs = qs.filter(voucherno__icontains=str(p["vno_contains"]))
        if p.get("desc_contains"):
            qs = qs.filter(desc__icontains=str(p["desc_contains"]))

        ah_ids = None
        if p.get("accounthead"):
            ah_ids = [int(x) for x in str(p["accounthead"]).split(",") if x.strip()]
            qs = qs.filter(account__accounthead_id__in=ah_ids)

        acc_ids = None
        if p.get("account"):
            acc_ids = [int(x) for x in str(p["account"]).split(",") if x.strip()]
            qs = qs.filter(account_id__in=acc_ids)

        # -----------------------------
        # 3) Aggregations
        # -----------------------------
        period_f = Q(entrydate__gte=startdate, entrydate__lte=enddate)

        opening_expr = Coalesce(
            Sum(
                Case(
                    When(drcr=True, then=F("amount")),
                    When(drcr=False, then=-F("amount")),
                    default=V(0),
                    output_field=DEC18_2,
                ),
                filter=Q(entrydate__lt=startdate),
                output_field=DEC18_2,
            ),
            ZERO,
        )
        debit_expr = Coalesce(
            Sum(F("amount"), filter=period_f & Q(drcr=True), output_field=DEC18_2),
            ZERO,
        )
        credit_expr = Coalesce(
            Sum(F("amount"), filter=period_f & Q(drcr=False), output_field=DEC18_2),
            ZERO,
        )
        period_net_expr = ExpressionWrapper(F("debit") - F("credit"), output_field=DEC18_2)
        abs_mov_expr    = ExpressionWrapper(F("debit") + F("credit"), output_field=DEC18_2)
        closing_expr    = ExpressionWrapper(F("openingbalance") + F("debit") - F("credit"), output_field=DEC18_2)

        if group_by == "head":
            group_vals = ["account__accounthead_id", "account__accounthead__name"]
        else:
            group_vals = [
                "account_id",
                "account__accountname",
                "account__accounthead_id",
                "account__accounthead__name",
            ]

        qs = (
            qs.values(*group_vals)
              .annotate(
                  openingbalance=opening_expr,
                  debit=debit_expr,
                  credit=credit_expr,
              )
              .annotate(
                  period_net=period_net_expr,
                  abs_movement=abs_mov_expr,
                  balancetotal=closing_expr,
                  txn_count=Count("id", filter=period_f),
                  last_txn_date=Max("entrydate", filter=period_f),
              )
              .annotate(
                  drcr=Case(When(balancetotal__lt=0, then=V("CR")), default=V("DR")),
                  obdrcr=Case(When(openingbalance__lt=0, then=V("CR")), default=V("DR")),
              )
        )

        # -----------------------------
        # 4) Summary-focused filters
        # -----------------------------
        sign = p.get("sign", "ALL")
        if sign == "DR":
            qs = qs.filter(balancetotal__gt=0)
        elif sign == "CR":
            qs = qs.filter(balancetotal__lt=0)

        if not p.get("include_zero", False):
            qs = qs.exclude(balancetotal=ZERO_D)

        if p.get("min_activity") is not None:
            qs = qs.filter(abs_movement__gte=p["min_activity"])

        if p.get("amount_min") is not None:
            qs = qs.filter(balancetotal__gte=p["amount_min"])
        if p.get("amount_max") is not None:
            qs = qs.filter(balancetotal__lte=p["amount_max"])

        if p.get("range_min") is not None:
            qs = qs.filter(balancetotal__gte=p["range_min"])
        if p.get("range_max") is not None:
            qs = qs.filter(balancetotal__lte=p["range_max"])

        # -----------------------------
        # 5) Ordering
        # -----------------------------
        order_key = (p.get("order_by") or "").strip() or ("head_name" if group_by == "head" else "accountname")
        order_field = self.ORDER_MAP.get(order_key)
        if group_by == "head" and order_field in ("account__accountname", "-account__accountname"):
            order_field = "account__accounthead__name"
        if not order_field:
            order_field = "account__accounthead__name" if group_by == "head" else "account__accountname"
        qs = qs.order_by(order_field)

        rows = list(qs)

        # -----------------------------
        # 6) Grand totals (same as JSON API, recomputed from base JournalLine)
        # -----------------------------
        totals_qs = JournalLine.objects.filter(
            entity_id=entity_id,
            account__isnull=False,
            entrydate__lte=enddate,
        )
        if include_txn:
            totals_qs = totals_qs.filter(transactiontype__in=include_txn)
        if p.get("voucherno"):
            totals_qs = totals_qs.filter(voucherno=str(p["voucherno"]))
        if p.get("vno_contains"):
            totals_qs = totals_qs.filter(voucherno__icontains=str(p["vno_contains"]))
        if p.get("desc_contains"):
            totals_qs = totals_qs.filter(desc__icontains=str(p["desc_contains"]))
        if ah_ids:
            totals_qs = totals_qs.filter(account__accounthead_id__in=ah_ids)
        if acc_ids:
            totals_qs = totals_qs.filter(account_id__in=acc_ids)

        opening_total_expr = Coalesce(
            Sum(
                Case(
                    When(drcr=True, then=F("amount")),
                    When(drcr=False, then=-F("amount")),
                    default=V(0),
                    output_field=DEC18_2,
                ),
                filter=Q(entrydate__lt=startdate),
                output_field=DEC18_2,
            ),
            ZERO,
        )
        debit_total_expr = Coalesce(
            Sum(F("amount"), filter=Q(entrydate__gte=startdate, entrydate__lte=enddate) & Q(drcr=True), output_field=DEC18_2),
            ZERO,
        )
        credit_total_expr = Coalesce(
            Sum(F("amount"), filter=Q(entrydate__gte=startdate, entrydate__lte=enddate) & Q(drcr=False), output_field=DEC18_2),
            ZERO,
        )
        totals_raw = totals_qs.aggregate(
            opening=opening_total_expr,
            debit=debit_total_expr,
            credit=credit_total_expr,
        )
        opening_t = totals_raw["opening"] or ZERO_D
        debit_t   = totals_raw["debit"]   or ZERO_D
        credit_t  = totals_raw["credit"]  or ZERO_D
        closing_t = opening_t + debit_t - credit_t

        # -----------------------------
        # 7) Build Excel
        # -----------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger Summary"

        # Styles
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        hdr_fill = PatternFill("solid", fgColor="F2F2F2")
        h1 = NamedStyle(name="h1")
        h1.font = Font(bold=True, size=14)
        h1.alignment = Alignment(horizontal="center")
        try:
            wb.add_named_style(h1)
        except ValueError:
            pass

        hdr = NamedStyle(name="hdr")
        hdr.font = Font(bold=True)
        hdr.fill = hdr_fill
        hdr.border = border
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        try:
            wb.add_named_style(hdr)
        except ValueError:
            pass

        money = NamedStyle(name="money")
        money.number_format = '#,##0.00;[Red]-#,##0.00'
        money.border = border
        try:
            wb.add_named_style(money)
        except ValueError:
            pass

        norm = NamedStyle(name="norm")
        norm.border = border
        try:
            wb.add_named_style(norm)
        except ValueError:
            pass

        # Header / Meta block
        title = f"Ledger Summary ({group_by.replace('_', ' ').title()})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        ws.cell(row=1, column=1, value=title).style = "h1"

        meta_rows = [
            ["Entity", entity_id, "", "Period From", str(startdate), "", "To", str(enddate)],
            ["Group By", group_by, "", "Sign", sign, "", "Include Zero", str(bool(p.get("include_zero", False)))],
            ["Filters",
             f"AH={','.join(map(str, ah_ids)) if ah_ids else '-'}; "
             f"AC={','.join(map(str, acc_ids)) if acc_ids else '-'}; "
             f"Txn={','.join(include_txn) if include_txn else '-'}; "
             f"VNo={p.get('voucherno') or '-'}; "
             f"VNo*={p.get('vno_contains') or '-'}; "
             f"Desc*={p.get('desc_contains') or '-'}; "
             f"MinAct={p.get('min_activity') if p.get('min_activity') is not None else '-'}; "
             f"Amt[min..max]={p.get('amount_min') or '-'}..{p.get('amount_max') or '-'}; "
             f"Range[min..max]={p.get('range_min') or '-'}..{p.get('range_max') or '-'}",
             "", "", "", ""]
        ]
        r = 3
        for mr in meta_rows:
            for c, val in enumerate(mr, start=1):
                ws.cell(row=r, column=c, value=val)
            r += 1
        r += 1  # blank line before table

        # Table headers (depend on group_by)
        cols = []
        if group_by == "head":
            cols = [
                ("Head ID", 10),
                ("Head Name", 35),
            ]
        else:
            cols = [
                ("Head ID", 10),
                ("Head Name", 35),
                ("Account ID", 12),
                ("Account Name", 40),
            ]
        # common numeric cols
        cols += [
            ("Opening", 16),
            ("O/B DRCR", 9),
            ("Debit", 16),
            ("Credit", 16),
            ("Period Net", 16),
            ("Abs Movement", 16),
            ("Closing", 16),
            ("DRCR", 8),
            ("Txn Count", 11),
            ("Last Txn Date", 16),
        ]

        # write header row
        for c_idx, (title, width) in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c_idx, value=title)
            cell.style = "hdr"
            ws.column_dimensions[get_column_letter(c_idx)].width = width
        ws.freeze_panes = ws.cell(row=r+1, column=1)
        r += 1

        # Data rows
        def to_number(x):
            # openpyxl handles Decimal, but explicit float keeps formatting consistent
            return float(x) if x is not None else 0.0

        for rec in rows:
            if group_by == "head":
                head_id   = rec.get("account__accounthead_id")
                head_name = rec.get("account__accounthead__name")
                account_id = None
                account_name = None
                fixed = [head_id, head_name]
            else:
                head_id    = rec.get("account__accounthead_id")
                head_name  = rec.get("account__accounthead__name")
                account_id = rec.get("account_id")
                account_nm = rec.get("account__accountname") or ""
                fixed = [head_id, head_name, account_id, account_nm]

            opening = rec["openingbalance"]
            debit   = rec["debit"]
            credit  = rec["credit"]
            periodn = rec["period_net"]
            absmov  = rec["abs_movement"]
            closing = rec["balancetotal"]
            drcr    = rec["drcr"]
            obdrcr  = rec["obdrcr"]
            txn_ct  = rec["txn_count"]
            last_dt = rec["last_txn_date"]

            row_vals = fixed + [
                to_number(opening), obdrcr,
                to_number(debit), to_number(credit),
                to_number(periodn), to_number(absmov),
                to_number(closing), drcr,
                int(txn_ct or 0),
                last_dt.isoformat() if last_dt else None
            ]

            for idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=idx, value=val)
                # numeric columns get money style
                if idx >= (len(fixed) + 1) and idx <= (len(fixed) + 7) and idx != (len(fixed) + 2):  # all amount cols except O/B DRCR text
                    cell.style = "money"
                else:
                    cell.style = "norm"
                if idx == len(cols):  # Last Txn Date - set number format for ISO text? keep as text
                    cell.number_format = numbers.FORMAT_TEXT
            r += 1

        # Totals row
        # We only show totals for amount columns (Opening, Debit, Credit, Period Net, Abs Movement, Closing)
        totals_row_start = r + 1
        ws.cell(row=totals_row_start, column=1, value="GRAND TOTALS").style = "hdr"
        # figure offsets for amount columns based on group_by
        base = 0 if group_by == "head" else 2  # extra two columns (Account ID/Name) when grouped by account/head_account
        opening_col = 1 + (2 + base)  # Opening is after Head/Account cols
        debit_col   = opening_col + 2
        credit_col  = debit_col + 1
        pnet_col    = credit_col + 1
        absm_col    = pnet_col + 1
        close_col   = absm_col + 1

        ws.cell(row=totals_row_start, column=opening_col, value=float(opening_t)).style = "money"
        ws.cell(row=totals_row_start, column=debit_col,   value=float(debit_t)).style   = "money"
        ws.cell(row=totals_row_start, column=credit_col,  value=float(credit_t)).style  = "money"
        ws.cell(row=totals_row_start, column=pnet_col,    value=float(debit_t - credit_t)).style = "money"
        # Abs Movement grand total is NOT simply debit+credit per-account summed signlessly;
        # it's the sum over accounts of (debit + credit). Compute explicitly if you want exact:
        # For performance, we can approximate as debit+credit across all lines in period:
        ws.cell(row=totals_row_start, column=absm_col,    value=float(debit_t + credit_t)).style = "money"
        ws.cell(row=totals_row_start, column=close_col,   value=float(closing_t)).style  = "money"

        # thin border for totals row cells
        for cidx in range(1, len(cols) + 1):
            ws.cell(row=totals_row_start, column=cidx).border = border

        # Autofilter
        header_row_idx = (totals_row_start - (len(rows) + 1)) + (len(meta_rows) + 2)  # safe: just set over the whole table area
        ws.auto_filter.ref = f"A{(header_row_idx)}:{get_column_letter(len(cols))}{r-1}"

        # -----------------------------
        # 8) Return XLSX
        # -----------------------------
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"ledger-summary_{group_by}_{startdate}_{enddate}_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


class LedgerSummaryPDFAPIView(APIView):
    """
    POST /api/reports/ledger-summary.pdf
    Body: same LedgerSummaryRequestSerializer payload as your JSON endpoint.
    """
    permission_classes = (permissions.IsAuthenticated,)

    ORDER_MAP = {
        "accountname": "account__accountname",
        "-accountname": "-account__accountname",
        "head_name": "account__accounthead__name",
        "-head_name": "-account__accounthead__name",
        "opening": "openingbalance",
        "-opening": "-openingbalance",
        "debit": "debit",
        "-debit": "-debit",
        "credit": "credit",
        "-credit": "-credit",
        "period_net": "period_net",
        "-period_net": "-period_net",
        "abs_movement": "abs_movement",
        "-abs_movement": "-abs_movement",
        "balancetotal": "balancetotal",
        "-balancetotal": "-balancetotal",
    }

    # ---------- helpers ----------
    def _fmt_money(self, x):
        if x is None:
            return "0.00"
        # Comma-separated with 2 decimals; negatives show with minus (you can switch to DR/CR if you prefer)
        return f"{x:,.2f}"

    def _make_header_footer(self, title, subline):
        def on_page(canvas, doc):
            canvas.saveState()
            w, h = landscape(A4)
            canvas.setFont("Helvetica-Bold", 11)
            canvas.drawString(0.6*inch, h - 0.55*inch, title)
            canvas.setFont("Helvetica", 9)
            canvas.drawRightString(w - 0.6*inch, h - 0.55*inch, subline)
            canvas.restoreState()

        def on_footer(canvas, doc):
            canvas.saveState()
            w, h = landscape(A4)
            canvas.setFont("Helvetica", 8)
            canvas.setFillGray(0.4)
            canvas.drawRightString(
                w - 0.6*inch, 0.4*inch,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Page {doc.page}"
            )
            canvas.restoreState()

        return on_page, on_footer

    # ---------- main ----------
    def post(self, request, *_args, **_kwargs):
        # 1) Validate (reuse serializer)
        in_ser = LedgerSummaryRequestSerializer(data=request.data)
        in_ser.is_valid(raise_exception=True)
        p = in_ser.validated_data

        entity_id = p["entity"]
        startdate = p["startdate"]
        enddate   = p["enddate"]
        group_by  = p["group_by"]

        # 2) Base queryset + filters
        qs = JournalLine.objects.filter(
            entity_id=entity_id,
            account__isnull=False,
            entrydate__lte=enddate,
        )

        include_txn = None
        if p.get("txn_in"):
            include_txn = [x.strip() for x in str(p["txn_in"]).split(",") if x.strip()]
            qs = qs.filter(transactiontype__in=include_txn)

        if p.get("voucherno"):
            qs = qs.filter(voucherno=str(p["voucherno"]))
        if p.get("vno_contains"):
            qs = qs.filter(voucherno__icontains=str(p["vno_contains"]))
        if p.get("desc_contains"):
            qs = qs.filter(desc__icontains=str(p["desc_contains"]))

        ah_ids = None
        if p.get("accounthead"):
            ah_ids = [int(x) for x in str(p["accounthead"]).split(",") if x.strip()]
            qs = qs.filter(account__accounthead_id__in=ah_ids)

        acc_ids = None
        if p.get("account"):
            acc_ids = [int(x) for x in str(p["account"]).split(",") if x.strip()]
            qs = qs.filter(account_id__in=acc_ids)

        # 3) Aggregations
        period_f = Q(entrydate__gte=startdate, entrydate__lte=enddate)

        opening_expr = Coalesce(
            Sum(
                Case(
                    When(drcr=True, then=F("amount")),
                    When(drcr=False, then=-F("amount")),
                    default=V(0),
                    output_field=DEC18_2,
                ),
                filter=Q(entrydate__lt=startdate),
                output_field=DEC18_2,
            ),
            ZERO,
        )
        debit_expr = Coalesce(
            Sum(F("amount"), filter=period_f & Q(drcr=True), output_field=DEC18_2),
            ZERO,
        )
        credit_expr = Coalesce(
            Sum(F("amount"), filter=period_f & Q(drcr=False), output_field=DEC18_2),
            ZERO,
        )
        period_net_expr = ExpressionWrapper(F("debit") - F("credit"), output_field=DEC18_2)
        abs_mov_expr    = ExpressionWrapper(F("debit") + F("credit"), output_field=DEC18_2)
        closing_expr    = ExpressionWrapper(F("openingbalance") + F("debit") - F("credit"), output_field=DEC18_2)

        if group_by == "head":
            group_vals = ["account__accounthead_id", "account__accounthead__name"]
        else:
            group_vals = [
                "account_id",
                "account__accountname",
                "account__accounthead_id",
                "account__accounthead__name",
            ]

        qs = (
            qs.values(*group_vals)
              .annotate(
                  openingbalance=opening_expr,
                  debit=debit_expr,
                  credit=credit_expr,
              )
              .annotate(
                  period_net=period_net_expr,
                  abs_movement=abs_mov_expr,
                  balancetotal=closing_expr,
                  txn_count=Count("id", filter=period_f),
                  last_txn_date=Max("entrydate", filter=period_f),
              )
              .annotate(
                  drcr=Case(When(balancetotal__lt=0, then=V("CR")), default=V("DR")),
                  obdrcr=Case(When(openingbalance__lt=0, then=V("CR")), default=V("DR")),
              )
        )

        # 4) Summary filters
        sign = p.get("sign", "ALL")
        if sign == "DR":
            qs = qs.filter(balancetotal__gt=0)
        elif sign == "CR":
            qs = qs.filter(balancetotal__lt=0)

        if not p.get("include_zero", False):
            qs = qs.exclude(balancetotal=ZERO_D)

        if p.get("min_activity") is not None:
            qs = qs.filter(abs_movement__gte=p["min_activity"])

        if p.get("amount_min") is not None:
            qs = qs.filter(balancetotal__gte=p["amount_min"])
        if p.get("amount_max") is not None:
            qs = qs.filter(balancetotal__lte=p["amount_max"])

        if p.get("range_min") is not None:
            qs = qs.filter(balancetotal__gte=p["range_min"])
        if p.get("range_max") is not None:
            qs = qs.filter(balancetotal__lte=p["range_max"])

        # 5) Ordering
        order_key = (p.get("order_by") or "").strip() or ("head_name" if group_by == "head" else "accountname")
        order_field = self.ORDER_MAP.get(order_key)
        if group_by == "head" and order_field in ("account__accountname", "-account__accountname"):
            order_field = "account__accounthead__name"
        if not order_field:
            order_field = "account__accounthead__name" if group_by == "head" else "account__accountname"
        qs = qs.order_by(order_field)

        rows = list(qs)

        # 6) Grand totals
        totals_qs = JournalLine.objects.filter(
            entity_id=entity_id,
            account__isnull=False,
            entrydate__lte=enddate,
        )
        if include_txn:
            totals_qs = totals_qs.filter(transactiontype__in=include_txn)
        if p.get("voucherno"):
            totals_qs = totals_qs.filter(voucherno=str(p["voucherno"]))
        if p.get("vno_contains"):
            totals_qs = totals_qs.filter(voucherno__icontains=str(p["vno_contains"]))
        if p.get("desc_contains"):
            totals_qs = totals_qs.filter(desc__icontains=str(p["desc_contains"]))
        if ah_ids:
            totals_qs = totals_qs.filter(account__accounthead_id__in=ah_ids)
        if acc_ids:
            totals_qs = totals_qs.filter(account_id__in=acc_ids)

        opening_total_expr = Coalesce(
            Sum(
                Case(
                    When(drcr=True, then=F("amount")),
                    When(drcr=False, then=-F("amount")),
                    default=V(0),
                    output_field=DEC18_2,
                ),
                filter=Q(entrydate__lt=startdate),
                output_field=DEC18_2,
            ),
            ZERO,
        )
        debit_total_expr = Coalesce(
            Sum(F("amount"), filter=Q(entrydate__gte=startdate, entrydate__lte=enddate) & Q(drcr=True), output_field=DEC18_2),
            ZERO,
        )
        credit_total_expr = Coalesce(
            Sum(F("amount"), filter=Q(entrydate__gte=startdate, entrydate__lte=enddate) & Q(drcr=False), output_field=DEC18_2),
            ZERO,
        )
        totals_raw = totals_qs.aggregate(
            opening=opening_total_expr,
            debit=debit_total_expr,
            credit=credit_total_expr,
        )
        opening_t = totals_raw["opening"] or ZERO_D
        debit_t   = totals_raw["debit"]   or ZERO_D
        credit_t  = totals_raw["credit"]  or ZERO_D
        closing_t = opening_t + debit_t - credit_t

        # 7) Build PDF
        buf = BytesIO()
        pagesize = landscape(A4)
        doc = BaseDocTemplate(
            buf, pagesize=pagesize,
            leftMargin=0.5*inch, rightMargin=0.5*inch,
            topMargin=0.8*inch, bottomMargin=0.6*inch
        )
        frame = Frame(
            doc.leftMargin, doc.bottomMargin,
            doc.width, doc.height - 0.1*inch,
            id="normal"
        )

        title = f"Ledger Summary ({group_by.replace('_', ' ').title()})"
        subline = f"Entity: {entity_id} | Period: {startdate} to {enddate}"
        on_page, on_footer = self._make_header_footer(title, subline)
        doc.addPageTemplates([PageTemplate(id="lpage", frames=[frame], onPage=on_page, onPageEnd=on_footer)])

        styles = getSampleStyleSheet()
        p_meta = ParagraphStyle(
            "meta",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#333333")
        )
        p_hdr = ParagraphStyle(
            "hdr",
            parent=styles["Normal"],
            fontSize=10,
            leading=12,
            textColor=colors.black
        )

        story = []
        # a) Meta block
        filters_text = (
            f"<b>Sign</b>: {p.get('sign','ALL')} &nbsp;&nbsp; "
            f"<b>Include Zero</b>: {bool(p.get('include_zero', False))} &nbsp;&nbsp; "
            f"<b>AH</b>: {','.join(map(str, ah_ids)) if ah_ids else '-'} &nbsp;&nbsp; "
            f"<b>AC</b>: {','.join(map(str, acc_ids)) if acc_ids else '-'} &nbsp;&nbsp; "
            f"<b>Txn</b>: {','.join(include_txn) if include_txn else '-'} &nbsp;&nbsp; "
            f"<b>VNo</b>: {p.get('voucherno') or '-'} &nbsp;&nbsp; "
            f"<b>VNo*</b>: {p.get('vno_contains') or '-'} &nbsp;&nbsp; "
            f"<b>Desc*</b>: {p.get('desc_contains') or '-'} &nbsp;&nbsp; "
            f"<b>MinAct</b>: {p.get('min_activity') if p.get('min_activity') is not None else '-'} &nbsp;&nbsp; "
            f"<b>Amt[min..max]</b>: {p.get('amount_min') or '-'}..{p.get('amount_max') or '-'} &nbsp;&nbsp; "
            f"<b>Range[min..max]</b>: {p.get('range_min') or '-'}..{p.get('range_max') or '-'}"
        )
        story.append(Paragraph(filters_text, p_meta))
        story.append(Spacer(1, 0.15*inch))

        # b) Build table header according to grouping
        if group_by == "head":
            headers = ["Head ID", "Head Name",
                       "Opening", "O/B DRCR", "Debit", "Credit",
                       "Period Net", "Abs Movement", "Closing", "DRCR",
                       "Txn Count", "Last Txn Date"]
            col_widths = [0.9*inch, 2.9*inch,
                          1.1*inch, 0.85*inch, 1.1*inch, 1.1*inch,
                          1.1*inch, 1.2*inch, 1.1*inch, 0.75*inch,
                          0.85*inch, 1.2*inch]
            fixed_keys = ("account__accounthead_id", "account__accounthead__name")
        else:
            headers = ["Head ID", "Head Name", "Account ID", "Account Name",
                       "Opening", "O/B DRCR", "Debit", "Credit",
                       "Period Net", "Abs Movement", "Closing", "DRCR",
                       "Txn Count", "Last Txn Date"]
            col_widths = [0.9*inch, 2.4*inch, 0.95*inch, 2.7*inch,
                          1.1*inch, 0.85*inch, 1.1*inch, 1.1*inch,
                          1.1*inch, 1.2*inch, 1.1*inch, 0.75*inch,
                          0.85*inch, 1.2*inch]
            fixed_keys = ("account__accounthead_id", "account__accounthead__name",
                          "account_id", "account__accountname")

        data = [headers]

        # c) Data rows
        for rec in rows:
            fixed_vals = [rec.get(k) or "" for k in fixed_keys]
            opening = rec["openingbalance"]
            debit   = rec["debit"]
            credit  = rec["credit"]
            periodn = rec["period_net"]
            absmov  = rec["abs_movement"]
            closing = rec["balancetotal"]
            drcr    = rec["drcr"]
            obdrcr  = rec["obdrcr"]
            txn_ct  = rec["txn_count"]
            last_dt = rec["last_txn_date"].isoformat() if rec["last_txn_date"] else ""

            row = list(fixed_vals) + [
                self._fmt_money(opening), obdrcr,
                self._fmt_money(debit), self._fmt_money(credit),
                self._fmt_money(periodn), self._fmt_money(absmov),
                self._fmt_money(closing), drcr,
                int(txn_ct or 0), last_dt
            ]
            data.append(row)

        # d) Totals row (band)
        totals_label = "GRAND TOTALS"
        # compute period net = debit - credit; abs movement proxy = debit + credit
        totals_row = []
        if group_by == "head":
            prefix_len = 2
            totals_row = ["", totals_label]
        else:
            prefix_len = 4
            totals_row = ["", totals_label, "", ""]

        totals_row += [
            self._fmt_money(opening_t), "",  # Opening, O/B DRCR is blank
            self._fmt_money(debit_t), self._fmt_money(credit_t),
            self._fmt_money(debit_t - credit_t),
            self._fmt_money(debit_t + credit_t),  # proxy for abs movement
            self._fmt_money(closing_t), "", "", ""  # DRCR/Txn/Date blank
        ]
        data.append(totals_row)

        # e) Table + styling
        tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        # figure out numeric columns to right-align
        # For both variants, numeric columns start after the fixed fields:
        # [Opening, O/B DRCR(text), Debit, Credit, Period Net, Abs Movement, Closing, DRCR(text), Txn Count, Last Date]
        ncols = len(headers)
        table_style_cmds = [
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F2F2F2")),  # header bg
            ("TEXTCOLOR",  (0,0), (-1,0), colors.HexColor("#000000")),
            ("GRID",       (0,0), (-1,-1), 0.25, colors.HexColor("#CCCCCC")),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,0), 9),
            ("FONTSIZE",   (0,1), (-1,-1), 8),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",      (0,0), (-1,0), "CENTER"),
        ]

        # Right align: all amounts + Txn Count
        first_amt_col = prefix_len  # 0-based: after fixed cols
        amt_cols = [
            first_amt_col + 0,  # Opening
            first_amt_col + 2,  # Debit
            first_amt_col + 3,  # Credit
            first_amt_col + 4,  # Period Net
            first_amt_col + 5,  # Abs Movement
            first_amt_col + 6,  # Closing
        ]
        for c in amt_cols + [ncols - 2]:  # txn count column
            table_style_cmds.append(("ALIGN", (c,1), (c,-1), "RIGHT"))

        # Totals band: bold + light background
        totals_row_idx = len(data) - 1
        table_style_cmds += [
            ("BACKGROUND", (0, totals_row_idx), (-1, totals_row_idx), colors.HexColor("#FFF8E1")),
            ("FONTNAME",   (0, totals_row_idx), (-1, totals_row_idx), "Helvetica-Bold"),
        ]

        tbl.setStyle(TableStyle(table_style_cmds))

        story.append(tbl)

        # 8) Build & respond
        doc.build(story)
        pdf = buf.getvalue()
        buf.close()

        fname = f"ledger-summary_{group_by}_{startdate}_{enddate}_{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

class ledgerjournaldetails(ListAPIView):
    """
    Optimized ledger details:
    - DB-side grouping & windowed running balance (non-aggregated)
    - FY-aware 'Y' aggregation
    - Total shows SUM(debits/credits); balance is final running balance
    """
    serializer_class = LedgerAccountSerializer
    permission_classes = (permissions.IsAuthenticated,)

    # -------- helpers that ignore None/blank/"null" -----------
    @staticmethod
    def _nz(v):
        if v is None: return None
        s = str(v).strip()
        return None if s == '' or s.lower() == 'null' else v

    @staticmethod
    def _split_ints(v):
        v = ledgerjournaldetails._nz(v)
        if v is None: return None
        return [int(x) for x in str(v).split(',') if x.strip()]

    @staticmethod
    def _split_strs(v):
        v = ledgerjournaldetails._nz(v)
        if v is None: return None
        return [x.strip() for x in str(v).split(',') if x.strip()]

    def post(self, request, *args, **kwargs):
        s_in = LedgerFilterSerializer(data=request.data)
        if not s_in.is_valid():
            return Response(s_in.errors, status=status.HTTP_400_BAD_REQUEST)
        p = s_in.validated_data

        entity = p['entity']
        sdate  = p['startdate']
        edate  = p['enddate']

        # Resolve FY that covers the requested range
        fy = (entityfinancialyear.objects
              .filter(entity=entity, finstartyear__lte=edate, finendyear__gte=sdate)
              .first())
        if not fy:
            return Response({"detail": "Financial year not found for the given range."}, status=400)

        fy_start = fy.finstartyear.date() if isinstance(fy.finstartyear, datetime) else fy.finstartyear
        fy_end   = fy.finendyear.date()   if isinstance(fy.finendyear,   datetime) else fy.finendyear

        # Base queryset (FY start..enddate)
        base = (JournalLine.objects
                .filter(entity=entity, entrydate__range=(fy_start, edate))
                .select_related('account'))

        # Only join entry if needed
        if p.get('include_entry_id'):
            base = base.select_related('entry')

        # ------- Null-safe filters (only apply if provided) -------
        ah_list = self._split_ints(p.get('accounthead'))
        if ah_list:
            base = base.filter(accounthead_id__in=ah_list)

        a_list = self._split_ints(p.get('account'))
        if a_list:
            base = base.filter(account_id__in=a_list)

        ttype_list = self._split_strs(p.get('transactiontype'))
        if ttype_list:
            base = base.filter(transactiontype__in=ttype_list)

        tid_list = self._split_ints(p.get('transactionid'))
        if tid_list:
            base = base.filter(transactionid__in=tid_list)

        vno = self._nz(p.get('voucherno'))
        if vno is not None:
            base = base.filter(voucherno__icontains=vno)

        drcr = self._nz(p.get('drcr'))
        if drcr in ('0', '1'):
            base = base.filter(drcr=(drcr == '1'))

        desc_txt = self._nz(p.get('desc'))  # if you later add desc in serializer
        if desc_txt is not None:
            base = base.filter(desc__icontains=desc_txt)

        # Amount range (applied after debit/credit split)
        amt_range = None
        if p.get('amountstart') is not None and p.get('amountend') is not None:
            try:
                lo = Decimal(p['amountstart'])
                hi = Decimal(p['amountend'])
                if lo > hi:
                    lo, hi = hi, lo
                amt_range = (lo, hi)
            except Exception:
                amt_range = None  # ignore bad numbers silently

        # Details window (inside start..end)
        details_base = base.filter(entrydate__gte=sdate)
        if p.get('sub_startdate') and p.get('sub_enddate'):
            details_base = details_base.filter(entrydate__range=(p['sub_startdate'], p['sub_enddate']))

        # Common annotations
        debit_case   = Case(When(drcr=True,  then=F('amount')))
        credit_case  = Case(When(drcr=False, then=F('amount')))
        debit_annot  = Coalesce(Sum(debit_case),  Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2))
        credit_annot = Coalesce(Sum(credit_case), Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2))

        include_opening   = bool(p.get('include_opening', True))
        include_total     = bool(p.get('include_total', True))
        include_zero      = bool(p.get('include_zero_balance', False))
        include_entry_id  = bool(p.get('include_entry_id', False))

        # --------------------------
        # 1) OPENING (FY start..sdate-1)
        # --------------------------
        opening_by_acct = {}
        if include_opening:
            open_qs = (base.filter(entrydate__lt=sdate)
                           .values('account_id', 'account__accountname', 'account__accountcode')
                           .annotate(debitamount=debit_annot, creditamount=credit_annot))
            for r in open_qs:
                acct_id   = r['account_id']
                acct_name = r['account__accountname']
                acct_code = r.get('account__accountcode')
                d = r['debitamount']; c = r['creditamount']
                bal = d - c
                opening_by_acct[acct_id] = {
                    'accountname': acct_name,
                    'accountid': acct_id,
                    'accountcode': acct_code,
                    'creditamount': (-bal) if bal < 0 else Decimal('0.00'),
                    'debitamount':  bal if bal >= 0 else Decimal('0.00'),
                    'desc': 'Opening',
                    'entrydate': sdate,
                    'transactiontype': 'O',
                    'transactionid': -1,
                    'drcr': (bal > 0),
                    'displaydate': sdate.strftime('%d-%m-%Y')
                }

        # --------------------------
        # 2) DETAILS
        # --------------------------
        aggby = (p.get('aggby') or '').strip().upper() if p.get('aggby') else ''
        detail_rows = []

        if aggby == 'Y':
            grouped = (details_base
                       .values('account_id', 'account__accountname', 'account__accountcode')
                       .annotate(debitamount=debit_annot, creditamount=credit_annot))
            for r in grouped:
                d, c = r['debitamount'], r['creditamount']
                if amt_range and not (amt_range[0] <= d <= amt_range[1] or amt_range[0] <= c <= amt_range[1]):
                    continue
                detail_rows.append({
                    'accountname': r['account__accountname'],
                    'accountid':   r['account_id'],
                    'accountcode': r.get('account__accountcode'),
                    'creditamount': c,
                    'debitamount':  d,
                    'desc': f"FY {fy_start.strftime('%Y')}-{str(fy_end.year % 100).zfill(2)}",
                    'entrydate': fy_end,
                    'transactiontype': 'Y',
                    'transactionid': -1,
                    'drcr': (d - c) > 0,
                    'displaydate': f"{fy_start.strftime('%d-%m-%Y')} to {fy_end.strftime('%d-%m-%Y')}"
                })

        elif aggby in ('D','M','Q'):
            trunc = {'D': TruncDay('entrydate'),
                     'M': TruncMonth('entrydate'),
                     'Q': TruncQuarter('entrydate')}[aggby]
            grouped = (details_base
                       .annotate(period=trunc)
                       .values('account_id', 'account__accountname', 'account__accountcode', 'period')
                       .annotate(debitamount=debit_annot, creditamount=credit_annot)
                       .order_by('period', 'account_id'))
            for r in grouped:
                d, c = r['debitamount'], r['creditamount']
                if amt_range and not (amt_range[0] <= d <= amt_range[1] or amt_range[0] <= c <= amt_range[1]):
                    continue
                period_dt = r['period'].date() if isinstance(r['period'], datetime) else r['period']
                desc = period_dt.strftime('%b') if aggby == 'M' else period_dt.strftime('%Y-%m-%d')
                detail_rows.append({
                    'accountname': r['account__accountname'],
                    'accountid':   r['account_id'],
                    'accountcode': r.get('account__accountcode'),
                    'creditamount': c,
                    'debitamount':  d,
                    'desc': desc,
                    'entrydate': period_dt,
                    'transactiontype': aggby,
                    'transactionid': -1,
                    'drcr': (d - c) > 0,
                    'displaydate': period_dt.strftime('%d-%m-%Y')
                })

        else:
            # Non-aggregated: one row per (account, date, typ, id); compute running in Python.
            values_fields = [
                'account_id', 'account__accountname', 'account__accountcode',
                'entrydate', 'transactiontype', 'transactionid', 'desc'
            ]
            if include_entry_id:
                values_fields.append('entry_id')

            base_qs = (
                details_base
                .values(*values_fields)               # <-- ensures GROUP BY only these fields
                .annotate(
                    debitamount=Coalesce(Sum(Case(When(drcr=True,  then=F('amount')))),
                                         Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2)),
                    creditamount=Coalesce(Sum(Case(When(drcr=False, then=F('amount')))),
                                          Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2)),
                )
                .order_by('entrydate', 'transactiontype', 'transactionid')
            )

            # Optional amount-range HAVING filter on the annotated sides
            if amt_range:
                base_qs = base_qs.filter(
                    Q(debitamount__gte=amt_range[0], debitamount__lte=amt_range[1]) |
                    Q(creditamount__gte=amt_range[0], creditamount__lte=amt_range[1])
                )

            running_by_acct = {}
            for r in base_qs:
                d = r['debitamount'] or Decimal('0.00')
                c = r['creditamount'] or Decimal('0.00')
                aid = r['account_id']
                running = running_by_acct.get(aid, Decimal('0.00')) + (d - c)
                running_by_acct[aid] = running

                row = {
                    'accountname': r['account__accountname'],
                    'accountid':   aid,
                    'accountcode': r.get('account__accountcode'),
                    'creditamount': c,
                    'debitamount':  d,
                    'desc': r.get('desc') or '',
                    'entrydate': r['entrydate'],
                    'transactiontype': r['transactiontype'],
                    'transactionid':   r['transactionid'],
                    'drcr': (d - c) > 0,
                    'displaydate': r['entrydate'].strftime('%d-%m-%Y') if isinstance(r['entrydate'], date) else str(r['entrydate']),
                    'balance': running,
                }
                if include_entry_id and 'entry_id' in r:
                    row['entry_id'] = r['entry_id']
                detail_rows.append(row)

        # --------------------------
        # 3) TOTALS & ASSEMBLY
        # --------------------------
        per_acct = {}

        # Seed opening
        if include_opening:
            for aid, row in opening_by_acct.items():
                per_acct.setdefault(aid, {
                    'accountname': row['accountname'],
                    'accountid':   aid,
                    'accountcode': row.get('accountcode'),
                    'rows': [],
                    'sum_debit':  Decimal('0.00'),
                    'sum_credit': Decimal('0.00'),
                })
                per_acct[aid]['rows'].append(row)
                per_acct[aid]['sum_debit']  += row['debitamount']
                per_acct[aid]['sum_credit'] += row['creditamount']

        # Add details
        for r in detail_rows:
            aid = r['accountid']
            per_acct.setdefault(aid, {
                'accountname': r['accountname'],
                'accountid':   aid,
                'accountcode': r.get('accountcode'),
                'rows': [],
                'sum_debit':  Decimal('0.00'),
                'sum_credit': Decimal('0.00'),
            })
            per_acct[aid]['rows'].append(r)
            per_acct[aid]['sum_debit']  += Decimal(str(r['debitamount']))
            per_acct[aid]['sum_credit'] += Decimal(str(r['creditamount']))

        # Sort rows & compute running balance for aggregated branches; add Totals
        type_rank = {'O': 0, 'T': 2}  # default 1 for details
        result_payload = []

        for aid, bucket in per_acct.items():
            total_debit  = bucket['sum_debit']
            total_credit = bucket['sum_credit']
            net_balance  = total_debit - total_credit
            net_drcr     = (net_balance > 0)

            if include_total:
                bucket['rows'].append({
                    'accountname': bucket['accountname'],
                    'accountid':   aid,
                    'accountcode': bucket.get('accountcode'),
                    'creditamount': total_credit,
                    'debitamount':  total_debit,
                    'desc': 'Total',
                    'entrydate': edate,
                    'transactiontype': 'T',
                    'transactionid': -1,
                    'drcr': net_drcr,
                    'displaydate': edate.strftime('%d-%m-%Y')
                })

            bucket['rows'].sort(
                key=lambda r: (
                    type_rank.get(r['transactiontype'], 1),
                    r['entrydate'],
                    r['transactiontype'],
                    r['transactionid']
                )
            )

            running = Decimal('0.00')
            for r in bucket['rows']:
                if 'balance' in r and r['transactiontype'] not in ('O','T'):
                    running = Decimal(str(r['balance']))
                elif r['transactiontype'] != 'T':
                    running += (Decimal(str(r['debitamount'])) - Decimal(str(r['creditamount'])))
                    r['balance'] = running
                else:
                    r['balance'] = running

            if not include_zero:
                has_detail = any(r['transactiontype'] not in ('O','T') for r in bucket['rows'])
                if abs(float(running)) < 1e-9 and not has_detail:
                    continue

            result_payload.append({
                'accountname': bucket['accountname'],
                'accountid':   aid,
                'accountcode': bucket.get('accountcode'),
                'accounts':    bucket['rows']
            })

        # Sorting across accounts (null-safe defaults)
        sort_by  = (p.get('sort_by')  or 'name')
        sort_dir = (p.get('sort_dir') or 'asc')
        reverse  = (sort_dir == 'desc')

        def acct_sort_key(a):
            if sort_by == 'name':
                return (a['accountname'] or '').lower()
            if sort_by == 'code':
                return a.get('accountcode') or 0
            tot_debit  = sum(Decimal(str(r['debitamount']))  for r in a['accounts'])
            tot_credit = sum(Decimal(str(r['creditamount'])) for r in a['accounts'])
            if sort_by == 'debit':  return (tot_debit,)
            if sort_by == 'credit': return (tot_credit,)
            if sort_by == 'net':    return (tot_debit - tot_credit,)
            return (a['accountname'] or '').lower()

        result_payload.sort(key=acct_sort_key, reverse=reverse)

        out = self.get_serializer(result_payload, many=True)
        return Response(out.data, status=200)


class tradingaccountstatementJournaline(ListAPIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, *args, **kwargs):
        entity_id  = int(self.request.query_params.get('entity'))
        start_date = self.request.query_params.get('startdate')  # 'YYYY-MM-DD'
        end_date   = self.request.query_params.get('enddate')    # 'YYYY-MM-DD'

        data = build_trading_account_dynamic(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            # include_move_types=('OUT',), exclude_move_types=('REV',)
        )
        return Response(data)


class profitandlossstatement(ListAPIView):
    permission_classes = (permissions.IsAuthenticated,)

    def _clean(self, s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    def get(self, request, *args, **kwargs):
        entity_id  = int(request.query_params.get('entity'))
        start_date = self._clean(request.query_params.get('startdate'))
        end_date   = self._clean(request.query_params.get('enddate'))

        level = (request.query_params.get('level') or 'head').lower()  # head|account|product|voucher
        valuation_method = (request.query_params.get('valuation_method') or 'fifo').lower()

        # Allow overriding the default group values via query if needed
        pl_detailsingroup_values = tuple(
            int(x) for x in (request.query_params.get('pl_detailsingroup_values') or '2').split(',')
        )
        trading_detailsingroup_values = tuple(
            int(x) for x in (request.query_params.get('trading_detailsingroup_values') or '1').split(',')
        )

        data = build_profit_and_loss_statement(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            level=level,
            pl_detailsingroup_values=pl_detailsingroup_values,          # default (2,)
            trading_detailsingroup_values=trading_detailsingroup_values, # default (1,)
            valuation_method=valuation_method
        )
        return Response(data)

# views.py
class balancesheetstatement(ListAPIView):
    permission_classes = (permissions.IsAuthenticated,)

    def _clean(self, s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    def get(self, request, *args, **kwargs):
        entity_id  = int(request.query_params.get('entity'))
        start_date = self._clean(request.query_params.get('startdate'))   # period start (for P&L)
        end_date   = self._clean(request.query_params.get('enddate'))     # as-of date for BS

        level = (request.query_params.get('level') or 'head').lower()             # head|account|voucher|product
        inventory_source = (request.query_params.get('inventory_source') or 'valuation').lower()  # valuation|gl
        valuation_method = (request.query_params.get('valuation_method') or 'fifo').lower()

        bs_dig = tuple(int(x) for x in (request.query_params.get('bs_detailsingroup_values') or '3').split(','))
        pl_dig = tuple(int(x) for x in (request.query_params.get('pl_detailsingroup_values') or '2').split(','))
        tr_dig = tuple(int(x) for x in (request.query_params.get('trading_detailsingroup_values') or '1').split(','))

        include_current_earnings = (request.query_params.get('include_current_earnings') or 'true').lower() != 'false'

        data = build_balance_sheet_statement(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            level=level,
            bs_detailsingroup_values=bs_dig,
            pl_detailsingroup_values=pl_dig,
            trading_detailsingroup_values=tr_dig,
            include_current_earnings=include_current_earnings,
            inventory_source=inventory_source,
            valuation_method=valuation_method
        )
        return Response(data)



def _q2f(x) -> float:
    return float(Decimal(str(x)).quantize(Decimal("0.01")))

def _auto_width(ws, col_min=1, col_max=8, padding=2):
    for col in range(col_min, col_max + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for cell in ws[letter]:
            l = len(str(cell.value)) if cell.value is not None else 0
            if l > maxlen:
                maxlen = l
        ws.column_dimensions[letter].width = maxlen + padding

def _money(cell):
    cell.number_format = '#,##0.00'
    return cell

def _title_cell(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='F3F4F6')
    cell.alignment = Alignment(horizontal='center')
    thin = Side(style='thin', color='D1D5DB')
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def _flatten_rows(rows: List[Dict], level=0):
    out = []
    for r in rows or []:
        out.append({"label": r.get("label"), "amount": r.get("amount", 0), "level": level})
        if r.get("children"):
            out.extend(_flatten_rows(r["children"], level + 1))
    return out


class BalanceSheetExcelAPIView(GenericAPIView):
    """
    GET /api/reports/balance-sheet.xlsx?entity=1&startdate=2025-04-01&enddate=2025-09-30
         [&level=head|account|voucher|product]
         [&inventory_source=valuation|gl]
         [&valuation_method=fifo|lifo|mwa|latest|wac]
         [&bs_detailsingroup_values=3]
         [&pl_detailsingroup_values=2]
         [&trading_detailsingroup_values=1]
         [&include_current_earnings=true|false]
    """
    permission_classes = (permissions.IsAuthenticated,)

    @staticmethod
    def _clean(s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    def get(self, request, *args, **kwargs):
        # Params
        entity_id  = int(request.query_params.get('entity'))
        start_date = self._clean(request.query_params.get('startdate'))
        end_date   = self._clean(request.query_params.get('enddate'))
        level = (request.query_params.get('level') or 'head').lower()
        inventory_source = (request.query_params.get('inventory_source') or 'valuation').lower()
        valuation_method = (request.query_params.get('valuation_method') or 'fifo').lower()
        bs_dig = tuple(int(x) for x in (request.query_params.get('bs_detailsingroup_values') or '3').split(','))
        pl_dig = tuple(int(x) for x in (request.query_params.get('pl_detailsingroup_values') or '2').split(','))
        tr_dig = tuple(int(x) for x in (request.query_params.get('trading_detailsingroup_values') or '1').split(','))
        include_current_earnings = (request.query_params.get('include_current_earnings') or 'true').lower() != 'false'

        # Data
        data = build_balance_sheet_statement(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            level=level,
            bs_detailsingroup_values=bs_dig,
            pl_detailsingroup_values=pl_dig,
            trading_detailsingroup_values=tr_dig,
            include_current_earnings=include_current_earnings,
            inventory_source=inventory_source,
            valuation_method=valuation_method
        )

        # Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        # Titles
        ws.cell(row=1, column=1, value="ASSETS"); _title_cell(ws.cell(row=1, column=1))
        ws.cell(row=1, column=3, value="LIABILITIES & EQUITY"); _title_cell(ws.cell(row=1, column=3))

        ws.cell(row=2, column=1, value="Particulars").font = Font(bold=True)
        ws.cell(row=2, column=2, value="Amount").font = Font(bold=True)
        ws.cell(row=2, column=3, value="Particulars").font = Font(bold=True)
        ws.cell(row=2, column=4, value="Amount").font = Font(bold=True)

        a_flat = _flatten_rows(data.get("assets_rows", []))
        l_flat = _flatten_rows(data.get("liabilities_rows", []))
        max_len = max(len(a_flat), len(l_flat))

        for i in range(max_len):
            if i < len(a_flat):
                a = a_flat[i]
                ws.cell(row=3 + i, column=1, value=(" " * (a["level"] * 2)) + str(a["label"]))
                _money(ws.cell(row=3 + i, column=2, value=_q2f(a["amount"])))
            if i < len(l_flat):
                l = l_flat[i]
                ws.cell(row=3 + i, column=3, value=(" " * (l["level"] * 2)) + str(l["label"]))
                _money(ws.cell(row=3 + i, column=4, value=_q2f(l["amount"])))

        total_row = 3 + max_len
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        _money(ws.cell(row=total_row, column=2, value=_q2f(data["assets_total"]))).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="Total").font = Font(bold=True)
        _money(ws.cell(row=total_row, column=4, value=_q2f(data["liabilities_total"]))).font = Font(bold=True)

        # Notes (optional)
        notes_start = total_row + 2
        ws.cell(row=notes_start, column=1, value="Notes").font = Font(bold=True)
        for i, note in enumerate(data.get("notes", []), start=notes_start + 1):
            ws.cell(row=i, column=1, value=f"• {note}")

        _auto_width(ws, 1, 4)

        # Response
        fname = f'BalanceSheet_entity{entity_id}_{data["period"]["start"]}_to_{data["period"]["end"]}.xlsx'
        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(resp)
        return resp



def _q2f(x) -> float:
    return float(round(Decimal(str(x or 0)), 2))

def _clean_str(s: str) -> str:
    return (s or "").strip().strip('"').strip("'")

def _flatten_rows(rows, *, level=0, out=None):
    """
    Flattens nested rows like:
      [{"label": "...", "amount": 123.45, "children": [...]}, ...]
    into: [{"label": "...", "amount": 123.45, "level": N}, ...]
    """
    if out is None:
        out = []
    for r in rows or []:
        out.append({"label": r.get("label", ""), "amount": _q2f(r.get("amount", 0)), "level": level})
        kids = r.get("children") or []
        if kids:
            _flatten_rows(kids, level=level + 1, out=out)
    return out

# Paragraph styles
_styles = getSampleStyleSheet()
STYLE_TITLE   = ParagraphStyle('TITLE', parent=_styles['Title'], alignment=TA_CENTER, fontSize=16, leading=20, spaceAfter=6)
STYLE_SUBT    = ParagraphStyle('SUBT', parent=_styles['Normal'], alignment=TA_CENTER, fontSize=9, leading=12, textColor=colors.grey)
STYLE_HDR     = ParagraphStyle('HDR', parent=_styles['Normal'], alignment=TA_LEFT, fontSize=10, leading=14, textColor=colors.black, spaceBefore=4, spaceAfter=2)
STYLE_CELL    = ParagraphStyle('CELL', parent=_styles['Normal'], alignment=TA_LEFT, fontSize=9, leading=12)
STYLE_CELL_R  = ParagraphStyle('CELL_R', parent=STYLE_CELL, alignment=TA_RIGHT)
STYLE_NOTE    = ParagraphStyle('NOTE', parent=_styles['Normal'], fontSize=8.5, leading=11, textColor=colors.grey)
STYLE_FOOT    = ParagraphStyle('FOOT', parent=_styles['Normal'], fontSize=8, leading=10, alignment=TA_CENTER, textColor=colors.grey)

def _money_fmt(v: float) -> str:
    return f"{v:,.2f}"

def _indent_label(text: str, level: int) -> Paragraph:
    # 9pt font ~ 3.2mm per indent step looks nice; tune as needed.
    left = max(0, level) * 5.5 * mm
    style = ParagraphStyle(name=f'CELL_L{level}', parent=STYLE_CELL, leftIndent=left)
    return Paragraph(text, style)

# Header/footer callbacks
def _on_page(canvas, doc, header_text: str):
    canvas.saveState()
    w, h = A4
    canvas.setFont("Helvetica", 8)
    canvas.setFillColorRGB(0.5, 0.5, 0.5)
    canvas.drawCentredString(w / 2.0, h - 12 * mm, header_text)
    canvas.drawCentredString(w / 2.0, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


# ---------------------------- PDF API view ------------------------------------

PALETTE = {
    "panel_assets": colors.Color(0.93, 0.97, 1.00),   # very light blue
    "panel_liabs":  colors.Color(1.00, 0.96, 0.93),   # very light peach
    "header_bg":    colors.Color(0.95, 0.95, 0.95),   # light grey
    "stripe_a":     colors.white,
    "stripe_b":     colors.Color(0.985, 0.985, 0.985),# ultra-light row
    "total_bg":     colors.Color(0.92, 0.92, 0.92),
    "note_bg":      colors.Color(0.98, 0.98, 0.98),
    "grid_line":    colors.Color(0.85, 0.85, 0.85),
    "text_dim":     colors.Color(0.45, 0.45, 0.45),
}

# ----------------------------- helpers -----------------------------
def _q2f(x) -> float:
    return float(round(Decimal(str(x or 0)), 2))

def _clean_str(s: str) -> str:
    return (s or "").strip().strip('"').strip("'")

def _flatten_rows(rows, *, level=0, out=None):
    if out is None:
        out = []
    for r in rows or []:
        out.append({"label": r.get("label", ""), "amount": _q2f(r.get("amount", 0)), "level": level})
        kids = r.get("children") or []
        if kids:
            _flatten_rows(kids, level=level + 1, out=out)
    return out

_styles = getSampleStyleSheet()
STYLE_TITLE   = ParagraphStyle('TITLE', parent=_styles['Title'], alignment=TA_CENTER, fontSize=16, leading=20, spaceAfter=6)
STYLE_SUBT    = ParagraphStyle('SUBT', parent=_styles['Normal'], alignment=TA_CENTER, fontSize=9, leading=12, textColor=PALETTE["text_dim"])
STYLE_HDR     = ParagraphStyle('HDR', parent=_styles['Normal'], alignment=TA_LEFT, fontSize=10, leading=14, spaceBefore=2, spaceAfter=1)
STYLE_CELL    = ParagraphStyle('CELL', parent=_styles['Normal'], alignment=TA_LEFT, fontSize=9, leading=12)
STYLE_CELL_R  = ParagraphStyle('CELL_R', parent=STYLE_CELL, alignment=TA_RIGHT)
STYLE_NOTE    = ParagraphStyle('NOTE', parent=_styles['Normal'], fontSize=8.5, leading=11, textColor=PALETTE["text_dim"])

def _money_fmt(v: float) -> str:
    return f"{v:,.2f}"

def _indent_label(text: str, level: int) -> Paragraph:
    left = max(0, level) * 5.5 * mm
    style = ParagraphStyle(name=f'CELL_L{level}', parent=STYLE_CELL, leftIndent=left)
    return Paragraph(text, style)

def _on_page(canvas, doc, header_text: str):
    canvas.saveState()
    w, h = A4
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(PALETTE["text_dim"])
    canvas.drawCentredString(w / 2.0, h - 12 * mm, header_text)
    canvas.drawCentredString(w / 2.0, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()

# ---------------------------- view --------------------------------
class BalanceSheetPDFAPIView(GenericAPIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, *args, **kwargs):
        # -------- query params --------
        try:
            entity_id  = int(request.query_params.get('entity'))
        except (TypeError, ValueError):
            return HttpResponse("Invalid or missing 'entity' parameter.", status=400)

        start_date = _clean_str(request.query_params.get('startdate'))
        end_date   = _clean_str(request.query_params.get('enddate'))
        if not start_date or not end_date:
            return HttpResponse("Both 'startdate' and 'enddate' (YYYY-MM-DD) are required.", status=400)

        level = (request.query_params.get('level') or 'head').lower()
        inventory_source = (request.query_params.get('inventory_source') or 'valuation').lower()
        valuation_method = (request.query_params.get('valuation_method') or 'fifo').lower()

        def _csv_ints(key, default_csv):
            raw = (request.query_params.get(key) or default_csv).split(',')
            vals = []
            for x in raw:
                x = x.strip()
                if not x:
                    continue
                try:
                    vals.append(int(x))
                except ValueError:
                    pass
            return tuple(vals) if vals else tuple(int(x) for x in default_csv.split(','))

        bs_dig = _csv_ints('bs_detailsingroup_values', '3')
        pl_dig = _csv_ints('pl_detailsingroup_values', '2')
        tr_dig = _csv_ints('trading_detailsingroup_values', '1')
        include_current_earnings = (request.query_params.get('include_current_earnings') or 'true').lower() != 'false'

        # -------- data --------
        data = build_balance_sheet_statement(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            level=level,
            bs_detailsingroup_values=bs_dig,
            pl_detailsingroup_values=pl_dig,
            trading_detailsingroup_values=tr_dig,
            include_current_earnings=include_current_earnings,
            inventory_source=inventory_source,
            valuation_method=valuation_method
        )

        assets_flat = _flatten_rows(data.get("assets_rows", []))
        liabs_flat  = _flatten_rows(data.get("liabilities_rows", []))
        max_len     = max(len(assets_flat), len(liabs_flat))

        # -------- build PDF --------
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=22 * mm,
            bottomMargin=16 * mm,
            title="Balance Sheet",
            author="YourApp",
        )

        story = []
        # Title / subtitle
        story.append(Paragraph("Balance Sheet", STYLE_TITLE))
        story.append(Paragraph(
            f"Entity: {entity_id}  •  Period: {data['period']['start']} to {data['period']['end']}  •  As of: {data['as_of']}",
            STYLE_SUBT
        ))
        story.append(Spacer(1, 4 * mm))

        # Header table (with shaded backgrounds)
        hdr = [
            [Paragraph("ASSETS", STYLE_HDR), "", Paragraph("LIABILITIES & EQUITY", STYLE_HDR), ""],
            [Paragraph("Particulars", STYLE_HDR), Paragraph("Amount", STYLE_HDR),
             Paragraph("Particulars", STYLE_HDR), Paragraph("Amount", STYLE_HDR)]
        ]
        col_widths = [70 * mm, 25 * mm, 70 * mm, 25 * mm]
        hdr_tbl = Table(hdr, colWidths=col_widths)
        hdr_tbl.setStyle(TableStyle([
            # Panel background blocks
            ('BACKGROUND', (0, 0), (1, 1), PALETTE["panel_assets"]),
            ('BACKGROUND', (2, 0), (3, 1), PALETTE["panel_liabs"]),
            # Header row subtle overlay
            ('BACKGROUND', (0, 1), (3, 1), PALETTE["header_bg"]),
            # Alignments
            ('ALIGN', (1, 0), (1, 1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, 1), 'RIGHT'),
            # Grid lines subtle
            ('LINEBELOW', (0, 1), (3, 1), 0.6, PALETTE["grid_line"]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 1.5 * mm))

        # Body rows (zebra striping + panel backgrounds)
        body = []
        for i in range(max_len):
            if i < len(assets_flat):
                a = assets_flat[i]
                a_label = _indent_label(str(a["label"]), a["level"])
                a_amt   = Paragraph(_money_fmt(a["amount"]), STYLE_CELL_R)
            else:
                a_label = Paragraph("", STYLE_CELL)
                a_amt   = Paragraph("", STYLE_CELL_R)

            if i < len(liabs_flat):
                l = liabs_flat[i]
                l_label = _indent_label(str(l["label"]), l["level"])
                l_amt   = Paragraph(_money_fmt(l["amount"]), STYLE_CELL_R)
            else:
                l_label = Paragraph("", STYLE_CELL)
                l_amt   = Paragraph("", STYLE_CELL_R)

            body.append([a_label, a_amt, l_label, l_amt])

        body_tbl = Table(body, colWidths=col_widths, repeatRows=0)

        # Build dynamic styles: panel backgrounds + zebra stripes
        body_styles = [
            # Panel soft background for entire body area
            ('BACKGROUND', (0, 0), (1, max_len - 1), PALETTE["panel_assets"]),
            ('BACKGROUND', (2, 0), (3, max_len - 1), PALETTE["panel_liabs"]),
            # Right align amount columns
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]

        # Zebra striping rows across all 4 columns
        for r in range(max_len):
            bg = PALETTE["stripe_b"] if (r % 2) else PALETTE["stripe_a"]
            body_styles.append(('BACKGROUND', (0, r), (3, r), bg))

        # Overlay panel colors lightly by drawing stripes first then panels (keep current order)
        body_tbl.setStyle(TableStyle(body_styles))
        story.append(body_tbl)

        # Totals (highlight bar)
        story.append(Spacer(1, 1.5 * mm))
        totals = [
            [Paragraph("<b>Total</b>", STYLE_CELL), Paragraph(f"<b>{_money_fmt(_q2f(data['assets_total']))}</b>", STYLE_CELL_R),
             Paragraph("<b>Total</b>", STYLE_CELL), Paragraph(f"<b>{_money_fmt(_q2f(data['liabilities_total']))}</b>", STYLE_CELL_R)]
        ]
        tot_tbl = Table(totals, colWidths=col_widths)
        tot_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (3, 0), PALETTE["total_bg"]),
            ('LINEABOVE', (0, 0), (3, 0), 0.8, PALETTE["grid_line"]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('ALIGN', (3, 0), (3, 0), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tot_tbl)

        # Notes block (soft background)
        notes = data.get("notes") or []
        if notes:
            story.append(Spacer(1, 3 * mm))
            story.append(Paragraph("<b>Notes</b>", STYLE_HDR))
            # Build a single-column table to get a background box
            note_rows = [[Paragraph(f"• {n}", STYLE_NOTE)] for n in notes]
            note_tbl = Table(note_rows, colWidths=[(70+25+70+25) * mm])  # full width of main table
            note_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), PALETTE["note_bg"]),
                ('BOX', (0, 0), (-1, -1), 0.6, PALETTE["grid_line"]),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(note_tbl)

        header_text = f"Balance Sheet • Entity {entity_id} • {data['period']['start']} to {data['period']['end']} • Generated {now().strftime('%Y-%m-%d %H:%M')}"
        doc.build(
            story,
            onFirstPage=lambda c, d: _on_page(c, d, header_text),
            onLaterPages=lambda c, d: _on_page(c, d, header_text),
        )

        # -------- response --------
        pdf = buf.getvalue()
        buf.close()
        fname = f'BalanceSheet_entity{entity_id}_{data["period"]["start"]}_to_{data["period"]["end"]}.pdf'
        resp = HttpResponse(content_type="application/pdf")
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        resp.write(pdf)
        return resp



class TrialbalanceExcelApiView(APIView):
    """
    GET /api/reports/trial-balance.xlsx?entity=1&startdate=YYYY-MM-DD&enddate=YYYY-MM-DD
    """
    permission_classes = (permissions.IsAuthenticated,)

    @staticmethod
    def _parse_ymd(s: str) -> date:
        s = str(s or "").strip().replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
        return date.fromisoformat(s)

    def get(self, request, *args, **kwargs):
        entity_id = request.query_params.get("entity")
        start_s   = request.query_params.get("startdate")
        end_s     = request.query_params.get("enddate")

        if entity_id is None or start_s is None or end_s is None:
            return Response(
                {"detail": "Query params required: entity, startdate, enddate (YYYY-MM-DD)"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            entity_id = int(str(entity_id).strip())
            startdate = self._parse_ymd(start_s)
            enddate   = self._parse_ymd(end_s)
        except Exception:
            return Response(
                {"detail": "Dates/entity invalid.", "received": {"entity": entity_id, "startdate": start_s, "enddate": end_s}},
                status=status.HTTP_400_BAD_REQUEST
            )

        if startdate > enddate:
            # still return an empty workbook
            wb = self._build_workbook([], startdate, enddate, entity_id)
            return self._xlsx_response(wb, startdate, enddate)

        # Clamp to FY boundaries, if any
        fy = (entityfinancialyear.objects
              .filter(entity_id=entity_id, finstartyear__date__lte=enddate, finendyear__date__gte=startdate)
              .order_by("-finstartyear").first())
        if fy:
            startdate = max(startdate, fy.finstartyear.date())
            enddate   = min(enddate, fy.finendyear.date())
            if startdate > enddate:
                wb = self._build_workbook([], startdate, enddate, entity_id)
                return self._xlsx_response(wb, startdate, enddate)

        # --- Opening balances by account ---
        opening_acct = (
            JournalLine.objects
            .filter(entity_id=entity_id, entrydate__lt=startdate)
            .values(
                "account_id",
                "account__accounthead_id", "account__accounthead__name",
                "account__creditaccounthead_id", "account__creditaccounthead__name",
                "accounthead_id", "accounthead__name",
            )
            .annotate(
                opening=Sum(
                    Case(
                        When(drcr=True, then=F("amount")),
                        When(drcr=False, then=-F("amount")),
                        default=V(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2),
                    ),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )
        )

        # --- Period activity by account ---
        period_acct = (
            JournalLine.objects
            .filter(entity_id=entity_id, entrydate__gte=startdate, entrydate__lte=enddate)
            .values(
                "account_id",
                "account__accounthead_id", "account__accounthead__name",
                "account__creditaccounthead_id", "account__creditaccounthead__name",
                "accounthead_id", "accounthead__name",
            )
            .annotate(
                debit=Sum(
                    Case(When(drcr=True, then=F("amount")), default=V(0),
                         output_field=DecimalField(max_digits=18, decimal_places=2)),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
                credit=Sum(
                    Case(When(drcr=False, then=F("amount")), default=V(0),
                         output_field=DecimalField(max_digits=18, decimal_places=2)),
                    default=V(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
            )
        )

        def _extract_heads(row):
            ah_id = row.get("account__accounthead_id") or row.get("accounthead_id")
            ah_nm = row.get("account__accounthead__name") or row.get("accounthead__name")
            ch_id = row.get("account__creditaccounthead_id") or ah_id
            ch_nm = row.get("account__creditaccounthead__name") or ah_nm
            return ah_id, ah_nm, ch_id, ch_nm

        per_acct = {}
        for r in opening_acct:
            aid = r["account_id"]
            ah_id, ah_nm, ch_id, ch_nm = _extract_heads(r)
            if ah_id is None and ch_id is None:
                continue
            per_acct[aid] = dict(
                opening=dec(r.get("opening")),
                debit=DZERO, credit=DZERO,
                ah_id=ah_id, ah_name=ah_nm or "",
                ch_id=ch_id, ch_name=ch_nm or "",
            )

        for r in period_acct:
            aid = r["account_id"]
            ah_id, ah_nm, ch_id, ch_nm = _extract_heads(r)
            if ah_id is None and ch_id is None:
                continue
            item = per_acct.setdefault(aid, dict(
                opening=DZERO, debit=DZERO, credit=DZERO,
                ah_id=ah_id, ah_name=ah_nm or "",
                ch_id=ch_id, ch_name=ch_nm or "",
            ))
            item["debit"]  = item["debit"]  + dec(r.get("debit"))
            item["credit"] = item["credit"] + dec(r.get("credit"))
            if not item["ah_name"] and ah_nm:
                item["ah_name"] = ah_nm
            if not item["ch_name"] and ch_nm:
                item["ch_name"] = ch_nm

        # Aggregate by head (positive closing -> natural head, negative -> credit head)
        rows = []
        if per_acct:
            by_head = {}
            for v in per_acct.values():
                opening = dec(v.get("opening"))
                debit   = dec(v.get("debit"))
                credit  = dec(v.get("credit"))
                closing = opening + debit - credit
                hid, hname = (v["ah_id"], v["ah_name"]) if (closing >= DZERO) else (v["ch_id"], v["ch_name"])
                if hid is None:
                    continue
                agg = by_head.setdefault(hid, dict(
                    accounthead=hid,
                    accountheadname=hname or "",
                    openingbalance=DZERO, debit=DZERO, credit=DZERO,
                ))
                agg["openingbalance"] = agg["openingbalance"] + opening
                agg["debit"]          = agg["debit"]          + debit
                agg["credit"]         = agg["credit"]         + credit

            for hid, v in by_head.items():
                opening = dec(v.get("openingbalance"))
                debit   = dec(v.get("debit"))
                credit  = dec(v.get("credit"))
                closing = opening + debit - credit
                rows.append({
                    "accounthead": v["accounthead"],
                    "accountheadname": v["accountheadname"],
                    "obdrcr": "CR" if opening < DZERO else "DR",
                    "openingbalance": abs(opening),
                    "debit": debit,
                    "credit": credit,
                    "drcr": "CR" if closing < DZERO else "DR",
                    "closingbalance": abs(closing),
                })

        # Sort by head name for a stable output
        rows.sort(key=lambda x: (x["accountheadname"] or "").lower())

        # Build and return Excel
        wb = self._build_workbook(rows, startdate, enddate, entity_id)
        return self._xlsx_response(wb, startdate, enddate)

    # ---------- Excel helpers ----------

    def _build_workbook(self, rows, startdate, enddate, entity_id):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        # Styles
        header_fill = PatternFill("solid", fgColor="E8EEF9")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        right  = Alignment(horizontal="right",  vertical="center")

        thin = Side(style="thin", color="DDDDDD")
        border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

        money = NamedStyle(name="money")
        money.number_format = "#,##0.00"
        money.alignment = right
        try:
            wb.add_named_style(money)
        except ValueError:
            # if style already exists
            pass

        # Title
        title = f"Trial Balance | Entity {entity_id} | Period {startdate.isoformat()} to {enddate.isoformat()}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        ws.row_dimensions[1].height = 20

        # Header
        headers = [
            "Account Head",
            "Opening DR/CR",
            "Opening Amount",
            "Debit",
            "Credit",
            "Closing DR/CR",
            "Closing Amount",
            "Head ID",
        ]
        ws.append(headers)
        header_row = 3
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border_all

        # Data
        start_data_row = header_row + 1
        for r in rows:
            ws.append([
                r.get("accountheadname") or "",
                r.get("obdrcr") or "",
                dec(r.get("openingbalance")),
                dec(r.get("debit")),
                dec(r.get("credit")),
                r.get("drcr") or "",
                dec(r.get("closingbalance")),
                r.get("accounthead"),
            ])

        # Apply number style & borders
        last_row = ws.max_row
        for row in ws.iter_rows(min_row=start_data_row, max_row=last_row, min_col=1, max_col=8):
            for idx, cell in enumerate(row, start=1):
                cell.border = border_all
                if idx in (3, 4, 5, 7):  # money columns
                    cell.style = "money"
                elif idx in (2, 6):      # DR/CR
                    cell.alignment = center
                elif idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = right

        # Totals row
        if last_row >= start_data_row:
            total_row = last_row + 1
            ws.cell(row=total_row, column=1, value="TOTAL").font = header_font
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")

            def _sum(col_idx):
                col_letter = get_column_letter(col_idx)
                ws.cell(
                    row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}{start_data_row}:{col_letter}{last_row})"
                ).style = "money"

            _sum(3)  # Opening Amount
            _sum(4)  # Debit
            _sum(5)  # Credit
            _sum(7)  # Closing Amount

            for c in range(1, 9):
                ws.cell(row=total_row, column=c).border = border_all
            ws.row_dimensions[total_row].height = 18

        # Usability: freeze panes & autofilter
        ws.freeze_panes = ws["A4"]
        ws.auto_filter.ref = f"A3:H{ws.max_row}"

        # Column widths
        widths = {
            1: 40,  # Account Head
            2: 12,  # Opening DR/CR
            3: 16,  # Opening Amount
            4: 14,  # Debit
            5: 14,  # Credit
            6: 12,  # Closing DR/CR
            7: 16,  # Closing Amount
            8: 10,  # Head ID
        }
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        return wb

    def _xlsx_response(self, workbook: Workbook, startdate: date, enddate: date) -> HttpResponse:
        buf = BytesIO()
        workbook.save(buf)
        buf.seek(0)
        filename = f"TrialBalance_{startdate.isoformat()}_to_{enddate.isoformat()}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class TradingAccountExcelAPIView(GenericAPIView):
    """
    GET /api/reports/trading-account.xlsx
        ?entity=1
        &startdate=2025-04-01
        &enddate=2025-09-30
        [&valuation_method=fifo|lifo|mwa|wac|latest]
        [&level=head|account|product|voucher]
        [&detailsingroup_values=1,2,3]
        [&inventory_breakdown=true|false]
        [&inventory_include_zero=false|true]
        [&inventory_product_ids=10,20,30]
        [&fold_returns=true|false]
        [&round=2]

    Returns a styled Excel with Debit/Credit schedules side-by-side and totals.
    """
    permission_classes = (permissions.IsAuthenticated,)

    # --------- helpers ----------
    @staticmethod
    def _clean(s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    @staticmethod
    def _to_bool(v: Optional[str], default=False) -> bool:
        if v is None:
            return default
        v = str(v).strip().lower()
        return v in ("1", "true", "yes", "y", "on")

    @staticmethod
    def _to_int_list(v: Optional[str]) -> Optional[List[int]]:
        if v is None or str(v).strip() == "":
            return None
        return [int(x) for x in str(v).split(",") if x.strip()]

    @staticmethod
    def _to_tuple_ints(v: Optional[str], default=(1,)) -> tuple:
        if v is None or str(v).strip() == "":
            return default
        return tuple(int(x) for x in str(v).split(",") if x.strip())

    # --------- excel styling ----------
    @staticmethod
    def _build_styles(wb: Workbook):
        # Number format style
        if "money" not in wb.named_styles:
            money = NamedStyle(name="money")
            money.number_format = "#,##0.00"
            money.font = Font(name="Calibri", size=11)
            wb.add_named_style(money)
        if "small" not in wb.named_styles:
            small = NamedStyle(name="small")
            small.font = Font(name="Calibri", size=10)
            wb.add_named_style(small)

        th_fill = PatternFill("solid", fgColor="E9EEF7")   # light blue header
        sec_fill = PatternFill("solid", fgColor="F8F9FA")  # soft gray for section band
        total_fill = PatternFill("solid", fgColor="FFF2CC")  # pale yellow for totals

        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        return {
            "th_fill": th_fill,
            "sec_fill": sec_fill,
            "total_fill": total_fill,
            "border": border,
            "h_left": Alignment(horizontal="left"),
            "h_right": Alignment(horizontal="right"),
            "h_center": Alignment(horizontal="center"),
            "wrap": Alignment(wrap_text=True),
            "money": "money",
            "small": "small",
        }

    @staticmethod
    def _autosize(ws):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)

    # Recursively write rows with optional children; indent children with prefix spaces
    def _write_rows(self, ws, start_row, col_label, col_amt, rows, styles, band=False, level=0):
        r = start_row
        indent = "  " * level  # simple indent
        for item in rows:
            label = f"{indent}{item.get('label','')}"
            ws.cell(r, col_label, label)
            ws.cell(r, col_amt, item.get("amount", 0))
            ws.cell(r, col_amt).style = styles["money"]

            # banded background for readability
            if band:
                ws.cell(r, col_label).fill = styles["sec_fill"]
                ws.cell(r, col_amt).fill = styles["sec_fill"]

            # light borders
            ws.cell(r, col_label).border = styles["border"]
            ws.cell(r, col_amt).border = styles["border"]

            # children (for opening/closing stock product-wise, and for 'account' grouped children)
            children = item.get("children") or []
            if children:
                r = self._write_rows(ws, r + 1, col_label, col_amt, children, styles, band=False, level=level + 1)
            else:
                r += 1
        return r

    def get(self, request, *args, **kwargs):
        # --------- parse params ----------
        try:
            entity_id = int(self.request.query_params.get("entity"))
            start_date = self._clean(self.request.query_params.get("startdate"))
            end_date = self._clean(self.request.query_params.get("enddate"))
            if not (entity_id and start_date and end_date):
                return Response(
                    {"detail": "Query params required: entity, startdate, enddate (YYYY-MM-DD)"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception:
            return Response({"detail": "Invalid 'entity' or dates."}, status=status.HTTP_400_BAD_REQUEST)

        valuation_method = (self.request.query_params.get("valuation_method") or "fifo").lower()
        level = (self.request.query_params.get("level") or "head").lower()

        detailsingroup_values = self._to_tuple_ints(
            self.request.query_params.get("detailsingroup_values"),
            default=(1,),
        )

        inventory_breakdown = self._to_bool(self.request.query_params.get("inventory_breakdown"), True)
        inventory_include_zero = self._to_bool(self.request.query_params.get("inventory_include_zero"), False)
        inventory_product_ids = self._to_int_list(self.request.query_params.get("inventory_product_ids"))
        fold_returns = self._to_bool(self.request.query_params.get("fold_returns"), True)

        round_decimals = int(self.request.query_params.get("round") or 2)

        # --------- build trading data ----------
        data = build_trading_account_dynamic(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            valuation_method=valuation_method,
            detailsingroup_values=detailsingroup_values,
            level=level,
            fold_returns=fold_returns,
            round_decimals=round_decimals,
            inventory_breakdown=inventory_breakdown,
            inventory_include_zero=inventory_include_zero,
            inventory_product_ids=inventory_product_ids,
        )

        # --------- workbook ----------
        wb = Workbook()
        ws = wb.active
        ws.title = "Trading Account"

        styles = self._build_styles(wb)

        # Header block
        title = "Trading Account"
        period = f"For the period {data['period']['start']} to {data['period']['end']}"
        entity_line = f"Entity: {data['entity_id']}"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

        ws["A1"] = title
        ws["A2"] = period
        ws["A3"] = entity_line

        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"].font = Font(size=12)
        ws["A3"].font = Font(size=11)
        ws["A1"].alignment = styles["h_center"]
        ws["A2"].alignment = styles["h_center"]
        ws["A3"].alignment = styles["h_center"]

        # Params row
        params_line = (
            f"Valuation: {data['params']['valuation_method'].upper()} | "
            f"Level: {data['params']['level']} | "
            f"Groups: {','.join(str(x) for x in data['params']['detailsingroup'])} | "
            f"Breakdown: {data['params']['inventory_breakdown']} | "
            f"Zero-items: {data['params']['inventory_include_zero']}"
        )
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)
        ws["A5"] = params_line
        ws["A5"].style = styles["small"]

        # Table headers (Debit | Credit)
        ws["A7"] = "Debit"
        ws["B7"] = "Amount"
        ws["D7"] = "Credit"
        ws["E7"] = "Amount"

        for cell in ("A7", "B7", "D7", "E7"):
            ws[cell].font = Font(bold=True)
            ws[cell].fill = styles["th_fill"]
            ws[cell].border = styles["border"]
            ws[cell].alignment = styles["h_center"]

        # Write debit rows (A,B) and credit rows (D,E)
        row_start = 8
        r_debit_end = self._write_rows(ws, row_start, 1, 2, data["debit_rows"], styles, band=True)
        r_credit_end = self._write_rows(ws, row_start, 4, 5, data["credit_rows"], styles, band=True)

        # Totals
        total_row = max(r_debit_end, r_credit_end) + 1

        ws.cell(total_row, 1, "Total")
        ws.cell(total_row, 2, data.get("debit_total", 0))
        ws.cell(total_row, 4, "Total")
        ws.cell(total_row, 5, data.get("credit_total", 0))

        for c in (1, 2, 4, 5):
            ws.cell(total_row, c).font = Font(bold=True)
            ws.cell(total_row, c).fill = styles["total_fill"]
            ws.cell(total_row, c).border = styles["border"]
            if c in (2, 5):
                ws.cell(total_row, c).style = styles["money"]

        # Summary box (right side, below totals)
        note_row = total_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
        ws.cell(note_row, 1, "Summary").font = Font(bold=True)

        ws.cell(note_row + 1, 1, "Opening Stock"); ws.cell(note_row + 1, 2, data.get("opening_stock", 0))
        ws.cell(note_row + 2, 1, "Closing Stock"); ws.cell(note_row + 2, 2, data.get("closing_stock", 0))
        ws.cell(note_row + 3, 1, "Gross Profit");  ws.cell(note_row + 3, 2, data.get("gross_profit", 0))
        ws.cell(note_row + 4, 1, "Gross Loss");    ws.cell(note_row + 4, 2, data.get("gross_loss", 0))
        ws.cell(note_row + 5, 1, "COGS (issues)"); ws.cell(note_row + 5, 2, data.get("cogs_from_issues", 0))

        for rr in range(note_row + 1, note_row + 6):
            ws.cell(rr, 1).border = styles["border"]
            ws.cell(rr, 2).border = styles["border"]
            ws.cell(rr, 2).style = styles["money"]

        # Notes (if any)
        notes = data.get("notes", [])
        warn = data.get("warnings", [])
        if notes or warn:
            notes_row = note_row + 7
            ws.cell(notes_row, 1, "Notes").font = Font(bold=True)
            r = notes_row + 1
            for n in notes:
                ws.cell(r, 1, f"• {n}")
                r += 1
            for w in warn:
                ws.cell(r, 1, f"⚠ {w.get('msg')}")
                r += 1

        # Freeze panes (keep headers visible)
        ws.freeze_panes = "A8"

        # Autosize
        self._autosize(ws)

        # --------- return file ----------
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"TradingAccount_{entity_id}_{start_date}_to_{end_date}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class TradingAccountPDFAPIView(GenericAPIView):
    """
    GET /api/reports/trading-account.pdf
        ?entity=1
        &startdate=2025-04-01
        &enddate=2025-09-30
        [&valuation_method=fifo|lifo|mwa|wac|latest]
        [&level=head|account|product|voucher]
        [&detailsingroup_values=1,2,3]
        [&inventory_breakdown=true|false]
        [&inventory_include_zero=false|true]
        [&inventory_product_ids=10,20,30]
        [&fold_returns=true|false]
        [&round=2]
    """
    permission_classes = (permissions.IsAuthenticated,)

    # ---------- helpers ----------
    @staticmethod
    def _clean(s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    @staticmethod
    def _to_bool(v: Optional[str], default=False) -> bool:
        if v is None:
            return default
        v = str(v).strip().lower()
        return v in ("1", "true", "yes", "y", "on")

    @staticmethod
    def _to_int_list(v: Optional[str]) -> Optional[List[int]]:
        if v is None or str(v).strip() == "":
            return None
        return [int(x) for x in str(v).split(",") if x.strip()]

    @staticmethod
    def _to_tuple_ints(v: Optional[str], default=(1,)) -> tuple:
        if v is None or str(v).strip() == "":
            return default
        return tuple(int(x) for x in str(v).split(",") if x.strip())

    @staticmethod
    def _fmt_money(x) -> str:
        try:
            return f"{Decimal(str(x)):.2f}"
        except Exception:
            return "0.00"

    def _flatten_rows(self, rows: List[dict], level: int = 0) -> List[Tuple[str, str]]:
        """
        Flatten nested rows into (label_with_indent, amount_str) tuples.
        Indent using non-breaking spaces; ensure consistent alignment in PDF.
        """
        out: List[Tuple[str, str]] = []
        prefix = "&nbsp;" * (level * 4)
        for r in rows:
            label = r.get("label", "")
            amount = self._fmt_money(r.get("amount", 0))
            out.append((f"{prefix}{label}", amount))
            children = r.get("children") or []
            if children:
                out.extend(self._flatten_rows(children, level + 1))
        return out

    # Header/Footer on each page
    def _on_page(self, canvas, doc, title: str, entity_line: str):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 11)
        canvas.drawString(15 * mm, doc.height + doc.topMargin + 6 * mm, title)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(
            doc.width + doc.leftMargin,
            doc.height + doc.topMargin + 6 * mm,
            entity_line
        )
        # Footer with page number
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(
            doc.width + doc.leftMargin,
            10 * mm,
            f"Page {doc.page}"
        )
        canvas.restoreState()

    def get(self, request, *args, **kwargs):
        # ---------- parse params ----------
        try:
            entity_id = int(self.request.query_params.get("entity"))
            start_date = self._clean(self.request.query_params.get("startdate"))
            end_date = self._clean(self.request.query_params.get("enddate"))
            if not (entity_id and start_date and end_date):
                return Response(
                    {"detail": "Query params required: entity, startdate, enddate (YYYY-MM-DD)"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception:
            return Response({"detail": "Invalid 'entity' or dates."}, status=status.HTTP_400_BAD_REQUEST)

        valuation_method = (self.request.query_params.get("valuation_method") or "fifo").lower()
        level = (self.request.query_params.get("level") or "head").lower()

        detailsingroup_values = self._to_tuple_ints(
            self.request.query_params.get("detailsingroup_values"),
            default=(1,),
        )
        inventory_breakdown = self._to_bool(self.request.query_params.get("inventory_breakdown"), True)
        inventory_include_zero = self._to_bool(self.request.query_params.get("inventory_include_zero"), False)
        inventory_product_ids = self._to_int_list(self.request.query_params.get("inventory_product_ids"))
        fold_returns = self._to_bool(self.request.query_params.get("fold_returns"), True)
        round_decimals = int(self.request.query_params.get("round") or 2)

        # ---------- fetch data ----------
        data = build_trading_account_dynamic(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            valuation_method=valuation_method,
            detailsingroup_values=detailsingroup_values,
            level=level,
            fold_returns=fold_returns,
            round_decimals=round_decimals,
            inventory_breakdown=inventory_breakdown,
            inventory_include_zero=inventory_include_zero,
            inventory_product_ids=inventory_product_ids,
        )

        # ---------- build PDF ----------
        buffer = BytesIO()
        pagesize = landscape(A4)  # wide table
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=18 * mm,
            bottomMargin=15 * mm,
            title="Trading Account",
            author="Finacc",
        )

        styles = getSampleStyleSheet()
        h1 = styles["Heading1"]
        h2 = styles["Heading2"]
        normal = styles["BodyText"]

        # Smaller paragraph for params
        params_style = ParagraphStyle(
            "params",
            parent=normal,
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#444444"),
        )

        story: List = []

        title = "Trading Account"
        period = f"For the period {data['period']['start']} to {data['period']['end']}"
        entity_line = f"Entity: {data['entity_id']}"

        # Top title block
        story.append(Paragraph(title, h1))
        story.append(Paragraph(period, h2))
        story.append(Paragraph(entity_line, normal))
        story.append(Spacer(1, 6))

        # Params ribbon
        params_line = (
            f"<b>Valuation:</b> {data['params']['valuation_method'].upper()} &nbsp;&nbsp; "
            f"<b>Level:</b> {data['params']['level']} &nbsp;&nbsp; "
            f"<b>Groups:</b> {', '.join(str(x) for x in data['params']['detailsingroup'])} &nbsp;&nbsp; "
            f"<b>Breakdown:</b> {data['params']['inventory_breakdown']} &nbsp;&nbsp; "
            f"<b>Zero-items:</b> {data['params']['inventory_include_zero']}"
        )
        story.append(Paragraph(params_line, params_style))
        story.append(Spacer(1, 8))

        # Build table body: we’ll align Debit and Credit columns side-by-side
        debit_flat = self._flatten_rows(data["debit_rows"])
        credit_flat = self._flatten_rows(data["credit_rows"])

        # Keep row counts aligned by padding the shorter side
        max_rows = max(len(debit_flat), len(credit_flat))
        debit_flat += [("", "")] * (max_rows - len(debit_flat))
        credit_flat += [("", "")] * (max_rows - len(credit_flat))

        # Table header
        table_data = [
            ["Debit", "Amount", "Credit", "Amount"]
        ]
        # Table rows
        for i in range(max_rows):
            dl, da = debit_flat[i]
            cl, ca = credit_flat[i]
            table_data.append([Paragraph(dl, normal), self._fmt_money(da),
                               Paragraph(cl, normal), self._fmt_money(ca)])

        # Totals row
        table_data.append([
            Paragraph("<b>Total</b>", normal),
            self._fmt_money(data.get("debit_total", 0)),
            Paragraph("<b>Total</b>", normal),
            self._fmt_money(data.get("credit_total", 0)),
        ])

        # Column widths: label wider, amount narrower
        page_width = pagesize[0] - (doc.leftMargin + doc.rightMargin)
        col_w = [page_width * 0.32, page_width * 0.18, page_width * 0.32, page_width * 0.18]

        t = Table(table_data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0B3D91")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9EEF7")),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -2), 0.25, colors.HexColor("#D0D7E2")),
            ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.HexColor("#A6A6A6")),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF2CC")),
            ("FONT", (0, -1), (-1, -1), "Helvetica-Bold", 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(t)
        story.append(Spacer(1, 10))

        # Summary box
        summary_data = [
            ["Opening Stock", self._fmt_money(data.get("opening_stock", 0))],
            ["Closing Stock", self._fmt_money(data.get("closing_stock", 0))],
            ["Gross Profit",  self._fmt_money(data.get("gross_profit", 0))],
            ["Gross Loss",    self._fmt_money(data.get("gross_loss", 0))],
            ["COGS (issues)", self._fmt_money(data.get("cogs_from_issues", 0))],
        ]
        st = Table(summary_data, colWidths=[page_width * 0.30, page_width * 0.20])
        st.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7E2")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F8F9FA")),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(KeepTogether(st))
        story.append(Spacer(1, 8))

        # Notes
        notes = data.get("notes", [])
        warnings = data.get("warnings", [])
        if notes or warnings:
            story.append(Paragraph("<b>Notes</b>", normal))
            for n in notes:
                story.append(Paragraph(f"• {n}", params_style))
            for w in warnings:
                story.append(Paragraph(f"⚠ {w.get('msg')}", params_style))

        # Build document with header/footer
        def on_page(canvas, doc_local):
            self._on_page(canvas, doc_local, "Trading Account", entity_line)

        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)

        # ---------- return response ----------
        pdf_bytes = buffer.getvalue()
        buffer.close()
        filename = f"TradingAccount_{entity_id}_{start_date}_to_{end_date}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class ProfitAndLossExcelAPIView(GenericAPIView):
    """
    GET /api/reports/profit-and-loss.xlsx
        ?entity=1
        &startdate=2025-04-01
        &enddate=2025-09-30
        [&level=head|account|product|voucher]
        [&valuation_method=fifo|lifo|mwa|wac|latest]      # used for Trading GP/GL
        [&pl_detailsingroup_values=2]
        [&trading_detailsingroup_values=1]
    """
    permission_classes = (permissions.IsAuthenticated,)

    # -------- helpers --------
    @staticmethod
    def _clean(s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    @staticmethod
    def _to_tuple_ints(v: Optional[str], default=(2,)) -> tuple:
        if v is None or str(v).strip() == "":
            return default
        return tuple(int(x) for x in str(v).split(",") if x.strip())

    # -------- styles --------
    @staticmethod
    def _build_styles(wb: Workbook):
        # number format
        if "money" not in wb.named_styles:
            money = NamedStyle(name="money")
            money.number_format = "#,##0.00"
            money.font = Font(name="Calibri", size=11)
            wb.add_named_style(money)
        if "small" not in wb.named_styles:
            small = NamedStyle(name="small")
            small.font = Font(name="Calibri", size=10)
            wb.add_named_style(small)

        th_fill    = PatternFill("solid", fgColor="E9EEF7")   # header
        band_fill  = PatternFill("solid", fgColor="F8F9FA")   # band rows
        total_fill = PatternFill("solid", fgColor="FFF2CC")   # totals

        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        return dict(
            th_fill=th_fill, band_fill=band_fill, total_fill=total_fill,
            border=border,
            h_center=Alignment(horizontal="center"),
            h_right=Alignment(horizontal="right"),
            money="money",
            small="small",
        )

    @staticmethod
    def _autosize(ws):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)

    # write rows with optional nested children (indented via leading spaces)
    def _write_rows(self, ws, start_row: int, col_label: int, col_amt: int, rows: list, styles, band=True, level=0):
        r = start_row
        indent = "  " * level
        for item in rows:
            label = f"{indent}{item.get('label','')}"
            ws.cell(r, col_label, label)
            ws.cell(r, col_amt, item.get("amount", 0))
            ws.cell(r, col_amt).style = styles["money"]

            if band:
                ws.cell(r, col_label).fill = styles["band_fill"]
                ws.cell(r, col_amt).fill = styles["band_fill"]

            ws.cell(r, col_label).border = styles["border"]
            ws.cell(r, col_amt).border = styles["border"]

            # nested children (present when level='account' due to grouping)
            children = item.get("children") or []
            if children:
                r = self._write_rows(ws, r + 1, col_label, col_amt, children, styles, band=False, level=level + 1)
            else:
                r += 1
        return r

    def get(self, request, *args, **kwargs):
        # ---- parse params ----
        try:
            entity_id  = int(request.query_params.get("entity"))
            start_date = self._clean(request.query_params.get("startdate"))
            end_date   = self._clean(request.query_params.get("enddate"))
            if not (entity_id and start_date and end_date):
                return Response({"detail": "Query params required: entity, startdate, enddate (YYYY-MM-DD)"},
                                status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return Response({"detail": "Invalid 'entity' or dates."}, status=status.HTTP_400_BAD_REQUEST)

        level = (request.query_params.get("level") or "head").lower()
        valuation_method = (request.query_params.get("valuation_method") or "fifo").lower()

        pl_detailsingroup_values = self._to_tuple_ints(
            request.query_params.get("pl_detailsingroup_values"), default=(2,)
        )
        trading_detailsingroup_values = self._to_tuple_ints(
            request.query_params.get("trading_detailsingroup_values"), default=(1,)
        )

        # ---- fetch PL data ----
        data = build_profit_and_loss_statement(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            level=level,
            pl_detailsingroup_values=pl_detailsingroup_values,
            trading_detailsingroup_values=trading_detailsingroup_values,
            valuation_method=valuation_method,
        )

        # ---- workbook ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"
        styles = self._build_styles(wb)

        # Header block
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

        ws["A1"] = "Profit & Loss"
        ws["A2"] = f"For the period {data['period']['start']} to {data['period']['end']}"
        ws["A3"] = f"Entity: {data['entity_id']}"

        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"].font = Font(size=12)
        ws["A3"].font = Font(size=11)
        ws["A1"].alignment = styles["h_center"]
        ws["A2"].alignment = styles["h_center"]
        ws["A3"].alignment = styles["h_center"]

        # Params line
        params_line = (
            f"Level: {data['params']['level']} | "
            f"P&L Groups: {','.join(str(x) for x in data['params']['pl_detailsingroup_values'])} | "
            f"Trading Groups (excluded): {','.join(str(x) for x in data['params']['trading_detailsingroup_values'])} | "
            f"Trading valuation: {data['params']['valuation_method'].upper()}"
        )
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)
        ws["A5"] = params_line
        ws["A5"].style = styles["small"]

        # Table headers
        ws["A7"] = "Debit"
        ws["B7"] = "Amount"
        ws["D7"] = "Credit"
        ws["E7"] = "Amount"
        for c in ("A7", "B7", "D7", "E7"):
            ws[c].font = Font(bold=True)
            ws[c].fill = styles["th_fill"]
            ws[c].border = styles["border"]
            ws[c].alignment = styles["h_center"]

        # Write rows
        row_start = 8
        r_debits_end = self._write_rows(ws, row_start, 1, 2, data["debit_rows"], styles, band=True)
        r_credits_end = self._write_rows(ws, row_start, 4, 5, data["credit_rows"], styles, band=True)

        # Totals
        total_row = max(r_debits_end, r_credits_end) + 1
        ws.cell(total_row, 1, "Total")
        ws.cell(total_row, 2, data.get("debit_total", 0)).style = styles["money"]
        ws.cell(total_row, 4, "Total")
        ws.cell(total_row, 5, data.get("credit_total", 0)).style = styles["money"]

        for c in (1, 2, 4, 5):
            ws.cell(total_row, c).font = Font(bold=True)
            ws.cell(total_row, c).fill = styles["total_fill"]
            ws.cell(total_row, c).border = styles["border"]

        # Summary box (GP/GL b/d and Net P/L)
        note_row = total_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
        ws.cell(note_row, 1, "Summary").font = Font(bold=True)

        ws.cell(note_row + 1, 1, "Gross Profit b/d")
        ws.cell(note_row + 1, 2, data.get("gross_profit_brought_down", 0)).style = styles["money"]
        ws.cell(note_row + 2, 1, "Gross Loss b/d")
        ws.cell(note_row + 2, 2, data.get("gross_loss_brought_down", 0)).style = styles["money"]
        ws.cell(note_row + 3, 1, "Net Profit")
        ws.cell(note_row + 3, 2, data.get("net_profit", 0)).style = styles["money"]
        ws.cell(note_row + 4, 1, "Net Loss")
        ws.cell(note_row + 4, 2, data.get("net_loss", 0)).style = styles["money"]

        for rr in range(note_row + 1, note_row + 5):
            ws.cell(rr, 1).border = styles["border"]
            ws.cell(rr, 2).border = styles["border"]

        # Notes / warnings
        notes = data.get("notes", [])
        warns = data.get("warnings", [])
        if notes or warns:
            notes_row = note_row + 6
            ws.cell(notes_row, 1, "Notes").font = Font(bold=True)
            r = notes_row + 1
            for n in notes:
                ws.cell(r, 1, f"• {n}")
                r += 1
            for w in warns:
                ws.cell(r, 1, f"⚠ {w.get('msg')}")
                r += 1

        # Freeze and autosize
        ws.freeze_panes = "A8"
        self._autosize(ws)

        # ---- return file ----
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"ProfitAndLoss_{entity_id}_{start_date}_to_{end_date}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class ProfitAndLossPDFAPIView(GenericAPIView):
    """
    GET /api/reports/profit-and-loss.pdf
        ?entity=1
        &startdate=2025-04-01
        &enddate=2025-09-30
        [&level=head|account|product|voucher]
        [&valuation_method=fifo|lifo|mwa|wac|latest]      # used to bring GP/GL from Trading
        [&pl_detailsingroup_values=2]
        [&trading_detailsingroup_values=1]
    """
    permission_classes = (permissions.IsAuthenticated,)

    # -------- helpers --------
    @staticmethod
    def _clean(s: str) -> str:
        return (s or "").strip().strip('"').strip("'")

    @staticmethod
    def _to_tuple_ints(v: Optional[str], default=(2,)) -> tuple:
        if v is None or str(v).strip() == "":
            return default
        return tuple(int(x) for x in str(v).split(",") if x.strip())

    @staticmethod
    def _fmt_money(x) -> str:
        try:
            return f"{Decimal(str(x)):.2f}"
        except Exception:
            return "0.00"

    # Flatten nested rows (for 'account' level children), indent using nbsp
    def _flatten_rows(self, rows: List[dict], level: int = 0) -> List[Tuple[str, str]]:
        out: List[Tuple[str, str]] = []
        prefix = "&nbsp;" * (level * 4)
        for r in rows:
            label = r.get("label", "")
            amount = self._fmt_money(r.get("amount", 0))
            out.append((f"{prefix}{label}", amount))
            children = r.get("children") or []
            if children:
                out.extend(self._flatten_rows(children, level + 1))
        return out

    # Header/Footer
    def _on_page(self, canvas, doc, title: str, entity_line: str):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 11)
        canvas.drawString(15 * mm, doc.height + doc.topMargin + 6 * mm, title)
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(doc.width + doc.leftMargin,
                               doc.height + doc.topMargin + 6 * mm,
                               entity_line)
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(doc.width + doc.leftMargin, 10 * mm, f"Page {doc.page}")
        canvas.restoreState()

    def get(self, request, *args, **kwargs):
        # ---- parse params ----
        try:
            entity_id  = int(request.query_params.get("entity"))
            start_date = self._clean(request.query_params.get("startdate"))
            end_date   = self._clean(request.query_params.get("enddate"))
            if not (entity_id and start_date and end_date):
                return Response(
                    {"detail": "Query params required: entity, startdate, enddate (YYYY-MM-DD)"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception:
            return Response({"detail": "Invalid 'entity' or dates."}, status=status.HTTP_400_BAD_REQUEST)

        level = (request.query_params.get("level") or "head").lower()
        valuation_method = (request.query_params.get("valuation_method") or "fifo").lower()

        pl_detailsingroup_values = self._to_tuple_ints(
            request.query_params.get("pl_detailsingroup_values"), default=(2,)
        )
        trading_detailsingroup_values = self._to_tuple_ints(
            request.query_params.get("trading_detailsingroup_values"), default=(1,)
        )

        # ---- fetch P&L ----
        data = build_profit_and_loss_statement(
            entity_id=entity_id,
            startdate=start_date,
            enddate=end_date,
            level=level,
            pl_detailsingroup_values=pl_detailsingroup_values,
            trading_detailsingroup_values=trading_detailsingroup_values,
            valuation_method=valuation_method,
        )

        # ---- build PDF ----
        buf = BytesIO()
        pagesize = landscape(A4)
        doc = SimpleDocTemplate(
            buf,
            pagesize=pagesize,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=18 * mm,
            bottomMargin=15 * mm,
            title="Profit & Loss",
            author="Finacc",
        )

        styles = getSampleStyleSheet()
        h1 = styles["Heading1"]
        h2 = styles["Heading2"]
        normal = styles["BodyText"]
        params_style = ParagraphStyle(
            "params",
            parent=normal,
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#444444"),
        )

        story: List = []
        title = "Profit & Loss"
        period = f"For the period {data['period']['start']} to {data['period']['end']}"
        entity_line = f"Entity: {data['entity_id']}"

        # Title block
        story.append(Paragraph(title, h1))
        story.append(Paragraph(period, h2))
        story.append(Paragraph(entity_line, normal))
        story.append(Spacer(1, 6))

        # Params ribbon
        params_line = (
            f"<b>Level:</b> {data['params']['level']} &nbsp;&nbsp; "
            f"<b>P&L Groups:</b> {', '.join(str(x) for x in data['params']['pl_detailsingroup_values'])} &nbsp;&nbsp; "
            f"<b>Trading Groups (excluded):</b> {', '.join(str(x) for x in data['params']['trading_detailsingroup_values'])} &nbsp;&nbsp; "
            f"<b>Trading valuation:</b> {data['params']['valuation_method'].upper()}"
        )
        story.append(Paragraph(params_line, params_style))
        story.append(Spacer(1, 8))

        # Flatten rows and align counts
        debit_flat = self._flatten_rows(data["debit_rows"])
        credit_flat = self._flatten_rows(data["credit_rows"])
        max_rows = max(len(debit_flat), len(credit_flat))
        debit_flat += [("", "")] * (max_rows - len(debit_flat))
        credit_flat += [("", "")] * (max_rows - len(credit_flat))

        # Table data (header + rows + totals)
        table_data = [["Debit", "Amount", "Credit", "Amount"]]
        for i in range(max_rows):
            dl, da = debit_flat[i]
            cl, ca = credit_flat[i]
            table_data.append([
                Paragraph(dl, normal), self._fmt_money(da),
                Paragraph(cl, normal), self._fmt_money(ca)
            ])
        table_data.append([
            Paragraph("<b>Total</b>", normal),
            self._fmt_money(data.get("debit_total", 0)),
            Paragraph("<b>Total</b>", normal),
            self._fmt_money(data.get("credit_total", 0)),
        ])

        # Column widths
        page_width = pagesize[0] - (doc.leftMargin + doc.rightMargin)
        col_w = [page_width * 0.32, page_width * 0.18, page_width * 0.32, page_width * 0.18]

        t = Table(table_data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0B3D91")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9EEF7")),
            ("ALIGN", (1, 1), (1, -2), "RIGHT"),
            ("ALIGN", (3, 1), (3, -2), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -2), 0.25, colors.HexColor("#D0D7E2")),
            ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.HexColor("#A6A6A6")),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF2CC")),
            ("FONT", (0, -1), (-1, -1), "Helvetica-Bold", 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

        # Summary / KPIs (GP/GL b/d & Net P/L)
        summary_data = [
            ["Gross Profit b/d",  self._fmt_money(data.get("gross_profit_brought_down", 0))],
            ["Gross Loss b/d",    self._fmt_money(data.get("gross_loss_brought_down", 0))],
            ["Net Profit",        self._fmt_money(data.get("net_profit", 0))],
            ["Net Loss",          self._fmt_money(data.get("net_loss", 0))],
        ]
        st = Table(summary_data, colWidths=[page_width * 0.30, page_width * 0.20])
        st.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7E2")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F8F9FA")),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(KeepTogether(st))
        story.append(Spacer(1, 8))

        # Notes / warnings
        notes = data.get("notes", [])
        warns = data.get("warnings", [])
        if notes or warns:
            story.append(Paragraph("<b>Notes</b>", normal))
            for n in notes:
                story.append(Paragraph(f"• {n}", params_style))
            for w in warns:
                story.append(Paragraph(f"⚠ {w.get('msg')}", params_style))

        # Build with header/footer
        def on_page(c, d):
            self._on_page(c, d, "Profit & Loss", entity_line)

        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)

        # ---- response ----
        pdf_bytes = buf.getvalue()
        buf.close()
        filename = f"ProfitAndLoss_{entity_id}_{start_date}_to_{end_date}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        return resp


class LedgerJournalExcelAPIView(APIView):
    """
    POST /api/reports/ledger-journal.xlsx

    Body (JSON) — same keys you already accept in `ledgerjournaldetails`, all optional except entity/startdate/enddate:
    {
      "entity": 50,
      "startdate": "2025-04-01",
      "enddate":   "2026-03-31",
      "account": "12,34",                 # <-- pass account(s) here (id or comma list)
      "accounthead": "10,11",
      "transactiontype": "S,P,R",
      "transactionid": "1001,1002",
      "voucherno": "INV-",
      "drcr": "0|1",
      "desc": "some text",
      "include_opening": true,
      "include_total": true,
      "include_zero_balance": false,
      "include_entry_id": false,
      "sub_startdate": null,
      "sub_enddate": null,
      "aggby": null,                      # Non-aggregated export (row detail) is implemented and recommended for Excel
      "sort_by": "name|code|debit|credit|net",
      "sort_dir": "asc|desc"
    }
    """

    permission_classes = (permissions.IsAuthenticated,)

    # ---------- helpers copied/simplified from your view ----------
    @staticmethod
    def _nz(v):
        if v is None:
            return None
        s = str(v).strip()
        return None if s == '' or s.lower() == 'null' else v

    @staticmethod
    def _split_ints(v):
        v = LedgerJournalExcelAPIView._nz(v)
        if v is None: return None
        return [int(x) for x in str(v).split(',') if x.strip()]

    @staticmethod
    def _split_strs(v):
        v = LedgerJournalExcelAPIView._nz(v)
        if v is None: return None
        return [x.strip() for x in str(v).split(',') if x.strip()]

    @staticmethod
    def _dateobj(d):
        if isinstance(d, (datetime, date)):
            return d.date() if isinstance(d, datetime) else d
        return date.fromisoformat(str(d))

    # ---------- Excel helpers ----------
    @staticmethod
    def _auto_width(ws, min_width=8, max_width=50):
        dims = {}
        for row in ws.iter_rows(values_only=True):
            for i, val in enumerate(row, 1):
                if val is None:
                    ln = 0
                else:
                    s = str(val)
                    # consider multi-line cell width
                    ln = max((len(part) for part in s.splitlines()), default=0)
                dims[i] = max(dims.get(i, 0), ln)
        for col_idx, width in dims.items():
            width = min(max(width + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    @staticmethod
    def _styles(wb: Workbook):
        # Header
        hdr = NamedStyle(name="hdr")
        hdr.font = Font(bold=True)
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        hdr.fill = PatternFill("solid", fgColor="F2F2F2")
        hdr.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))

        # Money
        money = NamedStyle(name="money")
        money.number_format = '#,##0.00'
        money.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))

        # Normal cell border
        cell = NamedStyle(name="cell")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        for st in (hdr, money, cell):
            if st.name not in wb.named_styles:
                wb.add_named_style(st)
        return hdr, money, cell

    def post(self, request, *args, **kwargs):
        p = request.data or {}

        # --- required params
        try:
            entity = int(p.get('entity'))
            sdate  = self._dateobj(p.get('startdate'))
            edate  = self._dateobj(p.get('enddate'))
        except Exception:
            return HttpResponse(
                content='{"detail":"Required JSON body: entity (int), startdate (YYYY-MM-DD), enddate (YYYY-MM-DD)"}',
                content_type="application/json",
                status=400
            )

        # Resolve FY
        fy = (entityfinancialyear.objects
              .filter(entity=entity, finstartyear__lte=edate, finendyear__gte=sdate)
              .first())
        if not fy:
            return HttpResponse(
                content='{"detail":"Financial year not found for the given range."}',
                content_type="application/json",
                status=400
            )

        fy_start = self._dateobj(fy.finstartyear)
        # fy_end   = self._dateobj(fy.finendyear)  # not needed in export

        # Base (FY start .. edate)
        base = (JournalLine.objects
                .filter(entity=entity, entrydate__range=(fy_start, edate))
                .select_related('account'))

        include_entry_id = bool(p.get('include_entry_id', False))
        if include_entry_id:
            base = base.select_related('entry')

        # Optional filters
        ah_list   = self._split_ints(p.get('accounthead'))
        acct_list = self._split_ints(p.get('account'))           # <--- pass account(s) here
        ttypes    = self._split_strs(p.get('transactiontype'))
        tids      = self._split_ints(p.get('transactionid'))
        vno       = self._nz(p.get('voucherno'))
        drcr      = self._nz(p.get('drcr'))
        desc_txt  = self._nz(p.get('desc'))

        if ah_list:
            base = base.filter(accounthead_id__in=ah_list)
        if acct_list:
            base = base.filter(account_id__in=acct_list)
        if ttypes:
            base = base.filter(transactiontype__in=ttypes)
        if tids:
            base = base.filter(transactionid__in=tids)
        if vno is not None:
            base = base.filter(voucherno__icontains=vno)
        if drcr in ('0', '1'):
            base = base.filter(drcr=(drcr == '1'))
        if desc_txt is not None:
            base = base.filter(desc__icontains=desc_txt)

        # Sub-window for details
        details_base = base.filter(entrydate__gte=sdate)
        if p.get('sub_startdate') and p.get('sub_enddate'):
            try:
                sub_s = self._dateobj(p['sub_startdate'])
                sub_e = self._dateobj(p['sub_enddate'])
                details_base = details_base.filter(entrydate__range=(sub_s, sub_e))
            except Exception:
                pass  # ignore bad sub dates

        # Annotations
        debit_case   = Case(When(drcr=True,  then=F('amount')))
        credit_case  = Case(When(drcr=False, then=F('amount')))
        debit_annot  = Coalesce(Sum(debit_case),  Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2))
        credit_annot = Coalesce(Sum(credit_case), Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2))

        include_opening   = bool(p.get('include_opening', True))
        include_total     = bool(p.get('include_total', True))
        include_zero      = bool(p.get('include_zero_balance', False))

        # 1) OPENING (FY start .. sdate-1)
        opening_by_acct = {}
        if include_opening:
            open_qs = (base.filter(entrydate__lt=sdate)
                           .values('account_id', 'account__accountname', 'account__accountcode')
                           .annotate(debitamount=debit_annot, creditamount=credit_annot))
            for r in open_qs:
                aid = r['account_id']
                d, c = r['debitamount'], r['creditamount']
                bal = (d or Decimal('0')) - (c or Decimal('0'))
                opening_by_acct[aid] = {
                    'accountname': r['account__accountname'],
                    'accountcode': r.get('account__accountcode'),
                    'debitamount':  bal if bal >= 0 else Decimal('0'),
                    'creditamount': -bal if bal < 0  else Decimal('0'),
                    'entrydate': sdate,
                    'desc': 'Opening',
                    'transactiontype': 'O',
                    'transactionid': -1,
                }

        # 2) DETAILS (non-aggregated; grouped per account/date/type/id)
        values_fields = [
            'account_id', 'account__accountname', 'account__accountcode',
            'entrydate', 'transactiontype', 'transactionid', 'desc'
        ]
        if include_entry_id:
            values_fields.append('entry_id')

        details_qs = (
            details_base
            .values(*values_fields)
            .annotate(
                debitamount=Coalesce(Sum(Case(When(drcr=True,  then=F('amount')))),
                                     Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2)),
                creditamount=Coalesce(Sum(Case(When(drcr=False, then=F('amount')))),
                                      Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2)),
            )
            .order_by('account_id', 'entrydate', 'transactiontype', 'transactionid')
        )

        # Bucket rows per account & compute running
        per_acct: Dict[int, Dict] = {}
        # seed opening
        if include_opening:
            for aid, op in opening_by_acct.items():
                per_acct.setdefault(aid, {
                    'accountname': op['accountname'],
                    'accountcode': op.get('accountcode'),
                    'rows': [],
                    'sum_debit':  Decimal('0'),
                    'sum_credit': Decimal('0')
                })
                per_acct[aid]['rows'].append({
                    **op,
                    'displaydate': op['entrydate'].strftime('%d-%m-%Y'),
                    'balance': (op['debitamount'] - op['creditamount'])
                })
                per_acct[aid]['sum_debit']  += op['debitamount']
                per_acct[aid]['sum_credit'] += op['creditamount']

        for r in details_qs:
            aid = r['account_id']
            per_acct.setdefault(aid, {
                'accountname': r['account__accountname'],
                'accountcode': r.get('account__accountcode'),
                'rows': [],
                'sum_debit':  Decimal('0'),
                'sum_credit': Decimal('0')
            })
            row = {
                'entrydate': r['entrydate'],
                'displaydate': (r['entrydate'].strftime('%d-%m-%Y')
                                if isinstance(r['entrydate'], (datetime, date))
                                else str(r['entrydate'])),
                'transactiontype': r['transactiontype'],
                'transactionid': r['transactionid'],
                'desc': r.get('desc') or '',
                'debitamount': r['debitamount'],
                'creditamount': r['creditamount'],
            }
            if include_entry_id and 'entry_id' in r:
                row['entry_id'] = r['entry_id']

            per_acct[aid]['rows'].append(row)
            per_acct[aid]['sum_debit']  += r['debitamount'] or Decimal('0')
            per_acct[aid]['sum_credit'] += r['creditamount'] or Decimal('0')

        # Totals row + running balance fixup + zero filtering
        result = []
        for aid, bucket in per_acct.items():
            # Running balance
            running = Decimal('0')
            for row in bucket['rows']:
                running += (row['debitamount'] or Decimal('0')) - (row['creditamount'] or Decimal('0'))
                row['balance'] = running

            # Totals row
            if include_total:
                bucket['rows'].append({
                    'entrydate': edate,
                    'displaydate': edate.strftime('%d-%m-%Y'),
                    'transactiontype': 'T',
                    'transactionid': -1,
                    'desc': 'Total',
                    'debitamount':  bucket['sum_debit'],
                    'creditamount': bucket['sum_credit'],
                    'balance': running
                })

            if not include_zero:
                if len(bucket['rows']) == (1 if include_opening else 0):  # only opening and nothing else
                    if abs(float(running)) < 1e-9:
                        continue

            result.append({
                'accountid': aid,
                'accountname': bucket['accountname'],
                'accountcode': bucket.get('accountcode'),
                'rows': bucket['rows']
            })

        # Sort accounts (optional)
        sort_by  = (p.get('sort_by')  or 'name').lower()
        sort_dir = (p.get('sort_dir') or 'asc').lower()
        reverse  = (sort_dir == 'desc')

        def acct_key(a):
            if sort_by == 'code':
                return a.get('accountcode') or 0
            if sort_by in ('debit', 'credit', 'net'):
                d = sum((r['debitamount']  or Decimal('0')) for r in a['rows'])
                c = sum((r['creditamount'] or Decimal('0')) for r in a['rows'])
                if sort_by == 'debit':  return d
                if sort_by == 'credit': return c
                return d - c
            return (a['accountname'] or '').lower()

        result.sort(key=acct_key, reverse=reverse)

        # ---------- Build Excel ----------
        wb = Workbook()
        hdr, money, cell = self._styles(wb)

        # Cover sheet
        ws0 = wb.active
        ws0.title = "Summary"
        ws0.append(["Ledger Journal", None, None, None])
        ws0.append([f"Entity: {entity}", None, None, None])
        ws0.append([f"Period: {sdate:%d-%m-%Y} to {edate:%d-%m-%Y}", None, None, None])
        ws0.append([])
        ws0.append(["Account", "Code", "Total Debit", "Total Credit", "Net (Dr - Cr)"])
        for c in ws0[5]:
            c.style = hdr

        for acct in result:
            d = sum((r['debitamount']  or Decimal('0')) for r in acct['rows'])
            c = sum((r['creditamount'] or Decimal('0')) for r in acct['rows'])
            ws0.append([acct['accountname'], acct.get('accountcode'), d, c, d - c])
            ws0.cell(ws0.max_row, 3).style = money
            ws0.cell(ws0.max_row, 4).style = money
            ws0.cell(ws0.max_row, 5).style = money

        self._auto_width(ws0)

        # One sheet per account
        for acct in result:
            title = (acct['accountname'] or f"Acct-{acct['accountid']}")[:31]
            ws = wb.create_sheet(title=title)

            ws.append([acct['accountname']])
            ws.append([f"Code: {acct.get('accountcode') or ''}"])
            ws.append([f"Period: {sdate:%d-%m-%Y} to {edate:%d-%m-%Y}"])
            ws.append([])

            ws.append(["Date", "Type", "Txn Id", "Narration", "Debit", "Credit", "Balance"])
            for c in ws[5]:
                c.style = hdr

            for r in acct['rows']:
                ws.append([
                    r['displaydate'],
                    r['transactiontype'],
                    r['transactionid'],
                    r.get('desc') or '',
                    r['debitamount'] or Decimal('0'),
                    r['creditamount'] or Decimal('0'),
                    r.get('balance') or Decimal('0'),
                ])
                # money styles
                ws.cell(ws.max_row, 5).style = money
                ws.cell(ws.max_row, 6).style = money
                ws.cell(ws.max_row, 7).style = money

            # Freeze header
            ws.freeze_panes = "A6"
            self._auto_width(ws)

        # Response
        fname = f"LedgerJournal_{entity}_{sdate:%Y%m%d}_{edate:%Y%m%d}.xlsx"
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(resp)
        return resp


class LedgerJournalPDFAPIView(APIView):
    """
    POST /api/reports/ledger-journal.pdf
    Body: same JSON as your ledger Excel/JSON APIs (entity, startdate, enddate, account,...)

    Prints ONLY the per-account Details section (no grand summary).
    """
    permission_classes =  (permissions.IsAuthenticated,)

    def post(self, request: Request, *args, **kwargs):
        # 1) Pull ledger rows by internally invoking your JSON view
        factory = APIRequestFactory()
        proxy_req = factory.post("/internal/ledgerjournaldetails", data=request.data, format='json')
        proxy_req.user = request.user
        # Forward the original auth header (keeps JWT intact)
        proxy_req.META["HTTP_AUTHORIZATION"] = request.META.get("HTTP_AUTHORIZATION", "")

        json_view = ledgerjournaldetails.as_view()
        resp = json_view(proxy_req)
        if hasattr(resp, "status_code") and resp.status_code != 200:
            return resp

        payload = resp.data  # list of accounts

        # 2) Collect header info (accounts & dates)
        def _D(x) -> Decimal:
            try:
                return Decimal(str(x or "0"))
            except Exception:
                return Decimal("0")

        entity = request.data.get("entity")
        sdate  = request.data.get("startdate")
        edate  = request.data.get("enddate")

        # Try to derive selected account names from payload (falls back to IDs filter)
        selected_names = [a.get("accountname") for a in payload if a.get("accountname")]
        if selected_names:
            if len(selected_names) <= 4:
                accounts_label = ", ".join(selected_names)
            else:
                accounts_label = f"Multiple accounts ({len(selected_names)})"
        else:
            accounts_label = str(request.data.get("account") or "All")

        # 3) Prepare PDF
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            rightMargin=28, leftMargin=28, topMargin=66, bottomMargin=40
        )

        styles = getSampleStyleSheet()
        st_title = ParagraphStyle("title", parent=styles["Heading1"], alignment=1, fontSize=14, leading=17, spaceAfter=4)
        st_sub   = ParagraphStyle("sub", parent=styles["Normal"], alignment=1, fontSize=9.5, leading=12, textColor=colors.grey)
        st_h3    = ParagraphStyle("h3", parent=styles["Heading3"], fontSize=12, spaceBefore=4, spaceAfter=4)
        st_th    = ParagraphStyle("th", parent=styles["Normal"], fontSize=10, alignment=1, leading=12)
        st_td    = ParagraphStyle("td", parent=styles["Normal"], fontSize=9, leading=11)

        elems = []
        # Top centered header (title + account + date filter)
        elems.append(Paragraph("Ledger Journal – Details", st_title))
        elems.append(Paragraph(f"Accounts: {accounts_label}", st_sub))
        elems.append(Paragraph(f"Period: {sdate} to {edate}", st_sub))
        elems.append(Spacer(1, 10))

        # 4) Per-account sections (ONLY details, page per account)
        for i, acct in enumerate(payload):
            name = acct.get("accountname") or f"Account {acct.get('accountid')}"
            code = acct.get("accountcode") or ""
            if i > 0:
                elems.append(PageBreak())

            header_line = ""
            elems.append(Paragraph(header_line, st_h3))
            elems.append(Spacer(1, 4))

            # Table header
            tdata = [[
                "Date", "Type", "Txn Id", "Narration",
                "Debit (₹)", "Credit (₹)", "Balance (₹)"
            ]]

            # Rows (keep only details you already provide)
            total_debit = Decimal("0")
            total_credit = Decimal("0")

            for r in acct.get("accounts", []):
                d = _D(r.get("debitamount"))
                c = _D(r.get("creditamount"))
                b = _D(r.get("balance"))

                total_debit += d
                total_credit += c

                tdata.append([
                    str(r.get("displaydate") or ""),
                    str(r.get("transactiontype") or ""),
                    str(r.get("transactionid") or ""),
                    Paragraph((r.get("desc") or ""), st_td),
                    f"{d:.2f}",
                    f"{c:.2f}",
                    f"{b:.2f}",
                ])

            # Optional Totals row (kept minimal; comment out if you truly want only raw rows)
            tdata.append([
                "", "", "", Paragraph("<b>Totals</b>", st_th),
                f"{total_debit:.2f}",
                f"{total_credit:.2f}",
                f"{(total_debit - total_credit):.2f}",
            ])

            table = Table(
                tdata,
                repeatRows=1,
                colWidths=[1.1*inch, 0.8*inch, 0.9*inch, 3.7*inch, 1.1*inch, 1.1*inch, 1.2*inch]
            )
            table.setStyle(TableStyle([
                # Header
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9.5),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#cfd8e3")),

                # Body
                ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 1), (-1, -2), 9),

                # Zebra rows for readability
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#fbfdff")]),

                # Grid & borders
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dde6")),

                # Totals row styling
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fafafa")),
            ]))
            elems.append(table)

        # 5) Build with nice header/footer
        doc.build(elems,
                  onFirstPage=lambda c, d: _pdf_header_footer(c, d, title="Ledger Journal – Details",
                                                              subtitle=f"Accounts: {accounts_label} | {sdate} → {edate}"),
                  onLaterPages=lambda c, d: _pdf_header_footer(c, d, title="Ledger Journal – Details",
                                                               subtitle=f"Accounts: {accounts_label} | {sdate} → {edate}"))

        buf.seek(0)
        http = HttpResponse(buf.getvalue(), content_type="application/pdf")
        fname = f"LedgerJournal_{entity}_{sdate}_{edate}.pdf"
        http["Content-Disposition"] = f'attachment; filename="{fname}"'
        return http


# ---------- centered header + page number footer ----------
def _pdf_header_footer(canvas, doc, title: str, subtitle: str):
    canvas.saveState()
    width, height = landscape(A4)

    # Top center: Title
    canvas.setFont("Helvetica-Bold", 10.5)
    t_w = stringWidth(title, "Helvetica-Bold", 10.5)
    canvas.drawString((width - t_w) / 2.0, height - 26, title)

    # Top center: Subtitle (accounts + date filter)
    canvas.setFont("Helvetica", 9)
    s_w = stringWidth(subtitle, "Helvetica", 9)
    canvas.setFillColor(colors.grey)
    canvas.drawString((width - s_w) / 2.0, height - 40, subtitle)
    canvas.setFillColor(colors.black)

    # Footer: page x
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(width - 28, 26, f"Page {canvas.getPageNumber()}")
    canvas.restoreState()




class DayBookAPIView(APIView):
    """
    Unified Day Book API (single or multi-FY via `sections[]`)

    GET /api/reports/day_book?entity=1&from=YYYY-MM-DD&to=YYYY-MM-DD
      Optional:
        - account_id=INT
        - accounthead=INT
        - voucherno=JV123,PV45
        - vouchertype=Sales,Receipt,...
        - desc_contains=rent
        - min_amount=100.00
        - max_amount=5000
        - posted_only=true|false         (default true)
        - include_empty_days=true|false  (default false)
        - strict_fy=true|false           (default false)

    Output mirrors Cash Book structure but lists ALL transactions.
    Adds CASH-only opening/receipts/payments/closing at grand, section, and per-day levels.
    """
    permission_classes = (permissions.IsAuthenticated,)

    # -------- cash helpers --------

    def _resolve_cash_account_id(self, entity_id: int) -> int:
        const = stocktransconstant()
        acct = const.getcashid(entity_id)
        return int(acct.pk) if hasattr(acct, "pk") else int(acct)

    def _sum_dc(self, qs):
        deb = qs.aggregate(
            s=Coalesce(Sum(
                Case(
                    When(Q(drcr=True) | Q(drcr__iexact='D'), then=F("amount")),
                    default=Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                )
            ), Value(Decimal("0.00")))
        )["s"] or Decimal("0.00")

        crd = qs.aggregate(
            s=Coalesce(Sum(
                Case(
                    When(Q(drcr=False) | Q(drcr__iexact='C'), then=F("amount")),
                    default=Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=16, decimal_places=2),
                )
            ), Value(Decimal("0.00")))
        )["s"] or Decimal("0.00")

        return _D(deb), _D(crd)

    def _cash_opening(self, entity_id: int, cash_id: int, start_date: date) -> Decimal:
        qs = JournalLine.objects.filter(
            entity_id=entity_id, account_id=cash_id, entrydate__lt=start_date
        )
        deb, crd = self._sum_dc(qs)
        return _D(deb - crd)

    def _cash_totals_in_range(self, entity_id: int, cash_id: int, d1: date, d2: date):
        qs = JournalLine.objects.filter(
            entity_id=entity_id, account_id=cash_id,
            entrydate__gte=d1, entrydate__lte=d2
        )
        deb, crd = self._sum_dc(qs)
        return _D(deb), _D(crd)

    # -------- data fetchers --------

    def _fetch_rows(self, qs, sub_from: date, sub_to: date):
        """
        Fetch journal lines for [sub_from, sub_to], annotate debit/credit from (drcr, amount).
        Works whether drcr is Boolean or 'D'/'C' Char.
        """
        return (
            qs.filter(entrydate__gte=sub_from, entrydate__lte=sub_to)
              .select_related("account")  # add accounthead if you need it
              .annotate(
                  debit=Coalesce(
                      Case(
                          When(Q(drcr=True) | Q(drcr__iexact='D'), then=F("amount")),
                          default=Value(Decimal("0.00")),
                          output_field=DecimalField(max_digits=16, decimal_places=2),
                      ),
                      Value(Decimal("0.00"))
                  ),
                  credit=Coalesce(
                      Case(
                          When(Q(drcr=False) | Q(drcr__iexact='C'), then=F("amount")),
                          default=Value(Decimal("0.00")),
                          output_field=DecimalField(max_digits=16, decimal_places=2),
                      ),
                      Value(Decimal("0.00"))
                  ),
                  account_name = F("account__accountname"),
                  voucher_name = F("transactiontype"),
              )
              .order_by("entrydate", "id")
        )

    def _build_day_sections_all(self, rows, include_empty: bool, sub_from: date, sub_to: date):
        """
        Group rows by entrydate; compute per-day debit/credit totals and return:
            (day_sections: list, section_total_debit: Decimal, section_total_credit: Decimal)
        """
        by_date = defaultdict(list)
        for r in rows:
            by_date[r.entrydate].append(r)

        day_cursor = sub_from
        day_sections = []
        total_deb = Decimal("0.00")
        total_cr  = Decimal("0.00")

        while day_cursor <= sub_to:
            items = []
            d_deb = Decimal("0.00")
            d_cr  = Decimal("0.00")

            for r in by_date.get(day_cursor, []):
                debit  = _D(r.debit or 0)
                credit = _D(r.credit or 0)
                d_deb += debit
                d_cr  += credit
                items.append({
                    "date": r.entrydate,                 # serializer expects 'date'
                    "voucherno": r.voucherno or "",
                    "voucher": r.voucher_name or "",
                    "account_id": r.account_id,
                    "account": r.account_name or "",
                    "contra_ac": "",                     # fill if you store it
                    "desc": r.desc or "",
                    "debit": debit,
                    "credit": credit,
                })

            if items or include_empty:
                day_sections.append({
                    "date": day_cursor,
                    "day_debits": _D(d_deb),
                    "day_credits": _D(d_cr),
                    "day_net": _D(d_deb - d_cr),
                    "items": items,
                })

            total_deb += d_deb
            total_cr  += d_cr
            day_cursor = date.fromordinal(day_cursor.toordinal() + 1)

        return day_sections, total_deb, total_cr

    # -------- main GET --------

    def get(self, request):
        # required params
        entity_raw = request.query_params.get("entity")
        from_raw   = request.query_params.get("from")
        to_raw     = request.query_params.get("to")
        if entity_raw is None or from_raw is None or to_raw is None:
            return Response({"detail": "Query params required: entity, from, to (YYYY-MM-DD)"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            entity_id = int(str(entity_raw).strip())
            from_dt   = _parse_iso_date(from_raw)
            to_dt     = _parse_iso_date(to_raw)
        except Exception:
            return Response({"detail": "Dates must be YYYY-MM-DD.",
                             "received": {"from": from_raw, "to": to_raw}}, status=400)
        if to_dt < from_dt:
            return Response({"detail": "'to' must be >= 'from'."}, status=400)

        # toggles
        posted_only        = _bool(request.query_params.get("posted_only"), default=True)
        include_empty_days = _bool(request.query_params.get("include_empty_days"), default=False)
        strict_fy          = _bool(request.query_params.get("strict_fy"), default=False)

        # base queryset for ALL entries in date range (entity scope)
        base = JournalLine.objects.filter(
            entity_id=entity_id,
            entrydate__gte=from_dt,
            entrydate__lte=to_dt
        )
        # if posted_only:
        #     base = base.filter(entry__is_posted=True)

        # optional filters
        account_q = request.query_params.get("account_id")
        if account_q:
            try:
                base = base.filter(account_id=int(account_q))
            except ValueError:
                return Response({"detail": "Invalid account_id."}, status=400)

        head_q = request.query_params.get("accounthead")
        if head_q:
            try:
                base = base.filter(accounthead_id=int(head_q))
            except ValueError:
                return Response({"detail": "Invalid accounthead."}, status=400)

        vno_q = request.query_params.get("voucherno")
        if vno_q:
            vnos = [x.strip() for x in vno_q.split(",") if x.strip()]
            if vnos:
                base = base.filter(voucherno__in=vnos)

        vtype_q = request.query_params.get("vouchertype")
        if vtype_q:
            vtypes = [x.strip() for x in vtype_q.split(",") if x.strip()]
            if vtypes:
                base = base.filter(transactiontype__in=vtypes)

        desc_contains = request.query_params.get("desc_contains")
        if desc_contains:
            base = base.filter(desc__icontains=desc_contains)

        min_amount = request.query_params.get("min_amount")
        if min_amount:
            try:
                base = base.filter(amount__gte=_D(min_amount))
            except Exception:
                return Response({"detail": "Invalid min_amount."}, status=400)

        max_amount = request.query_params.get("max_amount")
        if max_amount:
            try:
                base = base.filter(amount__lte=_D(max_amount))
            except Exception:
                return Response({"detail": "Invalid max_amount."}, status=400)

        # FY handling
        fy_from = _fy_covering_date(entity_id, from_dt)
        fy_to   = _fy_covering_date(entity_id, to_dt)

        # cash setup
        cash_id = self._resolve_cash_account_id(entity_id)

        # ---- single-section builder (with cash grand/section/day stats) ----
        def _single_section_payload():
            rows = self._fetch_rows(base, from_dt, to_dt)
            day_sections, sec_deb, sec_cr = self._build_day_sections_all(rows, include_empty_days, from_dt, to_dt)

            # cash grand/section
            cash_opening = self._cash_opening(entity_id, cash_id, from_dt)
            cash_rec, cash_pay = self._cash_totals_in_range(entity_id, cash_id, from_dt, to_dt)
            cash_closing = _D(cash_opening + cash_rec - cash_pay)

            # per-day cash (pre-aggregated)
            per_day_cash_qs = (
                JournalLine.objects
                .filter(
                    entity_id=entity_id,
                    account_id=cash_id,
                    entrydate__gte=from_dt,   # use from_dt in single-section; sub_from in multi-FY
                    entrydate__lte=to_dt      # use to_dt in single-section;   sub_to   in multi-FY
                )
                .values("entrydate")
                .annotate(
                    day_dr=Coalesce(Sum(
                        Case(
                            When(Q(drcr=True) | Q(drcr__iexact='D'), then=F("amount")),
                            default=Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=16, decimal_places=2),
                        )
                    ), Value(Decimal("0.00"))),
                    day_cr=Coalesce(Sum(
                        Case(
                            When(Q(drcr=False) | Q(drcr__iexact='C'), then=F("amount")),
                            default=Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=16, decimal_places=2),
                        )
                    ), Value(Decimal("0.00"))),
                )
                .values_list("entrydate", "day_dr", "day_cr")
            )

            per_day_cash = { d: (_D(dr or 0), _D(cr or 0)) for d, dr, cr in per_day_cash_qs }

            running = _D(cash_opening)
            for ds in day_sections:
                d = ds["date"]
                dr = _D(0); cr = _D(0)
                if d in per_day_cash:
                    dr = _D(per_day_cash[d][0] or 0)
                    cr = _D(per_day_cash[d][1] or 0)
                ds["cash_day_opening"]  = _D(running)
                ds["cash_day_receipts"] = _D(dr)
                ds["cash_day_payments"] = _D(cr)
                running = _D(running + dr - cr)
                ds["cash_day_closing"]  = _D(running)

            section_obj = {
                "fy_name": (_fy_name(fy_from) if fy_from and fy_to and fy_from.id == fy_to.id else ""),
                "fy_start": fy_from.finstartyear.date() if fy_from else None,
                "fy_end": fy_from.finendyear.date() if fy_from else None,
                "from_date": from_dt,
                "to_date": to_dt,
                "total_debits": _D(sec_deb),
                "total_credits": _D(sec_cr),
                "net_movement": _D(sec_deb - sec_cr),
                "day_sections": day_sections,
                # cash fields
                "cash_opening": _D(cash_opening),
                "cash_total_receipts": _D(cash_rec),
                "cash_total_payments": _D(cash_pay),
                "cash_closing": _D(cash_closing),
            }

            return {
                "entity": entity_id,
                "from_date": from_dt,
                "to_date": to_dt,
                "spans_multiple_fy": False,
                "grand_total_debits": _D(sec_deb),
                "grand_total_credits": _D(sec_cr),
                "grand_net_movement": _D(sec_deb - sec_cr),
                "sections": [section_obj],
                # cash grand (same as section in single-FY)
                "grand_cash_opening": _D(cash_opening),
                "grand_cash_total_receipts": _D(cash_rec),
                "grand_cash_total_payments": _D(cash_pay),
                "grand_cash_closing": _D(cash_closing),
            }

        # ---- choose single vs multi-FY ----
        if not fy_from or not fy_to:
            if strict_fy:
                return Response(
                    {"detail": "Requested dates are not fully covered by financial year setup.",
                     "hint": "Add FY rows for these dates or call without strict_fy."},
                    status=400,
                )
            payload = _single_section_payload()

        elif fy_from.id == fy_to.id:
            payload = _single_section_payload()

        else:
            # multi-FY
            fys = _fys_overlapping_range(entity_id, from_dt, to_dt)
            if not fys:
                return Response({"detail": "No financial year overlaps the requested range."}, status=400)

            sections = []
            grand_deb = Decimal("0.00")
            grand_cr  = Decimal("0.00")

            grand_cash_open  = self._cash_opening(entity_id, cash_id, from_dt)
            grand_cash_rece  = _D(0)
            grand_cash_pay   = _D(0)
            running_cash     = _D(grand_cash_open)

            for fy in fys:
                sub_from, sub_to = _clip_to_fy(from_dt, to_dt, fy)

                rows = self._fetch_rows(base, sub_from, sub_to)
                day_sections, sec_deb, sec_cr = self._build_day_sections_all(
                    rows, include_empty_days, sub_from, sub_to
                )

                # cash per-section
                sec_cash_open = _D(running_cash)
                sec_cash_rec, sec_cash_pay = self._cash_totals_in_range(entity_id, cash_id, sub_from, sub_to)
                sec_cash_close = _D(sec_cash_open + sec_cash_rec - sec_cash_pay)

                # per-day cash aggregate for this FY window
                per_day_cash_qs = (
                    JournalLine.objects
                    .filter(
                        entity_id=entity_id,
                        account_id=cash_id,
                        entrydate__gte=sub_from,   # use from_dt in single-section; sub_from in multi-FY
                        entrydate__lte=sub_to      # use to_dt in single-section;   sub_to   in multi-FY
                    )
                    .values("entrydate")
                    .annotate(
                        day_dr=Coalesce(Sum(
                            Case(
                                When(Q(drcr=True) | Q(drcr__iexact='D'), then=F("amount")),
                                default=Value(Decimal("0.00")),
                                output_field=DecimalField(max_digits=16, decimal_places=2),
                            )
                        ), Value(Decimal("0.00"))),
                        day_cr=Coalesce(Sum(
                            Case(
                                When(Q(drcr=False) | Q(drcr__iexact='C'), then=F("amount")),
                                default=Value(Decimal("0.00")),
                                output_field=DecimalField(max_digits=16, decimal_places=2),
                            )
                        ), Value(Decimal("0.00"))),
                    )
                    .values_list("entrydate", "day_dr", "day_cr")
                )

                per_day_cash = { d: (_D(dr or 0), _D(cr or 0)) for d, dr, cr in per_day_cash_qs }

                # attach per-day cash
                local_running = _D(sec_cash_open)
                for ds in day_sections:
                    d = ds["date"]
                    dr = _D(0); cr = _D(0)
                    if d in per_day_cash:
                        dr = _D(per_day_cash[d][0] or 0)
                        cr = _D(per_day_cash[d][1] or 0)
                    ds["cash_day_opening"]  = _D(local_running)
                    ds["cash_day_receipts"] = _D(dr)
                    ds["cash_day_payments"] = _D(cr)
                    local_running = _D(local_running + dr - cr)
                    ds["cash_day_closing"]  = _D(local_running)

                sections.append({
                    "fy_name": _fy_name(fy),
                    "fy_start": fy.finstartyear.date(),
                    "fy_end": fy.finendyear.date(),
                    "from_date": sub_from,
                    "to_date": sub_to,
                    "total_debits": _D(sec_deb),
                    "total_credits": _D(sec_cr),
                    "net_movement": _D(sec_deb - sec_cr),
                    "day_sections": day_sections,
                    # cash section
                    "cash_opening": _D(sec_cash_open),
                    "cash_total_receipts": _D(sec_cash_rec),
                    "cash_total_payments": _D(sec_cash_pay),
                    "cash_closing": _D(sec_cash_close),
                })

                # rollups
                grand_deb += sec_deb
                grand_cr  += sec_cr
                grand_cash_rece += sec_cash_rec
                grand_cash_pay  += sec_cash_pay
                running_cash     = _D(sec_cash_close)

            payload = {
                "entity": entity_id,
                "from_date": from_dt,
                "to_date": to_dt,
                "spans_multiple_fy": True,
                "grand_total_debits": _D(grand_deb),
                "grand_total_credits": _D(grand_cr),
                "grand_net_movement": _D(grand_deb - grand_cr),
                "sections": sections,
                # cash grand
                "grand_cash_opening": _D(grand_cash_open),
                "grand_cash_total_receipts": _D(grand_cash_rece),
                "grand_cash_total_payments": _D(grand_cash_pay),
                "grand_cash_closing": _D(running_cash),
            }

        return Response(DaybookUnifiedSerializer(payload).data, status=200)


class DayBookExcelAPIView(APIView):
    """
    Download Day Book as Excel (Summary + Detail)
    Mirrors DayBookAPIView logic including cash stats at grand/section/day levels.

    GET /api/reports/day_book.xlsx?entity=1&from=YYYY-MM-DD&to=YYYY-MM-DD
         [&account_id=..&accounthead=..&voucherno=..&vouchertype=..&desc_contains=..]
         [&min_amount=..&max_amount=..]
         [&posted_only=true|false]        (default true)
         [&include_empty_days=true|false] (default false)
         [&strict_fy=true|false]          (default false)
    """
    permission_classes = (permissions.IsAuthenticated,)

    # ---------- tiny copies of your common helpers (same signatures/behavior) ----------
    def _parse_iso_date(self, s: str) -> date:
        s = str(s or "").strip().replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
        return date.fromisoformat(s)  # strict YYYY-MM-DD

    def _bool(self, s, default=False) -> bool:
        if s is None:
            return default
        s = str(s).strip().lower()
        return s in ("1", "true", "t", "yes", "y", "on")

    # ---------- payload builder: delegates to DayBookAPIView methods to stay in sync ----------
    def _build_payload(self, request):
        """
        Rebuild the exact DayBook payload by reusing DayBookAPIView's helpers and logic.
        """
        # Parse core params
        entity_raw = request.query_params.get("entity")
        from_raw   = request.query_params.get("from")
        to_raw     = request.query_params.get("to")
        if entity_raw is None or from_raw is None or to_raw is None:
            return {"error": {"detail": "Query params required: entity, from, to (YYYY-MM-DD)"}}, 400

        try:
            entity_id = int(str(entity_raw).strip())
            from_dt   = self._parse_iso_date(from_raw)
            to_dt     = self._parse_iso_date(to_raw)
        except Exception:
            return {"error": {"detail": "Dates must be YYYY-MM-DD.", "received": {"from": from_raw, "to": to_raw}}}, 400

        if to_dt < from_dt:
            return {"error": {"detail": "'to' must be >= 'from'."}}, 400

        posted_only        = self._bool(request.query_params.get("posted_only"), default=True)
        include_empty_days = self._bool(request.query_params.get("include_empty_days"), default=False)
        strict_fy          = self._bool(request.query_params.get("strict_fy"), default=False)

        # Instantiate the JSON view so we can reuse its private methods 1:1
        core = DayBookAPIView()

        # Build the same base queryset and filters (copy from DayBookAPIView.get)
        base = JournalLine.objects.filter(
            entity_id=entity_id,
            entrydate__gte=from_dt,
            entrydate__lte=to_dt
        )
        # if posted_only:
        #     base = base.filter(entry__is_posted=True)

        account_q = request.query_params.get("account_id")
        if account_q:
            try:
                base = base.filter(account_id=int(account_q))
            except ValueError:
                return {"error": {"detail": "Invalid account_id."}}, 400

        head_q = request.query_params.get("accounthead")
        if head_q:
            try:
                base = base.filter(accounthead_id=int(head_q))
            except ValueError:
                return {"error": {"detail": "Invalid accounthead."}}, 400

        vno_q = request.query_params.get("voucherno")
        if vno_q:
            vnos = [x.strip() for x in vno_q.split(",") if x.strip()]
            if vnos:
                base = base.filter(voucherno__in=vnos)

        vtype_q = request.query_params.get("vouchertype")
        if vtype_q:
            vtypes = [x.strip() for x in vtype_q.split(",") if x.strip()]
            if vtypes:
                base = base.filter(transactiontype__in=vtypes)

        desc_contains = request.query_params.get("desc_contains")
        if desc_contains:
            base = base.filter(desc__icontains=desc_contains)

        min_amount = request.query_params.get("min_amount")
        if min_amount:
            try:
                base = base.filter(amount__gte=_D(min_amount))
            except Exception:
                return {"error": {"detail": "Invalid min_amount."}}, 400

        max_amount = request.query_params.get("max_amount")
        if max_amount:
            try:
                base = base.filter(amount__lte=_D(max_amount))
            except Exception:
                return {"error": {"detail": "Invalid max_amount."}}, 400

        # FY + Cash setup
        fy_from = _fy_covering_date(entity_id, from_dt)
        fy_to   = _fy_covering_date(entity_id, to_dt)
        cash_id = core._resolve_cash_account_id(entity_id)

        # Helper to build per-day cash dict (alias-safe)
        def per_day_cash_map(win_from, win_to):
            per_day_cash_qs = (
                JournalLine.objects
                .filter(entity_id=entity_id, account_id=cash_id,
                        entrydate__gte=win_from, entrydate__lte=win_to)
                .values("entrydate")
                .annotate(
                    day_dr=Coalesce(Sum(
                        Case(
                            When(Q(drcr=True) | Q(drcr__iexact='D'), then=F("amount")),
                            default=Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=16, decimal_places=2),
                        )
                    ), Value(Decimal("0.00"))),
                    day_cr=Coalesce(Sum(
                        Case(
                            When(Q(drcr=False) | Q(drcr__iexact='C'), then=F("amount")),
                            default=Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=16, decimal_places=2),
                        )
                    ), Value(Decimal("0.00"))),
                )
                .values_list("entrydate", "day_dr", "day_cr")
            )
            return {d: (_D(dr or 0), _D(cr or 0)) for d, dr, cr in per_day_cash_qs}

        # Build payload (same as DayBookAPIView)
        if not fy_from or not fy_to:
            if strict_fy:
                return {"error": {"detail": "Requested dates are not fully covered by financial year setup.",
                                  "hint": "Add FY rows or call without strict_fy."}}, 400

            rows = core._fetch_rows(base, from_dt, to_dt)
            day_sections, sec_deb, sec_cr = core._build_day_sections_all(rows, include_empty_days, from_dt, to_dt)
            cash_opening = core._cash_opening(entity_id, cash_id, from_dt)
            cash_rec, cash_pay = core._cash_totals_in_range(entity_id, cash_id, from_dt, to_dt)
            cash_closing = _D(cash_opening + cash_rec - cash_pay)

            pdc = per_day_cash_map(from_dt, to_dt)
            running = _D(cash_opening)
            for ds in day_sections:
                d = ds["date"]
                dr, cr = pdc.get(d, (_D(0), _D(0)))
                ds["cash_day_opening"]  = running
                ds["cash_day_receipts"] = dr
                ds["cash_day_payments"] = cr
                running = _D(running + dr - cr)
                ds["cash_day_closing"]  = running

            section_obj = {
                "fy_name": (_fy_name(fy_from) if fy_from and fy_to and fy_from.id == fy_to.id else ""),
                "fy_start": fy_from.finstartyear.date() if fy_from else None,
                "fy_end": fy_from.finendyear.date() if fy_from else None,
                "from_date": from_dt, "to_date": to_dt,
                "total_debits": _D(sec_deb), "total_credits": _D(sec_cr),
                "net_movement": _D(sec_deb - sec_cr),
                "day_sections": day_sections,
                "cash_opening": cash_opening, "cash_total_receipts": cash_rec,
                "cash_total_payments": cash_pay, "cash_closing": cash_closing,
            }

            payload = {
                "entity": entity_id, "from_date": from_dt, "to_date": to_dt,
                "spans_multiple_fy": False,
                "grand_total_debits": _D(sec_deb),
                "grand_total_credits": _D(sec_cr),
                "grand_net_movement": _D(sec_deb - sec_cr),
                "sections": [section_obj],
                "grand_cash_opening": cash_opening,
                "grand_cash_total_receipts": cash_rec,
                "grand_cash_total_payments": cash_pay,
                "grand_cash_closing": cash_closing,
            }
            return payload, 200

        if fy_from.id == fy_to.id:
            rows = core._fetch_rows(base, from_dt, to_dt)
            day_sections, sec_deb, sec_cr = core._build_day_sections_all(rows, include_empty_days, from_dt, to_dt)

            cash_opening = core._cash_opening(entity_id, cash_id, from_dt)
            cash_rec, cash_pay = core._cash_totals_in_range(entity_id, cash_id, from_dt, to_dt)
            cash_closing = _D(cash_opening + cash_rec - cash_pay)

            pdc = per_day_cash_map(from_dt, to_dt)
            running = _D(cash_opening)
            for ds in day_sections:
                d = ds["date"]
                dr, cr = pdc.get(d, (_D(0), _D(0)))
                ds["cash_day_opening"]  = running
                ds["cash_day_receipts"] = dr
                ds["cash_day_payments"] = cr
                running = _D(running + dr - cr)
                ds["cash_day_closing"]  = running

            section_obj = {
                "fy_name": (_fy_name(fy_from) if fy_from and fy_to and fy_from.id == fy_to.id else ""),
                "fy_start": fy_from.finstartyear.date() if fy_from else None,
                "fy_end": fy_from.finendyear.date() if fy_from else None,
                "from_date": from_dt, "to_date": to_dt,
                "total_debits": _D(sec_deb), "total_credits": _D(sec_cr),
                "net_movement": _D(sec_deb - sec_cr),
                "day_sections": day_sections,
                "cash_opening": cash_opening, "cash_total_receipts": cash_rec,
                "cash_total_payments": cash_pay, "cash_closing": cash_closing,
            }

            payload = {
                "entity": entity_id, "from_date": from_dt, "to_date": to_dt,
                "spans_multiple_fy": False,
                "grand_total_debits": _D(sec_deb),
                "grand_total_credits": _D(sec_cr),
                "grand_net_movement": _D(sec_deb - sec_cr),
                "sections": [section_obj],
                "grand_cash_opening": cash_opening,
                "grand_cash_total_receipts": cash_rec,
                "grand_cash_total_payments": cash_pay,
                "grand_cash_closing": cash_closing,
            }
            return payload, 200

        # multi-FY
        fys = _fys_overlapping_range(entity_id, from_dt, to_dt)
        if not fys:
            return {"error": {"detail": "No financial year overlaps the requested range."}}, 400

        sections = []
        grand_deb = Decimal("0.00")
        grand_cr  = Decimal("0.00")
        grand_cash_open  = core._cash_opening(entity_id, cash_id, from_dt)
        grand_cash_rece  = _D(0)
        grand_cash_pay   = _D(0)
        running_cash     = _D(grand_cash_open)

        for fy in fys:
            sub_from, sub_to = _clip_to_fy(from_dt, to_dt, fy)

            rows = core._fetch_rows(base, sub_from, sub_to)
            day_sections, sec_deb, sec_cr = core._build_day_sections_all(rows, include_empty_days, sub_from, sub_to)

            sec_cash_open = _D(running_cash)
            sec_cash_rec, sec_cash_pay = core._cash_totals_in_range(entity_id, cash_id, sub_from, sub_to)
            sec_cash_close = _D(sec_cash_open + sec_cash_rec - sec_cash_pay)

            pdc = per_day_cash_map(sub_from, sub_to)
            local_running = _D(sec_cash_open)
            for ds in day_sections:
                d = ds["date"]
                dr, cr = pdc.get(d, (_D(0), _D(0)))
                ds["cash_day_opening"]  = local_running
                ds["cash_day_receipts"] = dr
                ds["cash_day_payments"] = cr
                local_running = _D(local_running + dr - cr)
                ds["cash_day_closing"]  = local_running

            sections.append({
                "fy_name": _fy_name(fy),
                "fy_start": fy.finstartyear.date(),
                "fy_end": fy.finendyear.date(),
                "from_date": sub_from, "to_date": sub_to,
                "total_debits": _D(sec_deb), "total_credits": _D(sec_cr),
                "net_movement": _D(sec_deb - sec_cr),
                "day_sections": day_sections,
                "cash_opening": sec_cash_open,
                "cash_total_receipts": sec_cash_rec,
                "cash_total_payments": sec_cash_pay,
                "cash_closing": sec_cash_close,
            })

            grand_deb += sec_deb
            grand_cr  += sec_cr
            grand_cash_rece += sec_cash_rec
            grand_cash_pay  += sec_cash_pay
            running_cash     = _D(sec_cash_close)

        payload = {
            "entity": entity_id, "from_date": from_dt, "to_date": to_dt,
            "spans_multiple_fy": True,
            "grand_total_debits": _D(grand_deb),
            "grand_total_credits": _D(grand_cr),
            "grand_net_movement": _D(grand_deb - grand_cr),
            "sections": sections,
            "grand_cash_opening": _D(grand_cash_open),
            "grand_cash_total_receipts": _D(grand_cash_rece),
            "grand_cash_total_payments": _D(grand_cash_pay),
            "grand_cash_closing": _D(running_cash),
        }
        return payload, 200

    # ---------- Excel helpers ----------
    def _styles(self, wb: Workbook):
        thin = Side(style="thin", color="000000")
        hdr = NamedStyle(name="hdr")
        hdr.font = Font(bold=True)
        hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hdr.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        num = NamedStyle(name="num")
        num.number_format = numbers.BUILTIN_FORMATS[4]  # "#,##0.00"
        num.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        txt = NamedStyle(name="txt")
        txt.alignment = Alignment(vertical="top", wrap_text=True)
        txt.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for st in (hdr, num, txt):
            if st.name not in wb.named_styles:
                wb.add_named_style(st)
        return hdr, num, txt

    def _autosize(self, ws):
        for col in range(1, ws.max_column + 1):
            maxlen = 10
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                l = len(str(v))
                if l > maxlen:
                    maxlen = l
            ws.column_dimensions[get_column_letter(col)].width = min(maxlen + 2, 60)

    # ---------- GET ----------
    def get(self, request, *_args, **_kwargs):
        payload, code = self._build_payload(request)
        if code != 200:
            return Response(payload.get("error", payload), status=code)

        # If you want, validate via serializer to normalize types:
        # ser = DaybookUnifiedSerializer(data=payload)
        # ser.is_valid(raise_exception=True)
        # data = ser.validated_data
        data = payload

        wb = Workbook()
        hdr, num, txt = self._styles(wb)

        # ---------------- Summary sheet ----------------
        ws = wb.active
        ws.title = "Summary"

        ws.append(["Entity", data["entity"]])
        ws.append(["From", data["from_date"]])
        ws.append(["To",   data["to_date"]])
        ws.append([])
        ws.append(["Grand Totals"])
        ws.append(["Total Debits", "Total Credits", "Net Movement",
                   "Cash Opening", "Cash Receipts", "Cash Payments", "Cash Closing"])
        ws.append([
            data["grand_total_debits"],
            data["grand_total_credits"],
            data["grand_total_debits"] - data["grand_total_credits"],
            data.get("grand_cash_opening", Decimal("0.00")),
            data.get("grand_cash_total_receipts", Decimal("0.00")),
            data.get("grand_cash_total_payments", Decimal("0.00")),
            data.get("grand_cash_closing", Decimal("0.00")),
        ])
        for c in ws[6]:
            c.style = hdr
        for c in ws[7]:
            c.style = num

        ws.append([])
        ws.append(["Sections"])
        ws.append([
            "FY Name", "Period From", "Period To",
            "Total Debits", "Total Credits", "Net Movement",
            "Cash Opening", "Cash Receipts", "Cash Payments", "Cash Closing"
        ])
        for c in ws[9]:
            c.style = hdr

        for s in data["sections"]:
            ws.append([
                s.get("fy_name", ""),
                s["from_date"], s["to_date"],
                s["total_debits"], s["total_credits"],
                s["net_movement"],
                s.get("cash_opening", Decimal("0.00")),
                s.get("cash_total_receipts", Decimal("0.00")),
                s.get("cash_total_payments", Decimal("0.00")),
                s.get("cash_closing", Decimal("0.00")),
            ])
            # apply numeric style to numeric columns
            for idx in (4, 5, 6, 7, 8, 9, 10):
                ws.cell(row=ws.max_row, column=idx).style = num

        self._autosize(ws)

        # ---------------- Detail sheet ----------------
        wd = wb.create_sheet("Detail")
        wd.append([
            "Date", "Voucher No", "Voucher Type",
            "Account ID", "Account", "Description",
            "Debit", "Credit",
            "Cash Day Opening", "Cash Day Receipts", "Cash Day Payments", "Cash Day Closing"
        ])
        for c in wd[1]:
            c.style = hdr

        # Flatten day sections
        for s in data["sections"]:
            for ds in s["day_sections"]:
                # Row for the day header (with cash day numbers)
                wd.append([
                    ds["date"], "", "",
                    "", "", "Day Totals",
                    ds["day_debits"], ds["day_credits"],
                    ds.get("cash_day_opening", Decimal("0.00")),
                    ds.get("cash_day_receipts", Decimal("0.00")),
                    ds.get("cash_day_payments", Decimal("0.00")),
                    ds.get("cash_day_closing", Decimal("0.00")),
                ])
                # numeric styling
                for idx in (7, 8, 9, 10, 11, 12):
                    wd.cell(row=wd.max_row, column=idx).style = num

                # Lines for that day
                for it in ds["items"]:
                    wd.append([
                        it["date"], it.get("voucherno", ""), it.get("voucher", ""),
                        it.get("account_id", ""), it.get("account", ""), it.get("desc", ""),
                        it["debit"], it["credit"],
                        "", "", "", ""
                    ])
                    wd.cell(row=wd.max_row, column=7).style = num
                    wd.cell(row=wd.max_row, column=8).style = num

                # spacer
                wd.append([])

        self._autosize(wd)

        # finalize response
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"DayBook_{data['entity']}_{data['from_date']}_{data['to_date']}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class DayBookPDFAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def _fmt(self, x):
        if x is None or x == "":
            return ""
        try:
            return f"{Decimal(x):,.2f}"
        except Exception:
            return str(x)

    def _payload(self, request):
        core = DayBookExcelAPIView()
        return core._build_payload(request)   # -> (payload, status_code)

    # ---------- page chrome ----------
    def _header_footer(self, canvas, doc, entity, from_d, to_d):
        canvas.saveState()
        w, h = landscape(A4)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawString(doc.leftMargin, h - 0.45*inch, f"Entity: {entity}")
        canvas.drawCentredString(w/2,     h - 0.45*inch, "Day Book (Landscape)")
        canvas.drawRightString(w - doc.rightMargin, h - 0.45*inch, f"{from_d} to {to_d}")
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(w - doc.rightMargin, 0.35*inch, f"Page {doc.page}")
        canvas.restoreState()

    # ---------- building blocks ----------
    def _styles(self):
        styles = getSampleStyleSheet()
        h1 = ParagraphStyle("h1", parent=styles["Heading1"], alignment=1, spaceAfter=6)
        h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=6, spaceAfter=4)
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, leading=11)
        tiny  = ParagraphStyle("tiny",  parent=styles["Normal"], fontSize=8, leading=10)
        tiny_wrap = ParagraphStyle("tiny_wrap", parent=tiny, wordWrap='CJK',  # robust wrapping
                                   allowOrphans=1, allowWidows=1)
        return h1, h2, small, tiny_wrap

    def _p(self, s, style):
        return Paragraph("" if s is None else str(s).replace("\n", "<br/>"), style)

    def _make_summary(self, payload, styles, avail_w):
        _, h2, small, tiny = styles
        elems = []

        # Grand totals
        elems.append(Paragraph("Grand Totals", h2))
        data = [
            ["Total Debits", "Total Credits", "Net Movement",
             "Cash Opening", "Cash Receipts", "Cash Payments", "Cash Closing"],
            [ self._fmt(payload["grand_total_debits"]),
              self._fmt(payload["grand_total_credits"]),
              self._fmt(payload["grand_total_debits"] - payload["grand_total_credits"]),
              self._fmt(payload.get("grand_cash_opening", 0)),
              self._fmt(payload.get("grand_cash_total_receipts", 0)),
              self._fmt(payload.get("grand_cash_total_payments", 0)),
              self._fmt(payload.get("grand_cash_closing", 0)) ],
        ]
        # 7 columns -> distribute evenly
        col_w = [avail_w/7.0]*7
        t = LongTable(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("ALIGN",      (0,0), (-1,0), "CENTER"),
            ("ALIGN",      (0,1), (-1,-1), "RIGHT"),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
            ("VALIGN",     (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING",(0,0), (-1,-1), 3),
            ("RIGHTPADDING",(0,0),(-1,-1), 3),
        ]))
        elems += [t, Spacer(1, 6)]

        # Sections
        elems.append(Paragraph("Sections", h2))
        sec_head = ["FY", "From", "To", "Total Dr", "Total Cr", "Net",
                    "Cash Open", "Cash Rec", "Cash Pay", "Cash Close"]
        # fixed widths so it fits; a bit wider for dates
        col_w = [
            0.9*inch, 0.9*inch, 0.9*inch,
            0.9*inch, 0.9*inch, 0.8*inch,
            0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch
        ]
        data = [sec_head]
        for s in payload["sections"]:
            data.append([
                s.get("fy_name","") or "-",
                s["from_date"], s["to_date"],
                self._fmt(s["total_debits"]),
                self._fmt(s["total_credits"]),
                self._fmt(s["net_movement"]),
                self._fmt(s.get("cash_opening", 0)),
                self._fmt(s.get("cash_total_receipts", 0)),
                self._fmt(s.get("cash_total_payments", 0)),
                self._fmt(s.get("cash_closing", 0)),
            ])
        t2 = LongTable(data, colWidths=col_w, repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("ALIGN",      (0,0), (-1,0), "CENTER"),
            ("ALIGN",      (3,1), (-1,-1), "RIGHT"),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
            ("VALIGN",     (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING",(0,0), (-1,-1), 3),
            ("RIGHTPADDING",(0,0),(-1,-1), 3),
        ]))
        elems += [t2, Spacer(1, 8)]
        return elems

    def _make_detail(self, payload, styles, avail_w):
        h1, h2, small, tiny_wrap = styles
        elems = []

        # Column set chosen to fit on landscape with wrapping
        headers = ["Date", "Voucher No", "Type", "Account", "Description",
                   "Debit", "Credit", "Cash Open", "Cash Rec", "Cash Pay", "Cash Close"]
        # measured widths that fit within avail_w (~10.6-10.9in typically)
        col_w = [
            0.85*inch,  # Date
            0.95*inch,  # VNo
            1.00*inch,  # Type
            1.60*inch,  # Account
            2.80*inch,  # Description (largest; wraps)
            0.85*inch,  # Dr
            0.85*inch,  # Cr
            0.95*inch,  # Cash Open
            0.95*inch,  # Cash Rec
            0.95*inch,  # Cash Pay
            0.95*inch,  # Cash Close
        ]
        # Ensure they fit: if sum slightly over, shave a bit off description
        total_w = sum(col_w)
        if total_w > avail_w:
            delta = total_w - avail_w
            col_w[4] = max(1.4*inch, col_w[4] - delta)

        # Iterate sections and days
        for idx, s in enumerate(payload["sections"], start=1):
            sec_hdr = f"Section {idx} — FY: {s.get('fy_name','-')} | {s['from_date']} → {s['to_date']}"
            elems += [Paragraph(sec_hdr, small), Spacer(1, 3)]

            # Section cash line
            cdata = [
                ["Cash Opening", "Cash Receipts", "Cash Payments", "Cash Closing"],
                [ self._fmt(s.get("cash_opening", 0)),
                  self._fmt(s.get("cash_total_receipts", 0)),
                  self._fmt(s.get("cash_total_payments", 0)),
                  self._fmt(s.get("cash_closing", 0)) ],
            ]
            ctab = LongTable(cdata, colWidths=[1.2*inch]*4, repeatRows=1)
            ctab.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("ALIGN",      (0,0), (-1,0), "CENTER"),
                ("ALIGN",      (0,1), (-1,1), "RIGHT"),
                ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
                ("VALIGN",     (0,0), (-1,-1), "TOP"),
            ]))
            elems += [ctab, Spacer(1, 4)]

            # For each day, build a LongTable: header + 1 day total row + N item rows
            for ds in s["day_sections"]:
                rows = [headers]  # header (repeated)
                # day total row (first line)
                rows.append([
                    ds["date"], "", "", "", "Day Totals",
                    self._fmt(ds["day_debits"]), self._fmt(ds["day_credits"]),
                    self._fmt(ds.get("cash_day_opening", 0)),
                    self._fmt(ds.get("cash_day_receipts", 0)),
                    self._fmt(ds.get("cash_day_payments", 0)),
                    self._fmt(ds.get("cash_day_closing", 0)),
                ])

                # transaction lines
                for it in ds["items"]:
                    rows.append([
                        it["date"],
                        it.get("voucherno", ""),
                        it.get("voucher", ""),
                        self._p(it.get("account",""), tiny_wrap),
                        self._p(it.get("desc",""), tiny_wrap),
                        self._fmt(it["debit"]),
                        self._fmt(it["credit"]),
                        "", "", "", ""
                    ])

                lt = LongTable(rows, colWidths=col_w, repeatRows=1, splitByRow=1)
                lt.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                    ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                    ("ALIGN",      (0,0), (-1,0), "CENTER"),
                    ("GRID",       (0,0), (-1,-1), 0.3, colors.grey),
                    ("VALIGN",     (0,0), (-1,-1), "TOP"),
                    ("ALIGN",      (5,1), (6,-1), "RIGHT"),
                    ("ALIGN",      (7,1), (-1,-1), "RIGHT"),
                    ("LEFTPADDING",(0,0), (-1,-1), 3),
                    ("RIGHTPADDING",(0,0),(-1,-1), 3),
                ]))
                elems += [lt, Spacer(1, 6)]

            elems += [Spacer(1, 8)]

        return elems

    # ---------- GET ----------
    def get(self, request, *_args, **_kwargs):
        payload, code = self._payload(request)
        if code != 200:
            return HttpResponse(str(payload), status=code, content_type="application/json")

        entity = payload["entity"]
        from_d = payload["from_date"]
        to_d   = payload["to_date"]
        mode = (request.query_params.get("mode") or "both").lower()
        show_summary = mode in ("both", "summary")
        show_detail  = mode in ("both", "detail")

        buf = BytesIO()

        # Page setup (landscape A4)
        pagesize = landscape(A4)
        left, right, top, bottom = 0.5*inch, 0.5*inch, 0.7*inch, 0.5*inch
        doc = BaseDocTemplate(
            buf,
            pagesize=pagesize,
            leftMargin=left, rightMargin=right,
            topMargin=top, bottomMargin=bottom,
            title="Day Book (Landscape)"
        )
        frame = Frame(
            doc.leftMargin, doc.bottomMargin,
            doc.width, doc.height,
            id='normal'
        )
        template = PageTemplate(
            id='daybook',
            frames=[frame],
            onPage=lambda c, d: self._header_footer(c, d, entity, from_d, to_d),
        )
        doc.addPageTemplates([template])

        styles = self._styles()
        avail_w = doc.width

        elems = []
        # Title (flowable)
        h1, h2, small, _ = styles
        elems += [Paragraph("Day Book", h1),
                  Paragraph(f"Entity: <b>{entity}</b> &nbsp;&nbsp; Period: <b>{from_d}</b> to <b>{to_d}</b>", small),
                  Spacer(1, 6)]

        if show_summary:
            elems += self._make_summary(payload, styles, avail_w)
            if show_detail:
                elems.append(PageBreak())

        if show_detail:
            elems += self._make_detail(payload, styles, avail_w)

        doc.build(elems)

        buf.seek(0)
        filename = f"DayBook_{entity}_{from_d}_{to_d}.pdf"
        resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

DEC_QTY = Decimal("0.0000")
DEC_VAL = Decimal("0.00")


def q4(v): return Decimal(v).quantize(DEC_QTY, rounding=ROUND_HALF_UP)
def q2(v): return Decimal(v).quantize(DEC_VAL, rounding=ROUND_HALF_UP)


def fifo_valuation(moves, neg_policy, std_cost):
    layers = []
    qty = Decimal("0")
    val = Decimal("0")
    last_cost = Decimal("0")

    for m in moves:
        mqty = Decimal(m["qty"])
        cost = Decimal(m["unit_cost"] or 0)

        if mqty > 0:
            layers.append([mqty, cost])
            qty += mqty
            val += mqty * cost
            last_cost = cost or last_cost
        else:
            out = -mqty
            qty -= out
            while out > 0 and layers:
                lq, lc = layers[0]
                take = min(out, lq)
                val -= take * lc
                layers[0][0] -= take
                out -= take
                if layers[0][0] == 0:
                    layers.pop(0)

            if out > 0:
                if neg_policy == NegativeValuationPolicy.ERROR:
                    raise ValidationError("Negative stock encountered (FIFO).")
                cost_used = (
                    std_cost if neg_policy == NegativeValuationPolicy.STANDARD_COST
                    else last_cost if neg_policy == NegativeValuationPolicy.LAST_COST
                    else Decimal("0")
                )
                val -= out * cost_used

    return q4(qty), q2(val)


class StockSummaryAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        ser = StockSummaryRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        qs = InventoryMove.objects.filter(
            entity_id=p["entity"],
            entrydate__lte=p["as_on_date"]
        )

        # Filters
        if "location" in p:
            qs = qs.filter(location=p["location"])
        if "locations" in p:
            qs = qs.filter(location__in=p["locations"])
        if "product" in p:
            qs = qs.filter(product_id=p["product"])
        if "products" in p:
            qs = qs.filter(product_id__in=p["products"])
        if "category" in p:
            qs = qs.filter(product__category_id=p["category"])
        if "brand" in p:
            qs = qs.filter(product__brand_id=p["brand"])
        if "hsn" in p:
            qs = qs.filter(product__hsn_id=p["hsn"])
        if "include_txn_types" in p:
            qs = qs.filter(transactiontype__in=p["include_txn_types"])
        if "exclude_txn_types" in p:
            qs = qs.exclude(transactiontype__in=p["exclude_txn_types"])
        if p.get("search"):
            qs = qs.filter(product__productname__icontains=p["search"])

        qs = qs.order_by("product_id", "location", "entrydate", "id").values(
            "product_id",
            "product__productname",
            "location",
            "qty",
            "unit_cost",
            "entrydate",
        )

        grouped = defaultdict(list)
        names = {}
        last_move = {}

        for r in qs:
            key = (r["product_id"], r["location"])
            grouped[key].append(r)
            names[r["product_id"]] = r["product__productname"]
            last_move[key] = r["entrydate"]

        results = []
        for (pid, loc), moves in grouped.items():
            qty, val = fifo_valuation(
                moves,
                p["negative_valuation_policy"],
                Decimal("0")
            )

            if not p["include_zero"] and qty == DEC_QTY:
                continue
            if p["negative_only"] and qty >= DEC_QTY:
                continue

            row = {
                "product_id": pid,
                "product_name": names.get(pid),
                "location": loc,
                "stock_qty": qty,
                "stock_value": val,
            }

            if p["include_avg_cost"] and qty != DEC_QTY:
                row["avg_cost"] = q4(val / qty)

            if p["include_last_movement"]:
                row["last_movement_date"] = last_move.get((pid, loc))

            results.append(row)

        # Ordering
        if p["ordering"] == "-value":
            results.sort(key=lambda x: x["stock_value"], reverse=True)
        elif p["ordering"] == "value":
            results.sort(key=lambda x: x["stock_value"])
        elif p["ordering"] == "-qty":
            results.sort(key=lambda x: x["stock_qty"], reverse=True)
        elif p["ordering"] == "qty":
            results.sort(key=lambda x: x["stock_qty"])
        elif p["ordering"] == "-product":
            results.sort(key=lambda x: x["product_name"], reverse=True)
        else:
            results.sort(key=lambda x: x["product_name"])

        # Pagination
        page = p["page"]
        size = p["page_size"]
        start = (page - 1) * size
        end = start + size

        response = {
            "entity": p["entity"],
            "as_on_date": p["as_on_date"],
            "valuation_method": p["valuation_method"],
            "count": len(results[start:end]),
            "results": results[start:end],
        }

        if p["include_totals"]:
            response["totals"] = {
                "total_qty": q4(sum(r["stock_qty"] for r in results)),
                "total_value": q2(sum(r["stock_value"] for r in results)),
                "total_rows": len(results),
            }

        return Response(response)
