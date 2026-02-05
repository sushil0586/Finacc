from io import BytesIO

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from reports.serializers_stock_aging import StockAgingRequestSerializer
from reports.services.stock_aging_service import compute_stock_aging


class StockAgingExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockAgingRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        # export all (ignore pagination)
        p["page"] = 1
        p["page_size"] = 10_000_000

        rows, bucket_labels, totals = compute_stock_aging(p=p)

        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Aging"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append([f"Stock Aging Report (Entity: {p['entity']})"])
        ws.append([f"As on: {p['as_on_date']} | Group By Location: {p.get('group_by_location', True)}"])
        ws.append([])

        headers = ["Product", "Location", "Closing Qty"] + bucket_labels
        ws.append(headers)
        hr = ws.max_row

        for cidx in range(1, len(headers) + 1):
            c = ws.cell(row=hr, column=cidx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        for r in rows:
            ws.append(
                [(r.get("product_name") or str(r["product_id"])),
                 "" if r.get("location") is None else str(r.get("location")),
                 str(r["closing_qty"])]
                + [str(r["buckets"].get(lbl, 0)) for lbl in bucket_labels]
            )

        # Totals row
        ws.append([])
        ws.append(["TOTALS", "", str(totals["closing_qty"])] + [str(totals["buckets"][lbl]) for lbl in bucket_labels])

        widths = [40, 10, 14] + [12 for _ in bucket_labels]
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i - 1 < len(widths) else 12

        for row in ws.iter_rows(min_row=hr, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                if cell.row == hr:
                    continue
                if cell.column == 1:
                    cell.alignment = left
                elif cell.column == 2:
                    cell.alignment = center
                else:
                    cell.alignment = right

        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)

        filename = f"StockAging_Entity{p['entity']}_AsOn_{p['as_on_date']}.xlsx"
        resp = HttpResponse(
            buff.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class StockAgingPDFAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = StockAgingRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        p = ser.validated_data

        # PDF summary only
        p["page"] = 1
        p["page_size"] = 10_000_000

        rows, bucket_labels, totals = compute_stock_aging(p=p)

        buff = BytesIO()
        filename = f"StockAging_Entity{p['entity']}_AsOn_{p['as_on_date']}.pdf"

        doc = SimpleDocTemplate(
            buff,
            pagesize=landscape(A4),
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18,
            title=filename,
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph("<b>Stock Aging Report</b>", styles["Title"]),
            Spacer(1, 6),
            Paragraph(
                f"Entity: <b>{p['entity']}</b> | As on: <b>{p['as_on_date']}</b> | "
                f"Group By Location: <b>{p.get('group_by_location', True)}</b>",
                styles["Normal"]
            ),
            Spacer(1, 10),
        ]

        headers = ["Product", "Loc", "Closing"] + bucket_labels
        table_data = [headers]

        # Keep PDF readable: product name clipped
        for r in rows:
            table_data.append(
                [(r.get("product_name") or str(r["product_id"]))[:35],
                 "" if r.get("location") is None else str(r.get("location")),
                 str(r["closing_qty"])]
                + [str(r["buckets"].get(lbl, 0)) for lbl in bucket_labels]
            )

        table_data.append(["TOTALS", "", str(totals["closing_qty"])] + [str(totals["buckets"][lbl]) for lbl in bucket_labels])

        col_widths = [220, 40, 70] + [65 for _ in bucket_labels]
        tbl = Table(table_data, repeatRows=1, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        story.append(tbl)
        doc.build(story)

        pdf = buff.getvalue()
        buff.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
