from __future__ import annotations

import csv
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone
from entity.models import Entity, EntityFinancialYear, SubEntity
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from reports.gstr9.services.freeze import Gstr9FreezeService
from reports.gstr9.services.report import Gstr9ReportService
from reports.gstr9.views.utils import Gstr9ScopedReportMixin, parse_freeze_version


class Gstr9ExportAPIView(Gstr9ScopedReportMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]
    service_class = Gstr9ReportService
    freeze_service_class = Gstr9FreezeService

    def initialize_request(self, request, *args, **kwargs):
        request = super().initialize_request(request, *args, **kwargs)
        if "format" in request.query_params:
            request._gstr9_export_format = request.query_params.get("format")
            mutable = request._request.GET.copy()
            mutable.pop("format", None)
            request._request.GET = mutable
        return request

    def get(self, request):
        export_format = (getattr(request, "_gstr9_export_format", None) or request.query_params.get("format") or "json").lower()
        service = self.service_class()
        scope = service.build_scope(request.query_params)
        self.enforce_report_scope(request, scope)
        freeze_service = self.freeze_service_class(report_service=service)
        try:
            freeze_version = parse_freeze_version(request.query_params)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=400)
        if "freeze_version" in request.query_params:
            frozen = freeze_service.get_snapshot(scope, version=freeze_version)
            if not frozen:
                requested = request.query_params.get("freeze_version") or "latest"
                return Response({"detail": f"Frozen snapshot not found for freeze_version={requested}."}, status=404)
            payload = {
                "summary": frozen.get("payload", {}).get("summary") or service.summary(scope),
                "validations": frozen.get("payload", {}).get("validations") or [],
                "freeze": {
                    "version": frozen["version"],
                    "frozen_at": frozen["frozen_at"],
                    "frozen_by": frozen["frozen_by"],
                },
            }
        else:
            payload = service.export_payload(scope)

        if export_format == "json":
            return Response(payload)
        if export_format == "csv":
            content = self._export_csv(payload, scope=scope, freeze_version=freeze_version)
            return _file_response("GSTR9_Summary.csv", content, "text/csv")
        if export_format == "xlsx":
            content = self._export_xlsx(payload, scope=scope, freeze_version=freeze_version)
            return _file_response(
                "GSTR9_Summary.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return Response({"detail": "Unsupported export format."}, status=400)

    def _export_csv(self, payload: dict, *, scope, freeze_version) -> bytes:
        stream = StringIO()
        writer = csv.writer(stream)
        audit = _audit_context(scope, freeze_version=freeze_version)
        summary = payload.get("summary") or {}
        validations = payload.get("validations") or []

        writer.writerow(["Report", audit["report_title"]])
        writer.writerow(["Financial Year", audit["financial_year"]])
        writer.writerow(["Subentity", audit["subentity"]])
        writer.writerow(["Freeze Scope", audit["freeze_label"]])
        writer.writerow(["Generated On", audit["generated_on"]])
        writer.writerow([])

        writer.writerow(["Summary"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Phase", summary.get("phase")])
        writer.writerow(["Status", summary.get("status")])
        writer.writerow(["Message", summary.get("message")])
        writer.writerow([])

        writer.writerow(["Table Coverage"])
        writer.writerow(["Table Code", "Label", "Status"])
        for row in summary.get("tables") or []:
            writer.writerow([row.get("code"), row.get("label"), row.get("status")])

        writer.writerow([])
        writer.writerow(["Validation Warnings"])
        writer.writerow(["Severity", "Code", "Message"])
        for warning in validations:
            writer.writerow([warning.get("severity"), warning.get("code"), warning.get("message")])
        return stream.getvalue().encode("utf-8")

    def _export_xlsx(self, payload: dict, *, scope, freeze_version) -> bytes:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "GSTR9 Summary"
        summary = payload.get("summary") or {}
        validations = payload.get("validations") or []
        styles = _workbook_styles()
        audit = _audit_context(scope, freeze_version=freeze_version)

        _write_title_block(
            ws,
            title=audit["report_title"],
            subtitle=f"{audit['financial_year']} | {audit['subentity']} | {audit['freeze_label']}",
            generated_on=audit["generated_on"],
            width=3,
            styles=styles,
        )
        ws.append(["Metric", "Value"])
        _style_header_row(ws, ws.max_row, 2, styles)
        ws.append(["Phase", summary.get("phase")])
        ws.append(["Status", summary.get("status")])
        ws.append(["Message", summary.get("message")])
        _style_key_value_rows(ws, ws.max_row - 2, ws.max_row, styles)

        ws.append([])
        ws.append(["Table Code", "Table Label", "Status"])
        _style_header_row(ws, ws.max_row, 3, styles)
        table_start = ws.max_row + 1
        for row in summary.get("tables") or []:
            ws.append([row.get("code"), row.get("label"), row.get("status")])
        _style_body_rows(ws, table_start, ws.max_row, 3, numeric_columns=set(), styles=styles)

        if validations:
            warnings_ws = workbook.create_sheet("Warnings")
            _write_title_block(
                warnings_ws,
                title="GSTR-9 Validation Warnings",
                subtitle=f"{audit['financial_year']} | {audit['subentity']}",
                generated_on=audit["generated_on"],
                width=3,
                styles=styles,
            )
            warnings_ws.append(["Severity", "Code", "Message"])
            _style_header_row(warnings_ws, warnings_ws.max_row, 3, styles)
            warning_start = warnings_ws.max_row + 1
            for warning in validations:
                warnings_ws.append([warning.get("severity"), warning.get("code"), warning.get("message")])
            _style_body_rows(warnings_ws, warning_start, warnings_ws.max_row, 3, numeric_columns=set(), styles=styles)
            _apply_column_widths(warnings_ws, {1: 16, 2: 34, 3: 80})

        _apply_column_widths(ws, {1: 18, 2: 52, 3: 18})
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def _file_response(filename: str, content: bytes, content_type: str):
    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _audit_context(scope, *, freeze_version):
    entity_name = (
        Entity.objects.filter(id=scope.entity_id).values_list("entityname", flat=True).first()
        or f"Entity {scope.entity_id}"
    )
    financial_year = "All financial years"
    if scope.entityfinid_id:
        financial_year = (
            EntityFinancialYear.objects.filter(id=scope.entityfinid_id, entity_id=scope.entity_id)
            .values_list("desc", flat=True)
            .first()
            or f"FY {scope.entityfinid_id}"
        )
    subentity = "All subentities"
    if scope.subentity_id:
        subentity = (
            SubEntity.objects.filter(id=scope.subentity_id, entity_id=scope.entity_id)
            .values_list("subentityname", flat=True)
            .first()
            or f"Subentity {scope.subentity_id}"
        )
    freeze_label = f"Freeze v{freeze_version}" if freeze_version else "Live Data"
    return {
        "report_title": f"{entity_name} - GSTR-9 Annual Return",
        "financial_year": financial_year,
        "subentity": subentity,
        "freeze_label": freeze_label,
        "generated_on": timezone.localtime().strftime("%d %b %Y %I:%M %p"),
    }


def _workbook_styles():
    header_fill = PatternFill("solid", fgColor="2D4F83")
    section_fill = PatternFill("solid", fgColor="EEF4FB")
    border_color = "D6DEEA"
    thin = Side(style="thin", color=border_color)
    return {
        "title_font": Font(name="Calibri", bold=True, size=16, color="20324D"),
        "subtitle_font": Font(name="Calibri", size=10, color="5E6B80"),
        "header_font": Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
        "body_font": Font(name="Calibri", size=10, color="233248"),
        "section_fill": section_fill,
        "header_fill": header_fill,
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "left": Alignment(horizontal="left", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "wrap_left": Alignment(horizontal="left", vertical="top", wrap_text=True),
    }


def _write_title_block(ws, *, title, subtitle, generated_on, width, styles):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws["A1"] = title
    ws["A1"].font = styles["title_font"]
    ws["A1"].alignment = styles["left"]

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws["A2"] = subtitle
    ws["A2"].font = styles["subtitle_font"]
    ws["A2"].alignment = styles["left"]

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=width)
    ws["A3"] = f"Generated on {generated_on}"
    ws["A3"].font = styles["subtitle_font"]
    ws["A3"].alignment = styles["left"]
    ws.append([])


def _style_header_row(ws, row_number, width, styles):
    for col in range(1, width + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["left"]


def _style_key_value_rows(ws, start_row, end_row, styles):
    for row in range(start_row, end_row + 1):
        left = ws.cell(row=row, column=1)
        right = ws.cell(row=row, column=2)
        left.font = styles["body_font"]
        right.font = styles["body_font"]
        left.fill = styles["section_fill"]
        left.border = styles["border"]
        right.border = styles["border"]
        left.alignment = styles["left"]
        right.alignment = styles["wrap_left"]


def _style_body_rows(ws, start_row, end_row, width, *, numeric_columns, styles):
    for row in range(start_row, end_row + 1):
        for col in range(1, width + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = styles["body_font"]
            cell.border = styles["border"]
            cell.alignment = styles["right"] if col in numeric_columns else styles["wrap_left"]


def _apply_column_widths(ws, widths):
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width
