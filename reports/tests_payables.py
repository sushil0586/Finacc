from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient, APITestCase

from Authentication.models import User
from entity.models import Entity, EntityFinancialYear, GstRegistrationType, SubEntity
from financial.models import Ledger, account, accountHead, accounttype
from financial.profile_access import account_gstno
from financial.services import apply_normalized_profile_payload, create_account_with_synced_ledger
from geography.models import City, Country, District, State
from purchase.models.purchase_ap import VendorAdvanceBalance, VendorBillOpenItem, VendorSettlement, VendorSettlementLine
from posting.models import Entry, EntryStatus, JournalLine, PostingBatch, TxnType
from purchase.models.purchase_core import PurchaseInvoiceHeader, PurchaseInvoiceLine
from rbac.models import Permission, Role, RolePermission, UserRoleAssignment
from rbac.services import EffectiveMenuService


@override_settings(ROOT_URLCONF="FA.urls", AUTH_PASSWORD_VALIDATORS=[])
class PayableReportAPITests(APITestCase):
    def setUp(self):
        self.client = APIClient()
        suffix = uuid4().hex[:8]
        self.user = User.objects.create_user(username=f"payable-report-user-{suffix}", email=f"payable-{suffix}@example.com", password="pass123")
        self.client.force_authenticate(user=self.user)
        self.country = Country.objects.create(countryname="India", countrycode="IN")
        self.state = State.objects.create(statename="Punjab", statecode="PB", country=self.country)
        self.district = District.objects.create(districtname="District", districtcode="DT", state=self.state)
        self.city = City.objects.create(cityname="Ludhiana", citycode="LDH", pincode="141001", distt=self.district)
        self.gst_type = GstRegistrationType.objects.create(Name="Regular", Description="Regular")
        self.entity = Entity.objects.create(
            entityname="Payable Entity",
            legalname="Payable Entity Pvt Ltd",
            GstRegitrationType=self.gst_type,
            createdby=self.user,
        )
        self.subentity = SubEntity.objects.create(entity=self.entity, subentityname="Branch A")
        self.entityfin = EntityFinancialYear.objects.create(
            entity=self.entity,
            desc="FY 2025-26",
            finstartyear=timezone.make_aware(datetime(2025, 4, 1)),
            finendyear=timezone.make_aware(datetime(2026, 3, 31)),
            createdby=self.user,
        )
        self.report_role = Role.objects.create(
            entity=self.entity,
            name="Report Viewer",
            code="report_viewer_test",
            role_level=Role.LEVEL_ENTITY,
            is_system_role=False,
            is_assignable=True,
            priority=10,
            createdby=self.user,
        )
        report_permission_codes = [
            "reports.payables.view",
            "reports.vendoroutstanding.view",
            "reports.accountspayableaging.view",
            "reports.purchase_register.view",
            "reports.vendorledgerstatement.view",
            "reports.vendorsettlementhistory.view",
            "reports.vendornoteregister.view",
            "reports.payables.upcoming_payments_calendar.view",
            "reports.apglreconciliation.view",
            "reports.payablesclosepack.view",
            "reports.vendorbalanceexceptions.view",
            "reports.payables.settings.view",
        ]
        report_permission_ids = []
        for code in report_permission_codes:
            permission, _ = Permission.objects.get_or_create(
                code=code,
                defaults={
                    "name": code.replace(".", " ").replace("_", " ").title(),
                    "module": "reports",
                    "resource": "payables",
                    "action": "view",
                },
            )
            if not permission.isactive:
                permission.isactive = True
                permission.save(update_fields=["isactive"])
            report_permission_ids.append(permission.id)
        for permission_id in report_permission_ids:
            RolePermission.objects.get_or_create(
                role=self.report_role,
                permission_id=permission_id,
                defaults={
                    "effect": RolePermission.EFFECT_ALLOW,
                },
            )
        UserRoleAssignment.objects.create(
            user=self.user,
            entity=self.entity,
            role=self.report_role,
            assigned_by=self.user,
            is_primary=True,
        )
        self.acc_type = accounttype.objects.create(entity=self.entity, accounttypename="Liabilities", accounttypecode="L100", createdby=self.user)
        self.vendor_head = accountHead.objects.create(
            entity=self.entity,
            name="Sundry Creditors",
            code=400,
            balanceType="Credit",
            drcreffect="Credit",
            accounttype=self.acc_type,
            createdby=self.user,
        )
        self.expense_head = accountHead.objects.create(
            entity=self.entity,
            name="Purchase Expense",
            code=401,
            balanceType="Debit",
            drcreffect="Debit",
            accounttype=self.acc_type,
            createdby=self.user,
        )
        self.vendor_ledger = Ledger.objects.create(
            entity=self.entity,
            ledger_code=5001,
            name="ABC Traders",
            accounthead=self.vendor_head,
            createdby=self.user,
        )
        self.vendor = create_account_with_synced_ledger(
            account_data={
                "entity": self.entity,
                "ledger": self.vendor_ledger,
                "accountname": "ABC Traders",
                "createdby": self.user,
            },
            ledger_overrides={"ledger_code": 5001, "accounthead": self.vendor_head, "is_party": True},
        )
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={"gstno": "03ABCDE1234F1Z5"},
            commercial_data={
                "partytype": "Vendor",
                "currency": "INR",
                "agent": "Wholesale",
                "creditdays": 30,
                "creditlimit": Decimal("1000.00"),
            },
            primary_address_data={"state": self.state, "city": self.city},
        )
        self.other_vendor = self._create_vendor("Idle Vendor", 5002)
        self.invoice = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 1),
            due_date=date(2025, 4, 10),
            doc_code="PINV",
            doc_no=1001,
            purchase_number="PI-PINV-1001",
            supplier_invoice_number="SUP-001",
            amount=Decimal("1000.00"),
        )
        self.credit_note = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            bill_date=date(2025, 4, 15),
            due_date=date(2025, 4, 15),
            doc_code="PCN",
            doc_no=1002,
            purchase_number="PI-PCN-1002",
            supplier_invoice_number="SUP-CN-001",
            amount=Decimal("-100.00"),
            ref_document=self.invoice,
        )
        self.invoice_item = self._create_open_item(
            header=self.invoice,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 1),
            due_date=date(2025, 4, 10),
            purchase_number="PI-PINV-1001",
            supplier_invoice_number="SUP-001",
            amount=Decimal("1000.00"),
        )
        self.credit_item = self._create_open_item(
            header=self.credit_note,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            bill_date=date(2025, 4, 15),
            due_date=date(2025, 4, 15),
            purchase_number="PI-PCN-1002",
            supplier_invoice_number="SUP-CN-001",
            amount=Decimal("-100.00"),
        )
        self.advance = self._create_advance(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            credit_date=date(2025, 4, 20),
            reference_no="ADV-001",
            amount=Decimal("50.00"),
        )
        self.payment = VendorSettlement.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            settlement_type=VendorSettlement.SettlementType.PAYMENT,
            settlement_date=date(2025, 4, 5),
            reference_no="PAY-001",
            total_amount=Decimal("200.00"),
            status=VendorSettlement.Status.POSTED,
            posted_at=timezone.now(),
            posted_by=self.user,
        )
        VendorSettlementLine.objects.create(
            settlement=self.payment,
            open_item=self.invoice_item,
            amount=Decimal("200.00"),
            applied_amount_signed=Decimal("200.00"),
        )

    def _create_vendor(self, name, ledger_code, *, creditlimit=Decimal("1000.00"), creditdays=30):
        ledger = Ledger.objects.create(
            entity=self.entity,
            ledger_code=ledger_code,
            name=name,
            accounthead=self.vendor_head,
            createdby=self.user,
        )
        vendor = create_account_with_synced_ledger(
            account_data={
                "entity": self.entity,
                "ledger": ledger,
                "accountname": name,
                "createdby": self.user,
            },
            ledger_overrides={"ledger_code": ledger_code, "accounthead": self.vendor_head, "is_party": True},
        )
        apply_normalized_profile_payload(
            vendor,
            compliance_data={},
            commercial_data={
                "partytype": "Vendor",
                "currency": "INR",
                "creditdays": creditdays,
                "creditlimit": creditlimit,
            },
            primary_address_data={"state": self.state, "city": self.city},
        )
        return vendor

    def _create_purchase_header(
        self,
        *,
        vendor,
        vendor_ledger,
        doc_type,
        bill_date,
        due_date,
        doc_code,
        doc_no,
        purchase_number,
        supplier_invoice_number,
        amount,
        ref_document=None,
        status=PurchaseInvoiceHeader.Status.POSTED,
    ):
        return PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            doc_type=doc_type,
            ref_document=ref_document,
            bill_date=bill_date,
            due_date=due_date,
            doc_code=doc_code,
            doc_no=doc_no,
            purchase_number=purchase_number,
            supplier_invoice_number=supplier_invoice_number,
            vendor=vendor,
            vendor_ledger=vendor_ledger,
            vendor_name=vendor.accountname,
            vendor_gstin=account_gstno(vendor) or "",
            status=status,
            grand_total=amount,
            created_by=self.user,
        )

    def _create_open_item(
        self,
        *,
        header,
        vendor,
        vendor_ledger,
        doc_type,
        bill_date,
        due_date,
        purchase_number,
        supplier_invoice_number,
        amount,
    ):
        return VendorBillOpenItem.objects.create(
            header=header,
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=vendor,
            vendor_ledger=vendor_ledger,
            doc_type=doc_type,
            bill_date=bill_date,
            due_date=due_date,
            purchase_number=purchase_number,
            supplier_invoice_number=supplier_invoice_number,
            original_amount=amount,
            gross_amount=amount,
            net_payable_amount=amount,
            settled_amount=Decimal("0.00"),
            outstanding_amount=amount,
            is_open=True,
        )

    def _create_advance(self, *, vendor, vendor_ledger, credit_date, reference_no, amount):
        return VendorAdvanceBalance.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=vendor,
            vendor_ledger=vendor_ledger,
            source_type=VendorAdvanceBalance.SourceType.PAYMENT_ADVANCE,
            credit_date=credit_date,
            reference_no=reference_no,
            original_amount=amount,
            adjusted_amount=Decimal("0.00"),
            outstanding_amount=amount,
            is_open=True,
        )

    def _base_scope(self, **extra):
        params = {
            "entity": self.entity.id,
            "entityfinid": self.entityfin.id,
            "subentity": self.subentity.id,
        }
        params.update(extra)
        return params

    def _create_limited_report_user(self, permission_code):
        suffix = uuid4().hex[:8]
        user = User.objects.create_user(username=f"payable-limited-{suffix}", email=f"payable-limited-{suffix}@example.com", password="pass123")
        role = Role.objects.create(
            entity=self.entity,
            name=f"Limited Report Viewer {suffix}",
            code=f"limited_report_viewer_{suffix}",
            role_level=Role.LEVEL_ENTITY,
            is_system_role=False,
            is_assignable=True,
            priority=20,
            createdby=self.user,
        )
        base_permission = Permission.objects.get(code="reports.payables.view")
        if not base_permission.isactive:
            base_permission.isactive = True
            base_permission.save(update_fields=["isactive"])
        permission = Permission.objects.get(code=permission_code)
        if not permission.isactive:
            permission.isactive = True
            permission.save(update_fields=["isactive"])
        for permission_obj in (base_permission, permission):
            RolePermission.objects.get_or_create(
                role=role,
                permission=permission_obj,
                defaults={
                    "effect": RolePermission.EFFECT_ALLOW,
                },
            )
        UserRoleAssignment.objects.create(
            user=user,
            entity=self.entity,
            role=role,
            assigned_by=self.user,
            is_primary=True,
        )
        return user

    def test_vendor_outstanding_report_builds_vendor_totals_and_drilldowns(self):
        clean_vendor = self._create_vendor("Clean Vendor", 5003)
        clean_invoice = self._create_purchase_header(
            vendor=clean_vendor,
            vendor_ledger=clean_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 1),
            due_date=date(2025, 4, 10),
            doc_code="PINV",
            doc_no=1005,
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("1000.00"),
        )
        self._create_open_item(
            header=clean_invoice,
            vendor=clean_vendor,
            vendor_ledger=clean_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 1),
            due_date=date(2025, 4, 10),
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("1000.00"),
        )
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", vendor=clean_vendor.id),
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["report_code"], "vendor_outstanding")
        self.assertEqual(data["summary"]["vendor_count"], 1)
        self.assertEqual(data["totals"]["bill_amount"], "1000.00")
        self.assertEqual(data["totals"]["payment_amount"], "0.00")
        self.assertEqual(data["totals"]["credit_balance"], "0.00")
        self.assertEqual(data["totals"]["advance_balance"], "0.00")
        self.assertEqual(data["totals"]["outstanding"], "1000.00")
        self.assertEqual(data["pagination"]["paginated"], True)
        self.assertEqual(len(data["rows"]), 1)

        row = data["rows"][0]
        self.assertEqual(row["vendor_name"], "Clean Vendor")
        self.assertEqual(row["opening_balance"], "0.00")
        self.assertEqual(row["bill_amount"], "1000.00")
        self.assertEqual(row["payment_amount"], "0.00")
        self.assertEqual(row["credit_balance"], "0.00")
        self.assertEqual(row["advance_balance"], "0.00")
        self.assertEqual(row["outstanding"], "1000.00")
        self.assertEqual(row["overdue_amount"], "1000.00")
        self.assertEqual(
            row["drilldown_targets"],
            ["vendor_detail", "aging_summary", "aging_bill_list", "vendor_statement", "open_items", "payments"],
        )
        self.assertEqual(row["_meta"]["drilldown"]["aging_summary"]["target"], "ap_aging")

    def test_vendor_outstanding_reconciliation_warning_flags_difference(self):
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", reconcile_gl="true"),
        )
        self.assertEqual(response.status_code, 200)
        meta = response.json()["_meta"]
        self.assertTrue(meta["gl_reconciliation_warning"])
        self.assertEqual(meta["difference_amount"], "750.00")

    def test_settlement_application_respects_as_of_date(self):
        response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-04", view="invoice"),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertEqual(row["paid_amount"], "0.00")
        self.assertEqual(row["credit_applied_fifo"], "0.00")
        self.assertEqual(row["balance"], "1000.00")

    def test_ap_aging_report_supports_summary_and_invoice_views(self):
        summary_response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="summary", page_size=1),
        )
        self.assertEqual(summary_response.status_code, 200)
        summary_data = summary_response.json()
        self.assertEqual(summary_data["report_code"], "ap_aging")
        self.assertEqual(summary_data["summary"]["vendor_count"], 1)
        self.assertEqual(summary_data["totals"]["outstanding"], "650.00")
        self.assertEqual(summary_data["totals"]["bucket_1_30"], "650.00")
        self.assertEqual(summary_data["totals"]["unapplied_advance"], "0.00")
        self.assertFalse(summary_data["pagination"]["paginated"])
        summary_row = summary_data["rows"][0]
        self.assertEqual(summary_row["outstanding"], "650.00")
        self.assertEqual(summary_row["bucket_1_30"], "650.00")
        self.assertEqual(summary_row["bucket_90_plus"], "0.00")
        self.assertEqual(summary_row["unapplied_advance"], "0.00")

        invoice_response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="invoice"),
        )
        self.assertEqual(invoice_response.status_code, 200)
        invoice_data = invoice_response.json()
        self.assertEqual(invoice_data["view"], "invoice")
        self.assertTrue(invoice_data["pagination"]["paginated"])
        self.assertEqual(invoice_data["totals"]["balance"], "650.00")
        self.assertEqual(len(invoice_data["rows"]), 1)
        row = invoice_data["rows"][0]
        self.assertEqual(row["bill_number"], "PI-PINV-1001")
        self.assertEqual(row["bill_amount"], "1000.00")
        self.assertEqual(row["paid_amount"], "200.00")
        self.assertEqual(row["credit_applied_fifo"], "150.00")
        self.assertEqual(row["balance"], "650.00")
        self.assertEqual(row["bucket_1_30"], "650.00")
        self.assertEqual(row["_meta"]["drilldown"]["bill"]["target"], "purchase_document_detail")

    def test_vendor_outstanding_detailed_rows_expose_default_msme_due_date(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "gstno": "03ABCDE1234F1Z5",
                "msme": "legacy-msme",
                "msme_status": "micro",
                "udyam_no": "UDYAM-PB-1001",
                "has_written_payment_terms": False,
                "msme_credit_days": None,
            },
            commercial_data={
                "partytype": "Vendor",
                "currency": "INR",
                "agent": "Wholesale",
                "creditdays": 30,
                "creditlimit": Decimal("1000.00"),
            },
        )

        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(as_of_date="2025-04-30", view="detailed", vendor=self.vendor.id),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertTrue(row["is_msme_applicable"])
        self.assertEqual(row["msme_status"], "micro")
        self.assertEqual(row["udyam_no"], "UDYAM-PB-1001")
        self.assertFalse(row["has_written_payment_terms"])
        self.assertEqual(row["msme_allowed_credit_days"], 15)
        self.assertEqual(row["msme_due_date"], "2025-04-16")
        self.assertEqual(row["msme_days_overdue"], 14)
        self.assertTrue(row["is_msme_overdue"])

    def test_ap_aging_invoice_rows_cap_msme_written_terms_at_45_days(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "gstno": "03ABCDE1234F1Z5",
                "msme_status": "small",
                "udyam_no": "UDYAM-PB-2002",
                "has_written_payment_terms": True,
                "msme_credit_days": None,
            },
            commercial_data={
                "partytype": "Vendor",
                "currency": "INR",
                "agent": "Wholesale",
                "creditdays": 60,
                "creditlimit": Decimal("1000.00"),
            },
        )

        response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="invoice", vendor=self.vendor.id),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertTrue(row["is_msme_applicable"])
        self.assertEqual(row["msme_status"], "small")
        self.assertTrue(row["has_written_payment_terms"])
        self.assertEqual(row["msme_allowed_credit_days"], 45)
        self.assertEqual(row["msme_due_date"], "2025-05-16")
        self.assertEqual(row["msme_days_overdue"], 0)
        self.assertFalse(row["is_msme_overdue"])

    def test_msme_overdue_report_returns_only_msme_overdue_rows_with_drilldowns(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "msme_status": "micro",
                "udyam_no": "UDYAM-PB-0001",
                "has_written_payment_terms": False,
                "msme_credit_days": None,
            },
        )
        other_vendor_ledger = Ledger.objects.create(
            entity=self.entity,
            ledger_code=5003,
            name="Non MSME Vendor",
            accounthead=self.vendor_head,
            createdby=self.user,
        )
        other_vendor = create_account_with_synced_ledger(
            account_data={
                "entity": self.entity,
                "ledger": other_vendor_ledger,
                "accountname": "Non MSME Vendor",
                "createdby": self.user,
            },
            ledger_overrides={"ledger_code": 5003, "accounthead": self.vendor_head, "is_party": True},
        )
        apply_normalized_profile_payload(
            other_vendor,
            compliance_data={},
            commercial_data={"partytype": "Vendor", "currency": "INR", "creditdays": 30},
            primary_address_data={"state": self.state, "city": self.city},
        )
        non_msme_header = self._create_purchase_header(
            vendor=other_vendor,
            vendor_ledger=other_vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 2),
            due_date=date(2025, 4, 12),
            doc_code="PINV",
            doc_no=1003,
            purchase_number="PI-PINV-1003",
            supplier_invoice_number="SUP-003",
            amount=Decimal("300.00"),
        )
        self._create_open_item(
            header=non_msme_header,
            vendor=other_vendor,
            vendor_ledger=other_vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 2),
            due_date=date(2025, 4, 12),
            purchase_number="PI-PINV-1003",
            supplier_invoice_number="SUP-003",
            amount=Decimal("300.00"),
        )

        response = self.client.get(
            reverse("reports_api:msme-overdue-report"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "msme_overdue")
        self.assertEqual(payload["summary"]["bill_count"], 1)
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        self.assertEqual(payload["summary"]["overdue_amount"], "800.00")
        self.assertEqual(payload["summary"]["overdue_bill_count"], 1)
        self.assertEqual(payload["summary"]["overdue_vendor_count"], 1)
        self.assertEqual(payload["summary"]["oldest_overdue_days"], 14)
        self.assertIn("800.00", payload["summary"]["reporting_note"])
        row = payload["rows"][0]
        self.assertEqual(row["vendor_name"], "ABC Traders")
        self.assertEqual(row["msme_status"], "micro")
        self.assertEqual(row["udyam_no"], "UDYAM-PB-0001")
        self.assertEqual(row["msme_due_date"], "2025-04-16")
        self.assertEqual(row["msme_days_overdue"], 14)
        self.assertEqual(row["overdue_bucket"], "1-15")
        self.assertEqual(row["balance"], "800.00")
        self.assertEqual(row["_meta"]["drilldown"]["bill_detail"]["route"], "/purchaseinvoice")
        self.assertEqual(row["_meta"]["drilldown"]["vendor_outstanding"]["report_code"], "vendor_outstanding")

    def test_msme_overdue_export_endpoint_returns_csv(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "msme_status": "small",
                "has_written_payment_terms": False,
            },
        )
        response = self.client.get(
            reverse("reports_api:msme-overdue-report-csv"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("text/csv"))
        header_line = response.content.decode("utf-8-sig").splitlines()[0]
        self.assertIn("Vendor", header_line)
        self.assertIn("MSME Due Date", header_line)

    def test_fifo_credit_application_uses_oldest_vendor_invoice_first(self):
        second_invoice = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 28),
            doc_code="PINV",
            doc_no=1003,
            purchase_number="PI-PINV-1003",
            supplier_invoice_number="SUP-003",
            amount=Decimal("400.00"),
        )
        self._create_open_item(
            header=second_invoice,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 28),
            purchase_number="PI-PINV-1003",
            supplier_invoice_number="SUP-003",
            amount=Decimal("400.00"),
        )
        extra_advance = self._create_advance(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            credit_date=date(2025, 4, 25),
            reference_no="ADV-002",
            amount=Decimal("800.00"),
        )
        self.assertIsNotNone(extra_advance.id)

        response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="invoice"),
        )
        self.assertEqual(response.status_code, 200)
        rows = response.json()["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["bill_number"], "PI-PINV-1003")
        self.assertEqual(rows[0]["credit_applied_fifo"], "150.00")
        self.assertEqual(rows[0]["balance"], "250.00")

    def test_negative_vendor_balance_is_retained_in_vendor_outstanding(self):
        credit_vendor = self._create_vendor("Credit Vendor", 5003)
        self._create_advance(
            vendor=credit_vendor,
            vendor_ledger=credit_vendor.ledger,
            credit_date=date(2025, 4, 18),
            reference_no="ADV-CREDIT",
            amount=Decimal("75.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(
                from_date="2025-04-01",
                to_date="2025-04-30",
                include_credit_balances="true",
            ),
        )
        self.assertEqual(response.status_code, 200)
        rows = {row["vendor_name"]: row for row in response.json()["rows"]}
        self.assertEqual(rows["Credit Vendor"]["outstanding"], "-75.00")
        self.assertEqual(rows["Credit Vendor"]["advance_balance"], "75.00")

    def test_aging_bucket_placement_and_summary_not_paginated(self):
        current_invoice = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 6, 1),
            due_date=date(2025, 6, 20),
            doc_code="PINV",
            doc_no=1004,
            purchase_number="PI-PINV-1004",
            supplier_invoice_number="SUP-004",
            amount=Decimal("300.00"),
        )
        self._create_open_item(
            header=current_invoice,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 6, 1),
            due_date=date(2025, 6, 20),
            purchase_number="PI-PINV-1004",
            supplier_invoice_number="SUP-004",
            amount=Decimal("300.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-06-15", view="summary", page_size=1),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload["pagination"]["paginated"])
        row = payload["rows"][0]
        self.assertEqual(row["current"], "300.00")
        self.assertEqual(row["bucket_61_90"], "650.00")

    def test_vendor_outstanding_aging_basis_bill_date_can_mark_future_due_bill_as_overdue(self):
        future_due_vendor = self._create_vendor("Future Due Vendor", 5006)
        future_due_header = self._create_purchase_header(
            vendor=future_due_vendor,
            vendor_ledger=future_due_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 1),
            due_date=date(2025, 6, 30),
            doc_code="PINV",
            doc_no=1201,
            purchase_number="PI-PINV-1201",
            supplier_invoice_number="SUP-1201",
            amount=Decimal("400.00"),
        )
        self._create_open_item(
            header=future_due_header,
            vendor=future_due_vendor,
            vendor_ledger=future_due_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 1),
            due_date=date(2025, 6, 30),
            purchase_number="PI-PINV-1201",
            supplier_invoice_number="SUP-1201",
            amount=Decimal("400.00"),
        )

        due_date_response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(as_of_date="2025-04-30", vendor=future_due_vendor.id, aging_basis="due_date"),
        )
        self.assertEqual(due_date_response.status_code, 200)
        due_date_row = due_date_response.json()["rows"][0]
        self.assertEqual(due_date_row["outstanding"], "400.00")
        self.assertEqual(due_date_row["not_due"], "400.00")
        self.assertEqual(due_date_row["overdue_amount"], "0.00")
        self.assertEqual(due_date_row["bucket_0_30"], "0.00")

        bill_date_response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(as_of_date="2025-04-30", vendor=future_due_vendor.id, aging_basis="bill_date"),
        )
        self.assertEqual(bill_date_response.status_code, 200)
        bill_date_row = bill_date_response.json()["rows"][0]
        self.assertEqual(bill_date_row["outstanding"], "400.00")
        self.assertEqual(bill_date_row["not_due"], "0.00")
        self.assertEqual(bill_date_row["overdue_amount"], "400.00")
        self.assertEqual(bill_date_row["bucket_0_30"], "400.00")

    def test_ap_aging_overdue_only_excludes_current_vendor_and_invoice_rows(self):
        current_vendor = self._create_vendor("Current Vendor", 5007)
        current_header = self._create_purchase_header(
            vendor=current_vendor,
            vendor_ledger=current_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 25),
            due_date=date(2025, 5, 20),
            doc_code="PINV",
            doc_no=1202,
            purchase_number="PI-PINV-1202",
            supplier_invoice_number="SUP-1202",
            amount=Decimal("275.00"),
        )
        self._create_open_item(
            header=current_header,
            vendor=current_vendor,
            vendor_ledger=current_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 25),
            due_date=date(2025, 5, 20),
            purchase_number="PI-PINV-1202",
            supplier_invoice_number="SUP-1202",
            amount=Decimal("275.00"),
        )

        summary_response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="summary", overdue_only="true"),
        )
        self.assertEqual(summary_response.status_code, 200)
        summary_payload = summary_response.json()
        self.assertEqual(summary_payload["summary"]["vendor_count"], 1)
        self.assertEqual(len(summary_payload["rows"]), 1)
        self.assertEqual(summary_payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(summary_payload["totals"]["outstanding"], "650.00")

        invoice_response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="invoice", overdue_only="true"),
        )
        self.assertEqual(invoice_response.status_code, 200)
        invoice_payload = invoice_response.json()
        self.assertEqual(len(invoice_payload["rows"]), 1)
        self.assertEqual(invoice_payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(invoice_payload["totals"]["balance"], "650.00")

    def test_vendor_outstanding_and_invoice_aging_apply_pagination(self):
        second_vendor = self._create_vendor("Vendor Two", 5004)
        second_invoice = self._create_purchase_header(
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 6),
            due_date=date(2025, 4, 12),
            doc_code="PINV",
            doc_no=1101,
            purchase_number="PI-PINV-1101",
            supplier_invoice_number="SUP-1101",
            amount=Decimal("120.00"),
        )
        self._create_open_item(
            header=second_invoice,
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 6),
            due_date=date(2025, 4, 12),
            purchase_number="PI-PINV-1101",
            supplier_invoice_number="SUP-1101",
            amount=Decimal("120.00"),
        )
        third_invoice = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 23),
            due_date=date(2025, 4, 27),
            doc_code="PINV",
            doc_no=1005,
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("80.00"),
        )
        self._create_open_item(
            header=third_invoice,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 23),
            due_date=date(2025, 4, 27),
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("80.00"),
        )

        outstanding_response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", page=1, page_size=1),
        )
        self.assertEqual(outstanding_response.status_code, 200)
        outstanding_payload = outstanding_response.json()
        self.assertEqual(len(outstanding_payload["rows"]), 1)
        self.assertEqual(outstanding_payload["pagination"]["total_rows"], 2)

        aging_summary_response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="summary", page=1, page_size=1),
        )
        self.assertEqual(aging_summary_response.status_code, 200)
        aging_summary_payload = aging_summary_response.json()
        self.assertEqual(len(aging_summary_payload["rows"]), 2)
        self.assertFalse(aging_summary_payload["pagination"]["paginated"])

        invoice_response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="invoice", page=1, page_size=1),
        )
        self.assertEqual(invoice_response.status_code, 200)
        invoice_payload = invoice_response.json()
        self.assertEqual(len(invoice_payload["rows"]), 1)
        self.assertTrue(invoice_payload["pagination"]["paginated"])
        self.assertGreaterEqual(invoice_payload["pagination"]["total_rows"], 2)

    def test_vendor_outstanding_credit_limit_exceeded_filters_to_breached_vendors(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data=None,
            commercial_data={"creditlimit": Decimal("600.00")},
            primary_address_data=None,
        )
        within_limit_vendor = self._create_vendor("Within Limit Vendor", 5008, creditlimit=Decimal("500.00"))
        within_limit_header = self._create_purchase_header(
            vendor=within_limit_vendor,
            vendor_ledger=within_limit_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 8),
            due_date=date(2025, 4, 12),
            doc_code="PINV",
            doc_no=1203,
            purchase_number="PI-PINV-1203",
            supplier_invoice_number="SUP-1203",
            amount=Decimal("200.00"),
        )
        self._create_open_item(
            header=within_limit_header,
            vendor=within_limit_vendor,
            vendor_ledger=within_limit_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 8),
            due_date=date(2025, 4, 12),
            purchase_number="PI-PINV-1203",
            supplier_invoice_number="SUP-1203",
            amount=Decimal("200.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(as_of_date="2025-04-30", credit_limit_exceeded="true"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(payload["rows"][0]["outstanding"], "650.00")

    def test_ap_aging_credit_limit_exceeded_filters_to_breached_vendors(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data=None,
            commercial_data={"creditlimit": Decimal("600.00")},
            primary_address_data=None,
        )
        current_vendor = self._create_vendor("Current Limit Vendor", 5009, creditlimit=Decimal("500.00"))
        current_header = self._create_purchase_header(
            vendor=current_vendor,
            vendor_ledger=current_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 18),
            due_date=date(2025, 5, 10),
            doc_code="PINV",
            doc_no=1204,
            purchase_number="PI-PINV-1204",
            supplier_invoice_number="SUP-1204",
            amount=Decimal("200.00"),
        )
        self._create_open_item(
            header=current_header,
            vendor=current_vendor,
            vendor_ledger=current_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 18),
            due_date=date(2025, 5, 10),
            purchase_number="PI-PINV-1204",
            supplier_invoice_number="SUP-1204",
            amount=Decimal("200.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="summary", credit_limit_exceeded="true"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(payload["rows"][0]["outstanding"], "650.00")

    def test_vendor_outstanding_search_filters_by_vendor_name(self):
        other_vendor = self._create_vendor("Searchable Vendor", 5010)
        other_header = self._create_purchase_header(
            vendor=other_vendor,
            vendor_ledger=other_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 9),
            due_date=date(2025, 4, 15),
            doc_code="PINV",
            doc_no=1205,
            purchase_number="PI-PINV-1205",
            supplier_invoice_number="SUP-1205",
            amount=Decimal("180.00"),
        )
        self._create_open_item(
            header=other_header,
            vendor=other_vendor,
            vendor_ledger=other_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 9),
            due_date=date(2025, 4, 15),
            purchase_number="PI-PINV-1205",
            supplier_invoice_number="SUP-1205",
            amount=Decimal("180.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(as_of_date="2025-04-30", search="Searchable"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["vendor_name"], "Searchable Vendor")
        self.assertEqual(payload["rows"][0]["outstanding"], "180.00")

    def test_vendor_outstanding_detailed_include_advances_separately_shows_advance_row(self):
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(
                as_of_date="2025-04-30",
                vendor=self.vendor.id,
                view="detailed",
                include_advances_separately="true",
                include_credit_balances="true",
            ),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        self.assertGreaterEqual(len(payload["rows"]), 2)

        advance_rows = [row for row in payload["rows"] if row.get("is_advance")]
        self.assertEqual(len(advance_rows), 1)
        advance_row = advance_rows[0]
        self.assertEqual(advance_row["voucher_no"], "ADV-001")
        self.assertEqual(advance_row["aging_bucket"], "advance")
        self.assertEqual(advance_row["outstanding_amount"], "-50.00")

    def test_payables_meta_endpoint_exposes_filter_and_report_metadata(self):
        response = self.client.get(
            reverse("reports_api:payables-meta"),
            self._base_scope(),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("vendors", payload)
        self.assertIn("subentities", payload)
        self.assertIn("financial_years", payload)
        self.assertEqual(payload["choices"]["aging_view_modes"][0]["value"], "summary")
        report_codes = {row["code"] for row in payload["reports"]}
        self.assertIn("vendor_outstanding", report_codes)
        self.assertIn("ap_aging", report_codes)
        self.assertIn("msme_overdue", report_codes)
        self.assertIn("payables_dashboard_summary", report_codes)
        self.assertIn("upcoming_payments_calendar", report_codes)

    def test_upcoming_payments_calendar_returns_due_window_rows_and_exports(self):
        response = self.client.get(
            reverse("reports_api:upcoming-payments-calendar"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "upcoming_payments_calendar")
        self.assertEqual(payload["summary"]["bill_count"], 1)
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        self.assertEqual(payload["rows"][0]["payment_status"], "Due in 30 Days")
        self.assertEqual(payload["rows"][0]["balance"], "650.00")
        self.assertEqual(payload["actions"]["export_urls"]["excel"].split("?")[0], "/api/reports/payables/upcoming-payments-calendar/excel/")

        export_response = self.client.get(
            reverse("reports_api:upcoming-payments-calendar-csv"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertTrue(export_response["Content-Type"].startswith("text/csv"))
        header_line = export_response.content.decode("utf-8-sig").splitlines()[0]
        self.assertIn("Vendor", header_line)

    def test_upcoming_payments_calendar_bill_detail_uses_service_route_for_service_bills(self):
        PurchaseInvoiceLine.objects.create(
            header=self.invoice,
            line_no=1,
            is_service=True,
            purchase_behavior="expense",
            product_desc="Annual maintenance",
        )

        response = self.client.get(
            reverse("reports_api:upcoming-payments-calendar"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        drilldown = response.json()["rows"][0]["_meta"]["drilldown"]["bill_detail"]
        self.assertEqual(drilldown["route"], "/purchaseserviceinvoice")

    def test_ap_payment_forecast_applies_pagination_without_changing_totals(self):
        second_vendor = self._create_vendor("Forecast Vendor", 5005)
        second_header = self._create_purchase_header(
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 28),
            doc_code="PINV",
            doc_no=1005,
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("250.00"),
        )
        self._create_open_item(
            header=second_header,
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 28),
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("250.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-payment-forecast"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", as_of_date="2025-04-30", sort_by="due_date", sort_order="asc", page=1, page_size=1),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "ap_payment_forecast")
        self.assertEqual(payload["pagination"]["page"], 1)
        self.assertEqual(payload["pagination"]["page_size"], 1)
        self.assertEqual(payload["pagination"]["total_rows"], 2)
        self.assertEqual(payload["summary"]["date_bands"], 2)
        self.assertEqual(payload["totals"]["due_amount"], "1050.00")
        self.assertEqual(payload["totals"]["overdue_amount"], "1050.00")
        self.assertEqual(payload["totals"]["next_30_days_amount"], "0.00")
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["due_date"], "2025-04-10")
        self.assertEqual(payload["rows"][0]["due_amount"], "800.00")
        self.assertEqual(payload["rows"][0]["payment_band"], "Overdue")

        second_page = self.client.get(
            reverse("reports_api:ap-payment-forecast"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", as_of_date="2025-04-30", sort_by="due_date", sort_order="asc", page=2, page_size=1),
        )
        self.assertEqual(second_page.status_code, 200)
        second_payload = second_page.json()
        self.assertEqual(second_payload["pagination"]["total_rows"], 2)
        self.assertEqual(len(second_payload["rows"]), 1)
        self.assertEqual(second_payload["rows"][0]["due_date"], "2025-04-28")
        self.assertEqual(second_payload["rows"][0]["due_amount"], "250.00")

    def test_ap_payment_forecast_exposes_detail_and_related_report_drilldowns(self):
        response = self.client.get(
            reverse("reports_api:ap-payment-forecast"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", as_of_date="2025-04-30", sort_by="due_date", sort_order="asc"),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertEqual(row["can_drilldown"], True)
        self.assertIn("forecast_detail", row["_meta"]["drilldown"])
        self.assertIn("upcoming_payments_calendar", row["_meta"]["drilldown"])
        self.assertIn("vendor_outstanding", row["_meta"]["drilldown"])
        self.assertIn("ap_aging", row["_meta"]["drilldown"])
        self.assertEqual(row["_meta"]["drilldown"]["forecast_detail"]["kind"], "detail")
        self.assertEqual(row["_meta"]["drilldown"]["forecast_detail"]["params"]["due_date"], "2025-04-10")
        self.assertEqual(row["_meta"]["drilldown"]["upcoming_payments_calendar"]["report_code"], "upcoming_payments_calendar")
        self.assertEqual(row["_meta"]["drilldown"]["upcoming_payments_calendar"]["params"]["from_date"], "2025-04-10")
        self.assertEqual(row["_meta"]["drilldown"]["ap_aging"]["params"]["view"], "invoice")
        self.assertIn("ABC Traders", row["sample_vendor_names"])
        self.assertIn("PI-PINV-1001", row["sample_bill_numbers"])

    def test_vendor_reconciliation_statement_excel_export_returns_workbook(self):
        response = self.client.get(
            reverse("reports_api:vendor-reconciliation-statement-excel"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
        self.assertGreaterEqual(len(workbook.sheetnames), 1)
        self.assertNotRegex(workbook.sheetnames[0], r"[\\\\/*?:\\[\\]]")
        self.assertLessEqual(len(workbook.sheetnames[0]), 31)

    def test_vendor_reconciliation_statement_applies_pagination_without_changing_totals(self):
        second_vendor = self._create_vendor("XYZ Supplies", 5004)
        second_header = self._create_purchase_header(
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 20),
            due_date=date(2025, 4, 25),
            doc_code="PINV",
            doc_no=1004,
            purchase_number="PI-PINV-1004",
            supplier_invoice_number="SUP-004",
            amount=Decimal("400.00"),
        )
        self._create_open_item(
            header=second_header,
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 20),
            due_date=date(2025, 4, 25),
            purchase_number="PI-PINV-1004",
            supplier_invoice_number="SUP-004",
            amount=Decimal("400.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-reconciliation-statement"),
            self._base_scope(as_of_date="2025-04-30", sort_by="closing_balance", sort_order="desc", page=1, page_size=1),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "vendor_reconciliation_statement")
        self.assertEqual(payload["pagination"]["page"], 1)
        self.assertEqual(payload["pagination"]["page_size"], 1)
        self.assertEqual(payload["pagination"]["total_rows"], 2)
        self.assertEqual(payload["summary"]["vendor_count"], 2)
        self.assertEqual(payload["totals"]["invoiced"], "1400.00")
        self.assertEqual(payload["totals"]["notes"], "100.00")
        self.assertEqual(payload["totals"]["settled"], "200.00")
        self.assertEqual(payload["totals"]["closing_balance"], "1100.00")
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(payload["rows"][0]["closing_balance"], "700.00")
        self.assertEqual(payload["rows"][0]["status"], "Mismatch")

        second_page = self.client.get(
            reverse("reports_api:vendor-reconciliation-statement"),
            self._base_scope(as_of_date="2025-04-30", sort_by="closing_balance", sort_order="desc", page=2, page_size=1),
        )
        self.assertEqual(second_page.status_code, 200)
        second_payload = second_page.json()
        self.assertEqual(second_payload["pagination"]["total_rows"], 2)
        self.assertEqual(len(second_payload["rows"]), 1)
        self.assertEqual(second_payload["rows"][0]["vendor_name"], "XYZ Supplies")
        self.assertEqual(second_payload["rows"][0]["closing_balance"], "400.00")
        self.assertEqual(second_payload["rows"][0]["status"], "Mismatch")

    def test_vendor_reconciliation_statement_uses_opening_balance_and_period_movement(self):
        response = self.client.get(
            reverse("reports_api:vendor-reconciliation-statement"),
            self._base_scope(from_date="2025-04-10", to_date="2025-04-30", sort_by="vendor_name", sort_order="asc"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["vendor_count"], 1)
        row = payload["rows"][0]
        self.assertEqual(row["vendor_name"], "ABC Traders")
        self.assertEqual(row["opening_balance"], "800.00")
        self.assertEqual(row["invoiced"], "0.00")
        self.assertEqual(row["notes"], "100.00")
        self.assertEqual(row["settled"], "0.00")
        self.assertEqual(row["closing_balance"], "700.00")
        self.assertEqual(row["status"], "Mismatch")
        self.assertEqual(payload["totals"]["opening_balance"], "800.00")
        self.assertEqual(payload["totals"]["invoiced"], "0.00")
        self.assertEqual(payload["totals"]["notes"], "100.00")
        self.assertEqual(payload["totals"]["settled"], "0.00")
        self.assertEqual(payload["totals"]["closing_balance"], "700.00")

    def test_vendor_reconciliation_statement_exposes_detail_and_related_report_drilldowns(self):
        response = self.client.get(
            reverse("reports_api:vendor-reconciliation-statement"),
            self._base_scope(as_of_date="2025-04-30", sort_by="closing_balance", sort_order="desc"),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertEqual(row["can_drilldown"], True)
        self.assertIn("reconciliation_detail", row["_meta"]["drilldown"])
        self.assertIn("vendor_ledger_statement", row["_meta"]["drilldown"])
        self.assertIn("vendor_outstanding", row["_meta"]["drilldown"])
        self.assertIn("ap_aging", row["_meta"]["drilldown"])
        self.assertEqual(row["_meta"]["drilldown"]["reconciliation_detail"]["kind"], "detail")
        self.assertEqual(row["_meta"]["drilldown"]["reconciliation_detail"]["params"]["vendor_name"], "ABC Traders")
        self.assertEqual(row["_meta"]["drilldown"]["vendor_ledger_statement"]["params"]["vendor"], self.vendor.id)
        self.assertEqual(row["_meta"]["drilldown"]["ap_aging"]["params"]["view"], "invoice")

    def test_duplicate_anomalous_bill_detection_excel_export_returns_workbook(self):
        response = self.client.get(
            reverse("reports_api:duplicate-anomalous-bill-detection-excel"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
        self.assertGreaterEqual(len(workbook.sheetnames), 1)
        self.assertNotRegex(workbook.sheetnames[0], r"[\\\\/*?:\\[\\]]")
        self.assertLessEqual(len(workbook.sheetnames[0]), 31)

    def test_duplicate_anomalous_bill_detection_applies_pagination_without_changing_totals(self):
        duplicate_one = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 25),
            due_date=date(2025, 4, 25),
            doc_code="PINV",
            doc_no=1008,
            purchase_number="PI-PINV-1008",
            supplier_invoice_number="DUP-001",
            amount=Decimal("450.00"),
        )
        duplicate_two = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 26),
            due_date=date(2025, 4, 26),
            doc_code="PINV",
            doc_no=1009,
            purchase_number="PI-PINV-1009",
            supplier_invoice_number="DUP-001",
            amount=Decimal("450.00"),
        )

        response = self.client.get(
            reverse("reports_api:duplicate-anomalous-bill-detection"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", sort_by="bill_date", sort_order="desc", page=1, page_size=1),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["pagination"]["total_rows"], 2)
        self.assertEqual(payload["summary"]["anomaly_count_total"], 2)
        self.assertEqual(payload["summary"]["anomaly_counts"]["POSSIBLE_DUPLICATE"], 2)
        self.assertEqual(payload["totals"]["grand_total"], "900.00")
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["supplier_invoice_number"], duplicate_two.supplier_invoice_number)
        self.assertEqual(payload["rows"][0]["bill_date"], "2025-04-26")
        self.assertEqual(payload["rows"][0]["anomaly_type"], "POSSIBLE_DUPLICATE")

        second_page = self.client.get(
            reverse("reports_api:duplicate-anomalous-bill-detection"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", sort_by="bill_date", sort_order="desc", page=2, page_size=1),
        )
        self.assertEqual(second_page.status_code, 200)
        second_payload = second_page.json()
        self.assertEqual(second_payload["pagination"]["total_rows"], 2)
        self.assertEqual(len(second_payload["rows"]), 1)
        self.assertEqual(second_payload["rows"][0]["supplier_invoice_number"], duplicate_one.supplier_invoice_number)
        self.assertEqual(second_payload["rows"][0]["bill_date"], "2025-04-25")
        self.assertEqual(second_payload["rows"][0]["anomaly_type"], "POSSIBLE_DUPLICATE")

    def test_duplicate_anomalous_bill_detection_ignores_whitespace_supplier_invoice_numbers(self):
        self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 27),
            due_date=date(2025, 4, 27),
            doc_code="PINV",
            doc_no=1010,
            purchase_number="PI-PINV-1010",
            supplier_invoice_number="   ",
            amount=Decimal("600.00"),
        )
        self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 28),
            due_date=date(2025, 4, 28),
            doc_code="PINV",
            doc_no=1011,
            purchase_number="PI-PINV-1011",
            supplier_invoice_number="   ",
            amount=Decimal("600.00"),
        )

        response = self.client.get(
            reverse("reports_api:duplicate-anomalous-bill-detection"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["anomaly_count_total"], 0)
        self.assertEqual(payload["summary"]["anomaly_counts"], {})
        self.assertEqual(payload["rows"], [])
        self.assertEqual(payload["totals"]["grand_total"], "0.00")

    def test_duplicate_anomalous_bill_detection_exposes_detail_and_related_report_drilldowns(self):
        self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 25),
            due_date=date(2025, 4, 25),
            doc_code="PINV",
            doc_no=1008,
            purchase_number="PI-PINV-1008",
            supplier_invoice_number="DUP-001",
            amount=Decimal("450.00"),
        )
        duplicate_two = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 26),
            due_date=date(2025, 4, 26),
            doc_code="PINV",
            doc_no=1009,
            purchase_number="PI-PINV-1009",
            supplier_invoice_number="DUP-001",
            amount=Decimal("450.00"),
        )

        response = self.client.get(
            reverse("reports_api:duplicate-anomalous-bill-detection"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload["rows"] if item["purchase_number"] == duplicate_two.purchase_number)
        self.assertEqual(row["can_drilldown"], True)
        self.assertIn("duplicate_bill_detail", row["_meta"]["drilldown"])
        self.assertIn("bill_detail", row["_meta"]["drilldown"])
        self.assertIn("vendor_outstanding", row["_meta"]["drilldown"])
        self.assertIn("ap_aging", row["_meta"]["drilldown"])
        self.assertEqual(row["_meta"]["drilldown"]["duplicate_bill_detail"]["kind"], "detail")
        self.assertEqual(row["_meta"]["drilldown"]["duplicate_bill_detail"]["params"]["purchase_number"], duplicate_two.purchase_number)
        self.assertEqual(row["_meta"]["drilldown"]["bill_detail"]["params"]["id"], duplicate_two.id)
        self.assertEqual(row["_meta"]["drilldown"]["vendor_outstanding"]["params"]["vendor"], self.vendor.id)
        self.assertEqual(row["_meta"]["drilldown"]["ap_aging"]["params"]["view"], "invoice")

    def test_grn_invoice_posting_exceptions_applies_pagination_without_changing_totals(self):
        missing_supplier = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 18),
            due_date=date(2025, 4, 18),
            doc_code="PINV",
            doc_no=1006,
            purchase_number="PI-PINV-1006",
            supplier_invoice_number="",
            amount=Decimal("150.00"),
        )
        not_posted = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 21),
            due_date=date(2025, 4, 21),
            doc_code="PINV",
            doc_no=1007,
            purchase_number="PI-PINV-1007",
            supplier_invoice_number="SUP-007",
            amount=Decimal("200.00"),
            status=PurchaseInvoiceHeader.Status.DRAFT,
        )

        response = self.client.get(
            reverse("reports_api:grn-invoice-posting-exceptions"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", sort_by="bill_date", sort_order="desc", page=1, page_size=1),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["pagination"]["total_rows"], 2)
        self.assertEqual(payload["summary"]["issue_count_total"], 2)
        self.assertEqual(payload["totals"]["grand_total"], "350.00")
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["purchase_number"], not_posted.purchase_number)
        self.assertEqual(payload["rows"][0]["issue_type"], "NOT_POSTED")

        second_page = self.client.get(
            reverse("reports_api:grn-invoice-posting-exceptions"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", sort_by="bill_date", sort_order="desc", page=2, page_size=1),
        )
        self.assertEqual(second_page.status_code, 200)
        second_payload = second_page.json()
        self.assertEqual(second_payload["pagination"]["total_rows"], 2)
        self.assertEqual(len(second_payload["rows"]), 1)
        self.assertEqual(second_payload["rows"][0]["purchase_number"], missing_supplier.purchase_number)
        self.assertEqual(second_payload["rows"][0]["issue_type"], "MISSING_SUPPLIER_INVOICE")

    def test_grn_invoice_posting_exceptions_treats_whitespace_supplier_invoice_as_missing(self):
        header = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 22),
            doc_code="PINV",
            doc_no=1012,
            purchase_number="PI-PINV-1012",
            supplier_invoice_number="   ",
            amount=Decimal("175.00"),
        )

        response = self.client.get(
            reverse("reports_api:grn-invoice-posting-exceptions"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        target_row = next(row for row in payload["rows"] if row["purchase_number"] == header.purchase_number)
        self.assertEqual(target_row["supplier_invoice_number"], "-")
        self.assertEqual(target_row["issue_type"], "MISSING_SUPPLIER_INVOICE")

    def test_grn_invoice_posting_exceptions_exposes_detail_and_related_report_drilldowns(self):
        header = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 21),
            due_date=date(2025, 4, 21),
            doc_code="PINV",
            doc_no=1007,
            purchase_number="PI-PINV-1007",
            supplier_invoice_number="SUP-007",
            amount=Decimal("200.00"),
            status=PurchaseInvoiceHeader.Status.DRAFT,
        )

        response = self.client.get(
            reverse("reports_api:grn-invoice-posting-exceptions"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        row = next(item for item in payload["rows"] if item["purchase_number"] == header.purchase_number)
        self.assertEqual(row["can_drilldown"], True)
        self.assertIn("grn_exception_detail", row["_meta"]["drilldown"])
        self.assertIn("bill_detail", row["_meta"]["drilldown"])
        self.assertIn("vendor_outstanding", row["_meta"]["drilldown"])
        self.assertIn("ap_aging", row["_meta"]["drilldown"])
        self.assertEqual(row["_meta"]["drilldown"]["grn_exception_detail"]["kind"], "detail")
        self.assertEqual(row["_meta"]["drilldown"]["grn_exception_detail"]["params"]["purchase_number"], header.purchase_number)
        self.assertEqual(row["_meta"]["drilldown"]["bill_detail"]["params"]["id"], header.id)
        self.assertEqual(row["_meta"]["drilldown"]["vendor_outstanding"]["params"]["vendor"], self.vendor.id)
        self.assertEqual(row["_meta"]["drilldown"]["ap_aging"]["params"]["view"], "invoice")

    def test_ap_compliance_aging_applies_pagination_without_changing_totals(self):
        second_vendor = self._create_vendor("No GST Vendor", 5003)
        second_header = self._create_purchase_header(
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 1),
            due_date=date(2025, 7, 15),
            doc_code="PINV",
            doc_no=1003,
            purchase_number="PI-PINV-1003",
            supplier_invoice_number="SUP-003",
            amount=Decimal("300.00"),
        )
        self._create_open_item(
            header=second_header,
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 1),
            due_date=date(2025, 7, 15),
            purchase_number="PI-PINV-1003",
            supplier_invoice_number="SUP-003",
            amount=Decimal("300.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-compliance-aging"),
            self._base_scope(as_of_date="2025-08-01", sort_by="days_overdue", sort_order="desc", page=1, page_size=1),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "ap_compliance_aging")
        self.assertEqual(payload["pagination"]["page"], 1)
        self.assertEqual(payload["pagination"]["page_size"], 1)
        self.assertEqual(payload["pagination"]["total_rows"], 2)
        self.assertEqual(payload["summary"]["row_count"], 2)
        self.assertEqual(payload["totals"]["outstanding"], "1100.00")
        self.assertEqual(payload["summary"]["risk_counts"]["MEDIUM"], 1)
        self.assertEqual(payload["summary"]["risk_counts"]["HIGH"], 1)
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(payload["rows"][0]["days_overdue"], 113)
        self.assertEqual(payload["rows"][0]["outstanding"], "800.00")

        second_page = self.client.get(
            reverse("reports_api:ap-compliance-aging"),
            self._base_scope(as_of_date="2025-08-01", sort_by="days_overdue", sort_order="desc", page=2, page_size=1),
        )
        self.assertEqual(second_page.status_code, 200)
        second_payload = second_page.json()
        self.assertEqual(second_payload["pagination"]["total_rows"], 2)
        self.assertEqual(len(second_payload["rows"]), 1)
        self.assertEqual(second_payload["rows"][0]["vendor_name"], "No GST Vendor")
        self.assertEqual(second_payload["rows"][0]["compliance_risk"], "HIGH")
        self.assertEqual(second_payload["rows"][0]["outstanding"], "300.00")

    def test_ap_compliance_aging_search_matches_bill_number(self):
        second_vendor = self._create_vendor("Search Vendor", 5004)
        second_header = self._create_purchase_header(
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 5),
            due_date=date(2025, 7, 20),
            doc_code="PINV",
            doc_no=1004,
            purchase_number="PI-PINV-1004",
            supplier_invoice_number="FIND-ME-004",
            amount=Decimal("250.00"),
        )
        self._create_open_item(
            header=second_header,
            vendor=second_vendor,
            vendor_ledger=second_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 5),
            due_date=date(2025, 7, 20),
            purchase_number="PI-PINV-1004",
            supplier_invoice_number="FIND-ME-004",
            amount=Decimal("250.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-compliance-aging"),
            self._base_scope(as_of_date="2025-08-01", search="FIND-ME-004"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["row_count"], 1)
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["bill_number"], "PI-PINV-1004")

    def test_ap_compliance_aging_exposes_detail_and_related_report_drilldowns(self):
        response = self.client.get(
            reverse("reports_api:ap-compliance-aging"),
            self._base_scope(as_of_date="2025-08-01", sort_by="days_overdue", sort_order="desc"),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertEqual(row["can_drilldown"], True)
        self.assertIn("compliance_detail", row["_meta"]["drilldown"])
        self.assertIn("bill_detail", row["_meta"]["drilldown"])
        self.assertIn("vendor_outstanding", row["_meta"]["drilldown"])
        self.assertIn("ap_aging", row["_meta"]["drilldown"])
        self.assertIn("vendor_ledger_statement", row["_meta"]["drilldown"])
        self.assertEqual(row["_meta"]["drilldown"]["compliance_detail"]["kind"], "detail")
        self.assertEqual(row["_meta"]["drilldown"]["compliance_detail"]["params"]["vendor_id"], self.vendor.id)
        self.assertEqual(row["_meta"]["drilldown"]["bill_detail"]["params"]["id"], self.invoice.id)
        self.assertEqual(row["_meta"]["drilldown"]["ap_aging"]["params"]["view"], "invoice")

    def test_ap_compliance_aging_excludes_customer_only_open_items(self):
        customer_ledger = Ledger.objects.create(
            entity=self.entity,
            ledger_code=7001,
            name="Customer 1",
            accounthead=self.vendor_head,
            createdby=self.user,
        )
        customer = create_account_with_synced_ledger(
            account_data={
                "entity": self.entity,
                "ledger": customer_ledger,
                "accountname": "Customer 1",
                "createdby": self.user,
            },
            ledger_overrides={"ledger_code": 7001, "accounthead": self.vendor_head, "is_party": True},
        )
        apply_normalized_profile_payload(
            customer,
            compliance_data={"gstno": "27ABCDE1234F1Z5"},
            commercial_data={
                "partytype": "Customer",
                "currency": "INR",
                "creditdays": 30,
                "creditlimit": Decimal("1000.00"),
            },
            primary_address_data={"state": self.state, "city": self.city},
        )
        customer_header = self._create_purchase_header(
            vendor=customer,
            vendor_ledger=customer_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 5),
            due_date=date(2025, 7, 20),
            doc_code="PINV",
            doc_no=1005,
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("250.00"),
        )
        self._create_open_item(
            header=customer_header,
            vendor=customer,
            vendor_ledger=customer_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 5),
            due_date=date(2025, 7, 20),
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("250.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-compliance-aging"),
            self._base_scope(as_of_date="2025-08-01", sort_by="days_overdue", sort_order="desc", page=1, page_size=50),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        vendor_names = [row["vendor_name"] for row in payload["rows"]]
        self.assertNotIn("Customer 1", vendor_names)
        self.assertEqual(payload["summary"]["row_count"], 1)
        self.assertEqual(payload["rows"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(payload["totals"]["outstanding"], "800.00")

    def test_ap_compliance_aging_treats_whitespace_gstin_as_missing_across_sort_paths(self):
        whitespace_vendor = self._create_vendor("Whitespace GST Vendor", 5005)
        apply_normalized_profile_payload(
            whitespace_vendor,
            compliance_data={"gstno": "   "},
            commercial_data=None,
            primary_address_data=None,
        )
        header = self._create_purchase_header(
            vendor=whitespace_vendor,
            vendor_ledger=whitespace_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 10),
            due_date=date(2025, 7, 20),
            doc_code="PINV",
            doc_no=1005,
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("125.00"),
        )
        self._create_open_item(
            header=header,
            vendor=whitespace_vendor,
            vendor_ledger=whitespace_vendor.ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 10),
            due_date=date(2025, 7, 20),
            purchase_number="PI-PINV-1005",
            supplier_invoice_number="SUP-005",
            amount=Decimal("125.00"),
        )

        response = self.client.get(
            reverse("reports_api:ap-compliance-aging"),
            self._base_scope(as_of_date="2025-08-01", sort_by="vendor_name", sort_order="asc"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        target_row = next(row for row in payload["rows"] if row["vendor_name"] == "Whitespace GST Vendor")
        self.assertEqual(target_row["gstin"], "-")
        self.assertEqual(target_row["compliance_risk"], "HIGH")
        self.assertEqual(payload["summary"]["risk_counts"]["HIGH"], 1)
        self.assertEqual(payload["summary"]["risk_counts"]["MEDIUM"], 1)
        self.assertEqual(payload["summary"]["risk_counts"]["LOW"], 0)

    def test_payables_meta_filters_reports_and_report_endpoints_enforce_permissions(self):
        limited_user = self._create_limited_report_user("reports.vendoroutstanding.view")
        limited_client = APIClient()
        limited_client.force_authenticate(user=limited_user)

        meta_response = limited_client.get(reverse("reports_api:payables-meta"), self._base_scope())
        self.assertEqual(meta_response.status_code, 403)

        denied_response = limited_client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="summary"),
        )
        self.assertEqual(denied_response.status_code, 403)
        detail = denied_response.json()["detail"].lower()
        self.assertTrue("permission" in detail or "access to this entity" in detail)

    def test_accountspayableaging_alias_routes_to_canonical_ap_aging(self):
        response = self.client.get(
            reverse("reports_api:accountspayableaging-report"),
            self._base_scope(as_of_date="2025-04-30", view="summary"),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["report_code"], "ap_aging")

    def test_report_preferences_api_persists_and_meta_echoes_saved_state(self):
        pref_response = self.client.patch(
            reverse("reports_api:report-preferences"),
            {
                "entity": self.entity.id,
                "report_code": "ap_aging",
                "payload": {
                    "view": "invoice",
                    "sort_by": "balance",
                    "sort_order": "asc",
                    "page_size": 25,
                },
            },
            format="json",
        )
        self.assertEqual(pref_response.status_code, 200)
        self.assertEqual(pref_response.json()["report_code"], "ap_aging")

        meta_response = self.client.get(reverse("reports_api:payables-meta"), self._base_scope())
        self.assertEqual(meta_response.status_code, 200)
        prefs = meta_response.json().get("user_preferences", {})
        self.assertEqual(prefs.get("ap_aging", {}).get("view"), "invoice")
        self.assertEqual(prefs.get("ap_aging", {}).get("sort_by"), "balance")

    def test_payables_dashboard_summary_api_returns_totals_and_top_vendors(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "msme_status": "micro",
                "udyam_no": "UDYAM-PB-0001",
                "has_written_payment_terms": False,
            },
        )
        response = self.client.get(
            reverse("reports_api:payables-dashboard-summary"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "payables_dashboard_summary")
        self.assertEqual(payload["totals"]["vendor_outstanding"], "650.00")
        self.assertEqual(payload["totals"]["bucket_1_30"], "650.00")
        self.assertEqual(payload["totals"]["msme_overdue_amount"], "800.00")
        self.assertEqual(payload["vendor_count_with_open_balance"], 1)
        self.assertEqual(payload["summary"]["msme_overdue_bill_count"], 1)
        self.assertEqual(payload["summary"]["msme_overdue_vendor_count"], 1)
        self.assertEqual(payload["summary"]["msme_oldest_overdue_days"], 14)
        self.assertIn("800.00", payload["summary"]["msme_reporting_note"])
        self.assertEqual(payload["top_vendors"][0]["vendor_name"], "ABC Traders")
        self.assertEqual(payload["top_msme_overdue_vendors"][0]["vendor_name"], "ABC Traders")
        self.assertFalse(payload["actions"]["can_export_excel"])

    def test_report_metadata_and_drilldown_contracts_are_standardized(self):
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", reconcile_gl="true"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        meta = payload["_meta"]
        self.assertEqual(meta["required_menu_code"], "reports.vendoroutstanding")
        self.assertIn("reports.vendoroutstanding.view", meta["required_permission_codes"])
        self.assertIn("generated_at", meta)
        self.assertTrue(meta["supports_drilldown"])
        self.assertIn("excel", meta["exportable_formats"])
        self.assertTrue(meta["gl_reconciliation_warning"])
        row_drilldown = payload["rows"][0]["_meta"]["drilldown"]["aging_summary"]
        self.assertEqual(row_drilldown["report_code"], "ap_aging")
        self.assertEqual(row_drilldown["path"], "/api/reports/payables/aging/")

    def test_report_metadata_surfaces_payables_settings_display_and_overrides(self):
        settings_response = self.client.patch(
            reverse("reports_api:payables-settings"),
            {
                "entity": self.entity.id,
                "payload": {
                    "display_preferences": {
                        "amount_unit": "thousand",
                        "decimal_places": 3,
                    },
                    "report_overrides": {
                        "vendor_outstanding": {
                            "columns": ["vendor_name", "outstanding"],
                            "default_sort_by": "vendor_name",
                            "default_sort_order": "asc",
                        }
                    },
                },
            },
            format="json",
        )
        self.assertEqual(settings_response.status_code, 200)

        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        meta = response.json()["_meta"]
        self.assertEqual(meta["display_preferences"]["amount_unit"], "thousand")
        self.assertEqual(meta["display_preferences"]["decimal_places"], 3)
        self.assertEqual(meta["report_override"]["default_sort_by"], "vendor_name")
        self.assertEqual(meta["report_override"]["default_sort_order"], "asc")
        self.assertIn("drilldown", meta["report_override"]["columns"])
        self.assertIn("vendor_name", meta["effective_columns"])
        self.assertIn("outstanding", meta["effective_columns"])
        self.assertNotIn("vendor_code", meta["effective_columns"])

    def test_vendor_outstanding_csv_export_honors_selected_columns(self):
        settings_response = self.client.patch(
            reverse("reports_api:payables-settings"),
            {
                "entity": self.entity.id,
                "payload": {
                    "report_overrides": {
                        "vendor_outstanding": {
                            "columns": ["vendor_name", "outstanding", "drilldown"],
                        }
                    },
                },
            },
            format="json",
        )
        self.assertEqual(settings_response.status_code, 200)

        export_response = self.client.get(
            reverse("reports_api:vendor-outstanding-report-csv"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(export_response.status_code, 200)
        header_line = export_response.content.decode("utf-8-sig").splitlines()[0]
        self.assertEqual(header_line, "Vendor Name,Outstanding")

    def test_report_endpoints_require_authentication(self):
        client = APIClient()
        response = client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 403)

    def test_rbac_menu_registration_exposes_vendor_outstanding_and_ap_aging(self):
        tree = EffectiveMenuService.menu_tree_for_user(self.user, self.entity.id)

        def flatten(nodes):
            codes = []
            for node in nodes:
                codes.append(node["menu_code"])
                codes.extend(flatten(node["children"]))
            return codes

        def find_node(nodes, menu_code):
            for node in nodes:
                if node["menu_code"] == menu_code:
                    return node
                child = find_node(node["children"], menu_code)
                if child:
                    return child
            return None

        codes = set(flatten(tree))
        self.assertIn("reports.reports.payables", codes)
        self.assertIn("reports.accountspayableaging", codes)
        payables_menu = find_node(tree, "reports.reports.payables")
        self.assertIsNotNone(payables_menu)
        self.assertEqual(payables_menu["route_path"], "/reports/payables")
        child_codes = [child["menu_code"] for child in payables_menu["children"]]
        self.assertIn("reports.accountspayableaging", child_codes)
        ap_aging_menu = find_node(tree, "reports.accountspayableaging")
        self.assertIsNotNone(ap_aging_menu)
        self.assertEqual(ap_aging_menu["route_path"], "/reports/payables/ap_aging")

    def test_export_endpoints_return_expected_formats(self):
        vendor_scope = self._base_scope(from_date="2025-04-01", to_date="2025-04-30")
        ap_summary_scope = self._base_scope(as_of_date="2025-04-30", view="summary")
        ap_invoice_scope = self._base_scope(as_of_date="2025-04-30", view="invoice")

        export_checks = [
            ("reports_api:vendor-outstanding-report-csv", vendor_scope, "text/csv", b"Vendor Name"),
            ("reports_api:vendor-outstanding-report-excel", vendor_scope, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:vendor-outstanding-report-pdf", vendor_scope, "application/pdf", b"%PDF"),
            ("reports_api:vendor-outstanding-report-print", vendor_scope, "application/pdf", b"%PDF"),
            ("reports_api:ap-aging-report-csv", ap_summary_scope, "text/csv", b"Vendor"),
            ("reports_api:ap-aging-report-excel", ap_summary_scope, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:ap-aging-report-pdf", ap_summary_scope, "application/pdf", b"%PDF"),
            ("reports_api:ap-aging-report-print", ap_invoice_scope, "application/pdf", b"%PDF"),
            ("reports_api:upcoming-payments-calendar-excel", vendor_scope, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:upcoming-payments-calendar-pdf", vendor_scope, "application/pdf", b"%PDF"),
            ("reports_api:upcoming-payments-calendar-print", vendor_scope, "application/pdf", b"%PDF"),
        ]
        for route_name, params, content_type, prefix in export_checks:
            with self.subTest(route_name=route_name):
                response = self.client.get(reverse(route_name), params)
                self.assertEqual(response.status_code, 200)
                self.assertTrue(response["Content-Type"].startswith(content_type))
                if route_name == "reports_api:vendor-outstanding-report-csv":
                    header_line = response.content.decode("utf-8-sig").splitlines()[0]
                    self.assertIn("Vendor", header_line)
                else:
                    self.assertTrue(bytes(response.content).startswith(prefix))
                if route_name.endswith("-print"):
                    self.assertIn("inline;", response["Content-Disposition"])
                else:
                    self.assertIn("attachment;", response["Content-Disposition"])


    def _post_vendor_gl_balance(self, amount, *, posting_date, txn_id=9001, voucher_no="GL-AP-1"):
        batch = PostingBatch.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL,
            txn_id=txn_id,
            created_by=self.user,
        )
        entry = Entry.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL,
            txn_id=txn_id,
            voucher_no=voucher_no,
            voucher_date=posting_date,
            posting_date=posting_date,
            status=EntryStatus.POSTED,
            posted_at=timezone.now(),
            posted_by=self.user,
            posting_batch=batch,
            narration="AP reconciliation test entry",
            created_by=self.user,
        )
        amount = Decimal(amount)
        vendor_is_debit = amount < 0
        abs_amount = abs(amount)
        JournalLine.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL,
            txn_id=txn_id,
            voucher_no=voucher_no,
            account=self.vendor,
            ledger=self.vendor_ledger,
            drcr=vendor_is_debit,
            amount=abs_amount,
            description="Vendor control",
            posting_date=posting_date,
            posted_at=timezone.now(),
            created_by=self.user,
        )
        JournalLine.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=TxnType.JOURNAL,
            txn_id=txn_id,
            voucher_no=voucher_no,
            accounthead=self.expense_head,
            drcr=not vendor_is_debit,
            amount=abs_amount,
            description="Balancing expense",
            posting_date=posting_date,
            posted_at=timezone.now(),
            created_by=self.user,
        )
        return entry

    def test_ap_gl_reconciliation_report_matches_when_gl_balance_aligns(self):
        self._post_vendor_gl_balance(Decimal("650.00"), posting_date=date(2025, 4, 30))
        response = self.client.get(
            reverse("reports_api:ap-gl-reconciliation-report"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["overall_status"], "matched")
        self.assertEqual(payload["totals"]["subledger_balance"], "650.00")
        self.assertEqual(payload["totals"]["gl_balance"], "650.00")
        self.assertEqual(payload["totals"]["difference_amount"], "0.00")
        self.assertEqual(payload["rows"][0]["reconciliation_status"], "matched")

    def test_ap_gl_reconciliation_report_flags_mismatch_without_gl_balance(self):
        response = self.client.get(
            reverse("reports_api:ap-gl-reconciliation-report"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["summary"]["overall_status"], "mismatch")
        self.assertEqual(payload["rows"][0]["difference_amount"], "650.00")
        self.assertEqual(payload["rows"][0]["reconciliation_status"], "mismatch")

    def test_ap_gl_reconciliation_vendor_statement_drilldown_points_to_vendor_ledger_statement(self):
        response = self.client.get(
            reverse("reports_api:ap-gl-reconciliation-report"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        drilldown = payload["rows"][0]["_meta"]["drilldown"]["vendor_statement"]
        self.assertEqual(drilldown["target"], "vendor_ledger_statement")
        self.assertEqual(drilldown["report_code"], "vendor_ledger_statement")
        self.assertEqual(drilldown["kind"], "report")
        self.assertEqual(drilldown["params"]["vendor"], self.vendor.id)

    def test_vendor_balance_exception_report_detects_negative_balance_and_stale_advance(self):
        credit_vendor = self._create_vendor("Advance Heavy Vendor", 5010)
        self._create_advance(
            vendor=credit_vendor,
            vendor_ledger=credit_vendor.ledger,
            credit_date=date(2025, 1, 1),
            reference_no="ADV-OLD-1",
            amount=Decimal("125.00"),
        )
        response = self.client.get(
            reverse("reports_api:vendor-balance-exception-report"),
            self._base_scope(as_of_date="2025-08-01", stale_days_gt=30, min_amount="10.00", page_size=500),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        exception_types = set(payload["summary"]["by_type"].keys())
        self.assertIn("negative_vendor_balance", exception_types)
        self.assertIn("old_unapplied_advance", exception_types)

    def test_settlement_integrity_mismatch_surfaces_in_close_validation(self):
        broken = VendorSettlement.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            settlement_type=VendorSettlement.SettlementType.PAYMENT,
            settlement_date=date(2025, 4, 8),
            reference_no="PAY-BROKEN",
            total_amount=Decimal("300.00"),
            status=VendorSettlement.Status.POSTED,
            posted_at=timezone.now(),
            posted_by=self.user,
        )
        VendorSettlementLine.objects.create(
            settlement=broken,
            open_item=self.invoice_item,
            amount=Decimal("50.00"),
            applied_amount_signed=Decimal("50.00"),
        )
        response = self.client.get(
            reverse("reports_api:payables-close-validation"),
            self._base_scope(as_of_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        checks = {check["check_code"]: check for check in response.json()["checks"]}
        self.assertGreater(checks["settlement_integrity_errors"]["affected_count"], 0)
        self.assertIn("PAY-BROKEN", str(checks["settlement_integrity_errors"]["sample_references"]))

    def test_close_readiness_summary_returns_rollup_values(self):
        self._post_vendor_gl_balance(Decimal("650.00"), posting_date=date(2025, 8, 1), txn_id=9002, voucher_no="GL-AP-2")
        response = self.client.get(
            reverse("reports_api:payables-close-readiness-summary"),
            self._base_scope(as_of_date="2025-08-01"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["ap_gl_reconciliation_status"], "matched")
        self.assertEqual(payload["difference_amount"], "0.00")
        self.assertGreaterEqual(payload["open_vendor_count"], 1)
        self.assertGreaterEqual(payload["stale_advance_count"], 1)
        self.assertIn("top_critical_issues", payload)

    def test_payables_meta_and_rbac_include_control_reports(self):
        meta_response = self.client.get(reverse("reports_api:payables-meta"), self._base_scope())
        self.assertEqual(meta_response.status_code, 200)
        report_codes = {row["code"] for row in meta_response.json()["reports"]}
        self.assertIn("ap_gl_reconciliation", report_codes)
        self.assertIn("vendor_balance_exceptions", report_codes)
        self.assertIn("payables_close_validation", report_codes)
        self.assertIn("payables_close_readiness_summary", report_codes)
        self.assertIn("purchase_register", report_codes)
        self.assertIn("vendor_ledger_statement", report_codes)
        self.assertIn("payables_close_pack", report_codes)

        tree = EffectiveMenuService.menu_tree_for_user(self.user, self.entity.id)

        def flatten(nodes):
            codes = []
            for node in nodes:
                codes.append(node["menu_code"])
                codes.extend(flatten(node["children"]))
            return codes

        def find_node(nodes, menu_code):
            for node in nodes:
                if node["menu_code"] == menu_code:
                    return node
                child = find_node(node["children"], menu_code)
                if child:
                    return child
            return None

        codes = set(flatten(tree))
        self.assertIn("reports.reports.payables", codes)
        self.assertNotIn("reports.apglreconciliation", codes)
        self.assertNotIn("reports.vendorbalanceexceptions", codes)
        self.assertNotIn("reports.vendorledgerstatement", codes)
        self.assertNotIn("reports.payablesclosepack", codes)
        self.assertNotIn("reports.payables.payables_dashboard_summary", codes)
        self.assertNotIn("reports.payables.payables_close_validation", codes)
        self.assertNotIn("reports.payables.payables_close_readiness_summary", codes)
        self.assertIn("reports.payables.upcoming_payments_calendar", codes)
        payables_menu = find_node(tree, "reports.reports.payables")
        self.assertIsNotNone(payables_menu)
        self.assertEqual(payables_menu["route_path"], "/reports/payables")
        child_codes = {child["menu_code"] for child in payables_menu["children"]}
        self.assertIn("reports.accountspayableaging", child_codes)
        self.assertIn("reports.payables.msme_overdue", child_codes)
        self.assertIn("reports.payables.upcoming_payments_calendar", child_codes)

    def test_new_payables_control_export_endpoints_return_expected_formats(self):
        self._post_vendor_gl_balance(Decimal("650.00"), posting_date=date(2025, 4, 30), txn_id=9003, voucher_no="GL-AP-3")
        recon_scope = self._base_scope(as_of_date="2025-04-30")
        exception_scope = self._base_scope(as_of_date="2025-08-01")
        export_checks = [
            ("reports_api:ap-gl-reconciliation-report-csv", recon_scope, "text/csv", b"Vendor Name"),
            ("reports_api:ap-gl-reconciliation-report-excel", recon_scope, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:ap-gl-reconciliation-report-pdf", recon_scope, "application/pdf", b"%PDF"),
            ("reports_api:ap-gl-reconciliation-report-print", recon_scope, "application/pdf", b"%PDF"),
            ("reports_api:vendor-balance-exception-report-csv", exception_scope, "text/csv", b"Vendor Name"),
            ("reports_api:vendor-balance-exception-report-excel", exception_scope, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:vendor-balance-exception-report-pdf", exception_scope, "application/pdf", b"%PDF"),
            ("reports_api:vendor-balance-exception-report-print", exception_scope, "application/pdf", b"%PDF"),
        ]
        for route_name, params, content_type, prefix in export_checks:
            with self.subTest(route_name=route_name):
                response = self.client.get(reverse(route_name), params)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith(content_type))
        self.assertTrue(bytes(response.content).startswith(prefix))


    def _post_vendor_statement_entry(self, *, txn_type, txn_id, posting_date, amount, voucher_no, description):
        batch = PostingBatch.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            created_by=self.user,
        )
        entry = Entry.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            voucher_no=voucher_no,
            voucher_date=posting_date,
            posting_date=posting_date,
            status=EntryStatus.POSTED,
            posted_at=timezone.now(),
            posted_by=self.user,
            posting_batch=batch,
            narration=description,
            created_by=self.user,
        )
        amount = Decimal(amount)
        vendor_drcr = False if amount > 0 else True
        JournalLine.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            voucher_no=voucher_no,
            account=self.vendor,
            ledger=self.vendor_ledger,
            drcr=vendor_drcr,
            amount=abs(amount),
            description=description,
            posting_date=posting_date,
            posted_at=timezone.now(),
            created_by=self.user,
        )
        JournalLine.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            voucher_no=voucher_no,
            accounthead=self.expense_head,
            drcr=not vendor_drcr,
            amount=abs(amount),
            description="Balancing line",
            posting_date=posting_date,
            posted_at=timezone.now(),
            created_by=self.user,
        )
        return entry

    def test_vendor_ledger_statement_resolves_vendor_and_builds_running_balance(self):
        self._post_vendor_statement_entry(txn_type=TxnType.PURCHASE, txn_id=self.invoice.id, posting_date=date(2025, 4, 1), amount="1000.00", voucher_no="PI-PINV-1001", description="Purchase invoice")
        self._post_vendor_statement_entry(txn_type=TxnType.PAYMENT, txn_id=self.payment.id, posting_date=date(2025, 4, 5), amount="-200.00", voucher_no="PAY-001", description="Vendor payment")

        response = self.client.get(
            reverse("reports_api:vendor-ledger-statement"),
            self._base_scope(vendor=self.vendor.id, from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["vendor"]["vendor_id"], self.vendor.id)
        self.assertEqual(payload["opening_balance"], "0.00")
        self.assertEqual(payload["totals"]["closing_balance"], "800.00")
        self.assertEqual(len(payload["rows"]), 2)
        self.assertEqual(payload["rows"][0]["drilldown_target"], "purchase_invoice_detail")
        self.assertIn("vendor_settlements", payload["rows"][0]["_meta"]["drilldown"])

    def test_vendor_ledger_statement_purchase_document_drilldown_uses_service_route_for_service_bills(self):
        PurchaseInvoiceLine.objects.create(
            header=self.invoice,
            line_no=1,
            is_service=True,
            purchase_behavior="expense",
            product_desc="AMC service",
        )
        self._post_vendor_statement_entry(
            txn_type=TxnType.PURCHASE,
            txn_id=self.invoice.id,
            posting_date=date(2025, 4, 1),
            amount="1000.00",
            voucher_no="PI-PINV-1001",
            description="Purchase invoice",
        )

        response = self.client.get(
            reverse("reports_api:vendor-ledger-statement"),
            self._base_scope(vendor=self.vendor.id, from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        drilldown = response.json()["rows"][0]["_meta"]["drilldown"]["source_document"]
        self.assertEqual(drilldown["route"], "/purchaseserviceinvoice")

    def test_vendor_ledger_statement_export_endpoints_return_expected_formats(self):
        self._post_vendor_statement_entry(txn_type=TxnType.PURCHASE, txn_id=self.invoice.id, posting_date=date(2025, 4, 1), amount="1000.00", voucher_no="PI-PINV-1001", description="Purchase invoice")
        params = self._base_scope(vendor=self.vendor.id, from_date="2025-04-01", to_date="2025-04-30")
        export_checks = [
            ("reports_api:vendor-ledger-statement-csv", "text/csv", b"Transaction Date"),
            ("reports_api:vendor-ledger-statement-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:vendor-ledger-statement-pdf", "application/pdf", b"%PDF"),
            ("reports_api:vendor-ledger-statement-print", "application/pdf", b"%PDF"),
        ]
        for route_name, content_type, prefix in export_checks:
            with self.subTest(route_name=route_name):
                response = self.client.get(reverse(route_name), params)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith(content_type))
        self.assertTrue(bytes(response.content).startswith(prefix))

    def test_payables_close_pack_composes_existing_control_sections(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "msme_status": "micro",
                "has_written_payment_terms": False,
            },
        )
        self._post_vendor_gl_balance(Decimal("650.00"), posting_date=date(2025, 8, 1), txn_id=9010, voucher_no="GL-AP-10")
        response = self.client.get(
            reverse("reports_api:payables-close-pack"),
            self._base_scope(as_of_date="2025-08-01"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("overview", payload)
        self.assertIn("aging", payload)
        self.assertIn("reconciliation", payload)
        self.assertIn("validation", payload)
        self.assertEqual(payload["reconciliation"]["status"], "matched")
        self.assertEqual(payload["overview"]["msme_overdue_amount"], "800.00")
        self.assertEqual(payload["overview"]["msme_overdue_bill_count"], 1)
        self.assertEqual(payload["overview"]["msme_overdue_vendor_count"], 1)
        self.assertIn("800.00", payload["overview"]["msme_reporting_note"])
        self.assertIn("top_vendors", payload)
        self.assertEqual(payload["top_vendors"]["top_msme_overdue_vendors"][0]["vendor_name"], "ABC Traders")

    def test_payables_close_pack_top_overdue_vendors_are_not_limited_to_first_outstanding_page(self):
        for index in range(10):
            vendor = self._create_vendor(f"Future Vendor {index + 1}", 5100 + index)
            vendor_ledger = vendor.ledger
            header = self._create_purchase_header(
                vendor=vendor,
                vendor_ledger=vendor_ledger,
                doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
                bill_date=date(2025, 7, 20),
                due_date=date(2025, 9, 15),
                doc_code="PINV",
                doc_no=2000 + index,
                purchase_number=f"PI-FUT-{index + 1}",
                supplier_invoice_number=f"FUT-{index + 1}",
                amount=Decimal(str(1000 - (index * 10))),
            )
            self._create_open_item(
                header=header,
                vendor=vendor,
                vendor_ledger=vendor_ledger,
                doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
                bill_date=date(2025, 7, 20),
                due_date=date(2025, 9, 15),
                purchase_number=f"PI-FUT-{index + 1}",
                supplier_invoice_number=f"FUT-{index + 1}",
                amount=Decimal(str(1000 - (index * 10))),
            )

        overdue_vendor = self._create_vendor("Critical Overdue Vendor", 5200)
        overdue_vendor_ledger = overdue_vendor.ledger
        overdue_header = self._create_purchase_header(
            vendor=overdue_vendor,
            vendor_ledger=overdue_vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 1),
            due_date=date(2025, 7, 15),
            doc_code="PINV",
            doc_no=2999,
            purchase_number="PI-OVD-1",
            supplier_invoice_number="OVD-1",
            amount=Decimal("250.00"),
        )
        self._create_open_item(
            header=overdue_header,
            vendor=overdue_vendor,
            vendor_ledger=overdue_vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 7, 1),
            due_date=date(2025, 7, 15),
            purchase_number="PI-OVD-1",
            supplier_invoice_number="OVD-1",
            amount=Decimal("250.00"),
        )

        response = self.client.get(
            reverse("reports_api:payables-close-pack"),
            self._base_scope(as_of_date="2025-08-01"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        top_overdue = payload["top_vendors"]["top_overdue_vendors"]
        top_outstanding = payload["top_vendors"]["top_outstanding_vendors"]

        self.assertTrue(any(row["vendor_name"] == "Critical Overdue Vendor" for row in top_overdue))
        self.assertFalse(any(row["vendor_name"] == "Critical Overdue Vendor" for row in top_outstanding))

    def test_payables_close_pack_export_endpoints_return_expected_formats(self):
        apply_normalized_profile_payload(
            self.vendor,
            compliance_data={
                "msme_status": "micro",
                "has_written_payment_terms": False,
            },
        )
        self._post_vendor_gl_balance(Decimal("650.00"), posting_date=date(2025, 8, 1), txn_id=9011, voucher_no="GL-AP-11")
        params = self._base_scope(as_of_date="2025-08-01")
        export_checks = [
            ("reports_api:payables-close-pack-csv", "text/csv", b"Section"),
            ("reports_api:payables-close-pack-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:payables-close-pack-pdf", "application/pdf", b"%PDF"),
            ("reports_api:payables-close-pack-print", "application/pdf", b"%PDF"),
        ]
        for route_name, content_type, prefix in export_checks:
            with self.subTest(route_name=route_name):
                response = self.client.get(reverse(route_name), params)
                self.assertEqual(response.status_code, 200)
                self.assertTrue(response["Content-Type"].startswith(content_type))
                self.assertTrue(bytes(response.content).startswith(prefix))
        csv_response = self.client.get(reverse("reports_api:payables-close-pack-csv"), params)
        csv_text = csv_response.content.decode("utf-8-sig")
        self.assertIn("Overview,msme_reporting_note", csv_text)
        self.assertIn("top_msme_overdue_vendors[1].vendor_name", csv_text)
        self.assertIn("ABC Traders", csv_text)

    def test_payables_meta_exposes_centralized_report_definitions(self):
        response = self.client.get(reverse("reports_api:payables-meta"), self._base_scope())
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        definitions = {row["code"]: row for row in payload["report_definitions"]}
        self.assertIn("vendor_outstanding", definitions)
        self.assertIn("msme_overdue", definitions)
        self.assertIn("purchase_register", definitions)
        self.assertIn("payables_close_pack", definitions)
        self.assertIn("vendor_name", definitions["vendor_outstanding"]["enabled_columns"])
        self.assertIn("totals", definitions["vendor_outstanding"]["enabled_summary_blocks"])
        self.assertIn("msme_due_date", definitions["msme_overdue"]["enabled_columns"])
        self.assertIn("reporting_note", definitions["msme_overdue"]["enabled_summary_blocks"])
        self.assertIn("include_outstanding", {flag["code"] for flag in definitions["purchase_register"]["feature_flags"]})

    def test_vendor_outstanding_meta_reflects_config_when_reconciliation_enabled(self):
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", reconcile_gl="true"),
        )
        self.assertEqual(response.status_code, 200)
        meta = response.json()["_meta"]
        self.assertTrue(meta["feature_state"]["reconcile_gl"])
        self.assertIn("outstanding", meta["effective_columns"])
        self.assertIn("vendor_outstanding", [row["code"] for row in meta["related_reports"]])

    def test_vendor_ledger_statement_meta_respects_running_balance_flag(self):
        self._post_vendor_statement_entry(txn_type=TxnType.PURCHASE, txn_id=self.invoice.id, posting_date=date(2025, 4, 1), amount="1000.00", voucher_no="PI-PINV-1001", description="Purchase invoice")
        response = self.client.get(
            reverse("reports_api:vendor-ledger-statement"),
            self._base_scope(vendor=self.vendor.id, from_date="2025-04-01", to_date="2025-04-30", include_running_balance="false"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertNotIn("running_balance", payload["rows"][0])
        self.assertNotIn("running_balance", payload["_meta"]["effective_columns"])

    def test_payables_close_pack_section_order_is_config_driven(self):
        self._post_vendor_gl_balance(Decimal("650.00"), posting_date=date(2025, 8, 1), txn_id=9012, voucher_no="GL-AP-12")
        response = self.client.get(
            reverse("reports_api:payables-close-pack"),
            self._base_scope(
                as_of_date="2025-08-01",
                include_exceptions="false",
                include_top_vendors="false",
            ),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["included_sections"], ["overview", "aging", "reconciliation", "validation"])
        self.assertEqual(payload["section_order"], ["overview", "aging", "reconciliation", "validation"])
        self.assertEqual(payload["_meta"]["enabled_summary_blocks"], ["overview", "aging", "reconciliation", "validation"])

    def test_vendor_ledger_csv_header_tracks_running_balance_config(self):
        self._post_vendor_statement_entry(txn_type=TxnType.PURCHASE, txn_id=self.invoice.id, posting_date=date(2025, 4, 1), amount="1000.00", voucher_no="PI-PINV-1001", description="Purchase invoice")
        response = self.client.get(
            reverse("reports_api:vendor-ledger-statement-csv"),
            self._base_scope(vendor=self.vendor.id, from_date="2025-04-01", to_date="2025-04-30", include_running_balance="false"),
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("Running Balance", response.content.decode("utf-8").splitlines()[0])


    def test_vendor_settlement_history_report_totals_and_trace(self):
        response = self.client.get(
            reverse("reports_api:vendor-settlement-history"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "vendor_settlement_history")
        self.assertEqual(payload["totals"]["total_settled"], "200.00")
        self.assertEqual(payload["totals"]["total_unapplied"], "0.00")
        self.assertEqual(payload["summary"]["settlement_count"], 1)
        self.assertEqual(len(payload["rows"]), 1)
        row = payload["rows"][0]
        self.assertEqual(row["settlement_number"], "PAY-001")
        self.assertEqual(row["applied_amount"], "200.00")
        self.assertIn("settlement_detail", row["_meta"]["drilldown"])
        self.assertEqual(row["_trace"]["source_model"], "purchase.VendorSettlement")
        expected_line = VendorSettlementLine.objects.get(settlement=self.payment, open_item=self.invoice_item)
        self.assertEqual(row["_trace"]["settlement_line_id"], expected_line.id)

    def test_vendor_settlement_history_vendor_filter_and_unapplied_toggle(self):
        unapplied = VendorSettlement.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            settlement_type=VendorSettlement.SettlementType.MANUAL,
            settlement_date=date(2025, 4, 12),
            reference_no="SET-UNAPPLIED",
            total_amount=Decimal("25.00"),
            status=VendorSettlement.Status.POSTED,
            posted_at=timezone.now(),
            posted_by=self.user,
        )
        other_ledger = self.other_vendor.ledger
        other_header = self._create_purchase_header(
            vendor=self.other_vendor,
            vendor_ledger=other_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 2),
            due_date=date(2025, 4, 20),
            doc_code="PINV",
            doc_no=2001,
            purchase_number="PI-PINV-2001",
            supplier_invoice_number="SUP-2001",
            amount=Decimal("300.00"),
        )
        other_open = self._create_open_item(
            header=other_header,
            vendor=self.other_vendor,
            vendor_ledger=other_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 2),
            due_date=date(2025, 4, 20),
            purchase_number="PI-PINV-2001",
            supplier_invoice_number="SUP-2001",
            amount=Decimal("300.00"),
        )
        other_settlement = VendorSettlement.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=self.other_vendor,
            vendor_ledger=other_ledger,
            settlement_type=VendorSettlement.SettlementType.PAYMENT,
            settlement_date=date(2025, 4, 14),
            reference_no="PAY-OTHER",
            total_amount=Decimal("75.00"),
            status=VendorSettlement.Status.POSTED,
            posted_at=timezone.now(),
            posted_by=self.user,
        )
        VendorSettlementLine.objects.create(
            settlement=other_settlement,
            open_item=other_open,
            amount=Decimal("75.00"),
            applied_amount_signed=Decimal("75.00"),
        )

        filtered = self.client.get(
            reverse("reports_api:vendor-settlement-history"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", vendor=self.vendor.id, include_unapplied="false"),
        )
        self.assertEqual(filtered.status_code, 200)
        rows = filtered.json()["rows"]
        self.assertEqual({row["vendor_id"] for row in rows}, {self.vendor.id})
        self.assertNotIn("SET-UNAPPLIED", [row["settlement_number"] for row in rows])
        self.assertNotIn(unapplied.id, [row["settlement_id"] for row in rows])

    def test_vendor_note_register_totals_type_split_and_trace(self):
        debit_note = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.DEBIT_NOTE,
            bill_date=date(2025, 4, 18),
            due_date=date(2025, 4, 18),
            doc_code="PDN",
            doc_no=1004,
            purchase_number="PI-PDN-1004",
            supplier_invoice_number="SUP-DN-001",
            amount=Decimal("60.00"),
            ref_document=self.invoice,
        )
        self._create_open_item(
            header=debit_note,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.DEBIT_NOTE,
            bill_date=date(2025, 4, 18),
            due_date=date(2025, 4, 18),
            purchase_number="PI-PDN-1004",
            supplier_invoice_number="SUP-DN-001",
            amount=Decimal("60.00"),
        )
        response = self.client.get(
            reverse("reports_api:vendor-note-register"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_code"], "vendor_note_register")
        self.assertEqual(payload["totals"]["credit_note_total"], "100.00")
        self.assertEqual(payload["totals"]["debit_note_total"], "60.00")
        self.assertEqual(payload["totals"]["net_note_total"], "-40.00")
        note_types = {row["note_type"] for row in payload["rows"]}
        self.assertEqual(note_types, {"credit", "debit"})
        self.assertTrue(all("_trace" in row for row in payload["rows"]))

    def test_vendor_note_register_filters_by_type(self):
        response = self.client.get(
            reverse("reports_api:vendor-note-register"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", note_type="credit"),
        )
        self.assertEqual(response.status_code, 200)
        rows = response.json()["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["note_type"], "credit")
        self.assertEqual(rows[0]["_trace"]["source_model"], "purchase.PurchaseInvoiceHeader")

    def test_cancelled_purchase_documents_are_excluded_from_outstanding_default(self):
        cancelled_invoice = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 28),
            doc_code="PINV",
            doc_no=1999,
            purchase_number="PI-PINV-1999",
            supplier_invoice_number="SUP-CANC-001",
            amount=Decimal("500.00"),
            status=PurchaseInvoiceHeader.Status.CANCELLED,
        )
        self._create_open_item(
            header=cancelled_invoice,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            bill_date=date(2025, 4, 22),
            due_date=date(2025, 4, 28),
            purchase_number="PI-PINV-1999",
            supplier_invoice_number="SUP-CANC-001",
            amount=Decimal("500.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", vendor=self.vendor.id),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["rows"][0]
        self.assertEqual(Decimal(row["outstanding"]), Decimal("750.00"))
        self.assertEqual(Decimal(row["bill_amount"]), Decimal("1000.00"))

    def test_cancelled_settlements_are_excluded_from_settlement_history(self):
        cancelled_settlement = VendorSettlement.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            settlement_type=VendorSettlement.SettlementType.PAYMENT,
            settlement_date=date(2025, 4, 23),
            reference_no="PAY-CANCELLED-001",
            total_amount=Decimal("10.00"),
            status=VendorSettlement.Status.CANCELLED,
            posted_by=self.user,
        )
        VendorSettlementLine.objects.create(
            settlement=cancelled_settlement,
            open_item=self.invoice_item,
            amount=Decimal("10.00"),
            applied_amount_signed=Decimal("10.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-settlement-history"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", vendor=self.vendor.id),
        )
        self.assertEqual(response.status_code, 200)
        settlement_ids = {row["settlement_id"] for row in response.json()["rows"]}
        self.assertNotIn(cancelled_settlement.id, settlement_ids)

    def test_cancelled_notes_are_excluded_from_note_register_default(self):
        cancelled_note = self._create_purchase_header(
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.DEBIT_NOTE,
            bill_date=date(2025, 4, 24),
            due_date=date(2025, 4, 24),
            doc_code="PDN",
            doc_no=1888,
            purchase_number="PI-PDN-1888",
            supplier_invoice_number="SUP-CANC-DN-001",
            amount=Decimal("75.00"),
            ref_document=self.invoice,
            status=PurchaseInvoiceHeader.Status.CANCELLED,
        )
        self._create_open_item(
            header=cancelled_note,
            vendor=self.vendor,
            vendor_ledger=self.vendor_ledger,
            doc_type=PurchaseInvoiceHeader.DocType.DEBIT_NOTE,
            bill_date=date(2025, 4, 24),
            due_date=date(2025, 4, 24),
            purchase_number="PI-PDN-1888",
            supplier_invoice_number="SUP-CANC-DN-001",
            amount=Decimal("75.00"),
        )

        response = self.client.get(
            reverse("reports_api:vendor-note-register"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30", vendor=self.vendor.id),
        )
        self.assertEqual(response.status_code, 200)
        note_numbers = {row["note_number"] for row in response.json()["rows"]}
        self.assertNotIn("PI-PDN-1888", note_numbers)

    def test_trace_metadata_is_present_on_supported_payables_rows(self):
        outstanding = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(outstanding.status_code, 200)
        self.assertIn("_trace", outstanding.json()["rows"][0])

        aging = self.client.get(
            reverse("reports_api:ap-aging-report"),
            self._base_scope(as_of_date="2025-04-30", view="invoice"),
        )
        self.assertEqual(aging.status_code, 200)
        self.assertEqual(aging.json()["rows"][0]["_trace"]["source_model"], "purchase.VendorBillOpenItem")

    def test_payables_meta_exposes_settlement_and_note_reports(self):
        response = self.client.get(reverse("reports_api:payables-meta"), self._base_scope())
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        definitions = {row["code"]: row for row in payload["report_definitions"]}
        self.assertIn("vendor_settlement_history", definitions)
        self.assertIn("vendor_note_register", definitions)
        self.assertTrue(definitions["vendor_settlement_history"]["supports_traceability"])
        self.assertIn("include_unapplied", {flag["code"] for flag in definitions["vendor_settlement_history"]["feature_flags"]})
        self.assertIn("settlement_types", payload["choices"])
        self.assertIn("note_types", payload["choices"])

    def test_payables_rbac_tree_includes_audit_report_menus(self):
        tree = EffectiveMenuService.menu_tree_for_user(self.user, self.entity.id)

        def flatten(nodes):
            codes = []
            for node in nodes:
                codes.append(node["menu_code"])
                codes.extend(flatten(node["children"]))
            return codes

        codes = set(flatten(tree))
        self.assertIn("reports.reports.payables", codes)
        self.assertNotIn("reports.vendorsettlementhistory", codes)
        self.assertNotIn("reports.vendornoteregister", codes)

    def test_new_payables_audit_export_endpoints_return_expected_formats(self):
        params = self._base_scope(from_date="2025-04-01", to_date="2025-04-30")
        export_checks = [
            ("reports_api:vendor-settlement-history-csv", "text/csv", b"Settlement Number"),
            ("reports_api:vendor-settlement-history-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:vendor-settlement-history-pdf", "application/pdf", b"%PDF"),
            ("reports_api:vendor-settlement-history-print", "application/pdf", b"%PDF"),
            ("reports_api:vendor-note-register-csv", "text/csv", b"Note Number"),
            ("reports_api:vendor-note-register-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", b"PK"),
            ("reports_api:vendor-note-register-pdf", "application/pdf", b"%PDF"),
            ("reports_api:vendor-note-register-print", "application/pdf", b"%PDF"),
        ]
        for route_name, content_type, prefix in export_checks:
            with self.subTest(route_name=route_name):
                response = self.client.get(reverse(route_name), params)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith(content_type))
        self.assertTrue(bytes(response.content).startswith(prefix))


    def test_vendor_outstanding_accepts_date_aliases(self):
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(date_from="2025-04-01", date_to="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["applied_filters"]["from_date"], "2025-04-01")
        self.assertEqual(payload["applied_filters"]["to_date"], "2025-04-30")

    def test_payables_response_envelope_is_standardized(self):
        response = self.client.get(
            reverse("reports_api:vendor-outstanding-report"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("applied_filters", payload)
        self.assertIn("available_exports", payload)
        self.assertIn("available_drilldowns", payload)
        self.assertIn("summary", payload)
        self.assertIn("pagination", payload)
        self.assertEqual(payload["available_exports"], ["excel", "pdf", "csv", "print"])
        self.assertEqual(payload["_meta"]["pagination_mode"], "paged")

    def test_drilldown_payloads_follow_consistent_contract(self):
        response = self.client.get(
            reverse("reports_api:vendor-note-register"),
            self._base_scope(from_date="2025-04-01", to_date="2025-04-30"),
        )
        self.assertEqual(response.status_code, 200)
        drilldown = response.json()["rows"][0]["_meta"]["drilldown"]["document"]
        self.assertIn("target", drilldown)
        self.assertIn("label", drilldown)
        self.assertIn("params", drilldown)
        self.assertIn("kind", drilldown)
        self.assertEqual(drilldown["kind"], "document")

    def test_payables_meta_definitions_are_frontend_complete(self):
        response = self.client.get(reverse("reports_api:payables-meta"), self._base_scope())
        self.assertEqual(response.status_code, 200)
        definitions = {row["code"]: row for row in response.json()["report_definitions"]}
        for code in ["vendor_outstanding", "ap_aging", "msme_overdue", "vendor_settlement_history", "vendor_note_register", "vendor_ledger_statement", "payables_close_pack", "upcoming_payments_calendar"]:
            with self.subTest(code=code):
                definition = definitions[code]
                self.assertIn("supported_filters", definition)
        self.assertIn("pagination_mode", definition)
        self.assertIn("export_formats", definition)
        self.assertIn("drilldown_targets", definition)
        self.assertIn("related_reports", definition)
        self.assertIn("supports_traceability", definition)

    def test_payables_api_guide_mentions_frontend_reports(self):
        doc = Path("reports/PAYABLES_REPORTING_API.md").read_text()
        self.assertIn("Vendor Settlement History", doc)
        self.assertIn("Vendor Debit/Credit Note Register", doc)
        self.assertIn("/api/reports/payables/meta/", doc)
