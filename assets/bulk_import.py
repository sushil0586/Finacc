from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db.models import Q
from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook

from assets.models import AssetBulkJob, AssetCategory, FixedAsset
from assets.serializers import AssetCategorySerializer, FixedAssetWriteSerializer
from assets.services.asset_service import AssetService
from financial.models import Ledger, account
from entity.models import Entity, EntityFinancialYear, SubEntity

CATEGORY_SHEET = "asset_categories"
ASSET_SHEET = "fixed_assets"


@dataclass
class ImportResult:
    summary: dict[str, Any]
    errors: list[dict[str, Any]]


def _normalize_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _is_placeholder(value: str) -> bool:
    return bool(re.fullmatch(r"val-\d+", value, flags=re.IGNORECASE))


def _to_int(value: Any) -> int | None:
    text = _normalize_text(value)
    if not text or text in {"-", "--"} or _is_placeholder(text):
        return None
    return int(float(text))


def _to_decimal(value: Any, default: Decimal | None = None) -> Decimal | None:
    text = _normalize_text(value)
    if not text or text in {"-", "--"} or _is_placeholder(text):
        return default
    text = text.replace(",", "")
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def _to_bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return _normalize_text(value).lower() in {"1", "true", "yes", "y"}


def _to_date_string(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _normalize_text(value)
    if not text or _is_placeholder(text):
        return None

    parsed = parse_date(text)
    if parsed:
        return parsed.isoformat()

    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _read_xlsx(content: bytes, scope_type: str) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    lookup = {str(name).strip().lower(): name for name in wb.sheetnames}
    sheet_name = CATEGORY_SHEET if scope_type == AssetBulkJob.ScopeType.CATEGORY else ASSET_SHEET
    actual = lookup.get(sheet_name)
    if not actual:
        return {sheet_name: []}
    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {sheet_name: []}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    payload: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        payload.append({headers[i]: row[i] for i in range(len(headers))})
    return {sheet_name: payload}


def _read_csv_zip(content: bytes, scope_type: str) -> dict[str, list[dict[str, Any]]]:
    sheet_name = CATEGORY_SHEET if scope_type == AssetBulkJob.ScopeType.CATEGORY else ASSET_SHEET
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        lookup = {str(name).strip().lower(): name for name in zf.namelist()}
        fname = lookup.get(f"{sheet_name}.csv")
        if not fname:
            return {sheet_name: []}
        raw = zf.read(fname).decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw))
        return {sheet_name: [dict(r) for r in reader if any((v or "").strip() for v in r.values())]}


def _write_xlsx(data: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = Workbook()
    for idx, (sheet_name, rows) in enumerate(data.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name[:31] or f"sheet_{idx + 1}"
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(col) for col in headers])
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _write_csv_zip(data: dict[str, list[dict[str, Any]]]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        for sheet_name, rows in data.items():
            stream = io.StringIO()
            headers = list(rows[0].keys()) if rows else []
            writer = csv.DictWriter(stream, fieldnames=headers)
            if headers:
                writer.writeheader()
                writer.writerows(rows)
            zf.writestr(f"{sheet_name}.csv", stream.getvalue())
    return buff.getvalue()


def parse_payload(file_bytes: bytes, fmt: str, scope_type: str) -> dict[str, list[dict[str, Any]]]:
    if fmt == "xlsx":
        return _read_xlsx(file_bytes, scope_type)
    return _read_csv_zip(file_bytes, scope_type)


def render_payload(payload: dict[str, list[dict[str, Any]]], fmt: str) -> bytes:
    if fmt == "xlsx":
        return _write_xlsx(payload)
    return _write_csv_zip(payload)


def template_payload(scope_type: str) -> dict[str, list[dict[str, Any]]]:
    if scope_type == AssetBulkJob.ScopeType.CATEGORY:
        return {
            CATEGORY_SHEET: [
                {
                    "code": "COMP-LAPTOP",
                    "name": "Computer / Laptop",
                    "nature": "TANGIBLE",
                    "depreciation_method": "SLM",
                    "useful_life_months": 36,
                    "residual_value_percent": "5.0000",
                    "capitalization_threshold": "5000.00",
                    "asset_ledger": "",
                    "accumulated_depreciation_ledger": "",
                    "depreciation_expense_ledger": "",
                    "impairment_expense_ledger": "",
                    "impairment_reserve_ledger": "",
                    "cwip_ledger": "",
                    "gain_on_sale_ledger": "",
                    "loss_on_sale_ledger": "",
                    "subentity": "",
                }
            ]
        }

    return {
        ASSET_SHEET: [
            {
                "asset_code": "FA-000001",
                "asset_name": "Laptop - Dell Latitude",
                "category": "COMP-LAPTOP",
                "ledger": "",
                "entityfinid": "",
                "subentity": "",
                "asset_tag": "TAG-1001",
                "serial_number": "SN12345",
                "manufacturer": "Dell",
                "model_number": "Latitude 5520",
                "status": "DRAFT",
                "acquisition_date": "2026-04-01",
                "capitalization_date": "",
                "put_to_use_date": "",
                "depreciation_start_date": "",
                "disposal_date": "",
                "quantity": "1",
                "gross_block": "65000.00",
                "residual_value": "3250.00",
                "useful_life_months": "36",
                "depreciation_method": "SLM",
                "depreciation_rate": "0.00",
                "location_name": "Head Office",
                "department_name": "IT",
                "custodian_name": "Admin User",
                "vendor_account": "",
                "purchase_document_no": "PO-001",
                "external_reference": "LEGACY-FA-1",
                "notes": "Imported opening asset master row",
            }
        ]
    }


def _flatten_errors(errors: Any, prefix: str = "") -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    if isinstance(errors, dict):
        for key, value in errors.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            out.extend(_flatten_errors(value, next_prefix))
    elif isinstance(errors, list):
        for item in errors:
            if isinstance(item, (dict, list)):
                out.extend(_flatten_errors(item, prefix))
            else:
                out.append((prefix or "non_field_errors", str(item)))
    else:
        out.append((prefix or "non_field_errors", str(errors)))
    return out


def _resolve_subentity(entity: Entity, raw: Any, fallback: SubEntity | None = None) -> SubEntity | None:
    value = raw or fallback
    if value in (None, "", "null", "None"):
        return None
    if isinstance(value, SubEntity):
        return value if value.entity_id == entity.id else None
    try:
        subentity_id = int(value)
    except (TypeError, ValueError):
        text = _normalize_text(value)
        if not text:
            return None
        return SubEntity.objects.filter(entity=entity, subentityname__iexact=text).first()
    return SubEntity.objects.filter(entity=entity, id=subentity_id).first()


def _resolve_ledger(entity: Entity, raw: Any) -> Ledger | None:
    if raw in (None, "", "null", "None"):
        return None
    if isinstance(raw, Ledger):
        return raw if raw.entity_id == entity.id else None
    try:
        ledger_id = int(raw)
    except (TypeError, ValueError):
        ledger_id = None
    if ledger_id:
        ledger = Ledger.objects.filter(entity=entity, id=ledger_id).first()
        if ledger:
            return ledger
    try:
        ledger_code = int(str(raw).strip())
    except (TypeError, ValueError):
        ledger_code = None
    if ledger_code:
        ledger = Ledger.objects.filter(entity=entity, ledger_code=ledger_code).first()
        if ledger:
            return ledger
    text = _normalize_text(raw)
    if not text:
        return None
    return Ledger.objects.filter(entity=entity, name__iexact=text).first()


def _resolve_account(entity: Entity, raw: Any):
    if raw in (None, "", "null", "None"):
        return None
    if isinstance(raw, account):
        return raw if raw.entity_id == entity.id else None
    try:
        account_id = int(raw)
    except (TypeError, ValueError):
        account_id = None
    if account_id:
        acc = account.objects.filter(entity=entity, id=account_id).first()
        if acc:
            return acc
    text = _normalize_text(raw)
    if not text:
        return None
    return account.objects.filter(entity=entity, accountname__iexact=text).first()


def _resolve_category(entity: Entity, raw: Any, subentity: SubEntity | None = None) -> AssetCategory | None:
    if raw in (None, "", "null", "None"):
        return None
    if isinstance(raw, AssetCategory):
        return raw if raw.entity_id == entity.id else None
    try:
        category_id = int(raw)
    except (TypeError, ValueError):
        category_id = None
    if category_id:
        qs = AssetCategory.objects.filter(entity=entity, id=category_id)
        if subentity is not None:
            qs = qs.filter(Q(subentity=subentity) | Q(subentity__isnull=True))
        category = qs.first()
        if category:
            return category
    text = _normalize_text(raw)
    if not text:
        return None
    qs = AssetCategory.objects.filter(entity=entity)
    if subentity is not None:
        qs = qs.filter(Q(subentity=subentity) | Q(subentity__isnull=True))
    category = qs.filter(code__iexact=text).first()
    if category:
        return category
    return qs.filter(name__iexact=text).first()


def _existing_category(entity: Entity, row: dict[str, Any], subentity: SubEntity | None) -> AssetCategory | None:
    code = _normalize_text(row.get("code"))
    name = _normalize_text(row.get("name"))
    qs = AssetCategory.objects.filter(entity=entity)
    if subentity is not None:
        qs = qs.filter(Q(subentity=subentity) | Q(subentity__isnull=True))
    if code:
        existing = qs.filter(code__iexact=code).first()
        if existing:
            return existing
    if name:
        return qs.filter(name__iexact=name).first()
    return None


def _existing_asset(entity: Entity, row: dict[str, Any], subentity: SubEntity | None) -> FixedAsset | None:
    asset_code = _normalize_text(row.get("asset_code"))
    asset_tag = _normalize_text(row.get("asset_tag"))
    qs = FixedAsset.objects.filter(entity=entity)
    if subentity is not None:
        qs = qs.filter(Q(subentity=subentity) | Q(subentity__isnull=True))
    if asset_code:
        existing = qs.filter(asset_code__iexact=asset_code).first()
        if existing:
            return existing
    if asset_tag:
        return qs.filter(asset_tag__iexact=asset_tag).first()
    return None


def _build_category_payload(entity: Entity, row: dict[str, Any], fallback_subentity: SubEntity | None) -> dict[str, Any]:
    subentity = _resolve_subentity(entity, row.get("subentity"), fallback_subentity)
    payload = {
        "entity": entity.id,
        "subentity": getattr(subentity, "id", None),
        "code": _normalize_text(row.get("code")),
        "name": _normalize_text(row.get("name")),
        "nature": _normalize_text(row.get("nature")) or AssetCategory.AssetNature.TANGIBLE,
        "depreciation_method": _normalize_text(row.get("depreciation_method")) or AssetCategory.DepreciationMethod.SLM,
        "useful_life_months": _to_int(row.get("useful_life_months")) or 0,
        "residual_value_percent": _to_decimal(row.get("residual_value_percent"), Decimal("0.0000")),
        "capitalization_threshold": _to_decimal(row.get("capitalization_threshold"), Decimal("0.00")),
        "asset_ledger": getattr(_resolve_ledger(entity, row.get("asset_ledger")), "id", None),
        "accumulated_depreciation_ledger": getattr(_resolve_ledger(entity, row.get("accumulated_depreciation_ledger")), "id", None),
        "depreciation_expense_ledger": getattr(_resolve_ledger(entity, row.get("depreciation_expense_ledger")), "id", None),
        "impairment_expense_ledger": getattr(_resolve_ledger(entity, row.get("impairment_expense_ledger")), "id", None),
        "impairment_reserve_ledger": getattr(_resolve_ledger(entity, row.get("impairment_reserve_ledger")), "id", None),
        "cwip_ledger": getattr(_resolve_ledger(entity, row.get("cwip_ledger")), "id", None),
        "gain_on_sale_ledger": getattr(_resolve_ledger(entity, row.get("gain_on_sale_ledger")), "id", None),
        "loss_on_sale_ledger": getattr(_resolve_ledger(entity, row.get("loss_on_sale_ledger")), "id", None),
    }
    return payload


def _build_asset_payload(
    entity: Entity,
    row: dict[str, Any],
    fallback_subentity: SubEntity | None,
    fallback_entityfinid: EntityFinancialYear | None,
) -> dict[str, Any]:
    subentity = _resolve_subentity(entity, row.get("subentity"), fallback_subentity)
    entityfinid = row.get("entityfinid") or fallback_entityfinid
    category = _resolve_category(entity, row.get("category") or row.get("category_code") or row.get("category_name"), subentity=subentity)
    ledger = _resolve_ledger(entity, row.get("ledger") or row.get("ledger_code") or row.get("ledger_name"))
    vendor_account = _resolve_account(entity, row.get("vendor_account") or row.get("vendor_account_name"))
    payload = {
        "entity": entity.id,
        "entityfinid": getattr(entityfinid, "id", None) if isinstance(entityfinid, EntityFinancialYear) else _to_int(entityfinid),
        "subentity": getattr(subentity, "id", None),
        "category": getattr(category, "id", None),
        "ledger": getattr(ledger, "id", None),
        "asset_code": _normalize_text(row.get("asset_code")),
        "asset_name": _normalize_text(row.get("asset_name")),
        "asset_tag": _normalize_text(row.get("asset_tag")) or None,
        "serial_number": _normalize_text(row.get("serial_number")) or None,
        "manufacturer": _normalize_text(row.get("manufacturer")) or None,
        "model_number": _normalize_text(row.get("model_number")) or None,
        "status": _normalize_text(row.get("status")) or FixedAsset.AssetStatus.DRAFT,
        "acquisition_date": _to_date_string(row.get("acquisition_date")),
        "capitalization_date": _to_date_string(row.get("capitalization_date")),
        "put_to_use_date": _to_date_string(row.get("put_to_use_date")),
        "depreciation_start_date": _to_date_string(row.get("depreciation_start_date")),
        "disposal_date": _to_date_string(row.get("disposal_date")),
        "quantity": _to_decimal(row.get("quantity"), Decimal("1.0000")),
        "gross_block": _to_decimal(row.get("gross_block"), Decimal("0.00")),
        "residual_value": _to_decimal(row.get("residual_value"), Decimal("0.00")),
        "useful_life_months": _to_int(row.get("useful_life_months")) or 0,
        "depreciation_method": _normalize_text(row.get("depreciation_method")) or FixedAsset.DepreciationMethod.SLM,
        "depreciation_rate": _to_decimal(row.get("depreciation_rate"), Decimal("0.0000")),
        "location_name": _normalize_text(row.get("location_name")) or None,
        "department_name": _normalize_text(row.get("department_name")) or None,
        "custodian_name": _normalize_text(row.get("custodian_name")) or None,
        "vendor_account": getattr(vendor_account, "id", None),
        "purchase_document_no": _normalize_text(row.get("purchase_document_no")) or None,
        "external_reference": _normalize_text(row.get("external_reference")) or None,
        "notes": _normalize_text(row.get("notes")) or None,
    }
    return payload


def _category_serializer(existing: AssetCategory | None, payload: dict[str, Any]):
    return AssetCategorySerializer(existing, data=payload, partial=bool(existing))


def _asset_serializer(existing: FixedAsset | None, payload: dict[str, Any]):
    return FixedAssetWriteSerializer(existing, data=payload, partial=bool(existing))


def validate_payload(
    payload: dict[str, list[dict[str, Any]]],
    entity: Entity,
    scope_type: str,
    *,
    subentity: SubEntity | None = None,
    entityfinid: EntityFinancialYear | None = None,
    upsert_mode: str = AssetBulkJob.UpsertMode.UPSERT,
) -> ImportResult:
    errors: list[dict[str, Any]] = []
    sheet_name = CATEGORY_SHEET if scope_type == AssetBulkJob.ScopeType.CATEGORY else ASSET_SHEET
    rows = payload.get(sheet_name, [])
    if not rows:
        errors.append({"sheet": sheet_name, "row": 0, "field": "file", "message": "No rows found in upload."})
        return ImportResult(summary={"rows": {sheet_name: 0}, "error_count": 1}, errors=errors)

    for idx, row in enumerate(rows, start=2):
        existing = _existing_category(entity, row, subentity) if scope_type == AssetBulkJob.ScopeType.CATEGORY else _existing_asset(entity, row, subentity)
        if existing and upsert_mode == AssetBulkJob.UpsertMode.CREATE_ONLY:
            continue
        if not existing and upsert_mode == AssetBulkJob.UpsertMode.UPDATE_ONLY:
            continue

        try:
            serializer_payload = _build_category_payload(entity, row, subentity) if scope_type == AssetBulkJob.ScopeType.CATEGORY else _build_asset_payload(entity, row, subentity, entityfinid)
        except Exception as exc:
            errors.append({"sheet": sheet_name, "row": idx, "field": "row", "message": str(exc)})
            continue

        serializer = _category_serializer(existing, serializer_payload) if scope_type == AssetBulkJob.ScopeType.CATEGORY else _asset_serializer(existing, serializer_payload)
        if not serializer.is_valid():
            for field, message in _flatten_errors(serializer.errors):
                errors.append({"sheet": sheet_name, "row": idx, "field": field, "message": message})

    return ImportResult(summary={"rows": {sheet_name: len(rows)}, "error_count": len(errors)}, errors=errors)


def commit_payload(
    payload: dict[str, list[dict[str, Any]]],
    entity: Entity,
    scope_type: str,
    *,
    subentity: SubEntity | None = None,
    entityfinid: EntityFinancialYear | None = None,
    upsert_mode: str = AssetBulkJob.UpsertMode.UPSERT,
    duplicate_strategy: str = AssetBulkJob.DuplicateStrategy.FAIL,
    user_id: int | None = None,
) -> ImportResult:
    sheet_name = CATEGORY_SHEET if scope_type == AssetBulkJob.ScopeType.CATEGORY else ASSET_SHEET
    rows = payload.get(sheet_name, [])
    errors: list[dict[str, Any]] = []
    summary = {"created": 0, "updated": 0, "skipped": 0}

    for idx, row in enumerate(rows, start=2):
        existing = _existing_category(entity, row, subentity) if scope_type == AssetBulkJob.ScopeType.CATEGORY else _existing_asset(entity, row, subentity)
        if existing and upsert_mode == AssetBulkJob.UpsertMode.CREATE_ONLY:
            summary["skipped"] += 1
            continue
        if not existing and upsert_mode == AssetBulkJob.UpsertMode.UPDATE_ONLY:
            summary["skipped"] += 1
            continue

        try:
            serializer_payload = _build_category_payload(entity, row, subentity) if scope_type == AssetBulkJob.ScopeType.CATEGORY else _build_asset_payload(entity, row, subentity, entityfinid)
            serializer = _category_serializer(existing, serializer_payload) if scope_type == AssetBulkJob.ScopeType.CATEGORY else _asset_serializer(existing, serializer_payload)
            serializer.is_valid(raise_exception=True)
            if scope_type == AssetBulkJob.ScopeType.CATEGORY:
                serializer.save(created_by_id=user_id, updated_by_id=user_id)
            else:
                if existing:
                    AssetService.update_asset(instance=existing, data=serializer.validated_data, user_id=user_id)
                else:
                    AssetService.create_asset(data=serializer.validated_data, user_id=user_id)
            if existing:
                summary["updated"] += 1
            else:
                summary["created"] += 1
        except Exception as exc:
            errors.append({"sheet": sheet_name, "row": idx, "field": "row", "message": str(exc)})
            if duplicate_strategy == AssetBulkJob.DuplicateStrategy.FAIL:
                break
            summary["skipped"] += 1

    summary["error_count"] = len(errors)
    return ImportResult(summary=summary, errors=errors)

