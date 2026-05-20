from __future__ import annotations

import hashlib
import io
import json
from dataclasses import dataclass
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from gst_reconciliation.models import GstImportedReturn, GstImportedReturnRow, GstReconciliationActionLog, GstReconciliationItem, GstReconciliationRun
from gst_reconciliation.services.importing.base import BaseImportedReturnPipeline, ImportPipelineResult
from gst_reconciliation.services.item_workflow_service import GstReconciliationItemWorkflowService
from gst_reconciliation.services.normalization import (
    normalize_doc_type,
    normalize_gstin,
    normalize_invoice_number,
    normalize_return_period,
    parse_date_value,
    parse_decimal,
)


@dataclass(frozen=True)
class NormalizedGstr2bRow:
    row_no: int
    source_section: str | None
    source_row_reference: str | None
    doc_type_code: str
    counterparty_gstin: str | None
    counterparty_gstin_normalized: str | None
    counterparty_name: str | None
    invoice_number: str | None
    invoice_number_normalized: str | None
    invoice_date: Any
    taxable_value: Any
    cgst: Any
    sgst: Any
    igst: Any
    cess: Any
    total_amount: Any
    pos_state_name: str | None
    raw_row_json: dict[str, Any]

    def normalized_payload(self) -> dict[str, Any]:
        return {
            "row_no": self.row_no,
            "source_section": self.source_section,
            "source_row_reference": self.source_row_reference,
            "doc_type_code": self.doc_type_code,
            "counterparty_gstin": self.counterparty_gstin,
            "counterparty_gstin_normalized": self.counterparty_gstin_normalized,
            "counterparty_name": self.counterparty_name,
            "invoice_number": self.invoice_number,
            "invoice_number_normalized": self.invoice_number_normalized,
            "invoice_date": self.invoice_date.isoformat() if self.invoice_date else None,
            "taxable_value": str(self.taxable_value),
            "cgst": str(self.cgst),
            "sgst": str(self.sgst),
            "igst": str(self.igst),
            "cess": str(self.cess),
            "total_amount": str(self.total_amount),
            "pos_state_name": self.pos_state_name,
        }


def _safe_str(value: Any) -> str | None:
    token = str(value or "").strip()
    return token or None


def _row_hash(payload: dict[str, Any]) -> str:
    body = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(body).hexdigest()


def _json_checksum(payload: Any) -> str:
    body = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(body).hexdigest()


def _file_checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _extract_candidate_rows(payload: Any) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    key_markers = {
        "supplier_gstin",
        "gstin",
        "supplier_invoice_number",
        "invoice_number",
        "invoice_no",
        "taxable_value",
        "taxable",
    }

    def walk(node: Any) -> None:
        if isinstance(node, list):
            for item in node:
                walk(item)
            return
        if isinstance(node, dict):
            lowered = {str(key).strip().lower(): value for key, value in node.items()}
            if key_markers.intersection(lowered.keys()):
                candidates.append(node)
            for value in node.values():
                walk(value)

    walk(payload)
    return candidates


def normalize_gstr2b_row(row_no: int, payload: dict[str, Any], *, source_section: str | None = None) -> NormalizedGstr2bRow:
    lowered = {str(key).strip().lower(): value for key, value in payload.items()}

    def pick(*keys: str) -> Any:
        for key in keys:
            if key in lowered:
                return lowered[key]
        return None

    counterparty_gstin = _safe_str(pick("supplier_gstin", "gstin", "supplier gstin", "vendor gstin"))
    invoice_number = _safe_str(pick("supplier_invoice_number", "invoice_number", "invoice no", "invoice_no"))
    taxable_value = parse_decimal(pick("taxable_value", "taxable", "taxable value"))
    cgst = parse_decimal(pick("cgst", "cgst_amount", "cgst amount"))
    sgst = parse_decimal(pick("sgst", "sgst_amount", "sgst amount"))
    igst = parse_decimal(pick("igst", "igst_amount", "igst amount"))
    cess = parse_decimal(pick("cess", "cess_amount", "cess amount"))
    total_amount = parse_decimal(pick("total_amount", "invoice_value", "invoice value", "total"))
    if total_amount == 0:
        total_amount = taxable_value + cgst + sgst + igst + cess
    return NormalizedGstr2bRow(
        row_no=row_no,
        source_section=_safe_str(source_section or pick("source_section", "section", "table")),
        source_row_reference=_safe_str(pick("source_row_reference", "reference", "row_reference")),
        doc_type_code=normalize_doc_type(pick("doc_type", "document_type", "type")),
        counterparty_gstin=counterparty_gstin,
        counterparty_gstin_normalized=normalize_gstin(counterparty_gstin),
        counterparty_name=_safe_str(pick("supplier_name", "vendor_name", "supplier name", "vendor name")),
        invoice_number=invoice_number,
        invoice_number_normalized=normalize_invoice_number(invoice_number),
        invoice_date=parse_date_value(pick("supplier_invoice_date", "invoice_date", "invoice date")),
        taxable_value=taxable_value,
        cgst=cgst,
        sgst=sgst,
        igst=igst,
        cess=cess,
        total_amount=total_amount,
        pos_state_name=_safe_str(pick("pos_state", "place_of_supply", "place of supply")),
        raw_row_json=payload,
    )


class Gstr2bImportPipeline(BaseImportedReturnPipeline):
    code = "gstr2b"

    def supports(self, imported_return: GstImportedReturn) -> bool:
        return imported_return.return_type == GstImportedReturn.ReturnType.GSTR2B

    def validate(self, imported_return: GstImportedReturn) -> ImportPipelineResult:
        rows = list(imported_return.rows.order_by("row_no").values("row_no", "invoice_number_normalized", "counterparty_gstin_normalized"))
        return ImportPipelineResult(
            imported_return=imported_return,
            normalized_payload={"row_count": len(rows), "rows": rows[:10]},
            validation_summary={
                "row_count": len(rows),
                "has_rows": bool(rows),
                "source": imported_return.source,
            },
        )

    @classmethod
    @transaction.atomic
    def import_json(
        cls,
        *,
        entity_id: int,
        entityfinid_id: int,
        subentity_id: int | None,
        user,
        return_period: str,
        payload: Any,
        gst_registration_gstin: str | None = None,
        reference: str | None = None,
        create_run: bool = True,
        tolerance_config_json: dict[str, Any] | None = None,
    ) -> tuple[GstImportedReturn, GstReconciliationRun | None]:
        rows = [normalize_gstr2b_row(idx, row) for idx, row in enumerate(_extract_candidate_rows(payload), start=1)]
        imported_return = cls._persist_imported_return(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            user=user,
            return_period=return_period,
            gst_registration_gstin=gst_registration_gstin,
            source=GstImportedReturn.Source.JSON_UPLOAD,
            reference=reference,
            source_reference=f"gstr2b.json:{_json_checksum(payload)}",
            raw_payload_json=payload,
            normalized_rows=rows,
            checksum=_json_checksum(payload),
        )
        run = cls._create_run_from_imported_return(
            imported_return=imported_return,
            user=user,
            tolerance_config_json=tolerance_config_json or {},
        ) if create_run else None
        return imported_return, run

    @classmethod
    @transaction.atomic
    def import_excel(
        cls,
        *,
        entity_id: int,
        entityfinid_id: int,
        subentity_id: int | None,
        user,
        return_period: str,
        filename: str,
        content: bytes,
        gst_registration_gstin: str | None = None,
        create_run: bool = True,
        tolerance_config_json: dict[str, Any] | None = None,
    ) -> tuple[GstImportedReturn, GstReconciliationRun | None]:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(cell or "").strip().lower() for cell in header_cells]
        rows: list[NormalizedGstr2bRow] = []
        for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            payload = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values))) if headers[idx]}
            rows.append(normalize_gstr2b_row(row_no, payload, source_section=sheet.title))
        imported_return = cls._persist_imported_return(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            user=user,
            return_period=return_period,
            gst_registration_gstin=gst_registration_gstin,
            source=GstImportedReturn.Source.EXCEL_UPLOAD,
            reference=filename,
            source_reference=f"gstr2b.xlsx:{_file_checksum(content)}",
            raw_payload_json={"workbook_name": filename, "sheet": sheet.title},
            normalized_rows=rows,
            checksum=_file_checksum(content),
        )
        run = cls._create_run_from_imported_return(
            imported_return=imported_return,
            user=user,
            tolerance_config_json=tolerance_config_json or {},
        ) if create_run else None
        return imported_return, run

    @classmethod
    def _persist_imported_return(
        cls,
        *,
        entity_id: int,
        entityfinid_id: int,
        subentity_id: int | None,
        user,
        return_period: str,
        gst_registration_gstin: str | None,
        source: str,
        reference: str | None,
        source_reference: str,
        raw_payload_json: dict[str, Any] | list[Any] | Any,
        normalized_rows: list[NormalizedGstr2bRow],
        checksum: str,
    ) -> GstImportedReturn:
        imported_return = GstImportedReturn.objects.create(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            gst_registration_gstin=normalize_gstin(gst_registration_gstin),
            return_type=GstImportedReturn.ReturnType.GSTR2B,
            return_period=normalize_return_period(return_period),
            source=source,
            reference=reference,
            source_reference=source_reference,
            status=GstImportedReturn.Status.VALIDATED,
            checksum=checksum,
            raw_payload_json=raw_payload_json,
            normalized_payload_json={"row_count": len(normalized_rows)},
            validation_summary_json={"row_count": len(normalized_rows)},
            imported_by=user,
            imported_at=timezone.now(),
            created_by=user,
            updated_by=user,
        )
        for row in normalized_rows:
            payload = row.normalized_payload()
            GstImportedReturnRow.objects.create(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                imported_return=imported_return,
                row_no=row.row_no,
                source_section=row.source_section,
                source_row_reference=row.source_row_reference,
                row_hash=_row_hash(payload),
                doc_type_code=row.doc_type_code,
                counterparty_gstin=row.counterparty_gstin,
                counterparty_gstin_normalized=row.counterparty_gstin_normalized,
                counterparty_name=row.counterparty_name,
                invoice_number=row.invoice_number,
                invoice_number_normalized=row.invoice_number_normalized,
                invoice_date=row.invoice_date,
                taxable_value=row.taxable_value,
                cgst=row.cgst,
                sgst=row.sgst,
                igst=row.igst,
                cess=row.cess,
                total_amount=row.total_amount,
                pos_state_name=row.pos_state_name,
                raw_row_json=row.raw_row_json,
                normalized_row_json=payload,
                created_by=user,
                updated_by=user,
            )
        return imported_return

    @classmethod
    def _create_run_from_imported_return(
        cls,
        *,
        imported_return: GstImportedReturn,
        user,
        tolerance_config_json: dict[str, Any],
    ) -> GstReconciliationRun:
        run = GstReconciliationRun.objects.create(
            entity_id=imported_return.entity_id,
            entityfinid_id=imported_return.entityfinid_id,
            subentity_id=imported_return.subentity_id,
            gst_registration_gstin=imported_return.gst_registration_gstin,
            reconciliation_type=GstReconciliationRun.ReconciliationType.GSTR2B_PURCHASE,
            return_period=imported_return.return_period,
            source_mode=GstReconciliationRun.SourceMode.BOOKS_VS_IMPORTED,
            status=GstReconciliationRun.Status.IMPORTED,
            match_strategy_code="gstr2b_purchase_portal",
            tolerance_config_json=tolerance_config_json,
            imported_return=imported_return,
            source_reference=f"gst_imported_return:{imported_return.id}",
            summary_json={"total_items": imported_return.rows.count()},
            created_by=user,
            updated_by=user,
        )
        for row in imported_return.rows.order_by("row_no"):
            GstReconciliationItem.objects.create(
                entity_id=run.entity_id,
                entityfinid_id=run.entityfinid_id,
                subentity_id=run.subentity_id,
                run=run,
                item_type=GstReconciliationItem.ItemType.CREDIT_NOTE if row.doc_type_code == "CN" else (
                    GstReconciliationItem.ItemType.DEBIT_NOTE if row.doc_type_code == "DN" else GstReconciliationItem.ItemType.INVOICE
                ),
                direction=GstReconciliationItem.Direction.PURCHASE,
                match_key=f"{row.counterparty_gstin_normalized or ''}|{row.invoice_number_normalized or ''}",
                source_document_type="gst_imported_return_row",
                source_document_id=str(row.id),
                gstin=run.gst_registration_gstin,
                counterparty_gstin=row.counterparty_gstin,
                invoice_number=row.invoice_number,
                invoice_date=row.invoice_date,
                doc_type_code=row.doc_type_code,
                taxable_value_imported=row.taxable_value,
                cgst_imported=row.cgst,
                sgst_imported=row.sgst,
                igst_imported=row.igst,
                cess_imported=row.cess,
                resolution_status=GstReconciliationItem.ResolutionStatus.PENDING_REVIEW,
                metadata_json={"imported_return_id": imported_return.id, "imported_return_row_id": row.id},
                created_by=user,
                updated_by=user,
            )
        imported_return.status = GstImportedReturn.Status.CONSUMED
        imported_return.updated_by = user
        imported_return.save(update_fields=["status", "updated_by", "updated_at"])
        GstReconciliationActionLog.objects.create(
            entity_id=run.entity_id,
            entityfinid_id=run.entityfinid_id,
            subentity_id=run.subentity_id,
            run=run,
            action_type=GstReconciliationActionLog.ActionType.CREATED,
            actor=user,
            to_status=run.status,
            comment="Reconciliation run created from imported GSTR-2B portal data.",
            details_json={"imported_return_id": imported_return.id},
            created_by=user,
            updated_by=user,
        )
        return run
