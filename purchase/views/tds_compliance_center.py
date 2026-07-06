from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import logging
from time import perf_counter
from typing import Iterable, Optional

from django.conf import settings
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.response import Response

from entity.models import Entity, EntityFinancialYear, SubEntity
from financial.profile_access import account_pan
from purchase.models.purchase_core import PurchaseInvoiceHeader
from purchase.models.purchase_statutory import (
    PurchaseStatutoryChallan,
    PurchaseStatutoryChallanLine,
    PurchaseStatutoryForm16ADeducteeDocument,
    PurchaseStatutoryReturn,
    PurchaseStatutoryReturnLine,
    PurchaseStatutoryReviewNoteEvent,
)
from purchase.services.purchase_statutory_service import PurchaseStatutoryService
from purchase.views.purchase_meta import PurchaseMetaBaseAPIView
from purchase.views.purchase_statutory import _require_statutory_view
from reports.api.receivables_views import _safe_filename, _write_csv, _write_excel, _write_pdf

ZERO2 = Decimal("0.00")
logger = logging.getLogger(__name__)


class PurchaseTdsComplianceCenterAPIView(PurchaseMetaBaseAPIView):
    page_size = 25

    def get(self, request):
        request_started_at = perf_counter()
        stage_started_at = request_started_at
        stage_timings: dict[str, float] = {}

        def checkpoint(name: str) -> None:
            nonlocal stage_started_at
            now = perf_counter()
            stage_timings[name] = round((now - stage_started_at) * 1000, 2)
            stage_started_at = now

        entity_id, entityfinid_id, subentity_id = self._parse_scope(request, require_entityfinid=True)
        _require_statutory_view(request, entity_id)

        financial_year = EntityFinancialYear.objects.filter(
            pk=entityfinid_id,
            entity_id=entity_id,
        ).only("id", "desc", "finstartyear", "finendyear").first()
        if financial_year is None:
            return Response({"detail": "Financial year not found for the selected entity."}, status=404)

        period_from, period_to, quarter_code = self._resolve_period(
            request=request,
            financial_year=financial_year,
        )
        checkpoint("scope_resolution_ms")

        entity = Entity.objects.filter(pk=entity_id).only("entityname").first()
        subentity = (
            SubEntity.objects.filter(pk=subentity_id, entity_id=entity_id).only("id", "subentityname").first()
            if subentity_id is not None
            else None
        )

        posted_headers_qs = (
            PurchaseInvoiceHeader.objects.filter(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                status=PurchaseInvoiceHeader.Status.POSTED,
                bill_date__gte=period_from,
                bill_date__lte=period_to,
                tds_amount__gt=ZERO2,
            )
            .select_related("vendor", "vendor__compliance_profile", "tds_section", "subentity", "created_by")
            .order_by("-bill_date", "-id")
        )
        if subentity_id is not None:
            posted_headers_qs = posted_headers_qs.filter(subentity_id=subentity_id)
        headers = list(posted_headers_qs)
        checkpoint("headers_fetch_ms")

        challans_qs = (
            PurchaseStatutoryChallan.objects.filter(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
                challan_date__gte=period_from,
                challan_date__lte=period_to,
            )
            .prefetch_related("lines__header", "lines__section")
            .order_by("-challan_date", "-id")
        )
        if subentity_id is not None:
            challans_qs = challans_qs.filter(subentity_id=subentity_id)
        challans = list(challans_qs)
        checkpoint("challans_fetch_ms")

        returns_qs = (
            PurchaseStatutoryReturn.objects.filter(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                period_to__gte=period_from,
                period_from__lte=period_to,
            )
            .prefetch_related("lines__header", "lines__challan")
            .order_by("-period_to", "-id")
        )
        if subentity_id is not None:
            returns_qs = returns_qs.filter(subentity_id=subentity_id)
        returns = list(returns_qs)
        checkpoint("returns_fetch_ms")

        challan_lines_all = list(
            PurchaseStatutoryChallanLine.objects.filter(
                challan__entity_id=entity_id,
                challan__entityfinid_id=entityfinid_id,
                challan__tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            )
            .exclude(challan__status=PurchaseStatutoryChallan.Status.CANCELLED)
            .select_related("challan", "header", "header__tds_section", "section")
            .order_by("-challan__challan_date", "-id")
        )
        if subentity_id is not None:
            challan_lines_all = [line for line in challan_lines_all if line.challan.subentity_id == subentity_id]
        checkpoint("challan_lines_fetch_ms")

        summary = PurchaseStatutoryService.reconciliation_summary(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            date_from=period_from,
            date_to=period_to,
        )
        checkpoint("reconciliation_summary_ms")

        include_all_datasets = self._query_bool(request.query_params.get("include_datasets"), default=True)
        include_all_return_datasets = self._query_bool(
            request.query_params.get("include_return_datasets"),
            default=include_all_datasets,
        )
        requested_tabs = self._requested_tabs(request.query_params.get("tabs"))
        requested_return_tabs = self._requested_tabs(request.query_params.get("return_tabs"))
        requested_tab = str(request.query_params.get("tab") or "").strip().lower()
        requested_return_tab = str(request.query_params.get("return_tab") or "").strip().lower()
        if requested_tab and requested_tab != "return-filing":
            requested_tabs.add(requested_tab)
        if requested_tab == "return-filing" and requested_return_tab:
            requested_return_tabs.add(requested_return_tab)

        header_mapping = self._build_header_mapping(challan_lines_all)
        needs_header_rows = include_all_datasets or "deduction-register" in requested_tabs
        needs_section_rows = include_all_datasets or bool({"payable-report", "section-wise-summary"} & requested_tabs)
        needs_deductee_rows = include_all_datasets or "deductee-wise-summary" in requested_tabs
        needs_monthly_rows = include_all_datasets or "monthly-summary" in requested_tabs
        needs_challan_rows = include_all_datasets or "payment-register" in requested_tabs
        needs_challan_mapping_rows = include_all_datasets or "challan-mapping" in requested_tabs
        needs_pending_rows = include_all_datasets or "pending-payment" in requested_tabs
        needs_vendor_rows = include_all_datasets or "vendor-compliance" in requested_tabs
        needs_filing_rows = include_all_datasets or include_all_return_datasets or "return-filing" in requested_tabs or bool(requested_return_tabs)
        needs_form16a_rows = include_all_datasets or "form-16a" in requested_tabs
        needs_audit_rows = include_all_datasets or "audit-trail" in requested_tabs

        section_rows = self._build_section_summary_rows(
            headers=headers,
            challan_lines=challan_lines_all,
            period_from=period_from,
        )
        deductee_rows = self._build_deductee_summary_rows(headers=headers, header_mapping=header_mapping)
        monthly_rows = self._build_monthly_summary_rows(
            headers=headers,
            challan_lines=challan_lines_all,
            returns=returns,
        )
        challan_mapping_metrics = self._build_challan_mapping_metrics(headers, header_mapping)
        pending_dashboard_rows, pending_overdue_count = self._build_pending_liability_preview(headers, header_mapping)
        missing_pan_count = self._count_missing_pan_deductees(headers)
        filing_metrics = self._build_filing_metrics(returns)
        checkpoint("shared_builder_ms")
        challan_mapping_rows = self._build_challan_mapping_rows(headers, header_mapping) if needs_challan_mapping_rows else []
        pending_rows = self._build_pending_payment_rows(headers, header_mapping) if needs_pending_rows else []
        vendor_rows = self._build_vendor_compliance_rows(headers, header_mapping) if needs_vendor_rows else []
        filing_rows = self._build_return_rows(returns) if needs_filing_rows else {"24q": [], "26q": [], "27q": []}
        audit_rows = (
            self._build_audit_rows(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                period_from=period_from,
                period_to=period_to,
                challans=challans,
                returns=returns,
            )
            if needs_audit_rows
            else []
        )
        recent_activity_rows = audit_rows[:6] if audit_rows else self._build_recent_activity_rows(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            period_from=period_from,
            period_to=period_to,
            challans=challans,
            returns=returns,
            limit=6,
        )
        header_rows = self._build_deduction_rows(headers, header_mapping) if needs_header_rows else []
        challan_rows = self._build_payment_register_rows(challans) if needs_challan_rows else []
        form16a_rows = self._build_form16a_rows(returns) if needs_form16a_rows else []
        checkpoint("dataset_builder_ms")

        warnings = self._build_warning_chips(
            missing_pan_count=missing_pan_count,
            unmapped_count=challan_mapping_metrics["unmapped"],
            overdue_count=pending_overdue_count,
            filing_error_count=filing_metrics["filing_errors"],
        )

        payload = {
            "pageTitle": "TDS Compliance Center",
            "meta": self._workspace_meta(
                namespace="purchase.tds_compliance_center.meta",
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                entity_label=getattr(entity, "entityname", None) or f"Entity {entity_id}",
                branch_label=getattr(subentity, "subentityname", None) or "All subentities",
                tds_sections=self._tds_sections(),
            ),
            "tabs": self._tabs(),
            "returnTabs": self._return_tabs(),
            "headerChips": [
                {"label": getattr(entity, "entityname", None) or f"Entity {entity_id}"},
                {"label": self._financial_year_label(financial_year)},
                {"label": self._quarter_label(quarter_code)},
                {"label": getattr(subentity, "subentityname", None) or "All subentities"},
                {"label": f"{self._display_date(period_from)} to {self._display_date(period_to)}", "tone": "info"},
            ],
            "kpis": self._build_kpis(
                summary,
                missing_pan_count=missing_pan_count,
                pending_challans=challan_mapping_metrics["pending"],
                pending_returns=filing_metrics["pending_returns"],
                short_deduction_cases=filing_metrics["short_deduction_cases"],
            ),
            "warnings": warnings,
            "dashboard": self._build_dashboard(
                monthly_rows,
                section_rows,
                deductee_rows,
                pending_dashboard_rows,
                recent_activity_rows,
            ),
            "filters": {
                "financialYearId": entityfinid_id,
                "quarter": quarter_code,
                "fromDate": period_from.isoformat(),
                "toDate": period_to.isoformat(),
                "vendorId": None,
                "vendorLabel": "",
                "pan": "",
                "tdsSectionId": None,
                "expenseLedgerLabel": "",
                "voucherType": "",
                "paymentStatus": "",
                "challanStatus": "",
                "returnStatus": "",
                "minAmount": None,
                "maxAmount": None,
                "branchId": subentity_id,
                "entityId": entity_id,
                "searchText": request.query_params.get("search") or "",
            },
            "datasets": self._build_datasets_payload(
                include_all_datasets=include_all_datasets,
                requested_tabs=requested_tabs,
                summary=summary,
                monthly_rows=monthly_rows,
                header_rows=header_rows,
                section_rows=section_rows,
                challan_rows=challan_rows,
                challan_mapping_rows=challan_mapping_rows,
                deductee_rows=deductee_rows,
                pending_rows=pending_rows,
                vendor_rows=vendor_rows,
                filing_rows=filing_rows,
                form16a_rows=form16a_rows,
                audit_rows=audit_rows,
            ),
            "returnDatasets": self._build_return_datasets_payload(
                include_all_return_datasets=include_all_return_datasets,
                requested_return_tabs=requested_return_tabs,
                filing_rows=filing_rows,
            ),
        }
        checkpoint("payload_build_ms")
        logger.info(
            "purchase_tds_compliance_center_profile entity=%s entityfinid=%s subentity=%s quarter=%s headers=%s challans=%s returns=%s challan_lines=%s "
            "deduction_rows=%s section_rows=%s deductee_rows=%s monthly_rows=%s challan_mapping_rows=%s pending_rows=%s vendor_rows=%s filing_rows=%s "
            "audit_rows=%s form16a_rows=%s total_ms=%.2f stage_ms=%s",
            entity_id,
            entityfinid_id,
            subentity_id,
            quarter_code,
            len(headers),
            len(challans),
            len(returns),
            len(challan_lines_all),
            len(header_rows),
            len(section_rows),
            len(deductee_rows),
            len(monthly_rows),
            len(challan_mapping_rows),
            len(pending_rows),
            len(vendor_rows),
            sum(len(rows) for rows in filing_rows.values()),
            len(audit_rows),
            len(form16a_rows),
            (perf_counter() - request_started_at) * 1000,
            stage_timings,
        )
        return Response(payload)

    def _workspace_meta(
        self,
        *,
        namespace: str,
        entity_id: int,
        entityfinid_id: int,
        subentity_id: Optional[int],
        entity_label: str,
        branch_label: str,
        tds_sections: list[dict[str, object]],
    ) -> dict[str, object]:
        return self._get_cached_meta(
            namespace=namespace,
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            extra={},
            timeout=getattr(settings, "META_CACHE_FORM_TTL_SECONDS", 600),
            loader=lambda: {
                "entityId": entity_id,
                "entityLabel": entity_label,
                "branchLabel": branch_label,
                "financialYears": self._normalize_financial_years(self._financial_years(entity_id)),
                "subentities": self._subentities(entity_id),
                "quarters": self._quarter_options(),
                "voucherTypes": [
                    {"value": "purchase_invoice", "label": "Purchase Invoice"},
                    {"value": "purchase_credit_note", "label": "Purchase Credit Note"},
                    {"value": "purchase_debit_note", "label": "Purchase Debit Note"},
                ],
                "paymentStatuses": [
                    {"value": "pending", "label": "Pending"},
                    {"value": "paid", "label": "Paid"},
                    {"value": "overdue", "label": "Overdue"},
                    {"value": "partially_mapped", "label": "Partially Mapped"},
                ],
                "challanStatuses": [
                    {"value": "unmapped", "label": "Unmapped"},
                    {"value": "partially_mapped", "label": "Partially Mapped"},
                    {"value": "mapped", "label": "Mapped"},
                    {"value": "deposited", "label": "Deposited"},
                ],
                "returnStatuses": [
                    {"value": "draft", "label": "Draft"},
                    {"value": "approval_submitted", "label": "Approval Submitted"},
                    {"value": "approved_draft", "label": "Approved Draft"},
                    {"value": "rejected_draft", "label": "Rejected Draft"},
                    {"value": "filed", "label": "Filed"},
                    {"value": "revised", "label": "Revised"},
                    {"value": "validated", "label": "Validated"},
                ],
                "tdsSections": tds_sections,
            },
        )

    def _resolve_period(self, *, request, financial_year: EntityFinancialYear) -> tuple[date, date, str]:
        quarter_code = str(request.query_params.get("quarter") or "").strip().upper()
        today = timezone.localdate()
        fy_start = self._as_date(financial_year.finstartyear) or today
        fy_end = self._as_date(financial_year.finendyear) or today
        quarter_ranges = self._quarter_ranges(fy_start)

        if not quarter_code:
            quarter_code = self._quarter_for_date(today, fy_start, fy_end)

        requested_from = self._parse_optional_date(request.query_params.get("from_date") or request.query_params.get("fromDate"))
        requested_to = self._parse_optional_date(request.query_params.get("to_date") or request.query_params.get("toDate"))

        if requested_from and requested_to:
            return max(requested_from, fy_start), min(requested_to, fy_end), quarter_code

        if quarter_code in quarter_ranges:
            start, end = quarter_ranges[quarter_code]
            return max(start, fy_start), min(end, fy_end), quarter_code

        return fy_start, fy_end, "Q1"

    def _build_header_mapping(self, challan_lines: Iterable[PurchaseStatutoryChallanLine]) -> dict[int, dict[str, object]]:
        mapping: dict[int, dict[str, object]] = {}
        for line in challan_lines:
            header_id = int(line.header_id)
            bucket = mapping.setdefault(
                header_id,
                {
                    "mapped": ZERO2,
                    "deposited": ZERO2,
                    "challan_nos": set(),
                    "sections": set(),
                },
            )
            amount = self._decimal(line.amount)
            bucket["mapped"] = self._decimal(bucket["mapped"]) + amount
            challan_no = (getattr(line.challan, "challan_no", "") or "").strip()
            if challan_no:
                bucket["challan_nos"].add(challan_no)
            section_code = (
                (getattr(getattr(line, "section", None), "section_code", None) or "")
                or (getattr(getattr(line.header, "tds_section", None), "section_code", None) or "")
            ).strip()
            if section_code:
                bucket["sections"].add(section_code)
            if int(getattr(line.challan, "status", 0)) == int(PurchaseStatutoryChallan.Status.DEPOSITED):
                bucket["deposited"] = self._decimal(bucket["deposited"]) + amount
        return mapping

    def _build_deduction_rows(self, headers: list[PurchaseInvoiceHeader], mapping: dict[int, dict[str, object]]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for header in headers:
            row_mapping = mapping.get(int(header.id), {})
            tds_amount = self._decimal(header.tds_amount)
            deposited = self._decimal(row_mapping.get("deposited"))
            status = self._header_status_badge(tds_amount=tds_amount, deposited_amount=deposited, pan=account_pan(header.vendor))
            rows.append(
                {
                    "id": int(header.id),
                    "date": self._iso(header.bill_date),
                    "voucherNo": self._voucher_label(header),
                    "voucherType": self._voucher_type_label(header.doc_type),
                    "deductee": header.vendor_name or getattr(header.vendor, "accountname", "") or "-",
                    "pan": account_pan(header.vendor) or "-",
                    "section": getattr(getattr(header, "tds_section", None), "section_code", None) or "UNSPECIFIED",
                    "taxableAmount": self._money(header.tds_base_amount),
                    "tdsAmount": self._money(tds_amount),
                    "status": status,
                    "actions": self._actions("Voucher", "Deductee"),
                }
            )
        return rows

    def _build_section_summary_rows(
        self,
        *,
        headers: list[PurchaseInvoiceHeader],
        challan_lines: list[PurchaseStatutoryChallanLine],
        period_from: date,
    ) -> list[dict[str, object]]:
        section_map: dict[str, dict[str, object]] = {}

        prior_headers = (
            PurchaseInvoiceHeader.objects.filter(
                entity_id=headers[0].entity_id if headers else None,
                entityfinid_id=headers[0].entityfinid_id if headers else None,
                status=PurchaseInvoiceHeader.Status.POSTED,
                bill_date__lt=period_from,
                tds_amount__gt=ZERO2,
            )
            if headers
            else PurchaseInvoiceHeader.objects.none()
        )
        if headers and headers[0].subentity_id is not None:
            prior_headers = prior_headers.filter(subentity_id=headers[0].subentity_id)
        for row in prior_headers.values("tds_section__section_code", "tds_section__description").annotate(opening=Sum("tds_amount")):
            code = row.get("tds_section__section_code") or "UNSPECIFIED"
            bucket = section_map.setdefault(
                code,
                {
                    "nature": row.get("tds_section__description") or "TDS Section",
                    "opening": ZERO2,
                    "current": ZERO2,
                    "deposited": ZERO2,
                    "interest": ZERO2,
                    "transactions": 0,
                    "taxable": ZERO2,
                },
            )
            bucket["opening"] += self._decimal(row.get("opening"))

        for line in challan_lines:
            challan = line.challan
            section_code = (
                (getattr(getattr(line, "section", None), "section_code", None) or "")
                or (getattr(getattr(line.header, "tds_section", None), "section_code", None) or "")
                or "UNSPECIFIED"
            )
            bucket = section_map.setdefault(section_code, self._new_section_bucket(line.header))
            amount = self._decimal(line.amount)
            if challan.challan_date < period_from and int(challan.status) == int(PurchaseStatutoryChallan.Status.DEPOSITED):
                bucket["opening"] -= amount
            elif challan.challan_date >= period_from:
                if int(challan.status) == int(PurchaseStatutoryChallan.Status.DEPOSITED):
                    bucket["deposited"] += amount
                    bucket["interest"] += self._section_share(
                        amount=amount,
                        total=self._decimal(challan.amount),
                        distributed=self._decimal(challan.interest_amount),
                    )

        for header in headers:
            section_code = getattr(getattr(header, "tds_section", None), "section_code", None) or "UNSPECIFIED"
            bucket = section_map.setdefault(section_code, self._new_section_bucket(header))
            bucket["transactions"] += 1
            bucket["taxable"] += self._decimal(header.tds_base_amount)
            bucket["current"] += self._decimal(header.tds_amount)

        rows: list[dict[str, object]] = []
        for section_code, bucket in sorted(section_map.items(), key=lambda item: item[0]):
            closing = bucket["opening"] + bucket["current"] - bucket["deposited"]
            rows.append(
                {
                    "id": section_code.lower().replace("/", "-"),
                    "section": section_code,
                    "nature": bucket["nature"] or "TDS Section",
                    "transactions": int(bucket["transactions"]),
                    "taxableAmount": self._money(bucket["taxable"]),
                    "tdsAmount": self._money(bucket["current"]),
                    "pendingAmount": self._money(max(closing, ZERO2)),
                    "openingBalance": self._money(bucket["opening"]),
                    "currentDeduction": self._money(bucket["current"]),
                    "deposited": self._money(bucket["deposited"]),
                    "interest": self._money(bucket["interest"]),
                    "closingBalance": self._money(closing),
                    "status": self._balance_status(closing, bucket["interest"]),
                    "actions": self._actions("Section View", "Monthly Drilldown"),
                }
            )
        return rows

    def _build_deductee_summary_rows(
        self,
        *,
        headers: list[PurchaseInvoiceHeader],
        header_mapping: dict[int, dict[str, object]],
    ) -> list[dict[str, object]]:
        buckets: dict[str, dict[str, object]] = {}
        for header in headers:
            vendor_key = header.vendor_name or getattr(header.vendor, "accountname", "") or f"Vendor {header.vendor_id or header.id}"
            bucket = buckets.setdefault(
                vendor_key,
                {
                    "deductee": vendor_key,
                    "pan": account_pan(header.vendor) or "-",
                    "transactions": 0,
                    "taxableAmount": ZERO2,
                    "tdsAmount": ZERO2,
                    "pending": ZERO2,
                    "missing_pan": False,
                },
            )
            bucket["transactions"] += 1
            bucket["taxableAmount"] += self._decimal(header.tds_base_amount)
            bucket["tdsAmount"] += self._decimal(header.tds_amount)
            row_mapping = header_mapping.get(int(header.id), {})
            pending = max(self._decimal(header.tds_amount) - self._decimal(row_mapping.get("deposited")), ZERO2)
            bucket["pending"] += pending
            bucket["missing_pan"] = bucket["missing_pan"] or not bool((account_pan(header.vendor) or "").strip())

        rows: list[dict[str, object]] = []
        for index, bucket in enumerate(sorted(buckets.values(), key=lambda item: (-item["tdsAmount"], item["deductee"])), start=1):
            compliance = (
                self._badge("Missing PAN", "danger")
                if bucket["missing_pan"]
                else self._badge("Pending Deposit", "warning")
                if bucket["pending"] > ZERO2
                else self._badge("Compliant", "success")
            )
            rows.append(
                {
                    "id": index,
                    "deductee": bucket["deductee"],
                    "pan": bucket["pan"],
                    "transactions": int(bucket["transactions"]),
                    "taxableAmount": self._money(bucket["taxableAmount"]),
                    "tdsAmount": self._money(bucket["tdsAmount"]),
                    "pending": self._money(bucket["pending"]),
                    "complianceStatus": compliance,
                }
            )
        return rows

    def _build_monthly_summary_rows(
        self,
        *,
        headers: list[PurchaseInvoiceHeader],
        challan_lines: list[PurchaseStatutoryChallanLine],
        returns: list[PurchaseStatutoryReturn],
    ) -> list[dict[str, object]]:
        buckets: dict[str, dict[str, object]] = {}
        for header in headers:
            month_key = header.bill_date.strftime("%Y-%m")
            bucket = buckets.setdefault(
                month_key,
                {
                    "label": header.bill_date.strftime("%b %Y"),
                    "taxable": ZERO2,
                    "deducted": ZERO2,
                    "deposited": ZERO2,
                    "returnStatus": self._badge("Draft", "warning"),
                },
            )
            bucket["taxable"] += self._decimal(header.tds_base_amount)
            bucket["deducted"] += self._decimal(header.tds_amount)

        for line in challan_lines:
            if int(line.challan.status) != int(PurchaseStatutoryChallan.Status.DEPOSITED):
                continue
            month_key = line.challan.challan_date.strftime("%Y-%m")
            bucket = buckets.setdefault(
                month_key,
                {
                    "label": line.challan.challan_date.strftime("%b %Y"),
                    "taxable": ZERO2,
                    "deducted": ZERO2,
                    "deposited": ZERO2,
                    "returnStatus": self._badge("Draft", "warning"),
                },
            )
            bucket["deposited"] += self._decimal(line.amount)

        for filing in returns:
            filing_month = filing.period_to.strftime("%Y-%m")
            bucket = buckets.setdefault(
                filing_month,
                {
                    "label": filing.period_to.strftime("%b %Y"),
                    "taxable": ZERO2,
                    "deducted": ZERO2,
                    "deposited": ZERO2,
                    "returnStatus": self._badge("Draft", "warning"),
                },
            )
            bucket["returnStatus"] = self._return_status_badge(filing)

        rows: list[dict[str, object]] = []
        for month_key in sorted(buckets.keys()):
            bucket = buckets[month_key]
            pending = max(bucket["deducted"] - bucket["deposited"], ZERO2)
            rows.append(
                {
                    "id": month_key,
                    "month": bucket["label"],
                    "taxableAmount": self._money(bucket["taxable"]),
                    "tdsDeducted": self._money(bucket["deducted"]),
                    "deposited": self._money(bucket["deposited"]),
                    "pending": self._money(pending),
                    "returnStatus": bucket["returnStatus"],
                }
            )
        return rows

    def _build_payment_register_rows(self, challans: list[PurchaseStatutoryChallan]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for challan in challans:
            section_codes = sorted(
                {
                    (
                        getattr(getattr(line, "section", None), "section_code", None)
                        or getattr(getattr(line.header, "tds_section", None), "section_code", None)
                        or "UNSPECIFIED"
                    )
                    for line in challan.lines.all()
                }
            )
            rows.append(
                {
                    "id": int(challan.id),
                    "challanNo": challan.challan_no or f"Challan {challan.id}",
                    "cin": challan.cin_no or "-",
                    "bank": challan.bank_ref_no or challan.bsr_code or "-",
                    "depositDate": self._iso(challan.deposited_on or challan.challan_date),
                    "amount": self._money(challan.total_deposit_amount),
                    "section": " / ".join(section_codes),
                    "status": self._challan_status_badge(challan),
                    "actions": self._actions(self._challan_primary_action_label(challan), "Mapped Vouchers"),
                }
            )
        return rows

    def _build_challan_mapping_rows(
        self,
        headers: list[PurchaseInvoiceHeader],
        header_mapping: dict[int, dict[str, object]],
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for header in headers:
            row_mapping = header_mapping.get(int(header.id), {})
            tds_amount = self._decimal(header.tds_amount)
            mapped = self._decimal(row_mapping.get("mapped"))
            remaining = max(tds_amount - mapped, ZERO2)
            mapped_challan = ", ".join(sorted(row_mapping.get("challan_nos", set()))) or "-"
            rows.append(
                {
                    "id": int(header.id),
                    "voucherNo": self._voucher_label(header),
                    "deductee": header.vendor_name or getattr(header.vendor, "accountname", "") or "-",
                    "section": getattr(getattr(header, "tds_section", None), "section_code", None) or "UNSPECIFIED",
                    "tdsAmount": self._money(tds_amount),
                    "mappedChallan": mapped_challan,
                    "remainingAmount": self._money(remaining),
                    "mappingStatus": (
                        self._badge("Mapped", "success")
                        if remaining <= ZERO2
                        else self._badge("Partially Mapped", "warning")
                        if mapped > ZERO2
                        else self._badge("Unmapped", "danger")
                    ),
                    "actions": self._actions("Voucher", "Mapping Detail"),
                }
            )
        return rows

    def _build_pending_payment_rows(
        self,
        headers: list[PurchaseInvoiceHeader],
        header_mapping: dict[int, dict[str, object]],
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        today = timezone.localdate()
        for header in headers:
            row_mapping = header_mapping.get(int(header.id), {})
            remaining = max(self._decimal(header.tds_amount) - self._decimal(row_mapping.get("deposited")), ZERO2)
            if remaining <= ZERO2:
                continue
            due_date = self._deposit_due_date(header.bill_date)
            delay_days = max((today - due_date).days, 0)
            rows.append(
                {
                    "id": int(header.id),
                    "dueDate": self._iso(due_date),
                    "section": getattr(getattr(header, "tds_section", None), "section_code", None) or "UNSPECIFIED",
                    "deductee": header.vendor_name or getattr(header.vendor, "accountname", "") or "-",
                    "tdsAmount": self._money(remaining),
                    "delayDays": delay_days,
                    "interest": self._money(self._interest_estimate(remaining, delay_days)),
                    "status": self._badge("Overdue", "danger") if delay_days > 0 else self._badge("Pending", "warning"),
                }
            )
        rows.sort(key=lambda row: (row["dueDate"], row["section"], row["deductee"]))
        return rows

    def _build_vendor_compliance_rows(
        self,
        headers: list[PurchaseInvoiceHeader],
        header_mapping: dict[int, dict[str, object]],
    ) -> list[dict[str, object]]:
        buckets: dict[str, dict[str, object]] = {}
        for header in headers:
            key = header.vendor_name or getattr(header.vendor, "accountname", "") or f"Vendor {header.vendor_id or header.id}"
            bucket = buckets.setdefault(
                key,
                {
                    "id": key,
                    "deductee": key,
                    "pan": account_pan(header.vendor) or "-",
                    "defaultSection": getattr(getattr(header, "tds_section", None), "section_code", None) or "UNSPECIFIED",
                    "certificate": "Not Available",
                    "validity": None,
                    "pending": ZERO2,
                },
            )
            row_mapping = header_mapping.get(int(header.id), {})
            bucket["pending"] += max(self._decimal(header.tds_amount) - self._decimal(row_mapping.get("deposited")), ZERO2)

        rows: list[dict[str, object]] = []
        for bucket in sorted(buckets.values(), key=lambda item: item["deductee"]):
            has_pan = bool((bucket["pan"] or "").strip() and bucket["pan"] != "-")
            rows.append(
                {
                    "id": bucket["id"],
                    "deductee": bucket["deductee"],
                    "pan": bucket["pan"],
                    "panStatus": self._badge("Valid", "success") if has_pan else self._badge("Missing PAN", "danger"),
                    "defaultSection": bucket["defaultSection"],
                    "certificate": bucket["certificate"],
                    "validity": bucket["validity"],
                    "complianceStatus": (
                        self._badge("Review Required", "danger")
                        if not has_pan
                        else self._badge("Pending Deposit", "warning")
                        if bucket["pending"] > ZERO2
                        else self._badge("Compliant", "success")
                    ),
                }
            )
        return rows

    def _build_return_rows(self, returns: list[PurchaseStatutoryReturn]) -> dict[str, list[dict[str, object]]]:
        rows = {"24q": [], "26q": [], "27q": []}
        for filing in returns:
            code = (filing.return_code or "").strip().upper()
            if code not in {"24Q", "26Q", "27Q"}:
                continue
            target = code.lower()
            line_items = list(filing.lines.all())
            deductee_count = len({(line.deductee_pan_snapshot or line.deductee_tax_id_snapshot or line.header_id) for line in line_items})
            total_taxable = sum((self._decimal(getattr(line.header, "tds_base_amount", ZERO2)) for line in line_items), ZERO2)
            validation_errors = 0
            warnings = 0
            if code == "27Q":
                for line in line_items:
                    if not (line.deductee_tax_id_snapshot or "").strip():
                        validation_errors += 1
                    if not (line.deductee_country_name_snapshot or line.deductee_country_code_snapshot or "").strip():
                        warnings += 1
            status_badge = self._return_status_badge(filing)
            row = {
                "id": int(filing.id),
                "quarter": self._quarter_for_date(filing.period_from, filing.period_from, filing.period_to),
                "returnType": self._return_type_label(code),
                "transactions": len(line_items),
                "deductees": deductee_count,
                "totalTaxableAmount": self._money(total_taxable),
                "totalTds": self._money(filing.amount),
                "validationErrors": validation_errors,
                "warnings": warnings,
                "fvuStatus": self._badge("Validated", "success") if validation_errors == 0 else self._badge("Error", "danger"),
                "returnStatus": status_badge,
                "tokenNumber": filing.ack_no or filing.arn_no or "-",
                "filingDate": self._iso(filing.filed_on),
                "actions": self._actions(self._return_primary_action_label(filing), "Validation"),
            }
            if code == "27Q":
                countries = sorted({(line.deductee_country_name_snapshot or line.deductee_country_code_snapshot or "").strip() for line in line_items if (line.deductee_country_name_snapshot or line.deductee_country_code_snapshot or "").strip()})
                tax_ids = sorted({(line.deductee_tax_id_snapshot or "").strip() for line in line_items if (line.deductee_tax_id_snapshot or "").strip()})
                row.update(
                    {
                        "country": ", ".join(countries) or "-",
                        "taxResidencyCountry": ", ".join(countries) or "-",
                        "tin": ", ".join(tax_ids) or "-",
                        "currency": "INR",
                        "foreignAmount": self._money(total_taxable),
                        "exchangeRate": "1.00",
                        "inrAmount": self._money(total_taxable),
                        "dtaaApplicable": "Review",
                        "dtaaArticle": "-",
                        "surcharge": "0.00",
                        "cess": "0.00",
                    }
                )
            rows[target].append(row)
        return rows

    def _build_form16a_rows(self, returns: list[PurchaseStatutoryReturn]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        row_id = 1
        for filing in returns:
            code = (filing.return_code or "").strip().upper()
            if code not in {"26Q", "27Q"}:
                continue
            try:
                certificate_payload = PurchaseStatutoryService.list_form16a_certificates(filing_id=int(filing.id))
            except Exception:
                certificate_payload = {"certificates": []}
            certificate_docs = {
                str(doc.deductee_key): doc
                for doc in PurchaseStatutoryForm16ADeducteeDocument.objects.filter(filing_id=filing.id)
            }
            for certificate in certificate_payload.get("certificates", []):
                deductee_key = str(certificate.get("deductee_key") or "")
                uploaded_doc = certificate_docs.get(deductee_key)
                rows.append(
                    {
                        "id": row_id,
                        "deductee": deductee_key or certificate.get("invoice_label") or f"Filing {filing.id}",
                        "pan": certificate.get("pan") or "-",
                        "quarter": self._quarter_for_date(filing.period_from, filing.period_from, filing.period_to),
                        "certificateNo": certificate.get("certificate_no") or "-",
                        "generatedDate": self._iso(uploaded_doc.uploaded_at.date()) if uploaded_doc and uploaded_doc.uploaded_at else self._iso(filing.filed_on),
                        "emailStatus": self._badge("Pending", "warning"),
                        "downloadStatus": self._badge("Generated", "info") if uploaded_doc else self._badge("Pending", "warning"),
                        "actions": self._actions("Certificate", "Deductee"),
                    }
                )
                row_id += 1
        return rows

    def _build_audit_rows(
        self,
        *,
        entity_id: int,
        entityfinid_id: int,
        subentity_id: Optional[int],
        period_from: date,
        period_to: date,
        challans: list[PurchaseStatutoryChallan],
        returns: list[PurchaseStatutoryReturn],
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        row_id = 1
        for challan in challans:
            rows.append(
                {
                    "id": row_id,
                    "dateTime": self._iso_datetime(challan.created_at),
                    "user": self._user_label(challan.created_by),
                    "action": "Challan Created",
                    "voucherNo": challan.challan_no or f"Challan {challan.id}",
                    "field": "status",
                    "oldValue": "-",
                    "newValue": challan.get_status_display(),
                    "remarks": challan.remarks or "",
                    "ipAddress": "-",
                }
            )
            row_id += 1
            if challan.deposited_at:
                rows.append(
                    {
                        "id": row_id,
                        "dateTime": self._iso_datetime(challan.deposited_at),
                        "user": self._user_label(challan.deposited_by),
                        "action": "Challan Deposited",
                        "voucherNo": challan.challan_no or f"Challan {challan.id}",
                        "field": "status",
                        "oldValue": "Draft",
                        "newValue": "Deposited",
                        "remarks": challan.bank_ref_no or challan.cin_no or "",
                        "ipAddress": "-",
                    }
                )
                row_id += 1

        for filing in returns:
            rows.append(
                {
                    "id": row_id,
                    "dateTime": self._iso_datetime(filing.created_at),
                    "user": self._user_label(filing.created_by),
                    "action": "Return Created",
                    "voucherNo": filing.return_code or f"Return {filing.id}",
                    "field": "status",
                    "oldValue": "-",
                    "newValue": filing.get_status_display(),
                    "remarks": filing.remarks or "",
                    "ipAddress": "-",
                }
            )
            row_id += 1
            if filing.filed_at:
                rows.append(
                    {
                        "id": row_id,
                        "dateTime": self._iso_datetime(filing.filed_at),
                        "user": self._user_label(filing.filed_by),
                        "action": "Return Filed",
                        "voucherNo": filing.return_code or f"Return {filing.id}",
                        "field": "status",
                        "oldValue": "Draft",
                        "newValue": filing.get_status_display(),
                        "remarks": filing.ack_no or filing.arn_no or "",
                        "ipAddress": "-",
                    }
                )
                row_id += 1

        review_events = PurchaseStatutoryReviewNoteEvent.objects.filter(
            review_note__entity_id=entity_id,
            review_note__entityfinid_id=entityfinid_id,
            review_note__tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            changed_at__date__gte=period_from,
            changed_at__date__lte=period_to,
        ).select_related("review_note", "changed_by")
        if subentity_id is not None:
            review_events = review_events.filter(review_note__subentity_id=subentity_id)
        for event in review_events.order_by("-changed_at", "-id")[:20]:
            rows.append(
                {
                    "id": row_id,
                    "dateTime": self._iso_datetime(event.changed_at),
                    "user": self._user_label(event.changed_by),
                    "action": f"Review {event.action.title()}",
                    "voucherNo": event.review_note.tax_type or "IT_TDS",
                    "field": "closure_status",
                    "oldValue": "-",
                    "newValue": event.closure_status,
                    "remarks": event.closure_comment or event.review_summary or "",
                    "ipAddress": "-",
                }
            )
            row_id += 1

        rows.sort(key=lambda row: row["dateTime"] or "", reverse=True)
        return rows[:50]

    def _build_recent_activity_rows(
        self,
        *,
        entity_id: int,
        entityfinid_id: int,
        subentity_id: Optional[int],
        period_from: date,
        period_to: date,
        challans: list[PurchaseStatutoryChallan],
        returns: list[PurchaseStatutoryReturn],
        limit: int,
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []

        for challan in challans[:limit]:
            rows.append(
                {
                    "action": "Challan Created",
                    "voucherNo": challan.challan_no or f"Challan {challan.id}",
                    "remarks": challan.remarks or "",
                    "newValue": challan.get_status_display(),
                    "dateTime": self._iso_datetime(challan.created_at),
                }
            )
            if challan.deposited_at:
                rows.append(
                    {
                        "action": "Challan Deposited",
                        "voucherNo": challan.challan_no or f"Challan {challan.id}",
                        "remarks": challan.bank_ref_no or challan.cin_no or "",
                        "newValue": "Deposited",
                        "dateTime": self._iso_datetime(challan.deposited_at),
                    }
                )

        for filing in returns[:limit]:
            rows.append(
                {
                    "action": "Return Created",
                    "voucherNo": filing.return_code or f"Return {filing.id}",
                    "remarks": filing.remarks or "",
                    "newValue": filing.get_status_display(),
                    "dateTime": self._iso_datetime(filing.created_at),
                }
            )
            if filing.filed_at:
                rows.append(
                    {
                        "action": "Return Filed",
                        "voucherNo": filing.return_code or f"Return {filing.id}",
                        "remarks": filing.ack_no or filing.arn_no or "",
                        "newValue": filing.get_status_display(),
                        "dateTime": self._iso_datetime(filing.filed_at),
                    }
                )

        review_events = PurchaseStatutoryReviewNoteEvent.objects.filter(
            review_note__entity_id=entity_id,
            review_note__entityfinid_id=entityfinid_id,
            review_note__tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            changed_at__date__gte=period_from,
            changed_at__date__lte=period_to,
        ).select_related("review_note")
        if subentity_id is not None:
            review_events = review_events.filter(review_note__subentity_id=subentity_id)
        for event in review_events.order_by("-changed_at", "-id")[:limit]:
            rows.append(
                {
                    "action": f"Review {event.action.title()}",
                    "voucherNo": event.review_note.tax_type or "IT_TDS",
                    "remarks": event.closure_comment or event.review_summary or "",
                    "newValue": event.closure_status,
                    "dateTime": self._iso_datetime(event.changed_at),
                }
            )

        rows.sort(key=lambda row: row["dateTime"] or "", reverse=True)
        return rows[:limit]

    def _count_missing_pan_deductees(self, headers: list[PurchaseInvoiceHeader]) -> int:
        deductees: dict[str, bool] = {}
        for header in headers:
            key = header.vendor_name or getattr(header.vendor, "accountname", "") or f"Vendor {header.vendor_id or header.id}"
            deductees[key] = deductees.get(key, False) or not bool((account_pan(header.vendor) or "").strip())
        return sum(1 for missing_pan in deductees.values() if missing_pan)

    def _build_challan_mapping_metrics(
        self,
        headers: list[PurchaseInvoiceHeader],
        header_mapping: dict[int, dict[str, object]],
    ) -> dict[str, int]:
        pending = 0
        unmapped = 0
        for header in headers:
            row_mapping = header_mapping.get(int(header.id), {})
            tds_amount = self._decimal(header.tds_amount)
            mapped = self._decimal(row_mapping.get("mapped"))
            remaining = max(tds_amount - mapped, ZERO2)
            if remaining > ZERO2:
                pending += 1
                if mapped <= ZERO2:
                    unmapped += 1
        return {"pending": pending, "unmapped": unmapped}

    def _build_pending_liability_preview(
        self,
        headers: list[PurchaseInvoiceHeader],
        header_mapping: dict[int, dict[str, object]],
    ) -> tuple[list[dict[str, object]], int]:
        pending_by_section: dict[str, dict[str, object]] = {}
        overdue_count = 0
        today = timezone.localdate()
        for header in headers:
            row_mapping = header_mapping.get(int(header.id), {})
            remaining = max(self._decimal(header.tds_amount) - self._decimal(row_mapping.get("deposited")), ZERO2)
            if remaining <= ZERO2:
                continue
            due_date = self._deposit_due_date(header.bill_date)
            delay_days = max((today - due_date).days, 0)
            if delay_days > 0:
                overdue_count += 1
            section = getattr(getattr(header, "tds_section", None), "section_code", None) or "UNSPECIFIED"
            bucket = pending_by_section.setdefault(
                section,
                {
                    "section": section,
                    "pendingAmount": ZERO2,
                    "interestAmount": ZERO2,
                    "maxDelayDays": 0,
                },
            )
            bucket["pendingAmount"] += remaining
            bucket["interestAmount"] += self._interest_estimate(remaining, delay_days)
            bucket["maxDelayDays"] = max(int(bucket["maxDelayDays"]), delay_days)

        rows = [
            {
                "section": section,
                "pendingAmount": self._money(bucket["pendingAmount"]),
                "interestAmount": self._money(bucket["interestAmount"]),
                "dueBucket": (
                    f"Overdue by {bucket['maxDelayDays']} days"
                    if int(bucket["maxDelayDays"]) > 0
                    else "Pending"
                ),
                "tone": "danger" if int(bucket["maxDelayDays"]) > 0 else "warning",
                "_sort_pending": bucket["pendingAmount"],
            }
            for section, bucket in pending_by_section.items()
        ]
        rows.sort(key=lambda row: (self._decimal(row["_sort_pending"]) * Decimal("-1"), row["section"]))
        for row in rows:
            row.pop("_sort_pending", None)
        return rows[:5], overdue_count

    def _build_filing_metrics(self, returns: list[PurchaseStatutoryReturn]) -> dict[str, int]:
        pending_returns = 0
        short_deduction_cases = 0
        filing_errors = 0
        for filing in returns:
            code = (filing.return_code or "").strip().upper()
            if code not in {"26Q", "27Q"}:
                continue
            status_label = self._status_label(self._return_status_badge(filing))
            if status_label not in {"Filed", "Revised"}:
                pending_returns += 1
            validation_errors = 0
            if code == "27Q":
                for line in filing.lines.all():
                    if not (line.deductee_tax_id_snapshot or "").strip():
                        validation_errors += 1
            short_deduction_cases += validation_errors
            filing_errors += validation_errors
        return {
            "pending_returns": pending_returns,
            "short_deduction_cases": short_deduction_cases,
            "filing_errors": filing_errors,
        }

    def _build_warning_chips(
        self,
        *,
        missing_pan_count: int,
        unmapped_count: int,
        overdue_count: int,
        filing_error_count: int,
    ) -> list[dict[str, object]]:
        chips: list[dict[str, object]] = []
        if missing_pan_count:
            chips.append({"label": f"Missing PAN {missing_pan_count}", "tone": "danger"})
        if unmapped_count:
            chips.append({"label": f"Unmapped Challans {unmapped_count}", "tone": "warning"})
        if overdue_count:
            chips.append({"label": f"Pending Deposit {overdue_count}", "tone": "warning"})
        if filing_error_count:
            chips.append({"label": f"Return Errors {filing_error_count}", "tone": "danger"})
        if not chips:
            chips.append({"label": "Real Data Synced", "tone": "success"})
        return chips

    def _build_kpis(
        self,
        summary: dict[str, str],
        *,
        missing_pan_count: int,
        pending_challans: int,
        pending_returns: int,
        short_deduction_cases: int,
    ) -> list[dict[str, object]]:
        return [
            {"code": "total_deducted", "label": "Total TDS Deducted", "value": self._money(summary["deducted"]), "tone": "info", "hint": "Selected period"},
            {"code": "total_deposited", "label": "Total TDS Deposited", "value": self._money(summary["deposited"]), "tone": "success", "hint": "Deposited challans"},
            {"code": "pending_liability", "label": "Pending Liability", "value": self._money(summary["pending_deposit"]), "tone": "warning", "hint": "Deducted but not deposited"},
            {"code": "interest_payable", "label": "Interest Payable", "value": self._money(summary["deposited_interest"]), "tone": "danger", "hint": "Deposited interest in period"},
            {"code": "pending_challans", "label": "Pending Challans", "value": str(pending_challans), "tone": "warning", "hint": "Unmapped or partial"},
            {"code": "vendors_missing_pan", "label": "Vendors Missing PAN", "value": str(missing_pan_count), "tone": "danger", "hint": "Needs master cleanup"},
            {"code": "returns_pending", "label": "Returns Pending", "value": str(pending_returns), "tone": "warning", "hint": "Draft or not filed"},
            {"code": "short_deduction_cases", "label": "Short Deduction Cases", "value": str(short_deduction_cases), "tone": "danger", "hint": "Validation cases"},
        ]

    def _build_dashboard(
        self,
        monthly_rows: list[dict[str, object]],
        section_rows: list[dict[str, object]],
        deductee_rows: list[dict[str, object]],
        pending_rows: list[dict[str, object]],
        audit_rows: list[dict[str, object]],
    ) -> dict[str, object]:
        return {
            "monthlyTrend": [
                {
                    "period": row.get("month"),
                    "deducted": row.get("tdsDeducted"),
                    "deposited": row.get("deposited"),
                    "pending": row.get("pending"),
                }
                for row in monthly_rows
            ],
            "sectionDistribution": [
                {
                    "section": row.get("section"),
                    "taxableAmount": row.get("taxableAmount"),
                    "tdsAmount": row.get("tdsAmount"),
                    "shareLabel": row.get("nature") or "Section summary",
                    "tone": self._status_tone(row.get("status")),
                }
                for row in section_rows[:6]
            ],
            "topDeductees": [
                {
                    "deductee": row.get("deductee"),
                    "section": row.get("defaultSection") or row.get("section") or "-",
                    "tdsAmount": row.get("tdsAmount"),
                    "complianceStatus": row.get("complianceStatus"),
                }
                for row in deductee_rows[:5]
            ],
            "pendingLiability": pending_rows[:5],
            "recentActivities": [
                {
                    "title": row.get("action"),
                    "detail": f"{row.get('voucherNo')} • {row.get('remarks') or row.get('newValue') or ''}".strip(),
                    "timestamp": row.get("dateTime"),
                    "tone": self._audit_tone(row.get("action")),
                }
                for row in audit_rows[:6]
            ],
        }

    def _build_datasets_payload(
        self,
        *,
        include_all_datasets: bool,
        requested_tabs: set[str],
        summary: dict[str, str],
        monthly_rows: list[dict[str, object]],
        header_rows: list[dict[str, object]],
        section_rows: list[dict[str, object]],
        challan_rows: list[dict[str, object]],
        challan_mapping_rows: list[dict[str, object]],
        deductee_rows: list[dict[str, object]],
        pending_rows: list[dict[str, object]],
        vendor_rows: list[dict[str, object]],
        filing_rows: dict[str, list[dict[str, object]]],
        form16a_rows: list[dict[str, object]],
        audit_rows: list[dict[str, object]],
    ) -> dict[str, object]:
        datasets = {
            "dashboard": self._dataset(
                columns=[
                    {"key": "month", "label": "Month", "type": "text"},
                    {"key": "tdsDeducted", "label": "TDS Deducted", "type": "currency", "align": "right"},
                    {"key": "deposited", "label": "Deposited", "type": "currency", "align": "right"},
                    {"key": "pending", "label": "Pending Liability", "type": "currency", "align": "right"},
                    {"key": "returnStatus", "label": "Status", "type": "status"},
                ],
                rows=monthly_rows,
                totals={
                    "primaryLabel": "Period Deducted",
                    "primaryValue": self._money(summary["deducted"]),
                    "secondaryLabel": "Pending Deposit",
                    "secondaryValue": self._money(summary["pending_deposit"]),
                },
            ),
            "deduction-register": self._dataset(
                columns=[
                    {"key": "date", "label": "Date", "type": "date"},
                    {"key": "voucherNo", "label": "Voucher No", "type": "text"},
                    {"key": "voucherType", "label": "Voucher Type", "type": "text"},
                    {"key": "deductee", "label": "Deductee", "type": "text"},
                    {"key": "pan", "label": "PAN", "type": "text"},
                    {"key": "section", "label": "Section", "type": "text"},
                    {"key": "taxableAmount", "label": "Taxable Amount", "type": "currency", "align": "right"},
                    {"key": "tdsAmount", "label": "TDS Amount", "type": "currency", "align": "right"},
                    {"key": "status", "label": "Status", "type": "status"},
                    {"key": "actions", "label": "Actions", "type": "actions", "sortable": False},
                ],
                rows=header_rows,
                totals={
                    "primaryLabel": "TDS Total",
                    "primaryValue": self._money(summary["deducted"]),
                },
            ),
            "payable-report": self._dataset(
                columns=[
                    {"key": "section", "label": "Section", "type": "text"},
                    {"key": "openingBalance", "label": "Opening Balance", "type": "currency", "align": "right"},
                    {"key": "currentDeduction", "label": "Current Deduction", "type": "currency", "align": "right"},
                    {"key": "deposited", "label": "Deposited", "type": "currency", "align": "right"},
                    {"key": "interest", "label": "Interest", "type": "currency", "align": "right"},
                    {"key": "closingBalance", "label": "Closing Balance", "type": "currency", "align": "right"},
                    {"key": "status", "label": "Status", "type": "status"},
                    {"key": "actions", "label": "Actions", "type": "actions", "sortable": False},
                ],
                rows=section_rows,
                totals={
                    "primaryLabel": "Closing Balance",
                    "primaryValue": self._money(sum((self._decimal(row.get("closingBalance")) for row in section_rows), ZERO2)),
                    "secondaryLabel": "Interest",
                    "secondaryValue": self._money(sum((self._decimal(row.get("interest")) for row in section_rows), ZERO2)),
                },
            ),
            "payment-register": self._dataset(
                columns=[
                    {"key": "challanNo", "label": "Challan No", "type": "text"},
                    {"key": "cin", "label": "CIN", "type": "text"},
                    {"key": "bank", "label": "Bank", "type": "text"},
                    {"key": "depositDate", "label": "Deposit Date", "type": "date"},
                    {"key": "amount", "label": "Amount", "type": "currency", "align": "right"},
                    {"key": "section", "label": "Section", "type": "text"},
                    {"key": "status", "label": "Status", "type": "status"},
                    {"key": "actions", "label": "Actions", "type": "actions", "sortable": False},
                ],
                rows=challan_rows,
            ),
            "challan-mapping": self._dataset(
                columns=[
                    {"key": "voucherNo", "label": "Voucher No", "type": "text"},
                    {"key": "deductee", "label": "Deductee", "type": "text"},
                    {"key": "section", "label": "Section", "type": "text"},
                    {"key": "tdsAmount", "label": "TDS Amount", "type": "currency", "align": "right"},
                    {"key": "mappedChallan", "label": "Mapped Challan", "type": "text"},
                    {"key": "remainingAmount", "label": "Remaining Amount", "type": "currency", "align": "right"},
                    {"key": "mappingStatus", "label": "Mapping Status", "type": "status"},
                    {"key": "actions", "label": "Actions", "type": "actions", "sortable": False},
                ],
                rows=challan_mapping_rows,
            ),
            "section-wise-summary": self._dataset(
                columns=[
                    {"key": "section", "label": "Section", "type": "text"},
                    {"key": "nature", "label": "Nature Of Payment", "type": "text"},
                    {"key": "transactions", "label": "Transactions", "type": "number", "align": "right"},
                    {"key": "taxableAmount", "label": "Taxable Amount", "type": "currency", "align": "right"},
                    {"key": "tdsAmount", "label": "TDS Amount", "type": "currency", "align": "right"},
                    {"key": "pendingAmount", "label": "Pending Amount", "type": "currency", "align": "right"},
                ],
                rows=section_rows,
            ),
            "deductee-wise-summary": self._dataset(
                columns=[
                    {"key": "deductee", "label": "Deductee", "type": "text"},
                    {"key": "pan", "label": "PAN", "type": "text"},
                    {"key": "transactions", "label": "Transactions", "type": "number", "align": "right"},
                    {"key": "taxableAmount", "label": "Taxable Amount", "type": "currency", "align": "right"},
                    {"key": "tdsAmount", "label": "TDS Amount", "type": "currency", "align": "right"},
                    {"key": "pending", "label": "Pending", "type": "currency", "align": "right"},
                    {"key": "complianceStatus", "label": "Compliance", "type": "status"},
                ],
                rows=deductee_rows,
            ),
            "monthly-summary": self._dataset(
                columns=[
                    {"key": "month", "label": "Month", "type": "text"},
                    {"key": "taxableAmount", "label": "Taxable Amount", "type": "currency", "align": "right"},
                    {"key": "tdsDeducted", "label": "TDS Deducted", "type": "currency", "align": "right"},
                    {"key": "deposited", "label": "Deposited", "type": "currency", "align": "right"},
                    {"key": "pending", "label": "Pending", "type": "currency", "align": "right"},
                    {"key": "returnStatus", "label": "Return Status", "type": "status"},
                ],
                rows=monthly_rows,
            ),
            "pending-payment": self._dataset(
                columns=[
                    {"key": "dueDate", "label": "Due Date", "type": "date"},
                    {"key": "section", "label": "Section", "type": "text"},
                    {"key": "deductee", "label": "Deductee", "type": "text"},
                    {"key": "tdsAmount", "label": "TDS Amount", "type": "currency", "align": "right"},
                    {"key": "delayDays", "label": "Delay Days", "type": "number", "align": "right"},
                    {"key": "interest", "label": "Interest", "type": "currency", "align": "right"},
                    {"key": "status", "label": "Status", "type": "status"},
                ],
                rows=pending_rows,
            ),
            "vendor-compliance": self._dataset(
                columns=[
                    {"key": "deductee", "label": "Deductee", "type": "text"},
                    {"key": "pan", "label": "PAN", "type": "text"},
                    {"key": "panStatus", "label": "PAN Status", "type": "status"},
                    {"key": "defaultSection", "label": "Default TDS Section", "type": "text"},
                    {"key": "certificate", "label": "Lower / NIL Certificate", "type": "text"},
                    {"key": "validity", "label": "Certificate Validity", "type": "date"},
                    {"key": "complianceStatus", "label": "Compliance", "type": "status"},
                ],
                rows=vendor_rows,
            ),
            "return-filing": self._dataset(
                columns=self._return_columns("26q"),
                rows=filing_rows["26q"],
            ),
            "form-16a": self._dataset(
                columns=[
                    {"key": "deductee", "label": "Deductee", "type": "text"},
                    {"key": "pan", "label": "PAN", "type": "text"},
                    {"key": "quarter", "label": "Quarter", "type": "text"},
                    {"key": "certificateNo", "label": "Certificate No", "type": "text"},
                    {"key": "generatedDate", "label": "Generated Date", "type": "date"},
                    {"key": "emailStatus", "label": "Email Status", "type": "status"},
                    {"key": "downloadStatus", "label": "Download Status", "type": "status"},
                    {"key": "actions", "label": "Actions", "type": "actions", "sortable": False},
                ],
                rows=form16a_rows,
            ),
            "audit-trail": self._dataset(
                columns=[
                    {"key": "dateTime", "label": "Date Time", "type": "date"},
                    {"key": "user", "label": "User", "type": "text"},
                    {"key": "action", "label": "Action", "type": "text"},
                    {"key": "voucherNo", "label": "Voucher No", "type": "text"},
                    {"key": "field", "label": "Field", "type": "text"},
                    {"key": "oldValue", "label": "Old Value", "type": "text"},
                    {"key": "newValue", "label": "New Value", "type": "text"},
                    {"key": "remarks", "label": "Remarks", "type": "text"},
                    {"key": "ipAddress", "label": "IP Address", "type": "text"},
                ],
                rows=audit_rows,
            ),
        }
        if include_all_datasets:
            return datasets
        return {
            tab_id: dataset
            for tab_id, dataset in datasets.items()
            if tab_id in requested_tabs
        }

    def _build_return_datasets_payload(
        self,
        *,
        include_all_return_datasets: bool,
        requested_return_tabs: set[str],
        filing_rows: dict[str, list[dict[str, object]]],
    ) -> dict[str, object]:
        datasets = {
            "24q": self._dataset(
                columns=self._return_columns("24q"),
                rows=[],
                empty_message="No salary-side 24Q data is available in the purchase statutory TDS source.",
                return_tab="24q",
            ),
            "26q": self._dataset(
                columns=self._return_columns("26q"),
                rows=filing_rows["26q"],
                return_tab="26q",
            ),
            "27q": self._dataset(
                columns=self._return_columns("27q"),
                rows=filing_rows["27q"],
                return_tab="27q",
            ),
        }
        if include_all_return_datasets:
            return datasets
        return {
            tab_id: dataset
            for tab_id, dataset in datasets.items()
            if tab_id in requested_return_tabs
        }

    def _dataset(
        self,
        *,
        columns: list[dict[str, object]],
        rows: list[dict[str, object]],
        totals: Optional[dict[str, object]] = None,
        empty_message: Optional[str] = None,
        return_tab: Optional[str] = None,
    ) -> dict[str, object]:
        payload = {
            "columns": columns,
            "rows": rows,
            "pagination": {
                "page": 1,
                "pageSize": self.page_size,
                "totalRows": len(rows),
            },
            "totals": totals,
            "emptyMessage": empty_message or "No rows are available for the selected scope.",
        }
        if return_tab is not None:
            payload["returnTab"] = return_tab
        return payload

    def _query_bool(self, raw_value: Optional[str], *, default: bool) -> bool:
        if raw_value is None:
            return default
        token = str(raw_value).strip().lower()
        if token in {"1", "true", "yes", "y"}:
            return True
        if token in {"0", "false", "no", "n"}:
            return False
        return default

    def _requested_tabs(self, raw_value: Optional[str]) -> set[str]:
        if not raw_value:
            return set()
        return {
            token.strip().lower()
            for token in str(raw_value).split(",")
            if token.strip()
        }

    def _tabs(self) -> list[dict[str, str]]:
        return [
            {"id": "dashboard", "label": "Dashboard", "shortLabel": "Dashboard"},
            {"id": "deduction-register", "label": "Deduction Register", "shortLabel": "Deduction"},
            {"id": "payable-report", "label": "Payable Report", "shortLabel": "Payable"},
            {"id": "payment-register", "label": "Payment Register", "shortLabel": "Payment"},
            {"id": "challan-mapping", "label": "Challan Mapping", "shortLabel": "Mapping"},
            {"id": "section-wise-summary", "label": "Section Wise Summary", "shortLabel": "Section"},
            {"id": "deductee-wise-summary", "label": "Deductee Wise Summary", "shortLabel": "Deductee"},
            {"id": "monthly-summary", "label": "Monthly Summary", "shortLabel": "Monthly"},
            {"id": "pending-payment", "label": "Pending Payment", "shortLabel": "Pending"},
            {"id": "vendor-compliance", "label": "Vendor Compliance", "shortLabel": "Vendor"},
            {"id": "return-filing", "label": "Return Filing", "shortLabel": "Returns"},
            {"id": "form-16a", "label": "Form 16A", "shortLabel": "Form 16A"},
            {"id": "audit-trail", "label": "Audit Trail", "shortLabel": "Audit"},
        ]

    def _return_tabs(self) -> list[dict[str, str]]:
        return [
            {"id": "24q", "label": "24Q Salary", "shortLabel": "24Q"},
            {"id": "26q", "label": "26Q Resident Payments", "shortLabel": "26Q"},
            {"id": "27q", "label": "27Q Non-Resident Payments", "shortLabel": "27Q"},
        ]

    def _return_columns(self, return_tab: str) -> list[dict[str, object]]:
        base = [
            {"key": "quarter", "label": "Quarter", "type": "text"},
            {"key": "returnType", "label": "Return Type", "type": "text"},
            {"key": "transactions", "label": "Transactions", "type": "number", "align": "right"},
            {"key": "deductees", "label": "Deductees", "type": "number", "align": "right"},
            {"key": "totalTaxableAmount", "label": "Total Taxable Amount", "type": "currency", "align": "right"},
            {"key": "totalTds", "label": "Total TDS", "type": "currency", "align": "right"},
            {"key": "validationErrors", "label": "Validation Errors", "type": "number", "align": "right"},
            {"key": "warnings", "label": "Warnings", "type": "number", "align": "right"},
            {"key": "fvuStatus", "label": "FVU Status", "type": "status"},
            {"key": "returnStatus", "label": "Return Status", "type": "status"},
            {"key": "tokenNumber", "label": "Token Number", "type": "text"},
            {"key": "filingDate", "label": "Filing Date", "type": "date"},
            {"key": "actions", "label": "Actions", "type": "actions", "sortable": False},
        ]
        if return_tab != "27q":
            return base
        return base + [
            {"key": "country", "label": "Country", "type": "text"},
            {"key": "taxResidencyCountry", "label": "Tax Residency Country", "type": "text"},
            {"key": "tin", "label": "TIN", "type": "text"},
            {"key": "currency", "label": "Currency", "type": "text"},
            {"key": "foreignAmount", "label": "Foreign Amount", "type": "currency", "align": "right"},
            {"key": "exchangeRate", "label": "Exchange Rate", "type": "number", "align": "right"},
            {"key": "inrAmount", "label": "INR Amount", "type": "currency", "align": "right"},
            {"key": "dtaaApplicable", "label": "DTAA Applicable", "type": "text"},
            {"key": "dtaaArticle", "label": "DTAA Article", "type": "text"},
            {"key": "surcharge", "label": "Surcharge", "type": "currency", "align": "right"},
            {"key": "cess", "label": "Cess", "type": "currency", "align": "right"},
        ]

    def _new_section_bucket(self, header: PurchaseInvoiceHeader) -> dict[str, object]:
        return {
            "nature": getattr(getattr(header, "tds_section", None), "description", None) or "TDS Section",
            "opening": ZERO2,
            "current": ZERO2,
            "deposited": ZERO2,
            "interest": ZERO2,
            "transactions": 0,
            "taxable": ZERO2,
        }

    def _financial_year_label(self, financial_year: EntityFinancialYear) -> str:
        return str(getattr(financial_year, "desc", "") or f"FY {financial_year.finstartyear.year}-{financial_year.finendyear.year}")

    def _quarter_options(self) -> list[dict[str, str]]:
        return [
            {"value": "Q1", "label": "Q1 Apr-Jun"},
            {"value": "Q2", "label": "Q2 Jul-Sep"},
            {"value": "Q3", "label": "Q3 Oct-Dec"},
            {"value": "Q4", "label": "Q4 Jan-Mar"},
        ]

    def _quarter_ranges(self, fy_start: date) -> dict[str, tuple[date, date]]:
        fy_start = self._as_date(fy_start) or timezone.localdate()
        q1_start = fy_start
        q2_start = date(fy_start.year, 7, 1)
        q3_start = date(fy_start.year, 10, 1)
        q4_start = date(fy_start.year + 1, 1, 1)
        return {
            "Q1": (q1_start, date(fy_start.year, 6, 30)),
            "Q2": (q2_start, date(fy_start.year, 9, 30)),
            "Q3": (q3_start, date(fy_start.year, 12, 31)),
            "Q4": (q4_start, date(fy_start.year + 1, 3, 31)),
        }

    def _quarter_for_date(self, dt: date, fy_start: date, fy_end: date) -> str:
        dt = self._as_date(dt) or timezone.localdate()
        fy_start = self._as_date(fy_start) or dt
        fy_end = self._as_date(fy_end) or dt
        quarter_ranges = self._quarter_ranges(fy_start)
        for code, (start, end) in quarter_ranges.items():
            if start <= dt <= end:
                return code
        if dt < fy_start:
            return "Q1"
        if dt > fy_end:
            return "Q4"
        return "Q1"

    def _quarter_label(self, quarter_code: str) -> str:
        return next((option["label"] for option in self._quarter_options() if option["value"] == quarter_code), quarter_code)

    def _normalize_financial_years(self, rows: list[dict[str, object]]) -> list[dict[str, object]]:
        normalized: list[dict[str, object]] = []
        for row in rows:
            item = dict(row)
            item["financialyear"] = row.get("desc") or f"FY {row.get('finstartyear')} to {row.get('finendyear')}"
            normalized.append(item)
        return normalized

    def _parse_optional_date(self, raw_value: Optional[str]) -> Optional[date]:
        if raw_value in (None, "", "null", "None"):
            return None
        try:
            return date.fromisoformat(str(raw_value))
        except ValueError:
            return None

    def _as_date(self, value: object) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def _voucher_label(self, header: PurchaseInvoiceHeader) -> str:
        return header.purchase_number or f"{header.doc_code}-{header.doc_no}"

    def _voucher_type_label(self, doc_type: int) -> str:
        if int(doc_type) == int(PurchaseInvoiceHeader.DocType.CREDIT_NOTE):
            return "Purchase Credit Note"
        if int(doc_type) == int(PurchaseInvoiceHeader.DocType.DEBIT_NOTE):
            return "Purchase Debit Note"
        return "Purchase Invoice"

    def _return_type_label(self, code: str) -> str:
        labels = {
            "24Q": "24Q Salary",
            "26Q": "26Q Resident Payments",
            "27Q": "27Q Non-Resident Payments",
        }
        return labels.get(code, code)

    def _header_status_badge(self, *, tds_amount: Decimal, deposited_amount: Decimal, pan: Optional[str]) -> dict[str, str]:
        if not (pan or "").strip():
            return self._badge("Missing PAN", "danger")
        if deposited_amount >= tds_amount > ZERO2:
            return self._badge("Paid", "success")
        if deposited_amount > ZERO2:
            return self._badge("Partially Mapped", "warning")
        return self._badge("Pending Deposit", "warning")

    def _challan_status_badge(self, challan: PurchaseStatutoryChallan) -> dict[str, str]:
        if int(challan.status) == int(PurchaseStatutoryChallan.Status.DEPOSITED):
            return self._badge("Deposited", "success")
        if int(challan.status) == int(PurchaseStatutoryChallan.Status.CANCELLED):
            return self._badge("Cancelled", "neutral")
        return self._draft_approval_badge(getattr(challan, "payment_payload_json", None))

    def _return_status_badge(self, filing: PurchaseStatutoryReturn) -> dict[str, str]:
        if int(filing.status) == int(PurchaseStatutoryReturn.Status.FILED):
            return self._badge("Filed", "success")
        if int(filing.status) == int(PurchaseStatutoryReturn.Status.REVISED):
            return self._badge("Revised", "info")
        if int(filing.status) == int(PurchaseStatutoryReturn.Status.CANCELLED):
            return self._badge("Cancelled", "neutral")
        return self._draft_approval_badge(getattr(filing, "filed_payload_json", None))

    def _draft_approval_badge(self, payload: object) -> dict[str, str]:
        approval_state = PurchaseStatutoryService._approval_state(payload if isinstance(payload, dict) else None)
        status = str(approval_state.get("status") or "DRAFT").upper()
        if status == "SUBMITTED":
            return self._badge("Approval Submitted", "info")
        if status == "APPROVED":
            return self._badge("Approved Draft", "success")
        if status == "REJECTED":
            return self._badge("Rejected Draft", "danger")
        return self._badge("Draft", "warning")

    def _challan_primary_action_label(self, challan: PurchaseStatutoryChallan) -> str:
        if int(challan.status) == int(PurchaseStatutoryChallan.Status.CANCELLED):
            return "Audit only"
        if int(challan.status) == int(PurchaseStatutoryChallan.Status.DEPOSITED):
            return "Create return"
        approval_state = PurchaseStatutoryService._approval_state(getattr(challan, "payment_payload_json", None))
        approval_code = str(approval_state.get("status") or "DRAFT").upper()
        if approval_code == "SUBMITTED":
            return "Approve draft"
        if approval_code == "APPROVED":
            return "Deposit"
        if approval_code == "REJECTED":
            return "Review rejected"
        return "Review draft"

    def _return_primary_action_label(self, filing: PurchaseStatutoryReturn) -> str:
        if int(filing.status) == int(PurchaseStatutoryReturn.Status.CANCELLED):
            return "Audit only"
        if int(filing.status) in (int(PurchaseStatutoryReturn.Status.FILED), int(PurchaseStatutoryReturn.Status.REVISED)):
            return "NSDL / 16A" if self._has_it_tds_filed_follow_up(filing) else "Filing follow-up"
        approval_state = PurchaseStatutoryService._approval_state(getattr(filing, "filed_payload_json", None))
        approval_code = str(approval_state.get("status") or "DRAFT").upper()
        if approval_code == "SUBMITTED":
            return "Approve draft"
        if approval_code == "APPROVED":
            return "File"
        if approval_code == "REJECTED":
            return "Review rejected"
        return "Review draft"

    def _has_it_tds_filed_follow_up(self, filing: PurchaseStatutoryReturn) -> bool:
        if filing.tax_type != PurchaseStatutoryReturn.TaxType.IT_TDS:
            return False
        code = str(getattr(filing, "return_code", "") or "").strip().upper()
        return code in {"26Q", "27Q"}

    def _balance_status(self, closing: Decimal, interest: Decimal) -> dict[str, str]:
        if closing <= ZERO2:
            return self._badge("Paid", "success")
        if interest > ZERO2:
            return self._badge("Overdue", "danger")
        return self._badge("Pending", "warning")

    def _badge(self, label: str, tone: str) -> dict[str, str]:
        return {"label": label, "tone": tone}

    def _actions(self, primary: str, secondary: Optional[str] = None) -> list[dict[str, str]]:
        actions = [{"label": primary, "kind": "primary"}]
        if secondary:
            actions.append({"label": secondary})
        return actions

    def _status_label(self, badge: object) -> str:
        if isinstance(badge, dict):
            return str(badge.get("label") or "")
        return str(badge or "")

    def _status_tone(self, badge: object) -> str:
        if isinstance(badge, dict):
            return str(badge.get("tone") or "neutral")
        return "neutral"

    def _audit_tone(self, action: object) -> str:
        token = str(action or "").lower()
        if "filed" in token or "deposited" in token:
            return "success"
        if "review" in token:
            return "info"
        return "warning"

    def _deposit_due_date(self, bill_date: date) -> date:
        if bill_date.month == 12:
            next_month = date(bill_date.year + 1, 1, 1)
        else:
            next_month = date(bill_date.year, bill_date.month + 1, 1)
        return next_month + timedelta(days=6)

    def _interest_estimate(self, amount: Decimal, delay_days: int) -> Decimal:
        if delay_days <= 0 or amount <= ZERO2:
            return ZERO2
        monthly_interest = Decimal("0.015")
        return self._quantize(amount * monthly_interest * Decimal(delay_days) / Decimal("30"))

    def _section_share(self, *, amount: Decimal, total: Decimal, distributed: Decimal) -> Decimal:
        if total <= ZERO2 or distributed <= ZERO2 or amount <= ZERO2:
            return ZERO2
        return self._quantize(distributed * amount / total)

    def _money(self, value: object) -> str:
        return f"{self._decimal(value):.2f}"

    def _decimal(self, value: object) -> Decimal:
        if value in (None, "", "-"):
            return ZERO2
        if isinstance(value, Decimal):
            return self._quantize(value)
        return self._quantize(Decimal(str(value).replace(",", "")))

    def _quantize(self, value: Decimal) -> Decimal:
        return value.quantize(Decimal("0.01"))

    def _display_date(self, dt: date) -> str:
        return dt.strftime("%d %b %Y")

    def _iso(self, value: Optional[date]) -> Optional[str]:
        return value.isoformat() if value else None

    def _iso_datetime(self, value) -> Optional[str]:
        if value is None:
            return None
        try:
            return value.isoformat()
        except Exception:
            return str(value)

    def _user_label(self, user) -> str:
        if user is None:
            return "-"
        full_name = " ".join([part for part in [getattr(user, "first_name", ""), getattr(user, "last_name", "")] if part]).strip()
        return full_name or getattr(user, "email", None) or getattr(user, "username", None) or f"User {getattr(user, 'id', '')}"

class PurchaseTdsComplianceCenterExportAPIView(PurchaseTdsComplianceCenterAPIView):
    def get(self, request):
        export_format = str(request.query_params.get("format") or "xlsx").strip().lower()
        if export_format == "ca-pack":
            response = super().get(request)
            payload = response.data
            content = self._write_ca_pack(payload, request)
            return self._export_response(
                filename=f"{self._resolve_ca_pack_filename(payload)}.xlsx",
                content=content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        tab_id = str(request.query_params.get("tab") or "dashboard").strip()
        return_tab = str(request.query_params.get("return_tab") or "26q").strip().lower()
        selected_columns = [
            value.strip()
            for value in str(request.query_params.get("columns") or "").split(",")
            if value.strip()
        ]
        sort_field = str(request.query_params.get("sort_field") or "").strip()
        sort_direction = str(request.query_params.get("sort_direction") or "asc").strip().lower()

        response = super().get(request)
        payload = response.data

        dataset = self._resolve_export_dataset(payload, tab_id, return_tab)
        if dataset is None:
            return Response({"detail": "Requested TDS export tab is not available."}, status=404)

        columns = self._resolve_export_columns(dataset, selected_columns)
        rows = self._export_rows(payload=payload, dataset=dataset, request=request, sort_field=sort_field, sort_direction=sort_direction)
        title = self._resolve_export_title(payload, tab_id, return_tab)
        subtitle = self._resolve_export_subtitle(payload)
        filename_root = self._resolve_export_filename(payload, tab_id, return_tab)

        headers = [str(column.get("label") or column.get("key") or "") for column in columns]
        body_rows = [self._format_export_row(row, columns) for row in rows]
        numeric_columns = {
            index + 1
            for index, column in enumerate(columns)
            if str(column.get("type") or "").lower() in {"currency", "number"}
        }

        if export_format == "csv":
            content = _write_csv(headers, body_rows)
            return self._export_response(filename=f"{filename_root}.csv", content=content, content_type="text/csv")
        if export_format == "pdf":
            content = _write_pdf(title, subtitle, headers, body_rows)
            return self._export_response(filename=f"{filename_root}.pdf", content=content, content_type="application/pdf")

        content = _write_excel(title, subtitle, headers, body_rows, numeric_columns=numeric_columns)
        return self._export_response(
            filename=f"{filename_root}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _resolve_export_dataset(self, payload: dict[str, object], tab_id: str, return_tab: str) -> Optional[dict[str, object]]:
        if tab_id == "return-filing":
            return payload.get("returnDatasets", {}).get(return_tab)  # type: ignore[union-attr]
        return payload.get("datasets", {}).get(tab_id)  # type: ignore[union-attr]

    def _resolve_export_columns(self, dataset: dict[str, object], selected_columns: list[str]) -> list[dict[str, object]]:
        dataset_columns = [column for column in list(dataset.get("columns") or []) if column.get("type") != "actions"]
        if not selected_columns:
            return dataset_columns

        lookup = {str(column.get("key") or ""): column for column in dataset_columns}
        resolved: list[dict[str, object]] = []
        for key in selected_columns:
            column = lookup.get(key)
            if column is not None:
                resolved.append(column)
        return resolved or dataset_columns

    def _export_rows(
        self,
        *,
        payload: dict[str, object],
        dataset: dict[str, object],
        request,
        sort_field: str,
        sort_direction: str,
    ) -> list[dict[str, object]]:
        rows = list(dataset.get("rows") or [])
        rows = self._apply_export_filters(payload=payload, rows=rows, request=request)
        return self._apply_export_sort(rows=rows, sort_field=sort_field, sort_direction=sort_direction)

    def _apply_export_filters(self, *, payload: dict[str, object], rows: list[dict[str, object]], request) -> list[dict[str, object]]:
        search_text = str(request.query_params.get("search") or request.query_params.get("searchText") or "").strip().lower()
        pan = str(request.query_params.get("pan") or "").strip().lower()
        voucher_type = str(request.query_params.get("voucherType") or "").strip().lower()
        payment_status = str(request.query_params.get("paymentStatus") or "").strip().lower()
        challan_status = str(request.query_params.get("challanStatus") or "").strip().lower()
        return_status = str(request.query_params.get("returnStatus") or "").strip().lower()
        min_amount = self._parse_export_decimal(request.query_params.get("minAmount"))
        max_amount = self._parse_export_decimal(request.query_params.get("maxAmount"))
        section_lookup = self._resolve_export_section_lookup(payload, request.query_params.get("tdsSectionId"))

        filtered: list[dict[str, object]] = []
        for row in rows:
            if search_text and not self._row_matches_search(row, search_text):
                continue
            if pan and pan not in str(row.get("pan") or "").lower():
                continue
            if voucher_type and voucher_type not in str(row.get("voucherType") or "").lower():
                continue
            if payment_status and payment_status not in self._status_label(row.get("status")).lower():
                continue
            if challan_status and challan_status not in self._status_label(row.get("mappingStatus")).lower():
                continue
            if return_status and return_status not in self._status_label(row.get("returnStatus")).lower():
                continue
            if section_lookup and section_lookup not in str(row.get("section") or "").lower():
                continue

            numeric_value = self._resolve_export_numeric_value(row)
            if min_amount is not None and numeric_value is not None and numeric_value < min_amount:
                continue
            if max_amount is not None and numeric_value is not None and numeric_value > max_amount:
                continue
            filtered.append(row)

        return filtered

    def _apply_export_sort(self, *, rows: list[dict[str, object]], sort_field: str, sort_direction: str) -> list[dict[str, object]]:
        if not sort_field:
            return rows
        reverse = sort_direction == "desc"
        return sorted(rows, key=lambda row: self._sortable_export_value(row.get(sort_field)), reverse=reverse)

    def _resolve_export_title(self, payload: dict[str, object], tab_id: str, return_tab: str) -> str:
        if tab_id == "return-filing":
            for tab in list(payload.get("returnTabs") or []):
                if str(tab.get("id") or "").lower() == return_tab:
                    return f"TDS Compliance Center - {tab.get('label') or return_tab.upper()}"
        for tab in list(payload.get("tabs") or []):
            if str(tab.get("id") or "") == tab_id:
                return f"TDS Compliance Center - {tab.get('label') or tab_id}"
        return "TDS Compliance Center"

    def _resolve_export_subtitle(self, payload: dict[str, object]) -> str:
        return " | ".join(
            str(item.get("label") or "").strip()
            for item in list(payload.get("headerChips") or [])
            if str(item.get("label") or "").strip()
        )

    def _resolve_export_filename(self, payload: dict[str, object], tab_id: str, return_tab: str) -> str:
        quarter = str(payload.get("filters", {}).get("quarter") or "scope")  # type: ignore[union-attr]
        tab_token = f"{tab_id}-{return_tab}" if tab_id == "return-filing" else tab_id
        return _safe_filename(f"tds_compliance_{tab_token}_{quarter}")

    def _resolve_ca_pack_filename(self, payload: dict[str, object]) -> str:
        quarter = str(payload.get("filters", {}).get("quarter") or "scope")  # type: ignore[union-attr]
        return _safe_filename(f"{self._ca_pack_prefix()}_ca_pack_{quarter}")

    def _ca_pack_prefix(self) -> str:
        return "tds_compliance"

    def _ca_pack_title(self) -> str:
        return "TDS Compliance Center CA Pack"

    def _write_ca_pack(self, payload: dict[str, object], request) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        def create_sheet(title: str):
            return wb.create_sheet(title=title[:31])

        def autosize_sheet(ws):
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 12
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    value = row[0].value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

        def decorate_table(ws, header_row: int = 1):
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            for cell in ws[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            for row_idx in range(header_row + 1, ws.max_row + 1):
                for cell in ws[row_idx]:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top")
            ws.freeze_panes = f"A{header_row + 1}"
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            autosize_sheet(ws)

        def add_cover_sheet():
            ws = create_sheet("00_Cover")
            ws["A1"] = self._ca_pack_title()
            ws["A2"] = "Entity"
            ws["B2"] = self._chip_label(payload, 0)
            ws["A3"] = "Financial Year"
            ws["B3"] = self._chip_label(payload, 1)
            ws["A4"] = "Quarter"
            ws["B4"] = self._chip_label(payload, 2)
            ws["A5"] = "Branch"
            ws["B5"] = self._chip_label(payload, 3)
            ws["A6"] = "Period"
            ws["B6"] = self._chip_label(payload, 4)
            ws["A7"] = "Generated At"
            ws["B7"] = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
            ws["A8"] = "Generated By"
            ws["B8"] = getattr(request.user, "username", "") or getattr(request.user, "email", "") or getattr(request.user, "id", "")
            ws["A10"] = "Included Sheets"
            included = [
                "01_KPI_Summary",
                "02_Warnings",
                "03_Dashboard",
            ] + [self._sheet_title_from_label(index + 4, str(tab.get("label") or tab.get("id") or "Tab")) for index, tab in enumerate(list(payload.get("tabs") or []))] \
                + [self._sheet_title_from_label(index + 20, f"Return {str(tab.get('label') or tab.get('id') or 'Tab')}") for index, tab in enumerate(list(payload.get("returnTabs") or []))]
            for row_idx, name in enumerate(included, start=11):
                ws[f"A{row_idx}"] = name
            ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
            ws["A10"].font = Font(size=12, bold=True)
            for row_idx in range(2, 9):
                ws[f"A{row_idx}"].font = Font(bold=True)
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 42

        def add_kpi_sheet():
            ws = create_sheet("01_KPI_Summary")
            ws.append(["Metric", "Value", "Tone"])
            for kpi in list(payload.get("kpis") or []):
                ws.append([str(kpi.get("label") or ""), str(kpi.get("value") or ""), str(kpi.get("tone") or "")])
            decorate_table(ws)

        def add_warning_sheet():
            ws = create_sheet("02_Warnings")
            ws.append(["Warning", "Tone"])
            warnings = list(payload.get("warnings") or [])
            if warnings:
                for warning in warnings:
                    ws.append([str(warning.get("label") or ""), str(warning.get("tone") or "")])
            else:
                ws.append(["No active warnings in the selected scope.", ""])
            decorate_table(ws)

        def add_dataset_sheet(sheet_title: str, title: str, dataset: dict[str, object]):
            ws = create_sheet(sheet_title)
            ws["A1"] = title
            ws["A2"] = self._resolve_export_subtitle(payload)
            columns = self._resolve_export_columns(dataset, [])
            headers = [str(column.get("label") or column.get("key") or "") for column in columns]
            row_offset = 4
            for index, header in enumerate(headers, start=1):
                ws.cell(row=row_offset, column=index, value=header)
            export_rows = self._export_rows(payload=payload, dataset=dataset, request=request, sort_field="", sort_direction="asc")
            data_row = row_offset + 1
            for row in export_rows:
                for index, value in enumerate(self._format_export_row(row, columns), start=1):
                    ws.cell(row=data_row, column=index, value=value)
                data_row += 1
            if not export_rows:
                ws.append(["No rows available in the selected scope."])
            decorate_table(ws, header_row=row_offset)

        add_cover_sheet()
        add_kpi_sheet()

        dashboard_dataset = payload.get("datasets", {}).get("dashboard")  # type: ignore[union-attr]
        add_warning_sheet()
        if isinstance(dashboard_dataset, dict):
            add_dataset_sheet("03_Dashboard", f"{payload.get('pageTitle') or 'Compliance Center'} - Dashboard", dashboard_dataset)

        for index, tab in enumerate(list(payload.get("tabs") or []), start=4):
            tab_id = str(tab.get("id") or "")
            dataset = payload.get("datasets", {}).get(tab_id)  # type: ignore[union-attr]
            if not isinstance(dataset, dict):
                continue
            add_dataset_sheet(
                self._sheet_title_from_label(index, str(tab.get("label") or tab_id)),
                self._resolve_export_title(payload, tab_id, "26q"),
                dataset,
            )

        for index, tab in enumerate(list(payload.get("returnTabs") or []), start=20):
            return_tab = str(tab.get("id") or "").lower()
            dataset = payload.get("returnDatasets", {}).get(return_tab)  # type: ignore[union-attr]
            if not isinstance(dataset, dict):
                continue
            add_dataset_sheet(
                self._sheet_title_from_label(index, f"Return {str(tab.get('label') or return_tab)}"),
                self._resolve_export_title(payload, "return-filing", return_tab),
                dataset,
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    def _chip_label(self, payload: dict[str, object], index: int) -> str:
        chips = list(payload.get("headerChips") or [])
        if index >= len(chips):
            return ""
        chip = chips[index]
        return str(chip.get("label") or "") if isinstance(chip, dict) else str(chip or "")

    def _sheet_title_from_label(self, index: int, label: str) -> str:
        cleaned = "".join(char if char.isalnum() else "_" for char in str(label or "").strip()).strip("_")
        cleaned = cleaned[:26] or "Sheet"
        return f"{index:02d}_{cleaned}"[:31]

    def _format_export_row(self, row: dict[str, object], columns: list[dict[str, object]]) -> list[object]:
        return [self._format_export_cell(row.get(str(column.get("key") or "")), str(column.get("type") or "")) for column in columns]

    def _format_export_cell(self, value: object, column_type: str) -> object:
        normalized_type = column_type.lower()
        if normalized_type == "status":
            return self._status_label(value)
        if normalized_type == "date":
            parsed = self._parse_optional_date(str(value)) if value not in (None, "", "-") else None
            if parsed is not None:
                return parsed.strftime("%d-%m-%Y")
            return str(value or "")
        if normalized_type in {"currency", "number"}:
            return float(self._decimal(value))
        if isinstance(value, list):
            return ", ".join(str(item.get("label") or "") if isinstance(item, dict) else str(item or "") for item in value)
        if isinstance(value, dict):
            return str(value.get("label") or value.get("value") or "")
        return "" if value is None else str(value)

    def _resolve_export_section_lookup(self, payload: dict[str, object], raw_section_id: object) -> str:
        if raw_section_id in (None, "", "null", "None"):
            return ""
        try:
            section_id = int(str(raw_section_id))
        except (TypeError, ValueError):
            return ""
        for section in list(payload.get("meta", {}).get("tdsSections") or []):  # type: ignore[union-attr]
            if int(section.get("id") or 0) == section_id:
                return str(section.get("section_code") or "").lower()
        return ""

    def _row_matches_search(self, row: dict[str, object], search_text: str) -> bool:
        for value in row.values():
            if self._status_label(value).lower().find(search_text) >= 0:
                return True
            if isinstance(value, dict):
                if search_text in str(value.get("label") or "").lower():
                    return True
                continue
            if isinstance(value, list):
                joined = ", ".join(str(item.get("label") or "") if isinstance(item, dict) else str(item or "") for item in value)
                if search_text in joined.lower():
                    return True
                continue
            if search_text in str(value or "").lower():
                return True
        return False

    def _resolve_export_numeric_value(self, row: dict[str, object]) -> Optional[Decimal]:
        for key in ["amount", "tdsAmount", "taxableAmount", "pending", "closingBalance", "totalTds", "inrAmount"]:
            if key in row:
                return self._decimal(row.get(key))
        return None

    def _sortable_export_value(self, value: object):
        if isinstance(value, dict):
            return str(value.get("label") or "").lower()
        if value in (None, ""):
            return ""
        decimal_value = self._parse_export_decimal(value)
        if decimal_value is not None:
            return float(decimal_value)
        parsed_date = self._parse_optional_date(str(value))
        if parsed_date is not None:
            return parsed_date.toordinal()
        return str(value).lower()

    def _parse_export_decimal(self, value: object) -> Optional[Decimal]:
        if value in (None, "", "null", "None"):
            return None
        try:
            return Decimal(str(value).replace(",", ""))
        except Exception:
            return None

    def _export_response(self, *, filename: str, content: bytes, content_type: str) -> HttpResponse:
        response = HttpResponse(content, content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
