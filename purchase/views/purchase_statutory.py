from __future__ import annotations
from datetime import date
from calendar import monthrange
import csv
from io import BytesIO
from decimal import Decimal

from rest_framework import generics, permissions, status
from rest_framework.exceptions import ValidationError
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from pypdf import PdfReader, PdfWriter

from financial.profile_access import account_pan
from purchase.models.purchase_core import PurchaseInvoiceHeader
from purchase.models.purchase_statutory import PurchaseStatutoryChallan, PurchaseStatutoryReturn
from rbac.services import EffectivePermissionService
from purchase.serializers.purchase_statutory import (
    PurchaseStatutoryChallanSerializer,
    PurchaseStatutoryChallanCreateInputSerializer,
    PurchaseStatutoryReviewNoteInputSerializer,
    PurchaseStatutoryReviewNoteSerializer,
    PurchaseStatutoryReturnSerializer,
    PurchaseStatutoryReturnCreateInputSerializer,
)
from purchase.services.purchase_statutory_service import PurchaseStatutoryService


def _parse_scope(request):
    entity = request.query_params.get("entity")
    entityfinid = request.query_params.get("entityfinid")
    subentity = request.query_params.get("subentity")
    if not entity or not entityfinid:
        raise ValidationError({"detail": "entity and entityfinid query params are required."})
    try:
        entity_id = int(entity)
        entityfinid_id = int(entityfinid)
        subentity_id = int(subentity) if subentity not in (None, "", "null") else None
    except (TypeError, ValueError):
        raise ValidationError({"detail": "entity/entityfinid/subentity must be integers."})
    return entity_id, entityfinid_id, subentity_id


def _parse_required_period_and_tax_type(request):
    tax_type = request.query_params.get("tax_type")
    period_from_raw = request.query_params.get("period_from")
    period_to_raw = request.query_params.get("period_to")
    if not tax_type:
        raise ValidationError({"detail": "tax_type is required."})
    if not period_from_raw or not period_to_raw:
        raise ValidationError({"detail": "period_from and period_to are required."})
    try:
        period_from = date.fromisoformat(str(period_from_raw))
        period_to = date.fromisoformat(str(period_to_raw))
    except ValueError:
        raise ValidationError({"detail": "period_from and period_to must be YYYY-MM-DD."})
    return tax_type, period_from, period_to


def _require_any_permission(request, entity_id: int, permission_codes: list[str]) -> None:
    if not entity_id:
        raise PermissionDenied("Entity scope is required for permission check.")
    available = EffectivePermissionService.permission_codes_for_user(request.user, int(entity_id))
    for code in permission_codes:
        if code in available:
            return
    raise PermissionDenied(f"Missing permission: one of {', '.join(permission_codes)}")


def _require_statutory_view(request, entity_id: int) -> None:
    _require_any_permission(
        request,
        entity_id,
        [
            "purchase.statutory.view",
            "purchase.invoice.view",
            "purchase.invoice.list",
            "purchase.invoice.read",
        ],
    )


def _require_statutory_manage(request, entity_id: int) -> None:
    _require_any_permission(
        request,
        entity_id,
        [
            "purchase.statutory.manage",
            "purchase.invoice.update",
            "purchase.invoice.edit",
        ],
    )


def _require_statutory_approve(request, entity_id: int) -> None:
    _require_any_permission(
        request,
        entity_id,
        [
            "purchase.statutory.approve",
            "purchase.invoice.approve",
            "purchase.invoice.update",
            "purchase.invoice.edit",
        ],
    )


def _extract_entity_id_from_request(request):
    raw = None
    if hasattr(request, "data"):
        raw = request.data.get("entity")
    if raw in (None, "", "null"):
        raw = request.query_params.get("entity")
    if raw in (None, "", "null"):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _require_for_pk_if_resolvable(request, model, pk: int, level: str = "view") -> None:
    entity_id = _extract_entity_id_from_request(request)
    if entity_id is None:
        obj = model.objects.filter(pk=pk).only("entity_id").first()
        if obj is not None:
            entity_id = int(obj.entity_id)
    if entity_id is None:
        return
    if level == "view":
        _require_statutory_view(request, entity_id)
    elif level == "manage":
        _require_statutory_manage(request, entity_id)
    elif level == "approve":
        _require_statutory_approve(request, entity_id)
    else:
        raise ValueError(f"Unsupported RBAC level: {level}")


class PurchaseStatutoryChallanListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PurchaseStatutoryChallanSerializer

    def get_queryset(self):
        entity_id, entityfinid_id, subentity_id = _parse_scope(self.request)
        _require_statutory_view(self.request, entity_id)
        qs = PurchaseStatutoryChallan.objects.prefetch_related("lines").filter(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
        )
        if subentity_id is None:
            qs = qs.filter(subentity__isnull=True)
        else:
            qs = qs.filter(subentity_id=subentity_id)
        tax_type = self.request.query_params.get("tax_type")
        if tax_type:
            qs = qs.filter(tax_type=tax_type)
        status_q = self.request.query_params.get("status")
        if status_q not in (None, "", "null"):
            qs = qs.filter(status=int(status_q))
        return qs.order_by("-challan_date", "-id")

    def create(self, request, *args, **kwargs):
        inp = PurchaseStatutoryChallanCreateInputSerializer(data=request.data)
        inp.is_valid(raise_exception=True)
        data = inp.validated_data
        _require_statutory_manage(request, int(data["entity"]))
        lines = list(data.get("lines") or [])
        auto_populate = bool(data.get("auto_populate", False))
        if not lines and auto_populate:
            challan_date = data["challan_date"]
            period_from = data.get("period_from") or challan_date.replace(day=1)
            period_to = data.get("period_to") or challan_date.replace(
                day=monthrange(challan_date.year, challan_date.month)[1]
            )
            eligible_payload = PurchaseStatutoryService.challan_eligible_lines(
                entity_id=int(data["entity"]),
                entityfinid_id=int(data["entityfinid"]),
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                period_from=period_from,
                period_to=period_to,
            )
            lines = [
                {
                    "header_id": int(row["header_id"]),
                    "section_id": row.get("section_id"),
                    "amount": row["amount"],
                }
                for row in (eligible_payload.get("lines") or [])
            ]
            data["period_from"] = period_from
            data["period_to"] = period_to
        if not lines:
            raise ValidationError({"detail": "At least one line is required. Use eligible-lines or set auto_populate=true."})
        try:
            res = PurchaseStatutoryService.create_challan(
                entity_id=data["entity"],
                entityfinid_id=data["entityfinid"],
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                challan_no=data["challan_no"],
                challan_date=data["challan_date"],
                period_from=data.get("period_from"),
                period_to=data.get("period_to"),
                interest_amount=data.get("interest_amount"),
                late_fee_amount=data.get("late_fee_amount"),
                penalty_amount=data.get("penalty_amount"),
                bank_ref_no=data.get("bank_ref_no"),
                bsr_code=data.get("bsr_code"),
                cin_no=data.get("cin_no"),
                minor_head_code=data.get("minor_head_code"),
                payment_payload_json=data.get("payment_payload_json"),
                ack_document=data.get("ack_document"),
                remarks=data.get("remarks"),
                lines=lines,
                created_by_id=request.user.id,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        out = PurchaseStatutoryChallanSerializer(res.obj, context=self.get_serializer_context())
        return Response({"message": res.message, "data": out.data}, status=status.HTTP_201_CREATED)


class PurchaseStatutoryChallanDepositAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryChallan, pk, level="manage")
        deposited_on = request.data.get("deposited_on")
        bank_ref_no = request.data.get("bank_ref_no")
        bsr_code = request.data.get("bsr_code")
        cin_no = request.data.get("cin_no")
        minor_head_code = request.data.get("minor_head_code")
        payment_payload_json = request.data.get("payment_payload_json")
        ack_document = request.data.get("ack_document")
        try:
            res = PurchaseStatutoryService.deposit_challan(
                challan_id=pk,
                deposited_by_id=request.user.id,
                deposited_on=deposited_on,
                bank_ref_no=bank_ref_no,
                bsr_code=bsr_code,
                cin_no=cin_no,
                minor_head_code=minor_head_code,
                payment_payload_json=payment_payload_json,
                ack_document=ack_document,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": res.message, "data": PurchaseStatutoryChallanSerializer(res.obj).data})


class PurchaseStatutoryChallanCancelAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryChallan, pk, level="manage")
        reason = request.data.get("reason")
        try:
            res = PurchaseStatutoryService.cancel_challan(
                challan_id=pk,
                cancelled_by_id=request.user.id,
                reason=reason,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": res.message, "data": PurchaseStatutoryChallanSerializer(res.obj).data})


class PurchaseStatutoryChallanApprovalAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryChallan, pk, level="approve")
        action = (request.data.get("action") or "").strip().lower()
        remarks = request.data.get("remarks")
        try:
            if action == "submit":
                res = PurchaseStatutoryService.submit_challan_for_approval(
                    challan_id=pk, user_id=request.user.id, remarks=remarks
                )
            elif action == "approve":
                res = PurchaseStatutoryService.approve_challan(
                    challan_id=pk, user_id=request.user.id, remarks=remarks
                )
            elif action == "reject":
                res = PurchaseStatutoryService.reject_challan(
                    challan_id=pk, user_id=request.user.id, remarks=remarks
                )
            else:
                raise ValidationError({"detail": "action must be submit|approve|reject"})
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        out = PurchaseStatutoryChallanSerializer(res.obj).data
        return Response(
            {
                "message": res.message,
                "approval_status": out.get("approval_status", "DRAFT"),
                "approval_status_name": out.get("approval_status_name", "Draft"),
                "data": out,
            }
        )


class PurchaseStatutoryReturnListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PurchaseStatutoryReturnSerializer

    def get_queryset(self):
        entity_id, entityfinid_id, subentity_id = _parse_scope(self.request)
        _require_statutory_view(self.request, entity_id)
        qs = PurchaseStatutoryReturn.objects.prefetch_related("lines").filter(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
        )
        if subentity_id is None:
            qs = qs.filter(subentity__isnull=True)
        else:
            qs = qs.filter(subentity_id=subentity_id)
        tax_type = self.request.query_params.get("tax_type")
        if tax_type:
            qs = qs.filter(tax_type=tax_type)
        status_q = self.request.query_params.get("status")
        if status_q not in (None, "", "null"):
            qs = qs.filter(status=int(status_q))
        return qs.order_by("-period_to", "-id")

    def create(self, request, *args, **kwargs):
        inp = PurchaseStatutoryReturnCreateInputSerializer(data=request.data)
        inp.is_valid(raise_exception=True)
        data = inp.validated_data
        _require_statutory_manage(request, int(data["entity"]))
        lines = list(data.get("lines") or [])
        auto_populate = bool(data.get("auto_populate", False))
        if not lines and auto_populate:
            eligible_payload = PurchaseStatutoryService.return_eligible_lines(
                entity_id=int(data["entity"]),
                entityfinid_id=int(data["entityfinid"]),
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                period_from=data["period_from"],
                period_to=data["period_to"],
                return_code=data.get("return_code"),
            )
            lines = [
                {
                    "header_id": int(row["header_id"]),
                    "challan_id": row.get("challan_id"),
                    "amount": row["amount"],
                    "section_snapshot_code": row.get("section_snapshot_code"),
                    "section_snapshot_desc": row.get("section_snapshot_desc"),
                    "deductee_residency_snapshot": row.get("deductee_residency_snapshot"),
                    "deductee_country_snapshot": row.get("deductee_country_snapshot"),
                    "deductee_country_code_snapshot": row.get("deductee_country_code_snapshot"),
                    "deductee_country_name_snapshot": row.get("deductee_country_name_snapshot"),
                    "deductee_tax_id_snapshot": row.get("deductee_tax_id_snapshot"),
                    "deductee_pan_snapshot": row.get("deductee_pan_snapshot"),
                    "deductee_gstin_snapshot": row.get("deductee_gstin_snapshot"),
                    "cin_snapshot": row.get("cin_snapshot"),
                    "metadata_json": row.get("metadata_json") or {},
                }
                for row in (eligible_payload.get("lines") or [])
            ]
        if not lines:
            raise ValidationError({"detail": "At least one line is required. Use eligible-lines or set auto_populate=true."})
        try:
            res = PurchaseStatutoryService.create_return(
                entity_id=data["entity"],
                entityfinid_id=data["entityfinid"],
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                return_code=data["return_code"],
                period_from=data["period_from"],
                period_to=data["period_to"],
                ack_no=data.get("ack_no"),
                arn_no=data.get("arn_no"),
                interest_amount=data.get("interest_amount"),
                late_fee_amount=data.get("late_fee_amount"),
                penalty_amount=data.get("penalty_amount"),
                filed_payload_json=data.get("filed_payload_json"),
                ack_document=data.get("ack_document"),
                original_return_id=data.get("original_return_id"),
                revision_no=data.get("revision_no") or 0,
                remarks=data.get("remarks"),
                lines=lines,
                created_by_id=request.user.id,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        out = PurchaseStatutoryReturnSerializer(res.obj, context=self.get_serializer_context())
        return Response({"message": res.message, "data": out.data}, status=status.HTTP_201_CREATED)


class PurchaseStatutoryReturnFileAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="manage")
        filed_on = request.data.get("filed_on")
        ack_no = request.data.get("ack_no")
        arn_no = request.data.get("arn_no")
        filed_payload_json = request.data.get("filed_payload_json")
        ack_document = request.data.get("ack_document")
        try:
            res = PurchaseStatutoryService.file_return(
                filing_id=pk,
                filed_by_id=request.user.id,
                filed_on=filed_on,
                ack_no=ack_no,
                arn_no=arn_no,
                filed_payload_json=filed_payload_json,
                ack_document=ack_document,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": res.message, "data": PurchaseStatutoryReturnSerializer(res.obj).data})


class PurchaseStatutoryReturnCancelAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="manage")
        reason = request.data.get("reason")
        try:
            res = PurchaseStatutoryService.cancel_return(
                filing_id=pk,
                cancelled_by_id=request.user.id,
                reason=reason,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": res.message, "data": PurchaseStatutoryReturnSerializer(res.obj).data})


class PurchaseStatutoryReturnApprovalAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="approve")
        action = (request.data.get("action") or "").strip().lower()
        remarks = request.data.get("remarks")
        try:
            if action == "submit":
                res = PurchaseStatutoryService.submit_return_for_approval(
                    filing_id=pk, user_id=request.user.id, remarks=remarks
                )
            elif action == "approve":
                res = PurchaseStatutoryService.approve_return(
                    filing_id=pk, user_id=request.user.id, remarks=remarks
                )
            elif action == "reject":
                res = PurchaseStatutoryService.reject_return(
                    filing_id=pk, user_id=request.user.id, remarks=remarks
                )
            else:
                raise ValidationError({"detail": "action must be submit|approve|reject"})
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        out = PurchaseStatutoryReturnSerializer(res.obj).data
        return Response(
            {
                "message": res.message,
                "approval_status": out.get("approval_status", "DRAFT"),
                "approval_status_name": out.get("approval_status_name", "Draft"),
                "data": out,
            }
        )


class PurchaseStatutorySummaryAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        tax_type = request.query_params.get("tax_type")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        try:
            summary = PurchaseStatutoryService.reconciliation_summary(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                tax_type=tax_type or None,
                date_from=date_from or None,
                date_to=date_to or None,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"summary": summary})


class PurchaseStatutoryItcStatusRegisterAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        include_cancelled = str(request.query_params.get("include_cancelled", "false")).strip().lower() in {"1", "true", "yes", "y"}

        itc_claim_status_raw = request.query_params.get("itc_claim_status")
        gstr2b_status_raw = request.query_params.get("gstr2b_match_status")

        def _optional_int(value, field_name: str):
            if value in (None, "", "null"):
                return None
            try:
                return int(value)
            except (TypeError, ValueError):
                raise ValidationError({"detail": f"{field_name} must be an integer."})

        itc_claim_status = _optional_int(itc_claim_status_raw, "itc_claim_status")
        gstr2b_match_status = _optional_int(gstr2b_status_raw, "gstr2b_match_status")

        try:
            payload = PurchaseStatutoryService.itc_status_register(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                date_from=date_from or None,
                date_to=date_to or None,
                itc_claim_status=itc_claim_status,
                gstr2b_match_status=gstr2b_match_status,
                include_cancelled=include_cancelled,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(payload)


class PurchaseStatutoryReviewNoteAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        period_from_raw = request.query_params.get("period_from")
        period_to_raw = request.query_params.get("period_to")
        if not period_from_raw or not period_to_raw:
            raise ValidationError({"detail": "period_from and period_to are required."})
        try:
            period_from = date.fromisoformat(str(period_from_raw))
            period_to = date.fromisoformat(str(period_to_raw))
        except ValueError:
            raise ValidationError({"detail": "period_from and period_to must be YYYY-MM-DD."})
        tax_type = request.query_params.get("tax_type") or None
        try:
            note = PurchaseStatutoryService.get_review_note(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                tax_type=tax_type,
                period_from=period_from,
                period_to=period_to,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"data": PurchaseStatutoryReviewNoteSerializer(note).data if note else None})

    def post(self, request):
        inp = PurchaseStatutoryReviewNoteInputSerializer(data=request.data)
        inp.is_valid(raise_exception=True)
        data = inp.validated_data
        _require_statutory_manage(request, int(data["entity"]))
        try:
            res = PurchaseStatutoryService.save_review_note(
                entity_id=int(data["entity"]),
                entityfinid_id=int(data["entityfinid"]),
                subentity_id=data.get("subentity"),
                tax_type=data.get("tax_type"),
                period_from=data["period_from"],
                period_to=data["period_to"],
                reviewer_name=data.get("reviewer_name"),
                closure_status=data.get("closure_status"),
                review_summary=data.get("review_summary"),
                open_points=data.get("open_points"),
                closure_comment=data.get("closure_comment"),
                reviewed_by_id=request.user.id,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": res.message, "data": PurchaseStatutoryReviewNoteSerializer(res.obj).data})

    def delete(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_manage(request, entity_id)
        period_from_raw = request.query_params.get("period_from")
        period_to_raw = request.query_params.get("period_to")
        if not period_from_raw or not period_to_raw:
            raise ValidationError({"detail": "period_from and period_to are required."})
        try:
            period_from = date.fromisoformat(str(period_from_raw))
            period_to = date.fromisoformat(str(period_to_raw))
        except ValueError:
            raise ValidationError({"detail": "period_from and period_to must be YYYY-MM-DD."})
        tax_type = request.query_params.get("tax_type") or None
        try:
            PurchaseStatutoryService.delete_review_note(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                tax_type=tax_type,
                period_from=period_from,
                period_to=period_to,
                reviewed_by_id=request.user.id,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": "Review note deleted."}, status=status.HTTP_200_OK)


class PurchaseStatutoryGlReconciliationAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        period_from_raw = request.query_params.get("period_from")
        period_to_raw = request.query_params.get("period_to")
        if not period_from_raw or not period_to_raw:
            raise ValidationError({"detail": "period_from and period_to are required."})
        try:
            period_from = date.fromisoformat(str(period_from_raw))
            period_to = date.fromisoformat(str(period_to_raw))
        except ValueError:
            raise ValidationError({"detail": "period_from and period_to must be YYYY-MM-DD."})
        if period_from > period_to:
            raise ValidationError({"detail": "period_from cannot be greater than period_to."})
        payload = PurchaseStatutoryService.reconciliation_gl_status(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            period_from=period_from,
            period_to=period_to,
        )
        return Response(payload)


class _PdfJsonRenderer(JSONRenderer):
    format = "pdf"


class _XlsxJsonRenderer(JSONRenderer):
    format = "xlsx"


class _CsvJsonRenderer(JSONRenderer):
    format = "csv"


def _xlsx_response_from_rows(*, rows, filename: str, sheet_name: str, headers: list[str]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        col = get_column_letter(col_idx)
        max_len = len(str(header))
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            value = r[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 42)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _pdf_response_from_rows(*, rows, filename: str, title: str, headers: list[str]) -> HttpResponse:
    out = BytesIO()
    c = canvas.Canvas(out, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, y, title)
    y -= 18
    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Generated on {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 18
    c.setFont("Helvetica-Bold", 8)
    c.drawString(40, y, " | ".join(headers))
    y -= 10
    c.line(40, y, width - 40, y)
    y -= 10
    c.setFont("Helvetica", 8)
    for row in rows:
        if y < 50:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica-Bold", 8)
            c.drawString(40, y, " | ".join(headers))
            y -= 10
            c.line(40, y, width - 40, y)
            y -= 10
            c.setFont("Helvetica", 8)
        line = " | ".join(str(row.get(h) or "") for h in headers)
        c.drawString(40, y, line[:200])
        y -= 10
    c.showPage()
    c.save()
    out.seek(0)
    response = HttpResponse(out.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class PurchaseStatutoryChallanDetailAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def put(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryChallan, pk, level="manage")
        inp = PurchaseStatutoryChallanCreateInputSerializer(data=request.data)
        inp.is_valid(raise_exception=True)
        data = inp.validated_data
        lines = list(data.get("lines") or [])
        auto_populate = bool(data.get("auto_populate", False))
        if not lines and auto_populate:
            challan_date = data["challan_date"]
            period_from = data.get("period_from") or challan_date.replace(day=1)
            period_to = data.get("period_to") or challan_date.replace(
                day=monthrange(challan_date.year, challan_date.month)[1]
            )
            eligible_payload = PurchaseStatutoryService.challan_eligible_lines(
                entity_id=int(data["entity"]),
                entityfinid_id=int(data["entityfinid"]),
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                period_from=period_from,
                period_to=period_to,
            )
            lines = [
                {
                    "header_id": int(row["header_id"]),
                    "section_id": row.get("section_id"),
                    "amount": row["amount"],
                }
                for row in (eligible_payload.get("lines") or [])
            ]
            data["period_from"] = period_from
            data["period_to"] = period_to
        if not lines:
            raise ValidationError({"detail": "At least one line is required. Use eligible-lines or set auto_populate=true."})
        try:
            res = PurchaseStatutoryService.update_challan(
                challan_id=pk,
                entity_id=data["entity"],
                entityfinid_id=data["entityfinid"],
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                challan_no=data["challan_no"],
                challan_date=data["challan_date"],
                period_from=data.get("period_from"),
                period_to=data.get("period_to"),
                interest_amount=data.get("interest_amount"),
                late_fee_amount=data.get("late_fee_amount"),
                penalty_amount=data.get("penalty_amount"),
                bank_ref_no=data.get("bank_ref_no"),
                bsr_code=data.get("bsr_code"),
                cin_no=data.get("cin_no"),
                minor_head_code=data.get("minor_head_code"),
                payment_payload_json=data.get("payment_payload_json"),
                ack_document=data.get("ack_document"),
                remarks=data.get("remarks"),
                lines=lines,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        out = PurchaseStatutoryChallanSerializer(res.obj)
        return Response({"message": res.message, "data": out.data})

    def delete(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryChallan, pk, level="manage")
        try:
            msg = PurchaseStatutoryService.delete_challan(challan_id=pk)
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": msg}, status=status.HTTP_200_OK)


class PurchaseStatutoryReturnDetailAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def put(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="manage")
        inp = PurchaseStatutoryReturnCreateInputSerializer(data=request.data)
        inp.is_valid(raise_exception=True)
        data = inp.validated_data
        lines = list(data.get("lines") or [])
        auto_populate = bool(data.get("auto_populate", False))
        if not lines and auto_populate:
            eligible_payload = PurchaseStatutoryService.return_eligible_lines(
                entity_id=int(data["entity"]),
                entityfinid_id=int(data["entityfinid"]),
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                period_from=data["period_from"],
                period_to=data["period_to"],
            )
            lines = [
                {
                    "header_id": int(row["header_id"]),
                    "challan_id": row.get("challan_id"),
                    "amount": row["amount"],
                    "section_snapshot_code": row.get("section_snapshot_code"),
                    "section_snapshot_desc": row.get("section_snapshot_desc"),
                    "deductee_residency_snapshot": row.get("deductee_residency_snapshot"),
                    "deductee_country_snapshot": row.get("deductee_country_snapshot"),
                    "deductee_country_code_snapshot": row.get("deductee_country_code_snapshot"),
                    "deductee_country_name_snapshot": row.get("deductee_country_name_snapshot"),
                    "deductee_tax_id_snapshot": row.get("deductee_tax_id_snapshot"),
                    "deductee_pan_snapshot": row.get("deductee_pan_snapshot"),
                    "deductee_gstin_snapshot": row.get("deductee_gstin_snapshot"),
                    "cin_snapshot": row.get("cin_snapshot"),
                    "metadata_json": row.get("metadata_json") or {},
                }
                for row in (eligible_payload.get("lines") or [])
            ]
        if not lines:
            raise ValidationError({"detail": "At least one line is required. Use eligible-lines or set auto_populate=true."})
        try:
            res = PurchaseStatutoryService.update_return(
                filing_id=pk,
                entity_id=data["entity"],
                entityfinid_id=data["entityfinid"],
                subentity_id=data.get("subentity"),
                tax_type=data["tax_type"],
                return_code=data["return_code"],
                period_from=data["period_from"],
                period_to=data["period_to"],
                ack_no=data.get("ack_no"),
                arn_no=data.get("arn_no"),
                interest_amount=data.get("interest_amount"),
                late_fee_amount=data.get("late_fee_amount"),
                penalty_amount=data.get("penalty_amount"),
                filed_payload_json=data.get("filed_payload_json"),
                ack_document=data.get("ack_document"),
                original_return_id=data.get("original_return_id"),
                revision_no=data.get("revision_no") or 0,
                remarks=data.get("remarks"),
                lines=lines,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        out = PurchaseStatutoryReturnSerializer(res.obj)
        return Response({"message": res.message, "data": out.data})

    def delete(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="manage")
        try:
            msg = PurchaseStatutoryService.delete_return(filing_id=pk)
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": msg}, status=status.HTTP_200_OK)


class PurchaseStatutoryChallanExportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    renderer_classes = [JSONRenderer, _PdfJsonRenderer, _XlsxJsonRenderer, _CsvJsonRenderer]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        fmt = (request.query_params.get("format") or "xlsx").lower().strip()
        if fmt not in ("xlsx", "pdf", "csv"):
            raise ValidationError({"detail": "format must be xlsx|pdf|csv"})

        qs = PurchaseStatutoryChallan.objects.filter(entity_id=entity_id, entityfinid_id=entityfinid_id)
        if subentity_id is None:
            qs = qs.filter(subentity__isnull=True)
        else:
            qs = qs.filter(subentity_id=subentity_id)

        tax_type = request.query_params.get("tax_type")
        if tax_type:
            qs = qs.filter(tax_type=tax_type)
        headers = [
            "id", "tax_type", "challan_no", "challan_date", "period_from", "period_to",
            "amount", "interest_amount", "late_fee_amount", "penalty_amount", "status", "cin_no", "bsr_code",
        ]
        rows = list(qs.order_by("-challan_date", "-id").values(
            "id", "tax_type", "challan_no", "challan_date", "period_from", "period_to",
            "amount", "interest_amount", "late_fee_amount", "penalty_amount", "status", "cin_no", "bsr_code"
        ))

        if fmt == "xlsx":
            return _xlsx_response_from_rows(
                rows=rows,
                filename="purchase_statutory_challans.xlsx",
                sheet_name="challans",
                headers=headers,
            )
        if fmt == "pdf":
            return _pdf_response_from_rows(
                rows=rows,
                filename="purchase_statutory_challans.pdf",
                title="Purchase Statutory Challans",
                headers=headers,
            )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="purchase_statutory_challans.csv"'
        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        return response


class PurchaseStatutoryReturnExportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    renderer_classes = [JSONRenderer, _PdfJsonRenderer, _XlsxJsonRenderer, _CsvJsonRenderer]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        fmt = (request.query_params.get("format") or "xlsx").lower().strip()
        if fmt not in ("xlsx", "pdf", "csv"):
            raise ValidationError({"detail": "format must be xlsx|pdf|csv"})

        qs = PurchaseStatutoryReturn.objects.filter(entity_id=entity_id, entityfinid_id=entityfinid_id)
        if subentity_id is None:
            qs = qs.filter(subentity__isnull=True)
        else:
            qs = qs.filter(subentity_id=subentity_id)

        tax_type = request.query_params.get("tax_type")
        if tax_type:
            qs = qs.filter(tax_type=tax_type)
        headers = [
            "id", "tax_type", "return_code", "period_from", "period_to",
            "amount", "interest_amount", "late_fee_amount", "penalty_amount", "status", "ack_no", "arn_no",
        ]
        rows = list(qs.order_by("-period_to", "-id").values(
            "id", "tax_type", "return_code", "period_from", "period_to",
            "amount", "interest_amount", "late_fee_amount", "penalty_amount", "status", "ack_no", "arn_no"
        ))

        if fmt == "xlsx":
            return _xlsx_response_from_rows(
                rows=rows,
                filename="purchase_statutory_returns.xlsx",
                sheet_name="returns",
                headers=headers,
            )
        if fmt == "pdf":
            return _pdf_response_from_rows(
                rows=rows,
                filename="purchase_statutory_returns.pdf",
                title="Purchase Statutory Returns",
                headers=headers,
            )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="purchase_statutory_returns.csv"'
        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        return response


class PurchaseStatutoryChallanPreviewNoAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_manage(request, entity_id)
        tax_type = (request.query_params.get("tax_type") or "").strip().upper()
        if tax_type not in (
            PurchaseStatutoryChallan.TaxType.IT_TDS,
            PurchaseStatutoryChallan.TaxType.GST_TDS,
        ):
            raise ValidationError({"detail": "tax_type must be IT_TDS or GST_TDS."})

        challan_date_raw = request.query_params.get("challan_date")
        if challan_date_raw:
            try:
                challan_date = date.fromisoformat(str(challan_date_raw))
            except ValueError:
                raise ValidationError({"detail": "challan_date must be YYYY-MM-DD."})
        else:
            challan_date = date.today()

        qs = PurchaseStatutoryChallan.objects.filter(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            tax_type=tax_type,
            challan_date__year=challan_date.year,
            challan_date__month=challan_date.month,
        )
        if subentity_id is not None:
            qs = qs.filter(subentity_id=subentity_id)
        sequence = int(qs.count()) + 1
        prefix = "ITTDS" if tax_type == PurchaseStatutoryChallan.TaxType.IT_TDS else "GSTTDS"
        challan_no = f"{prefix}-{challan_date.strftime('%b').upper()}-{challan_date.year}-{sequence:03d}"
        return Response({"challan_no": challan_no})


class PurchaseStatutoryCaPackExportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_manage(request, entity_id)
        period_from_raw = request.query_params.get("period_from")
        period_to_raw = request.query_params.get("period_to")
        if not period_from_raw or not period_to_raw:
            raise ValidationError({"detail": "period_from and period_to are required."})
        try:
            period_from = date.fromisoformat(str(period_from_raw))
            period_to = date.fromisoformat(str(period_to_raw))
        except ValueError:
            raise ValidationError({"detail": "period_from and period_to must be YYYY-MM-DD."})
        if period_from > period_to:
            raise ValidationError({"detail": "period_from cannot be greater than period_to."})

        invoice_qs = PurchaseInvoiceHeader.objects.filter(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            bill_date__gte=period_from,
            bill_date__lte=period_to,
            status=PurchaseInvoiceHeader.Status.POSTED,
        ).select_related("vendor", "tds_section", "subentity")
        challan_qs = PurchaseStatutoryChallan.objects.filter(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            challan_date__gte=period_from,
            challan_date__lte=period_to,
        ).prefetch_related("lines__header", "lines__section")
        return_qs = PurchaseStatutoryReturn.objects.filter(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            period_to__gte=period_from,
            period_from__lte=period_to,
        ).prefetch_related("lines__header", "lines__challan")
        if subentity_id is not None:
            invoice_qs = invoice_qs.filter(subentity_id=subentity_id)
            challan_qs = challan_qs.filter(subentity_id=subentity_id)
            return_qs = return_qs.filter(subentity_id=subentity_id)

        wb = Workbook()
        wb.remove(wb.active)

        def add_cover_sheet():
            ws = wb.create_sheet(title="00_Cover", index=0)
            ws["A1"] = "Purchase Statutory CA Pack"
            ws["A2"] = "Entity ID"
            ws["B2"] = entity_id
            ws["A3"] = "Entity Financial Year ID"
            ws["B3"] = entityfinid_id
            ws["A4"] = "Subentity"
            ws["B4"] = subentity_id if subentity_id is not None else "All Subentities"
            ws["A5"] = "Period From"
            ws["B5"] = period_from
            ws["A6"] = "Period To"
            ws["B6"] = period_to
            ws["A7"] = "Generated At"
            ws["B7"] = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
            ws["A8"] = "Generated By"
            ws["B8"] = getattr(request.user, "username", "") or getattr(request.user, "email", "") or f"User#{request.user.id}"

            ws["A10"] = "Included Sheets"
            included = [
                "01_Management_Summary",
                "02_Reviewer_Signoff",
                "03_Action_Items",
                "04_IT_TDS_Deductions",
                "05_IT_TDS_Challans",
                "06_IT_TDS_Returns",
                "07_GST_TDS_Deductions",
                "08_GST_TDS_Challans",
                "09_GST_TDS_Returns",
                "10_Reconciliation",
                "11_Exceptions",
                "12_Supporting_Doc_Index",
            ]
            for idx, name in enumerate(included, start=11):
                ws[f"A{idx}"] = name

            ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
            ws["A10"].font = Font(size=12, bold=True)
            for row_idx in range(2, 9):
                ws[f"A{row_idx}"].font = Font(bold=True)
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 36
            ws.freeze_panes = "A11"

        add_cover_sheet()

        def add_sheet(title: str, headers, rows, total_columns=None, index=None):
            total_columns = total_columns or []
            rows = list(rows or [])
            ws = wb.create_sheet(title=title, index=index) if index is not None else wb.create_sheet(title=title)
            ws.append(headers)
            for row in rows:
                ws.append(list(row))
            if rows and total_columns:
                total_row = [""] * len(headers)
                total_row[0] = "TOTAL"
                for col_idx in total_columns:
                    total = Decimal("0.00")
                    for row in rows:
                        try:
                            total += Decimal(str(row[col_idx] or 0))
                        except Exception:
                            continue
                    total_row[col_idx] = total
                ws.append(total_row)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row_idx in range(2, ws.max_row + 1):
                for cell in ws[row_idx]:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top")

            if rows and total_columns:
                for cell in ws[ws.max_row]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)

            for col_idx, _ in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(headers[col_idx - 1]))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    v = row[0].value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 42)

        def decimal_or_zero(value):
            try:
                return Decimal(str(value or 0))
            except Exception:
                return Decimal("0.00")

        def review_note_rows(note, scope_label: str):
            if note is None:
                return [
                    (scope_label, "No saved signoff note", "", "", "", "", "", ""),
                ]
            reviewed_by = ""
            if getattr(note, "reviewed_by", None):
                reviewed_by = getattr(note.reviewed_by, "username", "") or getattr(note.reviewed_by, "email", "") or f"User#{note.reviewed_by_id}"
            return [
                (
                    scope_label,
                    note.get_closure_status_display(),
                    note.reviewer_name or "",
                    note.reviewed_at.strftime("%Y-%m-%d %H:%M:%S") if note.reviewed_at else "",
                    reviewed_by,
                    note.review_summary or "",
                    note.open_points or "",
                    note.closure_comment or "",
                )
            ]

        global_summary = PurchaseStatutoryService.reconciliation_summary(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            tax_type=None,
            date_from=period_from,
            date_to=period_to,
        )
        it_summary = PurchaseStatutoryService.reconciliation_summary(
            entity_id=entity_id, entityfinid_id=entityfinid_id, subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS, date_from=period_from, date_to=period_to
        )
        gst_summary = PurchaseStatutoryService.reconciliation_summary(
            entity_id=entity_id, entityfinid_id=entityfinid_id, subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS, date_from=period_from, date_to=period_to
        )
        it_exc_payload = PurchaseStatutoryService.reconciliation_exceptions(
            entity_id=entity_id, entityfinid_id=entityfinid_id, subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS, period_from=period_from, period_to=period_to
        )
        gst_exc_payload = PurchaseStatutoryService.reconciliation_exceptions(
            entity_id=entity_id, entityfinid_id=entityfinid_id, subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS, period_from=period_from, period_to=period_to
        )
        it_exc = it_exc_payload.get("exceptions", {})
        gst_exc = gst_exc_payload.get("exceptions", {})

        all_scope_note = PurchaseStatutoryService.get_review_note(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            tax_type=None,
            period_from=period_from,
            period_to=period_to,
        )
        it_scope_note = PurchaseStatutoryService.get_review_note(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            period_from=period_from,
            period_to=period_to,
        )
        gst_scope_note = PurchaseStatutoryService.get_review_note(
            entity_id=entity_id,
            entityfinid_id=entityfinid_id,
            subentity_id=subentity_id,
            tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS,
            period_from=period_from,
            period_to=period_to,
        )

        filed_or_revised_returns = [
            row for row in return_qs
            if int(row.status) in (
                int(PurchaseStatutoryReturn.Status.FILED),
                int(PurchaseStatutoryReturn.Status.REVISED),
            )
        ]
        evidence_follow_up = [
            row for row in filed_or_revised_returns
            if not (row.ack_no or "").strip() or not (row.arn_no or "").strip()
        ]
        open_draft_challans = [row for row in challan_qs if int(row.status) == int(PurchaseStatutoryChallan.Status.DRAFT)]
        open_draft_returns = [row for row in return_qs if int(row.status) == int(PurchaseStatutoryReturn.Status.DRAFT)]
        open_it_draft_challans = [row for row in open_draft_challans if row.tax_type == PurchaseStatutoryChallan.TaxType.IT_TDS]
        open_gst_draft_challans = [row for row in open_draft_challans if row.tax_type == PurchaseStatutoryChallan.TaxType.GST_TDS]
        open_it_draft_returns = [row for row in open_draft_returns if row.tax_type == PurchaseStatutoryReturn.TaxType.IT_TDS]
        open_gst_draft_returns = [row for row in open_draft_returns if row.tax_type == PurchaseStatutoryReturn.TaxType.GST_TDS]

        closure_rate = Decimal("100.00")
        deducted_value = decimal_or_zero(global_summary.get("deducted"))
        filed_value = decimal_or_zero(global_summary.get("filed"))
        pending_deposit_value = decimal_or_zero(global_summary.get("pending_deposit"))
        pending_filing_value = decimal_or_zero(global_summary.get("pending_filing"))
        if deducted_value > 0:
            closure_rate = min(Decimal("100.00"), (filed_value / deducted_value) * Decimal("100.00"))
        elif pending_deposit_value > 0 or pending_filing_value > 0:
            closure_rate = Decimal("0.00")

        deposit_coverage = Decimal("100.00")
        deposited_value = decimal_or_zero(global_summary.get("deposited"))
        if deducted_value > 0:
            deposit_coverage = min(Decimal("100.00"), (deposited_value / deducted_value) * Decimal("100.00"))
        elif pending_deposit_value > 0:
            deposit_coverage = Decimal("0.00")

        add_sheet(
            "01_Management_Summary",
            ["Metric", "Overall", "IT_TDS", "GST_TDS", "Comment"],
            [
                ("Deducted Value", global_summary.get("deducted", "0.00"), it_summary.get("deducted", "0.00"), gst_summary.get("deducted", "0.00"), "Source statutory value booked in the current period."),
                ("Deposited Value", global_summary.get("deposited", "0.00"), it_summary.get("deposited", "0.00"), gst_summary.get("deposited", "0.00"), "Value already covered through challan deposit."),
                ("Filed Value", global_summary.get("filed", "0.00"), it_summary.get("filed", "0.00"), gst_summary.get("filed", "0.00"), "Value already closed through return filing."),
                ("Pending Deposit", global_summary.get("pending_deposit", "0.00"), it_summary.get("pending_deposit", "0.00"), gst_summary.get("pending_deposit", "0.00"), "Deducted but not yet deposited."),
                ("Pending Filing", global_summary.get("pending_filing", "0.00"), it_summary.get("pending_filing", "0.00"), gst_summary.get("pending_filing", "0.00"), "Deposited but not yet filed."),
                ("Draft Challans", len(open_draft_challans), len(open_it_draft_challans), len(open_gst_draft_challans), "Open challan drafts still needing action."),
                ("Draft Returns", len(open_draft_returns), len(open_it_draft_returns), len(open_gst_draft_returns), "Open return drafts still needing action."),
                ("Reconciliation Exceptions", len(it_exc.get("invoices_pending_challan_mapping", {}).get("rows", [])) + len(it_exc.get("challan_lines_pending_return_mapping", {}).get("rows", [])) + len(gst_exc.get("invoices_pending_challan_mapping", {}).get("rows", [])) + len(gst_exc.get("challan_lines_pending_return_mapping", {}).get("rows", [])), len(it_exc.get("invoices_pending_challan_mapping", {}).get("rows", [])) + len(it_exc.get("challan_lines_pending_return_mapping", {}).get("rows", [])), len(gst_exc.get("invoices_pending_challan_mapping", {}).get("rows", [])) + len(gst_exc.get("challan_lines_pending_return_mapping", {}).get("rows", [])), "Open mapping exceptions visible in reconciliation helpers."),
                ("Evidence Follow-up", len(evidence_follow_up), len([row for row in evidence_follow_up if row.tax_type == PurchaseStatutoryReturn.TaxType.IT_TDS]), len([row for row in evidence_follow_up if row.tax_type == PurchaseStatutoryReturn.TaxType.GST_TDS]), "Filed or revised returns still missing ACK or ARN."),
                ("Closure Rate %", f"{closure_rate.quantize(Decimal('0.01'))}", "", "", "Filed value as a percentage of deducted value."),
                ("Deposit Coverage %", f"{deposit_coverage.quantize(Decimal('0.01'))}", "", "", "Deposited value as a percentage of deducted value."),
            ],
            index=1,
        )

        reviewer_rows = []
        reviewer_rows.extend(review_note_rows(all_scope_note, "All Tax Types"))
        reviewer_rows.extend(review_note_rows(it_scope_note, "IT_TDS"))
        reviewer_rows.extend(review_note_rows(gst_scope_note, "GST_TDS"))
        add_sheet(
            "02_Reviewer_Signoff",
            ["Scope", "Closure Status", "Reviewer Name", "Reviewed At", "Saved By", "Review Summary", "Open Points", "Closure Comment"],
            reviewer_rows,
            index=2,
        )

        review_timeline_rows = []
        for scope_label, note in (
            ("All Tax Types", all_scope_note),
            ("IT_TDS", it_scope_note),
            ("GST_TDS", gst_scope_note),
        ):
            if note is None:
                continue
            for event in note.events.select_related("changed_by").order_by("-changed_at")[:12]:
                changed_by = ""
                if getattr(event, "changed_by", None):
                    changed_by = getattr(event.changed_by, "username", "") or getattr(event.changed_by, "email", "") or f"User#{event.changed_by_id}"
                review_timeline_rows.append(
                    (
                        scope_label,
                        event.get_action_display(),
                        event.get_closure_status_display(),
                        event.reviewer_name or "",
                        changed_by,
                        event.changed_at.strftime("%Y-%m-%d %H:%M:%S") if event.changed_at else "",
                        event.review_summary or "",
                        event.open_points or "",
                        event.closure_comment or "",
                    )
                )
        if review_timeline_rows:
            add_sheet(
                "03_Action_Items",
                ["Category", "Tax Type", "Priority", "Reference", "Date / Period", "Amount", "Why Open", "Next Action"],
                [],
                index=3,
            )
            action_sheet = wb["03_Action_Items"]
            action_sheet.append([])
            action_sheet.append(["Reviewer Timeline", "", "", "", "", "", "", ""])
            action_sheet.append(["Scope", "Action", "Closure Status", "Reviewer", "Changed By", "Changed At", "Review Summary", "Open Points / Closure Comment"])
            for row in review_timeline_rows:
                action_sheet.append(
                    [
                        row[0], row[1], row[2], row[3], row[4], row[5],
                        row[6], f"{row[7]} {row[8]}".strip(),
                    ]
                )
        else:
            add_sheet(
                "03_Action_Items",
                ["Category", "Tax Type", "Priority", "Reference", "Date / Period", "Amount", "Why Open", "Next Action"],
                [],
                index=3,
            )

        action_rows = []
        for tax_label, payload in (("IT_TDS", it_exc_payload), ("GST_TDS", gst_exc_payload)):
            pending_challans = payload.get("exceptions", {}).get("invoices_pending_challan_mapping", {})
            for line in pending_challans.get("rows", []):
                action_rows.append(
                    (
                        "Pending Challan Mapping",
                        tax_label,
                        "High",
                        line.get("purchase_number") or f"Invoice#{line.get('header_id')}",
                        str(line.get("bill_date") or ""),
                        line.get("amount") or "0.00",
                        "Invoice still has statutory value not yet covered by a challan line.",
                        "Create or update challan mapping for this invoice.",
                    )
                )
            pending_returns = payload.get("exceptions", {}).get("challan_lines_pending_return_mapping", {})
            for line in pending_returns.get("rows", []):
                action_rows.append(
                    (
                        "Pending Return Mapping",
                        tax_label,
                        "High",
                        line.get("challan_no") or f"Challan#{line.get('challan_id')}",
                        "",
                        line.get("amount") or "0.00",
                        "Deposited challan value still needs to be consumed by a return line.",
                        "Prepare or update the relevant return before period closure.",
                    )
                )
            missing_evidence = payload.get("exceptions", {}).get("filed_returns_missing_ack_or_arn", {})
            for row in missing_evidence.get("rows", []):
                action_rows.append(
                    (
                        "Missing Filing Evidence",
                        tax_label,
                        "Medium",
                        row.get("return_code") or f"Return#{row.get('id')}",
                        f"{row.get('period_from')} to {row.get('period_to')}",
                        "",
                        "Filed return is still missing ACK or ARN details.",
                        "Capture acknowledgement references and attach supporting proof.",
                    )
                )
        for challan in open_draft_challans:
            action_rows.append(
                (
                    "Draft Challan",
                    challan.tax_type,
                    "Medium",
                    challan.challan_no or f"Draft Challan#{challan.id}",
                    str(challan.challan_date or ""),
                    str(challan.amount or "0.00"),
                    "Draft challan is still open and not yet ready for final closure.",
                    "Complete review, approval, and deposit if the challan is valid.",
                )
            )
        for filing in open_draft_returns:
            action_rows.append(
                (
                    "Draft Return",
                    filing.tax_type,
                    "Medium",
                    filing.return_code or f"Draft Return#{filing.id}",
                    f"{filing.period_from} to {filing.period_to}",
                    str(filing.amount or "0.00"),
                    "Draft return is still open and keeps the period from full closure.",
                    "Complete line review, approval, and filing once evidence is ready.",
                )
            )
        if action_rows:
            action_sheet = wb["03_Action_Items"]
            if action_sheet.max_row == 1:
                for row in action_rows:
                    action_sheet.append(list(row))
            else:
                action_sheet.append([])
                action_sheet.append(["Open Action Queue", "", "", "", "", "", "", ""])
                action_sheet.append(["Category", "Tax Type", "Priority", "Reference", "Date / Period", "Amount", "Why Open", "Next Action"])
                for row in action_rows:
                    action_sheet.append(list(row))

        it_invoices = [h for h in invoice_qs if (h.tds_amount or 0) > 0]
        add_sheet(
            "04_IT_TDS_Deductions",
            [
                "Invoice No", "Bill Date", "Vendor", "Vendor PAN", "Section Code", "Section Desc",
                "Base Amount", "TDS Rate", "TDS Amount", "Subentity", "Status",
            ],
            [
                (
                    h.purchase_number or f"{h.doc_code}-{h.doc_no}",
                    h.bill_date,
                    h.vendor_name or "",
                    (account_pan(h.vendor) or "") if h.vendor_id else "",
                    getattr(h.tds_section, "section_code", "") if h.tds_section_id else "",
                    getattr(h.tds_section, "description", "") if h.tds_section_id else "",
                    h.tds_base_amount, h.tds_rate, h.tds_amount,
                    getattr(h.subentity, "subentityname", "") if h.subentity_id else "",
                    h.get_status_display(),
                )
                for h in it_invoices
            ],
            total_columns=[6, 8],
        )

        it_challans = challan_qs.filter(tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS)
        add_sheet(
            "05_IT_TDS_Challans",
            ["Challan No", "Date", "Period From", "Period To", "Amount", "Interest", "Late Fee", "Penalty", "CIN", "BSR", "Minor Head", "Status"],
            [
                (c.challan_no, c.challan_date, c.period_from, c.period_to, c.amount, c.interest_amount, c.late_fee_amount, c.penalty_amount, c.cin_no, c.bsr_code, c.minor_head_code, c.get_status_display())
                for c in it_challans
            ],
            total_columns=[4, 5, 6, 7],
        )

        it_returns = return_qs.filter(tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS)
        it_return_rows = []
        for filing in it_returns:
            for line in filing.lines.all():
                it_return_rows.append(
                    (
                        filing.return_code, filing.period_from, filing.period_to,
                        line.header.purchase_number if line.header_id else "",
                        line.deductee_pan_snapshot or "",
                        line.section_snapshot_code or "",
                        line.amount,
                        line.challan.challan_no if line.challan_id else "",
                        line.cin_snapshot or "",
                        filing.ack_no or "", filing.arn_no or "",
                        filing.get_status_display(),
                    )
                )
        add_sheet(
            "06_IT_TDS_Returns",
            ["Return Code", "Period From", "Period To", "Invoice", "PAN", "Section", "Amount", "Challan", "CIN", "Ack No", "ARN No", "Status"],
            it_return_rows,
            total_columns=[6],
        )

        gst_invoices = [h for h in invoice_qs if (h.gst_tds_amount or 0) > 0]
        add_sheet(
            "07_GST_TDS_Deductions",
            ["Invoice No", "Bill Date", "Vendor", "Vendor GSTIN", "Contract Ref", "Base", "Rate", "CGST", "SGST", "IGST", "Total", "Subentity", "Status"],
            [
                (
                    h.purchase_number or f"{h.doc_code}-{h.doc_no}",
                    h.bill_date,
                    h.vendor_name or "",
                    h.vendor_gstin or "",
                    h.gst_tds_contract_ref or "",
                    h.gst_tds_base_amount, h.gst_tds_rate,
                    h.gst_tds_cgst_amount, h.gst_tds_sgst_amount, h.gst_tds_igst_amount, h.gst_tds_amount,
                    getattr(h.subentity, "subentityname", "") if h.subentity_id else "",
                    h.get_status_display(),
                )
                for h in gst_invoices
            ],
            total_columns=[5, 7, 8, 9, 10],
        )

        gst_challans = challan_qs.filter(tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS)
        add_sheet(
            "08_GST_TDS_Challans",
            ["Challan No", "Date", "Period From", "Period To", "Amount", "Interest", "Late Fee", "Penalty", "CIN", "BSR", "Minor Head", "Status"],
            [
                (c.challan_no, c.challan_date, c.period_from, c.period_to, c.amount, c.interest_amount, c.late_fee_amount, c.penalty_amount, c.cin_no, c.bsr_code, c.minor_head_code, c.get_status_display())
                for c in gst_challans
            ],
            total_columns=[4, 5, 6, 7],
        )

        gst_returns = return_qs.filter(tax_type=PurchaseStatutoryReturn.TaxType.GST_TDS)
        gst_return_rows = []
        for filing in gst_returns:
            for line in filing.lines.all():
                gst_return_rows.append(
                    (
                        filing.return_code, filing.period_from, filing.period_to,
                        line.header.purchase_number if line.header_id else "",
                        line.deductee_gstin_snapshot or "",
                        line.amount,
                        line.challan.challan_no if line.challan_id else "",
                        line.cin_snapshot or "",
                        filing.ack_no or "", filing.arn_no or "",
                        filing.get_status_display(),
                    )
                )
        add_sheet(
            "09_GST_TDS_Returns",
            ["Return Code", "Period From", "Period To", "Invoice", "Deductee GSTIN", "Amount", "Challan", "CIN", "Ack No", "ARN No", "Status"],
            gst_return_rows,
            total_columns=[5],
        )

        add_sheet(
            "10_Reconciliation",
            ["Metric", "IT_TDS", "GST_TDS"],
            [(k, it_summary.get(k, "0.00"), gst_summary.get(k, "0.00")) for k in ["deducted", "deposited", "filed", "pending_deposit", "pending_filing", "draft_challan", "draft_return"]],
        )

        add_sheet(
            "11_Exceptions",
            ["Exception", "IT_TDS", "GST_TDS"],
            [
                ("Pending challan mapping", it_exc.get("invoices_pending_challan_mapping", {}).get("line_count", 0), gst_exc.get("invoices_pending_challan_mapping", {}).get("line_count", 0)),
                ("Pending return mapping", it_exc.get("challan_lines_pending_return_mapping", {}).get("line_count", 0), gst_exc.get("challan_lines_pending_return_mapping", {}).get("line_count", 0)),
                ("Filed missing ACK/ARN", it_exc.get("filed_returns_missing_ack_or_arn", {}).get("count", 0), gst_exc.get("filed_returns_missing_ack_or_arn", {}).get("count", 0)),
            ],
        )

        add_sheet(
            "12_Supporting_Doc_Index",
            ["Doc Type", "Tax Type", "Number", "Date", "Attachment", "Status"],
            [
                ("Challan", c.tax_type, c.challan_no, c.challan_date, str(c.ack_document or ""), c.get_status_display())
                for c in challan_qs
            ] + [
                ("Return", r.tax_type, r.return_code, r.period_to, str(r.ack_document or ""), r.get_status_display())
                for r in return_qs
            ],
        )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="purchase_statutory_ca_pack.xlsx"'
        return response


class PurchaseStatutoryChallanEligibleLinesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        tax_type, period_from, period_to = _parse_required_period_and_tax_type(request)
        try:
            payload = PurchaseStatutoryService.challan_eligible_lines(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                tax_type=tax_type,
                period_from=period_from,
                period_to=period_to,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(payload)


class PurchaseStatutoryReturnEligibleLinesAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        tax_type, period_from, period_to = _parse_required_period_and_tax_type(request)
        return_code = request.query_params.get("return_code")
        try:
            payload = PurchaseStatutoryService.return_eligible_lines(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                tax_type=tax_type,
                period_from=period_from,
                period_to=period_to,
                return_code=return_code,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(payload)


class PurchaseStatutoryReconciliationExceptionsAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        entity_id, entityfinid_id, subentity_id = _parse_scope(request)
        _require_statutory_view(request, entity_id)
        tax_type, period_from, period_to = _parse_required_period_and_tax_type(request)
        try:
            payload = PurchaseStatutoryService.reconciliation_exceptions(
                entity_id=entity_id,
                entityfinid_id=entityfinid_id,
                subentity_id=subentity_id,
                tax_type=tax_type,
                period_from=period_from,
                period_to=period_to,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(payload)


class PurchaseStatutoryReturnNsdlExportAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="view")
        try:
            payload = PurchaseStatutoryService.generate_nsdl_payload(filing_id=pk)
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(payload)


class PurchaseStatutoryReturnForm16AIssueAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="view")
        try:
            payload = PurchaseStatutoryService.list_form16a_issues(filing_id=pk)
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response(payload)

    def post(self, request, pk: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="manage")
        issue_date = request.data.get("issue_date")
        remarks = request.data.get("remarks")
        try:
            payload = PurchaseStatutoryService.issue_form16a(
                filing_id=pk,
                issued_by_id=request.user.id,
                issue_date=issue_date,
                remarks=remarks,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": "Form16A issued.", "data": payload}, status=status.HTTP_201_CREATED)


class PurchaseStatutoryReturnForm16ADownloadAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _try_fill_pdf_form(self, file_field, cert: dict):
        try:
            src = file_field.open("rb")
            reader = PdfReader(src)
            writer = PdfWriter()
            if not reader.pages:
                return None

            rows = cert.get("lines") or []
            first = rows[0] if rows else {}
            form_map = {
                "ReturnID": str(cert.get("filing_id") or ""),
                "Return_Code": str(cert.get("return_code") or ""),
                "Period_From": str(cert.get("period_from") or ""),
                "Period_To": str(cert.get("period_to") or ""),
                "Issue_No": str(cert.get("issue_no") or ""),
                "Issue_Code": str(cert.get("issue_code") or ""),
                "Issued_On": str(cert.get("issued_on") or ""),
                "Total_Amount": str(cert.get("total_amount") or ""),
                "Line_Count": str(cert.get("line_count") or ""),
                "Invoice_No_1": str(first.get("invoice_no") or ""),
                "Bill_Date_1": str(first.get("bill_date") or ""),
                "Section_1": str(first.get("section_code") or ""),
                "PAN_1": str(first.get("pan") or ""),
                "GSTIN_1": str(first.get("gstin") or ""),
                "Challan_No_1": str(first.get("challan_no") or ""),
                "CIN_1": str(first.get("cin") or ""),
                "Amount_1": str(first.get("amount") or ""),
            }

            has_any_field = False
            for page in reader.pages:
                writer.add_page(page)
                try:
                    writer.update_page_form_field_values(page, form_map)
                    has_any_field = True
                except Exception:
                    continue
            if not has_any_field:
                src.close()
                return None

            out = BytesIO()
            writer.write(out)
            src.close()
            out.seek(0)
            return out
        except Exception:
            return None

    def _build_data_pdf(self, cert: dict) -> bytes:
        pdf_buffer = BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=A4)
        width, height = A4
        y = height - 40

        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, "Form 16A - System Generated Working Copy")
        y -= 16
        c.setFont("Helvetica", 9)
        c.drawString(40, y, "For final statutory certificate, use TRACES issued Form 16A.")
        y -= 22

        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, f"Return: {cert.get('return_code', '')}    Filing ID: {cert.get('filing_id', '')}")
        y -= 14
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Period: {cert.get('period_from', '')} to {cert.get('period_to', '')}")
        y -= 14
        c.drawString(40, y, f"Issue: {cert.get('issue_no', '')} ({cert.get('issue_code', '')}) on {cert.get('issued_on', '')}")
        y -= 14
        c.drawString(40, y, f"Line Count: {cert.get('line_count', 0)}    Total Amount: {cert.get('total_amount', '0.00')}")
        y -= 20

        headers = ["Invoice", "Date", "Section", "PAN", "GSTIN", "Challan", "CIN", "Amount"]
        col_x = [40, 110, 170, 230, 300, 380, 450, 520]
        c.setFont("Helvetica-Bold", 8)
        for idx, h in enumerate(headers):
            c.drawString(col_x[idx], y, h)
        y -= 10
        c.line(40, y, width - 40, y)
        y -= 10

        c.setFont("Helvetica", 8)
        for row in cert.get("lines") or []:
            if y < 60:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 8)
                for idx, h in enumerate(headers):
                    c.drawString(col_x[idx], y, h)
                y -= 10
                c.line(40, y, width - 40, y)
                y -= 10
                c.setFont("Helvetica", 8)
            values = [
                str(row.get("invoice_no") or "")[:12],
                str(row.get("bill_date") or "")[:10],
                str(row.get("section_code") or "")[:8],
                str(row.get("pan") or "")[:12],
                str(row.get("gstin") or "")[:15],
                str(row.get("challan_no") or "")[:10],
                str(row.get("cin") or "")[:10],
                str(row.get("amount") or ""),
            ]
            for idx, v in enumerate(values):
                c.drawString(col_x[idx], y, v)
            y -= 10

        c.showPage()
        c.save()
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()

    def _fill_nonfillable_template(self, file_field, cert: dict):
        try:
            src = file_field.open("rb")
            reader = PdfReader(src)
            writer = PdfWriter()
            if not reader.pages:
                src.close()
                return None

            overlays = []
            for page_index, page in enumerate(reader.pages):
                buf = BytesIO()
                c = canvas.Canvas(buf, pagesize=A4)
                c.setFont("Helvetica", 8)

                if page_index == 0:
                    # Top certificate meta
                    c.drawString(355, 742, str(cert.get("issue_code") or ""))
                    c.drawString(512, 742, str(cert.get("issued_on") or ""))

                    # Deductor/Deductee block (working approximation for govt template)
                    c.drawString(72, 682, f"Deductor: Entity #{cert.get('filing_id')}")
                    first = (cert.get("lines") or [{}])[0]
                    c.drawString(72, 603, f"Deductee PAN: {first.get('pan') or ''}")
                    c.drawString(72, 588, f"Deductee GSTIN: {first.get('gstin') or ''}")
                    c.drawString(72, 573, f"Section: {first.get('section_code') or ''}")

                    # Summary near lower section
                    c.drawString(68, 262, str(cert.get("period_from") or ""))
                    c.drawString(150, 262, str(cert.get("period_to") or ""))
                    c.drawString(240, 262, str(cert.get("line_count") or "0"))
                    c.drawString(300, 262, str(cert.get("total_amount") or "0.00"))

                    # First few line items in lower table area
                    y = 206
                    c.setFont("Helvetica", 7)
                    for row in (cert.get("lines") or [])[:3]:
                        c.drawString(42, y, str(row.get("invoice_no") or "")[:18])
                        c.drawString(140, y, str(row.get("bill_date") or "")[:10])
                        c.drawString(205, y, str(row.get("section_code") or "")[:8])
                        c.drawString(248, y, str(row.get("challan_no") or "")[:12])
                        c.drawRightString(350, y, str(row.get("amount") or "0.00"))
                        y -= 13

                elif page_index == 1:
                    c.setFont("Helvetica", 8)
                    c.drawString(78, 648, str(cert.get("total_amount") or "0.00"))
                    c.drawString(78, 634, f"Rupees {cert.get('total_amount') or '0.00'} only")
                    c.drawString(350, 648, str(cert.get("issued_on") or ""))
                    c.drawString(350, 634, str(cert.get("return_code") or ""))

                c.showPage()
                c.save()
                buf.seek(0)
                overlays.append(PdfReader(buf))

            for idx, page in enumerate(reader.pages):
                overlay_page = overlays[idx].pages[0]
                page.merge_page(overlay_page)
                writer.add_page(page)

            out = BytesIO()
            writer.write(out)
            src.close()
            out.seek(0)
            return out
        except Exception:
            return None

    def get(self, request, pk: int, issue_no: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="view")
        try:
            payload = PurchaseStatutoryService.form16a_download_payload(
                filing_id=pk, issue_no=issue_no
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        if payload.get("mode") == "file":
            f = payload.get("file_field")
            cert = payload.get("certificate_data") or {}
            filename = payload.get("filename") or f"form16a_{pk}_{issue_no}.pdf"
            if str(filename).lower().endswith(".pdf"):
                filled = self._try_fill_pdf_form(f, cert)
                if filled is not None:
                    response = HttpResponse(filled.getvalue(), content_type="application/pdf")
                    response["Content-Disposition"] = f'attachment; filename="{filename}"'
                    return response
                # Template has no fillable fields; try overlay-on-template.
                overlay_filled = self._fill_nonfillable_template(f, cert)
                if overlay_filled is not None:
                    response = HttpResponse(overlay_filled.getvalue(), content_type="application/pdf")
                    response["Content-Disposition"] = f'attachment; filename="form16a_{pk}_{issue_no}_filled_template.pdf"'
                    return response
                # Last fallback: generated data PDF.
                generated = self._build_data_pdf(cert)
                response = HttpResponse(generated, content_type="application/pdf")
                response["Content-Disposition"] = f'attachment; filename="form16a_{pk}_{issue_no}_generated.pdf"'
                return response
            response = FileResponse(f.open("rb"), content_type="application/octet-stream")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        # Backward-compatible service contract support (text payload).
        if payload.get("content") is not None:
            filename = payload.get("filename") or f"form16a_{pk}_{issue_no}.txt"
            content_type = "text/plain" if str(filename).lower().endswith(".txt") else "application/octet-stream"
            response = HttpResponse(payload.get("content"), content_type=content_type)
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        # Draft fallback as PDF so it opens in standard PDF readers.
        pdf_bytes = self._build_data_pdf(payload.get("certificate_data") or {})
        filename = f'form16a_{pk}_{issue_no}_draft.pdf'
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PurchaseStatutoryReturnForm16AOfficialUploadAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk: int, issue_no: int):
        _require_for_pk_if_resolvable(request, PurchaseStatutoryReturn, pk, level="manage")
        document = request.FILES.get("document")
        if not document:
            raise ValidationError({"detail": "document file is required."})
        source = request.data.get("source") or "TRACES"
        certificate_no = request.data.get("certificate_no")
        remarks = request.data.get("remarks")
        try:
            payload = PurchaseStatutoryService.attach_form16a_official_document(
                filing_id=pk,
                issue_no=issue_no,
                document=document,
                uploaded_by_id=request.user.id,
                source=source,
                certificate_no=certificate_no,
                remarks=remarks,
            )
        except ValueError as e:
            raise ValidationError({"detail": str(e)})
        return Response({"message": "Official Form16A document uploaded.", "data": payload}, status=status.HTTP_201_CREATED)
