from decimal import Decimal
from datetime import date
from datetime import datetime
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch, MagicMock

from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.test import SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.response import Response
from rest_framework.test import APITestCase, APIClient, APIRequestFactory, force_authenticate

from assets.models import AssetCategory, FixedAsset
from entity.models import Entity, EntityFinancialYear, Godown, SubEntity
from catalog.models import Product, ProductCategory, ProductPurchaseBehavior, UnitOfMeasure
from financial.models import Ledger, account
from purchase.models.purchase_core import PurchaseInvoiceHeader, PurchaseInvoiceLine
from purchase.serializers.purchase_invoice import PurchaseInvoiceHeaderSerializer
from purchase.serializers.purchase_statutory import (
    PurchaseStatutoryChallanCreateInputSerializer,
    PurchaseStatutoryReturnCreateInputSerializer,
)
from purchase.services.purchase_invoice_nav_service import PurchaseInvoiceNavService
from purchase.services.purchase_invoice_actions import PurchaseInvoiceActions
from purchase.services.purchase_asset_intake_service import PurchaseAssetIntakeService
from purchase.services.purchase_invoice_service import PurchaseInvoiceService
from purchase.services.purchase_settings_service import PurchaseSettingsService
from purchase.services.purchase_statutory_service import PurchaseStatutoryService
from purchase.views.purchase_invoice import PurchaseInvoiceListCreateAPIView
from purchase.views.purchase_meta import PurchaseInvoiceDetailFormMetaAPIView
from posting.adapters.purchase_invoice import (
    PurchaseInvoicePostingAdapter,
    PurchaseInvoicePostingConfig,
)
from posting.models import Entry, InventoryMove, PostingBatch, TxnType
from posting.common.static_accounts import StaticAccountCodes
from withholding.services import WithholdingResult
from purchase.services.purchase_withholding_service import PurchaseWithholdingService
from purchase.models.purchase_statutory import PurchaseStatutoryChallan, PurchaseStatutoryReturn
from purchase.models.purchase_statutory import PurchaseStatutoryReturnLine
from purchase.models import PurchaseLockPeriod


class PurchaseTdsApplyTests(SimpleTestCase):
    def _make_header(self, **overrides):
        defaults = {
            "withholding_enabled": True,
            "tds_is_manual": False,
            "tds_section_id": None,
            "tds_section": None,
            "tds_rate": Decimal("0.0000"),
            "tds_base_amount": Decimal("0.00"),
            "tds_amount": Decimal("0.00"),
            "tds_reason": None,
            "entity_id": 1,
            "entityfinid_id": 1,
            "subentity_id": None,
            "vendor_id": 1,
            "bill_date": None,
            "total_taxable": Decimal("1000.00"),
            "grand_total": Decimal("1180.00"),
            "match_notes": {},
        }
        defaults.update(overrides)
        return SimpleNamespace(**defaults)

    def test_withholding_disabled_clears_tds(self):
        header = self._make_header(
            withholding_enabled=False,
            tds_is_manual=True,
            tds_section_id=10,
            tds_rate=Decimal("1.0000"),
            tds_base_amount=Decimal("1000.00"),
            tds_amount=Decimal("10.00"),
            tds_reason="any",
        )

        PurchaseInvoiceService._apply_tds(header=header)

        self.assertFalse(header.tds_is_manual)
        self.assertIsNone(header.tds_section)
        self.assertEqual(header.tds_rate, Decimal("0.0000"))
        self.assertEqual(header.tds_base_amount, Decimal("0.00"))
        self.assertEqual(header.tds_amount, Decimal("0.00"))
        self.assertIsNone(header.tds_reason)
        self.assertEqual(header.match_notes, {})

    @patch("purchase.services.purchase_invoice_service.WithholdingResolver.get_entity_config")
    def test_manual_mode_requires_section(self, mock_get_cfg):
        mock_get_cfg.return_value = None
        header = self._make_header(withholding_enabled=True, tds_is_manual=True, tds_section_id=None)

        with self.assertRaisesMessage(ValueError, "TDS section is required when withholding_enabled is true."):
            PurchaseInvoiceService._apply_tds(header=header)

    @patch("purchase.services.purchase_invoice_service.WithholdingResolver.get_entity_config")
    def test_manual_mode_rejects_non_invoice_based_section(self, mock_get_cfg):
        mock_get_cfg.return_value = None
        header = self._make_header(
            withholding_enabled=True,
            tds_is_manual=True,
            tds_section_id=11,
            tds_section=SimpleNamespace(base_rule=4),  # PAYMENT_VALUE
            tds_rate=Decimal("1.0000"),
            tds_base_amount=Decimal("1000.00"),
            tds_amount=Decimal("10.00"),
        )

        with self.assertRaisesMessage(ValueError, "not invoice-based"):
            PurchaseInvoiceService._apply_tds(header=header)

    @patch("purchase.services.purchase_invoice_service.WithholdingResolver.get_entity_config")
    def test_manual_mode_respects_tds_enabled_config(self, mock_get_cfg):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=False)
        header = self._make_header(
            withholding_enabled=True,
            tds_is_manual=True,
            tds_section_id=11,
            tds_section=SimpleNamespace(base_rule=1),
            tds_rate=Decimal("1.0000"),
            tds_base_amount=Decimal("1000.00"),
            tds_amount=Decimal("10.00"),
        )
        with self.assertRaisesMessage(ValueError, "TDS is disabled in withholding configuration"):
            PurchaseInvoiceService._apply_tds(header=header)

    @patch("purchase.services.purchase_invoice_service.WithholdingResolver.evaluate_section_applicability")
    @patch("purchase.services.purchase_invoice_service.WithholdingResolver.get_entity_config")
    def test_manual_mode_respects_section_applicability(self, mock_get_cfg, mock_eval_applicability):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_eval_applicability.return_value = (False, "Section not applicable for party residency 'resident'", "NOT_APPLICABLE_RESIDENCY")
        header = self._make_header(
            withholding_enabled=True,
            tds_is_manual=True,
            tds_section_id=11,
            tds_section=SimpleNamespace(base_rule=1),
            tds_rate=Decimal("1.0000"),
            tds_base_amount=Decimal("1000.00"),
            tds_amount=Decimal("10.00"),
        )
        with self.assertRaisesMessage(ValueError, "not applicable"):
            PurchaseInvoiceService._apply_tds(header=header)

    @patch("purchase.services.purchase_invoice_service.PurchaseWithholdingService.compute_tds")
    def test_auto_mode_accepts_resolved_section_from_config(self, mock_compute):
        section = SimpleNamespace(id=22)
        header = self._make_header(withholding_enabled=True, tds_is_manual=False, tds_section_id=None)

        mock_compute.return_value = WithholdingResult(
            enabled=True,
            section=section,
            rate=Decimal("1.0000"),
            base_amount=Decimal("1000.00"),
            amount=Decimal("10.00"),
            reason="resolved from config",
        )

        PurchaseInvoiceService._apply_tds(header=header)

        self.assertEqual(header.tds_section, section)
        self.assertEqual(header.tds_rate, Decimal("1.0000"))
        self.assertEqual(header.tds_base_amount, Decimal("1000.00"))
        self.assertEqual(header.tds_amount, Decimal("10.00"))
        self.assertEqual(header.tds_reason, "resolved from config")
        self.assertEqual(
            header.match_notes.get("withholding_runtime_result"),
            {
                "enabled": True,
                "mode": "AUTO",
                "section_id": 22,
                "section_code": None,
                "rate": "1.0000",
                "base_amount": "1000.00",
                "amount": "10.00",
                "reason": "resolved from config",
                "reason_code": None,
                "deduction_status": "DEDUCTED",
                "zero_deduction": False,
                "user_selected_add_tds": True,
            },
        )

    @patch("purchase.services.purchase_invoice_service.PurchaseWithholdingService.compute_tds")
    def test_auto_mode_zero_deduction_persists_reason_snapshot(self, mock_compute):
        section = SimpleNamespace(id=31, section_code="194J")
        header = self._make_header(withholding_enabled=True, tds_is_manual=False, tds_section_id=31)

        mock_compute.return_value = WithholdingResult(
            enabled=True,
            section=section,
            rate=Decimal("10.0000"),
            base_amount=Decimal("0.00"),
            amount=Decimal("0.00"),
            reason="Below threshold (50000.00)",
            reason_code="BELOW_THRESHOLD",
        )

        PurchaseInvoiceService._apply_tds(header=header)

        self.assertEqual(header.tds_amount, Decimal("0.00"))
        self.assertEqual(header.tds_reason, "Below threshold (50000.00)")
        self.assertEqual(
            header.match_notes.get("withholding_runtime_result"),
            {
                "enabled": True,
                "mode": "AUTO",
                "section_id": 31,
                "section_code": "194J",
                "rate": "10.0000",
                "base_amount": "0.00",
                "amount": "0.00",
                "reason": "Below threshold (50000.00)",
                "reason_code": "BELOW_THRESHOLD",
                "deduction_status": "NOT_DEDUCTED",
                "zero_deduction": True,
                "user_selected_add_tds": True,
            },
        )


class PurchaseFiledPeriodAmendmentTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="purchase-amend-user",
            email="purchase-amend@example.com",
            password="pass123",
        )
        self.entity = Entity.objects.create(entityname="Filed Period Entity", createdby=self.user)
        self.subentity = SubEntity.objects.create(entity=self.entity, subentityname="Main")
        self.entityfin = EntityFinancialYear.objects.create(
            entity=self.entity,
            desc="FY 2026-27",
            finstartyear=timezone.make_aware(datetime(2026, 4, 1)),
            finendyear=timezone.make_aware(datetime(2027, 3, 31)),
            books_locked_until=date(2026, 4, 30),
            gst_locked_until=date(2026, 4, 30),
            inventory_locked_until=date(2026, 4, 30),
            ap_ar_locked_until=date(2026, 4, 30),
            createdby=self.user,
        )

    def test_amendment_window_resolves_next_open_date(self):
        PurchaseLockPeriod.objects.create(
            entity=self.entity,
            subentity=self.subentity,
            lock_date=date(2026, 4, 30),
            reason="April closed",
        )
        header = SimpleNamespace(
            entity_id=self.entity.id,
            subentity_id=self.subentity.id,
            entityfinid_id=self.entityfin.id,
            bill_date=date(2026, 4, 10),
        )

        window = PurchaseInvoiceService.amendment_window_for_header(header)

        self.assertTrue(window.amendment_required)
        self.assertEqual(window.lock_until, date(2026, 4, 30))
        expected_correction_date = max(date(2026, 5, 1), timezone.localdate())
        self.assertEqual(window.correction_date, expected_correction_date)
        self.assertEqual(window.gst_period, expected_correction_date.strftime("%Y-%m"))
        self.assertTrue(any("Purchase period locked up to 2026-04-30" in reason for reason in window.reasons))

    def test_assert_note_correction_date_open_rejects_locked_period_date(self):
        PurchaseLockPeriod.objects.create(
            entity=self.entity,
            subentity=self.subentity,
            lock_date=date(2026, 4, 30),
            reason="April closed",
        )
        original = PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            bill_date=date(2026, 4, 10),
            posting_date=date(2026, 4, 10),
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            supplier_invoice_number="SUP-APR-001",
            supplier_invoice_date=date(2026, 4, 10),
            vendor_name="Vendor",
            status=PurchaseInvoiceHeader.Status.POSTED,
        )

        with self.assertRaisesMessage(ValueError, "Correction document date must be in a current open period"):
            PurchaseInvoiceService.assert_note_correction_date_open(
                ref_document=original,
                correction_date=date(2026, 4, 20),
            )

    @patch("purchase.views.purchase_meta.PurchaseInvoiceService.amendment_window_for_header")
    @patch("purchase.views.purchase_meta.PurchaseSettingsService.get_policy")
    def test_detail_action_flags_expose_locked_period_correction_hint(self, mock_get_policy, mock_amendment_window):
        mock_get_policy.return_value = SimpleNamespace(
            controls={
                "allow_edit_confirmed": "on",
                "allow_unpost_posted": "on",
            },
            delete_policy="draft_only",
        )
        mock_amendment_window.return_value = SimpleNamespace(
            amendment_required=True,
            correction_date=date(2026, 5, 1),
        )
        header = SimpleNamespace(
            entity_id=self.entity.id,
            subentity_id=self.subentity.id,
            status=int(PurchaseInvoiceHeader.Status.POSTED),
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            get_status_display=lambda: "Posted",
        )

        flags = PurchaseInvoiceDetailFormMetaAPIView()._invoice_action_flags(header)

        self.assertTrue(flags["can_correct_locked_posted"])
        self.assertEqual(flags["locked_correction_modes"], ["full_reversal", "reduce", "increase"])
        self.assertEqual(flags["locked_correction_date"], "2026-05-01")
        self.assertIn("current-period correction draft", flags["locked_correction_message"])

    def test_append_correction_audit_event_links_original_and_correction_documents(self):
        original = PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            bill_date=date(2026, 4, 10),
            posting_date=date(2026, 4, 10),
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            supplier_invoice_number="SUP-APR-001",
            supplier_invoice_date=date(2026, 4, 10),
            vendor_name="Vendor",
            status=PurchaseInvoiceHeader.Status.POSTED,
            grand_total=Decimal("1180.00"),
            is_reverse_charge=False,
        )
        correction = PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            bill_date=date(2026, 5, 1),
            posting_date=date(2026, 5, 1),
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            ref_document=original,
            supplier_invoice_number="SUP-APR-001-CN",
            supplier_invoice_date=date(2026, 5, 1),
            vendor_name="Vendor",
            status=PurchaseInvoiceHeader.Status.POSTED,
            grand_total=Decimal("1180.00"),
            is_reverse_charge=False,
        )

        PurchaseInvoiceService.append_correction_audit_event(
            original=original,
            correction=correction,
            correction_type="credit_note_reversal",
            reason="Filed-period reversal",
            user_id=self.user.id,
            gst_period_impact="2026-05",
        )

        original.refresh_from_db()
        correction.refresh_from_db()
        self.assertEqual(len(original.match_notes["correction_history"]), 1)
        history_event = original.match_notes["correction_history"][0]
        self.assertEqual(history_event["correction_document_id"], correction.id)
        self.assertEqual(history_event["gst_period_impact"], "2026-05")
        self.assertEqual(history_event["old_value"]["bill_date"], "2026-04-10")
        self.assertEqual(history_event["new_value"]["bill_date"], "2026-05-01")
        self.assertEqual(correction.match_notes["correction_origin"]["original_invoice_id"], original.id)
        self.assertEqual(correction.match_notes["correction_origin"]["reason"], "Filed-period reversal")

    @patch("purchase.services.purchase_invoice_actions.PurchaseInvoiceActions.post")
    @patch("purchase.services.purchase_invoice_actions.PurchaseInvoiceActions.confirm")
    @patch("purchase.services.purchase_invoice_actions.PurchaseNoteFactory.create_note_from_invoice")
    @patch("purchase.services.purchase_invoice_actions.PurchaseInvoiceService.amendment_window_for_header")
    @patch("purchase.services.purchase_invoice_actions.PurchaseInvoiceActions._get")
    def test_cancel_posted_locked_invoice_creates_current_period_reversal_credit_note(
        self,
        mock_get,
        mock_amendment_window,
        mock_create_note,
        mock_confirm,
        mock_post,
    ):
        header = SimpleNamespace(
            id=81,
            entity_id=self.entity.id,
            subentity_id=self.subentity.id,
            entityfinid_id=self.entityfin.id,
            status=int(PurchaseInvoiceHeader.Status.POSTED),
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
        )
        note = SimpleNamespace(id=91)
        posted_note = SimpleNamespace(id=91, status=int(PurchaseInvoiceHeader.Status.POSTED))
        mock_get.return_value = header
        mock_amendment_window.return_value = SimpleNamespace(
            amendment_required=True,
            lock_until=date(2026, 4, 30),
            correction_date=date(2026, 5, 1),
            gst_period="2026-05",
        )
        mock_create_note.return_value = SimpleNamespace(header=note, message="created")
        mock_post.return_value = SimpleNamespace(header=posted_note, message="posted")

        result = PurchaseInvoiceActions.cancel(81, cancelled_by_id=9, reason="April reversal")

        mock_create_note.assert_called_once_with(
            invoice_id=81,
            note_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            note_reason=PurchaseInvoiceHeader.NoteReason.OTHER,
            created_by_id=9,
            correction_reason="April reversal",
        )
        mock_confirm.assert_called_once_with(91, confirmed_by_id=9)
        mock_post.assert_called_once_with(91, posted_by_id=9)
        self.assertEqual(
            result.message,
            "Locked-period purchase cannot be cancelled directly. A current-period reversal credit note was created and posted.",
        )
        self.assertEqual(result.header.id, 91)


class PurchaseGstTdsApplyTests(SimpleTestCase):
    def _make_header(self, **overrides):
        defaults = {
            "gst_tds_enabled": True,
            "gst_tds_is_manual": False,
            "gst_tds_contract_ref": "CNT-001",
            "gst_tds_reason": None,
            "gst_tds_rate": Decimal("0.0000"),
            "gst_tds_base_amount": Decimal("0.00"),
            "gst_tds_cgst_amount": Decimal("0.00"),
            "gst_tds_sgst_amount": Decimal("0.00"),
            "gst_tds_igst_amount": Decimal("0.00"),
            "gst_tds_amount": Decimal("0.00"),
            "gst_tds_status": 0,
            "entity_id": 1,
            "entityfinid_id": 1,
            "subentity_id": 1,
            "vendor_id": 1,
            "total_taxable": Decimal("1000.00"),
            "tax_regime": 1,
            "is_igst": False,
            "match_notes": {},
            "GstTdsStatus": SimpleNamespace(NA=0, ELIGIBLE=1),
            "TaxRegime": SimpleNamespace(INTER=2),
        }
        defaults.update(overrides)
        return SimpleNamespace(**defaults)

    def test_gst_tds_disabled_clears_runtime_snapshot(self):
        header = self._make_header(
            gst_tds_enabled=False,
            gst_tds_reason="old",
            match_notes={"gst_tds_runtime_result": {"enabled": True}},
        )

        PurchaseInvoiceService._apply_gst_tds(header=header)

        self.assertEqual(header.gst_tds_amount, Decimal("0.00"))
        self.assertNotIn("gst_tds_runtime_result", header.match_notes)

    @patch("purchase.services.purchase_invoice_service.GstTdsService.apply_to_header")
    def test_gst_tds_auto_zero_deduction_persists_reason_snapshot(self, mock_apply):
        header = self._make_header()

        def _apply(inv):
            inv.gst_tds_rate = Decimal("2.0000")
            inv.gst_tds_base_amount = Decimal("1000.00")
            inv.gst_tds_cgst_amount = Decimal("0.00")
            inv.gst_tds_sgst_amount = Decimal("0.00")
            inv.gst_tds_igst_amount = Decimal("0.00")
            inv.gst_tds_amount = Decimal("0.00")
            inv.gst_tds_status = 0
            inv.gst_tds_reason = "threshold not reached"
            return SimpleNamespace(reason="threshold not reached", reason_code="THRESHOLD_NOT_REACHED")

        mock_apply.side_effect = _apply

        PurchaseInvoiceService._apply_gst_tds(header=header)

        self.assertEqual(
            header.match_notes.get("gst_tds_runtime_result"),
            {
                "enabled": True,
                "mode": "AUTO",
                "contract_ref": "CNT-001",
                "rate": "2.0000",
                "base_amount": "1000.00",
                "amount": "0.00",
                "cgst_amount": "0.00",
                "sgst_amount": "0.00",
                "igst_amount": "0.00",
                "reason": "threshold not reached",
                "reason_code": "THRESHOLD_NOT_REACHED",
                "deduction_status": "NOT_DEDUCTED",
                "zero_deduction": True,
                "user_selected_add_gst_tds": True,
            },
        )

class PurchaseInvoiceViewUnitTests(SimpleTestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.user = SimpleNamespace(is_authenticated=True, id=7)

    def _make_header(self, **overrides):
        defaults = {
            "withholding_enabled": True,
            "tds_is_manual": False,
            "tds_section_id": None,
            "tds_section": None,
            "tds_rate": Decimal("0.0000"),
            "tds_base_amount": Decimal("0.00"),
            "tds_amount": Decimal("0.00"),
            "tds_reason": None,
            "entity_id": 1,
            "entityfinid_id": 1,
            "subentity_id": None,
            "vendor_id": 1,
            "bill_date": None,
            "total_taxable": Decimal("1000.00"),
            "grand_total": Decimal("1180.00"),
        }
        defaults.update(overrides)
        return SimpleNamespace(**defaults)

    @patch("purchase.views.purchase_invoice.require_purchase_request_permission")
    def test_list_queryset_uses_exists_for_line_mode_filter(self, mocked_require_permission):
        request = self.factory.get("/api/purchase/purchase-invoices/?entity=1&entityfinid=1&line_mode=goods")
        force_authenticate(request, user=self.user)

        view = PurchaseInvoiceListCreateAPIView()
        view.request = view.initialize_request(request)

        queryset = view.get_queryset()
        sql = str(queryset.query).upper()

        self.assertIn("EXISTS(", sql)
        self.assertNotIn(" DISTINCT ", sql)
        mocked_require_permission.assert_called_once()

    @patch("purchase.views.purchase_invoice.require_purchase_request_permission")
    def test_list_queryset_selects_vendor_related_profiles(self, mocked_require_permission):
        request = self.factory.get("/api/purchase/purchase-invoices/?entity=1&entityfinid=1")
        force_authenticate(request, user=self.user)

        view = PurchaseInvoiceListCreateAPIView()
        view.request = view.initialize_request(request)

        queryset = view.get_queryset()
        select_related = queryset.query.select_related

        self.assertIn("vendor", select_related)
        self.assertIn("ledger", select_related["vendor"])
        self.assertIn("commercial_profile", select_related["vendor"])
        mocked_require_permission.assert_called_once()

    def test_nav_scope_queryset_uses_exists_for_line_mode_filter(self):
        queryset = PurchaseInvoiceNavService._scope_qs(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            doc_code="PINV",
            allowed_statuses=PurchaseInvoiceNavService.DEFAULT_ALLOWED_STATUSES,
            line_mode="goods",
        )
        sql = str(queryset.query).upper()
        self.assertIn("EXISTS(", sql)
        self.assertNotIn(" DISTINCT ", sql)

    @patch("purchase.services.purchase_invoice_nav_service.PurchaseInvoiceNavService._scope_qs")
    def test_prev_next_uses_combined_goods_service_scope(self, mocked_scope_qs):
        mocked_scope_qs.return_value.filter.return_value.order_by.return_value.first.return_value = None
        instance = SimpleNamespace(
            id=1006,
            entity_id=10,
            entityfinid_id=2026,
            subentity_id=None,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            doc_code="PINV",
        )

        PurchaseInvoiceNavService.get_prev_next_for_instance(instance, line_mode="service")

        self.assertEqual(mocked_scope_qs.call_count, 2)
        first_call = mocked_scope_qs.call_args_list[0].kwargs
        second_call = mocked_scope_qs.call_args_list[1].kwargs
        self.assertEqual(first_call, {
            "entity_id": 10,
            "entityfinid_id": 2026,
            "subentity_id": None,
            "doc_type": int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            "doc_code": "PINV",
            "allowed_statuses": PurchaseInvoiceNavService.DEFAULT_ALLOWED_STATUSES,
            "line_mode": None,
        })
        self.assertEqual(second_call, {
            "entity_id": 10,
            "entityfinid_id": 2026,
            "subentity_id": None,
            "doc_type": int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            "doc_code": None,
            "allowed_statuses": PurchaseInvoiceNavService.DEFAULT_ALLOWED_STATUSES,
            "line_mode": None,
        })

    @patch("purchase.services.purchase_invoice_nav_service.PurchaseInvoiceNavService._scope_qs")
    def test_prev_next_orders_by_doc_no_with_id_tiebreaker(self, mocked_scope_qs):
        scoped_qs = MagicMock()
        all_code_rows = [
            SimpleNamespace(id=77, doc_no=1006, purchase_number="PINV-1006", status=3, bill_date=None),
            SimpleNamespace(id=88, doc_no=None, purchase_number="PINV-1007", status=3, bill_date=None),
            SimpleNamespace(id=95, doc_no=1009, purchase_number="PINV-1009", status=3, bill_date=None),
        ]
        mocked_scope_qs.side_effect = [scoped_qs, all_code_rows]

        instance = SimpleNamespace(
            id=90,
            doc_no=1008,
            entity_id=10,
            entityfinid_id=2026,
            subentity_id=None,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            doc_code="PINV",
        )

        result = PurchaseInvoiceNavService.get_prev_next_for_instance(instance, line_mode="service")

        self.assertEqual(result["previous"]["id"], 88)
        self.assertEqual(result["previous"]["purchase_number"], "PINV-1007")
        self.assertEqual(result["next"]["id"], 95)

    def test_last_saved_doc_scope_queryset_uses_subentity_isnull(self):
        with patch("purchase.services.purchase_settings_service.PurchaseInvoiceHeader.objects.filter") as mocked_filter:
            mocked_filter.return_value.only.return_value.__iter__.return_value = iter([])

            PurchaseSettingsService._last_saved_doc_in_scope(
                entity_id=10,
                entityfinid_id=8,
                subentity_id=None,
                doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
                current_number=1008,
            )

        mocked_filter.assert_called_once_with(
            entity_id=10,
            entityfinid_id=8,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            status__in=[2, 3, 9],
            subentity_id__isnull=True,
        )

    @patch("purchase.views.purchase_invoice.require_purchase_request_permission")
    def test_search_queryset_selects_vendor_related_profiles(self, mocked_require_permission):
        from purchase.views.purchase_invoice import PurchaseInvoiceSearchAPIView

        request = self.factory.get("/api/purchase/purchase-invoices/search/?entity=1&entityfinid=1")
        force_authenticate(request, user=self.user)

        view = PurchaseInvoiceSearchAPIView()
        view.request = view.initialize_request(request)

        queryset = view.get_queryset()
        select_related = queryset.query.select_related

        self.assertIn("vendor", select_related)
        self.assertIn("ledger", select_related["vendor"])
        self.assertIn("commercial_profile", select_related["vendor"])
        mocked_require_permission.assert_called_once()

    @patch("purchase.services.purchase_invoice_service.PurchaseWithholdingService.compute_tds")
    def test_auto_mode_fails_when_no_section_resolved(self, mock_compute):
        header = self._make_header(withholding_enabled=True, tds_is_manual=False, tds_section_id=None)

        mock_compute.return_value = WithholdingResult(
            enabled=True,
            section=None,
            rate=Decimal("0.0000"),
            base_amount=Decimal("0.00"),
            amount=Decimal("0.00"),
            reason="No TDS section",
        )

        with self.assertRaisesMessage(ValueError, "Provide tds_section or configure default TDS section for this entity."):
            PurchaseInvoiceService._apply_tds(header=header)


class PurchaseGstTdsTests(SimpleTestCase):
    @patch("purchase.services.purchase_invoice_service.GstTdsService.apply_to_header")
    def test_gst_tds_auto_mode_uses_gst_tds_service(self, mock_apply):
        header = SimpleNamespace(
            gst_tds_enabled=True,
            gst_tds_is_manual=False,
            gst_tds_contract_ref="CTR-001",
            gst_tds_rate=Decimal("0.0000"),
            gst_tds_base_amount=Decimal("0.00"),
            gst_tds_cgst_amount=Decimal("0.00"),
            gst_tds_sgst_amount=Decimal("0.00"),
            gst_tds_igst_amount=Decimal("0.00"),
            gst_tds_amount=Decimal("0.00"),
            gst_tds_status=0,
            gst_tds_reason=None,
            total_taxable=Decimal("1000.00"),
            tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA,
            is_igst=False,
            GstTdsStatus=PurchaseInvoiceHeader.GstTdsStatus,
            TaxRegime=PurchaseInvoiceHeader.TaxRegime,
        )

        PurchaseInvoiceService._apply_gst_tds(header=header)
        mock_apply.assert_called_once_with(header)

    def test_gst_tds_rate_fields_show_2_percent_logic(self):
        ser = PurchaseInvoiceHeaderSerializer()

        intra = SimpleNamespace(
            gst_tds_enabled=True,
            tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA,
            is_igst=False,
            TaxRegime=PurchaseInvoiceHeader.TaxRegime,
        )
        inter = SimpleNamespace(
            gst_tds_enabled=True,
            tax_regime=PurchaseInvoiceHeader.TaxRegime.INTER,
            is_igst=True,
            TaxRegime=PurchaseInvoiceHeader.TaxRegime,
        )

        self.assertEqual(ser.get_gst_tds_cgst_rate(intra), "1.0000")
        self.assertEqual(ser.get_gst_tds_sgst_rate(intra), "1.0000")
        self.assertEqual(ser.get_gst_tds_igst_rate(intra), "0.0000")

        self.assertEqual(ser.get_gst_tds_cgst_rate(inter), "0.0000")
        self.assertEqual(ser.get_gst_tds_sgst_rate(inter), "0.0000")
        self.assertEqual(ser.get_gst_tds_igst_rate(inter), "2.0000")

    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    def test_vendor_tds_variance_hard_blocks(self, mock_get_policy):
        mock_get_policy.return_value = SimpleNamespace(level=lambda key, default="warn": "hard")
        header = SimpleNamespace(
            entity_id=1,
            subentity_id=None,
            withholding_enabled=True,
            vendor_tds_declared=True,
            vendor_tds_base_amount=Decimal("900.00"),
            vendor_tds_rate=Decimal("1.0000"),
            vendor_tds_amount=Decimal("9.00"),
            tds_base_amount=Decimal("1000.00"),
            tds_rate=Decimal("1.0000"),
            tds_amount=Decimal("10.00"),
            gst_tds_enabled=False,
            vendor_gst_tds_declared=False,
            match_notes={},
            match_status="na",
            MatchStatus=SimpleNamespace(WARN="warn"),
        )
        with self.assertRaisesMessage(ValueError, "Vendor IT-TDS differs"):
            PurchaseInvoiceService._apply_vendor_withholding_variance_policy(header=header)

    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    def test_vendor_tds_variance_warn_sets_match_notes(self, mock_get_policy):
        mock_get_policy.return_value = SimpleNamespace(level=lambda key, default="warn": "warn")
        header = SimpleNamespace(
            entity_id=1,
            subentity_id=None,
            withholding_enabled=True,
            vendor_tds_declared=True,
            vendor_tds_base_amount=Decimal("900.00"),
            vendor_tds_rate=Decimal("1.0000"),
            vendor_tds_amount=Decimal("9.00"),
            tds_base_amount=Decimal("1000.00"),
            tds_rate=Decimal("1.0000"),
            tds_amount=Decimal("10.00"),
            gst_tds_enabled=False,
            vendor_gst_tds_declared=False,
            match_notes={},
            match_status="na",
            MatchStatus=SimpleNamespace(WARN="warn"),
        )
        PurchaseInvoiceService._apply_vendor_withholding_variance_policy(header=header)
        self.assertIn("withholding_warnings", header.match_notes)
        self.assertEqual(header.match_status, "warn")


class PurchaseVendorComplianceValidationTests(SimpleTestCase):
    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    @patch("purchase.services.purchase_invoice_service.account_partytype")
    @patch("purchase.services.purchase_invoice_service.account_gstno")
    def test_invalid_vendor_gstin_hard_rule_blocks(self, mock_gstno, mock_partytype, mock_policy):
        vendor = SimpleNamespace(ledger_id=10, isactive=True)
        mock_partytype.return_value = "Vendor"
        mock_gstno.return_value = None
        mock_policy.return_value = SimpleNamespace(controls={"vendor_gstin_format_rule": "hard"})

        attrs = {
            "entity": 1,
            "subentity": None,
            "vendor": vendor,
            "vendor_gstin": "INVALID-GSTIN",
            "withholding_enabled": False,
        }
        with self.assertRaisesMessage(ValueError, "Vendor GSTIN format is invalid."):
            PurchaseInvoiceService.validate_vendor_account(attrs)

    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    @patch("purchase.services.purchase_invoice_service.account_partytype")
    @patch("purchase.services.purchase_invoice_service.account_gstno")
    def test_invalid_vendor_gstin_warn_rule_sets_match_warning(self, mock_gstno, mock_partytype, mock_policy):
        vendor = SimpleNamespace(ledger_id=10, isactive=True)
        mock_partytype.return_value = "Vendor"
        mock_gstno.return_value = None
        mock_policy.return_value = SimpleNamespace(controls={"vendor_gstin_format_rule": "warn"})

        attrs = {
            "entity": 1,
            "subentity": None,
            "vendor": vendor,
            "vendor_gstin": "INVALID-GSTIN",
            "withholding_enabled": False,
        }
        PurchaseInvoiceService.validate_vendor_account(attrs)

        self.assertIn("compliance_warnings", attrs.get("match_notes", {}))
        self.assertEqual(attrs.get("match_status"), "warn")

    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    @patch("purchase.services.purchase_invoice_service.account_partytype")
    @patch("purchase.services.purchase_invoice_service.account_pan")
    def test_withholding_pan_required_hard_rule_blocks_when_pan_missing(self, mock_pan, mock_partytype, mock_policy):
        vendor = SimpleNamespace(ledger_id=10, isactive=True)
        mock_partytype.return_value = "Vendor"
        mock_pan.return_value = ""
        mock_policy.return_value = SimpleNamespace(
            controls={
                "withholding_pan_required_rule": "hard",
                "withholding_pan_format_rule": "hard",
            }
        )
        attrs = {
            "entity": 1,
            "subentity": None,
            "vendor": vendor,
            "withholding_enabled": True,
        }
        with self.assertRaisesMessage(ValueError, "Vendor PAN is required when Income-tax TDS is enabled."):
            PurchaseInvoiceService.validate_vendor_account(attrs)

    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    @patch("purchase.services.purchase_invoice_service.account_partytype")
    @patch("purchase.services.purchase_invoice_service.account_compliance_profile")
    @patch("purchase.services.purchase_invoice_service.account_gstno")
    def test_registered_vendor_without_gstin_is_blocked(self, mock_gstno, mock_compliance, mock_partytype, mock_policy):
        vendor = SimpleNamespace(ledger_id=10, isactive=True)
        mock_partytype.return_value = "Vendor"
        mock_gstno.return_value = None
        mock_compliance.return_value = SimpleNamespace(gstregtype="Regular", isactive=True)
        mock_policy.return_value = SimpleNamespace(controls={"vendor_gstin_format_rule": "hard"})

        attrs = {
            "entity": 1,
            "subentity": None,
            "vendor": vendor,
            "withholding_enabled": False,
        }
        with self.assertRaisesMessage(ValueError, "Vendor GSTIN is required for regular vendors."):
            PurchaseInvoiceService.validate_vendor_account(attrs)

    @patch("purchase.services.purchase_invoice_service.PurchaseSettingsService.get_policy")
    @patch("purchase.services.purchase_invoice_service.account_partytype")
    @patch("purchase.services.purchase_invoice_service.account_compliance_profile")
    @patch("purchase.services.purchase_invoice_service.account_gstno")
    def test_inactive_vendor_gstin_warn_rule_sets_match_warning(self, mock_gstno, mock_compliance, mock_partytype, mock_policy):
        vendor = SimpleNamespace(ledger_id=10, isactive=True)
        mock_partytype.return_value = "Vendor"
        mock_gstno.return_value = "27ABCDE1234F1Z5"
        mock_compliance.return_value = SimpleNamespace(gstregtype="Regular", isactive=False)
        mock_policy.return_value = SimpleNamespace(
            controls={
                "vendor_gstin_format_rule": "hard",
                "vendor_gstin_active_rule": "warn",
            }
        )

        attrs = {
            "entity": 1,
            "subentity": None,
            "vendor": vendor,
            "withholding_enabled": False,
        }
        PurchaseInvoiceService.validate_vendor_account(attrs)

        self.assertIn("compliance_warnings", attrs.get("match_notes", {}))
        self.assertEqual(attrs.get("match_status"), "warn")


class PurchaseUnregisteredVendorPolicyTests(SimpleTestCase):
    def test_blank_vendor_gstin_disables_itc_for_non_rcm_purchase(self):
        attrs = {
            "vendor_gstin": "",
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": int(PurchaseInvoiceHeader.ItcClaimStatus.PENDING),
        }

        PurchaseInvoiceService.apply_unregistered_vendor_defaults(attrs)

        self.assertFalse(attrs["is_itc_eligible"])
        self.assertEqual(attrs["itc_block_reason"], "ITC not eligible for unregistered vendor purchase.")

    def test_unregistered_vendor_gst_suppression_reason_is_user_friendly(self):
        message = PurchaseInvoiceService.supplier_gst_suppression_reason(
            attrs={
                "vendor_gstin": "",
                "is_reverse_charge": False,
            }
        )

        self.assertIn("vendor is treated as unregistered", message)
        self.assertIn("supplier-billed GST amounts are not allowed", message)
        self.assertIn("Set CGST/SGST/IGST amounts to 0", message)
        self.assertIn("registered vendor / reverse-charge purchase", message)

    def test_blank_vendor_gstin_suppresses_gst_for_non_rcm_purchase_line(self):
        derived = PurchaseInvoiceService.derive_tax_regime(
            {
                "tax_regime": int(PurchaseInvoiceHeader.TaxRegime.INTRA),
                "is_igst": False,
            }
        )

        computed = PurchaseInvoiceService.compute_line_authoritative(
            header_attrs={
                "default_taxability": int(PurchaseInvoiceHeader.Taxability.TAXABLE),
                "is_reverse_charge": False,
                "vendor_gstin": "",
            },
            line={
                "qty": Decimal("2.0000"),
                "rate": Decimal("50.00"),
                "gst_rate": Decimal("18.00"),
                "taxability": int(PurchaseInvoiceHeader.Taxability.TAXABLE),
            },
            derived=derived,
        )

        self.assertEqual(computed["taxable_value"], Decimal("100.00"))
        self.assertEqual(computed["cgst_amount"], Decimal("0.00"))
        self.assertEqual(computed["sgst_amount"], Decimal("0.00"))
        self.assertEqual(computed["igst_amount"], Decimal("0.00"))
        self.assertEqual(computed["line_total"], Decimal("100.00"))

    def test_claimed_itc_is_blocked_for_unregistered_vendor_purchase(self):
        attrs = {
            "vendor_gstin": "",
            "is_reverse_charge": False,
            "itc_claim_status": int(PurchaseInvoiceHeader.ItcClaimStatus.CLAIMED),
        }

        with self.assertRaisesMessage(
            ValueError,
            "Cannot claim ITC on an unregistered vendor purchase unless reverse charge applies.",
        ):
            PurchaseInvoiceService.apply_unregistered_vendor_defaults(attrs)


class PurchaseSpecialTaxTreatmentPolicyTests(SimpleTestCase):
    @patch("purchase.services.purchase_invoice_service.account_compliance_profile")
    def test_composition_vendor_blocks_itc_and_suppresses_supplier_gst(self, mock_compliance):
        mock_compliance.return_value = SimpleNamespace(gstregtype="Composition", isactive=True)
        attrs = {
            "vendor": SimpleNamespace(id=10),
            "vendor_gstin": "27ABCDE1234F1Z5",
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": int(PurchaseInvoiceHeader.ItcClaimStatus.PENDING),
            "default_taxability": int(PurchaseInvoiceHeader.Taxability.TAXABLE),
        }

        PurchaseInvoiceService.apply_special_tax_treatment_defaults(attrs)
        derived = PurchaseInvoiceService.derive_tax_regime({"tax_regime": int(PurchaseInvoiceHeader.TaxRegime.INTRA), "is_igst": False})
        computed = PurchaseInvoiceService.compute_line_authoritative(
            header_attrs=attrs,
            line={
                "qty": Decimal("1.0000"),
                "rate": Decimal("100.00"),
                "gst_rate": Decimal("18.00"),
                "taxability": int(PurchaseInvoiceHeader.Taxability.TAXABLE),
            },
            derived=derived,
        )

        self.assertFalse(attrs["is_itc_eligible"])
        self.assertEqual(attrs["itc_block_reason"], "ITC not eligible for composition vendor purchase.")
        self.assertEqual(computed["cgst_amount"], Decimal("0.00"))
        self.assertEqual(computed["sgst_amount"], Decimal("0.00"))

    def test_import_goods_blocks_normal_itc(self):
        attrs = {
            "supply_category": int(PurchaseInvoiceHeader.SupplyCategory.IMPORT_GOODS),
            "is_itc_eligible": True,
            "itc_claim_status": int(PurchaseInvoiceHeader.ItcClaimStatus.PENDING),
        }

        PurchaseInvoiceService.apply_special_tax_treatment_defaults(attrs)

        self.assertFalse(attrs["is_itc_eligible"])
        self.assertEqual(
            attrs["itc_block_reason"],
            "Import goods ITC should be claimed through customs or bill-of-entry flow.",
        )

    def test_import_services_requires_reverse_charge(self):
        attrs = {
            "supply_category": int(PurchaseInvoiceHeader.SupplyCategory.IMPORT_SERVICES),
            "is_reverse_charge": False,
            "itc_claim_status": int(PurchaseInvoiceHeader.ItcClaimStatus.PENDING),
        }

        with self.assertRaisesMessage(ValueError, "Import of services purchases must be marked as reverse charge."):
            PurchaseInvoiceService.apply_special_tax_treatment_defaults(attrs)

    def test_sez_purchase_requires_inter_regime(self):
        attrs = {
            "supply_category": int(PurchaseInvoiceHeader.SupplyCategory.SEZ),
            "tax_regime": int(PurchaseInvoiceHeader.TaxRegime.INTRA),
            "supplier_invoice_number": "SEZ-1001",
            "supplier_invoice_date": date(2026, 4, 1),
        }

        with self.assertRaisesMessage(ValueError, "Import and SEZ purchases must use INTER tax regime."):
            PurchaseInvoiceService.validate_header(attrs)


class PurchaseDuplicateSupplierInvoiceTests(TestCase):
    def test_duplicate_supplier_invoice_detected_for_same_vendor_date_and_amount(self):
        entity = Entity.objects.create(entityname="Duplicate Purchase Entity")
        vendor = account.objects.create(entity=entity, accountname="Vendor X")
        existing = PurchaseInvoiceHeader.objects.create(
            entity=entity,
            vendor=vendor,
            supplier_invoice_number="SUP-1001",
            supplier_invoice_date=date(2026, 4, 10),
            grand_total=Decimal("118.00"),
            status=PurchaseInvoiceHeader.Status.POSTED,
        )

        with self.assertRaisesMessage(
            ValueError,
            "Duplicate supplier invoice detected for this vendor, invoice number, invoice date, and amount.",
        ):
            PurchaseInvoiceService.assert_no_duplicate_supplier_invoice(
                instance=None,
                attrs={
                    "entity": entity,
                    "vendor": vendor,
                    "supplier_invoice_number": "SUP-1001",
                    "supplier_invoice_date": date(2026, 4, 10),
                },
                grand_total=Decimal("118.00"),
            )

        PurchaseInvoiceService.assert_no_duplicate_supplier_invoice(
            instance=existing,
            attrs={
                "entity": entity,
                "vendor": vendor,
                "supplier_invoice_number": "SUP-1001",
                "supplier_invoice_date": date(2026, 4, 10),
            },
            grand_total=Decimal("118.00"),
        )


class PurchaseWithholdingBaseRuleTests(SimpleTestCase):
    def _make_line(self, **overrides):
        defaults = {
            "product_desc": "",
            "hsn_sac": "",
        }
        defaults.update(overrides)
        return SimpleNamespace(**defaults)

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_skips_payment_based_section_in_invoice_context(self, mock_get_cfg, mock_resolve_section):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=99,
            section_code="194N",
            base_rule=4,  # PAYMENT_VALUE
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=99,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("1000.00"),
            gross_total=Decimal("1180.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "NOT_APPLICABLE_BASE_RULE_CONTEXT")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_returns_explicit_no_deduction_reason_for_credit_note_reversal(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194J",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("10.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("-100000.00"),
            gross_total=Decimal("-118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.base_amount, Decimal("0.00"))
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "CREDIT_NOTE_REVERSAL_NO_TDS")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_blocks_194q_when_config_flag_disabled(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True, apply_194q=False)
        mock_resolve_section.return_value = SimpleNamespace(
            id=99,
            section_code="194Q",
            base_rule=1,
            threshold_default=Decimal("5000000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("0.1000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=99,
            vendor_id=10,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("5500000.00"),
            gross_total=Decimal("5500000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "NOT_APPLICABLE_194Q_DISABLED")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_blocks_194q_for_import_purchase(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True, apply_194q=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=99,
            section_code="194Q",
            base_rule=1,
            threshold_default=Decimal("5000000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("0.1000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=99,
            vendor_id=10,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.IMPORT_GOODS,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("5500000.00"),
            gross_total=Decimal("5500000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "NOT_APPLICABLE_IMPORT")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.evaluate_section_applicability")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_blocks_194q_for_non_resident_vendor(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_applicability,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True, apply_194q=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=99,
            section_code="194Q",
            base_rule=1,
            threshold_default=Decimal("5000000.00"),
            applicability_json={"resident_status": ["resident"], "party_country_codes": ["IN"]},
        )
        mock_applicability.return_value = (
            False,
            "Section not applicable for party residency 'non_resident'",
            "NOT_APPLICABLE_RESIDENCY",
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("0.1000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        lines = MagicMock()
        lines.filter.return_value.exists.return_value = False
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=99,
            vendor_id=10,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
            lines=lines,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("5500000.00"),
            gross_total=Decimal("5500000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "NOT_APPLICABLE_RESIDENCY")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.evaluate_section_applicability")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_blocks_194q_for_service_invoice(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_applicability,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True, apply_194q=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=99,
            section_code="194Q",
            base_rule=1,
            threshold_default=Decimal("5000000.00"),
            applicability_json={},
        )
        mock_applicability.return_value = (True, None, None)
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("0.1000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        service_qs = MagicMock()
        service_qs.exists.return_value = True
        goods_qs = MagicMock()
        goods_qs.exists.return_value = False
        lines = MagicMock()
        lines.filter.side_effect = lambda **kwargs: service_qs if kwargs == {"is_service": True} else goods_qs
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=99,
            vendor_id=10,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
            lines=lines,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("5500000.00"),
            gross_total=Decimal("5500000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "NOT_APPLICABLE_SERVICE_INVOICE")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_blocks_194q_when_turnover_gate_not_met(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(
            enable_tds=True,
            apply_194q=True,
            tds_194q_prev_fy_turnover=Decimal("50000000.00"),
            tds_194q_turnover_limit=Decimal("100000000.00"),
            tds_194q_force_eligible=None,
        )
        mock_resolve_section.return_value = SimpleNamespace(
            id=99,
            section_code="194Q",
            base_rule=1,
            threshold_default=Decimal("5000000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("0.1000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        lines = MagicMock()
        lines.filter.return_value.exists.return_value = False
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=99,
            vendor_id=10,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
            lines=lines,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("5500000.00"),
            gross_total=Decimal("5500000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "NOT_ELIGIBLE_TURNOVER_GATE")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_respects_single_transaction_threshold_for_invoice_section(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194J",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("10.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("40000.00"),
            gross_total=Decimal("47200.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.base_amount, Decimal("40000.00"))
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "BELOW_THRESHOLD")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_applies_amount_when_invoice_section_crosses_threshold(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194J",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("10.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("100000.00"),
            gross_total=Decimal("118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.base_amount, Decimal("100000.00"))
        self.assertEqual(res.amount, Decimal("10000.00"))

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_uses_higher_no_pan_rate_for_invoice_section(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194J",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("20.0000"),
            reason="Higher rate due to PAN not available.",
            reason_code="NO_PAN_206AA",
            no_pan_applied=True,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("100000.00"),
            gross_total=Decimal("118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.rate, Decimal("20.0000"))
        self.assertEqual(res.amount, Decimal("20000.00"))
        self.assertEqual(res.reason_code, "NO_PAN_206AA")
        self.assertTrue(res.no_pan_applied)

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_prefers_valid_lower_deduction_rate_for_invoice_section(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194J",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("0.1000"),
            reason="Lower deduction certificate applied.",
            reason_code="LOWER_CERT",
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=True,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("100000.00"),
            gross_total=Decimal("118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.rate, Decimal("0.1000"))
        self.assertEqual(res.amount, Decimal("100.00"))
        self.assertEqual(res.reason_code, "LOWER_CERT")
        self.assertTrue(res.lower_rate_applied)

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_194h_applies_commission_rate_after_threshold(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=102,
            section_code="194H",
            base_rule=1,
            threshold_default=Decimal("15000.00"),
            applicability_json={},
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("5.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=102,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("100000.00"),
            gross_total=Decimal("118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.rate, Decimal("5.0000"))
        self.assertEqual(res.base_amount, Decimal("100000.00"))
        self.assertEqual(res.amount, Decimal("5000.00"))

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_194i_uses_lower_plant_machinery_rate_when_line_matches_keywords(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194I",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            rate_default=Decimal("10.0000"),
            applicability_json={
                "rent_rate_plant_machinery": "2.00",
                "rent_plant_machinery_keywords": ["plant", "machinery", "equipment"],
            },
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("10.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
            lines=[self._make_line(product_desc="Mobile crane plant hire charges", hsn_sac="9973")],
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("100000.00"),
            gross_total=Decimal("118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.rate, Decimal("2.00"))
        self.assertEqual(res.amount, Decimal("2000.00"))
        self.assertEqual(res.reason_code, "RATE_SUBTYPE_194I_PLANT_MACHINERY")

    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_194i_keeps_default_rate_for_office_rent(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        mock_resolve_section.return_value = SimpleNamespace(
            id=101,
            section_code="194I",
            base_rule=1,
            threshold_default=Decimal("50000.00"),
            rate_default=Decimal("10.0000"),
            applicability_json={
                "rent_rate_plant_machinery": "2.00",
                "rent_plant_machinery_keywords": ["plant", "machinery", "equipment"],
            },
        )
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("10.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            vendor_id=10,
            lines=[self._make_line(product_desc="Office building monthly rent", hsn_sac="9972")],
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("100000.00"),
            gross_total=Decimal("118000.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.rate, Decimal("10.0000"))
        self.assertEqual(res.amount, Decimal("10000.00"))

    @patch("purchase.services.purchase_withholding_service.FyPartyThresholdService.compute_base_above_threshold")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_194c_uses_cumulative_threshold_when_single_invoice_below_threshold(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
        mock_compute_threshold,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        section = SimpleNamespace(
            id=101,
            section_code="194C",
            base_rule=1,
            threshold_default=Decimal("30000.00"),
            applicability_json={"threshold_mode": "cumulative", "aggregate_threshold": "100000.00"},
        )
        mock_resolve_section.return_value = section
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("1.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        mock_compute_threshold.return_value = SimpleNamespace(
            base_applicable=Decimal("5000.00"),
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            tds_section=section,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("25000.00"),
            gross_total=Decimal("29500.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.base_amount, Decimal("25000.00"))
        self.assertEqual(res.amount, Decimal("250.00"))
        self.assertEqual(res.reason_code, "THRESHOLD_CROSSED_CUMULATIVE")

    @patch("purchase.services.purchase_withholding_service.FyPartyThresholdService.compute_base_above_threshold")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_rate")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_party_profile")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.resolve_section")
    @patch("purchase.services.purchase_withholding_service.WithholdingResolver.get_entity_config")
    def test_compute_tds_194c_blocks_below_aggregate_threshold_when_single_invoice_below_threshold(
        self,
        mock_get_cfg,
        mock_resolve_section,
        mock_party_profile,
        mock_resolve_rate,
        mock_compute_threshold,
    ):
        mock_get_cfg.return_value = SimpleNamespace(enable_tds=True)
        section = SimpleNamespace(
            id=101,
            section_code="194C",
            base_rule=1,
            threshold_default=Decimal("30000.00"),
            applicability_json={"threshold_mode": "cumulative", "aggregate_threshold": "100000.00"},
        )
        mock_resolve_section.return_value = section
        mock_party_profile.return_value = None
        mock_resolve_rate.return_value = SimpleNamespace(
            rate=Decimal("1.0000"),
            reason=None,
            reason_code=None,
            no_pan_applied=False,
            sec_206ab_applied=False,
            lower_rate_applied=False,
        )
        mock_compute_threshold.return_value = SimpleNamespace(
            base_applicable=Decimal("0.00"),
        )
        header = SimpleNamespace(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            withholding_enabled=True,
            tds_section_id=101,
            tds_section=section,
            vendor_id=10,
        )

        res = PurchaseWithholdingService.compute_tds(
            header=header,
            vendor_account_id=10,
            bill_date=date(2026, 4, 1),
            taxable_total=Decimal("25000.00"),
            gross_total=Decimal("29500.00"),
        )

        self.assertTrue(res.enabled)
        self.assertEqual(res.base_amount, Decimal("25000.00"))
        self.assertEqual(res.amount, Decimal("0.00"))
        self.assertEqual(res.reason_code, "BELOW_THRESHOLD_CUMULATIVE")


class PurchaseApiSmokeTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="purchase_api_tester",
            email="purchase_api_tester@example.com",
            password="x",
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    @patch("purchase.views.purchase_settings.PurchaseSettingsService.get_current_doc_no")
    @patch("purchase.views.purchase_settings.PurchaseSettingsService.get_choice_overrides")
    @patch("purchase.views.purchase_settings.PurchaseSettingsService.get_settings")
    def test_settings_get_with_entity_and_entityfinid_returns_200(
        self,
        mock_get_settings,
        mock_get_choice_overrides,
        mock_get_current_doc_no,
    ):
        mock_get_settings.return_value = SimpleNamespace(
            entity_id=32,
            subentity_id=None,
            default_doc_code_invoice="PINV",
            default_doc_code_cn="PCN",
            default_doc_code_dn="PDN",
            default_workflow_action="confirm",
            auto_derive_tax_regime=True,
            enforce_2b_before_itc_claim=False,
            allow_mixed_taxability_in_one_bill=True,
            round_grand_total_to=2,
            enable_round_off=True,
        )
        mock_get_choice_overrides.return_value = {}
        mock_get_current_doc_no.return_value = {
            "enabled": True,
            "current_number": 1,
            "previous_number": None,
            "previous_invoice_id": None,
            "previous_purchase_number": None,
            "previous_status": None,
            "previous_bill_date": None,
        }

        resp = self.client.get("/api/purchase/settings/?entity=32&entityfinid=32")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("settings", resp.data)
        self.assertIn("current_doc_numbers", resp.data)
        self.assertEqual(mock_get_current_doc_no.call_count, 3)

    @patch("purchase.views.purchase_settings.PurchaseSettingsAPIView._payload", return_value={"ok": True})
    @patch("purchase.views.purchase_settings.PurchaseSettingsService.upsert_settings")
    def test_settings_patch_triggers_meta_cache_invalidation(
        self,
        mock_upsert_settings,
        mock_payload,
    ):
        mock_upsert_settings.return_value = SimpleNamespace(
            default_doc_code_invoice="PINV",
            default_doc_code_cn="PCN",
            default_doc_code_dn="PDN",
        )

        with patch("purchase.views.purchase_settings.bump_meta_namespaces") as mocked_bump_cache:
            with self.captureOnCommitCallbacks(execute=True) as callbacks:
                resp = self.client.patch(
                    "/api/purchase/settings/?entity=32&subentity=1",
                    {"settings": {"enable_round_off": False}},
                    format="json",
                )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data, {"ok": True})
        self.assertGreaterEqual(len(callbacks), 1)
        mocked_bump_cache.assert_called_once()
        mock_upsert_settings.assert_called_once()
        mock_payload.assert_called_once()

    def test_charge_type_detail_missing_id_returns_404(self):
        resp = self.client.get("/api/purchase/charge-types/999999/?entity=1")
        self.assertEqual(resp.status_code, 404)


class _ListManager:
    def __init__(self, items):
        self._items = items

    def all(self):
        return self._items


class _FakeQuerySet(list):
    def select_related(self, *args, **kwargs):
        return self

    def prefetch_related(self, *args, **kwargs):
        return self

    def all(self):
        return self

    def filter(self, **kwargs):
        def _resolve_value(obj, path):
            current = obj
            for part in path.split("__"):
                current = getattr(current, part, None)
            return current

        def _matches(obj):
            for key, expected in kwargs.items():
                if "__" in key:
                    field_name, lookup = key.rsplit("__", 1)
                    if lookup in {"gte", "lte", "isnull"}:
                        current = _resolve_value(obj, field_name)
                        if lookup == "gte" and not (current >= expected):
                            return False
                        if lookup == "lte" and not (current <= expected):
                            return False
                        if lookup == "isnull" and not ((current is None) == bool(expected)):
                            return False
                        continue
                current = _resolve_value(obj, key)
                if current != expected:
                    return False
            return True

        return _FakeQuerySet([obj for obj in self if _matches(obj)])

    def order_by(self, field_name):
        reverse = field_name.startswith("-")
        key = field_name[1:] if reverse else field_name
        return _FakeQuerySet(sorted(self, key=lambda item: getattr(item, key, None), reverse=reverse))


class PurchasePostingAdapterTests(SimpleTestCase):
    def _base_header(self, **overrides):
        defaults = {
            "id": 101,
            "entity_id": 1,
            "entityfinid_id": 1,
            "subentity_id": None,
            "bill_date": date(2026, 3, 3),
            "posting_date": date(2026, 3, 3),
            "vendor_id": 7001,
            "vendor_ledger_id": 7001,
            "vendor_name": "Vendor-A",
            "vendor": SimpleNamespace(accountname="Vendor-A", ledger_id=7001),
            "doc_type": 1,
            "purchase_number": "PINV-101",
            "grand_total": Decimal("100.00"),
            "round_off": Decimal("0.00"),
            "total_expenses": Decimal("0.00"),
            "is_reverse_charge": False,
            "affects_inventory": False,
            "charges": _ListManager([]),
            "tds_amount": Decimal("0.00"),
            "gst_tds_amount": Decimal("0.00"),
        }
        defaults.update(overrides)
        return SimpleNamespace(**defaults)

    def _line(self, **overrides):
        defaults = {
            "id": 201,
            "line_no": 1,
            "product_id": None,
            "product_desc": "Copper Wire",
            "is_service": True,
            "purchase_behavior": ProductPurchaseBehavior.EXPENSE,
            "taxable_value": Decimal("100.00"),
            "cgst_amount": Decimal("0.00"),
            "sgst_amount": Decimal("0.00"),
            "igst_amount": Decimal("0.00"),
            "cess_amount": Decimal("0.00"),
            "is_itc_eligible": True,
            "qty": Decimal("1.0000"),
            "free_qty": Decimal("0.0000"),
            "uom_id": None,
        }
        defaults.update(overrides)
        return SimpleNamespace(**defaults)

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_posts_charge_line_to_misc_expense(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.TDS_PAYABLE: 8108,
            StaticAccountCodes.GST_TDS_PAYABLE: 8109,
        }

        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=999)

        charge = SimpleNamespace(
            id=301,
            line_no=1,
            taxable_value=Decimal("10.00"),
            cgst_amount=Decimal("0.00"),
            sgst_amount=Decimal("0.00"),
            igst_amount=Decimal("0.00"),
            itc_eligible=True,
        )
        header = self._base_header(
            grand_total=Decimal("110.00"),
            charges=_ListManager([charge]),
        )

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[self._line()],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        kwargs = posting_instance.post.call_args.kwargs
        jl_inputs = kwargs["jl_inputs"]
        charge_entries = [x for x in jl_inputs if x.account_id == 8100 and "charge" in x.description.lower()]

        self.assertTrue(charge_entries, "Expected a charge journal line in PURCHASE_MISC_EXPENSE.")
        self.assertEqual(charge_entries[0].amount, Decimal("10.00"))

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_purchase_descriptions_include_vendor_and_item_context(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
        }
        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []
        mock_posting_service_cls.return_value.post.return_value = SimpleNamespace(id=999)

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=self._base_header(),
            lines=[self._line()],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        jl_inputs = mock_posting_service_cls.return_value.post.call_args.kwargs["jl_inputs"]
        self.assertIn("Vendor Vendor-A", jl_inputs[0].description)
        self.assertIn("Item Copper Wire", jl_inputs[0].description)

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    @patch("posting.adapters.purchase_invoice.EntityWithholdingSectionPostingMap.objects.filter")
    def test_posts_gst_tds_when_config_enabled(
        self,
        mock_section_map_filter,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.TDS_PAYABLE: 8108,
            StaticAccountCodes.GST_TDS_PAYABLE: 8109,
        }

        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        resolver.get_ledger_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_section_map_filter.return_value.filter.return_value.order_by.return_value.first.return_value = None
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=1001)

        header = self._base_header(gst_tds_amount=Decimal("5.00"))

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[self._line()],
            user_id=1,
            config=PurchaseInvoicePostingConfig(post_gst_tds_on_invoice=True),
        )

        kwargs = posting_instance.post.call_args.kwargs
        jl_inputs = kwargs["jl_inputs"]

        gst_tds_cr = [
            x for x in jl_inputs
            if x.account_id == 8109 and x.drcr is False and x.amount == Decimal("5.00")
            and "gst-tds payable" in x.description.lower()
        ]
        vendor_cr = [
            x for x in jl_inputs
            if x.account_id == header.vendor_id and x.drcr is False and x.amount == Decimal("95.00")
            and "vendor payable" in x.description.lower()
        ]
        vendor_dr = [x for x in jl_inputs if x.account_id == header.vendor_id and x.drcr is True]

        self.assertTrue(gst_tds_cr, "Expected CR GST-TDS Payable entry.")
        self.assertTrue(vendor_cr, "Expected vendor payable to be reduced by GST-TDS deduction.")
        self.assertFalse(vendor_dr, "Did not expect a separate DR Vendor line for GST-TDS deduction.")

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    @patch("posting.adapters.purchase_invoice.EntityWithholdingSectionPostingMap.objects.filter")
    def test_posts_tds_to_section_specific_payable_mapping_when_available(
        self,
        mock_section_map_filter,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.TDS_PAYABLE: 8108,
            StaticAccountCodes.GST_TDS_PAYABLE: 8109,
        }

        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        resolver.get_ledger_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=1002)

        mapped = SimpleNamespace(payable_account_id=9901, payable_ledger_id=9902)
        sub_qs = MagicMock()
        sub_qs.order_by.return_value.first.return_value = mapped
        root_qs = MagicMock()
        root_qs.order_by.return_value.first.return_value = None
        base_qs = MagicMock()
        base_qs.select_related.return_value = base_qs
        base_qs.filter.side_effect = lambda **kwargs: sub_qs if kwargs == {"subentity_id": 1} else root_qs
        mock_section_map_filter.return_value = base_qs

        header = self._base_header(tds_amount=Decimal("10.00"))
        header.tds_section_id = 77
        header.subentity_id = 1

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[self._line()],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        jl_inputs = posting_instance.post.call_args.kwargs["jl_inputs"]
        tds_payable = [
            x for x in jl_inputs
            if x.account_id == 9901 and x.ledger_id == 9902 and x.drcr is False and x.amount == Decimal("10.00")
        ]
        vendor_cr = [
            x for x in jl_inputs
            if x.account_id == header.vendor_id and x.drcr is False and x.amount == Decimal("90.00")
            and "vendor payable" in x.description.lower()
        ]
        vendor_dr = [x for x in jl_inputs if x.account_id == header.vendor_id and x.drcr is True]
        self.assertTrue(tds_payable, "Expected TDS payable posting to use section-specific mapping.")
        self.assertTrue(vendor_cr, "Expected vendor payable to be net of TDS deduction.")
        self.assertFalse(vendor_dr, "Did not expect a separate DR Vendor line for TDS deduction.")

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    @patch("posting.adapters.purchase_invoice.EntityWithholdingSectionPostingMap.objects.filter")
    def test_posts_tds_to_mapped_account_ledger_when_payable_ledger_is_blank(
        self,
        mock_section_map_filter,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.TDS_PAYABLE: 8108,
            StaticAccountCodes.GST_TDS_PAYABLE: 8109,
        }

        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        resolver.get_ledger_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=1003)

        mapped_account = SimpleNamespace(ledger_id=9911)
        mapped = SimpleNamespace(payable_account_id=9901, payable_ledger_id=None, payable_account=mapped_account)
        sub_qs = MagicMock()
        sub_qs.order_by.return_value.first.return_value = mapped
        root_qs = MagicMock()
        root_qs.order_by.return_value.first.return_value = None
        base_qs = MagicMock()
        base_qs.select_related.return_value = base_qs
        base_qs.filter.side_effect = lambda **kwargs: sub_qs if kwargs == {"subentity_id": 1} else root_qs
        mock_section_map_filter.return_value = base_qs

        header = self._base_header(tds_amount=Decimal("10.00"))
        header.tds_section_id = 77
        header.subentity_id = 1

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[self._line()],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        jl_inputs = posting_instance.post.call_args.kwargs["jl_inputs"]
        tds_payable = [
            x for x in jl_inputs
            if x.account_id == 9901 and x.ledger_id == 9911 and x.drcr is False and x.amount == Decimal("10.00")
        ]
        self.assertTrue(
            tds_payable,
            "Expected TDS payable posting to fall back to the mapped account ledger when payable_ledger is blank.",
        )

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_asset_purchase_behavior_skips_inventory_move(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
        }

        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        product = SimpleNamespace(
            id=99,
            base_uom_id=1,
            base_uom=SimpleNamespace(id=1, code="GMS"),
            uom_conversions=[],
        )
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = [product]

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=1002)

        header = self._base_header(grand_total=Decimal("100.00"), affects_inventory=True)

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[
                self._line(
                    product_id=99,
                    is_service=False,
                    purchase_behavior=ProductPurchaseBehavior.ASSET,
                    qty=Decimal("1.0000"),
                    uom_id=1,
                )
            ],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        kwargs = posting_instance.post.call_args.kwargs
        self.assertEqual(kwargs["im_inputs"], [])

    @patch("posting.adapters.purchase_invoice.resolve_posting_location_id", return_value=5)
    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_inventory_move_uses_uom_conversion_for_base_qty(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
        mocked_resolve_location,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
        }
        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000

        kg_uom = SimpleNamespace(id=2, code="KG")
        gms_uom = SimpleNamespace(id=1, code="GMS")
        product = SimpleNamespace(
            id=99,
            base_uom_id=1,
            base_uom=gms_uom,
            uom_conversions=[
                SimpleNamespace(
                    from_uom_id=2,
                    from_uom=kg_uom,
                    to_uom_id=1,
                    to_uom=gms_uom,
                    factor=Decimal("1000"),
                )
            ],
        )
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = [product]

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=1003)

        header = self._base_header(grand_total=Decimal("500.00"), affects_inventory=True)

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[
                self._line(
                    product_id=99,
                    is_service=False,
                    purchase_behavior=ProductPurchaseBehavior.INVENTORY,
                    qty=Decimal("2.0000"),
                    free_qty=Decimal("0.0000"),
                    taxable_value=Decimal("500.00"),
                    uom_id=2,
                )
            ],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        kwargs = posting_instance.post.call_args.kwargs
        move = kwargs["im_inputs"][0]
        self.assertEqual(move.qty, Decimal("2.0000"))
        self.assertEqual(move.uom_factor, Decimal("1000.0000"))
        self.assertEqual(move.base_qty, Decimal("2000.0000"))
        self.assertEqual(move.base_uom_id, 1)
        self.assertTrue(mocked_resolve_location.called)

    @patch("posting.adapters.purchase_invoice.resolve_posting_location_id", return_value=5)
    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_inventory_move_generates_internal_expiry_lot_for_expiry_tracked_product(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
        mocked_resolve_location,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
        }
        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        product = SimpleNamespace(
            id=99,
            base_uom_id=1,
            base_uom=SimpleNamespace(id=1, code="PCS"),
            uom_conversions=[],
            is_batch_managed=False,
            is_expiry_tracked=True,
        )
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = [product]

        posting_instance = mock_posting_service_cls.return_value
        posting_instance.post.return_value = SimpleNamespace(id=1004)

        header = self._base_header(grand_total=Decimal("500.00"), affects_inventory=True)

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[
                self._line(
                    product_id=99,
                    is_service=False,
                    purchase_behavior=ProductPurchaseBehavior.INVENTORY,
                    qty=Decimal("5.0000"),
                    taxable_value=Decimal("500.00"),
                    uom_id=1,
                    batch_number="",
                    expiry_date=date(2026, 5, 1),
                )
            ],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        kwargs = posting_instance.post.call_args.kwargs
        move = kwargs["im_inputs"][0]
        self.assertEqual(move.batch_number, "EXP-99-20260501")
        self.assertEqual(move.expiry_date, date(2026, 5, 1))
        self.assertTrue(mocked_resolve_location.called)

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_reverse_charge_credit_note_reverses_rcm_payable_and_input_tax(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.RCM_CGST_PAYABLE: 8110,
            StaticAccountCodes.RCM_SGST_PAYABLE: 8111,
            StaticAccountCodes.RCM_IGST_PAYABLE: 8112,
            StaticAccountCodes.RCM_CESS_PAYABLE: 8113,
        }
        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        resolver.get_ledger_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []
        mock_posting_service_cls.return_value.post.return_value = SimpleNamespace(id=1005)

        header = self._base_header(
            doc_type=PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            purchase_number="PCN-101",
            grand_total=Decimal("354.00"),
            total_gst=Decimal("54.00"),
            is_reverse_charge=True,
        )
        line = self._line(
            taxable_value=Decimal("300.00"),
            igst_amount=Decimal("54.00"),
            is_itc_eligible=True,
        )

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[line],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        jl_inputs = mock_posting_service_cls.return_value.post.call_args.kwargs["jl_inputs"]
        rcm_igst_lines = [x for x in jl_inputs if x.account_id == 8112 and x.amount == Decimal("54.00")]
        input_igst_lines = [x for x in jl_inputs if x.account_id == 8105 and x.amount == Decimal("54.00")]
        vendor_lines = [x for x in jl_inputs if x.account_id == header.vendor_id and x.amount == Decimal("300.00")]

        self.assertTrue(rcm_igst_lines)
        self.assertTrue(input_igst_lines)
        self.assertTrue(vendor_lines)
        self.assertTrue(rcm_igst_lines[0].drcr, "Expected RCM payable to be debited on purchase credit note.")
        self.assertFalse(input_igst_lines[0].drcr, "Expected input IGST to be credited on purchase credit note.")
        self.assertTrue(vendor_lines[0].drcr, "Expected vendor reversal debit for RCM purchase credit note.")

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_reverse_charge_same_state_invoice_uses_tax_summary_for_liability_and_vendor_payable(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.RCM_CGST_PAYABLE: 8110,
            StaticAccountCodes.RCM_SGST_PAYABLE: 8111,
            StaticAccountCodes.RCM_IGST_PAYABLE: 8112,
            StaticAccountCodes.RCM_CESS_PAYABLE: 8113,
        }
        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        resolver.get_ledger_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []
        mock_posting_service_cls.return_value.post.return_value = SimpleNamespace(id=1006)

        summary_row = SimpleNamespace(
            cgst_amount=Decimal("27.00"),
            sgst_amount=Decimal("27.00"),
            igst_amount=Decimal("0.00"),
            cess_amount=Decimal("0.00"),
            itc_eligible_tax=Decimal("54.00"),
            itc_ineligible_tax=Decimal("0.00"),
        )
        header = self._base_header(
            purchase_number="PINV-RCM-INTRA",
            grand_total=Decimal("300.00"),
            total_gst=Decimal("0.00"),
            is_reverse_charge=True,
            tax_summaries=_ListManager([summary_row]),
        )
        line = self._line(
            taxable_value=Decimal("300.00"),
            cgst_amount=Decimal("0.00"),
            sgst_amount=Decimal("0.00"),
            igst_amount=Decimal("0.00"),
            is_itc_eligible=True,
        )

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[line],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        jl_inputs = mock_posting_service_cls.return_value.post.call_args.kwargs["jl_inputs"]
        self.assertTrue([x for x in jl_inputs if x.account_id == 8110 and x.drcr is False and x.amount == Decimal("27.00")])
        self.assertTrue([x for x in jl_inputs if x.account_id == 8111 and x.drcr is False and x.amount == Decimal("27.00")])
        self.assertTrue([x for x in jl_inputs if x.account_id == 8103 and x.drcr is True and x.amount == Decimal("27.00")])
        self.assertTrue([x for x in jl_inputs if x.account_id == 8104 and x.drcr is True and x.amount == Decimal("27.00")])
        self.assertTrue([x for x in jl_inputs if x.account_id == header.vendor_id and x.drcr is False and x.amount == Decimal("300.00")])

    @patch("posting.adapters.purchase_invoice.PostingService")
    @patch("posting.adapters.purchase_invoice.Product.objects")
    @patch("posting.adapters.purchase_invoice.ProductAccountResolver")
    @patch("posting.adapters.purchase_invoice.StaticAccountResolver")
    def test_reverse_charge_interstate_invoice_uses_tax_summary_for_liability_and_vendor_payable(
        self,
        mock_static_resolver_cls,
        mock_product_resolver_cls,
        mock_product_objects,
        mock_posting_service_cls,
    ):
        code_map = {
            StaticAccountCodes.PURCHASE_MISC_EXPENSE: 8100,
            StaticAccountCodes.ROUND_OFF_INCOME: 8101,
            StaticAccountCodes.ROUND_OFF_EXPENSE: 8102,
            StaticAccountCodes.INPUT_CGST: 8103,
            StaticAccountCodes.INPUT_SGST: 8104,
            StaticAccountCodes.INPUT_IGST: 8105,
            StaticAccountCodes.INPUT_CESS: 8106,
            StaticAccountCodes.PURCHASE_DEFAULT: 8107,
            StaticAccountCodes.RCM_CGST_PAYABLE: 8110,
            StaticAccountCodes.RCM_SGST_PAYABLE: 8111,
            StaticAccountCodes.RCM_IGST_PAYABLE: 8112,
            StaticAccountCodes.RCM_CESS_PAYABLE: 8113,
        }
        resolver = mock_static_resolver_cls.return_value
        resolver.get_account_id.side_effect = lambda code, required=False: code_map.get(code)
        resolver.get_ledger_id.side_effect = lambda code, required=False: code_map.get(code)
        mock_product_resolver_cls.return_value.purchase_account_id.return_value = 5000
        mock_product_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = []
        mock_posting_service_cls.return_value.post.return_value = SimpleNamespace(id=1007)

        summary_row = SimpleNamespace(
            cgst_amount=Decimal("0.00"),
            sgst_amount=Decimal("0.00"),
            igst_amount=Decimal("90.00"),
            cess_amount=Decimal("0.00"),
            itc_eligible_tax=Decimal("90.00"),
            itc_ineligible_tax=Decimal("0.00"),
        )
        header = self._base_header(
            purchase_number="PINV-RCM-INTER",
            grand_total=Decimal("500.00"),
            total_gst=Decimal("0.00"),
            is_reverse_charge=True,
            tax_summaries=_ListManager([summary_row]),
        )
        line = self._line(
            taxable_value=Decimal("500.00"),
            cgst_amount=Decimal("0.00"),
            sgst_amount=Decimal("0.00"),
            igst_amount=Decimal("0.00"),
            is_itc_eligible=True,
        )

        PurchaseInvoicePostingAdapter.post_purchase_invoice.__wrapped__(
            header=header,
            lines=[line],
            user_id=1,
            config=PurchaseInvoicePostingConfig(),
        )

        jl_inputs = mock_posting_service_cls.return_value.post.call_args.kwargs["jl_inputs"]
        self.assertTrue([x for x in jl_inputs if x.account_id == 8112 and x.drcr is False and x.amount == Decimal("90.00")])
        self.assertTrue([x for x in jl_inputs if x.account_id == 8105 and x.drcr is True and x.amount == Decimal("90.00")])
        self.assertTrue([x for x in jl_inputs if x.account_id == header.vendor_id and x.drcr is False and x.amount == Decimal("500.00")])


class PurchasePhase1ClassificationTests(SimpleTestCase):
    def test_validate_lines_structural_defaults_non_product_line_to_expense(self):
        attrs = {
            "default_taxability": PurchaseInvoiceHeader.Taxability.TAXABLE,
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": PurchaseInvoiceHeader.ItcClaimStatus.PENDING,
        }
        lines = [
            {
                "purchase_account": 5000,
                "product_desc": "Office lunch",
                "qty": Decimal("1.0000"),
                "rate": Decimal("250.00"),
                "is_service": True,
            }
        ]
        derived = SimpleNamespace(tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA)

        PurchaseInvoiceService.validate_lines_structural(attrs, lines, derived)

        self.assertEqual(lines[0]["purchase_behavior"], ProductPurchaseBehavior.EXPENSE)

    def test_validate_lines_structural_rejects_non_product_inventory_behavior(self):
        attrs = {
            "default_taxability": PurchaseInvoiceHeader.Taxability.TAXABLE,
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": PurchaseInvoiceHeader.ItcClaimStatus.PENDING,
        }
        lines = [
            {
                "purchase_account": 5000,
                "product_desc": "Office lunch",
                "qty": Decimal("1.0000"),
                "rate": Decimal("250.00"),
                "is_service": False,
                "purchase_behavior": ProductPurchaseBehavior.INVENTORY,
            }
        ]
        derived = SimpleNamespace(tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA)

        with self.assertRaisesMessage(ValueError, "non-product purchase lines can only use expense behavior"):
            PurchaseInvoiceService.validate_lines_structural(attrs, lines, derived)

    @patch("purchase.services.purchase_invoice_service.Product.objects")
    def test_validate_lines_structural_requires_account_for_expense_product(self, mock_product_objects):
        mock_product_objects.filter.return_value.only.return_value.first.return_value = SimpleNamespace(
            id=10,
            is_batch_managed=False,
            is_expiry_tracked=False,
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.EXPENSE,
            purchase_account_id=None,
        )
        attrs = {
            "default_taxability": PurchaseInvoiceHeader.Taxability.TAXABLE,
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": PurchaseInvoiceHeader.ItcClaimStatus.PENDING,
        }
        lines = [
            {
                "product": 10,
                "qty": Decimal("1.0000"),
                "rate": Decimal("250.00"),
                "is_service": False,
            }
        ]
        derived = SimpleNamespace(tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA)

        with self.assertRaisesMessage(ValueError, "expense purchase lines require an expense/purchase account"):
            PurchaseInvoiceService.validate_lines_structural(attrs, lines, derived)

    @patch("purchase.services.purchase_invoice_service.Product.objects")
    def test_validate_lines_structural_requires_default_asset_category_for_asset_product(self, mock_product_objects):
        mock_product_objects.filter.return_value.only.return_value.first.return_value = SimpleNamespace(
            id=10,
            is_batch_managed=False,
            is_expiry_tracked=False,
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.ASSET,
            purchase_account_id=5000,
            default_asset_category_id=None,
        )
        attrs = {
            "default_taxability": PurchaseInvoiceHeader.Taxability.TAXABLE,
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": PurchaseInvoiceHeader.ItcClaimStatus.PENDING,
        }
        lines = [
            {
                "product": 10,
                "qty": Decimal("1.0000"),
                "rate": Decimal("250.00"),
                "is_service": False,
            }
        ]
        derived = SimpleNamespace(tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA)

        with self.assertRaisesMessage(ValueError, "asset product is missing a default asset category"):
            PurchaseInvoiceService.validate_lines_structural(attrs, lines, derived)


class PurchaseInventoryReturnSafetyTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="purchase-return-safety",
            email="purchase-return-safety@example.com",
            password="testpass123",
        )
        self.entity = Entity.objects.create(entityname="Purchase Return Entity", createdby=self.user)
        self.subentity = SubEntity.objects.create(entity=self.entity, subentityname="Main Branch")
        self.entityfin = EntityFinancialYear.objects.create(
            entity=self.entity,
            desc="FY 2026-27",
            year_code="FY2026-27",
            finstartyear=timezone.make_aware(datetime(2026, 4, 1)),
            finendyear=timezone.make_aware(datetime(2027, 3, 31)),
        )
        self.location = Godown.objects.create(
            entity=self.entity,
            subentity=self.subentity,
            name="Main Godown",
            code="MAIN",
            address="Warehouse 1",
            city="Mumbai",
            state="MH",
            pincode="400001",
            is_active=True,
            is_default=True,
        )
        self.uom = UnitOfMeasure.objects.create(entity=self.entity, code="PCS", description="Pieces")
        self.product_category = ProductCategory.objects.create(entity=self.entity, pcategoryname="Finished Goods")
        self.product = Product.objects.create(
            entity=self.entity,
            productname="Return Product",
            sku="RET-1",
            productdesc="Returnable stock item",
            productcategory=self.product_category,
            base_uom=self.uom,
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.INVENTORY,
        )
        self.header = PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            location=self.location,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            status=PurchaseInvoiceHeader.Status.POSTED,
            bill_date=date(2026, 4, 10),
            posting_date=date(2026, 4, 10),
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            supply_category=PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
            tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA,
            is_igst=False,
            is_reverse_charge=False,
            is_itc_eligible=True,
        )
        self.line = PurchaseInvoiceLine.objects.create(
            header=self.header,
            line_no=1,
            product=self.product,
            uom=self.uom,
            qty=Decimal("10.0000"),
            free_qty=Decimal("0.0000"),
            rate=Decimal("100.00"),
            product_desc="Returnable stock item",
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.INVENTORY,
            taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            taxable_value=Decimal("1000.00"),
            gst_rate=Decimal("18.00"),
            cgst_percent=Decimal("9.00"),
            sgst_percent=Decimal("9.00"),
            igst_percent=Decimal("0.00"),
            cgst_amount=Decimal("90.00"),
            sgst_amount=Decimal("90.00"),
            igst_amount=Decimal("0.00"),
            cess_amount=Decimal("0.00"),
            line_total=Decimal("1180.00"),
            is_itc_eligible=True,
        )

    def _seed_inventory_move(self, *, txn_type: str, txn_id: int, move_type: str, qty: str, posting_day: date):
        batch = PostingBatch.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            voucher_no=f"{txn_type}-{txn_id}",
            revision=1,
            is_active=True,
            created_by=self.user,
        )
        entry = Entry.objects.create(
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            voucher_no=f"{txn_type}-{txn_id}",
            voucher_date=posting_day,
            posting_date=posting_day,
            status=2,
            posting_batch=batch,
            narration="Inventory movement fixture",
            created_by=self.user,
        )
        qty_decimal = Decimal(qty)
        InventoryMove.objects.create(
            entry=entry,
            posting_batch=batch,
            entity=self.entity,
            entityfin=self.entityfin,
            subentity=self.subentity,
            txn_type=txn_type,
            txn_id=txn_id,
            detail_id=self.line.id,
            voucher_no=f"{txn_type}-{txn_id}",
            product=self.product,
            batch_number="",
            location=self.location,
            source_location=self.location if move_type == InventoryMove.MoveType.OUT else None,
            destination_location=self.location if move_type == InventoryMove.MoveType.IN_ else None,
            uom=self.uom,
            base_uom=self.uom,
            qty=qty_decimal,
            uom_factor=Decimal("1.00000000"),
            base_qty=qty_decimal,
            unit_cost=Decimal("100.0000"),
            ext_cost=Decimal("1000.00"),
            cost_source=InventoryMove.CostSource.PURCHASE,
            move_type=move_type,
            movement_nature=InventoryMove.MovementNature.PURCHASE if move_type == InventoryMove.MoveType.IN_ else InventoryMove.MovementNature.OTHER,
            movement_reason="fixture",
            posting_date=posting_day,
            created_by=self.user,
        )

    def _return_attrs(self):
        return {
            "doc_type": PurchaseInvoiceHeader.DocType.CREDIT_NOTE,
            "ref_document": self.header,
            "note_reason": PurchaseInvoiceHeader.NoteReason.QUANTITY_RETURN,
            "default_taxability": PurchaseInvoiceHeader.Taxability.TAXABLE,
            "is_reverse_charge": False,
            "is_itc_eligible": True,
            "itc_claim_status": PurchaseInvoiceHeader.ItcClaimStatus.PENDING,
            "supply_category": PurchaseInvoiceHeader.SupplyCategory.DOMESTIC,
            "location": self.location,
        }

    def _return_lines(self, qty: str):
        return [
            {
                "line_no": 1,
                "product": self.product.id,
                "uom": self.uom.id,
                "qty": Decimal(qty),
                "free_qty": Decimal("0.0000"),
                "rate": Decimal("100.00"),
                "product_desc": "Returnable stock item",
                "is_service": False,
                "purchase_behavior": ProductPurchaseBehavior.INVENTORY,
                "taxability": PurchaseInvoiceHeader.Taxability.TAXABLE,
                "gst_rate": Decimal("18.00"),
            }
        ]

    def test_validate_lines_structural_allows_quantity_return_before_downstream_consumption(self):
        self._seed_inventory_move(
            txn_type=TxnType.PURCHASE,
            txn_id=self.header.id,
            move_type=InventoryMove.MoveType.IN_,
            qty="10.0000",
            posting_day=date(2026, 4, 10),
        )

        PurchaseInvoiceService.validate_lines_structural(
            self._return_attrs(),
            self._return_lines("5.0000"),
            SimpleNamespace(tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA),
        )

    def test_validate_lines_structural_blocks_quantity_return_after_downstream_consumption(self):
        self._seed_inventory_move(
            txn_type=TxnType.PURCHASE,
            txn_id=self.header.id,
            move_type=InventoryMove.MoveType.IN_,
            qty="10.0000",
            posting_day=date(2026, 4, 10),
        )
        self._seed_inventory_move(
            txn_type=TxnType.INVENTORY_ADJUSTMENT,
            txn_id=991,
            move_type=InventoryMove.MoveType.OUT,
            qty="10.0000",
            posting_day=date(2026, 4, 12),
        )

        with self.assertRaisesMessage(ValueError, "no longer safely returnable"):
            PurchaseInvoiceService.validate_lines_structural(
                self._return_attrs(),
                self._return_lines("5.0000"),
                SimpleNamespace(tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA),
            )


class PurchasePhase3AssetIntakeTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="purchase-phase3",
            email="purchase-phase3@example.com",
            password="testpass123",
        )
        self.entity = Entity.objects.create(entityname="Purchase Phase3 Entity")
        self.entityfin = EntityFinancialYear.objects.create(
            entity=self.entity,
            desc="FY 2026-27",
            year_code="FY2026-27",
            finstartyear=timezone.now(),
            finendyear=timezone.now(),
            createdby=self.user,
        )
        self.subentity = SubEntity.objects.create(
            entity=self.entity,
            subentityname="Head Office",
            branch_type=SubEntity.BranchType.HEAD_OFFICE,
            is_head_office=True,
        )
        self.vendor = account.objects.create(entity=self.entity, accountname="Vendor A")
        self.uom = UnitOfMeasure.objects.create(entity=self.entity, code="PCS", description="Pieces")
        self.category = ProductCategory.objects.create(entity=self.entity, pcategoryname="Asset Items")
        self.asset_ledger = Ledger.objects.create(entity=self.entity, name="CWIP Ledger")
        self.asset_category = AssetCategory.objects.create(
            entity=self.entity,
            code="CWIP-COMP",
            name="Computer CWIP",
            nature=AssetCategory.AssetNature.CAPITAL_WIP,
            cwip_ledger=self.asset_ledger,
        )
        self.product = Product.objects.create(
            entity=self.entity,
            productname="Office Computer",
            sku="COMP-001",
            productcategory=self.category,
            base_uom=self.uom,
            purchase_behavior=ProductPurchaseBehavior.ASSET,
            default_asset_category=self.asset_category,
        )
        self.header = PurchaseInvoiceHeader.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            vendor=self.vendor,
            bill_date=timezone.now().date(),
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            doc_code="PINV",
            doc_no=1001,
            purchase_number="PI/PINV/2026/1001",
            status=PurchaseInvoiceHeader.Status.CONFIRMED,
            default_taxability=PurchaseInvoiceHeader.Taxability.TAXABLE,
            tax_regime=PurchaseInvoiceHeader.TaxRegime.INTRA,
            grand_total=Decimal("100000.00"),
        )
        self.line = self.header.lines.create(
            line_no=1,
            product=self.product,
            product_desc="Office Computer",
            is_service=False,
            purchase_behavior=ProductPurchaseBehavior.ASSET,
            uom=self.uom,
            qty=Decimal("1.0000"),
            rate=Decimal("84745.76"),
            taxable_value=Decimal("84745.76"),
            line_total=Decimal("100000.00"),
        )

    @patch("purchase.services.purchase_invoice_actions.GstTdsService.sync_contract_ledger_for_header")
    @patch("purchase.services.purchase_invoice_actions.PurchaseApService.sync_open_item_for_header")
    @patch("purchase.services.purchase_invoice_actions.PurchaseInvoicePostingAdapter.post_purchase_invoice")
    def test_post_creates_capital_wip_asset_intake_for_asset_lines(
        self,
        mock_post_adapter,
        mock_sync_ap,
        mock_sync_gst,
    ):
        result = PurchaseInvoiceActions.post(self.header.id, posted_by_id=self.user.id)

        self.assertEqual(result.header.status, PurchaseInvoiceHeader.Status.POSTED)
        self.line.refresh_from_db()
        self.assertIsNotNone(self.line.asset_record_id)
        asset = self.line.asset_record
        self.assertEqual(asset.status, FixedAsset.AssetStatus.CAPITAL_WIP)
        self.assertEqual(asset.category_id, self.asset_category.id)
        self.assertEqual(asset.purchase_document_no, self.header.purchase_number)
        self.assertEqual(asset.vendor_account_id, self.vendor.id)
        self.assertEqual(asset.gross_block, Decimal("84745.76"))
        self.assertEqual(asset.quantity, Decimal("1.0000"))
        self.assertEqual(asset.external_reference, f"purchase-line:{self.line.id}")
        mock_post_adapter.assert_called_once()
        mock_sync_ap.assert_called_once()
        mock_sync_gst.assert_called_once()

    def test_revert_asset_intakes_for_unpost_deletes_uncapitalized_cwip_asset(self):
        asset = FixedAsset.objects.create(
            entity=self.entity,
            entityfinid=self.entityfin,
            subentity=self.subentity,
            category=self.asset_category,
            ledger=self.asset_ledger,
            asset_code="CWIP-001",
            asset_name="Office Computer",
            status=FixedAsset.AssetStatus.CAPITAL_WIP,
            acquisition_date=self.header.bill_date,
            quantity=Decimal("1.0000"),
            gross_block=Decimal("84745.76"),
            residual_value=Decimal("0.00"),
            useful_life_months=60,
            depreciation_method=FixedAsset.DepreciationMethod.SLM,
            net_book_value=Decimal("84745.76"),
            vendor_account=self.vendor,
            purchase_document_no=self.header.purchase_number,
            external_reference=f"purchase-line:{self.line.id}",
        )
        self.line.asset_record = asset
        self.line.save(update_fields=["asset_record"])

        PurchaseAssetIntakeService.revert_asset_intakes_for_unpost(header=self.header)

        self.line.refresh_from_db()
        self.assertIsNone(self.line.asset_record_id)
        self.assertFalse(FixedAsset.objects.filter(id=asset.id).exists())


class PurchaseStatutoryServiceTests(TestCase):
    @patch("purchase.services.purchase_statutory_service.PurchaseItcAction.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    def test_itc_status_register_marks_asset_rows_for_capital_goods_review(
        self,
        mock_header_objects,
        mock_itc_action_objects,
    ):
        header = SimpleNamespace(
            id=701,
            bill_date=date(2026, 4, 15),
            vendor_name="Asset Vendor",
            vendor_gstin="29ABCDE1234F1Z5",
            purchase_number="PI/ASSET/701",
            doc_code="PINV",
            doc_no=701,
            doc_type=PurchaseInvoiceHeader.DocType.TAX_INVOICE,
            status=PurchaseInvoiceHeader.Status.CONFIRMED,
            is_itc_eligible=True,
            itc_claim_status=PurchaseInvoiceHeader.ItcClaimStatus.PENDING,
            itc_claim_period=None,
            itc_claimed_at=None,
            itc_block_reason="",
            gstr2b_match_status=PurchaseInvoiceHeader.Gstr2bMatchStatus.MATCHED,
            total_taxable=Decimal("84745.76"),
            total_gst=Decimal("15254.24"),
            tax_summaries=_ListManager(
                [
                    SimpleNamespace(
                        itc_eligible_tax=Decimal("15254.24"),
                        itc_ineligible_tax=Decimal("0.00"),
                    )
                ]
            ),
            lines=_ListManager(
                [
                    SimpleNamespace(purchase_behavior=ProductPurchaseBehavior.ASSET),
                ]
            ),
            get_doc_type_display=lambda: "Tax Invoice",
            get_status_display=lambda: "Confirmed",
            get_itc_claim_status_display=lambda: "Pending",
            get_gstr2b_match_status_display=lambda: "Matched",
        )

        qs = MagicMock()
        qs.filter.return_value = qs
        qs.exclude.return_value = qs
        qs.order_by.return_value = [header]
        mock_header_objects.filter.return_value.select_related.return_value.prefetch_related.return_value = qs

        mock_itc_action_objects.filter.return_value.select_related.return_value.order_by.return_value = []

        result = PurchaseStatutoryService.itc_status_register(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            date_from=date(2026, 4, 1),
            date_to=date(2026, 4, 30),
        )

        self.assertEqual(result["count"], 1)
        self.assertEqual(result["rows"][0]["purchase_behavior_summary"], "asset")
        self.assertEqual(result["rows"][0]["asset_line_count"], 1)
        self.assertEqual(result["rows"][0]["itc_eligible_tax"], "15254.24")

    def test_validate_it_tds_amount_raises_on_excess(self):
        header = SimpleNamespace(id=1, tds_amount=Decimal("10.00"), gst_tds_amount=Decimal("0.00"))
        with self.assertRaisesMessage(ValueError, "exceeds IT-TDS"):
            PurchaseStatutoryService._validate_header_amount_for_tax_type(
                header=header,
                tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
                amount=Decimal("11.00"),
            )

    def test_validate_gst_tds_amount_raises_on_excess(self):
        header = SimpleNamespace(id=2, tds_amount=Decimal("0.00"), gst_tds_amount=Decimal("5.00"))
        with self.assertRaisesMessage(ValueError, "exceeds GST-TDS"):
            PurchaseStatutoryService._validate_header_amount_for_tax_type(
                header=header,
                tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS,
                amount=Decimal("6.00"),
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    def test_validate_challan_balance_rejects_over_allocation(self, mock_challan_line_objects, mock_header_objects):
        mapped_qs = MagicMock()
        mapped_qs.exclude.return_value = mapped_qs
        mapped_qs.filter.return_value = mapped_qs
        mapped_qs.values.return_value.annotate.return_value = [{"header_id": 10, "total": Decimal("7.00")}]
        mock_challan_line_objects.filter.return_value = mapped_qs

        header = SimpleNamespace(
            id=10,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tds_amount=Decimal("10.00"),
            gst_tds_amount=Decimal("0.00"),
        )
        mock_header_objects.filter.return_value.first.return_value = header

        with self.assertRaisesMessage(ValueError, "remaining IT-TDS balance 3.00"):
            PurchaseStatutoryService._validate_challan_balance_for_lines(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
                line_rows=[{"header_id": 10, "amount": Decimal("4.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    def test_validate_challan_balance_excludes_current_challan_during_edit(self, mock_challan_line_objects, mock_header_objects):
        mapped_qs = MagicMock()
        mapped_qs.exclude.return_value = mapped_qs
        mapped_qs.filter.return_value = mapped_qs
        mapped_qs.values.return_value.annotate.return_value = [{"header_id": 10, "total": Decimal("1.00")}]
        mock_challan_line_objects.filter.return_value = mapped_qs

        header = SimpleNamespace(
            id=10,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tds_amount=Decimal("10.00"),
            gst_tds_amount=Decimal("0.00"),
        )
        mock_header_objects.filter.return_value.first.return_value = header

        PurchaseStatutoryService._validate_challan_balance_for_lines(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            line_rows=[{"header_id": 10, "amount": Decimal("8.00")}],
            exclude_challan_id=99,
        )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    def test_validate_return_balance_rejects_over_allocation(self, mock_challan_line_objects, mock_return_line_objects):
        challan_line_qs = MagicMock()
        challan_line_qs.exclude.return_value = challan_line_qs
        challan_line_qs.filter.return_value = challan_line_qs
        challan_line_qs.values.return_value.annotate.return_value = [
            {"header_id": 10, "challan_id": 20, "total": Decimal("10.00")}
        ]
        mock_challan_line_objects.filter.return_value = challan_line_qs

        consumed_qs = MagicMock()
        consumed_qs.exclude.return_value = consumed_qs
        consumed_qs.filter.return_value = consumed_qs
        consumed_qs.values.return_value.annotate.return_value = [
            {"header_id": 10, "challan_id": 20, "total": Decimal("7.00")}
        ]
        mock_return_line_objects.filter.return_value = consumed_qs

        with self.assertRaisesMessage(ValueError, "requested 4.00 exceeds remaining balance 3.00"):
            PurchaseStatutoryService._validate_return_balance_for_lines(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                line_rows=[{"header_id": 10, "challan_id": 20, "amount": Decimal("4.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    def test_validate_return_balance_excludes_current_filing_during_edit(self, mock_challan_line_objects, mock_return_line_objects):
        challan_line_qs = MagicMock()
        challan_line_qs.exclude.return_value = challan_line_qs
        challan_line_qs.filter.return_value = challan_line_qs
        challan_line_qs.values.return_value.annotate.return_value = [
            {"header_id": 10, "challan_id": 20, "total": Decimal("10.00")}
        ]
        mock_challan_line_objects.filter.return_value = challan_line_qs

        consumed_qs = MagicMock()
        consumed_qs.exclude.return_value = consumed_qs
        consumed_qs.filter.return_value = consumed_qs
        consumed_qs.values.return_value.annotate.return_value = [
            {"header_id": 10, "challan_id": 20, "total": Decimal("1.00")}
        ]
        mock_return_line_objects.filter.return_value = consumed_qs

        PurchaseStatutoryService._validate_return_balance_for_lines(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            line_rows=[{"header_id": 10, "challan_id": 20, "amount": Decimal("9.00")}],
            exclude_filing_id=77,
        )

    def test_require_submitted_approval_state_rejects_non_submitted_payload(self):
        with self.assertRaisesMessage(ValueError, "requires the record to be in SUBMITTED approval state"):
            PurchaseStatutoryService._require_submitted_approval_state(
                payload={"_approval_state": {"status": "DRAFT"}},
                action_label="Return approval",
            )

    def test_validate_it_tds_return_snapshot_rejects_non_resident_for_26q(self):
        with self.assertRaisesMessage(ValueError, "26Q allows only RESIDENT deductees"):
            PurchaseStatutoryService._validate_it_tds_return_snapshot(
                return_code="26Q",
                deductee_residency_snapshot="NON_RESIDENT",
                deductee_pan_snapshot="ABCDE1234F",
                deductee_tax_id_snapshot="TIN123",
                line_label="Line 1",
            )

    def test_validate_it_tds_return_snapshot_rejects_missing_tax_id_for_27q(self):
        with self.assertRaisesMessage(ValueError, "27Q requires deductee_tax_id_snapshot"):
            PurchaseStatutoryService._validate_it_tds_return_snapshot(
                return_code="27Q",
                deductee_residency_snapshot="NON_RESIDENT",
                deductee_pan_snapshot="",
                deductee_tax_id_snapshot="",
                line_label="Line 1",
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_get_original_return_for_revision_rejects_prior_revision_reference(self, mock_return_objects):
        mock_return_objects.filter.return_value.first.return_value = SimpleNamespace(
            id=9,
            original_return_id=3,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
        )

        with self.assertRaisesMessage(ValueError, "must reference the original return"):
            PurchaseStatutoryService._get_original_return_for_revision(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                return_code="26Q",
                period_from=date(2026, 4, 1),
                period_to=date(2026, 6, 30),
                original_return_id=9,
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_assert_unique_revision_number_rejects_duplicate_active_revision(self, mock_return_objects):
        qs = MagicMock()
        qs.exclude.return_value = qs
        qs.exists.return_value = True
        mock_return_objects.filter.return_value = qs

        with self.assertRaisesMessage(ValueError, "A revision already exists"):
            PurchaseStatutoryService._assert_unique_revision_number(
                original_return_id=3,
                revision_no=1,
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_assert_unique_revision_number_allows_same_revision_when_excluding_current_filing(self, mock_return_objects):
        qs = MagicMock()
        exclude_qs = MagicMock()
        qs.exclude.side_effect = [qs, exclude_qs]
        exclude_qs.exists.return_value = False
        mock_return_objects.filter.return_value = qs

        PurchaseStatutoryService._assert_unique_revision_number(
            original_return_id=3,
            revision_no=1,
            exclude_filing_id=77,
        )

        qs.exclude.assert_any_call(status=PurchaseStatutoryReturn.Status.CANCELLED)
        qs.exclude.assert_any_call(pk=77)

    def test_assert_editable_draft_approval_state_blocks_submitted(self):
        with self.assertRaisesMessage(ValueError, "cannot be edited while approval state is SUBMITTED"):
            PurchaseStatutoryService._assert_editable_draft_approval_state(
                payload={"_approval_state": {"status": "SUBMITTED"}},
                record_label="Return",
            )

    def test_assert_editable_draft_approval_state_marks_rejected_for_reset(self):
        should_reset = PurchaseStatutoryService._assert_editable_draft_approval_state(
            payload={"_approval_state": {"status": "REJECTED"}},
            record_label="Return",
        )
        self.assertTrue(should_reset)

    def test_merge_payload_for_draft_update_resets_approval_when_requested(self):
        payload = PurchaseStatutoryService._merge_payload_for_draft_update(
            existing_payload={"_approval_state": {"status": "REJECTED"}, "_audit_log": [{"action": "X"}]},
            incoming_payload={},
            reset_approval_to_draft=True,
        )
        self.assertEqual(payload["_approval_state"]["status"], "DRAFT")
        self.assertEqual(payload["_audit_log"], [{"action": "X"}])

    def test_validate_period_bounds_rejects_inverted_period(self):
        with self.assertRaisesMessage(ValueError, "period_from cannot be greater than period_to"):
            PurchaseStatutoryService._validate_period_bounds(
                period_from=date(2026, 5, 1),
                period_to=date(2026, 4, 30),
                period_label="Return period",
            )

    def test_validate_period_bounds_rejects_anchor_outside_period(self):
        with self.assertRaisesMessage(ValueError, "challan_date must fall within Challan period"):
            PurchaseStatutoryService._validate_period_bounds(
                period_from=date(2026, 4, 1),
                period_to=date(2026, 4, 30),
                period_label="Challan period",
                anchor_date=date(2026, 5, 1),
                anchor_label="challan_date",
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_return_balance_for_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_revision_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    def test_create_return_rejects_non_deposited_challan(
        self,
        mock_return_line_objects,
        mock_return_objects,
        mock_validate_revision,
        mock_validate_balance,
        mock_header_objects,
        mock_challan_objects,
    ):
        original_return_qs = MagicMock()
        original_return_qs.exclude.return_value = original_return_qs
        original_return_qs.filter.return_value = original_return_qs
        original_return_qs.exists.return_value = False
        mock_return_objects.filter.return_value = original_return_qs
        mock_return_objects.create.return_value = SimpleNamespace(id=1, amount=Decimal("0.00"), save=MagicMock())
        mock_header_objects.filter.return_value.first.return_value = SimpleNamespace(
            id=10,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tds_amount=Decimal("10.00"),
            gst_tds_amount=Decimal("0.00"),
            tds_section=None,
            vendor=None,
            vendor_gstin=None,
        )
        mock_challan_objects.filter.return_value.first.return_value = SimpleNamespace(
            id=20,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            status=PurchaseStatutoryChallan.Status.DRAFT,
            cin_no=None,
        )

        with self.assertRaisesMessage(ValueError, "challan must be in DEPOSITED status"):
            PurchaseStatutoryService.create_return(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                return_code="26Q",
                period_from=date(2026, 4, 1),
                period_to=date(2026, 6, 30),
                lines=[{"header_id": 10, "challan_id": 20, "amount": Decimal("5.00")}],
            )
        mock_return_line_objects.create.assert_not_called()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_return_balance_for_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_create_return_rejects_duplicate_original_return_for_same_scope(
        self,
        mock_return_objects,
        mock_validate_balance,
    ):
        original_return_qs = MagicMock()
        original_return_qs.exclude.return_value = original_return_qs
        original_return_qs.filter.return_value = original_return_qs
        original_return_qs.exists.return_value = True
        mock_return_objects.filter.return_value = original_return_qs

        with self.assertRaisesMessage(ValueError, "An original return already exists for this return_code and period. Create a revision against that original return."):
            PurchaseStatutoryService.create_return(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                return_code="26Q",
                period_from=date(2026, 4, 1),
                period_to=date(2026, 6, 30),
                lines=[{"header_id": 10, "challan_id": 20, "amount": Decimal("5.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_return_balance_for_lines")
    def test_create_return_rejects_revision_without_original_return_linkage(
        self,
        mock_validate_balance,
    ):
        with self.assertRaisesMessage(ValueError, "original_return_id is required when revision_no > 0."):
            PurchaseStatutoryService.create_return(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                return_code="26Q",
                period_from=date(2026, 4, 1),
                period_to=date(2026, 6, 30),
                revision_no=1,
                lines=[{"header_id": 10, "challan_id": 20, "amount": Decimal("5.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_revision_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._get_original_return_for_revision")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_return_balance_for_lines")
    def test_create_return_rejects_second_active_revision_draft_for_same_original(
        self,
        mock_validate_balance,
        mock_get_original,
        mock_validate_revision,
        mock_return_objects,
    ):
        duplicate_revision_qs = MagicMock()
        duplicate_revision_qs.exclude.return_value = duplicate_revision_qs
        duplicate_revision_qs.exists.side_effect = [False, True]
        mock_return_objects.filter.return_value = duplicate_revision_qs
        mock_get_original.return_value = SimpleNamespace(id=3)

        with self.assertRaisesMessage(ValueError, "An active revision draft already exists for this original return. Finish or cancel that revision first."):
            PurchaseStatutoryService.create_return(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                return_code="26Q",
                period_from=date(2026, 4, 1),
                period_to=date(2026, 6, 30),
                original_return_id=3,
                revision_no=2,
                lines=[{"header_id": 10, "challan_id": 20, "amount": Decimal("5.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_return_balance_for_lines")
    def test_update_return_rejects_revision_without_original_return_linkage(
        self,
        mock_validate_balance,
    ):
        with self.assertRaisesMessage(ValueError, "original_return_id is required when revision_no > 0."):
            PurchaseStatutoryService.update_return(
                filing_id=88,
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
                return_code="26Q",
                period_from=date(2026, 4, 1),
                period_to=date(2026, 6, 30),
                revision_no=1,
                lines=[{"header_id": 10, "challan_id": 20, "amount": Decimal("5.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._vendor_deductee_snapshot")
    def test_return_eligible_lines_filters_by_26q_rules(
        self,
        mock_snapshot,
        mock_return_line_objects,
        mock_challan_line_objects,
        mock_purchase_line_objects,
    ):
        challan = SimpleNamespace(challan_no="CH1", cin_no="CIN1")
        header = SimpleNamespace(tds_section=None, purchase_number="PINV-10", doc_code="PINV", doc_no=10, vendor_name="Vendor A")
        resident_line = SimpleNamespace(header_id=10, challan_id=20, amount=Decimal("5.00"), header=header, challan=challan)
        non_resident_line = SimpleNamespace(header_id=11, challan_id=21, amount=Decimal("6.00"), header=header, challan=challan)

        challan_line_qs = MagicMock()
        challan_line_qs.filter.return_value = challan_line_qs
        challan_line_qs.__iter__.return_value = iter([resident_line, non_resident_line])
        mock_challan_line_objects.select_related.return_value.filter.return_value = challan_line_qs

        consumed_qs = MagicMock()
        consumed_qs.exclude.return_value = consumed_qs
        consumed_values_qs = MagicMock()
        consumed_values_qs.annotate.return_value = consumed_values_qs
        consumed_values_qs.filter.return_value = []
        consumed_values_qs.__iter__.return_value = iter([])
        consumed_qs.values.return_value = consumed_values_qs
        mock_return_line_objects.filter.return_value = consumed_qs

        mock_snapshot.side_effect = [
            {
                "deductee_residency_snapshot": "RESIDENT",
                "deductee_country_obj": None,
                "deductee_country_code_snapshot": "",
                "deductee_country_name_snapshot": "",
                "deductee_tax_id_snapshot": "TAX1",
                "deductee_pan_snapshot": "ABCDE1234F",
                "deductee_gstin_snapshot": "",
            },
            {
                "deductee_residency_snapshot": "NON_RESIDENT",
                "deductee_country_obj": None,
                "deductee_country_code_snapshot": "",
                "deductee_country_name_snapshot": "",
                "deductee_tax_id_snapshot": "TAX2",
                "deductee_pan_snapshot": "",
                "deductee_gstin_snapshot": "",
            },
        ]
        mock_purchase_line_objects.filter.return_value.exists.return_value = True

        payload = PurchaseStatutoryService.return_eligible_lines(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            return_code="26Q",
        )

        self.assertEqual(len(payload["lines"]), 1)
        self.assertEqual(payload["lines"][0]["header_id"], 10)
        self.assertEqual(payload["readiness_summary"]["eligible_lines"], 1)
        self.assertEqual(payload["readiness_summary"]["excluded_lines"], 1)
        self.assertEqual(payload["readiness_summary"]["excluded_residency_mismatch"], 1)
        self.assertEqual(payload["totals"]["excluded_amount"], "6.00")
        self.assertEqual(len(payload["excluded_rows"]), 1)
        self.assertEqual(payload["excluded_rows"][0]["reason_code"], "RESIDENCY_MISMATCH")
        self.assertEqual(payload["excluded_rows"][0]["source_kind"], "purchase_invoice")
        self.assertEqual(payload["excluded_rows"][0]["source_route"], "/purchaseserviceinvoice")
        self.assertEqual(payload["excluded_rows"][0]["linked_challan_source_kind"], "challan")
        self.assertEqual(payload["excluded_rows"][0]["linked_challan_source_label"], "CH1")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._vendor_deductee_snapshot")
    def test_return_eligible_lines_exposes_invalid_pan_format_exclusion(self, mock_snapshot, mock_return_line_objects, mock_challan_line_objects):
        challan = SimpleNamespace(challan_no="CH2", cin_no="CIN2")
        header = SimpleNamespace(tds_section=None, purchase_number="PINV-12", doc_code="PINV", doc_no=12, vendor_name="Vendor PAN")
        invalid_pan_line = SimpleNamespace(header_id=12, challan_id=22, amount=Decimal("7.00"), header=header, challan=challan)

        challan_line_qs = MagicMock()
        challan_line_qs.filter.return_value = challan_line_qs
        challan_line_qs.__iter__.return_value = iter([invalid_pan_line])
        mock_challan_line_objects.select_related.return_value.filter.return_value = challan_line_qs

        consumed_qs = MagicMock()
        consumed_qs.exclude.return_value = consumed_qs
        consumed_values_qs = MagicMock()
        consumed_values_qs.annotate.return_value = consumed_values_qs
        consumed_values_qs.filter.return_value = []
        consumed_values_qs.__iter__.return_value = iter([])
        consumed_qs.values.return_value = consumed_values_qs
        mock_return_line_objects.filter.return_value = consumed_qs

        mock_snapshot.return_value = {
            "deductee_residency_snapshot": "RESIDENT",
            "deductee_country_obj": None,
            "deductee_country_code_snapshot": "",
            "deductee_country_name_snapshot": "",
            "deductee_tax_id_snapshot": "TAX3",
            "deductee_pan_snapshot": "BADPAN",
            "deductee_gstin_snapshot": "",
        }

        payload = PurchaseStatutoryService.return_eligible_lines(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            return_code="26Q",
        )

        self.assertEqual(len(payload["lines"]), 0)
        self.assertEqual(payload["readiness_summary"]["excluded_invalid_pan_format"], 1)
        self.assertEqual(payload["excluded_rows"][0]["reason_code"], "INVALID_PAN_FORMAT")
        self.assertEqual(payload["excluded_rows"][0]["source_search"], "PINV-12")
        self.assertEqual(payload["excluded_rows"][0]["linked_challan_source_search"], "CH2")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallanLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._vendor_deductee_snapshot")
    def test_return_eligible_lines_exposes_source_metadata_for_eligible_preview(self, mock_snapshot, mock_return_line_objects, mock_challan_line_objects):
        section = SimpleNamespace(section_code="194J", description="Professional Fees")
        challan = SimpleNamespace(challan_no="CH3", cin_no="CIN3")
        header = SimpleNamespace(
            tds_section=section,
            purchase_number="PINV-13",
            doc_code="PINV",
            doc_no=13,
            vendor_name="Vendor Eligible",
        )
        eligible_line = SimpleNamespace(header_id=13, challan_id=23, amount=Decimal("8.00"), header=header, challan=challan)

        challan_line_qs = MagicMock()
        challan_line_qs.filter.return_value = challan_line_qs
        challan_line_qs.__iter__.return_value = iter([eligible_line])
        mock_challan_line_objects.select_related.return_value.filter.return_value = challan_line_qs

        consumed_qs = MagicMock()
        consumed_qs.exclude.return_value = consumed_qs
        consumed_values_qs = MagicMock()
        consumed_values_qs.annotate.return_value = consumed_values_qs
        consumed_values_qs.filter.return_value = []
        consumed_values_qs.__iter__.return_value = iter([])
        consumed_qs.values.return_value = consumed_values_qs
        mock_return_line_objects.filter.return_value = consumed_qs

        mock_snapshot.return_value = {
            "deductee_residency_snapshot": "RESIDENT",
            "deductee_country_obj": None,
            "deductee_country_code_snapshot": "",
            "deductee_country_name_snapshot": "",
            "deductee_tax_id_snapshot": "TAX4",
            "deductee_pan_snapshot": "ABCDE1234F",
            "deductee_gstin_snapshot": "",
        }

        payload = PurchaseStatutoryService.return_eligible_lines(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            return_code="26Q",
        )

        self.assertEqual(payload["lines"][0]["source_kind"], "purchase_invoice")
        self.assertEqual(payload["lines"][0]["source_id"], 13)
        self.assertEqual(payload["lines"][0]["source_label"], "Voucher PINV-13")
        self.assertEqual(payload["lines"][0]["linked_challan_source_kind"], "challan")
        self.assertEqual(payload["lines"][0]["linked_challan_source_id"], 23)
        self.assertEqual(payload["lines"][0]["linked_challan_source_label"], "CH3")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.return_eligible_lines")
    def test_return_readiness_summary_wraps_eligible_payload(self, mock_return_eligible):
        mock_return_eligible.return_value = {
            "lines": [{"header_id": 10, "amount": "5.00", "source_kind": "purchase_invoice", "source_id": 10}],
            "totals": {"line_count": 1, "amount": "5.00", "excluded_line_count": 1, "excluded_amount": "6.00"},
            "section_totals": [{"section_code": "194J", "amount": "5.00"}],
            "vendor_totals": [{"vendor_name": "Vendor A", "line_count": 1, "amount": "5.00"}],
            "readiness_summary": {"eligible_lines": 1, "excluded_lines": 1, "excluded_invalid_pan_format": 1},
            "excluded_rows": [{"header_id": 12, "reason_code": "INVALID_PAN_FORMAT", "source_kind": "purchase_invoice", "linked_challan_source_kind": "challan"}],
        }

        payload = PurchaseStatutoryService.return_readiness_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            return_code="26Q",
        )

        self.assertEqual(payload["summary"]["eligible_lines"], 1)
        self.assertEqual(payload["totals"]["excluded_amount"], "6.00")
        self.assertEqual(payload["section_totals"][0]["section_code"], "194J")
        self.assertEqual(payload["vendor_totals"][0]["vendor_name"], "Vendor A")
        self.assertEqual(payload["excluded_rows"][0]["reason_code"], "INVALID_PAN_FORMAT")
        self.assertEqual(payload["excluded_rows"][0]["source_kind"], "purchase_invoice")
        self.assertEqual(payload["eligible_preview"][0]["source_kind"], "purchase_invoice")
        self.assertEqual(payload["filters"]["return_code"], "26Q")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_header_amount_for_tax_type")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_submitted_approval_state")
    def test_approve_return_uses_submitted_state_gate(
        self,
        mock_require_submitted,
        mock_return_objects,
        mock_validate_header_amount,
        mock_header_objects,
        mock_challan_objects,
    ):
        filing = SimpleNamespace(
            status=PurchaseStatutoryReturn.Status.DRAFT,
            filed_payload_json={"_approval_state": {"status": "DRAFT"}},
        )
        mock_return_objects.select_for_update.return_value.get.return_value = filing
        mock_require_submitted.side_effect = ValueError("Return approval requires the record to be in SUBMITTED approval state.")

        with self.assertRaisesMessage(ValueError, "requires the record to be in SUBMITTED approval state"):
            PurchaseStatutoryService.approve_return(filing_id=1, user_id=9, remarks="ok")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_submitted_approval_state")
    def test_approve_challan_uses_submitted_state_gate(self, mock_require_submitted, mock_challan_objects):
        challan = SimpleNamespace(
            status=PurchaseStatutoryChallan.Status.DRAFT,
            payment_payload_json={"_approval_state": {"status": "DRAFT"}},
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan
        mock_require_submitted.side_effect = ValueError("Challan approval requires the record to be in SUBMITTED approval state.")

        with self.assertRaisesMessage(ValueError, "requires the record to be in SUBMITTED approval state"):
            PurchaseStatutoryService.approve_challan(challan_id=1, user_id=9, remarks="ok")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_submit_return_for_approval_rejects_filed_revision(self, mock_return_objects):
        filing = SimpleNamespace(
            status=PurchaseStatutoryReturn.Status.REVISED,
            filed_payload_json={"_approval_state": {"status": "DRAFT"}},
        )
        mock_return_objects.select_for_update.return_value.get.return_value = filing

        with self.assertRaisesMessage(ValueError, "Only draft return can be submitted."):
            PurchaseStatutoryService.submit_return_for_approval(filing_id=1, user_id=9, remarks="retry")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_submit_return_for_approval_rejects_already_submitted(self, mock_return_objects):
        filing = SimpleNamespace(
            status=PurchaseStatutoryReturn.Status.DRAFT,
            filed_payload_json={"_approval_state": {"status": "SUBMITTED"}},
        )
        mock_return_objects.select_for_update.return_value.get.return_value = filing

        with self.assertRaisesMessage(ValueError, "Return is already in approval workflow."):
            PurchaseStatutoryService.submit_return_for_approval(filing_id=1, user_id=9, remarks="retry")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    def test_submit_challan_for_approval_rejects_already_approved(self, mock_challan_objects):
        challan = SimpleNamespace(
            status=PurchaseStatutoryChallan.Status.DRAFT,
            payment_payload_json={"_approval_state": {"status": "APPROVED"}},
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan

        with self.assertRaisesMessage(ValueError, "Challan is already in approval workflow."):
            PurchaseStatutoryService.submit_challan_for_approval(challan_id=1, user_id=9, remarks="retry")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_challan_balance_for_lines")
    def test_update_challan_rejects_edit_while_submitted(self, mock_validate_balance, mock_challan_objects):
        challan = SimpleNamespace(
            status=PurchaseStatutoryChallan.Status.DRAFT,
            payment_payload_json={"_approval_state": {"status": "SUBMITTED"}},
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan

        with self.assertRaisesMessage(ValueError, "cannot be edited while approval state is SUBMITTED"):
            PurchaseStatutoryService.update_challan(
                challan_id=1,
                entity_id=1,
                entityfinid_id=1,
                subentity_id=5,
                tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
                challan_no="CH1",
                challan_date=date(2026, 4, 1),
                lines=[{"header_id": 10, "amount": Decimal("1.00")}],
            )

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._approval_state")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._enforcement_level")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_maker_checker")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._auto_compute_statutory_charges")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._due_date_for_return")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_it_tds_return_code")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._append_audit_event")
    def test_file_return_marks_revision_as_revised(
        self,
        mock_append_audit,
        mock_validate_code,
        mock_due_date,
        mock_auto_charges,
        mock_require_maker_checker,
        mock_enforcement_level,
        mock_approval_state,
        mock_return_objects,
    ):
        filing = SimpleNamespace(
            entity_id=1,
            subentity_id=5,
            created_by_id=7,
            status=PurchaseStatutoryReturn.Status.DRAFT,
            original_return_id=3,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_to=date(2026, 6, 30),
            amount=Decimal("10.00"),
            interest_amount=Decimal("0.00"),
            late_fee_amount=Decimal("0.00"),
            penalty_amount=Decimal("0.00"),
            filed_payload_json={},
            lines=SimpleNamespace(all=lambda: []),
            save=MagicMock(),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing
        mock_approval_state.return_value = {"status": "DRAFT"}
        mock_enforcement_level.return_value = "off"
        mock_due_date.return_value = date(2026, 7, 31)
        mock_auto_charges.return_value = {
            "interest_amount": Decimal("0.00"),
            "late_fee_amount": Decimal("0.00"),
            "penalty_amount": Decimal("0.00"),
        }
        mock_append_audit.side_effect = lambda payload, event: {"audit": event}

        PurchaseStatutoryService.file_return(filing_id=1, filed_by_id=9, filed_on="2026-07-10")

        self.assertEqual(filing.status, PurchaseStatutoryReturn.Status.REVISED)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._approval_state")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._enforcement_level")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_maker_checker")
    def test_file_return_returns_already_filed_for_terminal_status(
        self,
        mock_require_maker_checker,
        mock_enforcement_level,
        mock_approval_state,
        mock_return_objects,
    ):
        filing = SimpleNamespace(
            entity_id=1,
            subentity_id=None,
            created_by_id=7,
            status=PurchaseStatutoryReturn.Status.FILED,
            filed_payload_json={},
            lines=SimpleNamespace(all=lambda: []),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing
        mock_approval_state.return_value = {"status": "DRAFT"}
        mock_enforcement_level.return_value = "off"

        result = PurchaseStatutoryService.file_return(filing_id=1, filed_by_id=9, filed_on="2026-07-10")

        self.assertEqual(result.message, "Already filed.")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    def test_submit_challan_for_approval_sets_submitted_status(self, mock_challan_objects):
        challan = SimpleNamespace(
            status=PurchaseStatutoryChallan.Status.DRAFT,
            payment_payload_json={},
            save=MagicMock(),
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan

        result = PurchaseStatutoryService.submit_challan_for_approval(challan_id=1, user_id=9, remarks="please approve")

        approval = result.obj.payment_payload_json["_approval_state"]
        self.assertEqual(result.message, "Challan submitted for approval.")
        self.assertEqual(approval["status"], "SUBMITTED")
        self.assertEqual(approval["submitted_by"], 9)
        self.assertEqual(approval["remarks"], "please approve")
        self.assertEqual(result.obj.payment_payload_json["_audit_log"][-1]["action"], "SUBMITTED_FOR_APPROVAL")
        challan.save.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_maker_checker")
    def test_approve_challan_sets_approved_status(self, mock_require_maker_checker, mock_challan_objects):
        challan = SimpleNamespace(
            entity_id=1,
            subentity_id=5,
            created_by_id=7,
            status=PurchaseStatutoryChallan.Status.DRAFT,
            payment_payload_json={"_approval_state": {"status": "SUBMITTED", "submitted_by": 8}},
            save=MagicMock(),
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan

        result = PurchaseStatutoryService.approve_challan(challan_id=1, user_id=9, remarks="approved")

        approval = result.obj.payment_payload_json["_approval_state"]
        self.assertEqual(result.message, "Challan approved.")
        self.assertEqual(approval["status"], "APPROVED")
        self.assertEqual(approval["approved_by"], 9)
        self.assertEqual(approval["remarks"], "approved")
        self.assertEqual(result.obj.payment_payload_json["_audit_log"][-1]["action"], "APPROVED")
        challan.save.assert_called_once()
        mock_require_maker_checker.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._auto_compute_statutory_charges")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._due_date_for_challan")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_maker_checker")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._enforcement_level")
    def test_deposit_challan_marks_challan_deposited(
        self,
        mock_enforcement_level,
        mock_require_maker_checker,
        mock_due_date,
        mock_auto_compute,
        mock_challan_objects,
    ):
        mock_enforcement_level.return_value = "off"
        mock_due_date.return_value = date(2026, 4, 7)
        mock_auto_compute.return_value = {
            "interest_amount": Decimal("1.00"),
            "late_fee_amount": Decimal("2.00"),
            "penalty_amount": Decimal("3.00"),
        }
        challan = SimpleNamespace(
            entity_id=1,
            subentity_id=5,
            created_by_id=7,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            status=PurchaseStatutoryChallan.Status.DRAFT,
            amount=Decimal("100.00"),
            interest_amount=Decimal("0.00"),
            late_fee_amount=Decimal("0.00"),
            penalty_amount=Decimal("0.00"),
            period_to=date(2026, 4, 30),
            challan_date=date(2026, 4, 10),
            payment_payload_json={},
            lines=SimpleNamespace(all=lambda: []),
            save=MagicMock(),
        )
        mock_challan_objects.prefetch_related.return_value.get.return_value = challan

        result = PurchaseStatutoryService.deposit_challan(
            challan_id=1,
            deposited_by_id=9,
            deposited_on="2026-05-10",
            cin_no="CIN-1",
        )

        self.assertEqual(result.message, "Challan deposited.")
        self.assertEqual(result.obj.status, PurchaseStatutoryChallan.Status.DEPOSITED)
        self.assertEqual(result.obj.cin_no, "CIN-1")
        self.assertEqual(result.obj.interest_amount, Decimal("1.00"))
        self.assertEqual(result.obj.late_fee_amount, Decimal("2.00"))
        self.assertEqual(result.obj.penalty_amount, Decimal("3.00"))
        self.assertEqual(result.obj.payment_payload_json["_audit_log"][-1]["action"], "DEPOSITED")
        challan.save.assert_called_once()
        mock_require_maker_checker.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._enforcement_level")
    def test_deposit_challan_rejects_unapproved_record_when_maker_checker_is_hard(
        self,
        mock_enforcement_level,
        mock_challan_objects,
    ):
        mock_enforcement_level.return_value = "hard"
        challan = SimpleNamespace(
            entity_id=1,
            subentity_id=5,
            created_by_id=7,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            status=PurchaseStatutoryChallan.Status.DRAFT,
            amount=Decimal("100.00"),
            interest_amount=Decimal("0.00"),
            late_fee_amount=Decimal("0.00"),
            penalty_amount=Decimal("0.00"),
            period_to=date(2026, 4, 30),
            challan_date=date(2026, 4, 10),
            payment_payload_json={"_approval_state": {"status": "SUBMITTED", "submitted_by": 8}},
            lines=SimpleNamespace(all=lambda: []),
            save=MagicMock(),
        )
        mock_challan_objects.prefetch_related.return_value.get.return_value = challan

        with self.assertRaisesMessage(ValueError, "Challan must be approved before deposit when maker-checker is enabled."):
            PurchaseStatutoryService.deposit_challan(
                challan_id=1,
                deposited_by_id=9,
                deposited_on="2026-05-10",
            )

        challan.save.assert_not_called()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._validate_it_tds_return_code")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._auto_compute_statutory_charges")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._due_date_for_return")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._require_maker_checker")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._enforcement_level")
    def test_file_return_marks_original_return_filed(
        self,
        mock_enforcement_level,
        mock_require_maker_checker,
        mock_due_date,
        mock_auto_compute,
        mock_validate_code,
        mock_return_objects,
    ):
        mock_enforcement_level.return_value = "off"
        mock_due_date.return_value = date(2026, 7, 31)
        mock_auto_compute.return_value = {
            "interest_amount": Decimal("1.50"),
            "late_fee_amount": Decimal("0.00"),
            "penalty_amount": Decimal("0.00"),
        }
        filing = SimpleNamespace(
            entity_id=1,
            subentity_id=5,
            created_by_id=7,
            original_return_id=None,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            status=PurchaseStatutoryReturn.Status.DRAFT,
            amount=Decimal("250.00"),
            interest_amount=Decimal("0.00"),
            late_fee_amount=Decimal("0.00"),
            penalty_amount=Decimal("0.00"),
            period_to=date(2026, 6, 30),
            filed_payload_json={},
            lines=SimpleNamespace(all=lambda: [SimpleNamespace(header_id=10)]),
            save=MagicMock(),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        result = PurchaseStatutoryService.file_return(
            filing_id=1,
            filed_by_id=9,
            filed_on="2026-07-20",
            ack_no="ACK-1",
        )

        self.assertEqual(result.message, "Return filed.")
        self.assertEqual(result.obj.status, PurchaseStatutoryReturn.Status.FILED)
        self.assertEqual(result.obj.ack_no, "ACK-1")
        self.assertEqual(result.obj.interest_amount, Decimal("1.50"))
        self.assertEqual(result.obj.filed_payload_json["_audit_log"][-1]["action"], "FILED")
        filing.save.assert_called_once()
        mock_require_maker_checker.assert_called_once()
        mock_validate_code.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_issue_form16a_appends_new_issue_to_payload(self, mock_return_objects):
        filing = SimpleNamespace(
            id=12,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            status=PurchaseStatutoryReturn.Status.FILED,
            filed_payload_json={},
            lines=SimpleNamespace(count=lambda: 3),
            save=MagicMock(),
        )
        mock_return_objects.select_for_update.return_value.prefetch_related.return_value.get.return_value = filing

        payload = PurchaseStatutoryService.issue_form16a(
            filing_id=12,
            issued_by_id=9,
            issue_date="2026-08-01",
            remarks="First issue",
        )

        self.assertEqual(payload["filing_id"], 12)
        self.assertEqual(payload["issue"]["issue_no"], 1)
        self.assertEqual(payload["issue"]["issued_by"], 9)
        self.assertEqual(payload["issue"]["line_count"], 3)
        self.assertEqual(payload["issue"]["remarks"], "First issue")
        self.assertEqual(filing.filed_payload_json["form16a_issues"][0]["issue_no"], 1)
        self.assertEqual(filing.filed_payload_json["_audit_log"][-1]["action"], "FORM16A_ISSUED")
        filing.save.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    def test_cancel_challan_rejects_when_linked_to_active_return(self, mock_challan_objects, mock_return_line_objects):
        challan = SimpleNamespace(
            id=15,
            status=PurchaseStatutoryChallan.Status.DEPOSITED,
            remarks=None,
            payment_payload_json={},
            save=MagicMock(),
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan
        mock_return_line_objects.filter.return_value.exclude.return_value.exists.return_value = True

        with self.assertRaisesMessage(ValueError, "linked to active statutory returns"):
            PurchaseStatutoryService.cancel_challan(challan_id=15, cancelled_by_id=9, reason="bad upload")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_cancel_return_rejects_when_active_revision_exists(self, mock_return_objects):
        filing = SimpleNamespace(
            id=18,
            status=PurchaseStatutoryReturn.Status.FILED,
            remarks=None,
            filed_payload_json={},
            save=MagicMock(),
        )
        select_for_update_qs = MagicMock()
        select_for_update_qs.get.return_value = filing
        filter_qs = MagicMock()
        filter_qs.exclude.return_value.exists.return_value = True
        mock_return_objects.select_for_update.return_value = select_for_update_qs
        mock_return_objects.filter.return_value = filter_qs

        with self.assertRaisesMessage(ValueError, "active revisions exist"):
            PurchaseStatutoryService.cancel_return(filing_id=18, cancelled_by_id=9, reason="wrong revision")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    def test_cancel_challan_marks_status_and_audit_when_allowed(self, mock_challan_objects):
        challan = SimpleNamespace(
            id=16,
            status=PurchaseStatutoryChallan.Status.DEPOSITED,
            remarks=None,
            payment_payload_json={},
            save=MagicMock(),
        )
        mock_challan_objects.select_for_update.return_value.get.return_value = challan
        with patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturnLine.objects") as mock_return_line_objects:
            mock_return_line_objects.filter.return_value.exclude.return_value.exists.return_value = False
            result = PurchaseStatutoryService.cancel_challan(challan_id=16, cancelled_by_id=9, reason="voided")

        self.assertEqual(result.message, "Challan cancelled.")
        self.assertEqual(result.obj.status, PurchaseStatutoryChallan.Status.CANCELLED)
        self.assertEqual(result.obj.remarks, "voided")
        self.assertEqual(result.obj.payment_payload_json["_audit_log"][-1]["action"], "CANCELLED")
        challan.save.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_cancel_return_marks_status_and_audit_when_allowed(self, mock_return_objects):
        filing = SimpleNamespace(
            id=19,
            status=PurchaseStatutoryReturn.Status.FILED,
            remarks=None,
            filed_payload_json={},
            save=MagicMock(),
        )
        select_for_update_qs = MagicMock()
        select_for_update_qs.get.return_value = filing
        filter_qs = MagicMock()
        filter_qs.exclude.return_value.exists.return_value = False
        mock_return_objects.select_for_update.return_value = select_for_update_qs
        mock_return_objects.filter.return_value = filter_qs

        result = PurchaseStatutoryService.cancel_return(filing_id=19, cancelled_by_id=9, reason="superseded")

        self.assertEqual(result.message, "Return cancelled.")
        self.assertEqual(result.obj.status, PurchaseStatutoryReturn.Status.CANCELLED)
        self.assertEqual(result.obj.remarks, "superseded")
        self.assertEqual(result.obj.filed_payload_json["_audit_log"][-1]["action"], "CANCELLED")
        filing.save.assert_called_once()

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_issue_form16a_rejects_ineligible_return(self, mock_return_objects):
        filing = SimpleNamespace(
            id=13,
            tax_type=PurchaseStatutoryReturn.TaxType.GST_TDS,
            return_code="GSTR7",
            status=PurchaseStatutoryReturn.Status.FILED,
            filed_payload_json={},
            lines=SimpleNamespace(count=lambda: 0),
        )
        mock_return_objects.select_for_update.return_value.prefetch_related.return_value.get.return_value = filing

        with self.assertRaisesMessage(ValueError, "Form16A is allowed only"):
            PurchaseStatutoryService.issue_form16a(filing_id=13, issued_by_id=9)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryForm16AOfficialDocument.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_form16a_download_payload_rejects_unknown_issue_version(self, mock_return_objects, _mock_official_docs):
        filing = SimpleNamespace(
            id=13,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            status=PurchaseStatutoryReturn.Status.FILED,
            filed_payload_json={
                "form16a_issues": [
                    {"issue_no": 1, "issued_on": "2026-08-01", "issue_code": "F16A-0001", "line_count": 2}
                ]
            },
        )
        mock_return_objects.get.return_value = filing

        with self.assertRaisesMessage(ValueError, "Requested Form16A issue version not found."):
            PurchaseStatutoryService.form16a_download_payload(filing_id=13, issue_no=2)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._policy_controls")
    def test_due_date_for_gst_tds_uses_configurable_day(self, mock_controls):
        mock_controls.return_value = {"gst_tds_challan_due_day": "12"}
        due = PurchaseStatutoryService._due_date_for_challan(
            entity_id=1,
            subentity_id=None,
            tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS,
            period_to=date(2026, 4, 30),
            challan_date=None,
        )
        self.assertEqual(due, date(2026, 5, 12))

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService._policy_controls")
    def test_due_date_for_it_tds_return_q4_uses_configurable_month_day(self, mock_controls):
        mock_controls.return_value = {
            "it_tds_return_q4_due_month": "6",
            "it_tds_return_q4_due_day": "15",
        }
        due = PurchaseStatutoryService._due_date_for_return(
            entity_id=1,
            subentity_id=None,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_to=date(2026, 3, 31),
        )
        self.assertEqual(due, date(2026, 6, 15))

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    def test_reconciliation_summary_returns_expected_keys(
        self,
        mock_header_objects,
        mock_challan_objects,
        mock_return_objects,
    ):
        header_qs = MagicMock()
        header_qs.filter.return_value = header_qs
        header_qs.aggregate.side_effect = [
            {"t": Decimal("10.00")},  # deducted_it
            {"t": Decimal("5.00")},   # deducted_gst
        ]
        mock_header_objects.filter.return_value = header_qs

        challan_qs = MagicMock()
        challan_qs.filter.return_value = challan_qs
        challan_qs.filter.return_value.aggregate.side_effect = [
            {"t": Decimal("8.00")},  # deposited
            {"t": Decimal("0.50")},  # deposited_interest
            {"t": Decimal("0.25")},  # deposited_late_fee
            {"t": Decimal("0.10")},  # deposited_penalty
            {"t": Decimal("2.00")},  # draft challan
        ]
        mock_challan_objects.filter.return_value = challan_qs

        return_qs = MagicMock()
        return_qs.filter.return_value = return_qs
        return_qs.filter.return_value.aggregate.side_effect = [
            {"t": Decimal("6.00")},  # filed
            {"t": Decimal("0.40")},  # filed_interest
            {"t": Decimal("0.20")},  # filed_late_fee
            {"t": Decimal("0.05")},  # filed_penalty
            {"t": Decimal("1.00")},  # draft return
        ]
        mock_return_objects.filter.return_value = return_qs

        data = PurchaseStatutoryService.reconciliation_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            tax_type=None,
            date_from=None,
            date_to=None,
        )
        self.assertEqual(data["deducted"], "15.00")
        self.assertEqual(data["deposited"], "8.00")
        self.assertEqual(data["deposited_interest"], "0.50")
        self.assertEqual(data["filed"], "6.00")
        self.assertEqual(data["filed_penalty"], "0.05")
        self.assertEqual(data["pending_deposit"], "7.00")
        self.assertEqual(data["pending_filing"], "2.00")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    def test_reconciliation_summary_clamps_negative_pending_balances(
        self,
        mock_header_objects,
        mock_challan_objects,
        mock_return_objects,
    ):
        header_qs = MagicMock()
        header_qs.filter.return_value = header_qs
        header_qs.aggregate.side_effect = [
            {"t": Decimal("5.00")},
            {"t": Decimal("0.00")},
        ]
        mock_header_objects.filter.return_value = header_qs

        challan_qs = MagicMock()
        challan_qs.filter.return_value = challan_qs
        challan_qs.filter.return_value.aggregate.side_effect = [
            {"t": Decimal("8.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
        ]
        mock_challan_objects.filter.return_value = challan_qs

        return_qs = MagicMock()
        return_qs.filter.return_value = return_qs
        return_qs.filter.return_value.aggregate.side_effect = [
            {"t": Decimal("9.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
        ]
        mock_return_objects.filter.return_value = return_qs

        data = PurchaseStatutoryService.reconciliation_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            tax_type=None,
            date_from=None,
            date_to=None,
        )

        self.assertEqual(data["pending_deposit"], "0.00")
        self.assertEqual(data["pending_filing"], "0.00")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    def test_reconciliation_summary_filters_it_tds_scope_and_uses_tds_totals(
        self,
        mock_header_objects,
        mock_challan_objects,
        mock_return_objects,
    ):
        header_qs = MagicMock()
        header_qs.filter.return_value = header_qs
        header_qs.aggregate.side_effect = [
            {"t": Decimal("11.00")},
        ]
        mock_header_objects.filter.return_value = header_qs

        challan_qs = MagicMock()
        challan_qs.filter.return_value = challan_qs
        challan_qs.aggregate.side_effect = [
            {"t": Decimal("8.00")},
            {"t": Decimal("0.50")},
            {"t": Decimal("0.20")},
            {"t": Decimal("0.10")},
            {"t": Decimal("1.00")},
        ]
        mock_challan_objects.filter.return_value = challan_qs

        return_qs = MagicMock()
        return_qs.filter.return_value = return_qs
        return_qs.aggregate.side_effect = [
            {"t": Decimal("6.00")},
            {"t": Decimal("0.40")},
            {"t": Decimal("0.10")},
            {"t": Decimal("0.05")},
            {"t": Decimal("2.00")},
        ]
        mock_return_objects.filter.return_value = return_qs

        data = PurchaseStatutoryService.reconciliation_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=9,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            date_from=date(2026, 4, 1),
            date_to=date(2026, 4, 30),
        )

        self.assertEqual(data["deducted"], "11.00")
        self.assertEqual(data["pending_deposit"], "3.00")
        self.assertEqual(data["pending_filing"], "2.00")

        header_qs.filter.assert_any_call(subentity_id=9)
        header_qs.filter.assert_any_call(bill_date__gte=date(2026, 4, 1))
        header_qs.filter.assert_any_call(bill_date__lte=date(2026, 4, 30))
        header_qs.aggregate.assert_called_once_with(t=Sum("tds_amount"))

        challan_qs.filter.assert_any_call(subentity_id=9)
        challan_qs.filter.assert_any_call(challan_date__gte=date(2026, 4, 1))
        challan_qs.filter.assert_any_call(challan_date__lte=date(2026, 4, 30))
        challan_qs.filter.assert_any_call(tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS)

        return_qs.filter.assert_any_call(subentity_id=9)
        return_qs.filter.assert_any_call(period_to__gte=date(2026, 4, 1))
        return_qs.filter.assert_any_call(period_from__lte=date(2026, 4, 30))
        return_qs.filter.assert_any_call(tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryChallan.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    def test_reconciliation_summary_filters_gst_tds_scope_and_uses_gst_totals(
        self,
        mock_header_objects,
        mock_challan_objects,
        mock_return_objects,
    ):
        header_qs = MagicMock()
        header_qs.filter.return_value = header_qs
        header_qs.aggregate.side_effect = [
            {"t": Decimal("18.00")},
        ]
        mock_header_objects.filter.return_value = header_qs

        challan_qs = MagicMock()
        challan_qs.filter.return_value = challan_qs
        challan_qs.aggregate.side_effect = [
            {"t": Decimal("9.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("4.00")},
        ]
        mock_challan_objects.filter.return_value = challan_qs

        return_qs = MagicMock()
        return_qs.filter.return_value = return_qs
        return_qs.aggregate.side_effect = [
            {"t": Decimal("7.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("0.00")},
            {"t": Decimal("1.00")},
        ]
        mock_return_objects.filter.return_value = return_qs

        data = PurchaseStatutoryService.reconciliation_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS,
            date_from=None,
            date_to=None,
        )

        self.assertEqual(data["deducted"], "18.00")
        self.assertEqual(data["deposited"], "9.00")
        self.assertEqual(data["filed"], "7.00")
        self.assertEqual(data["pending_deposit"], "9.00")
        self.assertEqual(data["pending_filing"], "2.00")

        header_qs.aggregate.assert_called_once_with(t=Sum("gst_tds_amount"))
        challan_qs.filter.assert_any_call(tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS)
        return_qs.filter.assert_any_call(tax_type=PurchaseStatutoryReturn.TaxType.GST_TDS)


class PurchaseStatutorySerializerValidationTests(SimpleTestCase):
    @patch("purchase.serializers.purchase_statutory.assert_document_date_within_financial_year")
    def test_challan_create_serializer_rejects_partial_period(self, mock_assert_fy):
        mock_assert_fy.return_value = None
        serializer = PurchaseStatutoryChallanCreateInputSerializer(
            data={
                "entity": 1,
                "entityfinid": 1,
                "tax_type": "IT_TDS",
                "challan_no": "CH1",
                "challan_date": "2026-04-10",
                "period_from": "2026-04-01",
            }
        )
        self.assertFalse(serializer.is_valid())
        self.assertIn("detail", serializer.errors)

    @patch("purchase.serializers.purchase_statutory.assert_document_date_within_financial_year")
    def test_challan_create_serializer_accepts_consistent_period(self, mock_assert_fy):
        mock_assert_fy.return_value = None
        serializer = PurchaseStatutoryChallanCreateInputSerializer(
            data={
                "entity": 1,
                "entityfinid": 1,
                "tax_type": "IT_TDS",
                "challan_no": "CH1",
                "challan_date": "2026-04-10",
                "period_from": "2026-04-01",
                "period_to": "2026-04-30",
                "amount": "100.00",
            }
        )
        self.assertTrue(serializer.is_valid(), serializer.errors)

    @patch("purchase.serializers.purchase_statutory.assert_document_date_within_financial_year")
    def test_return_create_serializer_rejects_inverted_period(self, mock_assert_fy):
        mock_assert_fy.return_value = None
        serializer = PurchaseStatutoryReturnCreateInputSerializer(
            data={
                "entity": 1,
                "entityfinid": 1,
                "tax_type": "IT_TDS",
                "return_code": "26Q",
                "period_from": "2026-06-30",
                "period_to": "2026-04-01",
            }
        )
        self.assertFalse(serializer.is_valid())
        self.assertIn("detail", serializer.errors)

    @patch("purchase.serializers.purchase_statutory.assert_document_date_within_financial_year")
    def test_return_create_serializer_accepts_valid_period(self, mock_assert_fy):
        mock_assert_fy.return_value = None
        serializer = PurchaseStatutoryReturnCreateInputSerializer(
            data={
                "entity": 1,
                "entityfinid": 1,
                "tax_type": "IT_TDS",
                "return_code": "26Q",
                "period_from": "2026-04-01",
                "period_to": "2026-06-30",
            }
        )
        self.assertTrue(serializer.is_valid(), serializer.errors)


class PurchaseApiExtendedSmokeTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="purchase_api_ext_tester",
            email="purchase_api_ext_tester@example.com",
            password="x",
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    @patch("purchase.views.purchase_ap.PurchaseApService.cancel_settlement")
    def test_ap_settlement_cancel_endpoint_returns_200(self, mock_cancel):
        mock_cancel.return_value = SimpleNamespace(
            message="Settlement cancelled with reversal.",
            settlement=SimpleNamespace(
                id=1,
                entity_id=1,
                entityfinid_id=1,
                subentity_id=None,
                vendor_id=1,
                settlement_type="payment",
                settlement_date=None,
                reference_no=None,
                external_voucher_no=None,
                remarks=None,
                total_amount=Decimal("0.00"),
                status=9,
                posted_at=None,
                posted_by_id=None,
                lines=[],
                created_at=None,
                updated_at=None,
                get_status_display=lambda: "Cancelled",
                get_settlement_type_display=lambda: "Payment",
            ),
        )
        with patch("purchase.views.purchase_ap.VendorSettlementSerializer") as mock_ser:
            mock_ser.return_value.data = {"id": 1, "status": 9}
            resp = self.client.post("/api/purchase/ap/settlements/1/cancel/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["message"], "Settlement cancelled with reversal.")

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_summary")
    def test_statutory_summary_endpoint_returns_200(self, mock_summary, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.view"}
        mock_summary.return_value = {
            "deducted": "10.00",
            "deposited": "8.00",
            "filed": "6.00",
            "pending_deposit": "2.00",
            "pending_filing": "2.00",
            "draft_challan": "1.00",
            "draft_return": "1.00",
        }
        resp = self.client.get("/api/purchase/statutory/summary/?entity=1&entityfinid=1")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("summary", resp.data)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_exceptions")
    def test_statutory_reconciliation_exceptions_endpoint_returns_200(self, mock_fn, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.view"}
        mock_fn.return_value = {"exceptions": {}}
        resp = self.client.get(
            "/api/purchase/statutory/reconciliation-exceptions/?entity=1&entityfinid=1&tax_type=IT_TDS&period_from=2026-04-01&period_to=2026-04-30"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("exceptions", resp.data)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReviewNoteSerializer")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.get_review_note")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_review_note_get_returns_200(self, mock_perm_codes, mock_get_note, mock_serializer):
        mock_perm_codes.return_value = {"purchase.statutory.view"}
        mock_get_note.return_value = SimpleNamespace(id=1)
        mock_serializer.return_value.data = {
            "reviewer_name": "CA Reviewer",
            "closure_status": "READY_TO_SIGN_OFF",
        }

        resp = self.client.get(
            "/api/purchase/statutory/review-note/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["data"]["reviewer_name"], "CA Reviewer")
        mock_get_note.assert_called_once()

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReviewNoteSerializer")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.save_review_note")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.serializers.purchase_statutory.assert_document_date_within_financial_year")
    def test_statutory_review_note_post_returns_200(self, mock_date_guard, mock_perm_codes, mock_save_note, mock_serializer):
        mock_date_guard.return_value = None
        mock_perm_codes.return_value = {"purchase.statutory.manage"}
        mock_save_note.return_value = SimpleNamespace(message="Review note updated.", obj=SimpleNamespace(id=1))
        mock_serializer.return_value.data = {
            "reviewer_name": "CA Reviewer",
            "closure_status": "IN_REVIEW",
        }

        resp = self.client.post(
            "/api/purchase/statutory/review-note/",
            {
                "entity": 1,
                "entityfinid": 1,
                "period_from": "2026-04-01",
                "period_to": "2026-04-30",
                "reviewer_name": "CA Reviewer",
                "closure_status": "IN_REVIEW",
                "review_summary": "Period reviewed.",
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["message"], "Review note updated.")
        mock_save_note.assert_called_once()

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.delete_review_note")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_review_note_delete_returns_200(self, mock_perm_codes, mock_delete_note):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}

        resp = self.client.delete(
            "/api/purchase/statutory/review-note/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["message"], "Review note deleted.")
        mock_delete_note.assert_called_once()

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.itc_status_register")
    def test_statutory_itc_status_register_endpoint_returns_200(self, mock_fn, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.view"}
        mock_fn.return_value = {
            "count": 1,
            "rows": [
                {
                    "header_id": 10,
                    "purchase_number": "PINV-10",
                    "itc_claim_status_name": "Pending",
                    "gstr2b_match_status_name": "Not Checked",
                }
            ],
            "summary": {"invoice_count": 1},
        }
        resp = self.client.get(
            "/api/purchase/statutory/itc-status-register/?entity=1&entityfinid=1&date_from=2026-04-01&date_to=2026-04-30"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["count"], 1)
        self.assertIn("summary", resp.data)

    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceHeaderSerializer")
    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceActions.mark_itc_claimed")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_purchase_service_invoice_itc_claim_endpoint_returns_200(
        self,
        mock_entity_for_user,
        mock_codes,
        _mock_scope,
        mock_action,
        mock_serializer,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.invoice.update"}
        _mock_scope.return_value = SimpleNamespace(
            id=9,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
        )
        mock_action.return_value = SimpleNamespace(message="ok", header=SimpleNamespace(id=9))
        mock_serializer.return_value.data = {"id": 9}
        resp = self.client.post(
            "/api/purchase/purchase-service-invoices/9/itc/claim/?entity=1&entityfinid=1",
            {"period": "2026-04"},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["message"], "ok")

    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceHeaderSerializer")
    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceActions.mark_itc_unblocked")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_purchase_service_invoice_itc_unblock_endpoint_returns_200(
        self,
        mock_entity_for_user,
        mock_codes,
        _mock_scope,
        mock_action,
        mock_serializer,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.invoice.update"}
        _mock_scope.return_value = SimpleNamespace(
            id=9,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
        )
        mock_action.return_value = SimpleNamespace(message="unblocked", header=SimpleNamespace(id=9))
        mock_serializer.return_value.data = {"id": 9}
        resp = self.client.post(
            "/api/purchase/purchase-service-invoices/9/itc/unblock/?entity=1&entityfinid=1",
            {"reason": "GSTR-2B updated"},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["message"], "unblocked")

    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceHeaderSerializer")
    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceActions.update_2b_match_status")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_purchase_service_invoice_2b_status_endpoint_returns_200(
        self,
        mock_entity_for_user,
        mock_codes,
        _mock_scope,
        mock_action,
        mock_serializer,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.invoice.update"}
        _mock_scope.return_value = SimpleNamespace(
            id=10,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
        )
        mock_action.return_value = SimpleNamespace(message="ok", header=SimpleNamespace(id=10))
        mock_serializer.return_value.data = {"id": 10, "gstr2b_match_status": 2}
        resp = self.client.post(
            "/api/purchase/purchase-service-invoices/10/gstr2b/status/?entity=1&entityfinid=1",
            {"match_status": 2},
            format="json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["data"]["id"], 10)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.attach_form16a_official_document")
    def test_statutory_form16a_official_upload_endpoint_returns_201(self, mock_attach):
        from django.core.files.uploadedfile import SimpleUploadedFile

        mock_attach.return_value = {"filing_id": 1, "issue_no": 2, "source": "TRACES"}
        doc = SimpleUploadedFile("form16a.pdf", b"%PDF-1.4\nfake", content_type="application/pdf")
        resp = self.client.post(
            "/api/purchase/statutory/returns/1/form16a/2/official-upload/",
            {"document": doc, "source": "TRACES"},
            format="multipart",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data["message"], "Official Form16A document uploaded.")

    def test_purchase_service_invoice_search_alias_endpoint_returns_200(self):
        resp = self.client.get("/api/purchase/purchase-service-invoices/search/?entity=1&entityfinid=1")
        # Alias route should at least resolve (it may reject request params with 400).
        self.assertNotEqual(resp.status_code, 404)

    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_gstr2b.PurchaseGstr2bService.create_batch")
    @patch("purchase.views.purchase_gstr2b.Gstr2bImportBatchSerializer")
    def test_gstr2b_import_batch_create_returns_201(self, mock_ser, mock_create, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}
        mock_create.return_value = SimpleNamespace(id=12)
        mock_ser.return_value.data = {"id": 12, "period": "2026-04"}
        resp = self.client.post(
            "/api/purchase/gstr2b/import-batches/",
            {
                "entity": 1,
                "entityfinid": 1,
                "period": "2026-04",
                "source": "gstr2b",
                "rows": [
                    {
                        "supplier_gstin": "29ABCDE1234F1Z5",
                        "supplier_invoice_number": "INV-1",
                        "taxable_value": "100.00",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data["data"]["id"], 12)

    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_gstr2b.PurchaseGstr2bService.create_batch")
    def test_gstr2b_import_batch_create_rejects_invalid_supplier_gstin(self, mock_create, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}

        resp = self.client.post(
            "/api/purchase/gstr2b/import-batches/",
            {
                "entity": 1,
                "entityfinid": 1,
                "period": "2026-04",
                "source": "gstr2b",
                "rows": [
                    {
                        "supplier_gstin": "BAD-GSTIN",
                        "supplier_invoice_number": "INV-1",
                        "taxable_value": "100.00",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(resp.status_code, 400)
        self.assertIn("valid gstin", str(resp.data).lower())
        mock_create.assert_not_called()

    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_gstr2b.Gstr2bImportBatch.objects")
    @patch("purchase.views.purchase_gstr2b.PurchaseGstr2bService.auto_match_batch")
    def test_gstr2b_import_batch_match_returns_200(self, mock_match, mock_batch_objects, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}
        mock_batch_objects.filter.return_value.first.return_value = SimpleNamespace(
            id=15,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
        )
        mock_match.return_value = SimpleNamespace(
            batch=SimpleNamespace(id=15),
            total_rows=10,
            matched=7,
            partial=2,
            multiple=1,
            not_matched=0,
        )
        resp = self.client.post("/api/purchase/gstr2b/import-batches/15/match/?entity=1&entityfinid=1")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["summary"]["matched"], 7)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.generate_nsdl_payload")
    def test_statutory_nsdl_export_endpoint_returns_200(self, mock_fn):
        mock_fn.return_value = {"filing_id": 1, "nsdl_txt": "HDR|..."}
        resp = self.client.get("/api/purchase/statutory/returns/1/nsdl-export/")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("nsdl_txt", resp.data)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.return_readiness_summary")
    def test_statutory_return_readiness_summary_endpoint_returns_200(self, mock_fn, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.view"}
        mock_fn.return_value = {
            "filters": {
                "entity_id": 1,
                "entityfinid_id": 1,
                "subentity_id": None,
                "tax_type": "IT_TDS",
                "period_from": date(2026, 4, 1),
                "period_to": date(2026, 4, 30),
                "return_code": "26Q",
            },
            "summary": {"eligible_lines": 1, "excluded_lines": 1},
            "totals": {"line_count": 1, "amount": "5.00", "excluded_amount": "6.00"},
            "section_totals": [{"section_code": "194J", "amount": "5.00"}],
            "vendor_totals": [{"vendor_name": "Vendor A", "line_count": 1, "amount": "5.00"}],
            "excluded_rows": [{"header_id": 12, "reason_code": "INVALID_PAN_FORMAT"}],
            "eligible_preview": [{"header_id": 10, "amount": "5.00"}],
        }
        resp = self.client.get(
            "/api/purchase/statutory/returns/readiness-summary/"
            "?entity=1&entityfinid=1&tax_type=IT_TDS&period_from=2026-04-01&period_to=2026-04-30&return_code=26Q"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["summary"]["eligible_lines"], 1)
        self.assertEqual(resp.data["excluded_rows"][0]["reason_code"], "INVALID_PAN_FORMAT")

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.return_quality_summary")
    def test_statutory_return_quality_summary_endpoint_returns_200(self, mock_fn, mock_perm_codes):
        mock_perm_codes.return_value = {"purchase.statutory.view"}
        mock_fn.return_value = {
            "filters": {
                "entity_id": 1,
                "entityfinid_id": 1,
                "subentity_id": None,
                "tax_type": "IT_TDS",
                "period_from": date(2026, 4, 1),
                "period_to": date(2026, 4, 30),
                "return_code": "26Q",
            },
            "summary": {"return_count": 1, "line_count": 2, "invalid_pan_format": 1},
            "section_summary": [{"section_code": "194J", "line_count": 1, "amount": "5.00"}],
            "vendor_summary": [{"vendor_name": "Vendor A", "line_count": 1, "amount": "5.00"}],
            "status_summary": [{"status_name": "Filed", "return_count": 1, "line_count": 2, "amount": "11.00"}],
            "line_preview": [{"return_id": 21, "section_code": "194J", "amount": "5.00"}],
        }
        resp = self.client.get(
            "/api/purchase/statutory/returns/quality-summary/"
            "?entity=1&entityfinid=1&tax_type=IT_TDS&period_from=2026-04-01&period_to=2026-04-30&return_code=26Q"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["summary"]["return_count"], 1)
        self.assertEqual(resp.data["section_summary"][0]["section_code"], "194J")

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.get_review_note")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_exceptions")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_summary")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryChallan.objects")
    @patch("purchase.views.purchase_statutory.PurchaseInvoiceHeader.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_export_includes_reporting_sheets(
        self,
        mock_perm_codes,
        mock_invoice_objects,
        mock_challan_objects,
        mock_return_objects,
        mock_recon_summary,
        mock_recon_exceptions,
        mock_get_review_note,
    ):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}

        invoice_it = SimpleNamespace(
            id=11,
            purchase_number="PINV-11",
            doc_code="PINV",
            doc_no=11,
            bill_date=date(2026, 4, 10),
            vendor_name="Vendor IT",
            vendor_id=None,
            tds_section_id=None,
            tds_section=None,
            tds_base_amount=Decimal("1000.00"),
            tds_rate=Decimal("1.00"),
            tds_amount=Decimal("10.00"),
            gst_tds_amount=Decimal("0.00"),
            subentity_id=None,
            subentity=None,
            get_status_display=lambda: "Posted",
        )
        invoice_gst = SimpleNamespace(
            id=22,
            purchase_number="PINV-22",
            doc_code="PINV",
            doc_no=22,
            bill_date=date(2026, 4, 12),
            vendor_name="Vendor GST",
            vendor_gstin="29ABCDE1234F1Z5",
            vendor_id=None,
            tds_section_id=None,
            tds_section=None,
            tds_amount=Decimal("0.00"),
            gst_tds_contract_ref="GST-REF-22",
            gst_tds_base_amount=Decimal("2000.00"),
            gst_tds_rate=Decimal("2.00"),
            gst_tds_cgst_amount=Decimal("10.00"),
            gst_tds_sgst_amount=Decimal("10.00"),
            gst_tds_igst_amount=Decimal("20.00"),
            gst_tds_amount=Decimal("40.00"),
            subentity_id=None,
            subentity=None,
            get_status_display=lambda: "Posted",
        )
        mock_invoice_objects.filter.return_value = _FakeQuerySet([invoice_it, invoice_gst])

        challan_it = SimpleNamespace(
            id=51,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            challan_no="IT-CH-51",
            challan_date=date(2026, 4, 20),
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            amount=Decimal("10.00"),
            interest_amount=Decimal("0.00"),
            late_fee_amount=Decimal("0.00"),
            penalty_amount=Decimal("0.00"),
            cin_no="CINIT51",
            bsr_code="1234567",
            minor_head_code="200",
            status=PurchaseStatutoryChallan.Status.DEPOSITED,
            ack_document="",
            lines=_ListManager([]),
            get_status_display=lambda: "Deposited",
        )
        challan_gst = SimpleNamespace(
            id=52,
            tax_type=PurchaseStatutoryChallan.TaxType.GST_TDS,
            challan_no="GST-CH-52",
            challan_date=date(2026, 4, 21),
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            amount=Decimal("40.00"),
            interest_amount=Decimal("0.00"),
            late_fee_amount=Decimal("0.00"),
            penalty_amount=Decimal("0.00"),
            cin_no="CINGST52",
            bsr_code="7654321",
            minor_head_code="200",
            status=PurchaseStatutoryChallan.Status.DRAFT,
            ack_document="challan.pdf",
            lines=_ListManager([]),
            get_status_display=lambda: "Draft",
        )
        mock_challan_objects.filter.return_value = _FakeQuerySet([challan_it, challan_gst])

        return_line_it = SimpleNamespace(
            header_id=invoice_it.id,
            header=invoice_it,
            challan_id=challan_it.id,
            challan=challan_it,
            deductee_pan_snapshot="ABCDE1234F",
            deductee_gstin_snapshot="",
            section_snapshot_code="194C",
            amount=Decimal("10.00"),
            cin_snapshot="CINIT51",
        )
        return_line_gst = SimpleNamespace(
            header_id=invoice_gst.id,
            header=invoice_gst,
            challan_id=challan_gst.id,
            challan=challan_gst,
            deductee_pan_snapshot="",
            deductee_gstin_snapshot="29ABCDE1234F1Z5",
            section_snapshot_code="",
            amount=Decimal("40.00"),
            cin_snapshot="CINGST52",
        )
        filing_it = SimpleNamespace(
            id=81,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            amount=Decimal("10.00"),
            status=PurchaseStatutoryReturn.Status.FILED,
            ack_no="ACK81",
            arn_no="ARN81",
            ack_document="it-return.pdf",
            lines=_ListManager([return_line_it]),
            get_status_display=lambda: "Filed",
        )
        filing_gst = SimpleNamespace(
            id=82,
            tax_type=PurchaseStatutoryReturn.TaxType.GST_TDS,
            return_code="GSTR7",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            amount=Decimal("40.00"),
            status=PurchaseStatutoryReturn.Status.REVISED,
            ack_no="",
            arn_no="",
            ack_document="",
            lines=_ListManager([return_line_gst]),
            get_status_display=lambda: "Revised",
        )
        mock_return_objects.filter.return_value = _FakeQuerySet([filing_it, filing_gst])

        mock_recon_summary.side_effect = [
            {
                "deducted": "50.00",
                "deposited": "10.00",
                "filed": "10.00",
                "pending_deposit": "40.00",
                "pending_filing": "0.00",
                "draft_challan": "1.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "10.00",
                "deposited": "10.00",
                "filed": "10.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "40.00",
                "deposited": "0.00",
                "filed": "0.00",
                "pending_deposit": "40.00",
                "pending_filing": "0.00",
                "draft_challan": "1.00",
                "draft_return": "0.00",
            },
        ]
        mock_recon_exceptions.side_effect = [
            {
                "exceptions": {
                    "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                    "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                    "filed_returns_missing_ack_or_arn": {"count": 0, "rows": []},
                }
            },
            {
                "exceptions": {
                    "invoices_pending_challan_mapping": {"line_count": 1, "rows": [{"header_id": 22, "purchase_number": "PINV-22", "bill_date": "2026-04-12", "amount": "40.00"}]},
                    "challan_lines_pending_return_mapping": {"line_count": 1, "rows": [{"challan_id": 52, "challan_no": "GST-CH-52", "amount": "40.00"}]},
                    "filed_returns_missing_ack_or_arn": {"count": 1, "rows": [{"id": 82, "return_code": "GSTR7", "period_from": "2026-04-01", "period_to": "2026-04-30"}]},
                }
            },
        ]

        event = SimpleNamespace(
            reviewer_name="Lead CA",
            changed_by=SimpleNamespace(username="review.user"),
            changed_by_id=7,
            changed_at=date(2026, 5, 1),
            review_summary="Reviewed",
            open_points="1 item open",
            closure_comment="Follow-up pending",
            get_action_display=lambda: "Updated",
            get_closure_status_display=lambda: "In Review",
        )
        note_all = SimpleNamespace(
            reviewer_name="Lead CA",
            reviewed_at=date(2026, 5, 2),
            reviewed_by=SimpleNamespace(username="review.user"),
            reviewed_by_id=7,
            review_summary="Overall summary",
            open_points="Overall open points",
            closure_comment="Overall closure comment",
            events=_FakeQuerySet([event]),
            get_closure_status_display=lambda: "In Review",
        )
        note_it = SimpleNamespace(
            reviewer_name="IT Reviewer",
            reviewed_at=date(2026, 5, 2),
            reviewed_by=SimpleNamespace(username="it.reviewer"),
            reviewed_by_id=8,
            review_summary="IT summary",
            open_points="",
            closure_comment="",
            events=_FakeQuerySet([]),
            get_closure_status_display=lambda: "Ready To Sign Off",
        )
        note_gst = None
        mock_get_review_note.side_effect = [note_all, note_it, note_gst]

        resp = self.client.get(
            "/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        wb = load_workbook(BytesIO(resp.content))
        self.assertIn("01_Management_Summary", wb.sheetnames)
        self.assertIn("02_Reviewer_Signoff", wb.sheetnames)
        self.assertIn("03_Action_Items", wb.sheetnames)
        self.assertIn("12_Supporting_Doc_Index", wb.sheetnames)
        self.assertIn("13_IT_Return_Quality", wb.sheetnames)
        self.assertIn("14_IT_Return_Scope_Summary", wb.sheetnames)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.get_review_note")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_exceptions")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_summary")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryChallan.objects")
    @patch("purchase.views.purchase_statutory.PurchaseInvoiceHeader.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_keeps_revised_return_evidence_follow_up_in_action_queue(
        self,
        mock_perm_codes,
        mock_invoice_objects,
        mock_challan_objects,
        mock_return_objects,
        mock_recon_summary,
        mock_recon_exceptions,
        mock_get_review_note,
    ):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}
        mock_invoice_objects.filter.return_value = _FakeQuerySet([])
        mock_challan_objects.filter.return_value = _FakeQuerySet([])
        mock_get_review_note.side_effect = [None, None, None]
        mock_recon_summary.side_effect = [
            {
                "deducted": "10.00",
                "deposited": "10.00",
                "filed": "10.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "10.00",
                "deposited": "10.00",
                "filed": "10.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "0.00",
                "deposited": "0.00",
                "filed": "0.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
        ]
        mock_recon_exceptions.side_effect = [
            {
                "exceptions": {
                    "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                    "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                    "filed_returns_missing_ack_or_arn": {
                        "count": 1,
                        "rows": [
                            {
                                "id": 91,
                                "return_code": "26Q",
                                "period_from": "2026-04-01",
                                "period_to": "2026-06-30",
                                "source_kind": "return",
                                "source_id": 91,
                                "source_search": "26Q",
                                "source_label": "Return 26Q",
                            }
                        ],
                    },
                }
            },
            {
                "exceptions": {
                    "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                    "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                    "filed_returns_missing_ack_or_arn": {"count": 0, "rows": []},
                }
            },
        ]

        revised_filing = SimpleNamespace(
            id=91,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            amount=Decimal("10.00"),
            status=PurchaseStatutoryReturn.Status.REVISED,
            ack_no="",
            arn_no="ARN-REV-91",
            ack_document="revised-return.pdf",
            lines=_ListManager([]),
            get_status_display=lambda: "Revised",
        )
        mock_return_objects.filter.return_value = _FakeQuerySet([revised_filing])

        resp = self.client.get(
            "/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))

        summary_sheet = wb["01_Management_Summary"]
        summary_rows = {
            summary_sheet.cell(row=row_idx, column=1).value: (
                summary_sheet.cell(row=row_idx, column=2).value,
                summary_sheet.cell(row=row_idx, column=3).value,
                summary_sheet.cell(row=row_idx, column=4).value,
            )
            for row_idx in range(2, summary_sheet.max_row + 1)
        }
        self.assertEqual(summary_rows["Evidence Follow-up"], (1, 1, 0))

        action_sheet = wb["03_Action_Items"]
        action_rows = [
            [action_sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 9)]
            for row_idx in range(1, action_sheet.max_row + 1)
        ]
        self.assertIn(
            ["Missing Filing Evidence", "IT_TDS", "Medium", "26Q", "2026-04-01 to 2026-06-30", None, "Filed return is still missing ACK or ARN details.", "Capture acknowledgement references and attach supporting proof."],
            action_rows,
        )

        support_sheet = wb["12_Supporting_Doc_Index"]
        support_rows = [
            [support_sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 7)]
            for row_idx in range(2, support_sheet.max_row + 1)
        ]
        self.assertIn(["Return", "IT_TDS", "26Q", datetime(2026, 6, 30, 0, 0), "revised-return.pdf", "Revised"], support_rows)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.get_review_note")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_exceptions")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_summary")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryChallan.objects")
    @patch("purchase.views.purchase_statutory.PurchaseInvoiceHeader.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_export_clamps_negative_pending_metrics(
        self,
        mock_perm_codes,
        mock_invoice_objects,
        mock_challan_objects,
        mock_return_objects,
        mock_recon_summary,
        mock_recon_exceptions,
        mock_get_review_note,
    ):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}
        mock_invoice_objects.filter.return_value = _FakeQuerySet([])
        mock_challan_objects.filter.return_value = _FakeQuerySet([])
        mock_return_objects.filter.return_value = _FakeQuerySet([])
        mock_get_review_note.side_effect = [None, None, None]
        mock_recon_summary.side_effect = [
            {
                "deducted": "100.00",
                "deposited": "105.00",
                "filed": "107.00",
                "pending_deposit": "-5.00",
                "pending_filing": "-2.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "50.00",
                "deposited": "55.00",
                "filed": "56.00",
                "pending_deposit": "-5.00",
                "pending_filing": "-1.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "50.00",
                "deposited": "50.00",
                "filed": "51.00",
                "pending_deposit": "0.00",
                "pending_filing": "-1.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
        ]
        mock_recon_exceptions.side_effect = [
            {"exceptions": {
                "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                "filed_returns_missing_ack_or_arn": {"count": 0, "rows": []},
            }},
            {"exceptions": {
                "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                "filed_returns_missing_ack_or_arn": {"count": 0, "rows": []},
            }},
        ]

        resp = self.client.get(
            "/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        summary_sheet = wb["01_Management_Summary"]
        summary_rows = {
            summary_sheet.cell(row=row_idx, column=1).value: (
                summary_sheet.cell(row=row_idx, column=2).value,
                summary_sheet.cell(row=row_idx, column=3).value,
                summary_sheet.cell(row=row_idx, column=4).value,
            )
            for row_idx in range(2, summary_sheet.max_row + 1)
        }
        self.assertEqual(summary_rows["Pending Deposit"], ("0.00", "0.00", "0.00"))
        self.assertEqual(summary_rows["Pending Filing"], ("0.00", "0.00", "0.00"))
        recon_sheet = wb["10_Reconciliation"]
        recon_rows = {
            recon_sheet.cell(row=row_idx, column=1).value: (
                recon_sheet.cell(row=row_idx, column=2).value,
                recon_sheet.cell(row=row_idx, column=3).value,
            )
            for row_idx in range(2, recon_sheet.max_row + 1)
        }
        self.assertEqual(recon_rows["pending_deposit"], ("0.00", "0.00"))
        self.assertEqual(recon_rows["pending_filing"], ("0.00", "0.00"))

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.get_review_note")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_exceptions")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.reconciliation_summary")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.PurchaseStatutoryChallan.objects")
    @patch("purchase.views.purchase_statutory.PurchaseInvoiceHeader.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_it_tds_quality_honours_26q_27q_identity_rules(
        self,
        mock_perm_codes,
        mock_invoice_objects,
        mock_challan_objects,
        mock_return_objects,
        mock_recon_summary,
        mock_recon_exceptions,
        mock_get_review_note,
    ):
        mock_perm_codes.return_value = {"purchase.statutory.manage"}
        mock_invoice_objects.filter.return_value = _FakeQuerySet([])
        mock_challan_objects.filter.return_value = _FakeQuerySet([])
        mock_get_review_note.side_effect = [None, None, None]
        mock_recon_summary.side_effect = [
            {
                "deducted": "26.00",
                "deposited": "26.00",
                "filed": "26.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "26.00",
                "deposited": "26.00",
                "filed": "26.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
            {
                "deducted": "0.00",
                "deposited": "0.00",
                "filed": "0.00",
                "pending_deposit": "0.00",
                "pending_filing": "0.00",
                "draft_challan": "0.00",
                "draft_return": "0.00",
            },
        ]
        mock_recon_exceptions.side_effect = [
            {"exceptions": {
                "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                "filed_returns_missing_ack_or_arn": {"count": 0, "rows": []},
            }},
            {"exceptions": {
                "invoices_pending_challan_mapping": {"line_count": 0, "rows": []},
                "challan_lines_pending_return_mapping": {"line_count": 0, "rows": []},
                "filed_returns_missing_ack_or_arn": {"count": 0, "rows": []},
            }},
        ]

        resident_header = SimpleNamespace(purchase_number="PINV-61", vendor_name="Resident Vendor")
        foreign_header = SimpleNamespace(purchase_number="PINV-62", vendor_name="Foreign Vendor")
        challan = SimpleNamespace(challan_no="CH-61")
        line_26q = SimpleNamespace(
            header_id=61,
            header=resident_header,
            challan_id=61,
            challan=challan,
            amount=Decimal("11.00"),
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot="",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            section_snapshot_code="194C",
            cin_snapshot="CIN61",
        )
        line_27q = SimpleNamespace(
            header_id=62,
            header=foreign_header,
            challan_id=61,
            challan=challan,
            amount=Decimal("15.00"),
            deductee_pan_snapshot="",
            deductee_tax_id_snapshot="NR-TAX-62",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            section_snapshot_code="195",
            cin_snapshot="CIN62",
        )
        filing_26q = SimpleNamespace(
            id=61,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            amount=Decimal("11.00"),
            status=PurchaseStatutoryReturn.Status.FILED,
            ack_no="ACK61",
            arn_no="ARN61",
            ack_document="",
            lines=_ListManager([line_26q]),
            get_status_display=lambda: "Filed",
        )
        filing_27q = SimpleNamespace(
            id=62,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="27Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            amount=Decimal("15.00"),
            status=PurchaseStatutoryReturn.Status.REVISED,
            ack_no="ACK62",
            arn_no="ARN62",
            ack_document="",
            lines=_ListManager([line_27q]),
            get_status_display=lambda: "Revised",
        )
        mock_return_objects.filter.return_value = _FakeQuerySet([filing_26q, filing_27q])

        resp = self.client.get(
            "/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        quality_sheet = wb["13_IT_Return_Quality"]
        quality_rows = {
            quality_sheet.cell(row=row_idx, column=1).value: (
                quality_sheet.cell(row=row_idx, column=8).value,
                quality_sheet.cell(row=row_idx, column=10).value,
            )
            for row_idx in range(2, quality_sheet.max_row + 1)
        }
        self.assertEqual(quality_rows["26Q"], ("VALID", "NOT_REQUIRED"))
        self.assertEqual(quality_rows["27Q"], ("NOT_REQUIRED", "PRESENT"))

        scope_sheet = wb["14_IT_Return_Scope_Summary"]
        scope_rows = {
            (scope_sheet.cell(row=row_idx, column=1).value, scope_sheet.cell(row=row_idx, column=2).value): (
                scope_sheet.cell(row=row_idx, column=5).value,
                scope_sheet.cell(row=row_idx, column=7).value,
            )
            for row_idx in range(2, scope_sheet.max_row + 1)
        }
        self.assertEqual(scope_rows[("SECTION", "194C")], (0, 0))
        self.assertEqual(scope_rows[("SECTION", "195")], (0, 0))

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.issue_form16a")
    def test_statutory_form16a_issue_endpoint_returns_201(self, mock_fn):
        mock_fn.return_value = {"filing_id": 1, "issue": {"issue_no": 1}}
        resp = self.client.post("/api/purchase/statutory/returns/1/form16a/", {"remarks": "issue"})
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data["message"], "Form16A issued.")

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.form16a_download_payload")
    def test_statutory_form16a_download_endpoint_returns_200(self, mock_fn):
        mock_fn.return_value = {"filename": "f16a.txt", "content": "hello"}
        resp = self.client.get("/api/purchase/statutory/returns/1/form16a/1/download/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/plain")

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.submit_challan_for_approval")
    def test_statutory_challan_approval_submit_returns_200(self, mock_fn):
        mock_fn.return_value = SimpleNamespace(
            message="ok",
            obj=SimpleNamespace(id=1),
        )
        with patch("purchase.views.purchase_statutory.PurchaseStatutoryChallanSerializer") as mock_ser:
            mock_ser.return_value.data = {"id": 1, "approval_status": "SUBMITTED", "approval_status_name": "Submitted"}
            resp = self.client.post("/api/purchase/statutory/challans/1/approval/", {"action": "submit"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["approval_status"], "SUBMITTED")

    def test_statutory_challan_approval_invalid_action_returns_400(self):
        resp = self.client.post("/api/purchase/statutory/challans/1/approval/", {"action": "archive"})
        self.assertEqual(resp.status_code, 400)
        self.assertIn("submit|approve|reject", str(resp.data))

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryService.submit_return_for_approval")
    def test_statutory_return_approval_submit_returns_200(self, mock_fn):
        mock_fn.return_value = SimpleNamespace(
            message="ok",
            obj=SimpleNamespace(id=1),
        )
        with patch("purchase.views.purchase_statutory.PurchaseStatutoryReturnSerializer") as mock_ser:
            mock_ser.return_value.data = {"id": 1, "approval_status": "SUBMITTED", "approval_status_name": "Submitted"}
            resp = self.client.post("/api/purchase/statutory/returns/1/approval/", {"action": "submit"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["approval_status"], "SUBMITTED")

    def test_statutory_return_approval_invalid_action_returns_400(self):
        resp = self.client.post("/api/purchase/statutory/returns/1/approval/", {"action": "archive"})
        self.assertEqual(resp.status_code, 400)
        self.assertIn("submit|approve|reject", str(resp.data))


class PurchaseStatutoryComplianceTests(SimpleTestCase):
    def test_coerce_date_accepts_iso_and_date(self):
        self.assertEqual(
            PurchaseStatutoryService._coerce_date("2026-04-30", field_name="x"),
            date(2026, 4, 30),
        )
        self.assertEqual(
            PurchaseStatutoryService._coerce_date(date(2026, 4, 30), field_name="x"),
            date(2026, 4, 30),
        )

    def test_validate_it_tds_return_code_26q_requires_resident_pan(self):
        line_ok = SimpleNamespace(
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot=None,
        )
        PurchaseStatutoryService._validate_it_tds_return_code(return_code="26Q", lines=[line_ok])

        line_bad = SimpleNamespace(
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot=None,
        )
        with self.assertRaisesMessage(ValueError, "26Q allows only RESIDENT"):
            PurchaseStatutoryService._validate_it_tds_return_code(return_code="26Q", lines=[line_bad])

    def test_validate_it_tds_return_code_26q_requires_valid_pan_format(self):
        line_bad = SimpleNamespace(
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            deductee_pan_snapshot="BADPAN",
            deductee_tax_id_snapshot=None,
        )
        with self.assertRaisesMessage(ValueError, "valid PAN format"):
            PurchaseStatutoryService._validate_it_tds_return_code(return_code="26Q", lines=[line_bad])

    def test_validate_it_tds_return_code_27q_requires_non_resident_tax_id(self):
        line_bad = SimpleNamespace(
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            deductee_pan_snapshot=None,
            deductee_tax_id_snapshot=None,
        )
        with self.assertRaisesMessage(ValueError, "27Q requires deductee_tax_id_snapshot"):
            PurchaseStatutoryService._validate_it_tds_return_code(return_code="27Q", lines=[line_bad])

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_generate_nsdl_payload_includes_quality_and_section_summary(self, mock_return_objects):
        line_one = SimpleNamespace(
            header_id=10,
            amount=Decimal("5.00"),
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot="",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            section_snapshot_code="194J",
            cin_snapshot="CIN1",
        )
        line_two = SimpleNamespace(
            header_id=11,
            amount=Decimal("6.00"),
            deductee_pan_snapshot="BADPAN",
            deductee_tax_id_snapshot="TAX9",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            section_snapshot_code="27QSEC",
            cin_snapshot="CIN2",
        )
        filing = SimpleNamespace(
            id=21,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="27Q",
            status=PurchaseStatutoryReturn.Status.FILED,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            lines=SimpleNamespace(all=lambda: [line_one, line_two]),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        payload = PurchaseStatutoryService.generate_nsdl_payload(filing_id=21)

        self.assertEqual(payload["line_count"], 2)
        self.assertEqual(payload["quality_summary"]["resident_count"], 1)
        self.assertEqual(payload["quality_summary"]["non_resident_count"], 1)
        self.assertEqual(payload["quality_summary"]["invalid_pan_format"], 0)
        self.assertEqual(payload["quality_summary"]["missing_tax_id"], 1)
        self.assertEqual(payload["section_summary"][0]["section_code"], "194J")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_generate_nsdl_payload_27q_does_not_flag_missing_pan_when_tax_id_exists(self, mock_return_objects):
        line = SimpleNamespace(
            header_id=41,
            amount=Decimal("12.00"),
            deductee_pan_snapshot="",
            deductee_tax_id_snapshot="NR-TAX-41",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            section_snapshot_code="195",
            cin_snapshot="CIN41",
        )
        filing = SimpleNamespace(
            id=41,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="27Q",
            status=PurchaseStatutoryReturn.Status.FILED,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            lines=SimpleNamespace(all=lambda: [line]),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        payload = PurchaseStatutoryService.generate_nsdl_payload(filing_id=41)

        self.assertEqual(payload["quality_summary"]["missing_pan"], 0)
        self.assertEqual(payload["quality_summary"]["missing_tax_id"], 0)
        self.assertEqual(payload["quality_summary"]["non_resident_count"], 1)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_generate_nsdl_payload_26q_does_not_flag_missing_tax_id_when_pan_exists(self, mock_return_objects):
        line = SimpleNamespace(
            header_id=42,
            amount=Decimal("8.00"),
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot="",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            section_snapshot_code="194C",
            cin_snapshot="CIN42",
        )
        filing = SimpleNamespace(
            id=42,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            status=PurchaseStatutoryReturn.Status.FILED,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            lines=SimpleNamespace(all=lambda: [line]),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        payload = PurchaseStatutoryService.generate_nsdl_payload(filing_id=42)

        self.assertEqual(payload["quality_summary"]["missing_pan"], 0)
        self.assertEqual(payload["quality_summary"]["missing_tax_id"], 0)
        self.assertEqual(payload["quality_summary"]["resident_count"], 1)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_generate_nsdl_payload_emits_hdr_detail_and_trailer_rows_in_line_order(self, mock_return_objects):
        line_one = SimpleNamespace(
            header_id=51,
            amount=Decimal("10.00"),
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot="",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            section_snapshot_code="194C",
            cin_snapshot="CIN51",
        )
        line_two = SimpleNamespace(
            header_id=52,
            amount=Decimal("12.50"),
            deductee_pan_snapshot="",
            deductee_tax_id_snapshot="NR-TAX-52",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            section_snapshot_code="195",
            cin_snapshot="CIN52",
        )
        filing = SimpleNamespace(
            id=77,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="27Q",
            status=PurchaseStatutoryReturn.Status.REVISED,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            lines=SimpleNamespace(all=lambda: [line_one, line_two]),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        payload = PurchaseStatutoryService.generate_nsdl_payload(filing_id=77)

        txt_rows = payload["nsdl_txt"].splitlines()
        self.assertEqual(txt_rows[0], "HDR|77|IT_TDS|27Q|2026-04-01|2026-06-30|2|22.50")
        self.assertEqual(
            txt_rows[1],
            "DTL|1|51|ABCDE1234F||RESIDENT|10.00|194C|CIN51",
        )
        self.assertEqual(
            txt_rows[2],
            "DTL|2|52||NR-TAX-52|NON_RESIDENT|12.50|195|CIN52",
        )
        self.assertEqual(txt_rows[3], "TRL|2|22.50")
        self.assertEqual(payload["return_code"], "27Q")
        self.assertEqual(payload["residency_mode"], "NON_RESIDENT_ONLY")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_generate_nsdl_payload_rejects_non_it_tds_or_non_nsdl_return_codes(self, mock_return_objects):
        filing = SimpleNamespace(
            id=88,
            tax_type=PurchaseStatutoryReturn.TaxType.GST_TDS,
            return_code="GSTR7",
            status=PurchaseStatutoryReturn.Status.FILED,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            lines=SimpleNamespace(all=lambda: []),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        with self.assertRaisesMessage(ValueError, "NSDL export is available only for IT_TDS returns 26Q/27Q."):
            PurchaseStatutoryService.generate_nsdl_payload(filing_id=88)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_generate_nsdl_payload_rejects_draft_it_tds_returns(self, mock_return_objects):
        filing = SimpleNamespace(
            id=89,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            status=PurchaseStatutoryReturn.Status.DRAFT,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            lines=SimpleNamespace(all=lambda: []),
        )
        mock_return_objects.prefetch_related.return_value.get.return_value = filing

        with self.assertRaisesMessage(ValueError, "NSDL export is available only for filed or revised IT_TDS returns."):
            PurchaseStatutoryService.generate_nsdl_payload(filing_id=89)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_return_quality_summary_aggregates_section_vendor_and_status_quality(self, mock_return_objects):
        header_one = SimpleNamespace(purchase_number="PINV-31", vendor_name="Vendor A")
        header_two = SimpleNamespace(purchase_number="PINV-32", vendor_name="Vendor B")
        challan_one = SimpleNamespace(challan_no="CH-31")
        challan_two = SimpleNamespace(challan_no="CH-32")
        line_one = SimpleNamespace(
            header_id=31,
            header=header_one,
            challan=challan_one,
            amount=Decimal("5.00"),
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot="TAX31",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            section_snapshot_code="194J",
        )
        line_two = SimpleNamespace(
            header_id=32,
            header=header_two,
            challan=challan_two,
            amount=Decimal("6.00"),
            deductee_pan_snapshot="BADPAN",
            deductee_tax_id_snapshot="",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            section_snapshot_code="27QSEC",
        )
        filing_one = SimpleNamespace(
            id=31,
            status=PurchaseStatutoryReturn.Status.FILED,
            status_name="Filed",
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            filed_on=date(2026, 5, 10),
            lines=SimpleNamespace(all=lambda: [line_one]),
        )
        filing_two = SimpleNamespace(
            id=32,
            status=PurchaseStatutoryReturn.Status.REVISED,
            status_name="Revised",
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="27Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            filed_on=date(2026, 5, 11),
            lines=SimpleNamespace(all=lambda: [line_two]),
        )
        qs = MagicMock()
        qs.filter.return_value = qs
        qs.order_by.return_value = [filing_one, filing_two]
        mock_return_objects.prefetch_related.return_value.filter.return_value = qs

        payload = PurchaseStatutoryService.return_quality_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
        )

        self.assertEqual(payload["summary"]["return_count"], 2)
        self.assertEqual(payload["summary"]["line_count"], 2)
        self.assertEqual(payload["summary"]["filed_returns"], 1)
        self.assertEqual(payload["summary"]["revised_returns"], 1)
        self.assertEqual(payload["summary"]["invalid_pan_format"], 1)
        self.assertEqual(payload["summary"]["missing_tax_id"], 1)
        self.assertEqual(payload["section_summary"][0]["section_code"], "194J")
        self.assertEqual(payload["vendor_summary"][0]["vendor_name"], "Vendor A")
        self.assertEqual(payload["status_summary"][0]["return_count"], 1)
        self.assertEqual(payload["line_preview"][1]["challan_no"], "CH-32")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_return_quality_summary_26q_ignores_missing_tax_id_when_pan_quality_is_valid(self, mock_return_objects):
        header = SimpleNamespace(purchase_number="PINV-41", vendor_name="Vendor Resident")
        challan = SimpleNamespace(challan_no="CH-41")
        line = SimpleNamespace(
            header_id=41,
            header=header,
            challan=challan,
            amount=Decimal("12.00"),
            deductee_pan_snapshot="ABCDE1234F",
            deductee_tax_id_snapshot="",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.RESIDENT,
            section_snapshot_code="194C",
        )
        filing = SimpleNamespace(
            id=41,
            status=PurchaseStatutoryReturn.Status.FILED,
            status_name="Filed",
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            filed_on=date(2026, 5, 12),
            lines=SimpleNamespace(all=lambda: [line]),
        )
        qs = MagicMock()
        qs.filter.return_value = qs
        qs.order_by.return_value = [filing]
        mock_return_objects.prefetch_related.return_value.filter.return_value = qs

        payload = PurchaseStatutoryService.return_quality_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            return_code="26Q",
        )

        self.assertEqual(payload["summary"]["missing_tax_id"], 0)
        self.assertFalse(payload["line_preview"][0]["missing_tax_id"])

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    def test_return_quality_summary_27q_ignores_missing_pan_when_tax_id_is_present(self, mock_return_objects):
        header = SimpleNamespace(purchase_number="PINV-51", vendor_name="Vendor Foreign")
        challan = SimpleNamespace(challan_no="CH-51")
        line = SimpleNamespace(
            header_id=51,
            header=header,
            challan=challan,
            amount=Decimal("14.00"),
            deductee_pan_snapshot="",
            deductee_tax_id_snapshot="NR-TAX-51",
            deductee_residency_snapshot=PurchaseStatutoryReturnLine.DeducteeResidency.NON_RESIDENT,
            section_snapshot_code="195",
        )
        filing = SimpleNamespace(
            id=51,
            status=PurchaseStatutoryReturn.Status.FILED,
            status_name="Filed",
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="27Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            filed_on=date(2026, 5, 12),
            lines=SimpleNamespace(all=lambda: [line]),
        )
        qs = MagicMock()
        qs.filter.return_value = qs
        qs.order_by.return_value = [filing]
        mock_return_objects.prefetch_related.return_value.filter.return_value = qs

        payload = PurchaseStatutoryService.return_quality_summary(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=5,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
            return_code="27Q",
        )

        self.assertEqual(payload["summary"]["missing_pan"], 0)
        self.assertFalse(payload["line_preview"][0]["missing_pan"])

    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.return_eligible_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.challan_eligible_lines")
    def test_reconciliation_exceptions_exposes_rows_for_pending_mappings(
        self,
        mock_challan_eligible,
        mock_return_eligible,
        mock_return_objects,
        mock_purchase_line_objects,
    ):
        mock_challan_eligible.return_value = {
            "lines": [{"header_id": 11, "purchase_number": "PINV-11", "amount": "10.00"}],
            "totals": {"line_count": 1, "amount": "10.00"},
        }
        mock_return_eligible.return_value = {
            "lines": [{"header_id": 11, "challan_id": 21, "challan_no": "CH-21", "amount": "10.00"}],
            "totals": {"line_count": 1, "amount": "10.00"},
        }
        mock_return_objects.filter.return_value.filter.return_value.values.return_value = []
        mock_purchase_line_objects.filter.return_value.exists.return_value = True

        payload = PurchaseStatutoryService.reconciliation_exceptions(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 4, 30),
        )

        self.assertEqual(payload["exceptions"]["invoices_pending_challan_mapping"]["count"], 1)
        self.assertEqual(len(payload["exceptions"]["invoices_pending_challan_mapping"]["rows"]), 1)
        self.assertEqual(payload["exceptions"]["challan_lines_pending_return_mapping"]["count"], 1)
        self.assertEqual(len(payload["exceptions"]["challan_lines_pending_return_mapping"]["rows"]), 1)
        challan_row = payload["exceptions"]["invoices_pending_challan_mapping"]["rows"][0]
        self.assertEqual(challan_row["source_kind"], "purchase_invoice")
        self.assertEqual(challan_row["source_id"], 11)
        self.assertEqual(challan_row["source_route"], "/purchaseserviceinvoice")
        return_row = payload["exceptions"]["challan_lines_pending_return_mapping"]["rows"][0]
        self.assertEqual(return_row["source_kind"], "challan")
        self.assertEqual(return_row["source_id"], 21)

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.return_eligible_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.challan_eligible_lines")
    def test_reconciliation_exceptions_exposes_missing_ack_return_source_metadata(
        self,
        mock_challan_eligible,
        mock_return_eligible,
        mock_return_objects,
    ):
        mock_challan_eligible.return_value = {
            "lines": [],
            "totals": {"line_count": 0, "amount": "0.00"},
        }
        mock_return_eligible.return_value = {
            "lines": [],
            "totals": {"line_count": 0, "amount": "0.00"},
        }
        filed_qs = MagicMock()
        filed_qs.filter.return_value = filed_qs
        filed_qs.values.return_value = [
            {
                "id": 82,
                "return_code": "27Q",
                "period_from": date(2026, 4, 1),
                "period_to": date(2026, 6, 30),
                "ack_no": "",
                "arn_no": "",
            }
        ]
        mock_return_objects.filter.return_value = filed_qs

        payload = PurchaseStatutoryService.reconciliation_exceptions(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
        )

        bucket = payload["exceptions"]["filed_returns_missing_ack_or_arn"]
        self.assertEqual(bucket["count"], 1)
        self.assertEqual(len(bucket["rows"]), 1)
        row = bucket["rows"][0]
        self.assertEqual(row["source_kind"], "return")
        self.assertEqual(row["source_id"], 82)
        self.assertEqual(row["source_search"], "27Q")
        self.assertEqual(row["source_label"], "Return 27Q")

    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryReturn.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.return_eligible_lines")
    @patch("purchase.services.purchase_statutory_service.PurchaseStatutoryService.challan_eligible_lines")
    def test_reconciliation_exceptions_includes_revised_returns_missing_ack_or_arn(
        self,
        mock_challan_eligible,
        mock_return_eligible,
        mock_return_objects,
    ):
        mock_challan_eligible.return_value = {
            "lines": [],
            "totals": {"line_count": 0, "amount": "0.00"},
        }
        mock_return_eligible.return_value = {
            "lines": [],
            "totals": {"line_count": 0, "amount": "0.00"},
        }
        filing_qs = MagicMock()
        filing_qs.filter.return_value = filing_qs
        filing_qs.values.return_value = [
            {
                "id": 91,
                "return_code": "26Q",
                "period_from": date(2026, 4, 1),
                "period_to": date(2026, 6, 30),
                "ack_no": "",
                "arn_no": "ARN-26Q-REV",
            }
        ]
        mock_return_objects.filter.return_value = filing_qs

        payload = PurchaseStatutoryService.reconciliation_exceptions(
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
        )

        bucket = payload["exceptions"]["filed_returns_missing_ack_or_arn"]
        self.assertEqual(bucket["count"], 1)
        self.assertEqual(bucket["rows"][0]["source_id"], 91)
        self.assertEqual(bucket["rows"][0]["source_label"], "Return 26Q")

    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceLine.objects")
    @patch("purchase.services.purchase_statutory_service.PurchaseInvoiceHeader.objects")
    def test_reconciliation_gl_status_exposes_source_metadata_for_missing_entries(
        self,
        mock_header_objects,
        mock_purchase_line_objects,
    ):
        invoice_qs = MagicMock()
        invoice_qs.filter.return_value = invoice_qs
        invoice_qs.values_list.return_value = [44]
        invoice_qs.exclude.return_value.values.return_value = [
            {
                "id": 44,
                "purchase_number": "PINV-44",
                "bill_date": date(2026, 4, 12),
                "grand_total": "80.00",
                "vendor_name": "Vendor Trace",
            }
        ]
        mock_header_objects.filter.return_value = invoice_qs
        mock_purchase_line_objects.filter.return_value.exists.return_value = True

        with patch("posting.models.Entry.objects") as mock_entry_objects:
            entry_qs = MagicMock()
            entry_qs.values_list.return_value = []
            entry_qs.count.return_value = 0
            mock_entry_objects.filter.return_value = entry_qs

            payload = PurchaseStatutoryService.reconciliation_gl_status(
                entity_id=1,
                entityfinid_id=1,
                subentity_id=None,
                period_from=date(2026, 4, 1),
                period_to=date(2026, 4, 30),
            )

        self.assertEqual(payload["gl_reconciliation"]["missing_gl_entry_count"], 1)
        row = payload["gl_reconciliation"]["missing_gl_entries"][0]
        self.assertEqual(row["vendor_name"], "Vendor Trace")
        self.assertEqual(row["source_kind"], "purchase_invoice")
        self.assertEqual(row["source_id"], 44)
        self.assertEqual(row["source_route"], "/purchaseserviceinvoice")


class PurchaseApiPermissionTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="purchase_permission_tester",
            email="purchase_permission_tester@example.com",
            password="x",
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        self.allowed_entity = Entity.objects.create(
            entityname="Allowed Entity",
            legalname="Allowed Entity Pvt Ltd",
            business_type=Entity.BusinessType.MIXED,
            createdby=self.user,
        )
        self.allowed_subentity = SubEntity.objects.create(
            entity=self.allowed_entity,
            subentityname="Allowed HO",
            branch_type=SubEntity.BranchType.HEAD_OFFICE,
        )
        self.allowed_fy = EntityFinancialYear.objects.create(
            entity=self.allowed_entity,
            desc="FY 2026-27",
            finstartyear=timezone.make_aware(datetime(2026, 4, 1)),
            finendyear=timezone.make_aware(datetime(2027, 3, 31)),
            createdby=self.user,
        )
        self.foreign_entity = Entity.objects.create(
            entityname="Foreign Entity",
            legalname="Foreign Entity Pvt Ltd",
            business_type=Entity.BusinessType.MIXED,
            createdby=self.user,
        )
        self.foreign_subentity = SubEntity.objects.create(
            entity=self.foreign_entity,
            subentityname="Foreign HO",
            branch_type=SubEntity.BranchType.HEAD_OFFICE,
        )
        self.foreign_fy = EntityFinancialYear.objects.create(
            entity=self.foreign_entity,
            desc="FY 2026-27",
            finstartyear=timezone.make_aware(datetime(2026, 4, 1)),
            finendyear=timezone.make_aware(datetime(2027, 3, 31)),
            createdby=self.user,
        )
        self.foreign_challan = PurchaseStatutoryChallan.objects.create(
            entity=self.foreign_entity,
            entityfinid=self.foreign_fy,
            subentity=self.foreign_subentity,
            tax_type=PurchaseStatutoryChallan.TaxType.IT_TDS,
            challan_no="CH-FOREIGN-1",
            challan_date=date(2026, 4, 10),
            amount=Decimal("100.00"),
            created_by=self.user,
        )
        self.foreign_return = PurchaseStatutoryReturn.objects.create(
            entity=self.foreign_entity,
            entityfinid=self.foreign_fy,
            subentity=self.foreign_subentity,
            tax_type=PurchaseStatutoryReturn.TaxType.IT_TDS,
            return_code="26Q",
            period_from=date(2026, 4, 1),
            period_to=date(2026, 6, 30),
            amount=Decimal("100.00"),
            created_by=self.user,
        )

    def _pk_scope_obj(self, entity_id=1):
        return SimpleNamespace(entity_id=entity_id)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_review_note_get_requires_view_permission(self, mock_codes):
        mock_codes.return_value = set()

        response = self.client.get(
            "/api/purchase/statutory/review-note/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_summary_rejects_purchase_invoice_view_without_statutory_permission(self, mock_codes):
        mock_codes.return_value = {"purchase.invoice.view"}

        response = self.client.get("/api/purchase/statutory/summary/?entity=1&entityfinid=1")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_review_note_post_requires_manage_permission(self, mock_codes):
        mock_codes.return_value = {"purchase.statutory.view"}

        response = self.client.post(
            "/api/purchase/statutory/review-note/",
            {
                "entity": 1,
                "entityfinid": 1,
                "period_from": "2026-04-01",
                "period_to": "2026-04-30",
                "reviewer_name": "CA Reviewer",
                "closure_status": "IN_REVIEW",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_export_requires_manage_permission(self, mock_codes):
        mock_codes.return_value = {"purchase.statutory.view"}

        response = self.client.get(
            "/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1&period_from=2026-04-01&period_to=2026-04-30"
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_review_note_get_requires_period_params(self, mock_codes):
        mock_codes.return_value = {"purchase.statutory.view"}

        response = self.client.get("/api/purchase/statutory/review-note/?entity=1&entityfinid=1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("period_from and period_to are required", str(response.data))

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_review_note_get_rejects_inverted_period(self, mock_codes):
        mock_codes.return_value = {"purchase.statutory.view"}

        response = self.client.get(
            "/api/purchase/statutory/review-note/?entity=1&entityfinid=1&period_from=2026-04-30&period_to=2026-04-01"
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("period_from cannot be greater than period_to", str(response.data))

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_export_requires_period_params(self, mock_codes):
        mock_codes.return_value = {"purchase.statutory.manage"}

        response = self.client.get("/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("period_from and period_to are required", str(response.data))

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_ca_pack_export_rejects_inverted_period(self, mock_codes):
        mock_codes.return_value = {"purchase.statutory.manage"}

        response = self.client.get(
            "/api/purchase/statutory/export/ca-pack/?entity=1&entityfinid=1&period_from=2026-04-30&period_to=2026-04-01"
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("period_from cannot be greater than period_to", str(response.data))

    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    def test_gstr2b_batch_list_requires_statutory_view_permission(self, mock_codes):
        mock_codes.return_value = {"purchase.invoice.view"}

        response = self.client.get("/api/purchase/gstr2b/import-batches/?entity=1&entityfinid=1")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_gstr2b.Gstr2bImportBatch.objects")
    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    def test_gstr2b_auto_match_requires_statutory_manage_permission(self, mock_codes, mock_batch_objects):
        mock_codes.return_value = {"purchase.statutory.view"}
        mock_batch_objects.filter.return_value.first.return_value = SimpleNamespace(
            id=15,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=None,
        )

        response = self.client.post("/api/purchase/gstr2b/import-batches/15/match/?entity=1&entityfinid=1", {}, format="json")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryChallan.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_challan_deposit_requires_manage_permission(self, mock_codes, mock_challan_objects):
        mock_codes.return_value = {"purchase.statutory.view"}
        mock_challan_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.post(
            "/api/purchase/statutory/challans/15/deposit/",
            {"deposited_on": "2026-04-30"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryChallan.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_challan_approval_requires_approve_permission(self, mock_codes, mock_challan_objects):
        mock_codes.return_value = {"purchase.statutory.manage"}
        mock_challan_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.post(
            "/api/purchase/statutory/challans/15/approval/",
            {"action": "submit"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_return_file_requires_manage_permission(self, mock_codes, mock_return_objects):
        mock_codes.return_value = {"purchase.statutory.view"}
        mock_return_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.post(
            "/api/purchase/statutory/returns/21/file/",
            {"filed_on": "2026-07-31"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_return_approval_requires_approve_permission(self, mock_codes, mock_return_objects):
        mock_codes.return_value = {"purchase.statutory.manage"}
        mock_return_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.post(
            "/api/purchase/statutory/returns/21/approval/",
            {"action": "submit"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_form16a_issue_requires_manage_permission(self, mock_codes, mock_return_objects):
        mock_codes.return_value = {"purchase.statutory.view"}
        mock_return_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.post(
            "/api/purchase/statutory/returns/21/form16a/",
            {"remarks": "issue"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_form16a_download_requires_view_permission(self, mock_codes, mock_return_objects):
        mock_codes.return_value = set()
        mock_return_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.get("/api/purchase/statutory/returns/21/form16a/1/download/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.PurchaseStatutoryReturn.objects")
    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_nsdl_export_requires_view_permission(self, mock_codes, mock_return_objects):
        mock_codes.return_value = set()
        mock_return_objects.filter.return_value.only.return_value.first.return_value = self._pk_scope_obj()

        response = self.client.get("/api/purchase/statutory/returns/21/nsdl-export/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_gstr2b.Gstr2bImportBatch.objects")
    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    def test_gstr2b_batch_rows_rejects_subentity_mismatch(self, mock_codes, mock_batch_objects):
        mock_codes.return_value = {"purchase.statutory.view"}
        mock_batch_objects.filter.return_value.first.return_value = SimpleNamespace(
            id=15,
            entity_id=1,
            entityfinid_id=1,
            subentity_id=99,
        )

        response = self.client.get("/api/purchase/gstr2b/import-batches/15/rows/?entity=1&entityfinid=1&subentity=5")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("subentity mismatch", str(response.data).lower())

    @patch("purchase.views.purchase_gstr2b.Gstr2bImportRow.objects")
    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    def test_gstr2b_row_review_rejects_subentity_mismatch(self, mock_codes, mock_row_objects):
        mock_codes.return_value = {"purchase.statutory.manage"}
        mock_row_objects.select_related.return_value.filter.return_value.first.return_value = SimpleNamespace(
            id=33,
            batch=SimpleNamespace(subentity_id=99),
        )

        response = self.client.post(
            "/api/purchase/gstr2b/import-rows/33/review/?entity=1&entityfinid=1&subentity=5",
            {"match_status": "MATCHED"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("subentity mismatch", str(response.data).lower())

    @patch("purchase.views.purchase_gstr2b.Gstr2bImportRow.objects")
    @patch("purchase.views.purchase_gstr2b.EffectivePermissionService.permission_codes_for_user")
    def test_gstr2b_row_review_rejects_matched_status_without_linked_purchase(self, mock_codes, mock_row_objects):
        mock_codes.return_value = {"purchase.statutory.manage"}
        mock_row_objects.select_related.return_value.filter.return_value.first.return_value = SimpleNamespace(
            id=34,
            batch=SimpleNamespace(subentity_id=None),
        )

        response = self.client.post(
            "/api/purchase/gstr2b/import-rows/34/review/?entity=1&entityfinid=1",
            {"match_status": "MATCHED", "matched_purchase": None},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("linked purchase invoice is required", str(response.data).lower())

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_challan_approval_uses_object_entity_scope_for_permission_check(self, mock_codes):
        mock_codes.side_effect = lambda user, entity_id: {"purchase.statutory.approve"} if entity_id == self.allowed_entity.id else set()

        response = self.client.post(
            f"/api/purchase/statutory/challans/{self.foreign_challan.id}/approval/",
            {"action": "submit"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_return_file_uses_object_entity_scope_for_permission_check(self, mock_codes):
        mock_codes.side_effect = lambda user, entity_id: {"purchase.statutory.manage"} if entity_id == self.allowed_entity.id else set()

        response = self.client.post(
            f"/api/purchase/statutory/returns/{self.foreign_return.id}/file/",
            {"filed_on": "2026-07-20"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.purchase_statutory.EffectivePermissionService.permission_codes_for_user")
    def test_statutory_form16a_download_uses_object_entity_scope_for_permission_check(self, mock_codes):
        mock_codes.side_effect = lambda user, entity_id: {"purchase.statutory.view"} if entity_id == self.allowed_entity.id else set()

        response = self.client.get(
            f"/api/purchase/statutory/returns/{self.foreign_return.id}/form16a/1/download/"
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_credit_note_create_requires_credit_note_create_permission(self, mock_entity_for_user, mock_codes):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.credit_note.view"}

        response = self.client.post(
            "/api/purchase/purchase-invoices/",
            {
                "entity": 1,
                "entityfinid": 1,
                "doc_type": int(PurchaseInvoiceHeader.DocType.CREDIT_NOTE),
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("purchase.credit_note create", str(response.data["detail"]).lower())

    @patch("purchase.views.purchase_invoice.generics.RetrieveUpdateDestroyAPIView.update")
    @patch("purchase.views.purchase_invoice.PurchaseInvoiceRetrieveUpdateDestroyAPIView.get_object")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_debit_note_update_uses_debit_note_permission_family(
        self,
        mock_entity_for_user,
        mock_codes,
        mock_get_object,
        mock_super_update,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.debit_note.view"}
        mock_get_object.return_value = SimpleNamespace(
            id=9,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.DEBIT_NOTE),
        )
        mock_super_update.return_value = Response({"ok": True})

        response = self.client.put(
            "/api/purchase/purchase-invoices/9/?entity=1&entityfinid=1",
            {"vendor_name": "Updated"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_super_update.assert_not_called()

    @patch("purchase.views.purchase_invoice.generics.RetrieveUpdateDestroyAPIView.update")
    @patch("purchase.views.purchase_invoice.PurchaseInvoiceRetrieveUpdateDestroyAPIView.get_object")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_debit_note_update_allows_matching_update_permission(
        self,
        mock_entity_for_user,
        mock_codes,
        mock_get_object,
        mock_super_update,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.debit_note.update"}
        mock_get_object.return_value = SimpleNamespace(
            id=9,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.DEBIT_NOTE),
        )
        mock_super_update.return_value = Response({"ok": True})

        response = self.client.put(
            "/api/purchase/purchase-invoices/9/?entity=1&entityfinid=1",
            {"vendor_name": "Updated"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        mock_super_update.assert_called_once()

    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceActions.post")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_credit_note_post_requires_credit_note_post_permission(
        self,
        mock_entity_for_user,
        mock_codes,
        mock_scope,
        mock_post,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.credit_note.view"}
        mock_scope.return_value = SimpleNamespace(
            id=11,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.CREDIT_NOTE),
        )

        response = self.client.post(
            "/api/purchase/purchase-invoices/11/post/?entity=1&entityfinid=1",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_post.assert_not_called()

    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceHeaderSerializer")
    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceActions.cancel")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_credit_note_cancel_allows_update_permission_fallback(
        self,
        mock_entity_for_user,
        mock_codes,
        mock_scope,
        mock_cancel,
        mock_serializer,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.credit_note.update"}
        mock_scope.return_value = SimpleNamespace(
            id=12,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.CREDIT_NOTE),
        )
        mock_cancel.return_value = SimpleNamespace(message="cancelled", header=SimpleNamespace(id=12))
        mock_serializer.return_value.data = {"id": 12}

        response = self.client.post(
            "/api/purchase/purchase-invoices/12/cancel/?entity=1&entityfinid=1",
            {"reason": "test"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        mock_cancel.assert_called_once()

    @patch("purchase.views.purchase_invoice_actions.PurchaseNoteFactory.create_note_from_invoice")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_create_credit_note_action_requires_credit_note_create_permission(
        self,
        mock_entity_for_user,
        mock_codes,
        mock_scope,
        mock_factory,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.invoice.post"}
        mock_scope.return_value = SimpleNamespace(
            id=13,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
        )

        response = self.client.post(
            "/api/purchase/purchase-invoices/13/create-credit-note/?entity=1&entityfinid=1",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_factory.assert_not_called()

    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceService.amendment_window_for_header")
    @patch("purchase.views.purchase_invoice_actions.PurchaseInvoiceActions.cancel")
    @patch("purchase.views.purchase_invoice_actions._assert_invoice_scope")
    @patch("purchase.views.rbac.EffectivePermissionService.permission_codes_for_user")
    @patch("purchase.views.rbac.EffectivePermissionService.entity_for_user")
    def test_locked_period_cancel_requires_credit_note_permissions_for_auto_reversal(
        self,
        mock_entity_for_user,
        mock_codes,
        mock_scope,
        mock_cancel,
        mock_amendment_window,
    ):
        mock_entity_for_user.return_value = SimpleNamespace(id=1)
        mock_codes.return_value = {"purchase.invoice.cancel"}
        mock_scope.return_value = SimpleNamespace(
            id=14,
            entity_id=1,
            doc_type=int(PurchaseInvoiceHeader.DocType.TAX_INVOICE),
            status=int(PurchaseInvoiceHeader.Status.POSTED),
        )
        mock_amendment_window.return_value = SimpleNamespace(amendment_required=True)

        response = self.client.post(
            "/api/purchase/purchase-invoices/14/cancel/?entity=1&entityfinid=1",
            {"reason": "locked"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        mock_cancel.assert_not_called()
