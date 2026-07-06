from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.test import SimpleTestCase
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.test import APIRequestFactory, force_authenticate

from purchase.views.tds_compliance_center import PurchaseTdsComplianceCenterExportAPIView
from purchase.views.tds_compliance_center import PurchaseTdsComplianceCenterAPIView
from purchase.views.gst_tds_compliance_center import (
    PurchaseGstTdsComplianceCenterAPIView,
    PurchaseGstTdsComplianceCenterExportAPIView,
)


MOCK_TDS_EXPORT_PAYLOAD = {
    "tabs": [
        {"id": "deduction-register", "label": "Deduction Register"},
        {"id": "return-filing", "label": "Return Filing"},
    ],
    "returnTabs": [
        {"id": "26q", "label": "26Q Resident Payments"},
        {"id": "27q", "label": "27Q Non-Resident Payments"},
    ],
    "headerChips": [
        {"label": "Arnika G"},
        {"label": "FY 2026-27"},
        {"label": "Q1 Apr-Jun"},
        {"label": "Head Office"},
    ],
    "filters": {
        "quarter": "Q1",
    },
    "meta": {
        "tdsSections": [
            {"id": 9, "section_code": "194C"},
        ],
    },
    "datasets": {
        "deduction-register": {
            "columns": [
                {"key": "date", "label": "Date", "type": "date"},
                {"key": "voucherNo", "label": "Voucher No", "type": "text"},
                {"key": "deductee", "label": "Deductee", "type": "text"},
                {"key": "tdsAmount", "label": "TDS Amount", "type": "currency", "align": "right"},
                {"key": "status", "label": "Status", "type": "status"},
                {"key": "actions", "label": "Actions", "type": "actions"},
            ],
            "rows": [
                {
                    "id": 1,
                    "date": "2026-06-24",
                    "voucherNo": "PI/2026/1226",
                    "deductee": "Vendor-A",
                    "tdsAmount": "50.00",
                    "status": {"label": "Pending Deposit", "tone": "warning"},
                },
                {
                    "id": 2,
                    "date": "2026-06-23",
                    "voucherNo": "PI/2026/1220",
                    "deductee": "Vendor-B",
                    "tdsAmount": "10.00",
                    "status": {"label": "Paid", "tone": "success"},
                },
            ],
        },
    },
    "returnDatasets": {
        "26q": {
            "columns": [
                {"key": "returnType", "label": "Return Type", "type": "text"},
                {"key": "quarter", "label": "Quarter", "type": "text"},
                {"key": "totalTds", "label": "Total TDS", "type": "currency", "align": "right"},
                {"key": "returnStatus", "label": "Status", "type": "status"},
            ],
            "rows": [
                {
                    "id": 1,
                    "returnType": "26Q Resident Payments",
                    "quarter": "Q1 Apr-Jun",
                    "totalTds": "5820.00",
                    "returnStatus": {"label": "Draft", "tone": "warning"},
                }
            ],
        },
        "27q": {
            "columns": [
                {"key": "returnType", "label": "Return Type", "type": "text"},
                {"key": "quarter", "label": "Quarter", "type": "text"},
                {"key": "totalTds", "label": "Total TDS", "type": "currency", "align": "right"},
                {"key": "returnStatus", "label": "Status", "type": "status"},
            ],
            "rows": [],
        },
    },
}


class PurchaseTdsComplianceCenterExportTests(SimpleTestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.user = SimpleNamespace(is_authenticated=True, id=1)

    @patch("purchase.views.tds_compliance_center.PurchaseTdsComplianceCenterAPIView.get")
    def test_xlsx_export_returns_workbook_for_selected_tab(self, mock_workspace_get):
        mock_workspace_get.return_value = Response(MOCK_TDS_EXPORT_PAYLOAD)
        request = self.factory.get(
            "/api/purchase/statutory/tds-compliance-center/export/",
            {
                "entity": 1,
                "entityfinid": 1,
                "tab": "deduction-register",
                "format": "xlsx",
                "columns": "date,voucherNo,tdsAmount,status",
            },
        )
        force_authenticate(request, user=self.user)

        response = PurchaseTdsComplianceCenterExportAPIView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn('attachment; filename="tds_compliance_deduction-register_Q1.xlsx"', response["Content-Disposition"])

        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "TDS Compliance Center - Deduction Register")
        self.assertEqual(sheet["A4"].value, "Date")
        self.assertEqual(sheet["B4"].value, "Voucher No")
        self.assertEqual(sheet["C4"].value, "TDS Amount")
        self.assertEqual(sheet["D4"].value, "Status")
        self.assertEqual(sheet["B5"].value, "PI/2026/1226")
        self.assertEqual(sheet["D5"].value, "Pending Deposit")

    @patch("purchase.views.tds_compliance_center.PurchaseTdsComplianceCenterAPIView.get")
    def test_pdf_export_uses_active_return_subtab(self, mock_workspace_get):
        mock_workspace_get.return_value = Response(MOCK_TDS_EXPORT_PAYLOAD)
        request = self.factory.get(
            "/api/purchase/statutory/tds-compliance-center/export/",
            {
                "entity": 1,
                "entityfinid": 1,
                "tab": "return-filing",
                "return_tab": "26q",
                "format": "pdf",
            },
        )
        force_authenticate(request, user=self.user)

        response = PurchaseTdsComplianceCenterExportAPIView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn('attachment; filename="tds_compliance_return-filing-26q_Q1.pdf"', response["Content-Disposition"])

    @patch("purchase.views.tds_compliance_center.PurchaseTdsComplianceCenterAPIView.get")
    def test_ca_pack_export_includes_cover_and_tab_sheets(self, mock_workspace_get):
        mock_workspace_get.return_value = Response(MOCK_TDS_EXPORT_PAYLOAD)
        request = self.factory.get(
            "/api/purchase/statutory/tds-compliance-center/export/",
            {
                "entity": 1,
                "entityfinid": 1,
                "format": "ca-pack",
            },
        )
        force_authenticate(request, user=self.user)

        response = PurchaseTdsComplianceCenterExportAPIView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn('attachment; filename="tds_compliance_ca_pack_Q1.xlsx"', response["Content-Disposition"])

        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        self.assertIn("00_Cover", workbook.sheetnames)
        self.assertIn("01_KPI_Summary", workbook.sheetnames)
        self.assertIn("02_Warnings", workbook.sheetnames)
        self.assertIn("04_Deduction_Register", workbook.sheetnames)
        self.assertEqual(workbook["00_Cover"]["A1"].value, "TDS Compliance Center CA Pack")
        self.assertEqual(workbook["04_Deduction_Register"]["A1"].value, "TDS Compliance Center - Deduction Register")


class PurchaseGstTdsComplianceCenterExportTests(SimpleTestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.user = SimpleNamespace(is_authenticated=True, id=1)

    @patch("purchase.views.tds_compliance_center.PurchaseTdsComplianceCenterAPIView.get")
    def test_ca_pack_export_uses_gst_tds_naming(self, mock_workspace_get):
        gst_payload = dict(MOCK_TDS_EXPORT_PAYLOAD)
        gst_payload["pageTitle"] = "GST-TDS Compliance Center"
        mock_workspace_get.return_value = Response(gst_payload)
        request = self.factory.get(
            "/api/purchase/statutory/gst-tds-compliance-center/export/",
            {
                "entity": 1,
                "entityfinid": 1,
                "format": "ca-pack",
            },
        )
        force_authenticate(request, user=self.user)

        response = PurchaseGstTdsComplianceCenterExportAPIView.as_view()(request)

        self.assertEqual(response.status_code, 200)
        self.assertIn('attachment; filename="gst_tds_compliance_ca_pack_Q1.xlsx"', response["Content-Disposition"])
        workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
        self.assertEqual(workbook["00_Cover"]["A1"].value, "GST-TDS Compliance Center CA Pack")


class PurchaseTdsComplianceCenterStatusTests(SimpleTestCase):
    def setUp(self):
        self.view = PurchaseTdsComplianceCenterAPIView()

    def test_draft_status_badges_respect_approval_state(self):
        challan = SimpleNamespace(status=1, payment_payload_json={"_approval_state": {"status": "SUBMITTED"}})
        filing = SimpleNamespace(status=1, filed_payload_json={"_approval_state": {"status": "APPROVED"}})

        self.assertEqual(self.view._challan_status_badge(challan)["label"], "Approval Submitted")
        self.assertEqual(self.view._return_status_badge(filing)["label"], "Approved Draft")

    def test_primary_actions_follow_purchase_statutory_workflow_states(self):
        challan = SimpleNamespace(status=2)
        submitted_return = SimpleNamespace(
            status=1,
            filed_payload_json={"_approval_state": {"status": "SUBMITTED"}},
            tax_type="IT_TDS",
            return_code="26Q",
        )
        filed_return = SimpleNamespace(
            status=2,
            filed_payload_json={},
            tax_type="IT_TDS",
            return_code="26Q",
        )

        self.assertEqual(self.view._challan_primary_action_label(challan), "Create return")
        self.assertEqual(self.view._return_primary_action_label(submitted_return), "Approve draft")
        self.assertEqual(self.view._return_primary_action_label(filed_return), "NSDL / 16A")

    def test_compact_dataset_payload_only_includes_requested_tabs(self):
        payload = self.view._build_datasets_payload(
            include_all_datasets=False,
            requested_tabs={"payment-register"},
            summary={"deducted": "10.00", "pending_deposit": "2.00"},
            monthly_rows=[{"id": "m1", "month": "Apr 2026"}],
            header_rows=[{"id": 1, "voucherNo": "PI/1"}],
            section_rows=[{"id": "194c", "section": "194C", "closingBalance": "1.00", "interest": "0.00"}],
            challan_rows=[{"id": 2, "challanNo": "CH-1"}],
            challan_mapping_rows=[{"id": 3, "voucherNo": "PI/2"}],
            deductee_rows=[{"id": 4, "deductee": "Vendor-A"}],
            pending_rows=[{"id": 5, "deductee": "Vendor-B"}],
            vendor_rows=[{"id": 6, "deductee": "Vendor-C"}],
            filing_rows={"24q": [], "26q": [], "27q": []},
            form16a_rows=[{"id": 7, "deductee": "Vendor-D"}],
            audit_rows=[{"id": 8, "action": "Updated"}],
        )

        self.assertEqual(set(payload.keys()), {"payment-register"})

    def test_compact_return_dataset_payload_only_includes_requested_bucket(self):
        payload = self.view._build_return_datasets_payload(
            include_all_return_datasets=False,
            requested_return_tabs={"27q"},
            filing_rows={
                "24q": [],
                "26q": [{"id": 1, "returnType": "26Q"}],
                "27q": [{"id": 2, "returnType": "27Q"}],
            },
        )

        self.assertEqual(set(payload.keys()), {"27q"})


class PurchaseGstTdsComplianceCenterStatusTests(SimpleTestCase):
    def setUp(self):
        self.view = PurchaseGstTdsComplianceCenterAPIView()

    def test_compact_dataset_payload_only_includes_requested_tabs(self):
        payload = self.view._build_datasets_payload(
            include_all_datasets=False,
            requested_tabs={"payment-register"},
            summary={"deducted": "10.00", "pending_deposit": "2.00"},
            monthly_rows=[{"id": "m1", "month": "Apr 2026"}],
            header_rows=[{"id": 1, "voucherNo": "PI/1"}],
            section_rows=[{"id": "gst-tds", "section": "GST-TDS", "closingBalance": "1.00", "interest": "0.00"}],
            challan_rows=[{"id": 2, "challanNo": "CH-1"}],
            challan_mapping_rows=[{"id": 3, "voucherNo": "PI/2"}],
            deductee_rows=[{"id": 4, "deductee": "Vendor-A"}],
            pending_rows=[{"id": 5, "deductee": "Vendor-B"}],
            vendor_rows=[{"id": 6, "deductee": "Vendor-C"}],
            filing_rows={"gstr7": []},
            audit_rows=[{"id": 8, "action": "Updated"}],
        )

        self.assertEqual(set(payload.keys()), {"payment-register"})

    def test_compact_return_dataset_payload_only_includes_requested_bucket(self):
        payload = self.view._build_return_datasets_payload(
            include_all_return_datasets=False,
            requested_return_tabs={"gstr7"},
            filing_rows={"gstr7": [{"id": 1, "returnType": "GSTR-7"}]},
        )

        self.assertEqual(set(payload.keys()), {"gstr7"})
