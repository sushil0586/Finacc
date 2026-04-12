from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook

from entity.models import Entity, SubEntity
from financial.models import account

from .models import (
    Brand,
    HsnSac,
    OpeningStockByLocation,
    PriceList,
    Product,
    ProductBarcode,
    ProductCategory,
    ProductGstRate,
    ProductPrice,
    ProductUomConversion,
    UnitOfMeasure,
)

SHEETS = [
    "categories_master",
    "uoms_master",
    "products_basic",
    "gst_rates",
    "prices",
    "barcodes",
    "opening_stocks",
    "uom_conversions",
]


def _parse_date(value: Any):
    if value in (None, "", "-", "--"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date: {s}")


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in {"1", "true", "yes", "y"}


def _to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, "", "-", "--"):
        return default
    return Decimal(str(value))


def _to_int(value: Any, default: int | None = None) -> int | None:
    if value in (None, "", "-", "--"):
        return default
    return int(str(value))


def _read_xlsx(content: bytes) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    sheet_lookup = {str(name).strip().lower(): name for name in wb.sheetnames}
    out: dict[str, list[dict[str, Any]]] = {}
    for sheet in SHEETS:
        actual_sheet = sheet_lookup.get(sheet.lower())
        if not actual_sheet:
            out[sheet] = []
            continue
        ws = wb[actual_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[sheet] = []
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        payload = []
        for row in rows[1:]:
            if not any(v not in (None, "") for v in row):
                continue
            payload.append({headers[i]: row[i] for i in range(len(headers))})
        out[sheet] = payload
    return out


def _read_csv_zip(content: bytes) -> dict[str, list[dict[str, Any]]]:
    out = {name: [] for name in SHEETS}
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        name_lookup = {str(name).strip().lower(): name for name in zf.namelist()}
        for sheet in SHEETS:
            fname = name_lookup.get(f"{sheet}.csv")
            if not fname:
                continue
            data = zf.read(fname).decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(data))
            out[sheet] = [dict(r) for r in reader if any((v or "").strip() for v in r.values())]
    return out


def _write_xlsx(data: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in data.keys():
        rows = data.get(sheet, [])
        ws = wb.create_sheet(sheet)
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(col) for col in headers])
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _write_csv_zip(data: dict[str, list[dict[str, Any]]]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        for sheet in data.keys():
            rows = data.get(sheet, [])
            stream = io.StringIO()
            headers = list(rows[0].keys()) if rows else []
            writer = csv.DictWriter(stream, fieldnames=headers)
            if headers:
                writer.writeheader()
                writer.writerows(rows)
            zf.writestr(f"{sheet}.csv", stream.getvalue())
    return buff.getvalue()


def template_payload() -> dict[str, list[dict[str, Any]]]:
    return {
        "categories_master": [
            {
                "pcategoryname": "Raw Materials",
                "maincategory": "",
                "level": 1,
                "isactive": True,
            }
        ],
        "uoms_master": [
            {
                "code": "NOS",
                "description": "Numbers",
                "uqc": "NOS",
                "isactive": True,
            }
        ],
        "products_basic": [
            {
                "sku": "SKU-001",
                "productname": "Sample Product",
                "productdesc": "Optional description",
                "category": "Raw Materials",
                "brand": "Generic",
                "base_uom_code": "NOS",
                "sales_account_code": "",
                "purchase_account_code": "",
                "is_service": False,
                "item_classification": "trading_item",
                "is_batch_managed": False,
                "is_serialized": False,
                "is_expiry_tracked": False,
                "shelf_life_days": "",
                "expiry_warning_days": 30,
                "is_ecomm_9_5_service": False,
                "default_is_rcm": False,
                "is_itc_eligible": True,
                "product_status": "active",
                "launch_date": "01-04-2026",
                "discontinue_date": "",
                "isactive": True,
            }
        ],
        "gst_rates": [
            {
                "sku": "SKU-001",
                "hsn_code": "7203",
                "gst_type": "regular",
                "sgst": "9.00",
                "cgst": "9.00",
                "igst": "18.00",
                "cess": "0.00",
                "cess_type": "none",
                "cess_specific_amount": "",
                "valid_from": "01-04-2026",
                "valid_to": "",
                "isdefault": True,
            }
        ],
        "prices": [
            {
                "sku": "SKU-001",
                "pricelist": "Default",
                "uom_code": "NOS",
                "purchase_rate": "100.00",
                "purchase_rate_less_percent": "0.00",
                "mrp": "120.00",
                "mrp_less_percent": "0.00",
                "selling_price": "110.00",
                "effective_from": "01-04-2026",
                "effective_to": "",
            }
        ],
        "barcodes": [
            {
                "sku": "SKU-001",
                "barcode": "",
                "uom_code": "NOS",
                "pack_size": 1,
                "mrp": "120.00",
                "selling_price": "110.00",
                "isprimary": True,
            }
        ],
        "opening_stocks": [
            {
                "sku": "SKU-001",
                "location_code": "MAIN",
                "openingqty": "100.00",
                "openingrate": "100.00",
                "openingvalue": "10000.00",
                "as_of_date": "01-04-2026",
            }
        ],
        "uom_conversions": [
            {
                "sku": "SKU-001",
                "from_uom_code": "BOX",
                "to_uom_code": "NOS",
                "factor": "10.0000",
            }
        ],
    }


def export_payload(entity: Entity, search: str = "") -> dict[str, list[dict[str, Any]]]:
    products_qs = Product.objects.filter(entity=entity)
    if search:
        products_qs = products_qs.filter(productname__icontains=search)
    products = list(products_qs.select_related("productcategory", "brand", "base_uom", "sales_account", "purchase_account"))
    product_ids = [p.id for p in products]
    sku_by_id = {p.id: p.sku for p in products}

    gst_rows = ProductGstRate.objects.filter(product_id__in=product_ids).select_related("hsn")
    price_rows = ProductPrice.objects.filter(product_id__in=product_ids).select_related("pricelist", "uom")
    barcode_rows = ProductBarcode.objects.filter(product_id__in=product_ids).select_related("uom")
    opening_rows = OpeningStockByLocation.objects.filter(product_id__in=product_ids).select_related("location")
    uom_rows = ProductUomConversion.objects.filter(product_id__in=product_ids).select_related("from_uom", "to_uom")
    categories = ProductCategory.objects.filter(entity=entity).select_related("maincategory").order_by("pcategoryname")
    uoms = UnitOfMeasure.objects.filter(entity=entity).order_by("code")

    return {
        "categories_master": [
            {
                "pcategoryname": c.pcategoryname,
                "maincategory": c.maincategory.pcategoryname if c.maincategory else "",
                "level": c.level,
                "isactive": c.isactive,
            }
            for c in categories
        ],
        "uoms_master": [
            {
                "code": u.code,
                "description": u.description or "",
                "uqc": u.uqc or "",
                "isactive": u.isactive,
            }
            for u in uoms
        ],
        "products_basic": [
            {
                "sku": p.sku,
                "productname": p.productname,
                "productdesc": p.productdesc,
                "category": p.productcategory.pcategoryname if p.productcategory else "",
                "brand": p.brand.name if p.brand else "",
                "base_uom_code": p.base_uom.code if p.base_uom else "",
                "sales_account_code": getattr(p.sales_account, "ledger_code", "") or "",
                "purchase_account_code": getattr(p.purchase_account, "ledger_code", "") or "",
                "is_service": p.is_service,
                "item_classification": p.item_classification,
                "is_batch_managed": p.is_batch_managed,
                "is_serialized": p.is_serialized,
                "is_expiry_tracked": p.is_expiry_tracked,
                "shelf_life_days": p.shelf_life_days if p.shelf_life_days is not None else "",
                "expiry_warning_days": p.expiry_warning_days,
                "is_ecomm_9_5_service": p.is_ecomm_9_5_service,
                "default_is_rcm": p.default_is_rcm,
                "is_itc_eligible": p.is_itc_eligible,
                "product_status": p.product_status,
                "launch_date": p.launch_date.isoformat() if p.launch_date else "",
                "discontinue_date": p.discontinue_date.isoformat() if p.discontinue_date else "",
                "isactive": p.isactive,
            }
            for p in products
        ],
        "gst_rates": [
            {
                "sku": sku_by_id.get(r.product_id, ""),
                "hsn_code": r.hsn.code if r.hsn else "",
                "gst_type": r.gst_type,
                "sgst": str(r.sgst),
                "cgst": str(r.cgst),
                "igst": str(r.igst),
                "cess": str(r.cess),
                "cess_type": r.cess_type,
                "cess_specific_amount": str(r.cess_specific_amount or ""),
                "valid_from": r.valid_from.isoformat() if r.valid_from else "",
                "valid_to": r.valid_to.isoformat() if r.valid_to else "",
                "isdefault": r.isdefault,
            }
            for r in gst_rows
        ],
        "prices": [
            {
                "sku": sku_by_id.get(r.product_id, ""),
                "pricelist": r.pricelist.name if r.pricelist else "",
                "uom_code": r.uom.code if r.uom else "",
                "purchase_rate": str(r.purchase_rate or ""),
                "purchase_rate_less_percent": str(r.purchase_rate_less_percent or ""),
                "mrp": str(r.mrp or ""),
                "mrp_less_percent": str(r.mrp_less_percent or ""),
                "selling_price": str(r.selling_price),
                "effective_from": r.effective_from.isoformat() if r.effective_from else "",
                "effective_to": r.effective_to.isoformat() if r.effective_to else "",
            }
            for r in price_rows
        ],
        "barcodes": [
            {
                "sku": sku_by_id.get(r.product_id, ""),
                "barcode": r.barcode,
                "uom_code": r.uom.code if r.uom else "",
                "pack_size": r.pack_size or 1,
                "mrp": str(r.mrp or ""),
                "selling_price": str(r.selling_price or ""),
                "isprimary": r.isprimary,
            }
            for r in barcode_rows
        ],
        "opening_stocks": [
            {
                "sku": sku_by_id.get(r.product_id, ""),
                "location_code": r.location.subentity_code if r.location else "",
                "openingqty": str(r.openingqty or ""),
                "openingrate": str(r.openingrate or ""),
                "openingvalue": str(r.openingvalue or ""),
                "as_of_date": r.as_of_date.isoformat() if r.as_of_date else "",
            }
            for r in opening_rows
        ],
        "uom_conversions": [
            {
                "sku": sku_by_id.get(r.product_id, ""),
                "from_uom_code": r.from_uom.code if r.from_uom else "",
                "to_uom_code": r.to_uom.code if r.to_uom else "",
                "factor": str(r.factor),
            }
            for r in uom_rows
        ],
    }


@dataclass
class ImportResult:
    summary: dict[str, Any]
    errors: list[dict[str, Any]]


def parse_payload(file_bytes: bytes, fmt: str) -> dict[str, list[dict[str, Any]]]:
    if fmt == "xlsx":
        return _read_xlsx(file_bytes)
    return _read_csv_zip(file_bytes)


def render_payload(payload: dict[str, list[dict[str, Any]]], fmt: str) -> bytes:
    if fmt == "xlsx":
        return _write_xlsx(payload)
    return _write_csv_zip(payload)


def validate_payload(payload: dict[str, list[dict[str, Any]]], entity: Entity) -> ImportResult:
    errors: list[dict[str, Any]] = []
    total_rows = sum(len(payload.get(sheet, [])) for sheet in SHEETS)
    if total_rows == 0:
        errors.append(
            {
                "sheet": "workbook",
                "row": 0,
                "field": "file",
                "message": "No data rows found. Use the latest template and keep sheet/file names unchanged.",
            }
        )
        summary = {
            "rows": {sheet: len(payload.get(sheet, [])) for sheet in SHEETS},
            "error_count": len(errors),
        }
        return ImportResult(summary=summary, errors=errors)

    seen_sku: set[str] = set()
    existing_skus = set(Product.objects.filter(entity=entity).values_list("sku", flat=True))
    existing_categories = {c.strip().lower() for c in ProductCategory.objects.filter(entity=entity).values_list("pcategoryname", flat=True)}
    existing_uoms = {c.strip().lower() for c in UnitOfMeasure.objects.filter(entity=entity).values_list("code", flat=True)}
    file_categories = {(row.get("pcategoryname") or "").strip().lower() for row in payload.get("categories_master", []) if (row.get("pcategoryname") or "").strip()}
    file_uoms = {(row.get("code") or "").strip().lower() for row in payload.get("uoms_master", []) if (row.get("code") or "").strip()}

    for idx, row in enumerate(payload.get("products_basic", []), start=2):
        sku = (row.get("sku") or "").strip()
        name = (row.get("productname") or "").strip()
        if not sku:
            errors.append({"sheet": "products_basic", "row": idx, "field": "sku", "message": "SKU is required."})
            continue
        if not name:
            errors.append({"sheet": "products_basic", "row": idx, "field": "productname", "message": "Product name is required."})
        if sku in seen_sku:
            errors.append({"sheet": "products_basic", "row": idx, "field": "sku", "message": "Duplicate SKU in upload."})
        seen_sku.add(sku)

        try:
            _parse_date(row.get("launch_date"))
            _parse_date(row.get("discontinue_date"))
        except Exception as exc:
            errors.append({"sheet": "products_basic", "row": idx, "field": "launch_date", "message": str(exc)})

        try:
            shelf_life_days = _to_int(row.get("shelf_life_days"), default=None)
            expiry_warning_days = _to_int(row.get("expiry_warning_days"), default=None)
            if shelf_life_days is not None and shelf_life_days <= 0:
                raise ValueError("Shelf life days must be greater than 0.")
            if expiry_warning_days is not None and expiry_warning_days < 0:
                raise ValueError("Expiry warning days cannot be negative.")
            if shelf_life_days is not None and expiry_warning_days is not None and expiry_warning_days > shelf_life_days:
                raise ValueError("Expiry warning days cannot exceed shelf life days.")
        except Exception as exc:
            field = "shelf_life_days" if "Shelf life" in str(exc) else "expiry_warning_days"
            errors.append({"sheet": "products_basic", "row": idx, "field": field, "message": str(exc)})

        category = (row.get("category") or "").strip().lower()
        base_uom = (row.get("base_uom_code") or "").strip().lower()
        if not category:
            errors.append(
                {
                    "sheet": "products_basic",
                    "row": idx,
                    "field": "category",
                    "message": "Category is required.",
                }
            )
        if category and category not in existing_categories and category not in file_categories:
            errors.append(
                {
                    "sheet": "products_basic",
                    "row": idx,
                    "field": "category",
                    "message": f"Category '{row.get('category')}' not found in DB or categories_master sheet.",
                }
            )
        if not base_uom:
            errors.append(
                {
                    "sheet": "products_basic",
                    "row": idx,
                    "field": "base_uom_code",
                    "message": "base_uom_code is required.",
                }
            )
        if base_uom and base_uom not in existing_uoms and base_uom not in file_uoms:
            errors.append(
                {
                    "sheet": "products_basic",
                    "row": idx,
                    "field": "base_uom_code",
                    "message": f"UOM '{row.get('base_uom_code')}' not found in DB or uoms_master sheet.",
                }
            )

    for idx, row in enumerate(payload.get("categories_master", []), start=2):
        name = (row.get("pcategoryname") or "").strip()
        if not name:
            errors.append({"sheet": "categories_master", "row": idx, "field": "pcategoryname", "message": "Category name is required."})
            continue
        main_name = (row.get("maincategory") or "").strip().lower()
        if main_name and main_name not in existing_categories and main_name not in file_categories:
            errors.append(
                {
                    "sheet": "categories_master",
                    "row": idx,
                    "field": "maincategory",
                    "message": f"maincategory '{row.get('maincategory')}' not found in DB or categories_master sheet.",
                }
            )

    for idx, row in enumerate(payload.get("uoms_master", []), start=2):
        if not (row.get("code") or "").strip():
            errors.append({"sheet": "uoms_master", "row": idx, "field": "code", "message": "UOM code is required."})

    skus_in_file = seen_sku.union(existing_skus)
    for sheet in ("gst_rates", "prices", "barcodes", "opening_stocks", "uom_conversions"):
        for idx, row in enumerate(payload.get(sheet, []), start=2):
            sku = (row.get("sku") or "").strip()
            if not sku:
                errors.append({"sheet": sheet, "row": idx, "field": "sku", "message": "SKU is required."})
                continue
            if sku not in skus_in_file:
                errors.append({"sheet": sheet, "row": idx, "field": "sku", "message": f"SKU '{sku}' not found."})

    summary = {
        "rows": {sheet: len(payload.get(sheet, [])) for sheet in SHEETS},
        "error_count": len(errors),
    }
    return ImportResult(summary=summary, errors=errors)


def commit_payload(
    payload: dict[str, list[dict[str, Any]]],
    entity: Entity,
    *,
    upsert_mode: str = "upsert",
    duplicate_strategy: str = "fail",
) -> ImportResult:
    errors: list[dict[str, Any]] = []
    summary = {"created": 0, "updated": 0, "skipped": 0}

    cat_map = {c.pcategoryname.strip().lower(): c for c in ProductCategory.objects.filter(entity=entity)}
    brand_map = {b.name.strip().lower(): b for b in Brand.objects.filter(entity=entity)}
    uom_map = {u.code.strip().lower(): u for u in UnitOfMeasure.objects.filter(entity=entity)}
    hsn_map = {h.code.strip().lower(): h for h in HsnSac.objects.filter(entity=entity)}
    price_map = {p.name.strip().lower(): p for p in PriceList.objects.filter(entity=entity)}
    subentity_map = {s.subentity_code.strip().lower(): s for s in SubEntity.objects.filter(entity=entity)}
    acc_code_map = {str(a.ledger.ledger_code): a for a in account.objects.filter(entity=entity, ledger__isnull=False).select_related("ledger")}

    product_map = {p.sku: p for p in Product.objects.filter(entity=entity)}

    def _should_apply(created: bool) -> bool:
        if upsert_mode == "create_only":
            return created
        if upsert_mode == "update_only":
            return not created
        return True

    with transaction.atomic():
        for idx, row in enumerate(payload.get("categories_master", []), start=2):
            name = (row.get("pcategoryname") or "").strip()
            if not name:
                continue
            existing = cat_map.get(name.lower())
            if existing and upsert_mode == "create_only":
                summary["skipped"] += 1
                continue
            if not existing and upsert_mode == "update_only":
                summary["skipped"] += 1
                continue
            try:
                obj = existing or ProductCategory(entity=entity, pcategoryname=name)
                level_raw = row.get("level")
                obj.level = int(level_raw) if str(level_raw or "").strip() else 1
                obj.isactive = _to_bool(row.get("isactive"), default=True)
                obj.full_clean()
                obj.save()
                cat_map[name.lower()] = obj
                summary["updated" if existing else "created"] += 1
            except Exception as exc:
                errors.append({"sheet": "categories_master", "row": idx, "field": "row", "message": str(exc)})

        # Second pass so parent/child in same upload works regardless of row order.
        for idx, row in enumerate(payload.get("categories_master", []), start=2):
            name = (row.get("pcategoryname") or "").strip()
            if not name:
                continue
            main_name = (row.get("maincategory") or "").strip()
            if not main_name:
                continue
            try:
                obj = cat_map.get(name.lower())
                parent = cat_map.get(main_name.lower())
                if not obj:
                    continue
                if not parent:
                    raise ValueError(f"maincategory '{main_name}' not found.")
                obj.maincategory = parent
                obj.full_clean()
                obj.save(update_fields=["maincategory", "modifiedon"])
            except Exception as exc:
                errors.append({"sheet": "categories_master", "row": idx, "field": "maincategory", "message": str(exc)})

        for idx, row in enumerate(payload.get("uoms_master", []), start=2):
            code = (row.get("code") or "").strip()
            if not code:
                continue
            existing = uom_map.get(code.lower())
            if existing and upsert_mode == "create_only":
                summary["skipped"] += 1
                continue
            if not existing and upsert_mode == "update_only":
                summary["skipped"] += 1
                continue
            try:
                obj = existing or UnitOfMeasure(entity=entity, code=code)
                obj.description = (row.get("description") or "").strip()
                obj.uqc = (row.get("uqc") or "").strip() or None
                obj.isactive = _to_bool(row.get("isactive"), default=True)
                obj.full_clean()
                obj.save()
                uom_map[code.lower()] = obj
                summary["updated" if existing else "created"] += 1
            except Exception as exc:
                errors.append({"sheet": "uoms_master", "row": idx, "field": "row", "message": str(exc)})

        for idx, row in enumerate(payload.get("products_basic", []), start=2):
            sku = (row.get("sku") or "").strip()
            if not sku:
                continue
            existing = product_map.get(sku)
            if existing and upsert_mode == "create_only":
                summary["skipped"] += 1
                continue
            if not existing and upsert_mode == "update_only":
                summary["skipped"] += 1
                continue
            try:
                category_key = (row.get("category") or "").strip().lower()
                uom_key = (row.get("base_uom_code") or "").strip().lower()
                category = cat_map.get(category_key)
                base_uom = uom_map.get(uom_key)
                if not category and category_key and upsert_mode != "update_only":
                    category = ProductCategory.objects.create(
                        entity=entity,
                        pcategoryname=(row.get("category") or "").strip(),
                        level=1,
                        isactive=True,
                    )
                    cat_map[category_key] = category
                    summary["created"] += 1
                if not base_uom and uom_key and upsert_mode != "update_only":
                    base_uom = UnitOfMeasure.objects.create(
                        entity=entity,
                        code=(row.get("base_uom_code") or "").strip(),
                        description=(row.get("base_uom_code") or "").strip(),
                        isactive=True,
                    )
                    uom_map[uom_key] = base_uom
                    summary["created"] += 1
                if not category or not base_uom:
                    raise ValueError(
                        f"Category/base_uom_code not valid for SKU '{sku}' "
                        f"(category='{row.get('category')}', base_uom_code='{row.get('base_uom_code')}')."
                    )
                obj = existing or Product(entity=entity, sku=sku)
                obj.productname = (row.get("productname") or "").strip()
                obj.productdesc = (row.get("productdesc") or "").strip()
                obj.productcategory = category
                obj.brand = brand_map.get((row.get("brand") or "").strip().lower())
                obj.base_uom = base_uom
                obj.sales_account = acc_code_map.get(str(row.get("sales_account_code") or "").strip())
                obj.purchase_account = acc_code_map.get(str(row.get("purchase_account_code") or "").strip())
                obj.is_service = _to_bool(row.get("is_service"))
                obj.item_classification = (row.get("item_classification") or obj.item_classification or "trading_item")
                obj.is_batch_managed = _to_bool(row.get("is_batch_managed"))
                obj.is_serialized = _to_bool(row.get("is_serialized"))
                obj.is_expiry_tracked = _to_bool(row.get("is_expiry_tracked"))
                obj.shelf_life_days = _to_int(row.get("shelf_life_days"), default=None)
                obj.expiry_warning_days = _to_int(row.get("expiry_warning_days"), default=30)
                obj.is_ecomm_9_5_service = _to_bool(row.get("is_ecomm_9_5_service"))
                obj.default_is_rcm = _to_bool(row.get("default_is_rcm"))
                obj.is_itc_eligible = _to_bool(row.get("is_itc_eligible"), default=True)
                obj.product_status = (row.get("product_status") or obj.product_status or "active")
                obj.launch_date = _parse_date(row.get("launch_date"))
                obj.discontinue_date = _parse_date(row.get("discontinue_date"))
                obj.isactive = _to_bool(row.get("isactive"), default=True)
                obj.full_clean()
                obj.save()
                product_map[sku] = obj
                summary["updated" if existing else "created"] += 1
            except Exception as exc:
                errors.append({"sheet": "products_basic", "row": idx, "field": "row", "message": str(exc)})

        for idx, row in enumerate(payload.get("gst_rates", []), start=2):
            sku = (row.get("sku") or "").strip()
            product = product_map.get(sku)
            if not product:
                continue
            try:
                hsn = hsn_map.get((row.get("hsn_code") or "").strip().lower())
                if not hsn:
                    raise ValueError("Invalid hsn_code")
                valid_from = _parse_date(row.get("valid_from"))
                valid_to = _parse_date(row.get("valid_to"))
                obj, created = ProductGstRate.objects.get_or_create(
                    product=product,
                    hsn=hsn,
                    valid_from=valid_from,
                    defaults={},
                )
                if _should_apply(created):
                    obj.gst_type = (row.get("gst_type") or obj.gst_type or "regular")
                    obj.sgst = _to_decimal(row.get("sgst"))
                    obj.cgst = _to_decimal(row.get("cgst"))
                    obj.igst = _to_decimal(row.get("igst"))
                    obj.cess = _to_decimal(row.get("cess"))
                    obj.cess_type = (row.get("cess_type") or obj.cess_type or "none")
                    obj.cess_specific_amount = _to_decimal(row.get("cess_specific_amount"), Decimal("0"))
                    obj.valid_to = valid_to
                    obj.isdefault = _to_bool(row.get("isdefault"))
                    # ProductGstRate.save() normalizes IGST/gst_rate from CGST+SGST
                    # and then runs full_clean(); avoid pre-clean here.
                    obj.save()
                    summary["updated" if not created else "created"] += 1
                else:
                    summary["skipped"] += 1
            except Exception as exc:
                errors.append({"sheet": "gst_rates", "row": idx, "field": "row", "message": str(exc)})

        for idx, row in enumerate(payload.get("prices", []), start=2):
            sku = (row.get("sku") or "").strip()
            product = product_map.get(sku)
            if not product:
                continue
            try:
                pricelist = price_map.get((row.get("pricelist") or "").strip().lower())
                uom = uom_map.get((row.get("uom_code") or "").strip().lower())
                if not pricelist or not uom:
                    raise ValueError("Invalid pricelist or uom_code")
                effective_from = _parse_date(row.get("effective_from"))
                obj, created = ProductPrice.objects.get_or_create(
                    product=product,
                    pricelist=pricelist,
                    uom=uom,
                    effective_from=effective_from,
                    defaults={"selling_price": Decimal("0")},
                )
                if _should_apply(created):
                    obj.purchase_rate = _to_decimal(row.get("purchase_rate"))
                    obj.purchase_rate_less_percent = _to_decimal(row.get("purchase_rate_less_percent"))
                    obj.mrp = _to_decimal(row.get("mrp"))
                    obj.mrp_less_percent = _to_decimal(row.get("mrp_less_percent"))
                    obj.selling_price = _to_decimal(row.get("selling_price"))
                    obj.effective_to = _parse_date(row.get("effective_to"))
                    obj.full_clean()
                    obj.save()
                    summary["updated" if not created else "created"] += 1
                else:
                    summary["skipped"] += 1
            except Exception as exc:
                errors.append({"sheet": "prices", "row": idx, "field": "row", "message": str(exc)})

        for idx, row in enumerate(payload.get("barcodes", []), start=2):
            sku = (row.get("sku") or "").strip()
            product = product_map.get(sku)
            if not product:
                continue
            try:
                uom = uom_map.get((row.get("uom_code") or "").strip().lower())
                if not uom:
                    raise ValueError("Invalid uom_code")
                pack_size = int(row.get("pack_size") or 1)
                obj, created = ProductBarcode.objects.get_or_create(
                    product=product,
                    uom=uom,
                    pack_size=pack_size,
                    defaults={},
                )
                if _should_apply(created):
                    obj.mrp = _to_decimal(row.get("mrp"))
                    obj.selling_price = _to_decimal(row.get("selling_price"))
                    obj.isprimary = _to_bool(row.get("isprimary"))
                    provided_barcode = (row.get("barcode") or "").strip()
                    if provided_barcode:
                        obj.barcode = provided_barcode
                    obj.full_clean()
                    obj.save()
                    summary["updated" if not created else "created"] += 1
                else:
                    summary["skipped"] += 1
            except Exception as exc:
                errors.append({"sheet": "barcodes", "row": idx, "field": "row", "message": str(exc)})

        for idx, row in enumerate(payload.get("opening_stocks", []), start=2):
            sku = (row.get("sku") or "").strip()
            product = product_map.get(sku)
            if not product:
                continue
            try:
                location = subentity_map.get((row.get("location_code") or "").strip().lower())
                if not location:
                    raise ValueError("Invalid location_code (subentity_code)")
                as_of_date = _parse_date(row.get("as_of_date"))
                obj, created = OpeningStockByLocation.objects.get_or_create(
                    entity=entity,
                    product=product,
                    location=location,
                    as_of_date=as_of_date,
                    defaults={},
                )
                if _should_apply(created):
                    obj.openingqty = _to_decimal(row.get("openingqty"))
                    obj.openingrate = _to_decimal(row.get("openingrate"))
                    obj.openingvalue = _to_decimal(row.get("openingvalue"))
                    obj.save()
                    summary["updated" if not created else "created"] += 1
                else:
                    summary["skipped"] += 1
            except Exception as exc:
                errors.append({"sheet": "opening_stocks", "row": idx, "field": "row", "message": str(exc)})

        for idx, row in enumerate(payload.get("uom_conversions", []), start=2):
            sku = (row.get("sku") or "").strip()
            product = product_map.get(sku)
            if not product:
                continue
            try:
                from_uom = uom_map.get((row.get("from_uom_code") or "").strip().lower())
                to_uom = uom_map.get((row.get("to_uom_code") or "").strip().lower())
                if not from_uom or not to_uom:
                    raise ValueError("Invalid from_uom_code or to_uom_code")
                obj, created = ProductUomConversion.objects.get_or_create(
                    product=product,
                    from_uom=from_uom,
                    to_uom=to_uom,
                    defaults={"factor": Decimal("1")},
                )
                if _should_apply(created):
                    obj.factor = _to_decimal(row.get("factor"), Decimal("1"))
                    obj.full_clean()
                    obj.save()
                    summary["updated" if not created else "created"] += 1
                else:
                    summary["skipped"] += 1
            except Exception as exc:
                errors.append({"sheet": "uom_conversions", "row": idx, "field": "row", "message": str(exc)})

    summary["error_count"] = len(errors)
    return ImportResult(summary=summary, errors=errors)
