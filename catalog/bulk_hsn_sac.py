from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook

from entity.models import Entity

from .models import HsnSac

SHEET = "hsn_sac"


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, "", "-", "--"):
        return default
    text = str(value).strip()
    if not text:
        return default

    # Some user files carry placeholder tokens like "Val-1"; treat them as blank/default.
    if re.fullmatch(r"val-\d+", text, flags=re.IGNORECASE):
        return default

    # Accept common import formats like "18%", "18 %", or "1,234.50"
    if text.endswith("%"):
        text = text[:-1].strip()
    text = text.replace(",", "")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def _read_xlsx(content: bytes) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    sheet_lookup = {str(name).strip().lower(): name for name in wb.sheetnames}
    actual = sheet_lookup.get(SHEET)
    if not actual:
        return {SHEET: []}
    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {SHEET: []}
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    payload: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        payload.append({headers[i]: row[i] for i in range(len(headers))})
    return {SHEET: payload}


def _read_csv_zip(content: bytes) -> dict[str, list[dict[str, Any]]]:
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        name_lookup = {str(name).strip().lower(): name for name in zf.namelist()}
        fname = name_lookup.get(f"{SHEET}.csv")
        if not fname:
            return {SHEET: []}
        data = zf.read(fname).decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(data))
        return {SHEET: [dict(r) for r in reader if any((v or "").strip() for v in r.values())]}


def _write_xlsx(data: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    rows = data.get(SHEET, [])
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(col) for col in headers])
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _write_csv_zip(data: dict[str, list[dict[str, Any]]]) -> bytes:
    buff = io.BytesIO()
    with zipfile.ZipFile(buff, "w", zipfile.ZIP_DEFLATED) as zf:
        rows = data.get(SHEET, [])
        stream = io.StringIO()
        headers = list(rows[0].keys()) if rows else []
        writer = csv.DictWriter(stream, fieldnames=headers)
        if headers:
            writer.writeheader()
            writer.writerows(rows)
        zf.writestr(f"{SHEET}.csv", stream.getvalue())
    return buff.getvalue()


def template_payload() -> dict[str, list[dict[str, Any]]]:
    return {
        SHEET: [
            {
                "code": "7203",
                "description": "Steel ingot",
                "is_service": False,
                "default_sgst": "9.00",
                "default_cgst": "9.00",
                "default_igst": "18.00",
                "default_cess": "0.00",
                "is_exempt": False,
                "is_nil_rated": False,
                "is_non_gst": False,
                "isactive": True,
            }
        ]
    }


def export_payload(entity: Entity, search: str = "") -> dict[str, list[dict[str, Any]]]:
    qs = HsnSac.objects.filter(entity=entity)
    if search:
        qs = qs.filter(code__icontains=search)
    rows = [
        {
            "code": row.code,
            "description": row.description,
            "is_service": row.is_service,
            "default_sgst": str(row.default_sgst),
            "default_cgst": str(row.default_cgst),
            "default_igst": str(row.default_igst),
            "default_cess": str(row.default_cess),
            "is_exempt": row.is_exempt,
            "is_nil_rated": row.is_nil_rated,
            "is_non_gst": row.is_non_gst,
            "isactive": row.isactive,
        }
        for row in qs.order_by("code")
    ]
    return {SHEET: rows}


@dataclass
class ImportResult:
    summary: dict[str, Any]
    errors: list[dict[str, Any]]


def parse_payload(file_bytes: bytes, fmt: str) -> dict[str, list[dict[str, Any]]]:
    if fmt == "xlsx":
        return _read_xlsx(file_bytes)
    return _read_csv_zip(file_bytes)


def render_payload(payload: dict[str, list[dict[str, Any]]], fmt: str) -> bytes:
    if fmt == "xlsx":
        return _write_xlsx(payload)
    return _write_csv_zip(payload)


def validate_payload(payload: dict[str, list[dict[str, Any]]], entity: Entity) -> ImportResult:
    errors: list[dict[str, Any]] = []
    rows = payload.get(SHEET, [])
    if not rows:
        errors.append({"sheet": SHEET, "row": 0, "field": "file", "message": "No HSN/SAC rows found in upload."})
        return ImportResult(summary={"rows": {SHEET: 0}, "error_count": 1}, errors=errors)

    seen_codes: set[str] = set()
    for idx, row in enumerate(rows, start=2):
        code = (row.get("code") or "").strip()
        if not code:
            errors.append({"sheet": SHEET, "row": idx, "field": "code", "message": "Code is required."})
            continue
        if code.lower() in seen_codes:
            errors.append({"sheet": SHEET, "row": idx, "field": "code", "message": "Duplicate code in upload."})
        seen_codes.add(code.lower())
        for field_name in ("default_sgst", "default_cgst", "default_igst", "default_cess"):
            try:
                _to_decimal(row.get(field_name))
            except Exception as exc:
                errors.append({"sheet": SHEET, "row": idx, "field": field_name, "message": str(exc)})

    return ImportResult(summary={"rows": {SHEET: len(rows)}, "error_count": len(errors)}, errors=errors)


def commit_payload(payload: dict[str, list[dict[str, Any]]], entity: Entity, *, upsert_mode: str = "upsert", duplicate_strategy: str = "fail") -> ImportResult:
    rows = payload.get(SHEET, [])
    errors: list[dict[str, Any]] = []
    summary = {"created": 0, "updated": 0, "skipped": 0}

    existing_map = {obj.code.strip().lower(): obj for obj in HsnSac.objects.filter(entity=entity)}

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            code = (row.get("code") or "").strip()
            if not code:
                continue
            key = code.lower()
            existing = existing_map.get(key)

            if existing and upsert_mode == "create_only":
                summary["skipped"] += 1
                continue
            if not existing and upsert_mode == "update_only":
                summary["skipped"] += 1
                continue

            try:
                obj = existing or HsnSac(entity=entity, code=code)
                obj.code = code
                obj.description = (row.get("description") or "").strip()
                obj.is_service = _to_bool(row.get("is_service"))
                obj.default_sgst = _to_decimal(row.get("default_sgst"))
                obj.default_cgst = _to_decimal(row.get("default_cgst"))
                obj.default_igst = _to_decimal(row.get("default_igst"))
                obj.default_cess = _to_decimal(row.get("default_cess"))
                obj.is_exempt = _to_bool(row.get("is_exempt"))
                obj.is_nil_rated = _to_bool(row.get("is_nil_rated"))
                obj.is_non_gst = _to_bool(row.get("is_non_gst"))
                obj.isactive = _to_bool(row.get("isactive"), default=True)
                obj.full_clean()
                obj.save()
                existing_map[key] = obj
                summary["updated" if existing else "created"] += 1
            except Exception as exc:
                errors.append({"sheet": SHEET, "row": idx, "field": "row", "message": str(exc)})
                if duplicate_strategy == "fail":
                    # fail-fast mode still returns structured errors via API
                    break

    summary["error_count"] = len(errors)
    return ImportResult(summary=summary, errors=errors)
